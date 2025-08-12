#!/usr/bin/env python3
"""
Script Principal para Análisis Excel Modularizado
=================================================

Este script demuestra cómo usar el módulo analisis-excel para procesar
archivos de estados financieros de manera masiva.

Uso:
    python run_analisis_excel.py --mode single --file archivo.xlsx
    python run_analisis_excel.py --mode bulk --input-dir ./data/Reports --output-dir ./data/Analisis
"""

import argparse
import sys
from pathlib import Path

# Agregar el directorio padre al path para importar analisis-excel
sys.path.insert(0, str(Path(__file__).parent))

try:
    from analisis_excel import DataExtractor, RatioCalculator, FormulaBuilder, ExcelFormatter, BulkProcessor
except ImportError as e:
    print(f"❌ Error importando módulo analisis_excel: {e}")
    print("💡 Asegúrate de que la carpeta 'analisis-excel' existe en el directorio actual")
    print("💡 Ejecuta 'python migrate_to_modular.py' para verificar la instalación")
    sys.exit(1)


def process_single_file(file_path: str, output_dir: str = "./output", analysis_type: str = "formulas"):
    """
    Procesa un único archivo Excel.
    
    Args:
        file_path: Ruta al archivo Excel
        output_dir: Directorio de salida
        analysis_type: Tipo de análisis ("formulas" o "values")
    """
    print(f"\\n🔍 Procesando archivo único: {file_path}")
    print(f"📁 Directorio de salida: {output_dir}")
    print(f"📊 Tipo de análisis: {analysis_type}")
    
    # Crear directorio de salida
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    try:
        # 1. Extraer datos
        print("\\n1️⃣ Extrayendo datos...")
        extractor = DataExtractor(file_path)
        if not extractor.load_data():
            print("❌ Error: No se pudieron cargar los datos del archivo")
            return False
        
        financial_data = extractor.get_all_financial_data()
        years = financial_data.get("years", [])
        print(f"✅ Datos extraídos exitosamente")
        print(f"   📅 Años disponibles: {years}")
        print(f"   🏢 Conceptos del balance: {len(financial_data['balance'])}")
        print(f"   💰 Conceptos del P&L: {len(financial_data['income'])}")
        print(f"   💸 Conceptos del flujo: {len(financial_data['cash_flow'])}")
        
        if analysis_type == "formulas":
            print("\\n2️⃣ Procesando con fórmulas Excel...")
            return process_with_formulas(file_path, output_dir, financial_data, extractor)
        else:
            print("\\n2️⃣ Procesando con valores calculados...")
            return process_with_values(file_path, output_dir, financial_data)
            
    except Exception as e:
        print(f"❌ Error: {e}")
        return False


def process_with_formulas(file_path: str, output_dir: str, financial_data: dict, extractor: DataExtractor):
    """Procesa un archivo con fórmulas Excel."""
    from openpyxl import load_workbook
    
    # Cargar workbook original
    wb = load_workbook(file_path)
    
    # Agregar DataFrames para FormulaBuilder
    financial_data["_df_bal"] = extractor.df_bal
    financial_data["_df_pl"] = extractor.df_pl
    financial_data["_df_cfs"] = extractor.df_cfs
    
    # Construir fórmulas
    formula_builder = FormulaBuilder(wb, financial_data)
    formula_blocks = formula_builder.build_all_formulas()
    
    # Crear hoja de análisis
    sheet_name = "Análisis Avanzado (Fórmulas)"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)
    
    # Formatear hoja
    formatter = ExcelFormatter()
    years = financial_data.get("years", [])
    cols_total = 1 + len(years) + 3
    
    header_row = formatter.setup_worksheet_structure(ws, years, sheet_name)
    
    # Escribir fórmulas
    current_row = header_row + 1
    total_formulas = 0
    
    for section_name, formulas in formula_blocks:
        formatter.format_section_header(ws, current_row, cols_total, section_name)
        current_row += 1
        
        for name, ratio_type, func, description in formulas:
            ws.cell(row=current_row, column=1, value=name)
            
            formula_map = func()
            for j, year in enumerate(years, start=2):
                formula_str = formula_map.get(str(year))
                if formula_str:
                    ws.cell(row=current_row, column=j).value = f"={formula_str}"
                    total_formulas += 1
            
            formatter.format_ratio_row(ws, current_row, name, years, ratio_type)
            current_row += 1
    
    # Formateo condicional y tooltip
    data_start_row = header_row + 1
    data_end_row = current_row - 1
    formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
    
    tooltip_start = current_row + 2
    formatter.create_tooltip_section(ws, tooltip_start, formula_blocks, years, cols_total)
    formatter.add_freeze_panes(ws)
    
    # Guardar archivo
    filename = Path(file_path).stem
    output_file = Path(output_dir) / f"{filename}_Analisis_Formulas.xlsx"
    wb.save(str(output_file))
    
    print(f"✅ Archivo procesado exitosamente")
    print(f"   📊 Total de fórmulas creadas: {total_formulas}")
    print(f"   💾 Archivo guardado: {output_file}")
    
    return True


def process_with_values(file_path: str, output_dir: str, financial_data: dict):
    """Procesa un archivo con valores calculados."""
    from openpyxl import Workbook
    
    # Calcular ratios
    calculator = RatioCalculator(financial_data)
    all_ratios = calculator.calculate_all_ratios()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis Financiero (Valores)"
    
    # Formatear hoja
    formatter = ExcelFormatter()
    years = financial_data.get("years", [])
    cols_total = 1 + len(years) + 3
    
    header_row = formatter.setup_worksheet_structure(ws, years)
    
    # Escribir valores
    current_row = header_row + 1
    total_ratios = 0
    
    for section_name, ratios in all_ratios.items():
        formatter.format_section_header(ws, current_row, cols_total, section_name)
        current_row += 1
        
        for ratio_name, ratio_series in ratios.items():
            ws.cell(row=current_row, column=1, value=ratio_name)
            
            for j, year in enumerate(years, start=2):
                # Buscar columna correspondiente al año
                year_col = None
                for col in ratio_series.index:
                    if str(col).startswith(f"{year}-"):
                        year_col = col
                        break
                
                if year_col and year_col in ratio_series.index:
                    value = ratio_series[year_col]
                    if not pd.isna(value):
                        ws.cell(row=current_row, column=j).value = value
                        total_ratios += 1
            
            ratio_type = determine_ratio_type(ratio_name)
            formatter.format_ratio_row(ws, current_row, ratio_name, years, ratio_type)
            current_row += 1
    
    # Formateo condicional
    data_start_row = header_row + 1
    data_end_row = current_row - 1
    formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
    formatter.add_freeze_panes(ws)
    
    # Guardar archivo
    filename = Path(file_path).stem
    output_file = Path(output_dir) / f"{filename}_Analisis_Valores.xlsx"
    wb.save(str(output_file))
    
    print(f"✅ Archivo procesado exitosamente")
    print(f"   📊 Total de ratios calculados: {total_ratios}")
    print(f"   💾 Archivo guardado: {output_file}")
    
    return True


def determine_ratio_type(ratio_name: str) -> str:
    """Determina el tipo de ratio para formateo."""
    ratio_name_lower = ratio_name.lower()
    
    if any(word in ratio_name_lower for word in ["margen", "roe", "roa", "autonomía"]):
        return "pct"
    elif any(word in ratio_name_lower for word in ["días", "período", "ciclo"]):
        return "days"
    elif any(word in ratio_name_lower for word in ["capital", "free cash flow"]):
        return "number"
    else:
        return "ratio"


def process_bulk(input_dir: str, output_dir: str = "./output", analysis_type: str = "formulas", 
                max_workers: int = 4):
    """
    Procesa múltiples archivos Excel de manera masiva.
    
    Args:
        input_dir: Directorio con archivos Excel de entrada
        output_dir: Directorio de salida
        analysis_type: Tipo de análisis ("formulas" o "values")
        max_workers: Número de workers para procesamiento paralelo
    """
    print(f"\\n🚀 Iniciando procesamiento masivo")
    print(f"📁 Directorio de entrada: {input_dir}")
    print(f"📁 Directorio de salida: {output_dir}")
    print(f"📊 Tipo de análisis: {analysis_type}")
    print(f"⚡ Workers paralelos: {max_workers}")
    
    # Inicializar procesador masivo
    processor = BulkProcessor(input_dir, output_dir, max_workers)
    
    # Procesar archivos
    stats = processor.process_bulk(analysis_type)
    
    # Generar reporte resumen
    summary_file = processor.generate_summary_report()
    
    # Mostrar resultados
    print(f"\\n📊 Resumen del procesamiento:")
    print(f"   ✅ Exitosos: {stats['successful']}")
    print(f"   ❌ Fallidos: {stats['failed']}")
    print(f"   📈 Total procesados: {stats['processed']}")
    
    if stats['failed'] > 0:
        print(f"\\n⚠️  Archivos con errores:")
        for error in stats['errors'][:5]:  # Mostrar solo los primeros 5
            print(f"   - {error['file']}: {error.get('error', 'Error desconocido')}")
        if len(stats['errors']) > 5:
            print(f"   ... y {len(stats['errors']) - 5} más")
    
    print(f"\\n📋 Reporte detallado guardado en: {summary_file}")


def main():
    """Función principal del script."""
    parser = argparse.ArgumentParser(
        description="Análisis financiero modularizado para archivos Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:

  # Procesar un archivo único con fórmulas
  python run_analisis_excel.py --mode single --file ./data/demo/FinDataChile_Data_Demo.xlsx

  # Procesar un archivo único con valores calculados
  python run_analisis_excel.py --mode single --file mi_archivo.xlsx --analysis-type values

  # Procesar múltiples archivos de manera masiva
  python run_analisis_excel.py --mode bulk --input-dir ./data/Reports --output-dir ./data/Analisis

  # Procesamiento masivo con más workers
  python run_analisis_excel.py --mode bulk --input-dir ./data/Reports --workers 8
        """
    )
    
    parser.add_argument("--mode", choices=["single", "bulk"], required=True,
                       help="Modo de procesamiento: single (un archivo) o bulk (masivo)")
    
    parser.add_argument("--file", help="Archivo Excel a procesar (modo single)")
    
    parser.add_argument("--input-dir", help="Directorio con archivos Excel (modo bulk)")
    
    parser.add_argument("--output-dir", default="./output",
                       help="Directorio de salida (default: ./output)")
    
    parser.add_argument("--analysis-type", choices=["formulas", "values"], default="formulas",
                       help="Tipo de análisis: formulas (Excel dinámico) o values (calculado)")
    
    parser.add_argument("--workers", type=int, default=4,
                       help="Número de workers paralelos para modo bulk (default: 4)")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("📊 ANÁLISIS FINANCIERO EXCEL MODULARIZADO")
    print("=" * 60)
    
    if args.mode == "single":
        if not args.file:
            print("❌ Error: Se requiere --file para modo single")
            return 1
        
        if not Path(args.file).exists():
            print(f"❌ Error: El archivo {args.file} no existe")
            return 1
        
        success = process_single_file(args.file, args.output_dir, args.analysis_type)
        return 0 if success else 1
    
    elif args.mode == "bulk":
        if not args.input_dir:
            print("❌ Error: Se requiere --input-dir para modo bulk")
            return 1
        
        if not Path(args.input_dir).exists():
            print(f"❌ Error: El directorio {args.input_dir} no existe")
            return 1
        
        process_bulk(args.input_dir, args.output_dir, args.analysis_type, args.workers)
        return 0


if __name__ == "__main__":
    import pandas as pd
    exit(main())
