"""
Excel Formatter Module
=====================

Módulo para formatear y estilizar hojas de análisis financiero en Excel.
Incluye estilos, colores, formateo condicional y estructura visual.
"""

from typing import Dict, List, Optional, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


class ExcelFormatter:
    """
    Formateador de hojas Excel para análisis financiero.
    """
    
    def __init__(self):
        """Inicializa el formateador con estilos predefinidos."""
        self._setup_styles()
    
    def _setup_styles(self):
        """Configura los estilos predefinidos."""
        # Colores de fondo
        self.header_fill = PatternFill("solid", fgColor="0B2447")
        self.subheader_fill = PatternFill("solid", fgColor="19376D")
        
        self.section_fills = {
            "LIQUIDEZ": PatternFill("solid", fgColor="D1E7DD"),
            "SOLVENCIA Y ESTRUCTURA": PatternFill("solid", fgColor="FAD7A0"),
            "RENTABILIDAD": PatternFill("solid", fgColor="F8D7DA"),
            "EFICIENCIA OPERATIVA": PatternFill("solid", fgColor="D6EAF8"),
            "FLUJOS Y ADICIONALES": PatternFill("solid", fgColor="E8DAEF"),
        }
        
        # Fuentes
        self.bold_white = Font(bold=True, color="FFFFFF", size=13)
        self.bold_white_small = Font(bold=True, color="FFFFFF", size=11)
        self.bold_dark = Font(bold=True, color="000000", size=11)
        self.normal = Font(color="000000", size=10)
        
        # Alineaciones
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Bordes
        thin = Side(style="thin", color="DDDDDD")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def setup_worksheet_structure(self, ws, years: List[int], sheet_name: str = "RATIOS & KPIs"):
        """
        Configura la estructura básica de la hoja de análisis.
        
        Args:
            ws: Hoja de trabajo
            years: Lista de años disponibles
            sheet_name: Nombre de la hoja
        """
        num_years = len(years)
        cols_total = 1 + num_years + 2 + 1  # Indicador + años + Último + Promedio + Tendencia
        
        # Configurar ancho de columnas
        for c in range(1, cols_total + 1):
            ws.column_dimensions[get_column_letter(c)].width = 30 if c == 1 else 14
        
        # Título principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_total)
        title_cell = ws.cell(row=1, column=1, 
                           value="Análisis Financiero – Ratios y Evolución (celdas con FÓRMULAS)")
        title_cell.fill = self.header_fill
        title_cell.font = self.bold_white
        title_cell.alignment = self.center
        
        # Subtítulo
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_total)
        subtitle_cell = ws.cell(row=2, column=1,
                              value="Fechas: Balance (AAAA-12), Resultados (AAAA-01), Flujos (AAAA-01). Evolución alineada por AÑO.")
        subtitle_cell.fill = self.subheader_fill
        subtitle_cell.font = self.bold_white_small
        subtitle_cell.alignment = self.center
        
        # Fila de encabezados
        header_row = 4
        headers = ["Indicador"] + [str(y) for y in years] + ["Último", "Promedio", "Tendencia"]
        
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=header)
            cell.font = self.bold_white_small
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border
        
        return header_row
    
    def format_section_header(self, ws, row: int, cols_total: int, section_name: str):
        """
        Formatea una fila de encabezado de sección.
        
        Args:
            ws: Hoja de trabajo
            row: Número de fila
            cols_total: Total de columnas
            section_name: Nombre de la sección
        """
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_total)
        section_cell = ws.cell(row=row, column=1, value=section_name)
        section_cell.font = self.bold_dark
        section_cell.alignment = self.left
        section_cell.fill = self.section_fills.get(section_name, PatternFill("solid", fgColor="EFEFEF"))
        
        # Aplicar bordes a todas las celdas de la fila
        for c in range(1, cols_total + 1):
            ws.cell(row=row, column=c).border = self.border
    
    def format_ratio_row(self, ws, row: int, indicator_name: str, years: List[int], 
                        ratio_type: str = "ratio"):
        """
        Formatea una fila de ratio financiero.
        
        Args:
            ws: Hoja de trabajo
            row: Número de fila
            indicator_name: Nombre del indicador
            years: Lista de años
            ratio_type: Tipo de ratio (ratio, pct, number, days)
        """
        # Celda del indicador
        indicator_cell = ws.cell(row=row, column=1, value=indicator_name)
        indicator_cell.alignment = self.left
        indicator_cell.border = self.border
        
        # Celdas de datos por año
        for j in range(2, 2 + len(years)):
            cell = ws.cell(row=row, column=j)
            cell.alignment = self.center
            cell.border = self.border
            self._apply_number_format(cell, ratio_type)
        
        # Columnas adicionales: Último, Promedio, Tendencia
        last_col_idx = 1 + len(years) + 1
        avg_col_idx = 1 + len(years) + 2
        trend_col_idx = 1 + len(years) + 3
        
        year_start_col = 2
        year_end_col = 1 + len(years)
        rng = f"{get_column_letter(year_start_col)}{row}:{get_column_letter(year_end_col)}{row}"
        
        # Último valor
        last_cell = ws.cell(row=row, column=last_col_idx)
        last_cell.value = f"=LOOKUP(2,1/(--({rng}<>\"\")),{rng})"
        last_cell.alignment = self.center
        last_cell.border = self.border
        self._apply_number_format(last_cell, ratio_type)
        
        # Promedio
        avg_cell = ws.cell(row=row, column=avg_col_idx)
        avg_cell.value = f"=IFERROR(AVERAGE({rng}),\"\")"
        avg_cell.alignment = self.center
        avg_cell.border = self.border
        self._apply_number_format(avg_cell, ratio_type)
        
        # Tendencia
        trend_cell = ws.cell(row=row, column=trend_col_idx)
        trend_cell.value = f"=IFERROR(IF(" \
            f"(LOOKUP(2,1/(--({rng}<>\"\")),{rng}))>" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▲\"," \
            f"IF((LOOKUP(2,1/(--({rng}<>\"\")),{rng}))<" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▼\",\"→\"))," \
            f"\"→\")"
        trend_cell.alignment = self.center
        trend_cell.border = self.border
    
    def _apply_number_format(self, cell, ratio_type: str):
        """
        Aplica formato numérico según el tipo de ratio.
        
        Args:
            cell: Celda a formatear
            ratio_type: Tipo de formato
        """
        if ratio_type == "pct":
            cell.number_format = "0.0%"
        elif ratio_type == "number":
            cell.number_format = "#,##0"
        elif ratio_type == "days":
            cell.number_format = "0"
        else:  # ratio
            cell.number_format = "0.00"
    
    def apply_conditional_formatting(self, ws, data_start_row: int, data_end_row: int, 
                                   years: List[int]):
        """
        Aplica formateo condicional a la hoja.
        
        Args:
            ws: Hoja de trabajo
            data_start_row: Fila de inicio de datos
            data_end_row: Fila de fin de datos
            years: Lista de años
        """
        # Heatmap en columnas de años
        year_start_col = 2
        year_end_col = 1 + len(years)
        year_range = f"{get_column_letter(year_start_col)}{data_start_row}:{get_column_letter(year_end_col)}{data_end_row}"
        
        ws.conditional_formatting.add(
            year_range,
            ColorScaleRule(
                start_type="percentile", start_value=5, start_color="FDE725",
                mid_type="percentile", mid_value=50, mid_color="5DC863",
                end_type="percentile", end_value=95, end_color="2A788E"
            )
        )
        
        # Barras de datos en columnas Último y Promedio
        last_col_letter = get_column_letter(1 + len(years) + 1)
        avg_col_letter = get_column_letter(1 + len(years) + 2)
        
        ws.conditional_formatting.add(
            f"{last_col_letter}{data_start_row}:{last_col_letter}{data_end_row}",
            DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
        )
        
        ws.conditional_formatting.add(
            f"{avg_col_letter}{data_start_row}:{avg_col_letter}{data_end_row}",
            DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
        )
    
    def add_freeze_panes(self, ws):
        """Congela paneles en la posición apropiada."""
        ws.freeze_panes = "B5"
    
    def create_tooltip_section(self, ws, start_row: int, formula_blocks: List[Tuple], 
                             years: List[int], cols_total: int):
        """
        Crea la sección de tooltip con definiciones y fórmulas.
        
        Args:
            ws: Hoja de trabajo
            start_row: Fila de inicio para la sección
            formula_blocks: Bloques de fórmulas organizados por categoría
            years: Lista de años
            cols_total: Total de columnas
        """
        tip_title_row = start_row
        
        # Título de la sección tooltip
        ws.merge_cells(start_row=tip_title_row, start_column=1, 
                      end_row=tip_title_row, end_column=cols_total)
        tip_title = ws.cell(row=tip_title_row, column=1, 
                           value="Definición y Fórmula (tooltip)")
        tip_title.font = self.bold_dark
        tip_title.alignment = self.left
        tip_title.fill = PatternFill("solid", fgColor="EFEFEF")
        
        # Encabezados de la sección tooltip
        tip_header_row = tip_title_row + 1
        headers = ["Indicador", "Fórmula (texto)", "Ejemplo de Fórmula Excel (último año)"]
        
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=tip_header_row, column=i, value=header)
            cell.font = self.bold_white_small
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border
        
        # Contenido de la sección tooltip
        last_year = years[-1] if years else None
        tip_row = tip_header_row + 1
        
        for section_name, items in formula_blocks:
            # Encabezado de sección en tooltip
            section_cell = ws.cell(row=tip_row, column=1, value=section_name)
            section_cell.font = self.bold_dark
            section_cell.fill = self.section_fills.get(section_name, 
                                                     PatternFill("solid", fgColor="EFEFEF"))
            
            for c in range(1, 4):
                ws.cell(row=tip_row, column=c).border = self.border
            tip_row += 1
            
            # Items de la sección
            for name, kind, func, text_formula in items:
                # Nombre del indicador
                ws.cell(row=tip_row, column=1, value=name).alignment = self.left
                ws.cell(row=tip_row, column=1).border = self.border
                
                # Fórmula en texto
                ws.cell(row=tip_row, column=2, value=text_formula).alignment = self.left
                ws.cell(row=tip_row, column=2).border = self.border
                
                # Ejemplo de fórmula Excel
                example_formula = self._get_example_formula(func, last_year)
                ws.cell(row=tip_row, column=3, value=example_formula).alignment = self.left
                ws.cell(row=tip_row, column=3).border = self.border
                
                tip_row += 1
        
        return tip_row
    
    def _get_example_formula(self, func, year: Optional[int]) -> str:
        """
        Obtiene una fórmula de ejemplo para el último año.
        
        Args:
            func: Función que genera las fórmulas
            year: Año para el ejemplo
            
        Returns:
            Fórmula de ejemplo o string vacío
        """
        if year is None:
            return ""
        
        try:
            formula_map = func()
            formula = formula_map.get(str(year))
            if formula:
                return f"={formula}"
        except:
            pass
        
        return ""
