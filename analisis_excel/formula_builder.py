"""
Formula Builder Module
=====================

Módulo para construir fórmulas Excel que referencian directamente 
las celdas de los estados financieros originales.
"""

import re
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class FormulaBuilder:
    """
    Constructor de fórmulas Excel para análisis financiero.
    """
    
    def __init__(self, workbook: Workbook, financial_data: Dict[str, Any]):
        """
        Inicializa el constructor de fórmulas.
        
        Args:
            workbook: Libro de Excel abierto
            financial_data: Datos financieros extraídos
        """
        self.wb = workbook
        self.financial_data = financial_data
        self.years = financial_data.get("years", [])
        
        # Obtener referencias a las hojas
        self.sh_bal = workbook["Balance General"]
        self.sh_pl = workbook["Estado Resultados (Función)"]
        self.sh_cfs = workbook["Flujo Efectivo"]
        
        # Mapear conceptos a filas
        self.rows_bal = self._map_balance_rows()
        self.rows_pl = self._map_income_rows()
        self.rows_cfs = self._map_cash_flow_rows()
        
        # Fila de encabezados (siempre row 1)
        self.HEADER_ROW = 1
    
    def _find_row_in_sheet(self, sheet, df, concept_name: str) -> Optional[int]:
        """
        Encuentra la fila de un concepto en una hoja específica.
        
        Args:
            sheet: Hoja de Excel
            df: DataFrame correspondiente
            concept_name: Nombre del concepto a buscar
            
        Returns:
            Número de fila (1-based) o None si no se encuentra
        """
        # Búsqueda exacta
        mask = df["Concepto"].astype(str).str.strip().str.lower() == concept_name.strip().lower()
        if mask.any():
            idx = df.index[mask][0]
            return int(idx) + 2  # +2 porque df es 0-indexed y Excel tiene header en row 1
        
        # Búsqueda por contiene
        mask = df["Concepto"].astype(str).str.contains(re.escape(concept_name), case=False, na=False)
        if mask.any():
            idx = df.index[mask][0]
            return int(idx) + 2
        
        return None
    
    def _map_balance_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del balance a números de fila."""
        concepts = {
            # Activos básicos
            "AC": "Activos corrientes totales",
            "AT": "Total de activos",
            "ANC": "Total de activos no corrientes",
            "PPE": "Propiedades, planta y equipo",
            "Efec": "Efectivo y equivalentes al efectivo",
            "Inv": "Inventarios corrientes",
            "InvNC": "Inventarios, no corrientes",
            "CxC": "Deudores comerciales y otras cuentas por cobrar corrientes",
            "CxCNC": "Cuentas por cobrar no corrientes",
            "Intang": "Activos intangibles distintos de la plusvalía",
            "Plusvalia": "Plusvalía",
            "ActivDifImp": "Activos por impuestos diferidos",
            "PropInv": "Propiedad de inversión",
            "ActivDerUso": "Activos por derecho de uso",
            "InvAsoc": "Inversiones contabilizadas utilizando el método de la participación",
            
            # Pasivos básicos
            "PC": "Pasivos corrientes totales",
            "PT": "Total de pasivos",
            "PNC": "Total de pasivos no corrientes",
            "CxP": "Cuentas por pagar comerciales y otras cuentas por pagar",
            "CxPNC": "Cuentas por pagar no corrientes",
            "DeudaFinCorr": "Otros pasivos financieros corrientes",
            "DeudaFinNC": "Otros pasivos financieros no corrientes",
            "ArrCorr": "Pasivos por arrendamientos corrientes",
            "ArrNC": "Pasivos por arrendamientos no corrientes",
            "PasivDifImp": "Pasivo por impuestos diferidos",
            "ProvCorr": "Otras provisiones a corto plazo",
            "ProvNC": "Otras provisiones a largo plazo",
            "BenefEmpl": "Provisiones corrientes por beneficios a los empleados",
            "BenefEmplNC": "Provisiones no corrientes por beneficios a los empleados",
            
            # Patrimonio
            "Patr": "Patrimonio atribuible a los propietarios de la controladora",
            "PatrTotal": "Patrimonio total",
            "CapitalEmit": "Capital emitido y pagado",
            "GanAcum": "Ganancias (pérdidas) acumuladas",
            "PrimaEmis": "Prima de emisión",
            "AccPropias": "Acciones propias en cartera",
            "OtraPartPatr": "Otras participaciones en el patrimonio",
            "OtraReserv": "Otras reservas",
            "PartNoControl": "Participaciones no controladoras"
        }
        
        # Necesitamos el DataFrame del balance
        df_bal = self.financial_data.get("_df_bal")
        if df_bal is None:
            # Si no está disponible, crear un mapeo vacío
            return {key: None for key in concepts.keys()}
        
        return {key: self._find_row_in_sheet(self.sh_bal, df_bal, concept)
                for key, concept in concepts.items()}
    
    def _map_income_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del estado de resultados a números de fila."""
        concepts = {
            "Ventas": "Ingresos de actividades ordinarias",
            "COGS": "Costo de ventas",
            "Bruta": "Ganancia bruta",
            "EBIT": "Ganancias (pérdidas) de actividades operacionales",
            "Neta": "Ganancia (pérdida)",
            "NetaControl": "Ganancia (pérdida), atribuible a los propietarios de la controladora",
            "Interes": "Costos financieros",
            "IngFinanc": "Ingresos financieros",
            "Dep": "Depreciación",
            "Amort": "Amortización",
            "OtrosIng": "Otros ingresos",
            "CostDistrib": "Costos de distribución",
            "GastAdmin": "Gastos de administración",
            "OtrosGast": "Otros gastos, por función",
            "OtraGanPerd": "Otras ganancias (pérdidas)",
            "AntesImp": "Ganancia (pérdida), antes de impuestos",
            "ImpGanan": "Gasto por impuestos a las ganancias",
            "OperCont": "Ganancia (pérdida) procedente de operaciones continuadas",
            "OperDisc": "Ganancia (pérdida) procedente de operaciones discontinuadas",
            "DeterioBanco": "Deterioro de valor de ganancias y reversión de pérdidas por deterioro de valor (pérdidas por deterioro de valor) determinado de acuerdo con la NIIF 9",
            "PartAsoc": "Participación en las ganancias (pérdidas) de asociadas y negocios conjuntos que se contabilicen utilizando el método de la participación",
            "GanCambio": "Ganancias (pérdidas) de cambio en moneda extranjera",
            "ResUniReaj": "Resultados por unidades de reajuste",
            "GanBasAcc": "Ganancia (pérdida) por acción básica",
            "GanDilAcc": "Ganancias (pérdida) diluida por acción"
        }
        
        df_pl = self.financial_data.get("_df_pl")
        if df_pl is None:
            return {key: None for key in concepts.keys()}
        
        return {key: self._find_row_in_sheet(self.sh_pl, df_pl, concept)
                for key, concept in concepts.items()}
    
    def _map_cash_flow_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del flujo de efectivo a números de fila."""
        concepts = {
            "CFO": "Flujos de efectivo netos procedentes de (utilizados en) actividades de operación",
            "CFI": "Flujos de efectivo netos procedentes de (utilizados en) actividades de inversión",
            "CFF": "Flujos de efectivo netos procedentes de (utilizados en) actividades de financiación",
            "CapexBuy": "Compras de propiedades, planta y equipo",
            "CapexSale": "Importes procedentes de la venta de propiedades, planta y equipo",
            "IntangBuy": "Compras de activos intangibles",
            "IntangSale": "Importes procedentes de ventas de activos intangibles",
            "DivPag": "Dividendos pagados",
            "DivRec": "Dividendos recibidos",
            "IntPag": "Intereses pagados",
            "IntRec": "Intereses recibidos",
            "ImpPag": "Impuestos a las ganancias pagados (reembolsados)",
            "EmisPrest": "Importes procedentes de préstamos",
            "ReembPrest": "Reembolsos de préstamos",
            "EmisAcc": "Importes procedentes de la emisión de acciones",
            "CompAcc": "Pagos por adquirir o rescatar las acciones de la entidad",
            "PagArrend": "Pagos de pasivos por arrendamientos",
            "CobrosVenta": "Cobros procedentes de las ventas de bienes y prestación de servicios",
            "PagosProveed": "Pagos a proveedores por el suministro de bienes y servicios",
            "PagosEmpl": "Pagos a y por cuenta de los empleados",
            "EfecInic": "Efectivo y equivalentes al efectivo al principio del periodo",
            "EfecFinal": "Efectivo y equivalentes al efectivo al final del periodo",
            "VarEfec": "Incremento (disminución) neto de efectivo y equivalentes al efectivo"
        }
        
        df_cfs = self.financial_data.get("_df_cfs")
        if df_cfs is None:
            return {key: None for key in concepts.keys()}
        
        return {key: self._find_row_in_sheet(self.sh_cfs, df_cfs, concept)
                for key, concept in concepts.items()}
    
    def find_year_column(self, sheet, year: int) -> Optional[str]:
        """
        Encuentra la letra de columna para un año específico.
        
        Args:
            sheet: Hoja de Excel
            year: Año a buscar
            
        Returns:
            Letra de columna o None si no se encuentra
        """
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=self.HEADER_ROW, column=col).value
            if isinstance(val, str) and val.startswith(f"{year}-"):
                return get_column_letter(col)
        return None
    
    def create_cell_reference(self, sheet_name: str, col_letter: Optional[str], 
                            row_num: Optional[int]) -> Optional[str]:
        """
        Crea una referencia de celda Excel.
        
        Args:
            sheet_name: Nombre de la hoja
            col_letter: Letra de columna
            row_num: Número de fila
            
        Returns:
            Referencia de celda o None si falta información
        """
        if col_letter is None or row_num is None:
            return None
        return f"'{sheet_name}'!{col_letter}{row_num}"
    
    def create_average_reference(self, sheet, row: Optional[int], year: int) -> Optional[str]:
        """
        Crea referencia promedio entre año actual y anterior para items del balance.
        
        Args:
            sheet: Hoja de Excel
            row: Número de fila
            year: Año actual
            
        Returns:
            Fórmula de promedio o referencia simple
        """
        if row is None:
            return None
        
        col_now = self.find_year_column(sheet, year)
        col_prev = self.find_year_column(sheet, year - 1)
        
        now_ref = self.create_cell_reference(sheet.title, col_now, row) if col_now else None
        prev_ref = self.create_cell_reference(sheet.title, col_prev, row) if col_prev else None
        
        if now_ref and prev_ref:
            return f"AVERAGE({now_ref},{prev_ref})"
        elif now_ref:
            return now_ref
        elif prev_ref:
            return prev_ref
        
        return None
    
    def build_liquidity_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """
        Construye fórmulas para ratios de liquidez.
        
        Returns:
            Lista de tuplas (nombre, tipo, función_fórmula, descripción)
        """
        formulas = []
        
        # Liquidez Corriente = AC/PC
        def f_liq_corr():
            m = {}
            for y in self.years:
                ac = self.create_cell_reference(
                    self.sh_bal.title, 
                    self.find_year_column(self.sh_bal, y), 
                    self.rows_bal["AC"]
                )
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                if ac and pc:
                    m[str(y)] = f"IFERROR({ac}/{pc},\"\")"
            return m
        
        formulas.append(("Liquidez Corriente", "ratio", f_liq_corr, 
                        "Activo Corriente / Pasivo Corriente"))
        
        # Prueba Ácida = (AC - Inventarios) / PC
        def f_prueba_acida():
            m = {}
            for y in self.years:
                ac = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AC"]
                )
                inv = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Inv"]
                )
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                if ac and pc:
                    term_inv = inv if inv else "0"
                    m[str(y)] = f"IFERROR(({ac}-{term_inv})/{pc},\"\")"
            return m
        
        formulas.append(("Prueba Ácida", "ratio", f_prueba_acida,
                        "(Activo Corriente - Inventarios) / Pasivo Corriente"))
        
        # Cash Ratio = Efectivo / PC
        def f_cash_ratio():
            m = {}
            for y in self.years:
                ef = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Efec"]
                )
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                if ef and pc:
                    m[str(y)] = f"IFERROR({ef}/{pc},\"\")"
            return m
        
        formulas.append(("Cash Ratio", "ratio", f_cash_ratio,
                        "Efectivo y Equivalentes / Pasivo Corriente"))
        
        # Capital de Trabajo = AC - PC
        def f_cap_trabajo():
            m = {}
            for y in self.years:
                ac = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AC"]
                )
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                if ac and pc:
                    m[str(y)] = f"IFERROR({ac}-{pc},\"\")"
            return m
        
        formulas.append(("Capital de Trabajo", "number", f_cap_trabajo,
                        "Activo Corriente - Pasivo Corriente"))
        
        return formulas
    
    def build_solvency_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para ratios de solvencia."""
        formulas = []
        
        # Endeudamiento (D/E)
        def f_de():
            m = {}
            for y in self.years:
                pt = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PT"]
                )
                patr = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Patr"]
                )
                if pt and patr:
                    m[str(y)] = f"IFERROR({pt}/{patr},\"\")"
            return m
        
        formulas.append(("Endeudamiento (D/E)", "ratio", f_de,
                        "Deuda Total / Patrimonio"))
        
        # Apalancamiento (D/A)
        def f_da():
            m = {}
            for y in self.years:
                pt = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PT"]
                )
                at = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AT"]
                )
                if pt and at:
                    m[str(y)] = f"IFERROR({pt}/{at},\"\")"
            return m
        
        formulas.append(("Apalancamiento (D/A)", "ratio", f_da,
                        "Deuda Total / Activos Totales"))
        
        # Cobertura de Intereses
        def f_cover():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                interes = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Interes"]
                )
                if ebit and interes:
                    m[str(y)] = f"IFERROR({ebit}/ABS({interes}),\"\")"
            return m
        
        formulas.append(("Cobertura de Intereses", "ratio", f_cover,
                        "EBIT / |Gastos por Intereses|"))
        
        # Deuda / EBITDA
        def f_deuda_ebitda():
            m = {}
            for y in self.years:
                pt = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PT"]
                )
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                dep = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Dep"]
                ) if self.rows_pl["Dep"] else None
                amort = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Amort"]
                ) if self.rows_pl["Amort"] else None
                
                if pt and ebit:
                    if dep and amort:
                        ebitda = f"({ebit}+IFERROR({dep},0)+IFERROR({amort},0))"
                    else:
                        ebitda = f"({ebit})"
                    m[str(y)] = f"IFERROR({pt}/{ebitda},\"\")"
            return m
        
        formulas.append(("Deuda / EBITDA", "ratio", f_deuda_ebitda,
                        "Deuda Total / (EBIT + Depreciación + Amortización)"))
        
        # Autonomía Financiera
        def f_autonomia():
            m = {}
            for y in self.years:
                patr = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Patr"]
                )
                at = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AT"]
                )
                if patr and at:
                    m[str(y)] = f"IFERROR({patr}/{at},\"\")"
            return m
        
        formulas.append(("Autonomía Financiera", "pct", f_autonomia,
                        "Patrimonio / Activo Total"))
        
        return formulas
    
    def build_profitability_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para ratios de rentabilidad."""
        formulas = []
        
        # Margen Bruto
        def f_margen_bruto():
            m = {}
            for y in self.years:
                gb = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Bruta"]
                )
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                if gb and ven:
                    m[str(y)] = f"IFERROR({gb}/{ven},\"\")"
            return m
        
        formulas.append(("Margen Bruto", "pct", f_margen_bruto,
                        "Utilidad Bruta / Ventas"))
        
        # Margen Operativo (EBIT)
        def f_margen_operativo():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                if ebit and ven:
                    m[str(y)] = f"IFERROR({ebit}/{ven},\"\")"
            return m
        
        formulas.append(("Margen Operativo (EBIT)", "pct", f_margen_operativo,
                        "EBIT / Ventas"))
        
        # Margen EBITDA
        def f_margen_ebitda():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                dep = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Dep"]
                ) if self.rows_pl["Dep"] else None
                amort = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Amort"]
                ) if self.rows_pl["Amort"] else None
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                
                if ebit and ven:
                    ebitda = f"({ebit}+IFERROR({dep},0)+IFERROR({amort},0))" if (dep or amort) else f"({ebit})"
                    m[str(y)] = f"IFERROR({ebitda}/{ven},\"\")"
            return m
        
        formulas.append(("Margen EBITDA", "pct", f_margen_ebitda,
                        "EBITDA / Ventas"))
        
        # Margen Neto
        def f_margen_neto():
            m = {}
            for y in self.years:
                net = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Neta"]
                )
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                if net and ven:
                    m[str(y)] = f"IFERROR({net}/{ven},\"\")"
            return m
        
        formulas.append(("Margen Neto", "pct", f_margen_neto,
                        "Utilidad Neta / Ventas"))
        
        # ROE usando promedio del patrimonio
        def f_roe():
            m = {}
            for y in self.years:
                net = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Neta"]
                )
                patr_avg = self.create_average_reference(self.sh_bal, self.rows_bal["Patr"], y)
                if net and patr_avg:
                    m[str(y)] = f"IFERROR({net}/{patr_avg},\"\")"
            return m
        
        formulas.append(("ROE", "pct", f_roe,
                        "Utilidad Neta / Patrimonio Promedio"))
        
        # ROA usando promedio de activos totales
        def f_roa():
            m = {}
            for y in self.years:
                net = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Neta"]
                )
                at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                if net and at_avg:
                    m[str(y)] = f"IFERROR({net}/{at_avg},\"\")"
            return m
        
        formulas.append(("ROA", "pct", f_roa,
                        "Utilidad Neta / Activos Totales Promedio"))
        
        return formulas
    
    def build_efficiency_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para ratios de eficiencia operativa."""
        formulas = []
        
        # Rotación de Activos
        def f_rot_act():
            m = {}
            for y in self.years:
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                if ven and at_avg:
                    m[str(y)] = f"IFERROR({ven}/{at_avg},\"\")"
            return m
        
        formulas.append(("Rotación de Activos", "ratio", f_rot_act,
                        "Ventas / Activos Promedio"))
        
        # Rotación de Activos Fijos
        def f_rot_act_fijos():
            m = {}
            for y in self.years:
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                ppe_avg = self.create_average_reference(self.sh_bal, self.rows_bal["PPE"], y)
                if ven and ppe_avg:
                    m[str(y)] = f"IFERROR({ven}/{ppe_avg},\"\")"
            return m
        
        formulas.append(("Rotación de Activos Fijos", "ratio", f_rot_act_fijos,
                        "Ventas / PPE Promedio"))
        
        # Rotación de Inventarios y Días de Inventario
        def f_rot_inv_y_dias():
            m_rot = {}
            m_dias = {}
            for y in self.years:
                cogs = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["COGS"]
                )
                inv_now = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Inv"]
                )
                inv_prev = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y-1),
                    self.rows_bal["Inv"]
                )
                
                inv_avg = None
                if inv_now and inv_prev:
                    inv_avg = f"AVERAGE({inv_now},{inv_prev})"
                elif inv_now:
                    inv_avg = inv_now
                elif inv_prev:
                    inv_avg = inv_prev
                
                if cogs and inv_avg:
                    m_rot[str(y)] = f"IFERROR({cogs}/{inv_avg},\"\")"
                    m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
            return m_rot, m_dias
        
        rot_inv, dias_inv = f_rot_inv_y_dias()
        
        formulas.append(("Rotación de Inventarios", "ratio", lambda: rot_inv,
                        "Costo de Ventas / Inventario Promedio"))
        formulas.append(("Días de Inventario", "days", lambda: dias_inv,
                        "365 / Rotación de Inventarios"))
        
        # Rotación de Cuentas por Cobrar y Período Promedio de Cobro
        def f_rot_cxc_y_dias():
            m_rot = {}
            m_dias = {}
            for y in self.years:
                ven = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                cxc_avg = self.create_average_reference(self.sh_bal, self.rows_bal["CxC"], y)
                if ven and cxc_avg:
                    m_rot[str(y)] = f"IFERROR({ven}/{cxc_avg},\"\")"
                    m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
            return m_rot, m_dias
        
        rot_cxc, dias_cobro = f_rot_cxc_y_dias()
        
        formulas.append(("Rotación de Cuentas por Cobrar", "ratio", lambda: rot_cxc,
                        "Ventas / Cuentas por Cobrar Promedio"))
        formulas.append(("Período Promedio de Cobro", "days", lambda: dias_cobro,
                        "365 / Rotación de CxC"))
        
        # Rotación de Cuentas por Pagar y Período Promedio de Pago
        def f_rot_cxp_y_dias():
            m_rot = {}
            m_dias = {}
            for y in self.years:
                cogs = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["COGS"]
                )
                inv_now = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Inv"]
                )
                inv_prev = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y-1),
                    self.rows_bal["Inv"]
                )
                
                comp = None
                if cogs and inv_now and inv_prev:
                    comp = f"({cogs}+({inv_now}-{inv_prev}))"
                
                # Promedio de cuentas por pagar
                cxp_avg = self.create_average_reference(self.sh_bal, self.rows_bal["CxP"], y)
                if comp and cxp_avg:
                    m_rot[str(y)] = f"IFERROR({comp}/{cxp_avg},\"\")"
                    m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
            return m_rot, m_dias
        
        rot_cxp, dias_pago = f_rot_cxp_y_dias()
        
        formulas.append(("Rotación de Cuentas por Pagar", "ratio", lambda: rot_cxp,
                        "Compras (≈ COGS + ΔInventario) / Cuentas por Pagar Promedio"))
        formulas.append(("Período Promedio de Pago", "days", lambda: dias_pago,
                        "365 / Rotación de CxP"))
        
        # Ciclo de Conversión de Efectivo
        def f_cce():
            m = {}
            for y in self.years:
                # CCE = Días Inv + Días CxC - Días CxP
                di = dias_inv.get(str(y))
                dc = dias_cobro.get(str(y))
                dp = dias_pago.get(str(y))
                if di and dc and dp:
                    m[str(y)] = f"IFERROR({di}+{dc}-{dp},\"\")"
            return m
        
        formulas.append(("Ciclo de Conversión de Efectivo", "days", f_cce,
                        "Días Inventario + Días CxC - Días CxP"))
        
        return formulas
    
    def build_cash_flow_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para ratios de flujos y adicionales."""
        formulas = []
        
        # Conversión de caja (CFO/Utilidad Neta)
        def f_conv_caja():
            m = {}
            for y in self.years:
                cfo = self.create_cell_reference(
                    self.sh_cfs.title,
                    self.find_year_column(self.sh_cfs, y),
                    self.rows_cfs["CFO"]
                )
                net = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Neta"]
                )
                if cfo and net:
                    m[str(y)] = f"IFERROR({cfo}/{net},\"\")"
            return m
        
        formulas.append(("Conversión de caja (CFO/Utilidad Neta)", "ratio", f_conv_caja,
                        "Flujo Operativo / Utilidad Neta"))
        
        # Free Cash Flow (CFO - CAPEX)
        def f_fcf():
            m = {}
            for y in self.years:
                cfo = self.create_cell_reference(
                    self.sh_cfs.title,
                    self.find_year_column(self.sh_cfs, y),
                    self.rows_cfs["CFO"]
                )
                capex = self.create_cell_reference(
                    self.sh_cfs.title,
                    self.find_year_column(self.sh_cfs, y),
                    self.rows_cfs["CapexBuy"]
                ) if self.rows_cfs["CapexBuy"] else None
                
                if cfo:
                    if capex:
                        m[str(y)] = f"IFERROR({cfo}-ABS({capex}),\"\")"
                    else:
                        m[str(y)] = f"IFERROR({cfo},\"\")"
            return m
        
        formulas.append(("Free Cash Flow (CFO - CAPEX)", "number", f_fcf,
                        "CFO - CAPEX (Compras PPE)"))
        
        # AC / AT
        def f_ac_at():
            m = {}
            for y in self.years:
                ac = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AC"]
                )
                at = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AT"]
                )
                if ac and at:
                    m[str(y)] = f"IFERROR({ac}/{at},\"\")"
            return m
        
        formulas.append(("AC / AT", "pct", f_ac_at,
                        "Activo Corriente / Activo Total"))
        
        # PC / PT
        def f_pc_pt():
            m = {}
            for y in self.years:
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                pt = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PT"]
                )
                if pc and pt:
                    m[str(y)] = f"IFERROR({pc}/{pt},\"\")"
            return m
        
        formulas.append(("PC / PT", "pct", f_pc_pt,
                        "Pasivo Corriente / Pasivo Total"))
        
        return formulas
    
    def build_value_creation_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para indicadores de creación de valor."""
        formulas = []
        
        # ROA (ya existe en rentabilidad, pero lo incluimos aquí también)
        def f_roa_value():
            m = {}
            for y in self.years:
                net = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Neta"]
                )
                at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                if net and at_avg:
                    m[str(y)] = f"IFERROR({net}/{at_avg},\"\")"
            return m
        
        formulas.append(("ROA", "pct", f_roa_value,
                        "Utilidad Neta / Activos Totales Promedio"))
        
        # ROIC = NOPAT / Capital Invertido
        # NOPAT ≈ EBIT * (1 - Tax Rate)
        # Capital Invertido ≈ AT - Pasivos Sin Costo (efectivo, cuentas por pagar operativas)
        def f_roic():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                antes_imp = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["AntesImp"]
                )
                imp_ganan = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["ImpGanan"]
                )
                at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                efec_avg = self.create_average_reference(self.sh_bal, self.rows_bal["Efec"], y)
                cxp_avg = self.create_average_reference(self.sh_bal, self.rows_bal["CxP"], y)
                
                if ebit and antes_imp and imp_ganan and at_avg:
                    # Tax rate = Impuestos / EBT
                    tax_rate = f"IFERROR(ABS({imp_ganan})/{antes_imp},0.25)"
                    nopat = f"({ebit}*(1-{tax_rate}))"
                    
                    # Capital Invertido (simplificado)
                    cap_inv = f"({at_avg}"
                    if efec_avg:
                        cap_inv += f"-{efec_avg}"
                    if cxp_avg:
                        cap_inv += f"-{cxp_avg}"
                    cap_inv += ")"
                    
                    m[str(y)] = f"IFERROR({nopat}/{cap_inv},\"\")"
            return m
        
        formulas.append(("ROIC", "pct", f_roic,
                        "NOPAT / Capital Invertido"))
        
        # EVA = NOPAT - (Capital Invertido * WACC)
        # Asumimos WACC del 10% como proxy
        def f_eva():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                antes_imp = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["AntesImp"]
                )
                imp_ganan = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["ImpGanan"]
                )
                at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                efec_avg = self.create_average_reference(self.sh_bal, self.rows_bal["Efec"], y)
                cxp_avg = self.create_average_reference(self.sh_bal, self.rows_bal["CxP"], y)
                
                if ebit and antes_imp and imp_ganan and at_avg:
                    # Tax rate
                    tax_rate = f"IFERROR(ABS({imp_ganan})/{antes_imp},0.25)"
                    nopat = f"({ebit}*(1-{tax_rate}))"
                    
                    # Capital Invertido
                    cap_inv = f"({at_avg}"
                    if efec_avg:
                        cap_inv += f"-{efec_avg}"
                    if cxp_avg:
                        cap_inv += f"-{cxp_avg}"
                    cap_inv += ")"
                    
                    # EVA = NOPAT - (Cap_Inv * 10%)
                    m[str(y)] = f"IFERROR({nopat}-({cap_inv}*0.10),\"\")"
            return m
        
        formulas.append(("EVA (WACC=10%)", "number", f_eva,
                        "NOPAT - (Capital Invertido × WACC estimado)"))
        
        # Spread = ROIC - WACC (asumiendo WACC 10%)
        def f_spread():
            m = {}
            roic_map = f_roic()
            for y in self.years:
                roic_formula = roic_map.get(str(y))
                if roic_formula:
                    # Eliminar IFERROR para poder hacer cálculo
                    roic_clean = roic_formula.replace('IFERROR(', '').replace(',"")', '')
                    m[str(y)] = f"IFERROR(({roic_clean})-0.10,\"\")"
            return m
        
        formulas.append(("Spread (ROIC - WACC)", "pct", f_spread,
                        "ROIC - WACC estimado (10%)"))
        
        # MVA - requiere valor de mercado (no disponible en estados financieros)
        def f_mva_nota():
            m = {}
            for y in self.years:
                m[str(y)] = "\"Requiere Valor de Mercado\""
            return m
        
        formulas.append(("MVA (Nota)", "text", f_mva_nota,
                        "Valor de Mercado - Capital Invertido (requiere datos de mercado)"))
        
        return formulas
    
    def build_coverage_risk_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para indicadores de cobertura y riesgo."""
        formulas = []
        
        # Cobertura de Servicio de Deuda = EBITDA / (Intereses + Amortización de Deuda)
        def f_debt_service():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                dep = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Dep"]
                ) if self.rows_pl["Dep"] else None
                amort = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Amort"]
                ) if self.rows_pl["Amort"] else None
                interes = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Interes"]
                )
                reemb_prest = self.create_cell_reference(
                    self.sh_cfs.title,
                    self.find_year_column(self.sh_cfs, y),
                    self.rows_cfs["ReembPrest"]
                ) if self.rows_cfs["ReembPrest"] else None
                
                if ebit and interes:
                    # EBITDA
                    ebitda = f"({ebit}"
                    if dep:
                        ebitda += f"+IFERROR({dep},0)"
                    if amort:
                        ebitda += f"+IFERROR({amort},0)"
                    ebitda += ")"
                    
                    # Servicio de deuda = intereses + amortización
                    service = f"(ABS({interes})"
                    if reemb_prest:
                        service += f"+IFERROR(ABS({reemb_prest}),0)"
                    service += ")"
                    
                    m[str(y)] = f"IFERROR({ebitda}/{service},\"\")"
            return m
        
        formulas.append(("Cobertura Servicio Deuda", "ratio", f_debt_service,
                        "EBITDA / (Intereses + Amortización Deuda)"))
        
        # Cobertura de Gastos Fijos = (EBIT + Gastos Fijos) / Gastos Fijos
        # Gastos Fijos ≈ Gastos Admin + Costos Distribución + Gastos Financieros
        def f_fixed_coverage():
            m = {}
            for y in self.years:
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                gast_admin = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["GastAdmin"]
                ) if self.rows_pl["GastAdmin"] else None
                cost_distr = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["CostDistrib"]
                ) if self.rows_pl["CostDistrib"] else None
                interes = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Interes"]
                )
                
                if ebit:
                    # Gastos fijos estimados
                    gastos_fijos = "("
                    componentes = []
                    if gast_admin:
                        componentes.append(f"IFERROR(ABS({gast_admin}),0)")
                    if cost_distr:
                        componentes.append(f"IFERROR(ABS({cost_distr}),0)")
                    if interes:
                        componentes.append(f"IFERROR(ABS({interes}),0)")
                    
                    if componentes:
                        gastos_fijos += "+".join(componentes) + ")"
                        
                        m[str(y)] = f"IFERROR(({ebit}+{gastos_fijos})/{gastos_fijos},\"\")"
            return m
        
        formulas.append(("Cobertura Gastos Fijos", "ratio", f_fixed_coverage,
                        "(EBIT + Gastos Fijos) / Gastos Fijos"))
        
        # Altman Z-Score = 1.2*A + 1.4*B + 3.3*C + 0.6*D + 1.0*E
        # A = Capital Trabajo / AT
        # B = Utilidades Retenidas / AT  
        # C = EBIT / AT
        # D = Valor Mercado Patrimonio / PT (usaremos valor contable)
        # E = Ventas / AT
        def f_altman_zscore():
            m = {}
            for y in self.years:
                ac = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AC"]
                )
                pc = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PC"]
                )
                gan_acum = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["GanAcum"]
                ) if self.rows_bal["GanAcum"] else None
                ebit = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["EBIT"]
                )
                patr = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["Patr"]
                )
                pt = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["PT"]
                )
                ventas = self.create_cell_reference(
                    self.sh_pl.title,
                    self.find_year_column(self.sh_pl, y),
                    self.rows_pl["Ventas"]
                )
                at = self.create_cell_reference(
                    self.sh_bal.title,
                    self.find_year_column(self.sh_bal, y),
                    self.rows_bal["AT"]
                )
                
                if ac and pc and ebit and patr and pt and ventas and at:
                    # A = (AC - PC) / AT
                    comp_a = f"1.2*(({ac}-{pc})/{at})"
                    
                    # B = Utilidades Retenidas / AT
                    comp_b = f"1.4*({gan_acum}/{at})" if gan_acum else "0"
                    
                    # C = EBIT / AT
                    comp_c = f"3.3*({ebit}/{at})"
                    
                    # D = Patrimonio / Pasivos (valor contable)
                    comp_d = f"0.6*({patr}/{pt})"
                    
                    # E = Ventas / AT
                    comp_e = f"1.0*({ventas}/{at})"
                    
                    z_score = f"IFERROR({comp_a}+{comp_b}+{comp_c}+{comp_d}+{comp_e},\"\")"
                    m[str(y)] = z_score
            return m
        
        formulas.append(("Altman Z-Score", "ratio", f_altman_zscore,
                        "1.2×(CT/AT) + 1.4×(RE/AT) + 3.3×(EBIT/AT) + 0.6×(E/D) + 1.0×(S/AT)"))
        
        # Beta - requiere datos de mercado
        def f_beta_nota():
            m = {}
            for y in self.years:
                m[str(y)] = "\"Requiere datos de mercado\""
            return m
        
        formulas.append(("Beta (Nota)", "text", f_beta_nota,
                        "Requiere datos históricos de precios de mercado"))
        
        return formulas
    
    def build_all_formulas(self) -> List[Tuple[str, List]]:
        """
        Construye todas las fórmulas organizadas por categoría.
        
        Returns:
            Lista de tuplas (categoría, lista_fórmulas)
        """
        return [
            ("LIQUIDEZ", self.build_liquidity_formulas()),
            ("SOLVENCIA Y ESTRUCTURA", self.build_solvency_formulas()),
            ("RENTABILIDAD", self.build_profitability_formulas()),
            ("EFICIENCIA OPERATIVA", self.build_efficiency_formulas()),
            ("FLUJOS Y ADICIONALES", self.build_cash_flow_formulas()),
            ("CREACIÓN DE VALOR", self.build_value_creation_formulas()),
            ("COBERTURA Y RIESGO", self.build_coverage_risk_formulas()),
        ]
