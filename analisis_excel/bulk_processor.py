"""
Bulk Processor Module
====================

Módulo para procesamiento masivo de archivos Excel de estados financieros.
Permite analizar múltiples empresas y generar reportes consolidados.
"""

import os
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
from openpyxl import load_workbook

from .data_extractor import DataExtractor
from .ratio_calculator import RatioCalculator
from .formula_builder import FormulaBuilder
from .excel_formatter import ExcelFormatter


class BulkProcessor:
    """
    Procesador masivo de análisis financiero.
    """
    
    def __init__(self, input_directory: str, output_directory: str, 
                 max_workers: int = 4):
        """
        Inicializa el procesador masivo.
        
        Args:
            input_directory: Directorio con archivos Excel de entrada
            output_directory: Directorio para archivos de salida
            max_workers: Número máximo de workers para procesamiento paralelo
        """
        self.input_dir = Path(input_directory)
        self.output_dir = Path(output_directory)
        self.max_workers = max_workers
        
        # Crear directorio de salida si no existe
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Configurar logging
        self._setup_logging()
        
        # Inicializar componentes
        self.formatter = ExcelFormatter()
        
        # Estadísticas de procesamiento
        self.stats = {
            "processed": 0,
            "successful": 0,
            "failed": 0,
            "errors": []
        }
    
    def _setup_logging(self):
        """Configura el sistema de logging."""
        log_file = self.output_dir / "bulk_processing.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger(__name__)
    
    def find_excel_files(self, pattern: str = "*.xlsx") -> List[Path]:
        """
        Encuentra archivos Excel en el directorio de entrada.
        
        Args:
            pattern: Patrón de archivos a buscar
            
        Returns:
            Lista de rutas de archivos Excel
        """
        files = list(self.input_dir.glob(pattern))
        self.logger.info(f"Encontrados {len(files)} archivos Excel")
        return files
    
    def process_single_file(self, file_path: Path, 
                          analysis_type: str = "formulas") -> Dict[str, Any]:
        """
        Procesa un único archivo Excel.
        
        Args:
            file_path: Ruta del archivo Excel
            analysis_type: Tipo de análisis ("formulas" o "values")
            
        Returns:
            Diccionario con resultado del procesamiento
        """
        result = {
            "file": file_path.name,
            "success": False,
            "output_file": None,
            "error": None,
            "company_name": None,
            "years_analyzed": [],
            "ratios_calculated": 0
        }
        
        try:
            self.logger.info(f"Procesando: {file_path.name}")
            
            # Extraer nombre de empresa del archivo
            company_name = self._extract_company_name(file_path.name)
            result["company_name"] = company_name
            
            # Extraer datos
            extractor = DataExtractor(str(file_path))
            if not extractor.load_data():
                raise Exception("Error cargando datos del archivo")
            
            financial_data = extractor.get_all_financial_data()
            result["years_analyzed"] = financial_data.get("years", [])
            
            if analysis_type == "formulas":
                output_file = self._process_with_formulas(file_path, financial_data, company_name)
            else:
                output_file = self._process_with_values(file_path, financial_data, company_name)
            
            result["output_file"] = output_file
            result["success"] = True
            
            # Contar ratios calculados
            calculator = RatioCalculator(financial_data)
            all_ratios = calculator.calculate_all_ratios()
            result["ratios_calculated"] = sum(len(category) for category in all_ratios.values())
            
            self.logger.info(f"Éxito: {file_path.name} -> {output_file}")
            
        except Exception as e:
            error_msg = f"Error procesando {file_path.name}: {str(e)}"
            self.logger.error(error_msg)
            result["error"] = error_msg
            
        return result
    
    def _process_with_formulas(self, file_path: Path, financial_data: Dict, 
                             company_name: str) -> str:
        """
        Procesa archivo con fórmulas Excel (análisis dinámico).
        
        Args:
            file_path: Ruta del archivo original
            financial_data: Datos financieros extraídos
            company_name: Nombre de la empresa
            
        Returns:
            Nombre del archivo de salida
        """
        # Abrir workbook original
        wb = load_workbook(str(file_path))
        
        # Agregar DataFrames al financial_data para FormulaBuilder
        extractor = DataExtractor(str(file_path))
        extractor.load_data()
        financial_data["_df_bal"] = extractor.df_bal
        financial_data["_df_pl"] = extractor.df_pl
        financial_data["_df_cfs"] = extractor.df_cfs
        
        # Construir fórmulas
        formula_builder = FormulaBuilder(wb, financial_data)
        formula_blocks = formula_builder.build_all_formulas()
        
        # Crear hoja de análisis
        sheet_name = "RATIOS & KPIs"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, 0)
        
        years = financial_data.get("years", [])
        cols_total = 1 + len(years) + 3  # Indicador + años + Último + Promedio + Tendencia
        
        # Configurar estructura
        header_row = self.formatter.setup_worksheet_structure(ws, years, sheet_name)
        
        # Escribir datos con fórmulas
        current_row = header_row + 1
        
        for section_name, formulas in formula_blocks:
            # Encabezado de sección
            self.formatter.format_section_header(ws, current_row, cols_total, section_name)
            current_row += 1
            
            # Fórmulas de la sección
            for name, ratio_type, func, description in formulas:
                ws.cell(row=current_row, column=1, value=name)
                
                # Escribir fórmulas por año
                formula_map = func()
                for j, year in enumerate(years, start=2):
                    formula_str = formula_map.get(str(year))
                    if formula_str:
                        ws.cell(row=current_row, column=j).value = f"={formula_str}"
                
                # Formatear fila
                self.formatter.format_ratio_row(ws, current_row, name, years, ratio_type)
                current_row += 1
        
        # Aplicar formateo condicional
        data_start_row = header_row + 1
        data_end_row = current_row - 1
        self.formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
        
        # Agregar sección tooltip
        tooltip_start = current_row + 2
        self.formatter.create_tooltip_section(ws, tooltip_start, formula_blocks, years, cols_total)
        
        # Congelar paneles
        self.formatter.add_freeze_panes(ws)
        
        # Guardar archivo
        output_filename = f"{company_name}_Analisis_Formulas.xlsx"
        output_path = self.output_dir / output_filename
        wb.save(str(output_path))
        
        return output_filename
    
    def _process_with_values(self, file_path: Path, financial_data: Dict, 
                           company_name: str) -> str:
        """
        Procesa archivo con valores calculados (análisis estático).
        
        Args:
            file_path: Ruta del archivo original
            financial_data: Datos financieros extraídos
            company_name: Nombre de la empresa
            
        Returns:
            Nombre del archivo de salida
        """
        # Calcular ratios
        calculator = RatioCalculator(financial_data)
        all_ratios = calculator.calculate_all_ratios()
        
        # Crear nuevo workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis Financiero (Valores)"
        
        years = financial_data.get("years", [])
        cols_total = 1 + len(years) + 3
        
        # Configurar estructura
        header_row = self.formatter.setup_worksheet_structure(ws, years)
        
        # Escribir datos con valores
        current_row = header_row + 1
        
        for section_name, ratios in all_ratios.items():
            # Encabezado de sección
            self.formatter.format_section_header(ws, current_row, cols_total, section_name)
            current_row += 1
            
            # Ratios de la sección
            for ratio_name, ratio_series in ratios.items():
                ws.cell(row=current_row, column=1, value=ratio_name)
                
                # Escribir valores por año
                for j, year in enumerate(years, start=2):
                    year_col = f"{year}-12"  # Asumiendo formato YYYY-MM
                    if year_col in ratio_series.index:
                        value = ratio_series[year_col]
                        if pd.notna(value):
                            ws.cell(row=current_row, column=j).value = value
                
                # Determinar tipo de ratio para formateo
                ratio_type = self._determine_ratio_type(ratio_name)
                self.formatter.format_ratio_row(ws, current_row, ratio_name, years, ratio_type)
                current_row += 1
        
        # Aplicar formateo condicional
        data_start_row = header_row + 1
        data_end_row = current_row - 1
        self.formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
        
        # Congelar paneles
        self.formatter.add_freeze_panes(ws)
        
        # Guardar archivo
        output_filename = f"{company_name}_Analisis_Valores.xlsx"
        output_path = self.output_dir / output_filename
        wb.save(str(output_path))
        
        return output_filename
    
    def _extract_company_name(self, filename: str) -> str:
        """
        Extrae el nombre de la empresa del nombre del archivo.
        
        Args:
            filename: Nombre del archivo
            
        Returns:
            Nombre de la empresa extraído
        """
        # Remover extensión
        name = filename.replace(".xlsx", "").replace(".xls", "")
        
        # Intentar extraer usando patrones comunes
        # Formato: RUT_EMPRESA_SA_EEFF_...
        parts = name.split("_")
        if len(parts) >= 2:
            # Buscar las partes que forman el nombre de la empresa
            company_parts = []
            for part in parts[1:]:
                if part in ["SA", "EEFF", "Balance", "Resultados", "Flujos", "Anual"]:
                    break
                company_parts.append(part)
            
            if company_parts:
                return "_".join(company_parts)
        
        # Fallback: usar nombre completo sin extensión
        return name
    
    def _determine_ratio_type(self, ratio_name: str) -> str:
        """
        Determina el tipo de ratio basado en el nombre.
        
        Args:
            ratio_name: Nombre del ratio
            
        Returns:
            Tipo de ratio para formateo
        """
        ratio_name_lower = ratio_name.lower()
        
        if any(word in ratio_name_lower for word in ["margen", "roe", "roa", "autonomía", "ac / at", "pc / pt"]):
            return "pct"
        elif any(word in ratio_name_lower for word in ["días", "período", "ciclo"]):
            return "days"
        elif any(word in ratio_name_lower for word in ["capital", "free cash flow", "fcf"]):
            return "number"
        else:
            return "ratio"
    
    def process_bulk(self, analysis_type: str = "formulas", 
                    file_pattern: str = "*.xlsx") -> Dict[str, Any]:
        """
        Procesa múltiples archivos en paralelo.
        
        Args:
            analysis_type: Tipo de análisis ("formulas" o "values")
            file_pattern: Patrón de archivos a procesar
            
        Returns:
            Diccionario con estadísticas del procesamiento
        """
        self.logger.info(f"Iniciando procesamiento masivo con {self.max_workers} workers")
        
        # Encontrar archivos
        files = self.find_excel_files(file_pattern)
        
        if not files:
            self.logger.warning("No se encontraron archivos para procesar")
            return self.stats
        
        # Procesar archivos en paralelo
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Enviar trabajos
            future_to_file = {
                executor.submit(self.process_single_file, file_path, analysis_type): file_path
                for file_path in files
            }
            
            # Recopilar resultados
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                
                try:
                    result = future.result()
                    self.stats["processed"] += 1
                    
                    if result["success"]:
                        self.stats["successful"] += 1
                    else:
                        self.stats["failed"] += 1
                        self.stats["errors"].append(result)
                        
                except Exception as e:
                    self.stats["processed"] += 1
                    self.stats["failed"] += 1
                    error_info = {
                        "file": file_path.name,
                        "error": str(e),
                        "success": False
                    }
                    self.stats["errors"].append(error_info)
                    self.logger.error(f"Error inesperado procesando {file_path.name}: {e}")
        
        # Resumen final
        self.logger.info(f"Procesamiento completado:")
        self.logger.info(f"  Total: {self.stats['processed']}")
        self.logger.info(f"  Exitosos: {self.stats['successful']}")
        self.logger.info(f"  Fallidos: {self.stats['failed']}")
        
        return self.stats
    
    def generate_summary_report(self) -> str:
        """
        Genera un reporte resumen del procesamiento masivo.
        
        Returns:
            Ruta del archivo de reporte generado
        """
        summary_file = self.output_dir / "processing_summary.xlsx"
        
        # Crear DataFrame con resultados
        summary_data = []
        
        for error in self.stats["errors"]:
            summary_data.append({
                "Archivo": error["file"],
                "Estado": "Error",
                "Empresa": error.get("company_name", "N/A"),
                "Años": str(error.get("years_analyzed", [])),
                "Ratios": error.get("ratios_calculated", 0),
                "Error": error.get("error", "N/A")
            })
        
        df_summary = pd.DataFrame(summary_data)
        
        # Guardar en Excel
        with pd.ExcelWriter(str(summary_file), engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name="Resumen", index=False)
            
            # Hoja de estadísticas
            stats_data = {
                "Métrica": ["Total Procesados", "Exitosos", "Fallidos", "Tasa de Éxito"],
                "Valor": [
                    self.stats["processed"],
                    self.stats["successful"], 
                    self.stats["failed"],
                    f"{(self.stats['successful']/max(self.stats['processed'], 1))*100:.1f}%"
                ]
            }
            pd.DataFrame(stats_data).to_excel(writer, sheet_name="Estadísticas", index=False)
        
        self.logger.info(f"Reporte de resumen generado: {summary_file}")
        return str(summary_file)
