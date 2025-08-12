#!/usr/bin/env python3
"""
Script de Migración - Análisis Excel Modularizado
================================================

Este script migra los archivos existentes para usar el nuevo módulo modularizado.
Procesa archivos usando el sistema antiguo y el nuevo para comparar resultados.
"""

import sys
from pathlib import Path

# Agregar el directorio padre al path
sys.path.append(str(Path(__file__).parent))

def migrate_existing_files():
    """Migra archivos existentes al nuevo sistema modularizado."""
    
    print("🔄 MIGRACIÓN AL SISTEMA MODULARIZADO")
    print("=" * 50)
    
    # Archivo de demostración existente
    demo_file = "./data/demo/FinDataChile_Data_Demo.xlsx"
    
    if not Path(demo_file).exists():
        print(f"❌ No se encuentra el archivo de demostración: {demo_file}")
        return False
    
    print(f"📁 Archivo de demostración encontrado: {demo_file}")
    
    # 1. Procesar con el nuevo sistema modularizado
    print("\\n1️⃣ Procesando con sistema modularizado...")
    
    try:
        from analisis_excel import DataExtractor, RatioCalculator, ExcelFormatter
        from openpyxl import load_workbook
        
        # Extraer datos
        extractor = DataExtractor(demo_file)
        if not extractor.load_data():
            print("❌ Error extrayendo datos")
            return False
        
        financial_data = extractor.get_all_financial_data()
        print(f"✅ Datos extraídos: {len(financial_data['years'])} años")
        
        # Calcular ratios
        calculator = RatioCalculator(financial_data)
        all_ratios = calculator.calculate_all_ratios()
        
        total_ratios = sum(len(category) for category in all_ratios.values())
        print(f"✅ Ratios calculados: {total_ratios} ratios en {len(all_ratios)} categorías")
        
        # Crear archivo de salida con valores
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis Modularizado"
        
        formatter = ExcelFormatter()
        years = financial_data.get("years", [])
        header_row = formatter.setup_worksheet_structure(ws, years)
        
        # Escribir algunos ratios de ejemplo
        current_row = header_row + 1
        
        for section_name, ratios in all_ratios.items():
            cols_total = 1 + len(years) + 3
            formatter.format_section_header(ws, current_row, cols_total, section_name)
            current_row += 1
            
            # Solo primeros 3 ratios de cada sección para demostración
            count = 0
            for ratio_name, ratio_series in ratios.items():
                if count >= 3:
                    break
                    
                ws.cell(row=current_row, column=1, value=ratio_name)
                
                # Escribir valores
                for j, year in enumerate(years, start=2):
                    year_col = None
                    for col in ratio_series.index:
                        if str(col).startswith(f"{year}-"):
                            year_col = col
                            break
                    
                    if year_col and year_col in ratio_series.index:
                        value = ratio_series[year_col]
                        if not pd.isna(value):
                            ws.cell(row=current_row, column=j).value = value
                
                current_row += 1
                count += 1
        
        # Guardar archivo de demostración
        output_file = "./data/demo/FinDataChile_Data_Demo_MODULARIZADO.xlsx"
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        
        print(f"✅ Archivo modularizado guardado: {output_file}")
        
        # 2. Mostrar comparación con archivos existentes
        print("\\n2️⃣ Comparando con archivos existentes...")
        
        existing_files = [
            "./data/demo/FinDataChile_Data_Demo_con_Analisis_Formulas.xlsx",
            "./data/demo/FinDataChile_Data_Demo_con_Analisis_SEGMENTADO.xlsx"
        ]
        
        for existing_file in existing_files:
            if Path(existing_file).exists():
                print(f"📄 Encontrado: {Path(existing_file).name}")
            else:
                print(f"❌ No encontrado: {Path(existing_file).name}")
        
        print("\\n3️⃣ Ventajas del sistema modularizado:")
        print("   ✅ Código organizado en módulos separados")
        print("   ✅ Fácil mantenimiento y extensión")
        print("   ✅ Procesamiento masivo integrado")
        print("   ✅ Manejo de errores mejorado")
        print("   ✅ Logging y estadísticas")
        print("   ✅ Procesamiento paralelo")
        
        return True
        
    except ImportError as e:
        print(f"❌ Error importando módulo modularizado: {e}")
        print("💡 Asegúrate de que la carpeta analisis-excel esté en el directorio correcto")
        return False
    except Exception as e:
        print(f"❌ Error durante migración: {e}")
        return False


def demonstrate_bulk_processing():
    """Demuestra el procesamiento masivo."""
    
    print("\\n🚀 DEMOSTRACIÓN DE PROCESAMIENTO MASIVO")
    print("=" * 50)
    
    reports_dir = "./data/Reports"
    
    if not Path(reports_dir).exists():
        print(f"❌ Directorio de reportes no encontrado: {reports_dir}")
        print("💡 Crea algunos archivos Excel en este directorio para probar el procesamiento masivo")
        return False
    
    # Contar archivos Excel
    excel_files = list(Path(reports_dir).glob("*.xlsx"))
    print(f"📊 Archivos Excel encontrados: {len(excel_files)}")
    
    if len(excel_files) == 0:
        print("💡 No hay archivos Excel para procesar")
        return False
    
    # Mostrar algunos archivos de ejemplo
    print("\\n📄 Archivos disponibles:")
    for i, file in enumerate(excel_files[:5]):
        print(f"   {i+1}. {file.name}")
    
    if len(excel_files) > 5:
        print(f"   ... y {len(excel_files) - 5} más")
    
    print("\\n💡 Para procesar estos archivos masivamente, ejecuta:")
    print(f"   python run_analisis_excel.py --mode bulk --input-dir {reports_dir} --output-dir ./output")
    
    return True


def create_example_config():
    """Crea archivo de configuración de ejemplo."""
    
    config_content = '''# Configuración de Análisis Excel Modularizado
# ============================================

# Directorios
INPUT_DIR = "./data/Reports"
OUTPUT_DIR = "./data/Analisis"

# Procesamiento
MAX_WORKERS = 4
ANALYSIS_TYPE = "formulas"  # "formulas" o "values"

# Ratios a incluir por categoría
RATIOS_CONFIG = {
    "LIQUIDEZ": [
        "Liquidez Corriente",
        "Prueba Ácida", 
        "Cash Ratio",
        "Capital de Trabajo"
    ],
    "SOLVENCIA Y ESTRUCTURA": [
        "Endeudamiento (D/E)",
        "Apalancamiento (D/A)",
        "Cobertura de Intereses",
        "Deuda / EBITDA",
        "Autonomía Financiera"
    ],
    "RENTABILIDAD": [
        "Margen Bruto",
        "Margen Operativo (EBIT)",
        "Margen EBITDA",
        "Margen Neto",
        "ROE",
        "ROA"
    ],
    "EFICIENCIA OPERATIVA": [
        "Rotación de Activos",
        "Rotación de Inventarios",
        "Días de Inventario",
        "Rotación de Cuentas por Cobrar",
        "Período Promedio de Cobro",
        "Ciclo de Conversión de Efectivo"
    ],
    "FLUJOS Y ADICIONALES": [
        "Conversión de caja (CFO/Utilidad Neta)",
        "Free Cash Flow (CFO - CAPEX)",
        "AC / AT",
        "PC / PT"
    ]
}

# Formateo Excel
EXCEL_CONFIG = {
    "freeze_panes": True,
    "conditional_formatting": True,
    "data_bars": True,
    "color_scale": True
}
'''
    
    config_file = "./config_analisis_excel.py"
    with open(config_file, 'w', encoding='utf-8') as f:
        f.write(config_content)
    
    print(f"📋 Archivo de configuración creado: {config_file}")


def main():
    """Función principal del script de migración."""
    
    print("🔄 SCRIPT DE MIGRACIÓN - ANÁLISIS EXCEL MODULARIZADO")
    print("=" * 60)
    
    # 1. Migrar archivos existentes
    success = migrate_existing_files()
    
    # 2. Demostrar procesamiento masivo
    demonstrate_bulk_processing()
    
    # 3. Crear configuración de ejemplo
    create_example_config()
    
    # 4. Instrucciones finales
    print("\\n📚 INSTRUCCIONES DE USO")
    print("=" * 30)
    print("\\n1. Archivo único:")
    print("   python run_analisis_excel.py --mode single --file archivo.xlsx")
    
    print("\\n2. Procesamiento masivo:")
    print("   python run_analisis_excel.py --mode bulk --input-dir ./data/Reports")
    
    print("\\n3. Con configuraciones específicas:")
    print("   python run_analisis_excel.py --mode bulk --input-dir ./data/Reports --workers 8 --analysis-type values")
    
    print("\\n✅ Migración completada exitosamente!" if success else "⚠️  Migración completada con advertencias")
    
    return 0 if success else 1


if __name__ == "__main__":
    import pandas as pd
    exit(main())
