# Build the segmented analysis WITH REAL EXCEL FORMULAS + a "Definición y Fórmula" tooltip section.
import re, os
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ---------- Load source file ----------
# Source file (prefer the randomized-preserved formatting if present)
candidate_files = [ './data/demo/FinDataChile_Data_Demo.xlsx'
]
base_file = next((f for f in candidate_files if os.path.exists(f)), candidate_files[-1])

# Read to get sheet headers and row indexes
df_bal = pd.read_excel(base_file, sheet_name="Balance General")
df_pl  = pd.read_excel(base_file, sheet_name="Estado Resultados (Función)")
df_cfs = pd.read_excel(base_file, sheet_name="Flujo Efectivo")
df_bal.rename(columns={df_bal.columns[0]:"Concepto"}, inplace=True)
df_pl.rename(columns={df_pl.columns[0]:"Concepto"}, inplace=True)
df_cfs.rename(columns={df_cfs.columns[0]:"Concepto"}, inplace=True)

# Open workbook
wb = load_workbook(base_file)
sh_bal = wb["Balance General"]
sh_pl  = wb["Estado Resultados (Función)"]
sh_cfs = wb["Flujo Efectivo"]

# Helper: map concept text -> row number (1-based in Excel)
def find_row(sheet, concept_df, name):
    # exact match first
    mask = concept_df["Concepto"].astype(str).str.strip().str.lower() == name.strip().lower()
    if mask.any():
        idx = concept_df.index[mask][0]
    else:
        # contains fallback
        mask = concept_df["Concepto"].astype(str).str.contains(re.escape(name), case=False, na=False)
        if not mask.any():
            return None
        idx = concept_df.index[mask][0]
    # +2 because df has header row, Excel sheet starts at row 1 and header at row 1
    # df row index starts at 0; excel row = df_index + 2
    return int(idx) + 2

# Helper: map year -> column letter by looking up header that starts with "YYYY-"
def find_year_col(sheet, header_row, year):
    for col in range(2, sheet.max_column+1):
        val = sheet.cell(row=header_row, column=col).value
        if isinstance(val, str) and val.startswith(f"{year}-"):
            return get_column_letter(col)
    return None

# Determine header row is always row 1
HEADER_ROW = 1

# Build years present across sheets
def years_from_sheet(sheet):
    years = []
    for col in range(2, sheet.max_column+1):
        v = sheet.cell(row=HEADER_ROW, column=col).value
        if isinstance(v, str):
            m = re.match(r"^(\d{4})-", v)
            if m:
                years.append(int(m.group(1)))
    return years

years = sorted(set(years_from_sheet(sh_bal)) | set(years_from_sheet(sh_pl)) | set(years_from_sheet(sh_cfs)))

# Find necessary rows in each sheet
rows_bal = {
    "AC": find_row(sh_bal, df_bal, "Activos corrientes totales"),
    "PC": find_row(sh_bal, df_bal, "Pasivos corrientes totales"),
    "Efec": find_row(sh_bal, df_bal, "Efectivo y equivalentes al efectivo"),
    "Inv": find_row(sh_bal, df_bal, "Inventarios corrientes"),
    "AT": find_row(sh_bal, df_bal, "Total de activos"),
    "PT": find_row(sh_bal, df_bal, "Total de pasivos"),
    "Patr": find_row(sh_bal, df_bal, "Patrimonio atribuible a los propietarios de la controladora"),
    "CxC": find_row(sh_bal, df_bal, "Deudores comerciales y otras cuentas por cobrar corrientes"),
    "CxP": find_row(sh_bal, df_bal, "Cuentas por pagar comerciales y otras cuentas por pagar"),
}
rows_pl = {
    "Ventas":  find_row(sh_pl, df_pl, "Ingresos de actividades ordinarias"),
    "COGS":    find_row(sh_pl, df_pl, "Costo de ventas"),
    "Bruta":   find_row(sh_pl, df_pl, "Ganancia bruta"),
    "EBIT":    find_row(sh_pl, df_pl, "Ganancias (pérdidas) de actividades operacionales"),
    "Neta":    find_row(sh_pl, df_pl, "Ganancia (pérdida)"),
    "Interes": find_row(sh_pl, df_pl, "Costos financieros"),
    "Dep":     find_row(sh_pl, df_pl, "Depreciación"),
    "Amort":   find_row(sh_pl, df_pl, "Amortización"),
}
rows_cfs = {
    "CFO":      find_row(sh_cfs, df_cfs, "Flujos de efectivo netos procedentes de (utilizados en) operaciones"),
    "CapexBuy": find_row(sh_cfs, df_cfs, "Compras de propiedades, planta y equipo"),
}

# Utility: safe cell reference
def ref(sheet_name, col_letter, row_num):
    if col_letter is None or row_num is None:
        return None
    return f"'{sheet_name}'!{col_letter}{row_num}"

# Create/replace analysis sheet
sheet_name = "Análisis Avanzado (Fórmulas)"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name, 0)

# Styles
header_fill = PatternFill("solid", fgColor="0B2447")
subheader_fill = PatternFill("solid", fgColor="19376D")
section_fills = {
    "LIQUIDEZ": PatternFill("solid", fgColor="D1E7DD"),
    "SOLVENCIA Y ESTRUCTURA": PatternFill("solid", fgColor="FAD7A0"),
    "RENTABILIDAD": PatternFill("solid", fgColor="F8D7DA"),
    "EFICIENCIA OPERATIVA": PatternFill("solid", fgColor="D6EAF8"),
    "FLUJOS Y ADICIONALES": PatternFill("solid", fgColor="E8DAEF"),
}
bold_white = Font(bold=True, color="FFFFFF", size=13)
bold_white_small = Font(bold=True, color="FFFFFF", size=11)
bold_dark = Font(bold=True, color="000000", size=11)
normal = Font(color="000000", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

num_years = len(years)
cols_total = 1 + num_years + 2 + 1  # Indicador + años + Último + Promedio + Tendencia
for c in range(1, cols_total+1):
    ws.column_dimensions[get_column_letter(c)].width = 30 if c==1 else 14

# Title & subheader
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_total)
ws.cell(row=1, column=1, value="Análisis Financiero – Ratios y Evolución (celdas con FÓRMULAS)").fill = header_fill
ws.cell(row=1, column=1).font = bold_white
ws.cell(row=1, column=1).alignment = center

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_total)
ws.cell(row=2, column=1, value="Fechas: Balance (AAAA-12), Resultados (AAAA-01), Flujos (AAAA-01). Evolución alineada por AÑO.").fill = subheader_fill
ws.cell(row=2, column=1).font = bold_white_small
ws.cell(row=2, column=1).alignment = center

# Header row
header_row = 4
headers = ["Indicador"] + [str(y) for y in years] + ["Último", "Promedio", "Tendencia"]
for i,h in enumerate(headers, start=1):
    ws.cell(row=header_row, column=i, value=h).font = bold_white_small
    ws.cell(row=header_row, column=i).fill = subheader_fill
    ws.cell(row=header_row, column=i).alignment = center
    ws.cell(row=header_row, column=i).border = border

# Ratio specification: name, kind, formula builder (returns dict year->excel_formula), text formula
def build_formulas():
    blocks = []
    LIQ = []
    SOLV = []
    RENT = []
    EFF = []
    FLOW = []

    # Helper to average two balance cells (now, prev) if both exist
    def avg_ref(sheet, row, year):
        col_now = find_year_col(sheet, HEADER_ROW, year)
        col_prev = find_year_col(sheet, HEADER_ROW, year-1)
        now_ref = ref(sheet.title, col_now, row) if col_now else None
        prev_ref = ref(sheet.title, col_prev, row) if col_prev else None
        if now_ref and prev_ref:
            return f"AVERAGE({now_ref},{prev_ref})"
        elif now_ref:
            return now_ref
        elif prev_ref:
            return prev_ref
        return None

    # --- LIQUIDEZ ---
    # Liquidez Corriente = AC/PC
    def f_liq_corr():
        m = {}
        for y in years:
            ac = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AC"])
            pc = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PC"])
            if ac and pc:
                m[str(y)] = f"IFERROR({ac}/{pc},\"\")"
        return m
    LIQ.append(("Liquidez Corriente", "ratio", f_liq_corr, "Activo Corriente / Pasivo Corriente"))

    def f_prueba_acida():
        m = {}
        for y in years:
            ac = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AC"])
            inv = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Inv"])
            pc = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PC"])
            if ac and pc:
                term_inv = inv if inv else "0"
                m[str(y)] = f"IFERROR(({ac}-{term_inv})/{pc},\"\")"
        return m
    LIQ.append(("Prueba Ácida", "ratio", f_prueba_acida, "(Activo Corriente - Inventarios) / Pasivo Corriente"))

    def f_cash_ratio():
        m = {}
        for y in years:
            ef = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Efec"])
            pc = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PC"])
            if ef and pc:
                m[str(y)] = f"IFERROR({ef}/{pc},\"\")"
        return m
    LIQ.append(("Cash Ratio", "ratio", f_cash_ratio, "Efectivo y Equivalentes / Pasivo Corriente"))

    def f_cap_trabajo():
        m = {}
        for y in years:
            ac = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AC"])
            pc = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PC"])
            if ac and pc:
                m[str(y)] = f"IFERROR({ac}-{pc},\"\")"
        return m
    LIQ.append(("Capital de Trabajo", "number", f_cap_trabajo, "Activo Corriente - Pasivo Corriente"))

    # --- SOLVENCIA ---
    def f_de():
        m = {}
        for y in years:
            pt = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PT"])
            patr = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Patr"])
            if pt and patr:
                m[str(y)] = f"IFERROR({pt}/{patr},\"\")"
        return m
    SOLV.append(("Endeudamiento (D/E)", "ratio", f_de, "Deuda Total / Patrimonio"))

    def f_da():
        m = {}
        for y in years:
            pt = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PT"])
            at = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AT"])
            if pt and at:
                m[str(y)] = f"IFERROR({pt}/{at},\"\")"
        return m
    SOLV.append(("Apalancamiento (D/A)", "ratio", f_da, "Deuda Total / Activos Totales"))

    def f_cover():
        m = {}
        for y in years:
            ebit = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["EBIT"])
            interes = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Interes"])
            if ebit and interes:
                m[str(y)] = f"IFERROR({ebit}/ABS({interes}),\"\")"
        return m
    SOLV.append(("Cobertura de Intereses", "ratio", f_cover, "EBIT / |Gastos por Intereses|"))

    def f_deuda_ebitda():
        m = {}
        for y in years:
            pt = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PT"])
            ebit = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["EBIT"])
            dep = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Dep"]) if rows_pl["Dep"] else None
            amort = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Amort"]) if rows_pl["Amort"] else None
            if pt and ebit:
                ebitda = None
                if dep and amort:
                    ebitda = f"({ebit}+IFERROR({dep},0)+IFERROR({amort},0))"
                else:
                    ebitda = f"({ebit})"
                m[str(y)] = f"IFERROR({pt}/{ebitda},\"\")"
        return m
    SOLV.append(("Deuda / EBITDA", "ratio", f_deuda_ebitda, "Deuda Total / (EBIT + Depreciación + Amortización)"))

    def f_autonomia():
        m = {}
        for y in years:
            patr = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Patr"])
            at = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AT"])
            if patr and at:
                m[str(y)] = f"IFERROR({patr}/{at},\"\")"
        return m
    SOLV.append(("Autonomía Financiera", "pct", f_autonomia, "Patrimonio / Activo Total"))

    # --- RENTABILIDAD ---
    def f_margen_bruto():
        m = {}
        for y in years:
            gb = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Bruta"])
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            if gb and ven:
                m[str(y)] = f"IFERROR({gb}/{ven},\"\")"
        return m
    RENT.append(("Margen Bruto", "pct", f_margen_bruto, "Utilidad Bruta / Ventas"))

    def f_margen_operativo():
        m = {}
        for y in years:
            ebit = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["EBIT"])
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            if ebit and ven:
                m[str(y)] = f"IFERROR({ebit}/{ven},\"\")"
        return m
    RENT.append(("Margen Operativo (EBIT)", "pct", f_margen_operativo, "EBIT / Ventas"))

    def f_margen_ebitda():
        m = {}
        for y in years:
            ebit = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["EBIT"])
            dep = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Dep"]) if rows_pl["Dep"] else None
            amort = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Amort"]) if rows_pl["Amort"] else None
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            if ebit and ven:
                ebitda = f"({ebit}+IFERROR({dep},0)+IFERROR({amort},0))" if (dep or amort) else f"({ebit})"
                m[str(y)] = f"IFERROR({ebitda}/{ven},\"\")"
        return m
    RENT.append(("Margen EBITDA", "pct", f_margen_ebitda, "EBITDA / Ventas"))

    def f_margen_neto():
        m = {}
        for y in years:
            net = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Neta"])
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            if net and ven:
                m[str(y)] = f"IFERROR({net}/{ven},\"\")"
        return m
    RENT.append(("Margen Neto", "pct", f_margen_neto, "Utilidad Neta / Ventas"))

    def f_roe():
        m = {}
        for y in years:
            net = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Neta"])
            patr_avg = avg_ref(sh_bal, rows_bal["Patr"], y)
            if net and patr_avg:
                m[str(y)] = f"IFERROR({net}/{patr_avg},\"\")"
        return m
    RENT.append(("ROE", "pct", f_roe, "Utilidad Neta / Patrimonio Promedio"))

    def f_roa():
        m = {}
        for y in years:
            net = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Neta"])
            at_avg = avg_ref(sh_bal, rows_bal["AT"], y)
            if net and at_avg:
                m[str(y)] = f"IFERROR({net}/{at_avg},\"\")"
        return m
    RENT.append(("ROA", "pct", f_roa, "Utilidad Neta / Activos Totales Promedio"))

    # --- EFICIENCIA ---
    def f_rot_act():
        m = {}
        for y in years:
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            at_avg = avg_ref(sh_bal, rows_bal["AT"], y)
            if ven and at_avg:
                m[str(y)] = f"IFERROR({ven}/{at_avg},\"\")"
        return m
    EFF.append(("Rotación de Activos", "ratio", f_rot_act, "Ventas / Activos Promedio"))

    def f_rot_inv_y_dias():
        m_rot = {}
        m_dias = {}
        for y in years:
            cogs = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["COGS"])
            inv_now = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Inv"])
            inv_prev = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y-1), rows_bal["Inv"])
            inv_avg = None
            if inv_now and inv_prev:
                inv_avg = f"AVERAGE({inv_now},{inv_prev})"
            elif inv_now:
                inv_avg = inv_now
            elif inv_prev:
                inv_avg = inv_prev

            if cogs and inv_avg:
                m_rot[str(y)] = f"IFERROR({cogs}/{inv_avg},\"\")"
                m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
        return m_rot, m_dias
    rot_inv, dias_inv = f_rot_inv_y_dias()
    EFF.append(("Rotación de Inventarios", "ratio", lambda: rot_inv, "Costo de Ventas / Inventario Promedio"))
    EFF.append(("Días de Inventario", "days", lambda: dias_inv, "365 / Rotación de Inventarios"))

    def f_rot_cxc_y_dias():
        m_rot = {}
        m_dias = {}
        for y in years:
            ven = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Ventas"])
            cxc_avg = avg_ref(sh_bal, rows_bal["CxC"], y)
            if ven and cxc_avg:
                m_rot[str(y)] = f"IFERROR({ven}/{cxc_avg},\"\")"
                m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
        return m_rot, m_dias
    rot_cxc, dias_cobro = f_rot_cxc_y_dias()
    EFF.append(("Rotación de Cuentas por Cobrar", "ratio", lambda: rot_cxc, "Ventas / Cuentas por Cobrar Promedio"))
    EFF.append(("Período Promedio de Cobro", "days", lambda: dias_cobro, "365 / Rotación de CxC"))

    def f_rot_cxp_y_dias():
        m_rot = {}
        m_dias = {}
        for y in years:
            cogs = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["COGS"])
            inv_now = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["Inv"])
            inv_prev = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y-1), rows_bal["Inv"])
            comp = None
            if cogs and inv_now and inv_prev:
                comp = f"({cogs}+({inv_now}-{inv_prev}))"
            # average payables
            cxp_avg = avg_ref(sh_bal, rows_bal["CxP"], y)
            if comp and cxp_avg:
                m_rot[str(y)] = f"IFERROR({comp}/{cxp_avg},\"\")"
                m_dias[str(y)] = f"IFERROR(365/{m_rot[str(y)]},\"\")"
        return m_rot, m_dias
    rot_cxp, dias_pago = f_rot_cxp_y_dias()
    EFF.append(("Rotación de Cuentas por Pagar", "ratio", lambda: rot_cxp, "Compras (≈ COGS + ΔInventario) / Cuentas por Pagar Promedio"))
    EFF.append(("Período Promedio de Pago", "days", lambda: dias_pago, "365 / Rotación de CxP"))

    def f_cce():
        m = {}
        for y in years:
            cce = None
            # CCE = Días Inv + Días CxC - Días CxP
            di = dias_inv.get(str(y))
            dc = dias_cobro.get(str(y))
            dp = dias_pago.get(str(y))
            if di and dc and dp:
                m[str(y)] = f"IFERROR({di}+{dc}-{dp},\"\")"
        return m
    EFF.append(("Ciclo de Conversión de Efectivo", "days", f_cce, "Días Inventario + Días CxC - Días CxP"))

    # --- FLUJOS Y ADICIONALES ---
    def f_conv_caja():
        m = {}
        for y in years:
            cfo = ref(sh_cfs.title, find_year_col(sh_cfs, HEADER_ROW, y), rows_cfs["CFO"])
            net = ref(sh_pl.title, find_year_col(sh_pl, HEADER_ROW, y), rows_pl["Neta"])
            if cfo and net:
                m[str(y)] = f"IFERROR({cfo}/{net},\"\")"
        return m
    FLOW.append(("Conversión de caja (CFO/Utilidad Neta)", "ratio", f_conv_caja, "Flujo Operativo / Utilidad Neta"))

    def f_fcf():
        m = {}
        for y in years:
            cfo = ref(sh_cfs.title, find_year_col(sh_cfs, HEADER_ROW, y), rows_cfs["CFO"])
            capex = ref(sh_cfs.title, find_year_col(sh_cfs, HEADER_ROW, y), rows_cfs["CapexBuy"]) if rows_cfs["CapexBuy"] else None
            if cfo:
                if capex:
                    m[str(y)] = f"IFERROR({cfo}-ABS({capex}),\"\")"
                else:
                    m[str(y)] = f"IFERROR({cfo},\"\")"
        return m
    FLOW.append(("Free Cash Flow (CFO - CAPEX)", "number", f_fcf, "CFO - CAPEX (Compras PPE)"))

    def f_ac_at():
        m = {}
        for y in years:
            ac = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AC"])
            at = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["AT"])
            if ac and at:
                m[str(y)] = f"IFERROR({ac}/{at},\"\")"
        return m
    FLOW.append(("AC / AT", "pct", f_ac_at, "Activo Corriente / Activo Total"))

    def f_pc_pt():
        m = {}
        for y in years:
            pc = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PC"])
            pt = ref(sh_bal.title, find_year_col(sh_bal, HEADER_ROW, y), rows_bal["PT"])
            if pc and pt:
                m[str(y)] = f"IFERROR({pc}/{pt},\"\")"
        return m
    FLOW.append(("PC / PT", "pct", f_pc_pt, "Pasivo Corriente / Pasivo Total"))

    return [
        ("LIQUIDEZ", LIQ),
        ("SOLVENCIA Y ESTRUCTURA", SOLV),
        ("RENTABILIDAD", RENT),
        ("EFICIENCIA OPERATIVA", EFF),
        ("FLUJOS Y ADICIONALES", FLOW),
    ]

blocks = build_formulas()

# Write header row
r = header_row + 1
for section, items in blocks:
    # Section row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols_total)
    sec = ws.cell(row=r, column=1, value=section)
    sec.font = bold_dark
    sec.alignment = left
    sec.fill = section_fills.get(section)
    for c in range(1, cols_total+1):
        ws.cell(row=r, column=c).border = border
    r += 1

    # Items
    for (name, kind, func, text_formula) in items:
        ws.cell(row=r, column=1, value=name).alignment = left
        ws.cell(row=r, column=1).border = border
        # Per-year formulas
        row_vals = []
        for j, y in enumerate(years, start=2):
            formula_map = func()
            fstr = formula_map.get(str(y))
            cell = ws.cell(row=r, column=j)
            if fstr:
                cell.value = f"={fstr}"
            cell.alignment = center
            cell.border = border
            # number formats
            if kind == "pct":
                cell.number_format = "0.0%"
            elif kind == "number":
                cell.number_format = "#,##0"
            elif kind == "days":
                cell.number_format = "0"
            else:
                cell.number_format = "0.00"
        # Último = último año (col of last year)
        last_col_idx = 1 + len(years) + 1
        # Build formula to get last non-empty across the year range for that row
        year_start_col = 2
        year_end_col = 1 + len(years)
        rng = f"{get_column_letter(year_start_col)}{r}:{get_column_letter(year_end_col)}{r}"
        ws.cell(row=r, column=last_col_idx).value = f"=LOOKUP(2,1/(--({rng}<>\"\")),{rng})"
        ws.cell(row=r, column=last_col_idx).alignment = center
        ws.cell(row=r, column=last_col_idx).border = border
        if kind == "pct":
            ws.cell(row=r, column=last_col_idx).number_format = "0.0%"
        elif kind == "number":
            ws.cell(row=r, column=last_col_idx).number_format = "#,##0"
        elif kind == "days":
            ws.cell(row=r, column=last_col_idx).number_format = "0"
        else:
            ws.cell(row=r, column=last_col_idx).number_format = "0.00"

        # Promedio = AVERAGE ignoring blanks
        avg_col_idx = 1 + len(years) + 2
        ws.cell(row=r, column=avg_col_idx).value = f"=IFERROR(AVERAGE({rng}),\"\")"
        ws.cell(row=r, column=avg_col_idx).alignment = center
        ws.cell(row=r, column=avg_col_idx).border = border
        if kind == "pct":
            ws.cell(row=r, column=avg_col_idx).number_format = "0.0%"
        elif kind == "number":
            ws.cell(row=r, column=avg_col_idx).number_format = "#,##0"
        elif kind == "days":
            ws.cell(row=r, column=avg_col_idx).number_format = "0"
        else:
            ws.cell(row=r, column=avg_col_idx).number_format = "0.00"

        # Tendencia: compare last vs previous non-empty
        trend_col_idx = 1 + len(years) + 3
        # helper cells (not visible) could be used; we'll inline a safe approach:
        # Get last value L and previous P via LOOKUP tricks
        ws.cell(row=r, column=trend_col_idx).value = f"=IFERROR(IF(" \
            f"(LOOKUP(2,1/(--({rng}<>\"\")),{rng}))>" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▲\"," \
            f"IF((LOOKUP(2,1/(--({rng}<>\"\")),{rng}))<" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▼\",\"→\"))," \
            f"\"→\")"
        ws.cell(row=r, column=trend_col_idx).alignment = center
        ws.cell(row=r, column=trend_col_idx).border = border

        r += 1

# Heatmaps (across year columns for all rows)
def add_heatmap(r_start, r_end):
    year_start = 2
    year_end = 1 + len(years)
    rng = f"{get_column_letter(year_start)}{r_start}:{get_column_letter(year_end)}{r_end}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="percentile", start_value=5, start_color="FDE725",
            mid_type="percentile", mid_value=50, mid_color="5DC863",
            end_type="percentile", end_value=95, end_color="2A788E"
        )
    )

data_start_row = header_row + 1
data_end_row = ws.max_row
add_heatmap(data_start_row, data_end_row)

# Data bars on Último y Promedio
last_col_letter = get_column_letter(1 + len(years) + 1)
avg_col_letter  = get_column_letter(1 + len(years) + 2)
ws.conditional_formatting.add(
    f"{last_col_letter}{data_start_row}:{last_col_letter}{data_end_row}",
    DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
)
ws.conditional_formatting.add(
    f"{avg_col_letter}{data_start_row}:{avg_col_letter}{data_end_row}",
    DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
)

# Freeze panes
ws.freeze_panes = "B5"

# ---------- Tooltip section: Definición y Fórmula ----------
tip_title_row = ws.max_row + 2
ws.merge_cells(start_row=tip_title_row, start_column=1, end_row=tip_title_row, end_column=cols_total)
tip_title = ws.cell(row=tip_title_row, column=1, value="Definición y Fórmula (tooltip)")
tip_title.font = bold_dark
tip_title.alignment = left
tip_title.fill = PatternFill("solid", fgColor="EFEFEF")

# Headers
tip_header_row = tip_title_row + 1
ws.cell(row=tip_header_row, column=1, value="Indicador").font = bold_white_small
ws.cell(row=tip_header_row, column=2, value="Fórmula (texto)").font = bold_white_small
ws.cell(row=tip_header_row, column=3, value="Ejemplo de Fórmula Excel (último año)").font = bold_white_small
for c in range(1, 4):
    ws.cell(row=tip_header_row, column=c).fill = subheader_fill
    ws.cell(row=tip_header_row, column=c).alignment = center
    ws.cell(row=tip_header_row, column=c).border = border

# Compose the formula text + example for the last year
last_year = years[-1] if years else None
tip_row = tip_header_row + 1

def example_for_year(y, func):
    fmap = func()
    f = fmap.get(str(y))
    if not f:
        return ""
    return f"={f}"

for section, items in blocks:
    # section subheader within tooltip
    ws.cell(row=tip_row, column=1, value=section).font = bold_dark
    ws.cell(row=tip_row, column=1).fill = section_fills.get(section)
    for c in range(1, 4):
        ws.cell(row=tip_row, column=c).border = border
    tip_row += 1

    for (name, kind, func, text_formula) in items:
        ws.cell(row=tip_row, column=1, value=name).alignment = left
        ws.cell(row=tip_row, column=1).border = border
        ws.cell(row=tip_row, column=2, value=text_formula).alignment = left
        ws.cell(row=tip_row, column=2).border = border
        ws.cell(row=tip_row, column=3, value=example_for_year(last_year, func)).alignment = left
        ws.cell(row=tip_row, column=3).border = border
        tip_row += 1

# Save output
# Save
out_path = "./data/demo/FinDataChile_Data_Demo_con_Analisis_Formulas.xlsx"
Path(out_path).parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)
