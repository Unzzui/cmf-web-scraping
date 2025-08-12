# Rebuild the styled sheet using your preferred template, but with MANY more ratios,
# segmented by categories, computed across all years, with Último, Promedio y Tendencia (▲▼→).
import pandas as pd
import numpy as np
import re, os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# Source file (prefer the randomized-preserved formatting if present)
candidate_files = [ './data/demo/FinDataChile_Data_Demo.xlsx'
]
base_file = next((f for f in candidate_files if os.path.exists(f)), candidate_files[-1])

# Load dataframes
df_bal = pd.read_excel(base_file, sheet_name="Balance General")
df_pl  = pd.read_excel(base_file, sheet_name="Estado Resultados (Función)")
df_cfs = pd.read_excel(base_file, sheet_name="Flujo Efectivo")

# Ensure the first column is named Concepto
df_bal.rename(columns={df_bal.columns[0]:"Concepto"}, inplace=True)
df_pl.rename(columns={df_pl.columns[0]:"Concepto"}, inplace=True)
df_cfs.rename(columns={df_cfs.columns[0]:"Concepto"}, inplace=True)

# Helper: get series by exact row name
def row_series(df, exact):
    m = df["Concepto"].astype(str).str.strip().str.lower() == exact.strip().lower()
    if m.any():
        s = df[m].iloc[0].drop(labels=["Concepto"])
        s.index = s.index.astype(str)
        return pd.to_numeric(s, errors="coerce")
    # Fallback: contains
    m = df["Concepto"].astype(str).str.contains(re.escape(exact), case=False, na=False)
    if m.any():
        s = df[m].iloc[0].drop(labels=["Concepto"])
        s.index = s.index.astype(str)
        return pd.to_numeric(s, errors="coerce")
    return pd.Series(dtype=float)

# Extract balance items (per your list)
AC   = row_series(df_bal, "Activos corrientes totales")
PC   = row_series(df_bal, "Pasivos corrientes totales")
Efec = row_series(df_bal, "Efectivo y equivalentes al efectivo")
Inv  = row_series(df_bal, "Inventarios corrientes")
AT   = row_series(df_bal, "Total de activos")
PT   = row_series(df_bal, "Total de pasivos")
Patr = row_series(df_bal, "Patrimonio atribuible a los propietarios de la controladora")
CxC  = row_series(df_bal, "Deudores comerciales y otras cuentas por cobrar corrientes")
CxP  = row_series(df_bal, "Cuentas por pagar comerciales y otras cuentas por pagar")

# P&L
Ventas   = row_series(df_pl, "Ingresos de actividades ordinarias")
COGS     = row_series(df_pl, "Costo de ventas")
Bruta    = row_series(df_pl, "Ganancia bruta")
EBIT     = row_series(df_pl, "Ganancias (pérdidas) de actividades operacionales")
Neta     = row_series(df_pl, "Ganancia (pérdida)")
Interes  = row_series(df_pl, "Costos financieros")
# Try to get D&A if available
DA = row_series(df_pl, "Depreciación") + row_series(df_pl, "Amortización")
if DA.empty:
    # No D&A rows found
    DA = pd.Series(0.0, index=EBIT.index)

# Cash flows
CFO  = row_series(df_cfs, "Flujos de efectivo netos procedentes de (utilizados en) operaciones")
Capex_buy = row_series(df_cfs, "Compras de propiedades, planta y equipo")
if Capex_buy.empty:
    Capex_buy = pd.Series(0.0, index=CFO.index)
CAPEX = Capex_buy.abs()
FCF = CFO - CAPEX

# Align indices (years) by intersecting available columns
def years_from_cols(series):
    years = []
    for c in series.index:
        m = re.match(r"^(\d{4})-", str(c))
        if m: years.append(int(m.group(1)))
    return years

years = sorted(set(years_from_cols(AT)) | set(years_from_cols(Ventas)) | set(years_from_cols(CFO)))
# Build column labels "YYYY"
cols = [str(y) for y in years]

def col_for_year(series, year):
    # find column that starts with f"{year}-"
    for c in series.index:
        if str(c).startswith(f"{year}-"):
            return c
    return None

# Helper: average of two-year balance for ratios denominators
def avg_bal_item(series, y):
    c_now = col_for_year(series, y)
    c_prev = col_for_year(series, y-1)
    if c_now is None and c_prev is None: return np.nan
    v_now = series.get(c_now, np.nan)
    v_prev = series.get(c_prev, np.nan)
    if np.isnan(v_prev): return v_now
    if np.isnan(v_now): return v_prev
    return (v_now + v_prev)/2.0

# Purchases approximation for payables metrics
def compras_aprox(y):
    c = col_for_year(COGS, y)
    if c is None: return np.nan
    c_inv_now = col_for_year(Inv, y)
    c_inv_prev = col_for_year(Inv, y-1)
    inv_now = Inv.get(c_inv_now, np.nan)
    inv_prev = Inv.get(c_inv_prev, np.nan)
    delta_inv = 0 if (np.isnan(inv_now) or np.isnan(inv_prev)) else (inv_now - inv_prev)
    return COGS.get(c, np.nan) + delta_inv

# Compute all ratios per your taxonomy
data_blocks = []  # list of (section_title, [(ratio_name, series, kind)])

# 1) LIQUIDEZ
liq = {}
for y in years:
    liq.setdefault("Liquidez Corriente", {})[str(y)] = AC.get(col_for_year(AC,y), np.nan)/PC.get(col_for_year(PC,y), np.nan)
    liq.setdefault("Prueba Ácida", {})[str(y)] = (AC.get(col_for_year(AC,y), np.nan)-(Inv.get(col_for_year(Inv,y),0))) / PC.get(col_for_year(PC,y), np.nan)
    liq.setdefault("Cash Ratio", {})[str(y)] = Efec.get(col_for_year(Efec,y), np.nan) / PC.get(col_for_year(PC,y), np.nan)
    liq.setdefault("Capital de Trabajo", {})[str(y)] = AC.get(col_for_year(AC,y), np.nan) - PC.get(col_for_year(PC,y), np.nan)

liq_kinds = {"Liquidez Corriente":"ratio", "Prueba Ácida":"ratio", "Cash Ratio":"ratio", "Capital de Trabajo":"number"}
data_blocks.append(("LIQUIDEZ", liq, liq_kinds))

# 2) SOLVENCIA Y ESTRUCTURA
solv = {}
for y in years:
    solv.setdefault("Endeudamiento (D/E)", {})[str(y)] = PT.get(col_for_year(PT,y), np.nan) / Patr.get(col_for_year(Patr,y), np.nan)
    solv.setdefault("Apalancamiento (D/A)", {})[str(y)] = PT.get(col_for_year(PT,y), np.nan) / AT.get(col_for_year(AT,y), np.nan)
    # Cobertura intereses = EBIT / |Intereses| (intereses suelen ser negativos en P&L; forzamos abs)
    intereses_val = Interes.get(col_for_year(Interes,y), np.nan)
    intereses_abs = np.abs(intereses_val) if not np.isnan(intereses_val) else np.nan
    solv.setdefault("Cobertura de Intereses", {})[str(y)] = (EBIT.get(col_for_year(EBIT,y), np.nan)) / (intereses_abs if intereses_abs not in (0, np.nan) else np.nan)
    # EBITDA = EBIT + D&A si D&A disponible; de lo contrario, usa EBIT (marcará igual, pero no None)
    ebitda_y = (EBIT.get(col_for_year(EBIT,y), np.nan) + DA.get(col_for_year(DA,y), 0.0))
    solv.setdefault("Deuda / EBITDA", {})[str(y)] = PT.get(col_for_year(PT,y), np.nan) / (ebitda_y if ebitda_y not in (0,np.nan) else np.nan)
    solv.setdefault("Autonomía Financiera", {})[str(y)] = Patr.get(col_for_year(Patr,y), np.nan) / AT.get(col_for_year(AT,y), np.nan)

solv_kinds = {"Endeudamiento (D/E)":"ratio","Apalancamiento (D/A)":"ratio","Cobertura de Intereses":"ratio","Deuda / EBITDA":"ratio","Autonomía Financiera":"pct"}
data_blocks.append(("SOLVENCIA Y ESTRUCTURA", solv, solv_kinds))

# 3) RENTABILIDAD
rent = {}
for y in years:
    at_avg = avg_bal_item(AT, y)
    patr_avg = avg_bal_item(Patr, y)
    rent.setdefault("Margen Bruto", {})[str(y)] = Bruta.get(col_for_year(Bruta,y), np.nan) / Ventas.get(col_for_year(Ventas,y), np.nan)
    rent.setdefault("Margen Operativo (EBIT)", {})[str(y)] = EBIT.get(col_for_year(EBIT,y), np.nan) / Ventas.get(col_for_year(Ventas,y), np.nan)
    # EBITDA margin
    ebitda_y = (EBIT.get(col_for_year(EBIT,y), np.nan) + DA.get(col_for_year(DA,y), 0.0))
    rent.setdefault("Margen EBITDA", {})[str(y)] = ebitda_y / Ventas.get(col_for_year(Ventas,y), np.nan)
    rent.setdefault("Margen Neto", {})[str(y)] = Neta.get(col_for_year(Neta,y), np.nan) / Ventas.get(col_for_year(Ventas,y), np.nan)
    rent.setdefault("ROE", {})[str(y)] = Neta.get(col_for_year(Neta,y), np.nan) / (patr_avg if patr_avg not in (0,np.nan) else np.nan)
    rent.setdefault("ROA", {})[str(y)] = Neta.get(col_for_year(Neta,y), np.nan) / (at_avg if at_avg not in (0,np.nan) else np.nan)
    # ROIC (aprox): NOPAT ~ EBIT*(1 - tasa). Sin tasa, omitimos NOPAT -> dejar None
    rent.setdefault("ROIC", {})[str(y)] = np.nan

rent_kinds = {k:"pct" for k in ["Margen Bruto","Margen Operativo (EBIT)","Margen EBITDA","Margen Neto","ROE","ROA","ROIC"]}
data_blocks.append(("RENTABILIDAD", rent, rent_kinds))

# 4) EFICIENCIA OPERATIVA
eff = {}
for y in years:
    at_avg = avg_bal_item(AT, y)
    inv_avg = avg_bal_item(Inv, y)
    cxc_avg = avg_bal_item(CxC, y)
    cxp_avg = avg_bal_item(CxP, y)
    ventas_y = Ventas.get(col_for_year(Ventas,y), np.nan)
    cogs_y = COGS.get(col_for_year(COGS,y), np.nan)
    compras_y = compras_aprox(y)

    eff.setdefault("Rotación de Activos", {})[str(y)] = ventas_y / (at_avg if at_avg not in (0,np.nan) else np.nan)
    eff.setdefault("Rotación de Inventarios", {})[str(y)] = cogs_y / (inv_avg if inv_avg not in (0,np.nan) else np.nan)
    eff.setdefault("Días de Inventario", {})[str(y)] = 365.0 / (eff["Rotación de Inventarios"][str(y)] if eff["Rotación de Inventarios"][str(y)] not in (0,np.nan) else np.nan)
    eff.setdefault("Rotación de Cuentas por Cobrar", {})[str(y)] = ventas_y / (cxc_avg if cxc_avg not in (0,np.nan) else np.nan)
    eff.setdefault("Período Promedio de Cobro", {})[str(y)] = 365.0 / (eff["Rotación de Cuentas por Cobrar"][str(y)] if eff["Rotación de Cuentas por Cobrar"][str(y)] not in (0,np.nan) else np.nan)
    eff.setdefault("Rotación de Cuentas por Pagar", {})[str(y)] = (compras_y if compras_y not in (0,np.nan) else np.nan) / (cxp_avg if cxp_avg not in (0,np.nan) else np.nan)
    eff.setdefault("Período Promedio de Pago", {})[str(y)] = 365.0 / (eff["Rotación de Cuentas por Pagar"][str(y)] if eff["Rotación de Cuentas por Pagar"][str(y)] not in (0,np.nan) else np.nan)
    # Ciclo de Conversión de Efectivo
    cie = (eff["Días de Inventario"][str(y)] if not np.isnan(eff["Días de Inventario"][str(y)]) else np.nan) \
          + (eff["Período Promedio de Cobro"][str(y)] if not np.isnan(eff["Período Promedio de Cobro"][str(y)]) else np.nan) \
          - (eff["Período Promedio de Pago"][str(y)] if not np.isnan(eff["Período Promedio de Pago"][str(y)]) else np.nan)
    eff.setdefault("Ciclo de Conversión de Efectivo", {})[str(y)] = cie

eff_kinds = {
    "Rotación de Activos":"ratio",
    "Rotación de Inventarios":"ratio",
    "Días de Inventario":"days",
    "Rotación de Cuentas por Cobrar":"ratio",
    "Período Promedio de Cobro":"days",
    "Rotación de Cuentas por Pagar":"ratio",
    "Período Promedio de Pago":"days",
    "Ciclo de Conversión de Efectivo":"days",
}
data_blocks.append(("EFICIENCIA OPERATIVA", eff, eff_kinds))

# 5) FLUJOS Y ADICIONALES
flw = {}
for y in years:
    flw.setdefault("Conversión de caja (CFO/Utilidad Neta)", {})[str(y)] = CFO.get(col_for_year(CFO,y), np.nan) / (Neta.get(col_for_year(Neta,y), np.nan) if Neta.get(col_for_year(Neta,y), np.nan) not in (0,np.nan) else np.nan)
    flw.setdefault("Free Cash Flow (CFO - CAPEX)", {})[str(y)] = FCF.get(col_for_year(FCF,y), np.nan)
    flw.setdefault("AC / AT", {})[str(y)] = AC.get(col_for_year(AC,y), np.nan) / AT.get(col_for_year(AT,y), np.nan)
    flw.setdefault("PC / PT", {})[str(y)] = PC.get(col_for_year(PC,y), np.nan) / PT.get(col_for_year(PT,y), np.nan)

flw_kinds = {"Conversión de caja (CFO/Utilidad Neta)":"ratio","Free Cash Flow (CFO - CAPEX)":"number","AC / AT":"pct","PC / PT":"pct"}
data_blocks.append(("FLUJOS Y ADICIONALES", flw, flw_kinds))

# Build output table with sections and styling
wb = load_workbook(base_file)
sheet_name = "Análisis Avanzado"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name, 0)

# Styles
header_fill = PatternFill("solid", fgColor="0B2447")
subheader_fill = PatternFill("solid", fgColor="19376D")
section_fills = {
    "LIQUIDEZ": PatternFill("solid", fgColor="D1E7DD"),
    "SOLVENCIA Y ESTRUCTURA": PatternFill("solid", fgColor="FAD7A0"),
    "RENTABILIDAD": PatternFill("solid", fgColor="F8D7DA"),
    "EFICIENCIA OPERATIVA": PatternFill("solid", fgColor="D6EAF8"),
    "FLUJOS Y ADICIONALES": PatternFill("solid", fgColor="E8DAEF"),
}
band_fill = PatternFill("solid", fgColor="F7F9FC")
bold_white = Font(bold=True, color="FFFFFF", size=13)
bold_white_small = Font(bold=True, color="FFFFFF", size=11)
bold_dark = Font(bold=True, color="000000", size=11)
normal = Font(color="000000", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Columns: Indicador + years + Último + Promedio + Tendencia
num_years = len(years)
cols_total = 1 + num_years + 2 + 1
for c in range(1, cols_total+1):
    ws.column_dimensions[get_column_letter(c)].width = 30 if c==1 else 14

# Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_total)
ws.cell(row=1, column=1, value="Análisis Financiero – Ratios y Evolución").fill = header_fill
ws.cell(row=1, column=1).font = bold_white
ws.cell(row=1, column=1).alignment = center

# Subheader
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_total)
ws.cell(row=2, column=1, value="Fechas: Balance (AAAA-12), Resultados (AAAA-01), Flujos (AAAA-01). Evolución alineada por AÑO.").fill = subheader_fill
ws.cell(row=2, column=1).font = bold_white_small
ws.cell(row=2, column=1).alignment = center

# Header row (names)
header_row = 4
headers = ["Indicador"] + [str(y) for y in years] + ["Último", "Promedio", "Tendencia"]
for i,h in enumerate(headers, start=1):
    ws.cell(row=header_row, column=i, value=h).font = bold_white_small
    ws.cell(row=header_row, column=i).fill = subheader_fill
    ws.cell(row=header_row, column=i).alignment = center
    ws.cell(row=header_row, column=i).border = border

r = header_row + 1

def write_block(title, block_dict, kinds):
    global r
    # Section header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols_total)
    sec = ws.cell(row=r, column=1, value=title)
    sec.font = bold_dark
    sec.alignment = left
    sec.fill = section_fills.get(title, band_fill)
    for c in range(1, cols_total+1):
        ws.cell(row=r, column=c).border = border
    r += 1

    # Rows
    for ratio_name, series_map in block_dict.items():
        ws.cell(row=r, column=1, value=ratio_name).font = normal
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=1).border = border

        row_vals = []
        for j, y in enumerate(years, start=2):
            v = series_map.get(str(y), np.nan)
            row_vals.append(v)
            c = ws.cell(row=r, column=j, value=None if pd.isna(v) else float(v))
            kind = kinds.get(ratio_name, "ratio")
            if kind == "pct":
                c.number_format = "0.0%"
            elif kind == "number":
                c.number_format = "#,##0"
            elif kind == "days":
                c.number_format = "0"
            else:  # ratio
                c.number_format = "0.00"
            c.alignment = center
            c.border = border
        # Último & Promedio
        last_val = next((v for v in row_vals[::-1] if not pd.isna(v)), np.nan)
        avg_val = np.nanmean([v for v in row_vals if not pd.isna(v)]) if any([not pd.isna(v) for v in row_vals]) else np.nan
        col_last = 1 + num_years + 1
        col_avg  = 1 + num_years + 2
        c_last = ws.cell(row=r, column=col_last, value=None if pd.isna(last_val) else float(last_val))
        c_avg  = ws.cell(row=r, column=col_avg,  value=None if pd.isna(avg_val)  else float(avg_val))
        kind = kinds.get(ratio_name, "ratio")
        for cc in (c_last, c_avg):
            if kind == "pct":
                cc.number_format = "0.0%"
            elif kind == "number":
                cc.number_format = "#,##0"
            elif kind == "days":
                cc.number_format = "0"
            else:
                cc.number_format = "0.00"
            cc.alignment = center
            cc.border = border

        # Tendencia (último vs anterior)
        trend_col = 1 + num_years + 3
        arrow = "→"
        try:
            last_idx = next(i for i in range(len(row_vals)-1, -1, -1) if not pd.isna(row_vals[i]))
            prev_idx = next(i for i in range(last_idx-1, -1, -1) if not pd.isna(row_vals[i]))
            if not pd.isna(row_vals[last_idx]) and not pd.isna(row_vals[prev_idx]):
                if row_vals[last_idx] > row_vals[prev_idx]: arrow = "▲"
                elif row_vals[last_idx] < row_vals[prev_idx]: arrow = "▼"
        except StopIteration:
            arrow = "→"
        ws.cell(row=r, column=trend_col, value=arrow).alignment = center
        ws.cell(row=r, column=trend_col).border = border

        r += 1

# Write all blocks
for title, block, kinds in data_blocks:
    write_block(title, block, kinds)

# Heatmaps per block across the year columns only
def add_heatmap(r_start, r_end):
    year_start_col = 2
    year_end_col = 1 + num_years
    rng = f"{get_column_letter(year_start_col)}{r_start}:{get_column_letter(year_end_col)}{r_end}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="percentile", start_value=5, start_color="FDE725",
            mid_type="percentile", mid_value=50, mid_color="5DC863",
            end_type="percentile", end_value=95, end_color="2A788E"
        )
    )

# Identify ranges per block to apply heatmaps
row_ptr = header_row + 1
for title, block, kinds in data_blocks:
    # section row then len(block) rows
    sec_row = row_ptr
    start_rows = sec_row + 1
    end_rows = sec_row + len(block)
    add_heatmap(start_rows, end_rows)
    row_ptr = end_rows + 1

# Data bars on Último & Promedio (whole sheet)
last_col_letter = get_column_letter(1 + num_years + 1)
avg_col_letter  = get_column_letter(1 + num_years + 2)
total_rows = ws.max_row
for col_letter in (last_col_letter, avg_col_letter):
    ws.conditional_formatting.add(
        f"{col_letter}{header_row+1}:{col_letter}{total_rows}",
        DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
    )

# Freeze panes
ws.freeze_panes = "B5"

# Save
out_path = "./data/demo/FinDataChile_Data_Demo_con_Analisis_SEGMENTADO.xlsx"
Path(out_path).parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)


