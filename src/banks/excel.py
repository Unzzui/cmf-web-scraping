"""Generación de los Excel de producto de bancos desde las tablas bank_*.

Un libro por banco con la misma anatomía y la misma estética que los productos IFRS de
``Product_v1``: reutiliza ``cmf_extract.excel_style`` (paleta, tipografía, formatos y sus
guardas de contraste) en vez de replicarla, porque dos catálogos que se ven distinto se
ven mal.

Hojas: Inicio · Balance/Resultados por época · RATIOS & KPIs · Perfil · Ficha Técnica ·
METODOLOGÍA.

Dos diferencias que impone el dato de bancos:

* **Columnas mensuales, no trimestrales.** La CMF publica los bancos mes a mes.
* **Hojas separadas por época.** El Compendio de 2022 cambia el plan de cuentas *y* la
  unidad (MMCLP -> CLP): pre-2022 la cuenta de activos es ``1000000 ACTIVOS`` y desde 2022
  es ``100000000 TOTAL ACTIVOS``. No son la misma serie y no se concatenan; unirlas
  exigiría inventar un mapeo cuenta-vieja -> cuenta-nueva que la CMF nunca publicó.

Los ratios sólo se calculan sobre la época Compendio (ver ``src/banks/ratios.py``).
"""

import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from cmf_extract import excel_style as st  # noqa: E402
from src.banks import ratios as R  # noqa: E402

HOJA_BALANCE_NUEVA = "Balance 2022+"
HOJA_BALANCE_VIEJA = "Balance hasta 2021"
HOJA_RESULTADO_NUEVA = "Estado de Resultados 2022+"
HOJA_RESULTADO_VIEJA = "Estado de Resultados hasta 2021"
HOJA_RATIOS = "RATIOS & KPIs"

_EPOCAS = (
    # (epoch, unidad, hoja_balance, hoja_resultado)
    ("compendio_2022", "CLP", HOJA_BALANCE_NUEVA, HOJA_RESULTADO_NUEVA),
    ("pre_2022", "MMCLP", HOJA_BALANCE_VIEJA, HOJA_RESULTADO_VIEJA),
)

_ANCHO_CUENTA = 62
_ANCHO_PERIODO = 15
_FILA_CABECERA = 5   # las tablas arrancan con cabecera en la fila 5 y datos en la 6


@dataclass
class BankBook:
    codigo: str
    nombre: str
    rut: str | None
    periodos: list[tuple[int, int]]


def normalizar_rut(rut: str | None) -> str | None:
    """'97.004.000-5' -> '97004000-5', que es el formato que usan los productos IFRS."""
    if not rut:
        return None
    return rut.replace(".", "").strip().upper()


def nombre_archivo(nombre_banco: str, rut: str | None, periodos: list[tuple[int, int]]) -> str:
    """'BANCO DE CHILE - 97004000-5 - Análisis Financiero 2014-2026 [ES].xlsx'.

    Es el formato que parsea FinDataChile para deducir empresa/RUT/años (ver
    ``findatachile_uploader.parse_file_info``); sin trimestre, porque el dato de bancos es
    mensual y el server lee el rango anual como period_type='completo'.
    """
    ini = periodos[0][0] if periodos else 0
    fin = periodos[-1][0] if periodos else 0
    return (f"{nombre_banco} - {normalizar_rut(rut) or 'SIN-RUT'} - "
            f"Análisis Financiero {ini}-{fin} [ES].xlsx")


def etiqueta_periodo(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


# ── Lectura ───────────────────────────────────────────────────────────────────
def fetch_periodos(conn, cod: str) -> list[tuple[int, int]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT DISTINCT period_year, period_month FROM bank_financial_data
            WHERE codigo_institucion = %s ORDER BY period_year, period_month
            """,
            (cod,),
        )
        return [(r[0], r[1]) for r in cur.fetchall()]


def fetch_rut(conn, cod: str) -> str | None:
    """RUT desde el perfil más reciente: bank_institutions.rut está sin poblar."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT rut FROM bank_profiles
            WHERE codigo_institucion = %s AND rut IS NOT NULL
            ORDER BY period_year DESC, period_month DESC LIMIT 1
            """,
            (cod,),
        )
        row = cur.fetchone()
        return row[0] if row else None


def fetch_matriz(conn, cod: str, statement: str, epoch: str):
    """(cuentas, valores) de un statement/época: cuentas en orden de plan de cuentas."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT a.codigo_cuenta, a.descripcion_cuenta,
                   f.period_year, f.period_month, f.moneda_total
            FROM bank_financial_data f
            JOIN bank_accounts a ON a.id = f.account_id
            WHERE f.codigo_institucion = %s AND a.statement = %s AND f.taxonomy_epoch = %s
            ORDER BY a.codigo_cuenta, f.period_year, f.period_month
            """,
            (cod, statement, epoch),
        )
        cuentas: dict[str, str] = {}
        valores: dict[tuple[str, tuple[int, int]], float] = {}
        for codigo, desc, y, m, total in cur.fetchall():
            cuentas[codigo] = desc
            valores[(codigo, (y, m))] = total
    return list(cuentas.items()), valores


# ── Escritura ─────────────────────────────────────────────────────────────────
def _cabecera_hoja(ws, titulo: str, subtitulo: str) -> None:
    ws.cell(1, 1, titulo).font = st.SECCION
    ws.cell(2, 1, subtitulo).font = st.NOTA


def _escribir_estado(ws, titulo: str, unidad: str, banco: str, cuentas, valores,
                     periodos) -> dict[str, int]:
    """Escribe un estado financiero. Devuelve {codigo_cuenta: fila} para los ratios."""
    _cabecera_hoja(ws, f"{titulo} — {banco}",
                   f"Unidad: {unidad}    •    Períodos: "
                   f"{etiqueta_periodo(*periodos[0])} - {etiqueta_periodo(*periodos[-1])}"
                   if periodos else "sin datos")

    cols = list(reversed(periodos))  # más nuevo primero, como en Product_v1
    c = ws.cell(_FILA_CABECERA, 1, "Cuenta")
    c.font, c.fill, c.alignment = st.CABECERA, st.RELLENO_TINTA, st.IZQ
    for i, (y, m) in enumerate(cols, start=2):
        c = ws.cell(_FILA_CABECERA, i, etiqueta_periodo(y, m))
        c.font, c.fill, c.alignment = st.CABECERA, st.RELLENO_TINTA, st.DER

    filas: dict[str, int] = {}
    for n, (codigo, desc) in enumerate(cuentas):
        fila = _FILA_CABECERA + 1 + n
        filas[codigo] = fila
        # Los códigos de primer nivel (…000000 / …0000) son subtotales del plan: van en
        # negrita para que la jerarquía se lea sin tener que descifrar el código.
        es_total = codigo.endswith("000000") or (len(codigo) == 7 and codigo.endswith("0000"))
        celda = ws.cell(fila, 1, f"{codigo}  {desc}")
        celda.font = st.CUERPO_FUERTE if es_total else st.CUERPO
        celda.alignment = st.IZQ
        if n % 2:
            celda.fill = st.RELLENO_SUAVE
        for i, per in enumerate(cols, start=2):
            v = valores.get((codigo, per))
            cel = ws.cell(fila, i)
            if v is not None:
                cel.value = float(v)
            cel.font = st.CUERPO_FUERTE if es_total else st.CUERPO
            cel.number_format = st.FMT_NUM
            cel.alignment = st.DER
            if n % 2:
                cel.fill = st.RELLENO_SUAVE

    ws.column_dimensions["A"].width = _ANCHO_CUENTA
    for i in range(2, len(cols) + 2):
        ws.column_dimensions[ws.cell(_FILA_CABECERA, i).column_letter].width = _ANCHO_PERIODO
    st.preparar_hoja(ws, congelar=f"B{_FILA_CABECERA + 1}")
    return filas


def _escribir_ratios(ws, banco: str, periodos, filas_balance, filas_resultado) -> None:
    """Hoja de ratios con fórmulas vivas que apuntan a las hojas de estados."""
    from openpyxl.utils import get_column_letter

    _cabecera_hoja(ws, f"Ratios y KPIs — {banco}",
                   "Indicadores del negocio bancario. Los que cruzan un flujo (resultado, "
                   "acumulado del año) con un stock (balance) van anualizados x12/mes.")

    cols = list(reversed(periodos))
    c = ws.cell(_FILA_CABECERA, 1, "Indicador")
    c.font, c.fill, c.alignment = st.CABECERA, st.RELLENO_TINTA, st.IZQ
    for i, (y, m) in enumerate(cols, start=2):
        c = ws.cell(_FILA_CABECERA, i, etiqueta_periodo(y, m))
        c.font, c.fill, c.alignment = st.CABECERA, st.RELLENO_TINTA, st.DER

    fila = _FILA_CABECERA + 1
    for categoria in R.CATEGORIAS:
        cel = ws.cell(fila, 1, categoria.upper())
        cel.font, cel.alignment = st.ETIQUETA, st.IZQ
        fila += 1
        for ratio in (r for r in R.RATIOS if r.categoria == categoria):
            ws.cell(fila, 1, ratio.nombre).font = st.CUERPO
            ws.cell(fila, 1).alignment = st.IZQ
            for i, (y, m) in enumerate(cols, start=2):
                letra = get_column_letter(i)

                def ref(cuenta, _l=letra):
                    statement, codigo = cuenta
                    hoja = HOJA_BALANCE_NUEVA if statement == "balance" else HOJA_RESULTADO_NUEVA
                    mapa = filas_balance if statement == "balance" else filas_resultado
                    f = mapa.get(codigo)
                    return f"'{hoja}'!{_l}{f}" if f else "NA()"

                cel = ws.cell(fila, i, R.construir_formula(ratio, ref, m))
                cel.number_format = st.FMT_PCT if ratio.formato == "pct" else st.FMT_MULT
                cel.font, cel.alignment = st.CUERPO, st.DER
            fila += 1

    ws.column_dimensions["A"].width = 38
    for i in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(i)].width = _ANCHO_PERIODO
    st.preparar_hoja(ws, congelar=f"B{_FILA_CABECERA + 1}")


def _escribir_inicio(ws, banco: str, cod: str, rut: str | None, periodos, hojas) -> None:
    ws.cell(1, 1, "FinData Chile").font = st.ETIQUETA
    ws.cell(2, 1, f"ANÁLISIS FINANCIERO: {banco}").font = st.TITULO
    ws.cell(4, 1, "INFORMACIÓN CORPORATIVA").font = st.ETIQUETA
    datos = [
        ("Institución", banco),
        ("RUT", normalizar_rut(rut) or "—"),
        ("Código CMF", cod),
        ("Mercado", "Chile — Sistema bancario"),
        ("Períodos", f"{etiqueta_periodo(*periodos[0])} - {etiqueta_periodo(*periodos[-1])}"
                     if periodos else "—"),
        ("Frecuencia", "Mensual"),
    ]
    for i, (k, v) in enumerate(datos, start=5):
        ws.cell(i, 1, k).font = st.NOTA
        ws.cell(i, 2, v).font = st.CUERPO

    f = 5 + len(datos) + 1
    ws.cell(f, 1, "NAVEGACIÓN RÁPIDA").font = st.ETIQUETA
    for i, hoja in enumerate(hojas, start=f + 1):
        cel = ws.cell(i, 1, hoja)
        cel.font = st.fuente(10, color=st.EMBER)
        cel.hyperlink = f"#'{hoja}'!A1"

    f = f + len(hojas) + 2
    ws.cell(f, 1, "RESUMEN EJECUTIVO").font = st.ETIQUETA
    cel = ws.cell(f + 1, 1,
                  "Estados financieros mensuales publicados por la CMF (API de Bancos), con "
                  "ratios del negocio bancario calculados con fórmulas vivas sobre las hojas "
                  "de estados. El Compendio de Normas Contables (enero 2022) cambió el plan "
                  "de cuentas y la unidad monetaria, por lo que las series anterior y "
                  "posterior se publican en hojas separadas y no son comparables directamente.")
    cel.font, cel.alignment = st.NOTA, st.AJUSTE
    ws.merge_cells(start_row=f + 1, start_column=1, end_row=f + 4, end_column=6)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    st.preparar_hoja(ws, congelar=None)


def _escribir_perfil(ws, banco: str, cod: str, rut: str | None, conn) -> None:
    _cabecera_hoja(ws, f"Perfil — {banco}", "Ficha institucional del período más reciente")
    filas = [("Código institución CMF", cod), ("RUT", normalizar_rut(rut) or "—")]
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT period_year, period_month, direccion, telefono, sitio_web,
                   sucursales, empleados, codigo_swift
            FROM bank_profiles WHERE codigo_institucion = %s
            ORDER BY period_year DESC, period_month DESC LIMIT 1
            """,
            (cod,),
        )
        row = cur.fetchone()
    if row:
        y, m, direccion, telefono, web, sucursales, empleados, swift = row
        filas += [
            ("Perfil al período", etiqueta_periodo(y, m)), ("Dirección", direccion),
            ("Teléfono", telefono), ("Sitio web", web), ("Sucursales", sucursales),
            ("Empleados", empleados), ("SWIFT", swift),
        ]
    for i, (k, v) in enumerate(filas, start=_FILA_CABECERA):
        ws.cell(i, 1, k).font = st.NOTA
        ws.cell(i, 2, v if v is not None else "—").font = st.CUERPO
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 54
    st.preparar_hoja(ws, congelar=None)


def _escribir_ficha(ws, banco: str, cod: str, rut: str | None, periodos) -> None:
    _cabecera_hoja(ws, "FICHA TÉCNICA", "Resumen del archivo generado")
    filas = [
        ("Institución", banco), ("RUT", normalizar_rut(rut) or "—"), ("Código CMF", cod),
        ("Fuente", "CMF Chile — API de Bancos (api-sbifv3)"),
        ("Períodos", f"{etiqueta_periodo(*periodos[0])} - {etiqueta_periodo(*periodos[-1])}"
                     if periodos else "—"),
        ("Frecuencia", "Mensual"),
        ("Unidad", "CLP desde 2022-01 · MMCLP hasta 2021-12"),
        ("Web", "findatachile.com"),
    ]
    for i, (k, v) in enumerate(filas, start=_FILA_CABECERA):
        ws.cell(i, 1, k).font = st.NOTA
        ws.cell(i, 2, v).font = st.CUERPO
    f = _FILA_CABECERA + len(filas) + 1
    cel = ws.cell(f, 1,
                  "Datos publicados por la CMF. La adecuación de capital no se incluye: la "
                  "CMF dejó de publicarla en esta API en diciembre de 2020, con el paso a "
                  "Basilea III.")
    cel.font, cel.alignment = st.LEGAL, st.AJUSTE
    ws.merge_cells(start_row=f, start_column=1, end_row=f + 2, end_column=5)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 54
    st.preparar_hoja(ws, congelar=None)


def _escribir_metodologia(ws) -> None:
    _cabecera_hoja(ws, "METODOLOGÍA", "Cómo se calcula cada indicador de la hoja RATIOS & KPIs")
    for i, txt in enumerate(("Ratio / Indicador", "Categoría", "Fórmula")):
        c = ws.cell(_FILA_CABECERA, i + 1, txt)
        c.font, c.fill, c.alignment = st.CABECERA, st.RELLENO_TINTA, st.IZQ

    fila = _FILA_CABECERA + 1
    for categoria in R.CATEGORIAS:
        c = ws.cell(fila, 1, categoria.upper())
        c.font, c.alignment = st.ETIQUETA, st.IZQ
        fila += 1
        for ratio in (r for r in R.RATIOS if r.categoria == categoria):
            ws.cell(fila, 1, ratio.nombre).font = st.CUERPO
            ws.cell(fila, 2, ratio.categoria).font = st.NOTA
            ws.cell(fila, 3, ratio.formula_texto).font = st.CUERPO
            for cc in (1, 2, 3):
                ws.cell(fila, cc).alignment = st.IZQ
            fila += 1

    fila += 1
    for txt in (
        "Anualización: el estado de resultados de la CMF es acumulado del ejercicio (YTD), "
        "así que a mayo son cinco meses. Los ratios que cruzan un flujo con un stock del "
        "balance se anualizan multiplicando por 12/mes; los que cruzan dos flujos del mismo "
        "período (p. ej. eficiencia) no lo necesitan, porque el factor se cancela.",
        "Signos: los gastos y las provisiones se publican en negativo; las fórmulas usan "
        "valor absoluto donde el indicador se lee como magnitud.",
        "Alcance: los ratios se calculan sólo sobre la época del Compendio (desde 2022-01). "
        "El plan de cuentas anterior usa otros códigos y no publica los subtotales que estos "
        "indicadores requieren.",
    ):
        c = ws.cell(fila, 1, txt)
        c.font, c.alignment = st.NOTA, st.AJUSTE
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila + 2, end_column=5)
        fila += 4

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 58
    st.preparar_hoja(ws, congelar=None)


# ── Orquestación ──────────────────────────────────────────────────────────────
def construir_libro(conn, cod: str, nombre: str):
    """Arma el Workbook completo de un banco. Devuelve (workbook, BankBook)."""
    from openpyxl import Workbook

    periodos = fetch_periodos(conn, cod)
    rut = fetch_rut(conn, cod)
    wb = Workbook()
    wb.remove(wb.active)
    ws_inicio = wb.create_sheet("Inicio")

    filas: dict[str, dict[str, int]] = {}
    pers_compendio: list[tuple[int, int]] = []
    for epoch, unidad, hoja_bal, hoja_res in _EPOCAS:
        datos = {}
        for statement, titulo, hoja in (("balance", "Balance", hoja_bal),
                                        ("resultado", "Estado de Resultados", hoja_res)):
            cuentas, valores = fetch_matriz(conn, cod, statement, epoch)
            if cuentas:
                datos[hoja] = (titulo, cuentas, valores)
        if not datos:
            continue

        # Eje temporal ÚNICO para las dos hojas de la época: la UNIÓN de sus períodos, no
        # los de cada una por separado. La CMF tiene huecos de un solo report (p.ej. el
        # resultado de 2023-07 da 500 mientras su balance da 200), y con ejes propios la
        # hoja con el hueco corre todas sus columnas una posición: los ratios, que usan la
        # misma letra de columna en ambas hojas, terminarían cruzando el balance de un mes
        # con el resultado de OTRO. Silencioso y plausible — el peor tipo de error.
        # Alineadas por construcción, el hueco queda como celda vacía y el IFERROR lo tapa.
        pers = [p for p in periodos
                if any((c, p) in vals for _, cuentas, vals in datos.values() for c, _ in cuentas)]
        for hoja, (titulo, cuentas, valores) in datos.items():
            ws = wb.create_sheet(hoja)
            filas[hoja] = _escribir_estado(ws, f"{titulo} ({unidad})", unidad, nombre,
                                           cuentas, valores, pers)
        if epoch == "compendio_2022":
            pers_compendio = pers

    # Los ratios necesitan ambas hojas del Compendio: sin una de las dos, media fórmula
    # apuntaría a una hoja inexistente y el libro abriría con #REF!.
    if HOJA_BALANCE_NUEVA in filas and HOJA_RESULTADO_NUEVA in filas and pers_compendio:
        _escribir_ratios(wb.create_sheet(HOJA_RATIOS), nombre, pers_compendio,
                         filas[HOJA_BALANCE_NUEVA], filas[HOJA_RESULTADO_NUEVA])

    _escribir_perfil(wb.create_sheet("Perfil"), nombre, cod, rut, conn)
    _escribir_ficha(wb.create_sheet("Ficha Técnica"), nombre, cod, rut, periodos)
    _escribir_metodologia(wb.create_sheet("METODOLOGÍA"))

    hojas = [h for h in wb.sheetnames if h != "Inicio"]
    _escribir_inicio(ws_inicio, nombre, cod, rut, periodos, hojas)

    st.aplicar_tipografia_base(wb)
    return wb, BankBook(cod, nombre, rut, periodos)
