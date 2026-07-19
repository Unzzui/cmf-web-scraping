#!/usr/bin/env python3
"""Genera el `estados_<id>_<rango>_es.xlsx` de una empresa de EEUU DESDE LA BD.

Las empresas chilenas nacen del XBRL de la CMF → `xbrl_to_excel.py` arma los estados. Las de
EEUU ya están normalizadas en la BD (ingesta EDGAR, con los MISMOS labels ES que Chile). Este
script reconstruye el mismo `estados_*.xlsx` de 3 hojas (Balance General / Estado de Resultados
/ Flujo Efectivo) leyendo `financial_line_items` (label, role_code, display_order) +
`financial_data` (valor por período). Alimentado a `run_products_analysis.py`, produce un
análisis IDÉNTICO al chileno (RATIOS, DCF, etc.), porque el análisis busca por label.

El identificador del archivo es el CIK SIN ceros a la izquierda (encaja en el patrón de RUT
`\\d{4,8}` del pipeline). El nombre de la empresa va en el título de cada hoja, igual que Chile.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from cmf_extract import excel_style as est  # noqa: E402

FDC = Path(os.environ.get("FDC_DIR", "/home/unzzui/Proyectos/FinDataChile"))

# category (financial_line_items) → (nombre de hoja del pipeline, título)
_HOJAS = [
    ("balance_sheet", "Balance General"),
    ("income_statement", "Estado de Resultados"),
    ("cash_flow", "Flujo Efectivo"),
]


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for line in path.read_text().splitlines():
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def periodos(conn, company_id: int) -> list[tuple[int, int]]:
    """(year, quarter) disponibles, MÁS RECIENTE PRIMERO (igual que el estados chileno)."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT DISTINCT period_year, period_quarter FROM financial_data "
            "WHERE company_id = %s ORDER BY period_year DESC, period_quarter DESC",
            [company_id])
        return [(int(y), int(q)) for y, q in cur.fetchall()]


def filas(conn, company_id: int, category: str) -> list[tuple[str, int, str, dict]]:
    """[(label, display_order, subcategory, {(year,quarter): value})], ordenadas."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT fli.label, fli.display_order, COALESCE(fli.subcategory,''),
                   fd.period_year, fd.period_quarter, fd.value
            FROM financial_line_items fli
            JOIN financial_data fd ON fd.line_item_id = fli.id
            WHERE fli.company_id = %s AND fli.category = %s AND fli.is_active
              AND fd.value IS NOT NULL
            ORDER BY fli.display_order, fd.period_year DESC
            """,
            [company_id, category])
        rows = cur.fetchall()
    por_label: dict[str, tuple[int, str, dict]] = {}
    for label, order, sub, y, q, val in rows:
        entry = por_label.setdefault(label, (int(order or 0), sub, {}))
        entry[2][(int(y), int(q))] = float(val)
    return sorted(((lbl, o, sub, vals) for lbl, (o, sub, vals) in por_label.items()),
                  key=lambda t: t[1])


def etiqueta_periodo(y: int, q: int) -> str:
    return f"{y}Q{q}" if q else str(y)


def empresa(conn, company_id: int) -> tuple[str, str]:
    with conn.cursor() as cur:
        cur.execute("SELECT razon_social, COALESCE(financial_statements_currency,'USD') "
                    "FROM companies WHERE id = %s", [company_id])
        r = cur.fetchone()
    return (r[0] if r else f"US-{company_id}", r[1] if r else "USD")


def construir(conn, company_id: int, cik: str, out_dir: Path) -> Path | None:
    nombre, moneda = empresa(conn, company_id)
    pers = periodos(conn, company_id)
    if not pers:
        return None
    labels_periodo = [etiqueta_periodo(y, q) for y, q in pers]
    years = [y for y, _ in pers]
    rango = f"{max(years)}-{min(years)}"

    # Estilos = los MISMOS que el estados chileno (excel_style): Inter, encabezado en tinta
    # (#0B0D12) con texto papel, secciones de subcategoría destacadas, cifras en #,##0.
    title_font = est.fuente(16, bold=True, color=est.PAPER)
    unit_font = est.fuente(11, color=est.INK)
    hdr_font = est.fuente(11, bold=True, color=est.PAPER)
    sec_font = est.fuente(11, bold=True, color=est.PAPER)
    label_font = est.fuente(11, color=est.INK)
    val_font = est.fuente(11, color=est.INK)
    tinta = est.RELLENO_TINTA
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)
    ncols = 1 + len(pers)
    for category, hoja in _HOJAS:
        ws = wb.create_sheet(hoja)
        # Fila 1: título (el análisis lee el nombre de acá).
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws.cell(row=1, column=1, value=f"{hoja} — {nombre}")
        t.font, t.fill, t.alignment = title_font, tinta, center
        ws.row_dimensions[1].height = 26
        # Fila 2: unidad (el DCF detecta la moneda; "Miles USD").
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        u = ws.cell(row=2, column=1, value=f"Unidad: Miles {moneda}    •  Fuente: SEC/EDGAR")
        u.font, u.alignment = unit_font, center
        # Fila 3: encabezados "Cuenta" + períodos (más reciente primero).
        h = ws.cell(row=3, column=1, value="Cuenta")
        h.font, h.fill, h.alignment = hdr_font, tinta, center_wrap
        for c, lab in enumerate(labels_periodo, start=2):
            hc = ws.cell(row=3, column=c, value=lab)
            hc.font, hc.fill, hc.alignment = hdr_font, tinta, center_wrap
        ws.row_dimensions[3].height = 22
        # Filas 4+: encabezado de sección al cambiar de subcategoría + los conceptos.
        r = 4
        sub_actual = None
        for label, _order, sub, vals in filas(conn, company_id, category):
            if sub and sub != sub_actual:
                sub_actual = sub
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
                sc = ws.cell(row=r, column=1, value=sub)
                sc.font, sc.fill, sc.alignment = sec_font, tinta, left
                r += 1
            lc = ws.cell(row=r, column=1, value=label)
            lc.font, lc.alignment = label_font, left
            for c, (y, q) in enumerate(pers, start=2):
                v = vals.get((y, q))
                cell = ws.cell(row=r, column=c, value=v if v is not None else None)
                cell.font, cell.alignment, cell.number_format = val_font, right, '#,##0'
            r += 1
        ws.column_dimensions["A"].width = 80
        for c in range(2, ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 13

    # CIK sin ceros como "RUT" del pipeline (encaja en \d{4,8}).
    cik_token = str(int(cik))
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"estados_{cik_token}_{rango}_es.xlsx"
    wb.save(path)
    return path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", default="", help="IDs de empresa (coma). Vacío = todas las US.")
    ap.add_argument("--out-dir", default="cmf_extract/Products_US/Total")
    args = ap.parse_args()

    env = load_env(FDC / ".env")
    conn = psycopg2.connect(
        host=env["PGHOST"], port=env.get("PGPORT", 5432), dbname=env["PGDATABASE"],
        user=env["PGUSER"], password=env["PGPASSWORD"], sslmode="require")
    with conn.cursor() as cur:
        if args.only:
            ids = tuple(int(x) for x in args.only.split(",") if x.strip())
            cur.execute("SELECT id, cik FROM companies WHERE market='US' AND cik IS NOT NULL "
                        "AND id = ANY(%s) ORDER BY id", [list(ids)])
        else:
            cur.execute("SELECT id, cik FROM companies WHERE market='US' AND cik IS NOT NULL "
                        "ORDER BY id")
        targets = cur.fetchall()

    out_dir = Path(args.out_dir)
    hechos = 0
    for cid, cik in targets:
        path = construir(conn, int(cid), str(cik), out_dir)
        if path:
            hechos += 1
            print(f"  id={cid} → {path.name}")
    conn.close()
    print(f"\n{hechos} estados generados en {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
