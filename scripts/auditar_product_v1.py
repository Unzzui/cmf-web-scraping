#!/usr/bin/env python3
"""Audita los Excel de Product_v1 antes de publicarlos.

Es la compuerta entre "el Excel se generó" y "el Excel se vende". Existe porque las
dos veces que este producto salió mal, salió mal en silencio: el archivo se escribía,
pesaba lo normal, abría sin error — y adentro tenía un 5% de crecimiento inventado o
un título en letra blanca sobre fondo blanco. Nada de eso levanta una excepción.

Comprueba tres cosas, y las tres son cicatrices:

1. TEXTO ILEGIBLE. Contraste fuente/relleno < 1.6:1. El generador viejo pintaba
   "DEFINICIONES:" en blanco sin relleno (contraste 1,0) y el Enterprise Value en
   blanco sobre amarillo. Se mide el contraste real de cada celda con texto, no se
   confía en que la paleta esté bien usada.

2. CONSTANTES INVENTADAS. El DCF viejo escribía el literal 0.05 en "Crecimiento
   Ventas Y+1" y 0.27 en "Tasa efectiva de impuestos" para las 218 empresas por igual.
   Un número donde debería haber una fórmula es un número que alguien se inventó.

3. DEUDA NETA SIN ARRIENDOS. Bajo IFRS 16 un arriendo es deuda. Si la fórmula de deuda
   neta no referencia los pasivos por arrendamiento, el precio objetivo sale inflado
   (ESMAX, 1.236%).

Salida: 0 si todos los Excel pasan; 1 si alguno falla (y entonces NO se publica).
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
PRODUCT_V1 = RAIZ / "cmf_extract" / "Product_v1" / "Total"

CONTRASTE_MINIMO = 1.6


def _luminancia(argb: str | None) -> float | None:
    if not argb or not isinstance(argb, str) or len(argb) < 6:
        return None
    h = argb[-6:]
    try:
        r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None
    return (0.2126 * r + 0.7152 * g + 0.0722 * b) / 255


def celdas_ilegibles(wb) -> list[tuple[str, str, str]]:
    """Celdas con texto cuyo contraste contra su propio fondo es casi nulo."""
    malas = []
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for c in fila:
                if c.value is None or not str(c.value).strip():
                    continue
                fuente = (c.font.color.rgb
                          if c.font and c.font.color and c.font.color.type == "rgb" else None)
                relleno = (c.fill.fgColor.rgb
                           if c.fill and c.fill.fill_type == "solid"
                           and c.fill.fgColor and c.fill.fgColor.type == "rgb" else None)
                # Sin color de fuente => negro. Sin relleno => el papel, blanco.
                lf = _luminancia(fuente) if fuente else 0.0
                lb = _luminancia(relleno) if relleno else 1.0
                if lf is None or lb is None:
                    continue
                alto, bajo = max(lf, lb), min(lf, lb)
                if (alto + 0.05) / (bajo + 0.05) < CONTRASTE_MINIMO:
                    malas.append((ws.title, c.coordinate, str(c.value)[:40]))
    return malas


def problemas_del_dcf(wb) -> list[str]:
    """Constantes inventadas y deuda neta sin arriendos, en cada hoja DCF."""
    fallas = []
    for nombre in (h for h in wb.sheetnames if h.startswith("DCF")):
        ws = wb[nombre]
        for fila in ws.iter_rows(min_col=1, max_col=2):
            etiqueta, valor = fila[0].value, fila[1].value
            if not isinstance(etiqueta, str):
                continue
            es_formula = isinstance(valor, str) and valor.startswith("=")

            if "Crecimiento Ventas Y+1" in etiqueta and not es_formula:
                fallas.append(f"{nombre}: crecimiento Y+1 es la constante {valor!r}, no una fórmula")
            if "Tasa efectiva de impuestos" in etiqueta and not es_formula:
                fallas.append(f"{nombre}: tasa de impuestos es la constante {valor!r}, no una fórmula")
            if "Deuda neta" in etiqueta:
                if not es_formula:
                    fallas.append(f"{nombre}: deuda neta es la constante {valor!r}")
                elif "rrendamiento" not in valor:
                    fallas.append(f"{nombre}: la deuda neta NO incluye los arriendos (IFRS 16)")
    return fallas


def main() -> int:
    if not PRODUCT_V1.is_dir():
        print(f"No existe {PRODUCT_V1}", file=sys.stderr)
        return 1

    excels = sorted(p for p in PRODUCT_V1.glob("*.xlsx") if not p.name.startswith("~$"))
    if not excels:
        print(f"No hay ningún Excel en {PRODUCT_V1}", file=sys.stderr)
        return 1

    print(f"Auditando {len(excels)} Excel de {PRODUCT_V1}\n")

    reprobados: dict[str, list[str]] = {}
    motivos = Counter()

    for i, p in enumerate(excels, 1):
        try:
            wb = load_workbook(p)
        except Exception as exc:  # noqa: BLE001
            reprobados[p.name] = [f"no se pudo abrir: {exc}"]
            motivos["ilegible"] += 1
            continue

        fallas: list[str] = []

        ilegibles = celdas_ilegibles(wb)
        if ilegibles:
            motivos["texto ilegible"] += 1
            muestra = "; ".join(f"{h}!{c} {v!r}" for h, c, v in ilegibles[:3])
            fallas.append(f"{len(ilegibles)} celdas de texto ilegible → {muestra}")

        dcf = problemas_del_dcf(wb)
        if dcf:
            motivos["DCF"] += 1
            fallas.extend(dcf)

        wb.close()
        if fallas:
            reprobados[p.name] = fallas
        if i % 25 == 0:
            print(f"  … {i}/{len(excels)}", flush=True)

    print(f"\n{'=' * 72}")
    print(f"Auditados : {len(excels)}")
    print(f"Aprobados : {len(excels) - len(reprobados)}")
    print(f"Reprobados: {len(reprobados)}")
    if motivos:
        print(f"Motivos   : {dict(motivos)}")

    if reprobados:
        print(f"\n{'-' * 72}\nDETALLE (primeros 15):")
        for nombre, fallas in list(reprobados.items())[:15]:
            print(f"\n  {nombre}")
            for f in fallas[:4]:
                print(f"      · {f}")
        print(f"\n{'=' * 72}")
        print("REPROBADO — estos Excel NO se deben publicar.")
        return 1

    print(f"\n{'=' * 72}")
    print("APROBADO — los Excel se pueden publicar.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
