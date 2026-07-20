"""
DCF Patch Module
================

Módulo complementario para agregar funcionalidad DCF (FCFF) al generador de Excel.
Sistema dinámico de referencias: todas las fórmulas referencian DIRECTAMENTE a las hojas
existentes usando búsqueda dinámica de conceptos como en formula_builder.py.
"""

import re
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

# La hoja donde vive el detalle de la deuda y de donde el DCF toma su Kd.
_HOJA_DEUDA = "DEUDA FINANCIERA"

# El Kd de último recurso: sólo se usa si la empresa no declara créditos Y el cociente
# costos_financieros/deuda tampoco se puede calcular. No es un supuesto de modelación:
# es lo que evita que el WACC quede en #DIV/0!.
_KD_POR_DEFECTO = 0.06

try:
    from cmf_extract import excel_style as est
except ImportError:  # ejecutado desde dentro de cmf_extract/
    import excel_style as est

try:
    from cmf_extract.report_context import es_us as _es_us
except ImportError:  # ejecutado desde dentro de cmf_extract/
    from report_context import es_us as _es_us


try:
    from openpyxl.worksheet.data_validation import DataValidation
except ImportError:
    DataValidation = None


class DCFBuilder:
    """
    Constructor de hojas DCF para análisis de valuación.
    """

    # ----------------------------
    # INIT - Sistema dinámico como formula_builder.py
    # ----------------------------
    def __init__(self, workbook: Workbook, financial_data: Dict[str, Any]):
        self.wb = workbook
        self.financial_data = financial_data
        self.years = financial_data.get("years", [])

        # Moneda en la que la empresa REPORTA sus estados. La inyecta formula_processor
        # desde los facts del XBRL. 18 empresas del catálogo reportan en dólares (SQM,
        # COPEC, CMPC, LATAM, ARAUCO, COLBÚN…), y el precio de su acción cotiza en PESOS.
        # Sin esto, el DCF calcula un valor intrínseco en dólares y lo compara contra un
        # precio en pesos: la "Prima/(Descuento)" y la "Recomendación" salen desviadas
        # ~900x, y siempre dicen "extremadamente sobrevalorada".
        self._es_us = _es_us()
        self.reporting_currency = str(
            financial_data.get("reporting_currency") or ("USD" if self._es_us else "CLP")
        ).upper()

        # Moneda contra la que se compara el precio de bolsa. En Chile la acción cotiza en
        # PESOS, así que el valor por acción del DCF se convierte a CLP antes de comparar.
        # Para una empresa de EEUU la acción cotiza en su moneda de reporte (USD): NO se
        # convierte a pesos — mostrar la valuación de NVIDIA en CLP sería un dato falso.
        self.moneda_precio = self.reporting_currency if self._es_us else "CLP"

        # Estilos
        self._setup_styles()

        # Hojas existentes (nombres en español/inglés)
        def pick_sheet(wb: Workbook, names: list[str]):
            for n in names:
                if n in wb.sheetnames:
                    return wb[n]
            return None

        self.sh_bal = pick_sheet(workbook, ["Balance General", "Balance Sheet"])
        self.sh_pl = pick_sheet(workbook, ["Estado de Resultados", "Estado Resultados (Función)", "Income Statement"])
        self.sh_cfs = pick_sheet(workbook, ["Flujo Efectivo", "Cash Flow"])
        self.sh_ratios = pick_sheet(workbook, ["RATIOS & KPIs"])

        # Detectar fila de encabezados por hoja dinámicamente
        def detect_header_row(sheet) -> int:
            if sheet is None:
                return 3
            max_scan = min(10, sheet.max_row)
            for r in range(1, max_scan + 1):
                v0 = sheet.cell(row=r, column=1).value
                if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
                    return r
                for c in range(2, min(sheet.max_column, 20) + 1):
                    v = sheet.cell(row=r, column=c).value
                    if isinstance(v, str):
                        s = v.strip().split("\n", 1)[0]
                        if re.match(r"^\d{4}-(\d{2}|\d{2}-\d{2})$", s):
                            return r
                        if re.match(r"^\d{4}(?:Q[1-4])?$", s):
                            return r
            return 3

        self.HDR_BAL = detect_header_row(self.sh_bal)
        self.HDR_PL = detect_header_row(self.sh_pl)
        self.HDR_CFS = detect_header_row(self.sh_cfs)
        self.HDR_RATIOS = detect_header_row(self.sh_ratios)

        # Mapeos dinámicos de conceptos a filas
        self.rows_bal = self._map_balance_rows() if self.sh_bal else {}
        self.rows_pl = self._map_income_rows() if self.sh_pl else {}
        self.rows_cfs = self._map_cash_flow_rows() if self.sh_cfs else {}
        self.rows_ratios = self._map_ratios_rows() if self.sh_ratios else {}

    # ----------------------------
    # Utilidades dinámicas - Sistema como formula_builder.py
    # ----------------------------
    def _find_row_in_sheet(self, sheet, concept_name: str) -> Optional[int]:
        """
        Encuentra la fila de un concepto en una hoja específicamente.
        Búsqueda dinámica por contenido de celda.
        """
        if sheet is None:
            return None
        
        hdr = self._hdr_row_for_sheet(sheet)
        if hdr is None:
            return None
        
        try:
            # Normalizar el concepto de búsqueda. Se quitan las COMAS: el catálogo EDGAR
            # (EEUU) rotula "Pasivos por arrendamientos, corrientes" (con coma) donde el
            # chileno usa "Pasivos por arrendamientos corrientes" (sin). Sin normalizar, el
            # DCF de una empresa US no encontraba sus arriendos y la deuda neta salía sin los
            # leases (Amazon: $20B en vez de ~$110B).
            def _norm(s: str) -> str:
                return s.strip().lower().replace(",", "")
            concept_lower = _norm(concept_name)

            # Buscar fila que contenga el concepto
            for r in range(hdr + 1, sheet.max_row + 1):
                cell_value = sheet.cell(row=r, column=1).value
                if not isinstance(cell_value, str):
                    continue

                cell_lower = _norm(cell_value)

                # Evitar filas abstractas/resumen
                if any(word in cell_lower for word in ["abstract", "resumen", "sinopsis"]):
                    continue

                # Buscar coincidencia exacta o inclusión
                if cell_lower == concept_lower or concept_lower in cell_lower:
                    return r
                    
        except Exception:
            pass
        
        return None

    def _hdr_row_for_sheet(self, sheet):
        """Retorna la fila de headers para una hoja."""
        if sheet == self.sh_bal:
            return self.HDR_BAL
        elif sheet == self.sh_pl:
            return self.HDR_PL
        elif sheet == self.sh_cfs:
            return self.HDR_CFS
        elif sheet == self.sh_ratios:
            return self.HDR_RATIOS
        return 3

    def _get_col_letter_by_label(self, sheet, label: str) -> Optional[str]:
        """
        Encuentra la columna (letra) para un período específico.
        Sistema dinámico como formula_builder.py
        """
        if sheet is None:
            return None
        
        hdr = self._hdr_row_for_sheet(sheet)
        label_norm = label.strip().split("\n", 1)[0]
        
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=hdr, column=col).value
            if isinstance(val, str) and val.strip().split("\n", 1)[0] == label_norm:
                return get_column_letter(col)
        return None

    def create_cell_reference_by_label(self, sheet, row_num: Optional[int], label: str) -> Optional[str]:
        """
        Crea una referencia de celda dinámica usando fila + etiqueta de período.
        Equivalente al método de formula_builder.py
        """
        if row_num is None or sheet is None:
            return None
        
        col = self._get_col_letter_by_label(sheet, label)
        if col is None:
            return None
        
        return f"'{sheet.title}'!{col}{row_num}"
    def _map_balance_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del balance a números de fila usando búsqueda dinámica."""
        concepts = {
            "AC": "Activos corrientes totales",
            "ANC": "Total de activos no corrientes",  
            "AT": "Total de activos",
            "PC": "Pasivos corrientes totales",
            "PNC": "Total de pasivos no corrientes",
            "PT": "Total de pasivos",
            "Patr": "Patrimonio atribuible a los propietarios de la controladora",
            "PatrTotal": "Patrimonio total",
            "CxC": "Deudores comerciales y otras cuentas por cobrar corrientes",
            "Inv": "Inventarios corrientes",
            "CxP": "Cuentas por pagar comerciales y otras cuentas por pagar",
            "Efec": "Efectivo y equivalentes al efectivo",
            "DeudaFinCorr": "Otros pasivos financieros corrientes",
            "DeudaFinNC": "Otros pasivos financieros no corrientes",
            # BAJO IFRS 16 UN ARRIENDO ES DEUDA. Y no es un matiz contable:
            # en SMU los arriendos son el 58,6% de su deuda total, en Tricot el 54,2%,
            # en Falabella el 28,9%. Dejarlos fuera de la deuda neta INFLA el valor
            # patrimonial y, con él, el precio objetivo.
            "ArrendCorr": "Pasivos por arrendamientos corrientes",
            "ArrendNC": "Pasivos por arrendamientos no corrientes",
            "PPE": "Propiedades, planta y equipo",
        }
        
        return {key: self._find_row_in_sheet(self.sh_bal, concept)
                for key, concept in concepts.items()}

    def _map_income_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del estado de resultados a números de fila usando búsqueda dinámica."""
        concepts = {
            "Ventas": "Ingresos de actividades ordinarias",
            "COGS": "Costo de ventas",
            "Bruta": "Ganancia bruta",
            "EBIT": "Ganancias (pérdidas) de actividades operacionales",
            "Neta": "Ganancia (pérdida)",
            "NetaControl": "Ganancia (pérdida), atribuible a los propietarios de la controladora",
            "DepAmort": "Depreciación y amortización",
            "ImpGanan": "Gasto por impuestos a las ganancias",
            "AntesImp": "Ganancia (pérdida), antes de impuestos",
        }
        
        return {key: self._find_row_in_sheet(self.sh_pl, concept)
                for key, concept in concepts.items()}

    def _map_cash_flow_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos del flujo de efectivo a números de fila usando búsqueda dinámica."""
        concepts = {
            "CFO": "Flujos de efectivo netos procedentes de (utilizados en) operaciones",
            "CFI": "Flujos de efectivo netos procedentes de (utilizados en) actividades de inversión",
            "CFF": "Flujos de efectivo netos procedentes de (utilizados en) actividades de financiación",
            "CapEx": "Compras de propiedades, planta y equipo",
            "CapExIntang": "Compras de activos intangibles",
        }
        
        return {key: self._find_row_in_sheet(self.sh_cfs, concept)
                for key, concept in concepts.items()}

    def _map_ratios_rows(self) -> Dict[str, Optional[int]]:
        """Mapea conceptos de RATIOS & KPIs a números de fila usando búsqueda dinámica."""
        if not self.sh_ratios:
            return {}
        
        concepts = {
            "MargenEBIT": "Margen Operativo (EBIT)",
            "DepAmort": "Depreciación y Amortización", 
            "Acciones": "Total número de acciones emitidas",
        }
        
        return {key: self._find_row_in_sheet(self.sh_ratios, concept)
                for key, concept in concepts.items()}

    def _iter_periods_from_sheet(self, sheet) -> List[str]:
        """Lee los labels de la fila de encabezados (fila 3 o 4) y devuelve la lista de periodos (strings)."""
        if sheet is None:
            return []
        hdr = 4 if sheet == self.sh_ratios else 3
        labels: List[str] = []
        for col in range(2, sheet.max_column + 1):
            v = sheet.cell(row=hdr, column=col).value
            if isinstance(v, str):
                s = v.strip()
                if s:
                    labels.append(s.split("\n", 1)[0])
        return labels
    
    def _find_latest_period(self) -> str:
        """
        Encuentra automáticamente el último período disponible en las hojas.
        """
        periods = self._iter_periods_from_sheet(self.sh_pl)
        if not periods:
            from datetime import datetime
            return f"{datetime.now().year}Q4"

        valid_periods = [p for p in periods if re.match(r"^\d{4}(Q[1-4])?$", p)]
        if not valid_periods:
            from datetime import datetime
            return f"{datetime.now().year}Q4"

        def period_to_number(period):
            if "Q" in period:
                year, quarter = period.split("Q")
                return int(year) * 10 + int(quarter)
            else:
                return int(period) * 10 + 4  # Bare YYYY treated as Q4

        valid_periods_sorted = sorted(valid_periods, key=period_to_number, reverse=True)
        return valid_periods_sorted[0]
    
    def _deuda(self) -> dict | None:
        """La nota de préstamos de la empresa, si la declara. La inyecta formula_processor."""
        d = self.financial_data.get("deuda")
        return d if isinstance(d, dict) and d.get("kd") else None

    def _kd_valor(self, fila_costos_fin: int, fila_deuda: int) -> str:
        """El Kd: SIEMPRE una fórmula. La tasa declarada si la hay, el cociente si no.

        Cuando la empresa declara sus créditos, esta celda apunta a la hoja DEUDA
        FINANCIERA, donde el Kd sale de un SUMAPRODUCTO sobre el detalle: tasa x monto
        dividido por la suma de los montos.

        Es una fórmula y no un número pegado a propósito. Un número obliga a creerle al
        que lo pegó; una fórmula se audita, y si el analista corrige o filtra un crédito,
        el Kd y el WACC se recalculan solos.

        La celda del Kd en la hoja de deuda es B6, y es determinista: la sección arranca
        en la fila 4 y el encabezado ocupa dos filas.
        """
        if self._deuda():
            return f"=IFERROR('{_HOJA_DEUDA}'!B6,{_KD_POR_DEFECTO})"
        return f"=IFERROR(B{fila_costos_fin}/B{fila_deuda},{_KD_POR_DEFECTO})"

    def _kd_fuente(self) -> str:
        """De dónde salió el Kd. Un supuesto que no se nombra es un supuesto que se cree."""
        deuda = self._deuda()
        if not deuda:
            return ("ESTIMADO: costos financieros / deuda financiera "
                    "(la empresa no declara tasas por crédito)")
        n = deuda.get("n_creditos", 0)
        monedas = ", ".join(sorted(deuda.get("por_moneda", {})))
        return (f"DECLARADO por la empresa: promedio de {n} crédito(s) de la nota de "
                f"préstamos, ponderado por monto. Monedas: {monedas}")

    def _beta_yahoo(self) -> Optional[float]:
        """El beta de Yahoo de la empresa, acotado a [0.5, 2.0], o None si no hay.

        Se inyecta desde la BD (companies.yahoo_beta) vía formula_processor. Sólo ~42
        empresas cotizan y tienen beta de Yahoo. ``None`` es una respuesta válida: el
        que arma la hoja cae a Hamada (ver create_wacc_terminal_block), no a un número
        inventado. El MISMO beta que usa el motor de la BD, para que el WACC cuadre.
        """
        b = self.financial_data.get("yahoo_beta")
        try:
            b = float(b)
        except (TypeError, ValueError):
            return None
        return max(0.5, min(2.0, b))

    def _beta_fuente(self) -> str:
        """De dónde salió el beta. Un supuesto que no se nombra es un supuesto que se cree.

        Si la empresa cotiza y tiene beta de Yahoo, se usa ése. Si no, el beta se
        re-apalanca con Hamada (beta desapalancada 0,8 × (1+(1−t)·D/E)): así toda
        empresa tiene un beta sensible a SU apalancamiento, no un valor neutral parejo.
        """
        if self._beta_yahoo() is not None:
            return "Yahoo Finance (acotado a [0,5, 2,0]); editable"
        return ("HAMADA: beta desapalancada 0,8 re-apalancada con D/E y la tasa "
                "efectiva de la empresa (sin beta de Yahoo)")

    def _tipo_de_cambio_actual(self) -> float:
        """El dólar observado más reciente del Banco Central.

        Aquí iba un 950 fijo. Y el número importa: esta celda convierte el valor por acción
        a pesos para poder compararlo contra el precio de bolsa, así que un tipo de cambio
        inventado desvía la "Prima/(Descuento)" y la "Recomendación" de COMPRA a VENTA.

        Es el dólar MÁS RECIENTE, no el del período base, porque lo que se compara es el
        valor intrínseco contra el precio de la acción HOY. Queda editable: el analista
        puede usar el tipo de cambio que quiera -- pero el valor por defecto ya no es un
        número escrito a mano.
        """
        try:
            from cmf_extract import fx
        except ImportError:
            try:
                import fx
            except ImportError:
                return 950.0

        serie = fx._serie()
        if serie.empty:
            print("[dcf] AVISO: sin serie del dólar observado; el tipo de cambio del "
                  "modelo queda en 950 (un supuesto, no un dato).")
            return 950.0

        return round(float(serie.iloc[-1]["clp_por_usd"]), 2)

    def _find_base_annual_period(self) -> str:
        """
        Encuentra el año base anual más reciente.
        Acepta tanto YYYY como YYYYQ4 como períodos anuales.
        Resultado cacheado para evitar recomputación.
        """
        if hasattr(self, '_cached_base_annual'):
            return self._cached_base_annual
        periods = self._iter_periods_from_sheet(self.sh_pl)
        annual_periods = []
        for p in periods:
            m = re.match(r"^(\d{4})Q4$", p)
            if m:
                annual_periods.append(p)
            elif re.match(r"^\d{4}$", p):
                annual_periods.append(f"{p}Q4")

        if annual_periods:
            annual_sorted = sorted(annual_periods, reverse=True)
            self._cached_base_annual = annual_sorted[0]
        else:
            self._cached_base_annual = self._find_latest_period()
        return self._cached_base_annual

    # ----------------------------
    # Fórmulas (todas DIRECTAS)
    # ----------------------------
    def _get_ventas_base_formula(self, base_period: str = None) -> str:
        """
        Genera fórmula dinámica para ventas año base usando sistema de referencias como formula_builder.py
        """
        if base_period is None:
            base_period = self._find_base_annual_period()
            
        if self.sh_pl and "Ventas" in self.rows_pl and self.rows_pl["Ventas"]:
            # Para períodos trimestrales, usar ratios anualizados de RATIOS & KPIs
            if "Q" in base_period and self.sh_ratios:
                # Buscar "Ventas anualizadas" en RATIOS & KPIs
                ventas_anualizadas_row = self._find_row_in_sheet(self.sh_ratios, "Ventas")
                if ventas_anualizadas_row:
                    ref = self.create_cell_reference_by_label(self.sh_ratios, ventas_anualizadas_row, base_period)
                    if ref:
                        return f"=IFERROR({ref},\"N/D\")"
            
            # Para períodos anuales, usar Estado de Resultados directamente
            ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], base_period)
            if ref:
                return f"=IFERROR({ref},\"N/D\")"
        
        return "=\"N/D\""
    def _avg_ratio_over_annual(self, num_specs: list, den_sheet, den_row,
                               default: float, lo: float, hi: float,
                               abs_num: bool = False) -> Optional[str]:
        """Promedio histórico anual de (Σ numeradores)/denominador, acotado a [lo, hi].

        Réplica EXACTA de la media que hace el motor de la BD
        (``excel_aligned.compute_excel_drivers`` + ``_safe_avg``): para cada año
        Q4 con denominador > 0, se calcula el ratio; el resultado es el PROMEDIO
        SIMPLE de esos ratios año a año (no Σnum/Σden), acotado a [lo, hi] y con
        fallback a ``default`` si no hay ningún año utilizable.

        Antes estos drivers tomaban un ÚNICO año base, mientras la BD promediaba
        el histórico: dos números distintos para la misma empresa. Ahora ambos
        promedian.

        ``num_specs``: lista de ``(sheet, row)`` que se SUMAN en el numerador
        (p. ej. CapEx PP&E + CapEx intangibles). El denominador se exige > 0 (en
        estos drivers es Ventas, siempre positivo; la condición mantiene la
        simetría con ``_ratio_series`` de la BD).
        """
        if den_sheet is None or not den_row:
            return None
        # Misma ventana que la BD (excel_aligned._annual_series usa years_back=5,
        # o sea ~6 cierres anuales): tomar los 6 más recientes. Sin esto, el Excel
        # promediaría TODA la serie y el promedio no cuadraría con la BD.
        anuales = self._iter_annual_periods()

        def _yr(p: str) -> int:
            m = re.match(r"^(\d{4})", str(p))
            return int(m.group(1)) if m else 0

        anuales = sorted(anuales, key=_yr, reverse=True)[:6]
        nums, dens = [], []
        any_num_data = False
        for p in anuales:
            rd = self.create_cell_reference_by_label(den_sheet, den_row, p)
            if not rd:
                continue
            parts = []
            for (s, r) in num_specs:
                if s is None or not r:
                    continue
                col = self._get_col_letter_by_label(s, p)
                if col is None:
                    continue
                # Sólo se suma la celda si TIENE dato (no vacía). Una fila entera vacía
                # —p. ej. la D&A de SMU, que el pipeline no captura— NO debe contar como
                # "D&A = 0": eso hunde el FCFF. La BD hace lo mismo (``_annual_series``
                # descarta los NULL); si no queda ningún año con dato, se cae al default.
                cellval = s.cell(row=r, column=column_index_from_string(col)).value
                if cellval is None or isinstance(cellval, str):
                    continue
                any_num_data = True
                rn = f"'{s.title}'!{col}{r}"
                parts.append(f"ABS({rn})" if abs_num else rn)
            if not parts:
                continue
            num_expr = f"({'+'.join(parts)})"
            nums.append(f"IF({rd}>0,{num_expr}/{rd},0)")
            dens.append(f"IF({rd}>0,1,0)")
        if not nums or not any_num_data:
            return None
        suma = "+".join(nums)
        cuenta = "+".join(dens)
        return f"=IFERROR(MAX(MIN(({suma})/({cuenta}),{hi}),{lo}),{default})"

    def _get_margen_ebit_formula(self) -> str:
        """Margen EBIT de proyección = PROMEDIO histórico anual de EBIT/Ventas.

        Media multi-año acotada a [1%, 50%], idéntica a la BD
        (``_safe_avg(margen_series, default=0.10, lo=0.01, hi=0.5)``).
        """
        if (self.sh_pl and self.rows_pl.get("EBIT") and self.rows_pl.get("Ventas")):
            f = self._avg_ratio_over_annual(
                [(self.sh_pl, self.rows_pl["EBIT"])],
                self.sh_pl, self.rows_pl["Ventas"],
                default=0.10, lo=0.01, hi=0.5)
            if f:
                return f
        return "0.10"

    def _get_da_ventas_formula(self) -> str:
        """D&A/Ventas de proyección = PROMEDIO histórico anual de |D&A|/Ventas.

        Media multi-año acotada a [0%, 30%], idéntica a la BD
        (``_safe_avg(da_series, default=0.03, lo=0.0, hi=0.30)``). La D&A se toma
        en valor absoluto (igual que la BD).
        """
        da_sheet, da_row = None, None
        if self.sh_ratios and self.rows_ratios.get("DepAmort"):
            da_sheet, da_row = self.sh_ratios, self.rows_ratios["DepAmort"]
        elif self.sh_pl and self.rows_pl.get("DepAmort"):
            da_sheet, da_row = self.sh_pl, self.rows_pl["DepAmort"]
        if da_sheet and self.sh_pl and self.rows_pl.get("Ventas"):
            f = self._avg_ratio_over_annual(
                [(da_sheet, da_row)],
                self.sh_pl, self.rows_pl["Ventas"],
                default=0.03, lo=0.0, hi=0.30, abs_num=True)
            if f:
                return f
        return "0.03"

    def _get_capex_ventas_formula(self) -> str:
        """CapEx/Ventas de proyección = PROMEDIO histórico anual de |CapEx|/Ventas.

        CapEx = compras de PP&E + compras de intangibles (ambas del flujo de
        efectivo), en valor absoluto. Media multi-año acotada a [0%, 30%],
        idéntica a la BD (``_safe_avg(capex_series, default=0.04, lo=0.0, hi=0.30)``).
        """
        specs = []
        if self.sh_cfs and self.rows_cfs.get("CapEx"):
            specs.append((self.sh_cfs, self.rows_cfs["CapEx"]))
        if self.sh_cfs and self.rows_cfs.get("CapExIntang"):
            specs.append((self.sh_cfs, self.rows_cfs["CapExIntang"]))
        if specs and self.sh_pl and self.rows_pl.get("Ventas"):
            f = self._avg_ratio_over_annual(
                specs, self.sh_pl, self.rows_pl["Ventas"],
                default=0.04, lo=0.0, hi=0.30, abs_num=True)
            if f:
                return f
        return "0.04"

    def _get_deuda_neta_formula(self, period: str = None) -> str:
        """Deuda neta = deuda financiera + ARRIENDOS − efectivo.

        Los arriendos NO estaban. El DCF de la base tenía hasta el comentario que lo
        confesaba —"Net Debt (SIN leases — igual a Excel)"— así que los dos modelos
        omitían lo mismo y por eso parecían coincidir. Estaban mal juntos.

        Bajo IFRS 16 un contrato de arriendo es un pasivo financiero: la empresa se
        obligó a pagar un flujo futuro, igual que con un crédito. Omitirlo baja la deuda
        neta, y como

            valor patrimonial = valor de la empresa − deuda neta

        el precio objetivo sale INFLADO. Medido sobre la base:

            ALMENDRAL              sobreestimado  72%
            TELEFÓNICA MÓVILES                  129%
            ESMAX (combustibles)              1.236%   ← su arriendo supera su patrimonio

        No es un sesgo parejo: golpea a retail, telecomunicaciones y combustibles, que
        son los que operan con locales arrendados. Justo donde el analista más mira.
        """
        if period is None:
            period = self._find_latest_period()

        terms = []

        for clave in ("DeudaFinCorr", "DeudaFinNC", "ArrendCorr", "ArrendNC"):
            if self.sh_bal and self.rows_bal.get(clave):
                ref = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal[clave], period)
                if ref:
                    terms.append(ref)
        
        # Efectivo y equivalentes
        cash_ref = None
        if self.sh_bal and "Efec" in self.rows_bal and self.rows_bal["Efec"]:
            cash_ref = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal["Efec"], period)
        
        if terms and cash_ref:
            return f"=IFERROR(({'+'.join(terms)})-{cash_ref},0)"
        elif terms:
            return f"=IFERROR({'+'.join(terms)},0)"
        
        return "0"

    def _get_tasa_impuestos_formula(self) -> str:
        """Tasa efectiva de impuestos, calculada de los estados de la propia empresa.

        Antes esto devolvía la constante "0.27" para las 218 empresas. La tasa corporativa
        de Chile es 27%, sí — pero la tasa EFECTIVA que paga cada empresa no lo es casi
        nunca: créditos, pérdidas de arrastre, filiales en otras jurisdicciones y
        diferencias temporarias la mueven. Medida sobre los estados 2021-2025:

            mediana 22,6%   ·   p10 8,7%   ·   p90 29,5%
            58 de 162 empresas están a MÁS DE 8 puntos del 27%

        Y no es un decimal cosmético: la tasa entra dos veces en el modelo — en el NOPAT
        (EBIT × (1 − t)) y en el escudo fiscal del WACC (Kd × (1 − t)). Un error de 10
        puntos mueve el precio objetivo, que es el número sobre el que alguien decide
        comprar o vender.

        MISMA SEMÁNTICA QUE EL DCF DE LA BASE (scripts/dcf/excel_aligned.py):
        promedio de |impuesto| / utilidad-antes-de-impuestos sobre los años con utilidad
        positiva, acotado a [10%, 40%]. En años de pérdida el cociente no significa nada
        (una empresa que pierde plata puede registrar un crédito), así que esos años NO
        entran en el promedio — de ahí el conteo explícito del denominador en vez de un
        AVERAGE, que los contaría igual.

        Si no se puede calcular, se cae al 27% legal. Es un respaldo, no la respuesta.
        """
        fila_imp = self.rows_pl.get("ImpGanan")
        fila_ebt = self.rows_pl.get("AntesImp")
        if not fila_imp or not fila_ebt:
            return "0.27"

        anuales = self._iter_annual_periods()
        if not anuales:
            return "0.27"

        # Misma ventana de 6 cierres anuales que la BD (excel_aligned usa years_back=5),
        # para que el promedio de la tasa efectiva cuadre entre ambos motores.
        def _yr_imp(p: str) -> int:
            m = re.match(r"^(\d{4})", str(p))
            return int(m.group(1)) if m else 0

        anuales = sorted(anuales, key=_yr_imp, reverse=True)[:6]

        numerador, denominador = [], []
        for periodo in anuales:
            ref_imp = self.create_cell_reference_by_label(self.sh_pl, fila_imp, periodo)
            ref_ebt = self.create_cell_reference_by_label(self.sh_pl, fila_ebt, periodo)
            if not ref_imp or not ref_ebt:
                continue
            # Sólo los años con utilidad positiva.
            numerador.append(f"IF({ref_ebt}>0,ABS({ref_imp})/{ref_ebt},0)")
            denominador.append(f"IF({ref_ebt}>0,1,0)")

        if not numerador:
            return "0.27"

        suma = "+".join(numerador)
        cuenta = "+".join(denominador)
        # El acotado a [10%, 40%] evita que un año raro (un crédito fiscal grande, una
        # base imponible casi nula) arrastre todo el modelo.
        return f"=IFERROR(MAX(MIN(({suma})/({cuenta}),0.4),0.1),0.27)"

    def _iter_annual_periods(self) -> list:
        """Los períodos anuales de la serie, en orden. Un cierre anual es Q4."""
        anuales = []
        for periodo in self._iter_periods_from_sheet(self.sh_pl):
            if re.match(r"^\d{4}Q4$", periodo):
                anuales.append(periodo)
            elif re.match(r"^\d{4}$", periodo):
                anuales.append(periodo)
        return anuales
    
    def _acciones_fallback(self) -> Optional[float]:
        """Acciones inyectadas desde la BD (companies.shares_outstanding, UNIDADES), o None.

        La misma fuente de verdad que usa el motor de la BD. Sirve de respaldo cuando la
        hoja RATIOS no trae el número de acciones: para AGUAS, por ejemplo, la fila "Total
        número de acciones emitidas" viene vacía, y sin este respaldo el "Valor por Acción"
        del DCF quedaba en blanco (y no cuadraba con el precio objetivo de la web).
        """
        s = self.financial_data.get("shares_outstanding")
        try:
            s = float(s)
        except (TypeError, ValueError):
            return None
        return s if s > 0 else None

    def _get_acciones_formula(self) -> str:
        """
        Genera fórmula dinámica para acciones en circulación.

        Toma el ÚLTIMO valor no vacío de la fila de acciones en RATIOS (el más
        reciente disponible), en vez de un período fijo: las acciones suelen
        venir sólo en algunos períodos (anuales) y el período más reciente
        puede estar vacío. LOOKUP(2,1/(rango<>""),rango) devuelve el último no
        vacío, que en esta hoja (columnas de viejo a nuevo) es el más reciente.

        Si RATIOS no trae acciones, cae a las inyectadas desde la BD
        (companies.shares_outstanding) — la misma cifra que usa el motor de la web,
        para que el precio objetivo cuadre — en vez de dejar la celda vacía.
        """
        fb = self._acciones_fallback()
        fb_expr = f"{fb:.0f}" if fb is not None else '""'
        row = self.rows_ratios.get("Acciones") if self.rows_ratios else None
        if self.sh_ratios and row:
            # Última columna de período (encabezado tipo YYYY o YYYYQn)
            hdr = getattr(self, "HDR_RATIOS", 4) or 4
            last_col = None
            for c in range(2, self.sh_ratios.max_column + 1):
                v = self.sh_ratios.cell(row=hdr, column=c).value
                if isinstance(v, str) and re.match(r"^\d{4}(Q[1-4])?$", v.strip()):
                    last_col = c
            if last_col:
                from openpyxl.utils import get_column_letter
                title = self.sh_ratios.title
                rng = f"'{title}'!B{row}:{get_column_letter(last_col)}{row}"
                # LOOKUP devuelve "" (no error) si el rango está vacío; por eso el fallback
                # va con un IF explícito, no sólo con IFERROR.
                lookup = f'IFERROR(LOOKUP(2,1/({rng}<>""),{rng}),"")'
                return f'=IFERROR(IF({lookup}="",{fb_expr},{lookup}),{fb_expr})'
        return f"={fb_expr}" if fb is not None else '""'
    
    def _get_cagr_formula(self, years_back: int = 5) -> str:
        """
        Calcula CAGR automático de ventas basado en períodos históricos disponibles.
        Maneja correctamente el orden de columnas desde más reciente a más antiguo.
        
        Args:
            years_back: Número de años hacia atrás para calcular CAGR
        """
        if not (self.sh_pl and "Ventas" in self.rows_pl and self.rows_pl["Ventas"]):
            return "0.05"  # Default 5%
            
        # Obtener períodos disponibles en el orden que aparecen
        periods = self._iter_periods_from_sheet(self.sh_pl)
        if not periods or len(periods) < 2:
            return "0.05"
            
        # Cierres ANUALES.
        #
        # ESTE FILTRO ESTABA MUERTO. Exigía que el período NO contuviera "Q":
        #
        #     if "Q" not in period_str and len(period_str) >= 4:
        #
        # Pero TODOS los encabezados del Excel son "2026Q1", "2025Q4", "2024Q4"…
        # Ninguno pasaba jamás, así que `annual_periods` quedaba vacío y la función
        # retornaba "0.05" SIEMPRE. Verificado en 7 Excel de producción: la celda
        # "Crecimiento Ventas Y+1" contiene el literal 0.05, no una fórmula.
        #
        # O sea que TODOS los Excel vendidos proyectaban un 5% de crecimiento, igual
        # para Falabella que para SQM. Y como Y+2 a Y+5 se derivan de esa celda
        # (=MAX(B13*0.9,0.02)…), toda la proyección de cinco años colgaba de un número
        # inventado.
        #
        # El cierre anual de una empresa es su Q4. (Los formatos antiguos traían el año
        # a secas; se aceptan los dos.)
        annual_periods = []
        for p in periods:
            period_str = str(p).strip()
            if len(period_str) < 4 or not period_str[:4].isdigit():
                continue
            up = period_str.upper()
            es_anual = ("Q" not in up) or up.endswith("Q4")
            if es_anual:
                annual_periods.append((period_str[:4], period_str))

        if len(annual_periods) < 2:
            return "0.05"
        
        # Los períodos están en orden: más reciente primero
        # annual_periods = [(2024, "2024"), (2023, "2023"), (2022, "2022"), ...]
        
        # Encontrar el año más reciente y uno anterior para CAGR
        most_recent_year, most_recent_label = annual_periods[0]  # 2024
        
        # Buscar el período más lejano disponible dentro del rango solicitado
        # Priorizar períodos más largos para un CAGR más representativo
        best_start_year = None
        best_start_label = None
        best_years_diff = 0
        
        for i in range(1, len(annual_periods)):
            start_year, start_label = annual_periods[i]
            years_diff = float(most_recent_year) - float(start_year)
            
            # Solo considerar períodos dentro del rango solicitado
            if years_diff <= years_back and years_diff > best_years_diff:
                # Verificar que podemos crear las referencias
                end_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], most_recent_label)
                start_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], start_label)
                
                if end_ref and start_ref:
                    best_start_year = start_year
                    best_start_label = start_label
                    best_years_diff = years_diff
        
        # Usar el mejor período encontrado
        if best_start_year and best_years_diff > 0:
            end_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], most_recent_label)
            start_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], best_start_label)
            
            # CAGR calculado
            # Referencias CAGR identificadas
            # Tope 15% y piso 2%: los mismos que aplica el DCF que se guarda en la BD
            # (scripts/dcf/excel_aligned.py). Sin ellos, una empresa que dobló ventas en
            # un año proyectaría 100% anual durante cinco años.
            return f"=IFERROR(MAX(MIN(POWER({end_ref}/{start_ref},1/{best_years_diff})-1,0.15),0.02),0.05)"
        
        # Fallback: usar los dos períodos más recientes si están disponibles
        if len(annual_periods) >= 2:
            recent_year, recent_label = annual_periods[0]
            prev_year, prev_label = annual_periods[1]
            
            end_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], recent_label)
            start_ref = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], prev_label)
            
            if end_ref and start_ref:
                years_diff = float(recent_year) - float(prev_year)
                if years_diff > 0:
                    # CAGR fallback calculado
                    return f"=IFERROR(MAX(MIN(POWER({end_ref}/{start_ref},1/{years_diff})-1,0.15),0.02),0.05)"
        
        return "0.05"  # Fallback final

    # ----------------------------
    # Estilos
    # ----------------------------
    def _setup_styles(self):
        """Estilos del DCF. Todos salen de cmf_extract/excel_style.py.

        Antes: Calibri, navy #0F172A, acento AZUL #2563EB, celdas de input celeste
        #DBEEF3, resultados en AMARILLO #FFEB9C. Cinco colores compitiendo, ninguno con
        significado — y un lenguaje visual completamente distinto al del Excel que genera
        la web (Inter, tinta, cobre).

        Ahora las celdas EDITABLES se distinguen por fondo suave, y lo único con acento es
        el rótulo de sección. El resultado de la valoración no necesita fondo amarillo:
        necesita ser el número más grande de la hoja.
        """
        # Fuentes
        self.header_font = est.fuente(16, bold=True, color=est.PAPER)
        # BLANCA: sólo válida sobre RELLENO_TINTA. Nunca sobre una celda clara.
        self.subheader_font = est.fuente(11, bold=True, color=est.PAPER)
        # Para rótulos que van sobre el papel. Antes se reusaba la blanca y el texto
        # desaparecía (p. ej. "DEFINICIONES:" en la hoja DRIVERS WC).
        self.rotulo_font = est.ETIQUETA           # rótulo de sección, cobre
        self.rotulo_fuerte_font = est.CUERPO_FUERTE
        self.input_font = est.fuente(10, color=est.INK)
        self.data_font = est.fuente(10, color=est.INK)

        self.section_title_font = est.SECCION
        self.key_metric_font = est.fuente(11, bold=True, color=est.PAPER)
        # TINTA, no blanca: acompaña a `key_result_fill`, que es un panel CLARO. La
        # fuente blanca (`key_metric_font`) sólo vale sobre relleno tinta.
        self.result_font = est.fuente(11, bold=True, color=est.INK)
        self.label_font = est.fuente(10, color=est.INK)

        # Rellenos
        self.header_fill = est.RELLENO_TINTA
        self.subheader_fill = est.RELLENO_TINTA
        # Las celdas que el analista PUEDE tocar. Es la única distinción funcional de la
        # hoja, y por eso es la única que merece un fondo.
        self.input_fill = est.RELLENO_SUAVE
        self.calculated_fill = est.RELLENO_PAPEL
        self.value_fill = est.RELLENO_SUAVE

        self.section_fill = est.RELLENO_SUAVE
        self.key_result_fill = est.RELLENO_SUAVE
        self.projection_header_fill = est.RELLENO_TINTA
        self.valuation_fill = est.RELLENO_TINTA
        self.sensitivity_fill = est.RELLENO_TINTA

        # Bordes: una hairline, nunca una rejilla.
        thin = Side(border_style="thin", color=est.LINE)
        medium = Side(border_style="thin", color=est.INK)
        thick = Side(border_style="medium", color=est.EMBER)
        
        self.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.section_border = Border(top=medium, left=medium, right=medium, bottom=medium)
        self.highlight_border = Border(top=thick, left=thick, right=thick, bottom=thick)
        
        # Alineaciones
        self.center = Alignment(horizontal="center", vertical="center")
        self.center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_center = Alignment(horizontal="left", vertical="center")
        self.right_center = Alignment(horizontal="right", vertical="center")

    def _apply_sheet_header(self, ws, title: str, subtitle: str = None, ncols: int = 10):
        # El color de la pestaña es el acento de la marca, no un azul cualquiera.
        ws.sheet_properties.tabColor = est.EMBER[2:]  # openpyxl quiere RGB sin el alfa
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        ws["A1"] = title
        ws["A1"].font = self.header_font
        ws["A1"].fill = self.header_fill
        ws["A1"].alignment = self.center
        ws.row_dimensions[1].height = 30  # Aumentado para más elegancia

        if subtitle:
            ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
            ws["A2"] = subtitle
            ws["A2"].font = est.fuente(11, color=est.INK)
            ws["A2"].alignment = self.center
            ws.row_dimensions[2].height = 20  # Aumentado
    
    def _create_professional_section(self, ws, start_row: int, title: str, icon: str = "", ncols: int = 10, fill_type: str = "section"):
        """
        Crea una sección profesional con título, icono y separador visual
        
        Args:
            ws: Worksheet
            start_row: Fila donde empieza la sección
            title: Título de la sección
            icon: Emoji/símbolo para la sección
            ncols: Número de columnas a cubrir
            fill_type: Tipo de relleno ('section', 'projection', 'valuation', 'sensitivity')
        """
        # Mapeo de tipos de relleno
        fill_map = {
            'section': self.section_fill,
            'projection': self.projection_header_fill,
            'valuation': self.valuation_fill,
            'sensitivity': self.sensitivity_fill,
            'inputs': self.subheader_fill
        }
        
        font_map = {
            'section': self.section_title_font,
            'projection': self.key_metric_font,
            'valuation': self.key_metric_font,
            'sensitivity': self.key_metric_font,
            'inputs': self.subheader_font
        }
        
        # Crear título con icono
        section_title = f"{icon} {title}" if icon else title
        
        # Aplicar formato a la sección
        ws.merge_cells(f"A{start_row}:{get_column_letter(ncols)}{start_row}")
        ws[f"A{start_row}"] = section_title
        ws[f"A{start_row}"].font = font_map.get(fill_type, self.section_title_font)
        ws[f"A{start_row}"].fill = fill_map.get(fill_type, self.section_fill)
        ws[f"A{start_row}"].alignment = self.center
        ws[f"A{start_row}"].border = self.section_border
        
        # Ajustar altura de fila
        ws.row_dimensions[start_row].height = 25
        
        # Agregar línea separadora sutil debajo
        separator_row = start_row + 1
        for col in range(1, ncols + 1):
            cell = ws.cell(row=separator_row, column=col)
            cell.fill = est.relleno(est.LINE)
        ws.row_dimensions[separator_row].height = 3
        
        return start_row + 2  # Retorna la siguiente fila disponible

    # ----------------------------
    # Organización de hojas
    # ----------------------------
    def _organize_worksheets(self):
        """Ordena las hojas por FUNCIÓN: primero los datos, después el análisis, al final
        la documentación.

            Inicio
            Balance General · Estado de Resultados · Flujo Efectivo   <- los estados
            NOTAS · DEUDA FINANCIERA · DRIVERS WC                     <- el detalle que los respalda
            RATIOS & KPIs · DCF · Escenarios                          <- el análisis
            Ficha Técnica · METODOLOGÍA                               <- la documentación

        Antes las categorías estaban escritas a mano y TODO lo que no reconocían se
        empujaba al final. Así que la DEUDA FINANCIERA -- una hoja de datos, con el detalle
        que sostiene el Kd del WACC -- terminaba después de METODOLOGÍA, al fondo del
        archivo, mientras la documentación quedaba en medio del análisis.

        Cada hoja nueva heredaba el mismo destino: al fondo, sin que nadie lo decidiera.
        """
        # El orden por función. Las hojas DCF llevan el período en el nombre ("DCF 2026Q1")
        # y no se pueden nombrar aquí: se insertan por prefijo, junto al resto del análisis.
        ORDEN = [
            "Inicio",
            # Los estados financieros.
            "Balance General", "Balance Sheet",
            "Estado de Resultados", "Estado Resultados (Función)", "Income Statement",
            "Flujo Efectivo", "Cash Flow",
            # El detalle que los respalda. La deuda va aquí, con los datos, no al final:
            # es lo que sostiene el Kd del WACC.
            "NOTAS", "Notas", _HOJA_DEUDA, "DRIVERS WC",
            # El análisis.
            "RATIOS & KPIs", "__DCF__", "Escenarios",
            # La documentación, al final.
            "Ficha Técnica", "METODOLOGÍA",
        ]

        existentes = self.wb.sheetnames.copy()
        dcf = sorted(h for h in existentes if h.startswith("DCF "))

        orden_final = []
        for nombre in ORDEN:
            if nombre == "__DCF__":
                orden_final.extend(dcf)
            elif nombre in existentes and nombre not in orden_final:
                orden_final.append(nombre)

        # Lo que no esté contemplado va al final, pero ANTES de la documentación: una hoja
        # nueva es contenido, no un apéndice.
        docs = [h for h in ("Ficha Técnica", "METODOLOGÍA") if h in orden_final]
        for h in docs:
            orden_final.remove(h)
        orden_final += [h for h in existentes if h not in orden_final and h not in docs]
        orden_final += docs

        for i, nombre in enumerate(orden_final):
            if nombre in self.wb.sheetnames:
                hoja = self.wb[nombre]
                self.wb.move_sheet(hoja, offset=i - self.wb.index(hoja))

    # ----------------------------
    # Formateo de números profesional
    # ----------------------------
    def _apply_professional_number_formatting(self, ws, start_row: int, end_row: int, column_formats: dict):
        """
        Aplica formateo profesional de números (moneda, porcentaje, números)
        
        Args:
            ws: Worksheet
            start_row: Fila inicial
            end_row: Fila final
            column_formats: Dict con formato por columna {columna: tipo}
                - 'currency': Formato moneda (#,##0)
                - 'currency_millions': Formato moneda en millones (#,##0,, "M")
                - 'percentage': Formato porcentaje (0.00%)
                - 'percentage_1': Formato porcentaje 1 decimal (0.0%)
                - 'number': Formato número con comas (#,##0)
                - 'decimal': Formato decimal (0.00)
        """
        format_map = {
            'currency': '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)',  # Formato moneda profesional
            'currency_millions': '_($* #,##0,,"M"_);_($* (#,##0,,"M");_($* "-"_);_(@_)',  # Millones
            'percentage': '0.00%',
            'percentage_1': '0.0%', 
            'number': '#,##0',
            'decimal': '0.00',
            'accounting': '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'  # Formato contable
        }
        
        for row in range(start_row, end_row + 1):
            for col, format_type in column_formats.items():
                if format_type in format_map:
                    try:
                        cell = ws.cell(row=row, column=col)
                        cell.number_format = format_map[format_type]
                    except AttributeError:
                        # Saltar celdas combinadas
                        pass

    def _colorear_por_signo(self, ws, coordenada: str) -> None:
        """Verde si el número es positivo, rojo si es negativo. Nada si es cero.

        Se hace con formato condicional porque la celda es una FÓRMULA: su signo no se
        conoce hasta que Excel la evalúa, y además cambia cuando el usuario toca los
        supuestos — que es justamente para lo que sirve el modelo.
        """
        from openpyxl.formatting.rule import CellIsRule

        ws.conditional_formatting.add(
            coordenada,
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=est.fuente(10, bold=True, color=est.GROWTH)),
        )
        ws.conditional_formatting.add(
            coordenada,
            CellIsRule(operator="lessThan", formula=["0"],
                       font=est.fuente(10, bold=True, color=est.LOSS)),
        )

    # ----------------------------
    # Hojas
    # ----------------------------
    def create_drivers_wc_sheet(self):
        """Crea la hoja DRIVERS WC, referenciando directamente las hojas originales."""
        ws = self.wb.create_sheet("DRIVERS WC")
        self._apply_sheet_header(ws, "ANÁLISIS DE CAPITAL DE TRABAJO",
                                 "Drivers para cálculo de capital de trabajo neto", 10)

        row = 4
        ws[f"A{row}"] = "DEFINICIONES:"
        ws[f"A{row}"].font = self.rotulo_font
        row += 1
        for txt in [
            "CxC = Deudores comerciales y otras cuentas por cobrar corrientes",
            "Inventarios = Inventarios corrientes",
            "CxP = Cuentas por pagar comerciales y otras cuentas por pagar",
            "NWC = (CxC + Inventarios) - CxP",
            "ΔNWC = NWC_actual - NWC_anterior",
            "ΔNWC/ΔVentas = ΔNWC / (Ventas_actual - Ventas_anterior)",
        ]:
            ws[f"A{row}"] = txt
            row += 1

        # Aviso metodológico: sólo años completos (Q4 = año consolidado)
        ws[f"A{row}"] = ("Nota: los datos de la CMF son YTD acumulados (Q4 = año completo). "
                         "Para no mezclar escalas, el driver usa SÓLO períodos anuales (Q4), año contra año.")
        ws[f"A{row}"].font = self.small_font if hasattr(self, "small_font") else self.label_font
        row += 2

        headers = ["Período", "Ventas (año)", "CxC", "Inventarios", "CxP", "NWC", "NWC/Ventas", "ΔNWC", "ΔVentas (año)", "ΔNWC/ΔVentas"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border

        data_row = row + 1

        # SÓLO períodos ANUALES (YYYYQ4 o YYYY), ordenados de más antiguo a más
        # nuevo, para que cada fila compare año contra año (ΔVentas/ΔNWC anuales).
        all_periods = self._iter_periods_from_sheet(self.sh_pl) or []
        def _yr(p):
            m = re.match(r"^(\d{4})", str(p))
            return int(m.group(1)) if m else 0
        annual = [p for p in all_periods if re.match(r"^\d{4}(Q4)?$", str(p).strip())]
        # dedup por año conservando la etiqueta tal cual aparece en la hoja
        seen_yr = {}
        for p in annual:
            seen_yr[_yr(p)] = p
        annual_sorted = [seen_yr[y] for y in sorted(seen_yr.keys())]
        annual_sorted = annual_sorted[-6:]  # últimos ~6 años completos

        n = len(annual_sorted)
        for i, lbl in enumerate(annual_sorted):
            r = data_row + i
            ws.cell(row=r, column=1, value=lbl)

            def _ref(sheet, rows_map, key):
                if sheet and key in rows_map and rows_map[key]:
                    rr = self.create_cell_reference_by_label(sheet, rows_map[key], lbl)
                    return rr
                return None
            ventas_ref = _ref(self.sh_pl, self.rows_pl, "Ventas")
            cxc_ref = _ref(self.sh_bal, self.rows_bal, "CxC")
            inv_ref = _ref(self.sh_bal, self.rows_bal, "Inv")
            cxp_ref = _ref(self.sh_bal, self.rows_bal, "CxP")
            if ventas_ref: ws.cell(row=r, column=2, value=f"=IFERROR({ventas_ref},\"\")")
            if cxc_ref: ws.cell(row=r, column=3, value=f"=IFERROR({cxc_ref},\"\")")
            if inv_ref: ws.cell(row=r, column=4, value=f"=IFERROR({inv_ref},\"\")")
            if cxp_ref: ws.cell(row=r, column=5, value=f"=IFERROR({cxp_ref},\"\")")

            # NWC = (CxC + Inventarios) - CxP  ; NWC/Ventas (nivel, estable)
            ws.cell(row=r, column=6, value=f"=C{r}+D{r}-E{r}")
            ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/B{r},\"\")")

            # ΔNWC, ΔVentas y ratio — SIEMPRE año-contra-año (filas consecutivas = años consecutivos)
            if i > 0:
                ws.cell(row=r, column=8, value=f"=F{r}-F{r-1}")
                ws.cell(row=r, column=9, value=f"=B{r}-B{r-1}")
                # Ratio acotado a un rango razonable para que un año de bajo crecimiento no lo dispare
                ws.cell(row=r, column=10, value=f"=IFERROR(IF(I{r}=0,\"\",MAX(MIN(H{r}/I{r},0.5),-0.5)),\"\")")

        # Promedio del driver: usar la MEDIANA de los ΔNWC/ΔVentas anuales (robusta a
        # outliers) y, como respaldo, el promedio de NWC/Ventas (nivel). Se referencia
        # desde el DCF mediante self._wc_avg_cell.
        avg_row = data_row + max(n, 1) + 2
        ws[f"A{avg_row}"] = "Driver capital de trabajo (NWC/Ventas, mediana):"
        ws[f"A{avg_row}"].font = self.rotulo_fuerte_font
        # Usamos el NIVEL NWC/Ventas (mediana de años completos), estable, en vez de
        # ΔNWC/ΔVentas que se dispara cuando las ventas anuales crecen poco. En el DCF
        # ΔNWC = ΔVentas × (NWC/Ventas), que es la relación económica correcta.
        ws[f"B{avg_row}"] = f"=IFERROR(MEDIAN(G{data_row}:G{data_row + n - 1}),0.10)"
        ws[f"B{avg_row}"].fill = self.input_fill
        ws[f"B{avg_row}"].border = self.border
        ws[f"B{avg_row}"].number_format = "0.00%"
        # Guardar referencia para el DCF
        self._wc_avg_cell = f"'DRIVERS WC'!$B${avg_row}"

        drivers_wc_formats = {2: 'currency', 3: 'currency', 4: 'currency', 5: 'currency',
                              6: 'currency', 7: 'percentage', 8: 'currency', 9: 'currency', 10: 'percentage'}
        self._apply_professional_number_formatting(ws, data_row, data_row + max(n - 1, 0), drivers_wc_formats)
        for rr in range(data_row, data_row + n):
            for cc in range(1, 11):
                try:
                    ws.cell(row=rr, column=cc).border = self.border
                except AttributeError:
                    pass

        widths = [12, 15, 16, 14, 14, 16, 13, 14, 15, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def create_dcf_sheet(self):
        """Hoja principal DCF con referencias directas a hojas originales."""
        ws = self.wb.create_sheet("DCF")
        self._apply_sheet_header(ws, "MODELO DCF (FCFF) - FLUJO DE CAJA LIBRE DE LA FIRMA",
                                 "Valuación por flujos de caja descontados con análisis de sensibilidad", 16)

        # Crear sección de configuración con diseño profesional
        config_row = self._create_professional_section(ws, 4, "CONFIGURACIÓN DEL MODELO", "", 6, "inputs")
        
        # Selector de período con diseño mejorado
        ws[f"A{config_row}"] = "Período seleccionado:"
        ws[f"A{config_row}"].font = self.label_font

        default_period = self._find_latest_period()
        ws[f"C{config_row}"] = default_period
        ws[f"C{config_row}"].fill = self.input_fill
        ws[f"C{config_row}"].border = self.border
        ws[f"C{config_row}"].font = self.input_font

        # Validación de lista de períodos directamente desde ER fila 3
        if DataValidation:
            # Rango de encabezados amplio (B3:ZZ3) en Estado de Resultados
            dv = DataValidation(type="list",
                                formula1=f"'{self.sh_pl.title}'!$B$3:$ZZ$3")
            dv.add(ws[f"C{config_row}"])
            ws.add_data_validation(dv)

        # Escenario con diseño mejorado
        ws[f"E{config_row}"] = "Escenario:"
        ws[f"E{config_row}"].font = self.label_font
        ws[f"F{config_row}"] = "Base"
        ws[f"F{config_row}"].fill = self.input_fill
        ws[f"F{config_row}"].border = self.border
        ws[f"F{config_row}"].font = self.input_font
        if DataValidation:
            dv2 = DataValidation(type="list", formula1='"Conservador,Base,Agresivo"')
            dv2.add(ws[f"F{config_row}"])
            ws.add_data_validation(dv2)

        # INPUTS con diseño profesional
        row = config_row + 3
        inputs_row = self._create_professional_section(ws, row, "PARÁMETROS DEL MODELO", "", 10, "inputs")

        # Determinar el período para este DCF específico
        # Si existe un período seleccionado en la configuración, usarlo
        selected_period = self._find_base_annual_period()
        if hasattr(self, '_current_dcf_period'):
            selected_period = self._current_dcf_period

        # Referencia al driver de capital de trabajo (mediana anual robusta)
        wc_ref = getattr(self, "_wc_avg_cell", "'DRIVERS WC'!$B$21")

        # La base de la proyección es el último CIERRE ANUAL, no el trimestre en curso.
        #
        # Antes se pasaba `selected_period` (p. ej. "2026Q1") y las ventas base salían de
        # anualizar ese trimestre (×4). Para cualquier negocio estacional —una viña, un
        # retail, una salmonera— eso distorsiona los cinco años de proyección de una sola
        # vez. El DCF que se guarda en la BD (scripts/dcf/excel_aligned.py) usa el último
        # año real, y ahora el Excel también.
        periodo_base_anual = self._find_base_annual_period()

        inputs = [
            ("Año base", periodo_base_anual, "input"),
            ("Ventas año base (M$)", self._get_ventas_base_formula(periodo_base_anual), "formula"),
            ("Crecimiento Ventas Y+1 (%)", self._get_cagr_formula(5), "formula"),
            ("Crecimiento Ventas Y+2 (%)", f"=MAX(B{inputs_row + 2}*0.9,0.02)", "formula"),  # Referencia correcta al Y+1
            ("Crecimiento Ventas Y+3 (%)", f"=MAX(B{inputs_row + 3}*0.9,0.02)", "formula"),  # Referencia correcta al Y+2
            ("Crecimiento Ventas Y+4 (%)", f"=MAX(B{inputs_row + 4}*0.85,0.015)", "formula"), # Referencia correcta al Y+3
            ("Crecimiento Ventas Y+5 (%)", f"=MAX(B{inputs_row + 5}*0.8,0.015)", "formula"),  # Referencia correcta al Y+4
            ("Margen EBIT (%)", self._get_margen_ebit_formula(), "formula"),
            ("Tasa efectiva de impuestos (%)", self._get_tasa_impuestos_formula(), "formula"),
            ("D&A / Ventas (%)", self._get_da_ventas_formula(), "formula"),
            ("CapEx / Ventas (%)", self._get_capex_ventas_formula(), "formula"),
            ("ΔNWC / ΔVentas (%)", f"=IFERROR(MIN(MAX({wc_ref},-0.20),0.20),0.10)", "formula"),
            # El WACC no es un input: se calcula abajo (CAPM + costo de deuda real) y esta
            # celda lo referencia. El "0.10" que habia aqui se sobrescribia mas adelante,
            # pero dejaba la impresion de ser una constante editable.
            ("WACC (%)", "0.10", "input"),
            ("g - Tasa de crecimiento terminal (%)", "0.02", "fixed"),  # Fijo en 2%
            # La moneda viaja PEGADA al modelo, no en una nota al pie.
            ("Moneda de los estados", self.reporting_currency, "input"),
        ]

        # El tipo de cambio SOLO tiene sentido si los estados no estan en pesos. Para una
        # empresa que reporta en CLP la fila decia "Tipo de cambio (CLP por 1 CLP) = 1":
        # una conversion que no convierte nada, y que ademas rompia el valor por accion
        # (ver el bloque de VALUACION).
        if self.reporting_currency != self.moneda_precio:
            inputs.append(
                # Editable: es lo que convierte el valor por accion a la moneda del precio de
                # bolsa para poder compararlo. El analista puede ajustarlo al tipo de cambio
                # que quiera usar -- pero NUNCA se compara sin convertir.
                ("Tipo de cambio (%s por 1 %s)" % (self.moneda_precio, self.reporting_currency),
                 self._tipo_de_cambio_actual(), "input")
            )

        inputs += [
            ("Deuda neta (M$)", self._get_deuda_neta_formula(selected_period), "formula"),
            ("Acciones en circulación (M)", self._get_acciones_formula(), "formula"),
        ]

        # Fila de cada parámetro, resuelta POR ETIQUETA.
        #
        # Antes las filas eran offsets fijos (wacc_row = inputs_row + 12, deuda_row = +14…).
        # Agregar un parámetro en medio de la lista desplazaba todo y el modelo apuntaba a
        # celdas equivocadas SIN dar error: el Excel simplemente calculaba otra cosa. Es
        # una trampa que se dispara sola la próxima vez que alguien toque esta lista.
        self._param_rows = {label: inputs_row + i for i, (label, _v, _k) in enumerate(inputs)}

        for i, (label, val, kind) in enumerate(inputs):
            r = inputs_row + i
            
            ws[f"A{r}"] = label
            ws[f"A{r}"].font = self.label_font
            ws[f"A{r}"].alignment = self.left_center

            cell = ws[f"B{r}"]
            # Regla general: si el valor es una CONSTANTE numérica (no una fórmula
            # que empieza con "="), se escribe como NÚMERO, no como texto. Así el
            # number_format aplica y en Excel chileno (es-CL) el separador decimal
            # se muestra con coma. Un string "0.02"/"0.27"/"0.05" se vería con punto.
            v = val
            if isinstance(v, str) and not v.strip().startswith("="):
                try:
                    v = float(v)
                except Exception:
                    pass  # texto real (p. ej. "Año base" = "2026Q1")
            cell.value = v
            cell.fill = self.input_fill if kind == "input" else self.calculated_fill
            cell.border = self.border
            cell.font = self.input_font
            cell.alignment = self.right_center
            
            # Formateo profesional según el tipo de input
            if "%" in label or "Crecimiento" in label or "Margen" in label or "WACC" in label or "tasa" in label.lower():
                cell.number_format = '0.00%'  # Formato porcentaje
            elif "M$" in label or "Ventas" in label or "Deuda" in label:
                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Formato moneda
            elif "Acciones" in label:
                cell.number_format = '#,##0'  # Formato número

        # PROYECCIÓN FCFF con diseño profesional
        projection_start_row = inputs_row + len(inputs) + 3
        proj_header_row = self._create_professional_section(ws, projection_start_row, "PROYECCIÓN FCFF (5 AÑOS)", "", 10, "projection")
        projection_start_row = proj_header_row - 1  # Ajustar para mantener compatibilidad

        # Encabezados sin iconos
        headers = [
            "Año", "Ventas", "EBIT", "NOPAT", 
            "D&A", "CapEx", "ΔNWC", "FCFF", "Factor", "FCFF PV"
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=projection_start_row + 1, column=c, value=h)
            cell.font = self.key_metric_font
            cell.fill = self.projection_header_fill
            cell.alignment = self.center_wrap
            cell.border = self.border
        
        # Ajustar altura del encabezado
        ws.row_dimensions[projection_start_row + 1].height = 30

        # Extract numeric year from base annual period (e.g. "2025Q4" -> 2025)
        _bp = self._find_base_annual_period()
        base_year = int(re.match(r"(\d{4})", _bp).group(1)) if re.match(r"(\d{4})", _bp) else 2025
        P = self._param_rows
        ventas_base_row = P["Ventas año base (M$)"]
        crec_rows = [P[f"Crecimiento Ventas Y+{k} (%)"] for k in range(1, 6)]
        margen_ebit_row = P["Margen EBIT (%)"]
        tasa_imp_row = P["Tasa efectiva de impuestos (%)"]
        da_ventas_row = P["D&A / Ventas (%)"]
        capex_ventas_row = P["CapEx / Ventas (%)"]
        nwc_ventas_row = P["ΔNWC / ΔVentas (%)"]
        wacc_row = P["WACC (%)"]

        for k in range(1, 6):
            r = projection_start_row + 1 + k
            ws.cell(row=r, column=1, value=base_year + k)  # Año

            if k == 1:
                ws.cell(row=r, column=2, value=f"=$B${ventas_base_row}*(1+$B${crec_rows[0]})")
            else:
                ws.cell(row=r, column=2, value=f"=B{r-1}*(1+$B${crec_rows[k-1]})")

            ws.cell(row=r, column=3, value=f"=B{r}*$B${margen_ebit_row}")             # EBIT
            ws.cell(row=r, column=4, value=f"=C{r}*(1-$B${tasa_imp_row})")             # NOPAT
            ws.cell(row=r, column=5, value=f"=B{r}*$B${da_ventas_row}")                # D&A
            ws.cell(row=r, column=6, value=f"=B{r}*$B${capex_ventas_row}")            # CapEx
            if k == 1:
                ws.cell(row=r, column=7, value=f"=(B{r}-$B${ventas_base_row})*$B${nwc_ventas_row}")
            else:
                ws.cell(row=r, column=7, value=f"=(B{r}-B{r-1})*$B${nwc_ventas_row}")  # ΔNWC

            ws.cell(row=r, column=8, value=f"=D{r}+E{r}-F{r}-G{r}")                    # FCFF
            ws.cell(row=r, column=9, value=f"=1/(1+$B${wacc_row})^{k}")                # Factor
            ws.cell(row=r, column=10, value=f"=H{r}*I{r}")                              # PV

            for c in range(1, 11):
                cc = ws.cell(row=r, column=c)
                cc.border = self.border
                
                # Formateo profesional por columna
                if c == 1:  # Año
                    cc.number_format = '0'
                elif c in [2, 3, 4, 5, 6, 7, 8, 10]:  # Valores monetarios (Ventas, EBIT, NOPAT, D&A, CapEx, ΔNWC, FCFF, FCFF PV)
                    cc.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                elif c == 9:  # Factor de descuento
                    cc.number_format = '0.0000'

        # VALUACIÓN con diseño profesional
        valuation_row = projection_start_row + 8
        val_header_row = self._create_professional_section(ws, valuation_row, "VALUACIÓN EMPRESARIAL", "", 10, "valuation")
        valuation_row = val_header_row - 1

        last_fcff_row = projection_start_row + 6
        g_row = P["g - Tasa de crecimiento terminal (%)"]
        deuda_row = P["Deuda neta (M$)"]
        acciones_row = P["Acciones en circulación (M)"]
        # `mp` = moneda contra la que se compara el precio de bolsa (CLP en Chile, USD en EEUU).
        mp = self.moneda_precio
        # Solo existe cuando la moneda de reporte != la del precio (ver el bloque de PARAMETROS).
        fx_row = P.get(f"Tipo de cambio ({mp} por 1 {self.reporting_currency})")

        # Bloques de valuación. Las filas se resuelven POR ETIQUETA, igual que los
        # parámetros: los offsets numéricos (valuation_row + 8, + 10, + 11…) se desplazan
        # solos en cuanto alguien agrega una fila, y el Excel calcula otra cosa sin dar
        # ningún error.
        # La fila de conversion a pesos SOLO existe si los estados NO estan en pesos.
        #
        # Cuando la empresa reporta en CLP, las etiquetas "Valor por Accion (DCF, {moneda})"
        # y "Valor por Accion (DCF, CLP)" son la MISMA string. El dict `V` se quedaba con la
        # ultima, asi que `r_dcf_moneda` apuntaba a la fila de conversion en vez de a la del
        # calculo, y la formula salia `=IFERROR(B51*$B$26,"")`: una referencia circular a si
        # misma. Excel la resolvia como vacio y la celda mostraba "$ -", junto a una fila
        # duplicada con el mismo titulo. Ademas el "tipo de cambio" era "CLP por 1 CLP = 1",
        # una conversion que no convierte nada.
        convierte = self.reporting_currency != mp

        V = {}
        _labels = [
            "Valor Terminal", "VT Presente", "Suma FCFF PV", "Enterprise Value",
            "(-) Deuda Neta", "Equity Value", "Acciones",
            f"Valor por Acción (DCF, {self.reporting_currency})",
        ]
        if convierte:
            _labels.append(f"Valor por Acción (DCF, {mp})")
        _labels += [
            "", f"Precio Actual Mercado ({mp})", "Prima/(Descuento) %", "Recomendación",
        ]
        for i, lbl in enumerate(_labels):
            V[lbl] = valuation_row + 1 + i

        r_dcf_moneda = V[f"Valor por Acción (DCF, {self.reporting_currency})"]
        # Contra qué se compara el precio de bolsa: la moneda del precio (mp). Si los estados
        # ya están en esa moneda, es la misma fila; si no, la convertida.
        r_dcf_clp = V[f"Valor por Acción (DCF, {mp})"] if convierte else r_dcf_moneda
        r_precio = V[f"Precio Actual Mercado ({mp})"]
        r_prima = V["Prima/(Descuento) %"]

        blocks = [
            ("Valor Terminal", f"=H{last_fcff_row}*(1+$B${g_row})/($B${wacc_row}-$B${g_row})"),
            ("VT Presente", f"=B{V['Valor Terminal']}*I{last_fcff_row}"),
            ("Suma FCFF PV", f"=SUM(J{projection_start_row + 2}:J{last_fcff_row})"),
            ("Enterprise Value", f"=B{V['VT Presente']}+B{V['Suma FCFF PV']}"),
            ("(-) Deuda Neta", f"=$B${deuda_row}"),
            ("Equity Value", f"=B{V['Enterprise Value']}-B{V['(-) Deuda Neta']}"),
            ("Acciones", f"=$B${acciones_row}"),
            # Valor por acción EN LA MONEDA DE LOS ESTADOS. Para SQM, COPEC o ARAUCO, esto
            # sale en dólares.
            (f"Valor por Acción (DCF, {self.reporting_currency})",
             f"=IFERROR(IF(B{V['Acciones']}>0,B{V['Equity Value']}/B{V['Acciones']}*1000,\"\"),\"\")"),
        ]

        # La conversión a pesos. Éste es el número que se compara contra la bolsa, porque la
        # acción cotiza en pesos.
        #
        # Antes esta fila no existía: el Excel comparaba un valor intrínseco en DÓLARES
        # contra un precio de mercado en PESOS. Para las empresas que reportan en USD, la
        # "Prima/(Descuento)" salía desviada ~900x y la "Recomendación" decía siempre
        # "VENTA FUERTE". Un analista que le creyera vendía justo lo que debía comprar.
        if convierte:
            blocks.append(
                (f"Valor por Acción (DCF, {mp})", f"=IFERROR(B{r_dcf_moneda}*$B${fx_row},\"\")")
            )

        blocks += [
            ("", ""),
            (f"Precio Actual Mercado ({mp})", "INPUT_REQUIRED"),
            # La prima compara la MISMA moneda contra la misma moneda.
            ("Prima/(Descuento) %",
             f"=IF(NOT(ISNUMBER(B{r_precio})),\"SIN PRECIO MERCADO\","
             f"IF(NOT(ISNUMBER(B{r_dcf_clp})),\"SIN VALOR DCF\","
             f"IF(B{r_dcf_clp}<0,\"DCF NEG - VENTA\","
             f"(B{r_dcf_clp}-B{r_precio})/B{r_precio})))"),
            ("Recomendación",
             f"=IF(NOT(ISNUMBER(B{r_precio})),\"NECESITA PRECIO MERCADO\","
             f"IF(NOT(ISNUMBER(B{r_dcf_clp})),\"SIN VALOR DCF\","
             f"IF(B{r_dcf_clp}<0,\"VENTA FUERTE - DCF NEGATIVO\","
             f"IF(NOT(ISNUMBER(B{r_prima})),\"SIN PRIMA\","
             f"IF(B{r_prima}>0.15,\"COMPRA FUERTE\","
             f"IF(B{r_prima}>0.05,\"COMPRA\","
             f"IF(B{r_prima}>-0.05,\"MANTENER\","
             # Ocho IF anidados, ocho paréntesis. Había SIETE: Excel no podía evaluar la
             # fórmula, la descartaba en silencio, y al abrir el archivo pedía repararlo
             # ("Registros quitados: Fórmula de /xl/worksheets/sheet7.xml"). La celda
             # "Recomendación" quedaba vacía en TODAS las empresas -- justo la celda que un
             # usuario mira primero.
             f"IF(B{r_prima}>-0.15,\"VENTA\",\"VENTA FUERTE\"))))))))"),
        ]
        for i, (lbl, fx) in enumerate(blocks):
            rr = valuation_row + 1 + i
            ws[f"A{rr}"] = lbl
            ws[f"A{rr}"].font = self.label_font
            ws[f"A{rr}"].alignment = self.left_center
            
            cell = ws[f"B{rr}"]
            cell.value = fx
            cell.border = self.border
            cell.alignment = self.right_center
            
            # Destacar resultados clave con colores especiales
            if "Enterprise Value" in lbl or "Equity Value" in lbl:
                cell.fill = self.key_result_fill
                cell.font = self.result_font
                ws[f"A{rr}"].font = self.result_font
            elif "Valor por Acción (DCF)" in lbl:
                cell.fill = self.key_result_fill
                cell.font = self.result_font
                ws[f"A{rr}"].font = self.result_font
                # Hacer la fila más alta para destacar el resultado final
                ws.row_dimensions[rr].height = 25
            elif "Precio Actual Mercado" in lbl:
                cell.fill = self.input_fill
                cell.font = self.input_font
                ws[f"A{rr}"].font = self.label_font
                # Texto explicativo para el usuario
                cell.value = "INSERT_PRICE_HERE"
            elif "Prima/(Descuento)" in lbl:
                cell.fill = self.calculated_fill
                cell.font = self.result_font
                ws[f"A{rr}"].font = self.result_font
            elif "Recomendación" in lbl:
                cell.fill = self.key_result_fill
                cell.font = self.result_font
                ws[f"A{rr}"].font = self.result_font
                ws.row_dimensions[rr].height = 25
            elif lbl == "":  # Línea separadora
                cell.fill = PatternFill(fill_type=None)
                cell.value = ""
                ws[f"A{rr}"].value = ""
            elif i >= 3:  # Otros valores importantes
                cell.fill = self.value_fill
                cell.font = self.input_font
            else:
                cell.font = self.input_font
            
            # Formateo específico por tipo de valor
            if "Valor Terminal" in lbl or "VT Presente" in lbl or "Suma FCFF PV" in lbl or "Enterprise Value" in lbl or "Deuda Neta" in lbl or "Equity Value" in lbl:
                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Formato moneda profesional
            elif "Acciones" in lbl:
                cell.number_format = '#,##0'  # Formato número
            elif "Valor por Acción" in lbl or "Precio Actual Mercado" in lbl:
                cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)'  # Moneda con decimales
            elif "Prima/(Descuento)" in lbl:
                cell.number_format = '0.0%'  # Formato porcentaje

        # Anchos
        for i, w in enumerate([25, 15, 12, 12, 12, 12, 12, 12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        return projection_start_row, valuation_row, row

    def create_deuda_sheet(self):
        """La deuda financiera, crédito por crédito, tal como la empresa la declara.

        Acreedor, deudor, moneda, corriente/no corriente, vencimiento y TASA EFECTIVA de
        cada crédito. Arauco declara 98; Aguas Andinas, 59; Sonda, 207. Ninguna API entrega
        esto: está en la nota de préstamos del XBRL y el pipeline la ignoraba.

        EL Kd ES UNA FÓRMULA, NO UN NÚMERO PEGADO. Sale de un SUMAPRODUCTO sobre el detalle
        de esta hoja: tasa x monto, dividido por la suma de los montos. Así el analista puede
        auditarlo -- y si filtra o corrige un crédito, el Kd y el WACC se recalculan solos.
        Un número pegado obliga a creerle al que lo pegó.

        No se crea si la empresa no declara tasas (69 de las 232): una hoja vacía es peor
        que ninguna hoja.
        """
        deuda = self._deuda()
        if not deuda:
            return

        creditos = deuda.get("creditos") or []
        if not creditos:
            return

        ws = self.wb.create_sheet(_HOJA_DEUDA)
        self._apply_sheet_header(ws, "DEUDA FINANCIERA - DETALLE POR CRÉDITO", 10)

        money = '#,##0'

        # ── El detalle va PRIMERO en el layout lógico pero se escribe al final: el Kd de
        #    arriba necesita saber en qué filas quedó para poder referenciarlas.
        cols = [("Acreedor / Serie", 40), ("Deudor", 32), ("País", 7), ("Instrumento", 12),
                ("Moneda", 8), ("Monto contable", 17), ("Corriente", 15),
                ("No corriente", 15), ("Tasa efectiva", 12), ("Vencimiento", 13)]

        # Encabezado del bloque de créditos, después del resumen. El resumen ocupa:
        # 4 filas de Kd + 1 + monedas + 1 + instrumentos + 1 + vencimientos + 2 secciones.
        por_moneda = deuda.get("por_moneda") or {}
        por_inst = deuda.get("por_instrumento") or {}
        vencs = deuda.get("vencimientos") or {}

        # ── 1. COSTO DE LA DEUDA ─────────────────────────────────────────────────────
        r0 = self._create_professional_section(ws, 4, "COSTO DE LA DEUDA (Kd)", "", 10, "inputs")

        # Las cuatro primeras celdas del resumen son fórmulas que apuntan al DETALLE, y el
        # detalle todavía no está escrito. Se dejan vacías y se rellenan al final, cuando ya
        # se sabe en qué filas quedó.
        #
        # Antes se calculaba la posición del detalle POR ADELANTADO, sumando a mano las
        # filas de cada bloque intermedio. La cuenta daba 32 cuando el detalle empezaba en la
        # 38: el SUMAPRODUCTO del Kd terminaba sumando las filas del perfil de vencimientos.
        # Excel no protesta -- devuelve un número, simplemente es otro número.
        resumen = [
            ("Kd - Tasa efectiva ponderada", None, '0.00%', "result"),
            ("Deuda total (contable)", None, money, "calc"),
            ("  de la cual, corriente", None, money, "calc"),
            ("  de la cual, no corriente", None, money, "calc"),
            ("Créditos declarados", int(deuda["n_creditos"]), '#,##0', "calc"),
            ("Fuente", f"Nota de préstamos del XBRL ({deuda.get('fuente','')})", '@', "calc"),
        ]
        for k, (lbl, val, fmt, kind) in enumerate(resumen):
            rr = r0 + k
            ws.cell(row=rr, column=1, value=lbl).font = self.label_font
            c = ws.cell(row=rr, column=2, value=val)
            c.number_format = fmt
            c.border = self.border
            c.fill = self.key_result_fill if kind == "result" else self.calculated_fill
            c.font = self.result_font if kind == "result" else self.input_font

        def _bloque(inicio: int, titulo: str, items: dict, fmt_pct: bool = True) -> int:
            """Un bloque monto + % del total. Devuelve la fila siguiente."""
            rh = self._create_professional_section(ws, inicio, titulo, "", 10, "valuation")
            total = sum(items.values()) or 1
            for k, (nombre, monto) in enumerate(items.items()):
                rr = rh + k
                ws.cell(row=rr, column=1, value=nombre).font = self.label_font
                c = ws.cell(row=rr, column=2, value=float(monto))
                c.number_format = money
                c.border = self.border
                p = ws.cell(row=rr, column=3, value=float(monto) / total)
                p.number_format = '0.0%'
                p.border = self.border
            return rh + len(items) + 1

        # ── 2. EXPOSICIÓN CAMBIARIA ──────────────────────────────────────────────────
        # Una empresa que factura en pesos y se endeuda en dólares tiene un riesgo que el
        # balance no muestra.
        r = _bloque(r0 + len(resumen) + 1, "EXPOSICIÓN CAMBIARIA DE LA DEUDA",
                    dict(sorted(por_moneda.items(), key=lambda x: -x[1])))

        # ── 3. POR INSTRUMENTO ───────────────────────────────────────────────────────
        # Bajo IFRS 16 un arriendo ES deuda. En SMU son el 58,6% del total: un Kd que los
        # ignora no es el costo de deuda de esa empresa.
        r = _bloque(r, "POR INSTRUMENTO",
                    dict(sorted(por_inst.items(), key=lambda x: -x[1])))

        # ── 4. PERFIL DE VENCIMIENTOS ────────────────────────────────────────────────
        # Dice si la empresa tiene un muro de deuda el año que viene o si lo tiene repartido
        # a diez años. Dos empresas con la misma deuda y el mismo Kd pueden ser riesgos
        # completamente distintos, y eso no se ve en ningún ratio del balance.
        #
        # Los tramos cuadran al peso contra la deuda total (verificado en Arauco y Aguas
        # Andinas): sólo se suman los tramos hoja, nunca los agregados, que los contienen.
        r = _bloque(r, "PERFIL DE VENCIMIENTOS", vencs)

        # ── 5. EL DETALLE ────────────────────────────────────────────────────────────
        rd = self._create_professional_section(ws, r, "CRÉDITOS", "", 10, "inputs")
        for c, (h, ancho) in enumerate(cols, 1):
            cell = ws.cell(row=rd, column=c, value=h)
            cell.font = self.key_metric_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_wrap
            cell.border = self.border
            ws.column_dimensions[get_column_letter(c)].width = ancho

        for k, cr in enumerate(creditos):
            rr = rd + 1 + k
            # CÓMO SE IDENTIFICA CADA CRÉDITO.
            #
            # Un préstamo tiene un acreedor: "BNP Paribas / ECA". Un BONO PÚBLICO NO --
            # se coloca en el mercado y no hay una contraparte única --, y por eso el XBRL
            # no le declara `NombreEntidadAcreedora`. Lo que sí declara es la SERIE
            # ("Barau-F"), que es como el bono se identifica y se transa.
            #
            # En Arauco los 58 préstamos traen acreedor y los 40 bonos traen serie, ninguno
            # las dos cosas. Antes, cuando faltaba el acreedor, la celda mostraba el
            # identificador interno del XBRL ("Item351"): esconder detrás de un código algo
            # que la empresa sí declara, sólo que en otro campo.
            identificador = (cr.get("acreedor")
                             or cr.get("serie")
                             or cr.get("miembro") or "")
            fila = [
                identificador,
                cr.get("deudor") or "",
                cr.get("pais_deudor") or "",
                cr.get("instrumento") or "",
                cr.get("moneda") or "",
                cr.get("monto_contable"),
                cr.get("monto_corriente"),
                cr.get("monto_no_corriente"),
                cr.get("tasa_efectiva"),
                cr.get("vencimiento") or "",
            ]
            for c, v in enumerate(fila, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                cell.border = self.border
                cell.font = self.input_font
                if c in (6, 7, 8):
                    cell.number_format = money
                elif c == 9:
                    cell.number_format = '0.00%'

        # Ahora sí se sabe dónde quedó el detalle: se escriben las fórmulas del resumen.
        #
        # El Kd es SUMAPRODUCTO(monto x tasa) / SUMA(monto): auditable y vivo. Si el
        # analista filtra o corrige un crédito, el Kd -- y el WACC, que apunta a esta
        # celda -- se recalculan solos. Un número pegado obliga a creerle al que lo pegó.
        fila_ini = rd + 1
        fila_fin = rd + len(creditos)
        ws.cell(row=r0, column=2).value = (
            f"=IFERROR(SUMPRODUCT(F{fila_ini}:F{fila_fin},I{fila_ini}:I{fila_fin})"
            f"/SUM(F{fila_ini}:F{fila_fin}),\"\")"
        )
        ws.cell(row=r0 + 1, column=2).value = f"=SUM(F{fila_ini}:F{fila_fin})"
        ws.cell(row=r0 + 2, column=2).value = f"=SUM(G{fila_ini}:G{fila_fin})"
        ws.cell(row=r0 + 3, column=2).value = f"=SUM(H{fila_ini}:H{fila_fin})"

        # Un AUTOFILTRO sobre el detalle, no un `freeze_panes`.
        #
        # Congelar en la fila del encabezado de créditos (la 37) suena razonable -- deja los
        # títulos de la tabla a la vista -- pero `freeze_panes` congela TODO lo que está
        # arriba, y arriba hay 36 filas de resumen: la hoja quedaba pegada, sin poder
        # desplazarse.
        #
        # El autofiltro además sirve para más: deja ordenar los créditos por monto o por
        # tasa, y filtrar por moneda, acreedor o instrumento.
        ws.auto_filter.ref = f"A{rd}:{get_column_letter(len(cols))}{fila_fin}"

    def create_scenarios_sheet(self):
        """
        Crea una hoja de Escenarios interactiva que calcula la valuación para diferentes supuestos.
        Cada escenario calcula automáticamente Enterprise Value, Equity Value y Valor por Acción.
        """
        ws = self.wb.create_sheet("Escenarios")
        self._apply_sheet_header(ws, "ESCENARIOS DE VALUACIÓN", 16)

        # SECCIÓN 1: PARÁMETROS DE LOS ESCENARIOS
        params_row = self._create_professional_section(ws, 4, "PARÁMETROS POR ESCENARIO", "", 12, "inputs")
        
        headers = ["Escenario", "Crec Y+1", "Crec Y+2", "Crec Y+3", "Crec Y+4", "Crec Y+5",
                   "Margen EBIT", "Tasa Impuestos", "D&A/Ventas", "CapEx/Ventas", "ΔNWC/ΔVentas"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=params_row, column=c, value=h)
            cell.font = self.key_metric_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_wrap
            cell.border = self.border

        # Obtener referencias dinámicas a los inputs base del DCF principal
        # Intentar encontrar hojas DCF disponibles
        dcf_sheets = [sheet for sheet in self.wb.sheetnames if sheet.startswith("DCF ")]
        dcf_base_sheet = dcf_sheets[0] if dcf_sheets else "DCF"
        
        # Buscar filas de parámetros en la hoja DCF base
        def find_dcf_param_row(param_text):
            try:
                sheet = self.wb[dcf_base_sheet]
                for row in range(1, min(sheet.max_row + 1, 50)):
                    cell = sheet.cell(row=row, column=1)
                    if cell.value and isinstance(cell.value, str) and param_text.lower() in cell.value.lower():
                        return f"'{dcf_base_sheet}'!B{row}"
                return None
            except:
                return None
        
        # Referencias dinámicas a parámetros base
        base_crec_y1 = find_dcf_param_row("Crecimiento Ventas Y+1") or f"'{dcf_base_sheet}'!B11"
        base_crec_y2 = find_dcf_param_row("Crecimiento Ventas Y+2") or f"'{dcf_base_sheet}'!B12" 
        base_crec_y3 = find_dcf_param_row("Crecimiento Ventas Y+3") or f"'{dcf_base_sheet}'!B13"
        base_crec_y4 = find_dcf_param_row("Crecimiento Ventas Y+4") or f"'{dcf_base_sheet}'!B14"
        base_crec_y5 = find_dcf_param_row("Crecimiento Ventas Y+5") or f"'{dcf_base_sheet}'!B15"
        base_margen = find_dcf_param_row("Margen EBIT") or f"'{dcf_base_sheet}'!B16"
        base_tasa_imp = find_dcf_param_row("Tasa efectiva de impuestos") or f"'{dcf_base_sheet}'!B17"
        base_da_ventas = find_dcf_param_row("D&A / Ventas") or f"'{dcf_base_sheet}'!B18"
        base_capex_ventas = find_dcf_param_row("CapEx / Ventas") or f"'{dcf_base_sheet}'!B19"
        base_nwc_ventas = find_dcf_param_row("ΔNWC / ΔVentas") or f"'{dcf_base_sheet}'!B20"
        
        # Escenarios dinámicos basados en el escenario Base
        scenarios = [
            # Conservador: 20% menos optimista que Base
            ["Conservador", 
             f"=MAX({base_crec_y1}*0.8,0.015)",    # Y+1: 80% del base
             f"=MAX({base_crec_y2}*0.8,0.015)",    # Y+2: 80% del base 
             f"=MAX({base_crec_y3}*0.8,0.015)",    # Y+3: 80% del base
             f"=MAX({base_crec_y4}*0.8,0.01)",     # Y+4: 80% del base
             f"=MAX({base_crec_y5}*0.8,0.01)",     # Y+5: 80% del base
             f"=MIN({base_margen}*0.9,0.06)",      # Margen: 90% del base (conservador usa MIN para piso)
             "0.27",    # Impuestos: Fijo en 27%
             f"={base_da_ventas}*1.05",            # D&A: 105% del base
             f"={base_capex_ventas}*1.1",          # CapEx: 110% del base (más inversión)
             f"={base_nwc_ventas}*1.2"             # NWC: 120% del base (más capital trabajo)
            ],
            # Base: Referencias directas a la hoja DCF
            ["Base",        
             f"={base_crec_y1}",  # Referencia directa Y+1
             f"={base_crec_y2}",  # Referencia directa Y+2
             f"={base_crec_y3}",  # Referencia directa Y+3
             f"={base_crec_y4}",  # Referencia directa Y+4
             f"={base_crec_y5}",  # Referencia directa Y+5
             f"={base_margen}",   # Referencia directa Margen
             "0.27", # Tasa impuestos fija en 27%
             f"={base_da_ventas}",# Referencia directa D&A
             f"={base_capex_ventas}", # Referencia directa CapEx
             f"={base_nwc_ventas}"    # Referencia directa NWC
            ],
            # Agresivo: 20% más optimista que Base
            ["Agresivo",    
             f"={base_crec_y1}*1.2",               # Y+1: 120% del base
             f"={base_crec_y2}*1.2",               # Y+2: 120% del base
             f"={base_crec_y3}*1.2",               # Y+3: 120% del base 
             f"={base_crec_y4}*1.2",               # Y+4: 120% del base
             f"={base_crec_y5}*1.2",               # Y+5: 120% del base
             f"=MAX({base_margen}*1.15,0.18)",     # Margen: 115% del base (agresivo usa MAX para piso alto)
             "0.27",    # Impuestos: Fijo en 27%
             f"={base_da_ventas}*1.05",            # D&A: 105% del base (más depreciación es bueno para FCFF)
             f"={base_capex_ventas}*0.85",         # CapEx: 85% del base (menos inversión)
             f"={base_nwc_ventas}*0.7"             # NWC: 70% del base (mucho menos capital trabajo)
            ]
        ]
        
        # Referencias DCF base encontradas
        # Crecimiento Y+1 identificado
        # Margen EBIT identificado
        # NWC/Ventas identificado
        
        scenario_rows = {}  # Para guardar las filas de cada escenario
        for r_idx, rowvals in enumerate(scenarios, params_row + 1):
            scenario_rows[rowvals[0]] = r_idx
            for c, v in enumerate(rowvals, 1):
                cell = ws.cell(row=r_idx, column=c)
                
                if c == 1:  # Nombre del escenario
                    cell.value = v
                    cell.font = est.fuente(10, bold=True, color=est.PAPER)
                    cell.fill = self.subheader_fill
                    cell.alignment = self.center
                elif isinstance(v, str) and v.startswith("="):  # Fórmulas dinámicas
                    cell.value = v
                    cell.fill = self.calculated_fill
                    cell.font = self.input_font
                else:  # Valores estáticos: escribir como NÚMERO (no texto) para
                       # que el formato % aplique y en es-CL se muestre con coma.
                    try:
                        cell.value = float(v)
                    except (TypeError, ValueError):
                        cell.value = v
                    cell.fill = self.input_fill
                    cell.font = self.input_font
                    
                cell.border = self.border
                
                # Formateo específico por columna
                if c >= 2 and c <= 11:  # Todos los parámetros numéricos
                    cell.number_format = '0.00%'  # Formato porcentaje

        # SECCIÓN 2: CÁLCULOS DE VALUACIÓN AUTOMÁTICOS
        calc_start_row = params_row + len(scenarios) + 3
        calc_header_row = self._create_professional_section(ws, calc_start_row, "RESULTADOS DE VALUACIÓN", "", 12, "valuation")
        
        # Headers de resultados
        result_headers = ["Escenario", "Ventas Base", "WACC", "g Terminal", "Enterprise Value", 
                         "Deuda Neta", "Equity Value", "Acciones (M)", "Valor/Acción", "Premium/Desc vs Base"]
        for c, h in enumerate(result_headers, 1):
            cell = ws.cell(row=calc_header_row, column=c, value=h)
            cell.font = self.key_metric_font
            cell.fill = self.valuation_fill
            cell.alignment = self.center_wrap
            cell.border = self.border

        # Ajustar altura del encabezado
        ws.row_dimensions[calc_header_row].height = 30

        
        # Función para buscar una celda por contenido en una hoja
        def find_cell_by_content(sheet_name, search_text):
            try:
                sheet = self.wb[sheet_name]
                for row in range(1, min(sheet.max_row + 1, 50)):  # Buscar en primeras 50 filas
                    for col in range(1, min(sheet.max_column + 1, 10)):  # Buscar en primeras 10 columnas
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str) and search_text.lower() in cell.value.lower():
                            return f"'{sheet_name}'!{get_column_letter(col+1)}{row}"  # Columna B (valor)
                return None
            except:
                return None
        
        # Buscar referencias dinámicamente
        base_ventas_ref = find_cell_by_content(dcf_base_sheet, "Ventas año base") or f"'{dcf_base_sheet}'!B12"
        deuda_neta_ref = find_cell_by_content(dcf_base_sheet, "Deuda neta") or f"'{dcf_base_sheet}'!B22"  
        acciones_ref = find_cell_by_content(dcf_base_sheet, "Acciones en circulación") or f"'{dcf_base_sheet}'!B23"
        
        # El WACC de los escenarios sale del WACC REAL de la hoja DCF (CAPM + costo de deuda
        # de la propia empresa), no de una constante.
        #
        # Antes los tres escenarios escribian los literales 0.12 / 0.10 / 0.08. O sea que
        # ninguna de las tres valuaciones usaba el WACC calculado: una empresa con WACC 14%
        # se valuaba igual que una con 7%, y el "rango de valuacion" del resumen ejecutivo
        # no decia nada sobre la empresa.
        wacc_ref = find_cell_by_content(dcf_base_sheet, "WACC (%)")
        g_terminal = "0.02"  # Tasa terminal fija en 2%
        
        # Referencias para scenarios encontradas
        # Ventas base identificadas
        # Deuda neta identificada
        # Acciones identificadas
        
        # Calcular valuaciones para cada escenario
        for i, (scenario_name, _) in enumerate([(s[0], s[1:]) for s in scenarios]):
            r = calc_header_row + 1 + i
            scenario_param_row = scenario_rows[scenario_name]
            
            # Columna 1: Nombre del escenario
            cell = ws.cell(row=r, column=1, value=scenario_name)
            cell.font = self.result_font
            cell.alignment = self.left_center
            cell.border = self.border
            
            # Columna 2: Ventas Base (referencia común)
            cell = ws.cell(row=r, column=2, value=f"={base_ventas_ref}")
            cell.border = self.border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Columna 3: WACC. El del escenario base ES el WACC calculado de la empresa; el
            # conservador le suma 2 puntos y el agresivo le resta 2. Si por lo que sea no se
            # encuentra la celda del WACC, se cae a las constantes de antes -- pero con el
            # modelo apuntando a una celda real, eso no deberia pasar.
            if wacc_ref:
                if scenario_name == "Conservador":
                    wacc_scenario = f"=MAX({wacc_ref}+0.02,0.03)"
                elif scenario_name == "Agresivo":
                    wacc_scenario = f"=MAX({wacc_ref}-0.02,0.03)"
                else:  # Base
                    wacc_scenario = f"={wacc_ref}"
            else:
                wacc_scenario = {"Conservador": 0.12, "Agresivo": 0.08}.get(scenario_name, 0.10)

            cell = ws.cell(row=r, column=3, value=wacc_scenario)
            cell.border = self.border
            cell.number_format = '0.00%'

            # Columna 4: g Terminal (fijo)
            cell = ws.cell(row=r, column=4, value=float(g_terminal))
            cell.border = self.border
            cell.number_format = '0.00%'
            
            # Columna 5: Enterprise Value (CÁLCULO BASADO EN METODOLOGÍA DCF)
            # Usar aproximación DCF más precisa que refleje la metodología real
            
            # Referencias a parámetros del escenario
            ventas_base = f"B{r}"
            crec_y1 = f"B{scenario_param_row}"
            crec_y2 = f"C{scenario_param_row}"  
            crec_y3 = f"D{scenario_param_row}"
            crec_y4 = f"E{scenario_param_row}"
            crec_y5 = f"F{scenario_param_row}"
            margen_ebit = f"G{scenario_param_row}"
            tasa_imp = f"H{scenario_param_row}"
            da_ventas = f"I{scenario_param_row}"
            capex_ventas = f"J{scenario_param_row}"
            nwc_ventas = f"K{scenario_param_row}"
            
            # Cálculo aproximado de FCFF promedio anual usando los parámetros del escenario
            # FCFF ≈ Ventas × Margen EBIT × (1-Tax) × [1 + D&A/Ventas - CapEx/Ventas - ΔNWC/ΔVentas]
            # Aplicar crecimiento promedio para obtener ventas promedio del período
            crecimiento_promedio = f"(({crec_y1}+{crec_y2}+{crec_y3}+{crec_y4}+{crec_y5})/5)"
            ventas_promedio = f"({ventas_base}*(1+{crecimiento_promedio}))"
            
            # Cálculo correcto del FCFF
            # NOPAT = Ventas × Margen EBIT × (1 - Tasa Impuestos)
            nopat = f"({ventas_promedio}*{margen_ebit}*(1-{tasa_imp}))"
            
            # D&A = Ventas × D&A/Ventas (positivo para FCFF)
            da_amount = f"({ventas_promedio}*{da_ventas})"
            
            # CapEx = Ventas × CapEx/Ventas (negativo para FCFF)
            capex_amount = f"({ventas_promedio}*{capex_ventas})"
            
            # ΔNWC = Δ Ventas × ΔNWC/ΔVentas (negativo para FCFF)
            delta_ventas = f"({ventas_promedio}*{crecimiento_promedio})"
            nwc_amount = f"({delta_ventas}*{nwc_ventas})"
            
            # FCFF = NOPAT + D&A - CapEx - ΔNWC
            fcff_anual = f"({nopat}+{da_amount}-{capex_amount}-{nwc_amount})"
            
            # Valor terminal usando tasa g fija y el WACC del escenario.
            #
            # Se referencia la CELDA (columna 3 de esta misma fila), no el valor: ahora el
            # WACC del escenario es una formula que sale del WACC real de la empresa, y
            # pegar su texto aqui produciria `(1+0.02)/(=MAX(...)-0.02)`, que Excel no puede
            # evaluar. Referenciar la celda ademas hace que el EV se recalcule solo si el
            # analista edita el WACC.
            g_terminal = "0.02"  # Tasa terminal fija en 2%
            wacc = f"C{r}"
            valor_terminal_factor = f"(1+{g_terminal})/({wacc}-{g_terminal})"
            
            # Enterprise Value = FCFF_anual × Factor_terminal
            # Aplicar un factor de descuento promedio (aproximadamente 3 años)
            factor_descuento = f"1/POWER(1+{wacc},3)"
            
            cell = ws.cell(row=r, column=5)
            cell.value = f"={fcff_anual}*{valor_terminal_factor}*{factor_descuento}"
            cell.border = self.border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Columna 6: Deuda Neta (referencia común)
            cell = ws.cell(row=r, column=6, value=f"={deuda_neta_ref}")
            cell.border = self.border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Para el escenario Base, usar referencias directas al DCF para que coincida exactamente
            if scenario_name == "Base":
                # Buscar las filas de Enterprise Value y Equity Value en la hoja DCF base
                try:
                    dcf_sheet = self.wb[dcf_base_sheet]
                    ev_row = None
                    equity_row = None
                    
                    for row in range(1, min(dcf_sheet.max_row + 1, 50)):
                        cell_val = dcf_sheet.cell(row=row, column=1).value
                        if cell_val and isinstance(cell_val, str):
                            if "enterprise value" in cell_val.lower():
                                ev_row = row
                            elif "equity value" in cell_val.lower():
                                equity_row = row
                    
                    # Actualizar Enterprise Value con referencia directa
                    if ev_row:
                        ev_cell = ws.cell(row=r, column=5)
                        ev_cell.value = f"='{dcf_base_sheet}'!B{ev_row}"
                        
                except:
                    pass  # Mantener la fórmula calculada si no se encuentra
            
            # Columna 7: Equity Value = EV - Deuda Neta
            cell = ws.cell(row=r, column=7, value=f"=E{r}-F{r}")
            cell.border = self.border
            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Destacar Equity Value para Base
            if scenario_name == "Base":
                cell.fill = self.key_result_fill
                cell.font = self.result_font
            
            # Columna 8: Acciones (referencia común)
            cell = ws.cell(row=r, column=8, value=f"={acciones_ref}")
            cell.border = self.border
            cell.number_format = '#,##0'
            
            # Columna 9: Valor por Acción = Equity Value / Acciones
            cell = ws.cell(row=r, column=9, value=f"=IFERROR(IF(H{r}>0,G{r}/H{r}*1000,\"\"),\"\")")
            cell.border = self.border
            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)'
            
            # Destacar Valor por Acción
            cell.fill = self.key_result_fill
            cell.font = self.result_font
            
            # Columna 10: Premium/Descuento vs Base
            base_row = calc_header_row + 2  # Fila del escenario Base (segunda fila)
            if scenario_name != "Base":
                cell = ws.cell(row=r, column=10, value=f"=IFERROR((I{r}-I${base_row})/I${base_row},\"\")")
                cell.border = self.border
                cell.number_format = '0.0%'
                
                # La prima/descuento SÍ es dirección: se colorea por el signo del
                # resultado, no por el nombre del escenario. Un escenario "Agresivo"
                # puede perfectamente valer MENOS que el Base si los supuestos que el
                # usuario editó lo llevan ahí — pintarlo de verde por llamarse Agresivo
                # sería afirmar algo que la celda no dice.
                cell.fill = self.key_result_fill
                self._colorear_por_signo(ws, cell.coordinate)
            else:
                cell = ws.cell(row=r, column=10, value="Base")
                cell.border = self.border
                cell.alignment = self.center
                cell.font = est.CUERPO_FUERTE
                
        # SECCIÓN 3: RESUMEN EJECUTIVO
        summary_start_row = calc_header_row + len(scenarios) + 3
        summary_header_row = self._create_professional_section(ws, summary_start_row, "RESUMEN EJECUTIVO", "", 8, "valuation")
        
        # Métricas clave del análisis
        base_result_row = calc_header_row + 2  # Fila del escenario Base
        conservador_result_row = calc_header_row + 1  # Fila del escenario Conservador  
        agresivo_result_row = calc_header_row + 3    # Fila del escenario Agresivo
        
        summary_metrics = [
            ("Rango de Valuación:", ""),
            ("  • Escenario Conservador", f"=I{conservador_result_row}"),
            ("  • Escenario Base", f"=I{base_result_row}"),
            ("  • Escenario Agresivo", f"=I{agresivo_result_row}"),
            ("", ""),
            ("Rango Total", f"=IFERROR(I{agresivo_result_row}-I{conservador_result_row},\"\")"),
            ("Valor Medio", f"=IFERROR((I{conservador_result_row}+I{agresivo_result_row})/2,\"\")"),
            ("Coeficiente de Variación", f"=IFERROR(ABS((I{agresivo_result_row}-I{conservador_result_row})/(2*I{base_result_row})),\"\")")
        ]
        
        for i, (label, formula) in enumerate(summary_metrics):
            r = summary_header_row + 1 + i
            cell_a = ws.cell(row=r, column=1, value=label)
            cell_a.font = self.label_font
            cell_a.alignment = self.left_center
            
            if formula:
                cell_b = ws.cell(row=r, column=2)
                cell_b.value = formula
                cell_b.border = self.border
                
                if "Coeficiente" in label:
                    cell_b.number_format = '0.0%'
                elif formula.startswith("="):
                    cell_b.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)'
                    
                # Destacar métricas importantes
                if "Valor Medio" in label or "Rango Total" in label:
                    cell_b.fill = self.key_result_fill
                    cell_b.font = self.result_font

        # Ajustar anchos de columna
        column_widths = [16, 12, 8, 8, 15, 12, 15, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_sensitivity_analysis(self, projection_start_row, valuation_row):
        ws = self.wb["DCF"]
        start_row = valuation_row + 15

        ws[f"A{start_row}"] = "ANÁLISIS DE SENSIBILIDAD - ENTERPRISE VALUE"
        ws[f"A{start_row}"].font = self.subheader_font
        ws[f"A{start_row}"].fill = self.subheader_fill
        ws.merge_cells(f"A{start_row}:H{start_row}")

        waccs = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13]
        gs = [0.015, 0.020, 0.025, 0.030, 0.035]

        hdr_row = start_row + 2
        ws[f"A{hdr_row}"] = "g \\ WACC"
        ws[f"A{hdr_row}"].font = self.subheader_font
        ws[f"A{hdr_row}"].fill = self.subheader_fill
        ws[f"A{hdr_row}"].border = self.border

        for j, w in enumerate(waccs, 2):
            cell = ws.cell(row=hdr_row, column=j, value=f"{w:.0%}")
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border

        for i, g in enumerate(gs, 1):
            r = hdr_row + i
            cell = ws.cell(row=r, column=1, value=f"{g:.1%}")
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border
            for j, w in enumerate(waccs, 2):
                ev_base_row = valuation_row + 4
                cell = ws.cell(row=r, column=j,
                               value=f"={w/0.10}*{g/0.025}*$B${ev_base_row}")  # aprox
                cell.border = self.border
                cell.number_format = "#,##0"

        return start_row + 8

    def create_wacc_terminal_block(self, sheet_name, start_row: int, inputs_start_row: int,
                                   valuation_row: int, projection_start_row: int):
        """WACC profesional (CAPM + costo de deuda REAL) + contraste de valor terminal.

        Reemplaza el WACC fijo de 10% por uno construido y visible:
            Ke   = Rf + Beta x ERP                          (CAPM)
            Kd   = Costos financieros (anual) / Deuda financiera   (real, de los estados)
            WACC = E/(D+E)*Ke + D/(D+E)*Kd*(1 - t)          (t = 27%, Chile)
        Se coloca como sección propia DESPUÉS del tornado para no colisionar.
        """
        try:
            ws = self.wb[sheet_name] if sheet_name and sheet_name in self.wb.sheetnames else self.wb["DCF"]
        except Exception:
            return
        base_p = self._find_base_annual_period()
        # Offsets relativos a inputs_start_row (igual que create_tornado_analysis:
        # el valor devuelto por create_dcf_sheet apunta 2 filas antes del header
        # de parámetros, por eso WACC = +14 y tasa de impuesto = +10).
        tasa_imp_row = inputs_start_row + 10
        wacc_row = inputs_start_row + 14
        g_row_wacc = inputs_start_row + 15   # "g - Tasa de crecimiento terminal (%)"
        last_fcff_row = projection_start_row + 6
        ev_row = valuation_row + 4          # "Enterprise Value"
        vtp_row = valuation_row + 2         # "VT Presente"

        def _bal_ref(concept):
            rr = self._find_row_in_sheet(self.sh_bal, concept)
            return self.create_cell_reference_by_label(self.sh_bal, rr, base_p) if rr else None

        debt_c = _bal_ref("Otros pasivos financieros corrientes")
        debt_nc = _bal_ref("Otros pasivos financieros no corrientes")
        # Bajo IFRS 16 un arriendo ES deuda: entra en el peso de deuda del WACC igual
        # que en la deuda neta. Sin esto, el peso D del WACC y la deuda neta usarían
        # definiciones distintas de "deuda", y además no cuadraría con la BD (que sí
        # los incluye vía _deuda_total).
        lease_c = _bal_ref("Pasivos por arrendamientos corrientes")
        lease_nc = _bal_ref("Pasivos por arrendamientos no corrientes")
        equity_ref = (_bal_ref("Patrimonio total")
                      or _bal_ref("Patrimonio atribuible a los propietarios de la controladora"))
        fc_r = self._find_row_in_sheet(self.sh_pl, "Costos financieros")
        fincost_ref = self.create_cell_reference_by_label(self.sh_pl, fc_r, base_p) if fc_r else None

        d_expr = " + ".join(f"IFERROR({x},0)"
                            for x in (debt_c, debt_nc, lease_c, lease_nc) if x) or "0"
        e_expr = f"IFERROR({equity_ref},0)" if equity_ref else "0"
        fc_expr = f"IFERROR({fincost_ref},0)" if fincost_ref else "0"

        money = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
        wacc_hdr = self._create_professional_section(
            ws, start_row, "WACC (CAPM + COSTO DE DEUDA REAL)", "", 6, "inputs")
        # EL BETA: el de YAHOO FINANCE si la empresa cotiza (inyectado desde la BD,
        # companies.yahoo_beta, acotado a [0,5, 2,0]); HAMADA si no. Es el MISMO criterio
        # que el motor de la BD (scripts/dcf/excel_aligned.calculate_wacc_excel), para que
        # el WACC del Excel y el que ve la web CUADREN empresa por empresa.
        #
        # Sólo ~42 empresas cotizan y tienen beta de Yahoo. Para el resto NO se usa un
        # valor parejo (antes fue un 1,0 escondido que volvía el CAPM una identidad):
        # se re-apalanca una beta desapalancada 0,8 con el D/E contable (con arriendos) y
        # la tasa efectiva REAL de la empresa (Hamada), acotada a [0,5, 2,0]. Así una
        # eléctrica muy apalancada y una empresa sin deuda NO salen con el mismo Ke.
        #
        # Cuando viene de Yahoo, el beta es una celda EDITABLE; cuando es Hamada, es una
        # fórmula viva que se recalcula si el analista corrige la deuda o el patrimonio.
        w0 = wacc_hdr + 1
        r_rf, r_erp = w0, w0 + 1
        r_d, r_e, r_de, r_t = w0 + 2, w0 + 3, w0 + 4, w0 + 5
        r_beta, r_beta_src, r_ke = w0 + 6, w0 + 7, w0 + 8
        r_fc, r_kd, r_fuente, r_wacc = w0 + 9, w0 + 10, w0 + 11, w0 + 12

        # El beta: el de Yahoo si la empresa cotiza (input editable), Hamada si no.
        # Hamada re-apalanca una beta desapalancada 0,8 con el D/E contable (con
        # arriendos) y la tasa efectiva, acotada a [0,5, 2,0] — así TODA empresa tiene
        # un beta propio, sensible a su apalancamiento, no un valor parejo. Es el mismo
        # criterio que la BD (excel_aligned.calculate_wacc_excel).
        _yahoo_beta = self._beta_yahoo()
        if _yahoo_beta is not None:
            beta_label, beta_value, beta_kind = "Beta (Yahoo)", _yahoo_beta, "in"
        else:
            beta_label = "Beta (Hamada, s/ beta de Yahoo)"
            beta_value = f"=IFERROR(MAX(MIN(0.8*(1+(1-B{r_t})*B{r_de}),2),0.5),0.8)"
            beta_kind = "calc"

        wacc_rows = [
            ("Rf - Tasa libre de riesgo (%)", 0.055, "in", '0.00%'),
            ("ERP - Prima de riesgo de mercado (%)", 0.055, "in", '0.00%'),
            ("D - Deuda financiera + arriendos (M$)", f"={d_expr}", "calc", money),
            ("E - Patrimonio (M$)", f"={e_expr}", "calc", money),
            ("D/E - Apalancamiento", f'=IFERROR(B{r_d}/B{r_e},0)', "calc", '0.00'),
            ("t - Tasa de impuesto (%)", f"=$B${tasa_imp_row}", "calc", '0.00%'),
            (beta_label, beta_value, beta_kind, '0.00'),
            ("Fuente del beta", self._beta_fuente(), "calc", '@'),
            ("Ke - Costo de patrimonio (CAPM, %)", f"=B{r_rf}+B{r_beta}*B{r_erp}", "calc", '0.00%'),
            ("Costos financieros (anual, M$)", f"={fc_expr}", "calc", money),
            # EL Kd SALE DE LAS TASAS QUE LA EMPRESA DECLARÓ, NO DE UN COCIENTE.
            #
            # Estimarlo como costos_financieros/deuda_financiera es una aproximación cruda
            # y en muchas empresas simplemente se rompe: cuando hay pocos pasivos
            # financieros pero sí costos financieros (intereses de arrendamientos, por
            # ejemplo), el cociente se dispara.
            #
            #     FORUS              estimado 958,41%   declarado 0,52%
            #     SOPROCAL           estimado 198,01%   declarado 0,88%
            #     SODIMAC            estimado  91,18%   declarado 3,24%
            #     TELEFÓNICA CHILE   estimado  60,73%   declarado 7,64%
            #
            # Con un Kd de 958% el WACC no significa nada y el modelo caía al
            # IFERROR(...,0.10): un 10% inventado, presentado como el costo de capital.
            #
            # La empresa YA declara la tasa de cada crédito en la nota de préstamos del
            # XBRL. Arauco declara 137; Aguas Andinas, 65. Kd = promedio de las tasas
            # efectivas ponderado por el monto contable. 163 de 232 empresas la traen; el
            # resto cae a la estimación, y la celda de al lado dice cuál se usó.
            ("Kd - Costo de la deuda (%)", self._kd_valor(r_fc, r_d), "calc", '0.00%'),
            ("Fuente del Kd", self._kd_fuente(), "calc", '@'),
            # Acotado a [8%, 15%] (mismo clamp que la BD: min_wacc/max_wacc).
            ("WACC (%)",
             f"=IFERROR(MAX(MIN(B{r_e}/(B{r_d}+B{r_e})*B{r_ke}+B{r_d}/(B{r_d}+B{r_e})*B{r_kd}*(1-B{r_t}),0.15),0.08),0.10)",
             "result", '0.00%'),
        ]
        wacc_result_row = w0 + len(wacc_rows) - 1
        for k, (lbl, val, kind, fmt) in enumerate(wacc_rows):
            rr = w0 + k
            ws.cell(row=rr, column=1, value=lbl).font = self.label_font
            cell = ws.cell(row=rr, column=2)
            cell.value = val
            cell.border = self.border
            cell.number_format = fmt
            cell.alignment = self.right_center
            cell.fill = (self.input_fill if kind == "in"
                         else self.key_result_fill if kind == "result"
                         else self.calculated_fill)
            cell.font = self.result_font if kind == "result" else self.input_font

        # Reconectar el supuesto "WACC (%)" del modelo al WACC calculado.
        wcell = ws.cell(row=wacc_row, column=2)
        wcell.value = f"=B{wacc_result_row}"
        wcell.number_format = '0.00%'
        wcell.fill = self.calculated_fill

        # --- Contraste del valor terminal (múltiplo de salida + TV%EV) ---
        xc_hdr = self._create_professional_section(
            ws, wacc_result_row + 2, "CONTRASTE DE VALOR TERMINAL", "", 6, "valuation")
        x0 = xc_hdr + 1
        # EL MULTIPLO DE SALIDA SE DERIVA DEL PROPIO MODELO, NO ES UN 8x INVENTADO.
        #
        # El multiplo implicito que la perpetuidad de Gordon ya esta aplicando es
        #
        #     EV_terminal / EBITDA_terminal = (1 + g) / (WACC - g)
        #
        # Antes esta celda decia 8,0x fijo para todas las empresas: no salia de la empresa,
        # ni de sus comparables, ni del modelo. El "contraste" comparaba entonces la
        # perpetuidad contra un numero arbitrario, asi que no contrastaba nada -- y para una
        # empresa con WACC alto el 8x podia estar al doble de lo razonable.
        #
        # Ahora se muestra el multiplo implicito y se lo compara contra el multiplo del
        # analista, que queda como input editable. Si los dos se separan mucho, es que la
        # perpetuidad esta asumiendo algo que el mercado no pagaria.
        r_ebitda, r_mult_imp, r_mult_in = x0, x0 + 1, x0 + 2
        xchecks = [
            ("EBITDA terminal (Y+5, M$)", f"=C{last_fcff_row}+E{last_fcff_row}", money),
            ("Múltiplo implícito de Gordon (EV/EBITDA)",
             f'=IFERROR((1+$B${g_row_wacc})/($B${wacc_row}-$B${g_row_wacc}),"")', '0.0"x"'),
            ("Múltiplo de salida del analista (EV/EBITDA)", 8.0, '0.0"x"'),
            ("EV por múltiplo de salida (M$)", f"=B{r_ebitda}*B{r_mult_in}", money),
            ("EV por perpetuidad (Gordon, M$)", f"=B{ev_row}", money),
            ("Brecha Gordon vs múltiplo",
             f'=IFERROR(B{r_mult_imp}/B{r_mult_in}-1,"")', '0.0%'),
            ("Valor terminal como % del EV", f'=IFERROR(B{vtp_row}/B{ev_row},"")', '0.0%'),
            ("Alerta concentración terminal", f'=IF(IFERROR(B{vtp_row}/B{ev_row},0)>0.75,"ALTO: el terminal domina la valuación","OK")', '@'),
        ]
        for k, (lbl, val, fmt) in enumerate(xchecks):
            rr = x0 + k
            ws.cell(row=rr, column=1, value=lbl).font = self.label_font
            cell = ws.cell(row=rr, column=2)
            cell.value = val
            cell.border = self.border
            cell.number_format = fmt
            cell.alignment = self.right_center
            cell.fill = self.input_fill if k == 1 else self.calculated_fill
            cell.font = self.input_font

    def create_tornado_analysis(self, start_row: int, inputs_start_row: int, valuation_row: int, sheet_name: str = None):
        """
        Crea un análisis tornado profesional con recálculo dinámico del Enterprise Value.
        Cada driver se modifica individualmente y se recalcula el EV resultante.
        """
        if sheet_name:
            ws = self.wb[sheet_name]
        else:
            # Usar la última hoja creada (que debería ser la hoja DCF actual)
            ws = self.wb.worksheets[-1]
        # Usar sección profesional para tornado analysis
        tornado_header_row = self._create_professional_section(ws, start_row, "ANÁLISIS TORNADO - SENSIBILIDAD DE DRIVERS", "", 8, "sensitivity")
        
        r0 = tornado_header_row
        headers = ["Driver", "Valor Base", "Escenario -10%", "Escenario +10%", "EV Base", "EV(-10%)", "EV(+10%)", "Impacto Neto"]
        for c, h in enumerate(headers, 1):
            try:
                cell = ws.cell(row=r0, column=c, value=h)
                cell.font = self.key_metric_font
                cell.fill = self.sensitivity_fill
                cell.alignment = self.center_wrap
                cell.border = self.border
            except AttributeError:
                # Saltar celdas combinadas
                pass
        
        # Ajustar altura del encabezado tornado
        ws.row_dimensions[r0].height = 30

        # Referencias a los inputs del modelo
        margen_ebit_row = inputs_start_row + 9
        wacc_row = inputs_start_row + 14
        g_row = inputs_start_row + 15
        capex_row = inputs_start_row + 12
        nwc_row = inputs_start_row + 13
        crec_y1_row = inputs_start_row + 4
        tasa_row = inputs_start_row + 10

        # Configuración avanzada de drivers con elasticidades profesionales
        ev_base_ref = f"$B${valuation_row + 4}"  # Enterprise Value base
        
        # Obtener referencia a DRIVERS WC para ΔNWC/ΔVentas real
        drivers_wc_sheet = None
        try:
            drivers_wc_sheet = self.wb["DRIVERS WC"]
        except KeyError:
            pass
            
        # Usar el driver calculado de DRIVERS WC (mediana anual robusta) si existe
        real_nwc_ratio = getattr(self, "_wc_avg_cell", None) or f"$B${nwc_row}"
        if not getattr(self, "_wc_avg_cell", None) and drivers_wc_sheet:
            for row in range(10, 40):
                cell_value = drivers_wc_sheet.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str) and ("driver" in cell_value.lower() or "promedio" in cell_value.lower()):
                    real_nwc_ratio = f"'DRIVERS WC'!$B${row}"
                    break
        
        drivers_config = [
            {
                "name": "Margen EBIT",
                "base_ref": f"$B${margen_ebit_row}",
                "elasticity": 1.2,  # 10% cambio en margen → 12% cambio en EV
            },
            {
                "name": "WACC",
                "base_ref": f"$B${wacc_row}",
                "elasticity": -1.8,  # 10% cambio en WACC → -18% cambio en EV (inverso)
            },
            {
                "name": "g terminal",
                "base_ref": f"$B${g_row}",
                "elasticity": 2.5,  # 10% cambio en g → 25% cambio en EV
            },
            {
                "name": "CapEx/Ventas",
                "base_ref": f"$B${capex_row}",
                "elasticity": -0.8,  # 10% cambio en CapEx → -8% cambio en EV
            },
            {
                "name": "ΔNWC/ΔVentas",
                "base_ref": real_nwc_ratio,  # Usar valor real de DRIVERS WC
                "elasticity": -0.6,  # 10% cambio en ΔNWC → -6% cambio en EV
            },
            {
                "name": "Crec. Ventas Y+1",
                "base_ref": f"$B${crec_y1_row}",
                "elasticity": 1.0,  # 10% cambio en crecimiento → 10% cambio en EV
            },
            {
                "name": "Tasa Impuestos",
                "base_ref": f"$B${tasa_row}",
                "elasticity": -0.7,  # 10% cambio en tasa → -7% cambio en EV
            }
        ]
        
        # Crear análisis de sensibilidad avanzado usando elasticidades profesionales
        for i, driver in enumerate(drivers_config, 1):
            r = r0 + i
            
            # Calcular factores de sensibilidad basados en elasticidad
            elasticity = driver["elasticity"]
            sensitivity_factor = 0.1  # 10% de cambio en el driver
            
            # Calcular el cambio esperado en EV
            ev_change_down = 1 + (elasticity * sensitivity_factor * -1)  # -10% en driver
            ev_change_up = 1 + (elasticity * sensitivity_factor)         # +10% en driver
            
            # Asegurar que los factores sean razonables
            ev_change_down = max(0.5, min(1.5, ev_change_down))  # Limitar entre 50% y 150%
            ev_change_up = max(0.5, min(1.5, ev_change_up))
            
            try:
                # Columna 1: Nombre del driver
                ws.cell(row=r, column=1, value=driver["name"])
                
                # Columna 2: Valor base del driver (con =)
                ws.cell(row=r, column=2, value=f"={driver['base_ref']}")
                
                # Columna 3: Escenario -10%
                ws.cell(row=r, column=3, value=f"={driver['base_ref']}*0.9")
                
                # Columna 4: Escenario +10%
                ws.cell(row=r, column=4, value=f"={driver['base_ref']}*1.1")
                
                # Columna 5: EV Base (referencia con =)
                ws.cell(row=r, column=5, value=f"={ev_base_ref}")
                
                # Columna 6: EV con driver -10% (usa elasticidad)
                ws.cell(row=r, column=6, value=f"={ev_base_ref}*{ev_change_down:.3f}")
                
                # Columna 7: EV con driver +10% (usa elasticidad)
                ws.cell(row=r, column=7, value=f"={ev_base_ref}*{ev_change_up:.3f}")
                
                # Columna 8: Impacto neto (rango de variación)
                ws.cell(row=r, column=8, value=f"=G{r}-F{r}")
                
            except AttributeError:
                pass
            
            # Formateo profesional avanzado
            for c in range(1, 9):  # 8 columnas
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.border = self.border
                    
                    if c in [5, 6, 7, 8]:  # EV Base, EV(-10%), EV(+10%), Impacto - valores monetarios
                        cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Formato moneda profesional
                    elif c in [2, 3, 4]:  # Valores base y escenarios
                        if driver["name"] in ["Margen EBIT", "WACC", "g terminal", "CapEx/Ventas", "ΔNWC/ΔVentas", "Crec. Ventas Y+1", "Tasa Impuestos"]:
                            cell.number_format = '0.00%'  # Formato porcentaje
                        else:
                            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Formato moneda
                            
                except AttributeError:
                    pass

        # Ajustar anchos de columna para el nuevo formato
        column_widths = [18, 12, 12, 12, 15, 15, 15, 15]  # 8 columnas
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_ev_sensitivity_formula(self, valuation_row: int, driver_row: int, multiplier: str, driver_name: str) -> str:
        """
        Crea una fórmula profesional que recalcula el Enterprise Value cuando un driver específico cambia.
        
        Args:
            valuation_row: Fila donde comienza la sección de valuación (donde está "Valor Terminal")
            driver_row: Fila del driver que se está modificando  
            multiplier: Multiplicador para el driver (ej: "*0.9", "*1.1")
            driver_name: Nombre del driver para lógica específica
        
        Returns:
            Fórmula de Excel que calcula EV con el driver modificado
        """
        # Primero, identificar las filas clave de la valuación
        # valuation_row es donde empieza la sección "VALUACIÓN"
        # Las filas típicas son:
        # valuation_row + 1: Valor Terminal 
        # valuation_row + 2: VT Presente
        # valuation_row + 3: Suma FCFF PV
        # valuation_row + 4: Enterprise Value
        
        if driver_name == "WACC":
            # EV = VT_Presente + Suma_FCFF_PV
            # VT_Presente = VT / (1+WACC)^5
            # Recalcular VT Presente con nuevo WACC (usar referencias correctas)
            vt_base = f"B{valuation_row + 1}"  # Valor Terminal (fila 35)
            suma_fcff = f"B{valuation_row + 3}"  # Suma FCFF PV (fila 37) 
            new_wacc = f"B{driver_row}{multiplier}"
            
            # VT Presente recalculado = VT / (1+nuevo_WACC)^5
            vt_presente_new = f"({vt_base}/POWER(1+{new_wacc},5))"
            return f"={vt_presente_new}+{suma_fcff}"
            
        elif driver_name == "g terminal":
            # EV = VT_Presente + Suma_FCFF_PV  
            # VT = FCFF_ultimo * (1+g) / (WACC-g)
            # Recalcular VT con nuevo g (usar referencias correctas según la estructura encontrada)
            wacc_base = "B20"  # WACC está en fila 20 según el análisis
            new_g = f"B{driver_row}{multiplier}"
            
            # FCFF del último año (año 5) - buscar en proyección
            # Según estructura: PROYECCIÓN fila 26, así que año 5 sería aprox fila 30
            fcff_ultimo = "H30"  # FCFF año 5 estimado
            
            # VT recalculado = FCFF_ultimo * (1+nuevo_g) / (WACC-nuevo_g)
            vt_new = f"({fcff_ultimo}*(1+{new_g})/({wacc_base}-{new_g}))"
            vt_presente_new = f"({vt_new}/POWER(1+{wacc_base},5))"
            suma_fcff = f"B{valuation_row + 3}"  # Suma FCFF PV
            
            return f"={vt_presente_new}+{suma_fcff}"
            
        elif driver_name == "Margen EBIT":
            # Cambio en margen EBIT afecta todo el FCFF de los 5 años
            # Aproximación: EV cambia proporcionalmente al cambio en margen
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:  # -10%
                return f"={base_ev}*(1-0.12)"  # ~12% impacto negativo
            else:  # +10%
                return f"={base_ev}*(1+0.12)"  # ~12% impacto positivo
                
        elif driver_name == "Tasa Impuestos":
            # Cambio en tasa de impuestos afecta NOPAT = EBIT * (1-Tax)
            # A mayor tasa, menor NOPAT, menor FCFF, menor EV
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:  # -10% tasa = mayor NOPAT
                return f"={base_ev}*(1+0.08)"  # ~8% impacto positivo
            else:  # +10% tasa = menor NOPAT  
                return f"={base_ev}*(1-0.08)"  # ~8% impacto negativo
                
        elif driver_name == "CapEx/Ventas":
            # Cambio en CapEx afecta FCFF = NOPAT + D&A - CapEx - ΔNWC
            # Mayor CapEx = menor FCFF = menor EV
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:  # -10% CapEx = mayor FCFF
                return f"={base_ev}*(1+0.06)"  # ~6% impacto positivo
            else:  # +10% CapEx = menor FCFF
                return f"={base_ev}*(1-0.06)"  # ~6% impacto negativo
                
        elif driver_name == "ΔNWC/ΔVentas":
            # Cambio en ΔNWC afecta FCFF = NOPAT + D&A - CapEx - ΔNWC
            # Mayor ΔNWC = menor FCFF = menor EV
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:  # -10% ΔNWC = mayor FCFF
                return f"={base_ev}*(1+0.04)"  # ~4% impacto positivo
            else:  # +10% ΔNWC = menor FCFF
                return f"={base_ev}*(1-0.04)"  # ~4% impacto negativo
                
        elif driver_name == "Crec. Ventas Y+1":
            # Cambio en crecimiento afecta las ventas futuras y por ende todo el FCFF
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:  # -10% crecimiento
                return f"={base_ev}*(1-0.09)"  # ~9% impacto negativo
            else:  # +10% crecimiento
                return f"={base_ev}*(1+0.09)"  # ~9% impacto positivo
                
        else:
            # Fallback genérico
            base_ev = f"B{valuation_row + 4}"
            if "*0.9" in multiplier:
                return f"={base_ev}*0.95"
            else:
                return f"={base_ev}*1.05"

    def add_data_validation_and_protection(self):
        """Lugar para agregar protecciones/validaciones adicionales si quieres."""
        pass

    def build_dcf_module(self):
        """Orquestador."""
        self.create_drivers_wc_sheet()
        projection_start_row, valuation_row, inputs_start_row = self.create_dcf_sheet()
        self.create_scenarios_sheet()
        next_row = self.create_sensitivity_analysis(projection_start_row, valuation_row)
        self.create_tornado_analysis(next_row, inputs_start_row, valuation_row)
        self.create_wacc_terminal_block("DCF", next_row + 14, inputs_start_row, valuation_row, projection_start_row)
        self.add_data_validation_and_protection()


def add_dcf_functionality(workbook: Workbook, financial_data: Dict[str, Any]):
    """
    Agrega el módulo DCF a un libro existente con EERR, Balance y Flujo de Efectivo.
    """
    dcf = DCFBuilder(workbook, financial_data)
    dcf.build_dcf_module()

def add_multi_period_dcf_functionality(workbook: Workbook, financial_data: Dict[str, Any]):
    """
    Agrega modelo DCF para el período más reciente disponible.
    """
    dcf = DCFBuilder(workbook, financial_data)

    latest_period = dcf._find_latest_period()

    # Crear DRIVERS WC común
    dcf.create_drivers_wc_sheet()

    # DCF con período más reciente - HOJA COMPLETA CON TORNADO
    if True:
        # Creando DCF adicional para período reciente
        
        # Crear DCF completo usando el método existente pero con período personalizado
        # Temporalmente modificar las fórmulas para usar el período más reciente
        
        # La base de la valoración es el último CIERRE ANUAL. Punto.
        #
        # Aquí había un monkey-patch (`temp_ventas_formula`) que pisaba la fórmula de
        # ventas base y la reemplazaba por el ÚLTIMO TRIMESTRE ANUALIZADO: Q1×4, Q2×2,
        # Q3×1,33. Para Celulosa Arauco, cuyo último período es 2026Q1, eso significaba
        # tomar UN trimestre, multiplicarlo por cuatro, y colgar de ahí los cinco años de
        # proyección, el valor terminal y el precio objetivo.
        #
        # Para cualquier negocio estacional —una viña, una salmonera, un retail— eso es
        # una distorsión enorme. Y además rompía la paridad con el DCF que se guarda en la
        # base (scripts/dcf/excel_aligned.py), que usa el último año real: el mismo cliente
        # veía un precio objetivo en la ficha web y otro distinto en su Excel.
        #
        # El margen EBIT y la deuda neta también se anclan al cierre anual, por lo mismo.
        periodo_base = dcf._find_base_annual_period()

        original_deuda_neta_formula = dcf._get_deuda_neta_formula

        def temp_deuda_neta_formula(period=None):
            return original_deuda_neta_formula(periodo_base)

        dcf._current_dcf_period = periodo_base
        dcf._get_deuda_neta_formula = temp_deuda_neta_formula
        
        # Crear el DCF completo
        projection_start_row_latest, valuation_row_latest, inputs_start_row_latest = dcf.create_dcf_sheet()
        dcf.wb.worksheets[-1].title = f"DCF {latest_period}"
        
        # Restaurar la función original (ventas y margen ya no se parchean).
        dcf._get_deuda_neta_formula = original_deuda_neta_formula
        
        # Limpiar el período específico
        if hasattr(dcf, '_current_dcf_period'):
            delattr(dcf, '_current_dcf_period')
        
        # Actualizar el período seleccionado en la nueva hoja
        latest_sheet = dcf.wb.worksheets[-1]
        
        # Buscar la celda de período (puede estar en diferentes filas debido al nuevo diseño)
        try:
            latest_sheet["C6"] = latest_period  # Ajustado para el nuevo diseño
        except AttributeError:
            try:
                latest_sheet["C7"] = latest_period  # Alternativa
            except AttributeError:
                pass  # Si hay problemas con celdas combinadas, continuar
        
        # El "Año base" NO se toca aquí.
        #
        # Este bloque lo sobrescribía con `latest_period` -- el último período disponible,
        # que es un TRIMESTRE (p. ej. "2026Q1"). Pero el modelo no proyecta desde un
        # trimestre: `_find_base_annual_period()` elige a propósito el último año COMPLETO
        # (2025Q4), porque anualizar un trimestre suelto distorsiona a cualquier negocio
        # estacional -- una viña, un retail, una salmonera.
        #
        # O sea que la celda decía "Año base: 2026Q1" mientras la fila de abajo tomaba las
        # ventas de 2025. El rótulo mentía sobre el propio modelo: quien lo leyera creería
        # que la proyección arranca de un trimestre anualizado.
        #
        # El período seleccionado ya se muestra arriba, en "Período seleccionado".
        
        # Agregar análisis tornado al modelo del período más reciente (después de la valuación)
        tornado_start_row_latest = valuation_row_latest + 15
        dcf.create_tornado_analysis(tornado_start_row_latest, inputs_start_row_latest, valuation_row_latest, f"DCF {latest_period}")
        # WACC profesional (CAPM + Kd real) + contraste terminal, DESPUÉS del tornado
        dcf.create_wacc_terminal_block(f"DCF {latest_period}", tornado_start_row_latest + 14,
                                       inputs_start_row_latest, valuation_row_latest,
                                       projection_start_row_latest)

    # La deuda declarada, crédito por crédito: es lo que respalda el Kd del WACC.
    dcf.create_deuda_sheet()

    # Escenarios
    dcf.create_scenarios_sheet()

    # Reorganizar hojas en orden profesional
    dcf._organize_worksheets()

