#!/usr/bin/env python3
"""
Módulo para mejorar los datos de WATTS SA antes del procesamiento
Se integra automáticamente en el flujo de cmf_total_cli_v2.py
"""

import pandas as pd
import openpyxl
import numpy as np
import re
from pathlib import Path
from typing import Dict, Any, Optional

def is_watts_company(company_dir: Path) -> bool:
    """Verifica si es WATTS SA (76455830-8)"""
    return company_dir.name.startswith('76455830-8')

def clean_numeric_value(value, from_facts=False):
    """Limpia y convierte valores numéricos"""
    if pd.isna(value) or value is None or value == 'None':
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ['none', 'nan', 'false', 'true']:
        return None
    
    if from_facts:
        value_str = value_str.replace(',', '')
    
    try:
        return float(value_str)
    except:
        return None

def normalize_account_name(name):
    """Normaliza nombres de cuentas para matching"""
    if pd.isna(name) or not name:
        return ""
    
    name = str(name).strip()
    name = re.sub(r'\[.*?\]', '', name)
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'[,;:]', '', name)
    
    # Remover palabras comunes
    for word in [' totales', ' total', ' corriente', ' corrientes', ' no corriente', ' no corrientes']:
        name = name.replace(word, '')
    
    return name.strip().lower()

def find_watts_historical_excel() -> Optional[Path]:
    """Encuentra el Excel con datos históricos de WATTS"""
    # Buscar primero en la carpeta especial
    special_case_dir = Path('/home/unzzui/Documents/coding/CMF_extract/Products/Total/Special_Case')
    products_dir = Path('/home/unzzui/Documents/coding/CMF_extract/Products/Total')
    
    # Archivos específicos conocidos (en orden de preferencia)
    specific_files = [
        special_case_dir / 'estados_76455830-8_2025-2023_es.xlsx',
        special_case_dir / 'estados_76455830-8_2025-2023_es.backup.xlsx',
        special_case_dir / 'estados_76455830-8_2025-2023_es.backup (Copy).xlsx',
        products_dir / 'estados_76455830-8_2025-2021_es.xlsx'
    ]
    
    print(f"    🔍 Buscando Excel histórico de WATTS SA...")
    
    best_file = None
    best_score = 0
    
    # Primero verificar archivos específicos
    for excel_file in specific_files:
        if excel_file.exists():
            try:
                print(f"    📋 Analizando: {excel_file}")
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                if 'Balance General' in wb.sheetnames:
                    ws = wb['Balance General']
                    # Buscar columnas históricas y contar valores reales
                    score = 0
                    headers = []
                    historical_data_count = 0
                    
                    for col in range(2, min(20, ws.max_column + 1)):
                        cell_value = str(ws.cell(row=3, column=col).value or '')
                        headers.append(cell_value)
                        
                        # Puntos por columnas históricas
                        if any(year in cell_value for year in ['2021', '2022', '2023', '2020', '2019', '2018', '2017', '2016']):
                            score += 2
                        
                        # Bonus extra por contar valores reales en columnas históricas
                        if any(year in cell_value for year in ['2021', '2022']):
                            for row in range(4, min(50, ws.max_row + 1)):
                                if ws.cell(row=row, column=col).value:
                                    historical_data_count += 1
                    
                    # Bonus por archivos de Special_Case
                    if 'Special_Case' in str(excel_file):
                        score += 10
                    
                    # Bonus por cantidad de datos históricos
                    score += historical_data_count // 10
                    
                    print(f"      Encabezados: {headers[:10]}")
                    print(f"      Score histórico: {score}")
                    
                    if score > best_score:
                        best_score = score
                        best_file = excel_file
                wb.close()
            except Exception as e:
                print(f"      ❌ Error analizando {excel_file}: {e}")
                continue
    
    # Si no encontramos archivos específicos, buscar en directorios
    if best_file is None:
        search_dirs = [special_case_dir, products_dir]
        for search_dir in search_dirs:
            if not search_dir.exists():
                continue
                
            print(f"    🔍 Buscando en: {search_dir}")
            excel_files = list(search_dir.glob('estados_76455830-8_*_es.xlsx'))
            print(f"    📊 Archivos encontrados: {[f.name for f in excel_files]}")
            
            for excel_file in excel_files:
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    if 'Balance General' in wb.sheetnames:
                        ws = wb['Balance General']
                        score = 0
                        for col in range(2, min(15, ws.max_column + 1)):
                            cell_value = str(ws.cell(row=3, column=col).value or '')
                            if any(year in cell_value for year in ['2021', '2022', '2023']):
                                score += 1
                        
                        if score > best_score:
                            best_score = score
                            best_file = excel_file
                    wb.close()
                except Exception:
                    continue
    
    print(f"    ✓ Excel histórico seleccionado: {best_file.name if best_file else 'Ninguno'} (score: {best_score})")
    return best_file

def extract_excel_data(excel_path: Path) -> Dict[str, Any]:
    """Extrae datos del Excel existente"""
    print(f"  📊 Extrayendo datos históricos de: {excel_path.name}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    account_data = {}
    
    for sheet_name in ['Balance General', 'Estado de Resultados', 'Flujo Efectivo']:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # Encontrar fechas
        date_columns = {}
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=3, column=col).value
            if cell_value:
                date_columns[col] = str(cell_value)
        
        # Extraer cuentas
        for row in range(4, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            
            if not account_name or str(account_name).startswith('='):
                continue
            
            account_key = f"{sheet_name}|{account_name}"
            
            if account_key not in account_data:
                account_data[account_key] = {
                    'sheet': sheet_name,
                    'original_name': account_name,
                    'normalized': normalize_account_name(account_name),
                    'values': {}
                }
            
            for col, date in date_columns.items():
                value = ws.cell(row=row, column=col).value
                clean_value = clean_numeric_value(value)
                
                if clean_value is not None:
                    account_data[account_key]['values'][date] = clean_value
    
    wb.close()
    
    # Contar datos históricos (2021-2022)
    historical_count = 0
    for account in account_data.values():
        for date in account['values']:
            if '2021' in date or '2022' in date:
                historical_count += 1
    
    print(f"    ✓ Extraídas {len(account_data)} cuentas con {historical_count} valores históricos")
    
    return account_data

def match_and_populate_facts(facts_path: Path, excel_data: Dict[str, Any]) -> int:
    """Hace matching y puebla el facts CSV - INCLUYENDO NUEVAS COLUMNAS HISTÓRICAS"""
    print(f"  🔄 Poblando facts CSV con datos históricos...")
    
    facts_df = pd.read_csv(facts_path)
    original_columns = [col for col in facts_df.columns 
                       if col not in ['LabelKeyId', 'LabelKeyIdExt', 'SectionKey', 'Label', 'RoleCode']]
    
    # MAPEO COMPLETO incluyendo años históricos del Excel
    date_mapping = {
        '2025Q2': '2025-06-30', '2025Q1': '2025-03-31',
        '2024': '2024-12-31', '2024Q3': '2024-09-30', '2024Q2': '2024-06-30', '2024Q1': '2024-03-31',
        '2023': '2023-12-31', '2023Q3': '2023-09-30', '2023Q2': '2023-06-30', '2023Q1': '2023-03-31',
        '2022': '2022-12-31', '2021': '2021-12-31',
        # AÑOS HISTÓRICOS ADICIONALES
        '2020': '2020-12-31', '2019': '2019-12-31', '2018': '2018-12-31', 
        '2017': '2017-12-31', '2016': '2016-12-31'
    }
    
    # Identificar qué columnas históricas necesitamos agregar
    years_in_excel = set()
    for account_data in excel_data.values():
        for date in account_data['values'].keys():
            years_in_excel.add(date)
    
    historical_years = [year for year in ['2020', '2019', '2018', '2017', '2016'] if year in years_in_excel]
    
    if historical_years:
        print(f"    🆕 Agregando columnas históricas: {historical_years}")
        
        # AGREGAR NUEVAS COLUMNAS al facts DataFrame
        for year in historical_years:
            facts_column = date_mapping[year]
            if facts_column not in facts_df.columns:
                # Insertar la nueva columna después de 2021-12-31
                insert_pos = len(facts_df.columns)  # Al final
                if '2021-12-31' in facts_df.columns:
                    insert_pos = facts_df.columns.get_loc('2021-12-31') + 1
                
                facts_df.insert(insert_pos, facts_column, pd.NA)
                print(f"      ✅ Columna agregada: {facts_column}")
    
    # Actualizar la lista de columnas de fecha
    date_columns = [col for col in facts_df.columns 
                   if col not in ['LabelKeyId', 'LabelKeyIdExt', 'SectionKey', 'Label', 'RoleCode']]
    
    matches = {}
    
    # Matching por nombre y valor
    for excel_key, excel_account in excel_data.items():
        excel_norm = excel_account['normalized']
        
        for idx, facts_row in facts_df.iterrows():
            facts_label = facts_row['Label']
            facts_norm = normalize_account_name(facts_label)
            
            # Score de similitud
            score = 0
            if excel_norm and facts_norm:
                if excel_norm == facts_norm:
                    score = 100
                elif excel_norm in facts_norm or facts_norm in excel_norm:
                    score = 80
                else:
                    excel_words = set(excel_norm.split())
                    facts_words = set(facts_norm.split())
                    common = excel_words & facts_words
                    if len(common) >= 2:
                        score = 60 * len(common) / max(len(excel_words), len(facts_words))
            
            if score >= 60:
                # Verificar coincidencia de valor
                for excel_date, excel_value in excel_account['values'].items():
                    if excel_date in date_mapping:
                        facts_date = date_mapping[excel_date]
                        facts_value = clean_numeric_value(facts_row.get(facts_date, None), from_facts=True)
                        
                        if facts_value is not None and excel_value is not None:
                            excel_value_units = excel_value * 1000  # Convertir de miles a unidades
                            tolerance = max(abs(facts_value * 0.01), 1000)
                            
                            if abs(facts_value - excel_value_units) < tolerance:
                                if idx not in matches or matches[idx]['score'] < score + 50:
                                    matches[idx] = {
                                        'excel_key': excel_key,
                                        'excel_account': excel_account,
                                        'facts_label': facts_label,
                                        'score': score + 50
                                    }
                                break
    
    # Matching adicional solo por nombre para cuentas no mapeadas
    unmapped_excel = set(excel_data.keys())
    for match in matches.values():
        if match['excel_key'] in unmapped_excel:
            unmapped_excel.remove(match['excel_key'])
    
    for excel_key in unmapped_excel:
        excel_account = excel_data[excel_key]
        excel_norm = excel_account['normalized']
        
        if not excel_norm:
            continue
        
        for idx, facts_row in facts_df.iterrows():
            if idx in matches:
                continue
                
            facts_norm = normalize_account_name(facts_row['Label'])
            
            if excel_norm == facts_norm:
                matches[idx] = {
                    'excel_key': excel_key,
                    'excel_account': excel_account,
                    'facts_label': facts_row['Label'],
                    'score': 100
                }
                break
            elif len(excel_norm) > 10 and len(facts_norm) > 10:
                if excel_norm[:15] == facts_norm[:15]:
                    matches[idx] = {
                        'excel_key': excel_key,
                        'excel_account': excel_account,
                        'facts_label': facts_row['Label'],
                        'score': 70
                    }
                    break
    
    # Poblar datos
    populated_count = 0
    
    for facts_idx, match_info in matches.items():
        excel_account = match_info['excel_account']
        
        for excel_date, excel_value in excel_account['values'].items():
            if excel_date not in date_mapping:
                continue
                
            facts_date = date_mapping[excel_date]
            if facts_date not in date_columns:
                continue
            
            # Poblar TODOS los períodos históricos disponibles
            if excel_date not in ['2021', '2022', '2020', '2019', '2018', '2017', '2016']:
                continue
            
            current_value = facts_df.at[facts_idx, facts_date]
            
            if pd.isna(current_value) or current_value == '' or current_value is None:
                value_units = int(excel_value * 1000)
                formatted_value = f"{value_units:,}"
                facts_df.at[facts_idx, facts_date] = formatted_value
                populated_count += 1
    
    # Crear backup del facts original antes de sobrescribir
    backup_path = facts_path.with_suffix('.backup.csv')
    if not backup_path.exists():  # Solo crear backup si no existe
        import shutil
        shutil.copy2(facts_path, backup_path)
        print(f"    💾 Backup creado: {backup_path.name}")
    
    # SOBRESCRIBIR el facts original con los datos mejorados
    facts_df.to_csv(facts_path, index=False)
    print(f"    ✅ Facts original actualizado con datos históricos: {facts_path.name}")
    
    print(f"    ✓ Poblados {populated_count} valores históricos con {len(matches)} coincidencias")
    
    return populated_count

def enhance_watts_data(company_dir: Path) -> bool:
    """
    Función principal para mejorar datos de WATTS SA
    Retorna True si se realizaron mejoras
    """
    if not is_watts_company(company_dir):
        return False
    
    print(f"\n  🔧 MEJORA ESPECIAL DE DATOS: WATTS SA")
    
    # 1. Encontrar Excel con datos históricos
    historical_excel = find_watts_historical_excel()
    if not historical_excel:
        print(f"    ⚠️  No se encontró Excel histórico de WATTS")
        return False
    
    # 2. Encontrar facts CSV consolidado (usar el de end-period más reciente)
    import re as _re_oc
    consolidated_dirs = [p for p in company_dir.glob('out_consolidated_*') if p.is_dir()]
    if not consolidated_dirs:
        print(f"    ⚠️  No se encontró directorio consolidado")
        return False

    def _end_period(p: Path) -> str:
        m = _re_oc.search(r'(\d{6})(?=[/\\]|$)', p.name)
        return m.group(1) if m else ""
    consolidated_dirs.sort(key=_end_period, reverse=True)
    consolidated_dir = consolidated_dirs[0]
    facts_files = list(consolidated_dir.glob('facts_*_es.csv'))
    if not facts_files:
        print(f"    ⚠️  No se encontró facts CSV")
        return False
    
    facts_path = facts_files[0]
    
    # 3. Extraer datos del Excel
    excel_data = extract_excel_data(historical_excel)
    if not excel_data:
        print(f"    ⚠️  No se pudieron extraer datos del Excel")
        return False
    
    # 4. Poblar facts CSV
    populated = match_and_populate_facts(facts_path, excel_data)
    
    if populated > 0:
        print(f"    ✅ Mejora completada: {populated} valores históricos añadidos")
        return True
    else:
        print(f"    ℹ️  No se necesitaron mejoras")
        return False

if __name__ == "__main__":
    # Test con WATTS SA
    test_dir = Path('/home/unzzui/Documents/coding/CMF_extract/data/XBRL/Total/76455830-8_WATTS_SA')
    if test_dir.exists():
        result = enhance_watts_data(test_dir)
        print(f"Resultado: {'Mejorado' if result else 'Sin cambios'}")