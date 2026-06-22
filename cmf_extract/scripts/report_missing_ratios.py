#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook


def find_header_row(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower() in ("indicador", "indicator"):
            return r
        # también detectar fila que contenga años/Último
        hits = 0
        for c in range(2, min(ws.max_column, 30) + 1):
            hv = ws.cell(row=r, column=c).value
            if isinstance(hv, str) and hv.strip() and (hv.strip().isdigit() or hv.strip().lower() in ("último", "ultimo", "promedio", "tendencia", "last", "average", "trend")):
                hits += 1
        if hits >= 3:
            return r
    return 4


def year_columns(ws, hdr: int) -> List[Tuple[int, str]]:
    stop_tokens = {"último", "ultimo", "promedio", "tendencia", "last", "average", "trend"}
    cols: List[Tuple[int, str]] = []
    for c in range(2, ws.max_column + 1):
        hv = ws.cell(row=hdr, column=c).value
        if not isinstance(hv, str):
            continue
        txt = hv.strip()
        if txt.lower() in stop_tokens:
            break
        cols.append((c, txt))
    return cols


def is_section_row(name: object) -> bool:
    if not isinstance(name, str):
        return False
    s = name.strip()
    su = s.upper()
    # Secciones ES
    if su in {
        "LIQUIDEZ", "SOLVENCIA Y ESTRUCTURA", "RENTABILIDAD", "EFICIENCIA OPERATIVA",
        "FLUJOS Y ADICIONALES", "CREACIÓN DE VALOR", "COBERTURA Y RIESGO",
        "UTILIDADES", "UTILS",
    }:
        return True
    # Secciones EN
    if su in {
        "LIQUIDITY", "SOLVENCY & CAPITAL STRUCTURE", "PROFITABILITY", "OPERATING EFFICIENCY",
        "CASH FLOWS & OTHER", "VALUE CREATION", "COVERAGE & RISK",
    }:
        return True
    # Bloque tooltip
    if s.startswith("Definición y Fórmula (tooltip)") or s.startswith("Definition and Formula (tooltip)"):
        return True
    return False


def scan_workbook(path: Path) -> str:
    wb = load_workbook(str(path), read_only=True, data_only=False)
    if "RATIOS & KPIs" not in wb.sheetnames:
        return f"- {path.name}: hoja 'RATIOS & KPIs' no encontrada\n"
    ws = wb["RATIOS & KPIs"]
    hdr = find_header_row(ws)
    ycols = year_columns(ws, hdr)
    if not ycols:
        return f"- {path.name}: sin columnas de períodos detectadas\n"
    miss_rows: List[Tuple[str, int]] = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if is_section_row(name):
            continue
        if not isinstance(name, str):
            continue
        missing = 0
        has_any = False
        for c, _ in ycols:
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                has_any = True
                continue
            if v not in (None, ""):
                has_any = True
                continue
            missing += 1
        if missing and not has_any:
            miss_rows.append((name.strip(), missing))
    if not miss_rows:
        return f"- {path.name}: OK (todas las filas tienen fórmula o valor en algún período)\n"
    lines = [f"- {path.name}: {len(miss_rows)} fila(s) con celdas totalmente vacías en períodos"]
    for nm, cnt in miss_rows[:20]:
        lines.append(f"    • {nm}  (vacías: {cnt})")
    if len(miss_rows) > 20:
        lines.append(f"    … y {len(miss_rows)-20} más")
    return "\n".join(lines) + "\n"


def main() -> int:
    base = Path("Product_v1/Total")
    outputs: List[str] = []
    if not base.exists():
        print("No existe Product_v1/Total")
        return 2
    for xlsx in sorted(base.glob("*.xlsx")):
        try:
            outputs.append(scan_workbook(xlsx))
        except Exception as ex:
            outputs.append(f"- {xlsx.name}: error {ex}\n")
    report = "".join(outputs)
    print(report)
    out_path = base / "_ratios_formula_report.txt"
    try:
        out_path.write_text(report, encoding="utf-8")
        print(f"Reporte escrito en {out_path}")
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


