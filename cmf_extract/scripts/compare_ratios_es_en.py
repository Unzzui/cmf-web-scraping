#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Compara los ratios (RATIOS & KPIs) entre archivos ES y EN en Product_v1/Total.

- Recorre pares ES/EN por empresa (nombres pretty) en Product_v1/Total
- Evalúa las celdas de las columnas de periodos (YYYY o YYYYQn) en la hoja
  "RATIOS & KPIs" resolviendo las fórmulas con un evaluador ligero (IFERROR,
  AVERAGE, ABS y aritmética básica), usando los valores numéricos de las hojas
  base (Balance/Resultados/Flujo) y referencias dentro de la misma hoja
  (p.ej., EBITDA en UTILIDADES).
- Reporta discrepancias con contexto (empresa, indicador, periodo).

Uso:
  python scripts/compare_ratios_es_en.py [--base Product_v1/Total] [--limit 50]

Salida: código 0 si no hay diferencias; 2 si se detectan diferencias.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook


PRETTY_RE = re.compile(
    r"^(?P<company>.+) - (?P<rut>\d{7,8}(?:-[0-9Kk])?) - (?P<title>.+) (?P<range>\d{4}-\d{4}(?:Q[1-4])?) \[(?P<lang>ES|EN)\]\.xlsx$"
)

KPIS_SHEET = "RATIOS & KPIs"


def _normalize_label(raw: Any) -> Optional[str]:
    if not isinstance(raw, str):
        return None
    s = raw.strip().split("\n", 1)[0]
    if re.match(r"^\d{4}Q[1-4]$", s):
        return s
    m = re.match(r"^(\d{4})-(\d{2})(?:-\d{2})?$", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        q = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(mo)
        return f"{y}{q}" if q else str(y)
    # Bare YYYY -> YYYYQ4 for backward compat
    if re.match(r"^\d{4}$", s):
        return f"{s}Q4"
    return None


def _period_cols(ws, hdr_row: int) -> List[Tuple[int, str]]:
    out: List[Tuple[int, str]] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        lb = _normalize_label(v)
        if lb:
            out.append((c, lb))
    return out


def _is_numberlike(x: Any) -> bool:
    try:
        if x in (None, ""):
            return False
        float(str(x).replace(",", "").replace(" ", ""))
        return True
    except Exception:
        return False


def _to_number(x: Any) -> Optional[float]:
    if x in (None, ""):
        return None
    try:
        return float(str(x).replace(",", "").replace(" ", ""))
    except Exception:
        return None


def _find_header_row(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("indicador", "indicator"):
            return r
    return 4


class FormulaEvaluator:
    def __init__(self, wb):
        self.wb = wb
        self.memo: Dict[Tuple[str, int, int], Optional[float]] = {}

    def eval_cell(self, ws_name: str, row: int, col: int) -> Optional[float]:
        key = (ws_name, row, col)
        if key in self.memo:
            return self.memo[key]
        ws = self.wb[ws_name]
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            res = float(v)
            self.memo[key] = res
            return res
        if isinstance(v, str) and v.startswith("="):
            res = self._eval_formula(ws_name, row, col, v)
            self.memo[key] = res
            return res
        # empty or text
        self.memo[key] = _to_number(v)
        return self.memo[key]

    def _eval_formula(self, ws_name: str, row: int, col: int, formula: str) -> Optional[float]:
        expr = formula.lstrip("=").strip()

        # Capturar referencias 'Hoja'!A1 y A1 (misma hoja)
        # Orden: primero con hoja, luego sin hoja
        ref_pat_with_sheet = re.compile(r"'([^']+)'!([A-Z]{1,3})(\d{1,7})")
        ref_pat_no_sheet = re.compile(r"\b([A-Z]{1,3})(\d{1,7})\b")

        locals_map: Dict[str, Any] = {}
        fn_map = {
            'IFERROR': lambda x, y: x if (x is not None and not (isinstance(x, float) and (math.isnan(x) or math.isinf(x)))) else (None if y == "" else y),
            'AVERAGE': lambda *args: (sum(a for a in args if a is not None) / max(1, sum(1 for a in args if a is not None))) if any(a is not None for a in args) else None,
            'ABS': lambda x: abs(x) if x is not None else None,
        }

        # Resolver referencias con hoja
        idx = 0
        def _coord_to_rc(a1: str) -> Tuple[int, int]:
            from openpyxl.utils.cell import coordinate_to_tuple
            r, c = coordinate_to_tuple(a1)
            return r, c

        def _sub_with_sheet(m):
            nonlocal idx
            sheet = m.group(1)
            a1 = f"{m.group(2)}{m.group(3)}"
            r, c = _coord_to_rc(a1)
            val = self.eval_cell(sheet, r, c)
            name = f"__v{idx}"
            locals_map[name] = val
            idx += 1
            return name

        expr = ref_pat_with_sheet.sub(_sub_with_sheet, expr)

        # Resolver referencias sin hoja (misma hoja KPIs)
        def _sub_no_sheet(m):
            nonlocal idx
            a1 = f"{m.group(1)}{m.group(2)}"
            r, c = _coord_to_rc(a1)
            val = self.eval_cell(ws_name, r, c)
            name = f"__v{idx}"
            locals_map[name] = val
            idx += 1
            return name

        expr = ref_pat_no_sheet.sub(_sub_no_sheet, expr)

        # Reemplazar comas por comas de Python en llamadas a funciones; mantener formato
        # No se requiere sustitución especial: usaremos eval con funciones exponiendo nombres en locals.
        safe_locals = {**locals_map, **fn_map}

        try:
            val = eval(expr, {"__builtins__": {}}, safe_locals)
        except Exception:
            val = None
        return float(val) if _to_number(val) is not None else None


def find_pairs(base_dir: Path) -> List[Tuple[Path, Path]]:
    files = list(base_dir.glob("*.xlsx"))
    by_key: Dict[Tuple[str, str], Dict[str, Path]] = {}
    for p in files:
        m = PRETTY_RE.match(p.name)
        if not m:
            continue
        key = (m.group("rut"), m.group("range"))
        lang = m.group("lang").upper()
        by_key.setdefault(key, {})[lang] = p
    pairs: List[Tuple[Path, Path]] = []
    for mp in by_key.values():
        es = mp.get("ES")
        en = mp.get("EN")
        if es and en:
            pairs.append((es, en))
    return pairs


def compare_pair(es_path: Path, en_path: Path, limit: int = 50, tol: float = 1e-6) -> Tuple[int, List[str]]:
    wb_es = load_workbook(str(es_path), data_only=False)
    wb_en = load_workbook(str(en_path), data_only=False)
    if KPIS_SHEET not in wb_es.sheetnames or KPIS_SHEET not in wb_en.sheetnames:
        return 0, []
    ws_es = wb_es[KPIS_SHEET]
    ws_en = wb_en[KPIS_SHEET]
    hdr_es = _find_header_row(ws_es)
    hdr_en = _find_header_row(ws_en)
    cols_es = _period_cols(ws_es, hdr_es)
    cols_en = _period_cols(ws_en, hdr_en)
    labels_es = {lb for (_, lb) in cols_es}
    labels_en = {lb for (_, lb) in cols_en}
    common = sorted(labels_es & labels_en, key=lambda s: (int(s[:4]), int(s[5]) if re.match(r"^\d{4}Q[1-4]$", s) else 5))
    cidx_es: Dict[str, int] = {lb: c for (c, lb) in cols_es}
    cidx_en: Dict[str, int] = {lb: c for (c, lb) in cols_en}

    ev_es = FormulaEvaluator(wb_es)
    ev_en = FormulaEvaluator(wb_en)

    mismatches: List[str] = []
    # Recorrer filas de datos (omitir headers)
    start_es = hdr_es + 1
    start_en = hdr_en + 1
    max_rows = max(ws_es.max_row, ws_en.max_row)
    for r in range(0, max_rows - start_es + 1):
        rr_es = start_es + r
        rr_en = start_en + r
        if rr_es > ws_es.max_row or rr_en > ws_en.max_row:
            break
        name_es = ws_es.cell(row=rr_es, column=1).value
        name_en = ws_en.cell(row=rr_en, column=1).value
        # Saltar filas de sección (títulos) que no contienen datos
        if isinstance(name_es, str) and name_es.strip().upper() in {
            "LIQUIDEZ", "SOLVENCIA Y ESTRUCTURA", "RENTABILIDAD", "EFICIENCIA OPERATIVA", "FLUJOS Y ADICIONALES", "CREACIÓN DE VALOR", "COBERTURA Y RIESGO", "UTILIDADES"
        }:
            continue
        if isinstance(name_en, str) and name_en.strip().upper() in {
            "LIQUIDITY", "SOLVENCY & CAPITAL STRUCTURE", "PROFITABILITY", "OPERATING EFFICIENCY", "CASH FLOWS & OTHER", "VALUE CREATION", "COVERAGE & RISK", "UTILS"
        }:
            continue
        # Si la fila está completamente vacía en ambas, saltar
        any_val = False
        for lb in common:
            ce = cidx_es[lb]
            cv = ws_es.cell(row=rr_es, column=ce).value
            cn = cidx_en[lb]
            cv2 = ws_en.cell(row=rr_en, column=cn).value
            if (cv not in (None, "")) or (cv2 not in (None, "")):
                any_val = True
                break
        if not any_val:
            continue
        # Comparar valores evaluados por periodo
        for lb in common:
            ce = cidx_es[lb]
            cn = cidx_en[lb]
            ve = ev_es.eval_cell(KPIS_SHEET, rr_es, ce)
            vn = ev_en.eval_cell(KPIS_SHEET, rr_en, cn)
            ok = (ve is None and vn in (None, 0.0)) or (vn is None and ve in (None, 0.0))
            if not ok:
                try:
                    ok = (ve is not None and vn is not None and abs(float(ve) - float(vn)) <= tol)
                except Exception:
                    ok = False
            if not ok:
                if len(mismatches) < limit:
                    mismatches.append(
                        f"{es_path.name} vs {en_path.name} :: fila={rr_es} ES='{name_es}' EN='{name_en}' periodo={lb}  ES={ve} EN={vn}"
                    )
                else:
                    # cortar temprano si supera un umbral razonable
                    mismatches.append("…")
                    return len(mismatches), mismatches
    return len(mismatches), mismatches


def main() -> int:
    ap = argparse.ArgumentParser(description="Comparar ratios ES vs EN en Product_v1/Total")
    ap.add_argument("--base", type=Path, default=Path("Product_v1/Total"))
    ap.add_argument("--limit", type=int, default=50)
    args = ap.parse_args()

    base = args.base
    if not base.exists():
        print(f"No existe {base}")
        return 2

    pairs = find_pairs(base)
    if not pairs:
        print(f"No se encontraron pares ES/EN en {base}")
        return 2

    total_pairs = 0
    total_mism = 0
    for es, en in pairs:
        total_pairs += 1
        n, lines = compare_pair(es, en, limit=args.limit)
        total_mism += n
        if n > 0:
            print(f"✖ {es.name} vs {en.name}: {n} diferencia(s)")
            for ln in lines:
                print(f"   - {ln}")
        else:
            print(f"✔ {es.name} vs {en.name}: sin diferencias")

    if total_mism > 0:
        print(f"TOTAL: {total_mism} diferencia(s) en {total_pairs} pares")
        return 2
    print(f"OK. {total_pairs} pares verificados, sin diferencias.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


