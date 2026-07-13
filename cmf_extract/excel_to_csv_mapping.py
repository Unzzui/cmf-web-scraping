#!/usr/bin/env python3
"""
Script para extraer datos de archivos Excel en Product_v1/Total y mapearlos con role codes
del archivo new_eeff_estructura.json para generar un CSV con Label y RoleCode.
"""

import os
import json
import pandas as pd
import re
from pathlib import Path
import csv
from openpyxl import load_workbook

def extract_rut_from_filename(filename):
    """Extraer RUT del nombre del archivo."""
    rut_match = re.search(r'(\d{8,9}-[\dkK])', filename)
    return rut_match.group(1) if rut_match else None

def load_estructura_json(json_path):
    """Cargar el archivo de estructura y crear un diccionario de mapeo RUT -> roles."""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    rut_to_roles = {}
    for empresa_data in data['empresas']:
        rut = empresa_data['empresa']['rut']
        roles = {}
        for role in empresa_data['roles']:
            role_id = role['id']
            titulo = role['titulo']
            lineas = role['lineas']
            roles[role_id] = {
                'titulo': titulo,
                'lineas': set(lineas)  # Usar set para búsqueda rápida
            }
        rut_to_roles[rut] = roles
    
    return rut_to_roles

def map_sheet_to_role_id(sheet_name, roles_data):
    """Mapear nombre de hoja a role_id basado en el título."""
    sheet_mappings = {
        'Balance General': ['situación financiera'],
        'Estado de Resultados': ['resultado', 'estado del resultado'],
        'Flujo Efectivo': ['flujos de efectivo', 'flujo de efectivo']
    }

    search_terms = sheet_mappings.get(sheet_name, [])
    if not search_terms:
        return None

    for role_id, role_info in roles_data.items():
        titulo_lower = role_info['titulo'].lower()
        for search_term in search_terms:
            if search_term in titulo_lower:
                return role_id

    return None


# Marcadores para distinguir el estado de resultados por FUNCIÓN (310000) del
# estado por NATURALEZA (320000) leyendo las propias cuentas de la hoja.
_ER_FUNCION_MARKERS = ('costo de ventas', 'ganancia bruta')
_ER_NATURALEZA_MARKERS = (
    'gastos por beneficios a los empleados',
    'materias primas y consumibles',
    'otros gastos, por naturaleza',
)


def derive_role_id(sheet_name, labels):
    """Deducir el role_id de una hoja sin depender de la plantilla JSON.

    ``new_eeff_estructura.json`` sólo cubre 53 empresas, pero el JSON se usa
    únicamente para saber qué RoleCode estampar por hoja. Las demás empresas se
    descartaban en silencio y nunca llegaban a la BD. Balance y Flujo tienen
    código fijo; para el ER se distingue por función vs por naturaleza según las
    cuentas presentes.
    """
    if sheet_name == 'Balance General':
        return '210000'
    if sheet_name == 'Flujo Efectivo':
        return '510000'
    if sheet_name != 'Estado de Resultados':
        return None

    low = [str(l).strip().lower() for l in labels if l is not None]
    if any(any(m in l for m in _ER_FUNCION_MARKERS) for l in low):
        return '310000'
    if any(any(m in l for m in _ER_NATURALEZA_MARKERS) for l in low):
        return '320000'
    return '310000'  # mismo default que detect_income_statement_role_from_facts

def extract_data_from_excel(excel_path, sheet_name, rut_to_roles, rut):
    """Extraer labels y datos de todas las fechas de una hoja específica del Excel."""
    try:
        # Leer el archivo Excel sin header
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        
        # Buscar la fila que contiene "Cuenta" (normalmente fila 2)
        cuenta_row = None
        for i, row in df.iterrows():
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() == 'Cuenta':
                cuenta_row = i
                break
        
        if cuenta_row is None:
            print(f"  ⚠️  No se encontró fila 'Cuenta' en {sheet_name}")
            return [], []
        
        # Obtener role_id para esta hoja. Si la empresa no está en la plantilla
        # JSON (o la plantilla no cubre su rol), se deduce de las propias cuentas.
        roles_data = rut_to_roles.get(rut, {})
        role_id = map_sheet_to_role_id(sheet_name, roles_data)

        if not role_id:
            sheet_labels = df.iloc[cuenta_row + 1:, 0].tolist()
            role_id = derive_role_id(sheet_name, sheet_labels)

        if not role_id:
            print(f"  ⚠️  No se encontró role_id para {sheet_name} (RUT: {rut})")
            return [], []
        
        # Extraer nombres de fechas desde la fila 'Cuenta' manteniendo orden original
        fecha_row = df.iloc[cuenta_row]
        date_columns = []
        for col_idx in range(1, len(fecha_row)):  # Empezar desde col 1 (después de 'Cuenta')
            if pd.notna(fecha_row.iloc[col_idx]):
                date_val = str(fecha_row.iloc[col_idx]).strip()
                if date_val and date_val != 'Cuenta':
                    # Limpiar formato de fecha para mantener formato original
                    if date_val.endswith('.0'):
                        date_val = date_val[:-2]  # Remover .0 al final
                    date_columns.append((col_idx, date_val))
            else:
                break  # Parar cuando encontremos una celda vacía
        
        # Extraer datos desde la fila después de 'Cuenta'
        data_rows = []
        for i in range(cuenta_row + 1, len(df)):
            label = df.iloc[i, 0]  # Primera columna contiene los labels
            if pd.notna(label) and str(label).strip():
                clean_label = str(label).strip()
                # Filtrar filas vacías o que no son cuentas reales
                if clean_label and not clean_label.startswith('Unidad:'):
                    # Crear diccionario base para esta fila
                    row_data = {
                        'Label': clean_label,
                        'RoleCode': role_id
                    }
                    
                    # Agregar valores para cada fecha
                    for col_idx, date_name in date_columns:
                        if col_idx < len(df.iloc[i]):
                            value = df.iloc[i, col_idx]
                            if pd.isna(value):
                                row_data[date_name] = None
                            else:
                                # Convertir a string para mantener formato original
                                row_data[date_name] = str(value) if value != '' else None
                        else:
                            row_data[date_name] = None
                    
                    data_rows.append(row_data)
        
        # Extraer solo los nombres de fechas para el header
        date_names = [date_name for _, date_name in date_columns]
        
        print(f"  ✓ {len(data_rows)} filas extraídas de {sheet_name} (Role: {role_id}, {len(date_names)} fechas)")
        return data_rows, date_names
        
    except Exception as e:
        print(f"  ❌ Error procesando {sheet_name}: {str(e)}")
        return [], []

def extract_ratios_kpis_data(excel_path, rut):
    """Extraer todos los ratios y KPIs de la hoja RATIOS & KPIs."""
    try:
        # Cargar workbook con openpyxl para evaluar fórmulas
        wb = load_workbook(excel_path, data_only=True)
        ws = wb['RATIOS & KPIs']
        
        # Convertir worksheet a DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data)
        
        # Buscar la fila que contiene "Indicador"
        indicador_row = None
        for i, row in df.iterrows():
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() == 'Indicador':
                indicador_row = i
                break
        
        if indicador_row is None:
            print(f"  ⚠️  No se encontró fila 'Indicador' en RATIOS & KPIs")
            return [], []
        
        # Extraer nombres de fechas desde la fila 'Indicador'
        fecha_row = df.iloc[indicador_row]
        date_columns = []
        for col_idx in range(1, len(fecha_row)):  # Empezar desde col 1 (después de 'Indicador')
            if pd.notna(fecha_row.iloc[col_idx]):
                date_val = str(fecha_row.iloc[col_idx]).strip()
                if date_val and date_val != 'Indicador':
                    # Limpiar formato de fecha
                    if date_val.endswith('.0'):
                        date_val = date_val[:-2]
                    date_columns.append((col_idx, date_val))
            else:
                break
        
        # Solo extraer estos dos indicadores específicos
        target_indicators = [
            'Depreciación y Amortización',
            'Total número de acciones emitidas',
        ]
        
        # Extraer solo los indicadores específicos de la hoja RATIOS & KPIs
        data_rows = []
        for i in range(indicador_row + 1, len(df)):
            label = df.iloc[i, 0]  # Primera columna contiene los labels
            if pd.notna(label) and str(label).strip():
                clean_label = str(label).strip()
                
                # Solo procesar los indicadores específicos que nos interesan
                if any(target in clean_label for target in target_indicators):
                    # Verificar que la fila tenga al menos un valor no vacío
                    has_data = False
                    for col_idx, _ in date_columns:
                        if col_idx < len(df.iloc[i]):
                            value = df.iloc[i, col_idx]
                            if pd.notna(value) and str(value).strip() != '':
                                has_data = True
                                break
                    
                    # Solo procesar si tiene datos reales
                    if has_data:
                        # Crear diccionario base para esta fila con rol 000000
                        row_data = {
                            'Label': clean_label,
                            'RoleCode': '000000'
                        }
                        
                        # Agregar valores para cada fecha
                        for col_idx, date_name in date_columns:
                            if col_idx < len(df.iloc[i]):
                                value = df.iloc[i, col_idx]
                                if pd.isna(value):
                                    row_data[date_name] = None
                                else:
                                    # Convertir a string para mantener formato original
                                    row_data[date_name] = str(value) if value != '' else None
                            else:
                                row_data[date_name] = None
                        
                        data_rows.append(row_data)
        
        # Extraer solo los nombres de fechas para el header
        date_names = [date_name for _, date_name in date_columns]
        
        print(f"  ✓ {len(data_rows)} ratios/KPIs extraídos de RATIOS & KPIs (Role: 000000, {len(date_names)} fechas)")
        return data_rows, date_names
        
    except Exception as e:
        print(f"  ❌ Error procesando RATIOS & KPIs: {str(e)}")
        return [], []

def process_excel_files(input_dir, json_path, output_dir, progress_callback=None,
                        filter_ruts=None):
    """Procesar todos los archivos Excel y generar CSV por empresa.

    Parameters
    ----------
    input_dir : str or Path
        Directory containing analysis Excel files.
    json_path : str or Path
        Path to the ``new_eeff_estructura.json`` role structure file.
    output_dir : str or Path
        Directory where CSV files will be written.
    progress_callback : callable, optional
        Called as ``progress_callback(message, current, total)`` to report progress.
    filter_ruts : set[str] or None, optional
        If given, only process Excel files whose RUT matches one in this set.
    """

    def _log(msg):
        if progress_callback:
            progress_callback(msg, 0, 0)
        else:
            print(msg)

    # Cargar estructura de roles
    _log("Cargando estructura de roles...")
    rut_to_roles = load_estructura_json(json_path)
    _log(f"  {len(rut_to_roles)} empresas cargadas")

    # Crear directorio de salida
    os.makedirs(output_dir, exist_ok=True)

    # Hojas a procesar
    sheets_to_process = ['Balance General', 'Estado de Resultados', 'Flujo Efectivo']

    # Procesar cada archivo Excel
    excel_files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx') and not f.startswith('.~lock')]

    # Filter by RUT if requested
    if filter_ruts:
        excel_files = [
            f for f in excel_files
            if extract_rut_from_filename(f) in filter_ruts
        ]

    _log(f"Procesando {len(excel_files)} archivos Excel...")
    
    total_files_processed = 0
    total_records = 0

    for idx, excel_file in enumerate(excel_files, 1):
        if progress_callback:
            progress_callback(f"[{idx}/{len(excel_files)}] {excel_file}", idx, len(excel_files))
        _log(f"Procesando: {excel_file}")

        # Extraer RUT y nombre de empresa del archivo
        rut = extract_rut_from_filename(excel_file)
        if not rut:
            _log(f"  No se pudo extraer RUT de {excel_file}")
            continue

        # Extraer nombre de empresa del archivo
        company_name = excel_file.split(' - ')[0] if ' - ' in excel_file else f"Empresa_{rut}"
        company_name = company_name.replace(' ', '_').replace('/', '_')

        _log(f"  RUT: {rut} | Empresa: {company_name}")

        # La plantilla JSON sólo cubre 53 empresas. Antes, las demás se
        # descartaban aquí en silencio (y la fase reportaba "ok"), así que nunca
        # podían llegar a la BD. Ahora se procesan igual: el RoleCode de cada
        # hoja se deduce de las cuentas (ver derive_role_id).
        if rut not in rut_to_roles:
            _log(f"  RUT {rut} sin plantilla JSON; se deduce el rol de cada hoja")

        excel_path = os.path.join(input_dir, excel_file)
        
        # Recopilar todas las fechas únicas de todas las hojas
        all_date_columns = set()
        company_data = []
        
        # Procesar cada hoja para recopilar datos y fechas
        for sheet_name in sheets_to_process:
            try:
                sheet_data, date_names = extract_data_from_excel(excel_path, sheet_name, rut_to_roles, rut)
                company_data.extend(sheet_data)
                all_date_columns.update(date_names)
            except Exception as e:
                _log(f"  Error en hoja {sheet_name}: {str(e)}")

        # Procesar hoja RATIOS & KPIs por separado
        try:
            ratios_data, ratios_date_names = extract_ratios_kpis_data(excel_path, rut)
            company_data.extend(ratios_data)
            all_date_columns.update(ratios_date_names)
        except Exception as e:
            _log(f"  Error en hoja RATIOS & KPIs: {str(e)}")
        
        # Generar CSV para esta empresa
        if company_data:
            csv_filename = f"{company_name}_{rut}_financial_data.csv"
            csv_path = os.path.join(output_dir, csv_filename)
            
            # Crear fieldnames: Label, RoleCode, luego fechas en orden original del Excel
            # No ordenamos las fechas, mantenemos el orden original
            ordered_dates = []
            # Tomar las fechas de la primera hoja procesada para mantener orden consistente
            for sheet_name in sheets_to_process:
                try:
                    df_temp = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                    cuenta_row_temp = None
                    for i, row in df_temp.iterrows():
                        if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() == 'Cuenta':
                            cuenta_row_temp = i
                            break
                    if cuenta_row_temp is not None:
                        fecha_row_temp = df_temp.iloc[cuenta_row_temp]
                        for col_idx in range(1, len(fecha_row_temp)):
                            if pd.notna(fecha_row_temp.iloc[col_idx]):
                                date_val = str(fecha_row_temp.iloc[col_idx]).strip()
                                if date_val and date_val != 'Cuenta':
                                    if date_val.endswith('.0'):
                                        date_val = date_val[:-2]
                                    if date_val not in ordered_dates:
                                        ordered_dates.append(date_val)
                            else:
                                break
                        break  # Solo necesitamos el orden de la primera hoja
                except:
                    continue
            
            # También agregar fechas de RATIOS & KPIs si no están ya incluidas
            try:
                df_ratios = pd.read_excel(excel_path, sheet_name='RATIOS & KPIs', header=None)
                indicador_row_temp = None
                for i, row in df_ratios.iterrows():
                    if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() == 'Indicador':
                        indicador_row_temp = i
                        break
                if indicador_row_temp is not None:
                    fecha_row_temp = df_ratios.iloc[indicador_row_temp]
                    for col_idx in range(1, len(fecha_row_temp)):
                        if pd.notna(fecha_row_temp.iloc[col_idx]):
                            date_val = str(fecha_row_temp.iloc[col_idx]).strip()
                            if date_val and date_val != 'Indicador':
                                if date_val.endswith('.0'):
                                    date_val = date_val[:-2]
                                if date_val not in ordered_dates:
                                    ordered_dates.append(date_val)
                        else:
                            break
            except:
                pass
            
            fieldnames = ['Label', 'RoleCode'] + ordered_dates

            # Blindaje: `ordered_dates` sale de escanear la fila 'Cuenta' de la primera
            # hoja, y ese escaneo corta al primer hueco de la cabecera. Si una fila trae
            # un período que el escaneo no vio, DictWriter aborta con ValueError y se cae
            # la corrida COMPLETA (una empresa mata a las 230). Se completan los períodos
            # faltantes a partir de las filas reales, ordenados de más reciente a más
            # antiguo, igual que el orden natural del Excel.
            faltantes = {k for row in company_data for k in row} - set(fieldnames)
            if faltantes:
                def _orden_periodo(p: str):
                    m = re.match(r'^(\d{4})(?:Q([1-4]))?$', str(p))
                    return (-int(m.group(1)), -int(m.group(2) or 4)) if m else (0, 0)
                fieldnames += sorted(faltantes, key=_orden_periodo)
                _log(f"  ⚠ {len(faltantes)} período(s) no detectados en la cabecera, "
                     f"recuperados de las filas: {sorted(faltantes, key=_orden_periodo)[:4]}...")

            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(company_data)
            
            _log(f"  CSV generado: {csv_filename} ({len(company_data)} registros, {len(ordered_dates)} fechas)")
            total_files_processed += 1
            total_records += len(company_data)
        else:
            _log(f"  No se pudieron extraer datos para {company_name}")

    _log(f"Proceso completado: {total_files_processed} empresas, {total_records} registros")
    return total_files_processed, total_records

def main():
    """Función principal."""
    
    # Rutas
    base_dir = "/home/unzzui/Documents/coding/CMF_extract"
    input_dir = os.path.join(base_dir, "Product_v1", "Total")
    json_path = os.path.join(base_dir, "new_eeff_estructura.json")
    output_dir = os.path.join(base_dir, "Product_v1", "TO_SQL")
    
    print("🚀 Iniciando extracción de Labels y RoleCodes")
    print(f"📂 Input: {input_dir}")
    print(f"📋 JSON: {json_path}")
    print(f"💾 Output: {output_dir}")
    
    # Verificar que existan los archivos/directorios necesarios
    if not os.path.exists(input_dir):
        print(f"❌ No existe directorio: {input_dir}")
        return
    
    if not os.path.exists(json_path):
        print(f"❌ No existe archivo JSON: {json_path}")
        return
    
    # Procesar archivos
    process_excel_files(input_dir, json_path, output_dir)
    
    print("\n🎉 Proceso completado!")

if __name__ == "__main__":
    main()