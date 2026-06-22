#!/usr/bin/env python3
"""
Script para poblar datos históricos de WATTS SA desde Excel existente a facts CSV
Hace matching exacto por nombre de cuenta y valores para asegurar mapeo correcto
"""

import pandas as pd
import openpyxl
import numpy as np
import re
from pathlib import Path

def clean_numeric_value(value, from_facts=False):
    """Limpia y convierte valores numéricos"""
    if pd.isna(value) or value is None or value == 'None':
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    # Si es string, limpiar formato
    value_str = str(value).strip()
    if not value_str or value_str.lower() == 'none' or value_str == 'nan':
        return None
    
    # Remover formato de miles si es del facts CSV
    if from_facts:
        value_str = value_str.replace(',', '')
    
    try:
        return float(value_str)
    except:
        return None

def normalize_account_name(name):
    """Normaliza nombres de cuentas para matching"""
    if pd.isna(name) or not name:
        return ""
    
    # Convertir a string y limpiar
    name = str(name).strip()
    
    # Remover [sinopsis] y otros tags
    name = re.sub(r'\[.*?\]', '', name)
    
    # Normalizar espacios y puntuación
    name = re.sub(r'\s+', ' ', name)
    name = name.strip()
    
    return name.lower()

def extract_excel_data(excel_path):
    """Extrae datos del Excel existente"""
    print(f"Leyendo Excel: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    all_data = {}
    
    for sheet_name in ['Balance General', 'Estado de Resultados', 'Flujo Efectivo']:
        if sheet_name not in wb.sheetnames:
            continue
            
        print(f"\nProcesando hoja: {sheet_name}")
        ws = wb[sheet_name]
        
        # Encontrar fila de fechas (típicamente fila 3)
        date_row = 3
        dates = {}
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value:
                dates[col] = str(cell_value)
        
        print(f"  Columnas de fecha encontradas: {list(dates.values())}")
        
        # Extraer datos por cuenta
        for row in range(4, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            
            if not account_name or str(account_name).startswith('='):
                continue
            
            account_norm = normalize_account_name(account_name)
            if not account_norm:
                continue
            
            # Extraer valores para cada fecha
            for col, date in dates.items():
                value = ws.cell(row=row, column=col).value
                clean_value = clean_numeric_value(value)
                
                if clean_value is not None:
                    key = (sheet_name, account_name, date)
                    all_data[key] = {
                        'sheet': sheet_name,
                        'account': account_name,
                        'account_norm': account_norm,
                        'date': date,
                        'value': clean_value,
                        'row': row,
                        'col': col
                    }
    
    wb.close()
    
    print(f"\nTotal de datos extraídos: {len(all_data)}")
    
    # Convertir a DataFrame para análisis
    df = pd.DataFrame.from_dict(all_data, orient='index')
    
    return df

def load_facts_csv(facts_path):
    """Carga el facts CSV actual"""
    print(f"\nCargando facts CSV: {facts_path}")
    facts_df = pd.read_csv(facts_path)
    
    # Identificar columnas de fecha
    date_columns = [col for col in facts_df.columns 
                   if col not in ['LabelKeyId', 'LabelKeyIdExt', 'SectionKey', 'Label', 'RoleCode']]
    
    print(f"  Columnas de fecha en facts: {date_columns}")
    print(f"  Total de filas: {len(facts_df)}")
    
    return facts_df, date_columns

def find_matching_accounts(excel_df, facts_df, date_columns):
    """Encuentra coincidencias entre cuentas del Excel y facts CSV"""
    print("\n=== BUSCANDO COINCIDENCIAS DE CUENTAS ===")
    
    matches = []
    checked_pairs = 0
    name_matches = 0
    
    # Para cada cuenta en facts
    for idx, facts_row in facts_df.iterrows():
        facts_label = facts_row['Label']
        facts_norm = normalize_account_name(facts_label)
        
        if not facts_norm:
            continue
        
        # Buscar en Excel por nombre similar y valor coincidente
        for _, excel_row in excel_df.iterrows():
            excel_norm = excel_row['account_norm']
            checked_pairs += 1
            
            # Comparar nombres normalizados
            if facts_norm == excel_norm or facts_norm in excel_norm or excel_norm in facts_norm:
                name_matches += 1
                if checked_pairs <= 5:  # Mostrar solo los primeros para debug
                    print(f"  Nombre match: '{facts_label[:40]}' <-> '{excel_row['account'][:40]}'")
                # Verificar coincidencia de valor en algún período común
                excel_date = excel_row['date']
                excel_value = excel_row['value']
                
                # Mapear fecha Excel a fecha facts
                date_mapping = {
                    '2025Q2': '2025-06-30',
                    '2025Q1': '2025-03-31',
                    '2024': '2024-12-31',
                    '2024Q3': '2024-09-30',
                    '2024Q2': '2024-06-30',
                    '2024Q1': '2024-03-31',
                    '2023': '2023-12-31',
                    '2023Q3': '2023-09-30',
                    '2023Q2': '2023-06-30',
                    '2023Q1': '2023-03-31',
                    '2022': '2022-12-31',
                    '2021': '2021-12-31'
                }
                
                if excel_date in date_mapping:
                    facts_date = date_mapping[excel_date]
                    
                    if facts_date in date_columns:
                        facts_value_str = str(facts_row[facts_date])
                        facts_value = clean_numeric_value(facts_value_str, from_facts=True)
                        
                        # Los valores del Excel están en miles, el facts en unidades
                        # Convertir Excel a unidades multiplicando por 1000
                        excel_value_units = excel_value * 1000
                        
                        # Comparar valores (con tolerancia para redondeo)
                        if facts_value is not None and excel_value_units is not None:
                            # Tolerancia del 1% o 1000 unidades
                            tolerance = max(abs(facts_value * 0.01), 1000)
                            if abs(facts_value - excel_value_units) < tolerance:
                                matches.append({
                                    'facts_idx': idx,
                                    'facts_label': facts_label,
                                    'excel_account': excel_row['account'],
                                    'excel_sheet': excel_row['sheet'],
                                    'matched_date': excel_date,
                                    'matched_value': excel_value,
                                    'confidence': 'HIGH'
                                })
                                break
    
    print(f"\nTotal pares revisados: {checked_pairs}")
    print(f"Coincidencias de nombre: {name_matches}")
    print(f"Coincidencias de alta confianza (nombre + valor): {len(matches)}")
    
    # Crear mapeo único por cuenta facts
    unique_matches = {}
    for match in matches:
        facts_idx = match['facts_idx']
        if facts_idx not in unique_matches:
            unique_matches[facts_idx] = match
    
    print(f"Cuentas únicas mapeadas: {len(unique_matches)}")
    
    return unique_matches

def populate_historical_data(facts_df, excel_df, matches, date_columns):
    """Pobla los datos históricos en el facts DataFrame"""
    print("\n=== POBLANDO DATOS HISTÓRICOS ===")
    
    # Mapeo de fechas Excel a facts
    date_mapping = {
        '2022': '2022-12-31',
        '2021': '2021-12-31'
    }
    
    populated_count = 0
    
    for facts_idx, match in matches.items():
        excel_account = match['excel_account']
        excel_sheet = match['excel_sheet']
        
        # Obtener todos los valores de esta cuenta del Excel
        excel_data = excel_df[
            (excel_df['account'] == excel_account) & 
            (excel_df['sheet'] == excel_sheet)
        ]
        
        for _, excel_row in excel_data.iterrows():
            excel_date = excel_row['date']
            
            # Solo procesar 2021 y 2022 (anuales)
            if excel_date in date_mapping:
                facts_date = date_mapping[excel_date]
                
                if facts_date in date_columns:
                    current_value = facts_df.at[facts_idx, facts_date]
                    
                    # Solo poblar si está vacío o es NaN
                    if pd.isna(current_value) or current_value == '' or current_value is None:
                        # Los valores del Excel están en miles, convertir a unidades
                        value = excel_row['value']
                        if value is not None:
                            value_units = int(value * 1000)
                            # Formatear valor con separador de miles
                            formatted_value = f"{value_units:,}"
                            facts_df.at[facts_idx, facts_date] = formatted_value
                            populated_count += 1
                            
                            print(f"  Poblado: {match['facts_label'][:50]} - {facts_date}: {formatted_value}")
    
    print(f"\nTotal de valores poblados: {populated_count}")
    
    return facts_df

def verify_population(facts_df, date_columns):
    """Verifica cuántos datos se poblaron"""
    print("\n=== VERIFICACIÓN DE POBLACIÓN ===")
    
    for date_col in ['2022-12-31', '2021-12-31']:
        if date_col in date_columns:
            non_empty = facts_df[date_col].notna().sum()
            print(f"  {date_col}: {non_empty} valores no vacíos")

def main():
    # Rutas
    excel_path = '/home/unzzui/Documents/coding/CMF_extract/Products/Total/estados_76455830-8_2025-2021_es.xlsx'
    facts_path = '/home/unzzui/Documents/coding/CMF_extract/data/XBRL/Total/76455830-8_WATTS_SA/out_consolidated_76455830-8_202312-202506/facts_76455830-8_202312-202506_es.csv'
    output_path = '/home/unzzui/Documents/coding/CMF_extract/data/XBRL/Total/76455830-8_WATTS_SA/out_consolidated_76455830-8_202312-202506/facts_76455830-8_202312-202506_es_populated.csv'
    
    # 1. Extraer datos del Excel
    excel_df = extract_excel_data(excel_path)
    
    # Mostrar resumen de datos Excel
    print("\n=== RESUMEN DE DATOS EXCEL ===")
    for date in ['2021', '2022']:
        count = len(excel_df[excel_df['date'] == date])
        print(f"  {date}: {count} valores")
    
    # 2. Cargar facts CSV
    facts_df, date_columns = load_facts_csv(facts_path)
    
    # 3. Encontrar coincidencias
    matches = find_matching_accounts(excel_df, facts_df, date_columns)
    
    # 4. Poblar datos históricos
    populated_df = populate_historical_data(facts_df.copy(), excel_df, matches, date_columns)
    
    # 5. Guardar resultado
    populated_df.to_csv(output_path, index=False)
    print(f"\n✓ Facts poblado guardado en: {output_path}")
    
    # 6. Verificar población
    verify_population(populated_df, date_columns)
    
    # También actualizar el original para respaldo
    backup_path = facts_path.replace('.csv', '_backup.csv')
    facts_df.to_csv(backup_path, index=False)
    print(f"✓ Backup del original guardado en: {backup_path}")
    
    # Actualizar el facts original
    populated_df.to_csv(facts_path, index=False)
    print(f"✓ Facts original actualizado: {facts_path}")

if __name__ == "__main__":
    main()