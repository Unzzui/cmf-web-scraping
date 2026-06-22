"""
Commercial Quality Tests
========================

Tests that verify the commercial quality of generated financial Excel files
for findatachile.com. These tests check:
- IFERROR fallbacks use "N/A" for ratios (not empty strings)
- Metadata sheet exists with required fields
- Data quality section is present
- Methodology sheet exists
- Unusual value flags column works
"""

import re
import pytest
from pathlib import Path
from unittest.mock import MagicMock


class TestIFERRORFallbacks:
    """Verify that formula_builder uses 'N/A' fallback for ratio IFERRORs."""

    def test_no_blank_iferror_in_ratio_formulas(self):
        """No ratio formula should use IFERROR(...,'') — all must use 'N/A'."""
        fb_dir = Path(__file__).parent.parent / "analisis_excel" / "formula_builder"
        source = "\n".join(p.read_text(encoding="utf-8") for p in sorted(fb_dir.glob("*.py")))

        # Find all IFERROR(...,"") occurrences
        blank_iferrors = re.findall(r'IFERROR\([^)]*,\s*\\?"\\?"\s*\)', source)

        # Filter: only flag lines that contain a division (/) — those are ratio formulas
        lines = source.splitlines()
        violations = []
        for i, line in enumerate(lines, 1):
            if 'IFERROR(' in line and ',""' in line and '/' in line:
                violations.append(f"  Line {i}: {line.strip()[:100]}")

        assert not violations, (
            f"Found {len(violations)} ratio formula(s) using IFERROR(...,'') "
            f"instead of IFERROR(...,'N/A'):\n" + "\n".join(violations[:10])
        )

    def test_absolute_value_formulas_keep_empty_string(self):
        """Absolute value formulas (subtraction only) should keep '' fallback."""
        fb_dir = Path(__file__).parent.parent / "analisis_excel" / "formula_builder"
        source = "\n".join(p.read_text(encoding="utf-8") for p in sorted(fb_dir.glob("*.py")))

        lines = source.splitlines()
        empty_iferror_lines = []
        for i, line in enumerate(lines, 1):
            if 'IFERROR(' in line and ',""' in line:
                empty_iferror_lines.append((i, line.strip()))

        # These should exist (Capital de Trabajo, CFO passthrough, EVA, Spread)
        assert len(empty_iferror_lines) > 0, "Expected some IFERROR(...,'') for absolute values"
        # But they should NOT contain division
        for lineno, line in empty_iferror_lines:
            # Extract the IFERROR content (before the fallback)
            # Allow division only in nested IFERROR (already handled)
            assert '/' not in line.split('IFERROR(')[1].split(',')[0] or 'IFERROR' in line.split('IFERROR(')[1].split(',')[0], (
                f"Line {lineno} has division in IFERROR(...,'') — should use 'N/A': {line[:100]}"
            )


class TestMetadataSheet:
    """Verify metadata_sheet module creates correct structure."""

    def test_metadata_sheet_exists_after_creation(self):
        """create_metadata_sheet should create a 'Ficha Técnica' sheet at index 0."""
        from openpyxl import Workbook
        from analisis_excel.metadata_sheet import create_metadata_sheet

        wb = Workbook()
        create_metadata_sheet(
            wb=wb,
            company_name="Test Company S.A.",
            rut="76129263-3",
            periods=["2020", "2021", "2022", "2023Q1"],
            frequency="Total",
            lang="es",
            currency="Miles de CLP",
        )

        assert "Ficha Técnica" in wb.sheetnames
        assert wb.sheetnames[0] == "Ficha Técnica"

    def test_metadata_sheet_has_required_fields(self):
        """The metadata sheet must contain all required field labels."""
        from openpyxl import Workbook
        from analisis_excel.metadata_sheet import create_metadata_sheet

        wb = Workbook()
        create_metadata_sheet(
            wb=wb,
            company_name="AGROSUPER S.A.",
            rut="76129263-3",
            periods=["2014", "2015", "2024", "2025Q1", "2025Q2"],
            frequency="Total (Anual + Trimestral)",
            lang="es",
            currency="Miles de CLP",
        )

        ws = wb["Ficha Técnica"]
        # Collect all values in column A
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value]

        required_labels = ["Empresa", "RUT", "Generado", "Fuente", "Períodos", "Moneda", "Frecuencia", "Versión", "Web"]
        for label in required_labels:
            assert any(label in str(v) for v in labels), f"Missing required field: {label}"

    def test_metadata_sheet_english(self):
        """English version should use 'Data Sheet' title."""
        from openpyxl import Workbook
        from analisis_excel.metadata_sheet import create_metadata_sheet

        wb = Workbook()
        create_metadata_sheet(
            wb=wb,
            company_name="Test Corp",
            rut="12345678-9",
            periods=["2022", "2023"],
            frequency="Annual",
            lang="en",
        )

        assert "Data Sheet" in wb.sheetnames

    def test_metadata_company_name_and_rut(self):
        """Company name and RUT should appear as cell values."""
        from openpyxl import Workbook
        from analisis_excel.metadata_sheet import create_metadata_sheet

        wb = Workbook()
        create_metadata_sheet(
            wb=wb,
            company_name="CENCOSUD S.A.",
            rut="93834000-5",
            periods=["2020", "2021"],
            frequency="Anual",
            lang="es",
        )

        ws = wb["Ficha Técnica"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "CENCOSUD S.A." in values
        assert "93834000-5" in values


class TestDataExtractorQuality:
    """Verify data_extractor tracks quality information."""

    def test_quality_fields_exist(self):
        """DataExtractor should have quality tracking attributes."""
        from analisis_excel.data_extractor import DataExtractor

        de = DataExtractor("dummy.xlsx")
        assert hasattr(de, 'missing_sheets')
        assert hasattr(de, 'warnings')
        assert hasattr(de, 'found_accounts')
        assert hasattr(de, 'missing_accounts')
        assert hasattr(de, 'estimated_accounts')
        assert hasattr(de, 'income_role')

    def test_income_role_default(self):
        """Default income role should be '310000' (function)."""
        from analisis_excel.data_extractor import DataExtractor

        de = DataExtractor("dummy.xlsx")
        assert de.income_role == "310000"


class TestExcelFormatterQuality:
    """Verify excel_formatter has quality section and N/A formatting."""

    def test_format_na_style_exists(self):
        """ExcelFormatter should have format_na_style method."""
        from analisis_excel.excel_formatter import ExcelFormatter

        ef = ExcelFormatter()
        assert hasattr(ef, 'format_na_style')

    def test_format_quality_section_exists(self):
        """ExcelFormatter should have format_quality_section method."""
        from analisis_excel.excel_formatter import ExcelFormatter

        ef = ExcelFormatter()
        assert hasattr(ef, 'format_quality_section')

    def test_format_quality_section_output(self):
        """format_quality_section should write quality data to a worksheet."""
        from openpyxl import Workbook
        from analisis_excel.excel_formatter import ExcelFormatter

        ef = ExcelFormatter()
        wb = Workbook()
        ws = wb.active

        quality_data = {
            "found_accounts": ["AC", "PC", "AT", "PT"],
            "missing_accounts": ["D&A", "CAPEX"],
            "estimated_accounts": [],
            "warnings": ["D&A not available"],
            "total_expected": 6,
        }

        next_row = ef.format_quality_section(ws, 1, quality_data, cols_total=5, lang="es")
        assert next_row > 1

        # Check title was written
        assert ws.cell(row=1, column=1).value == "CALIDAD DE DATOS"


class TestRatioCalculatorEdgeCases:
    """Verify ratio_calculator handles edge cases correctly."""

    def test_roe_negative_equity_returns_nan(self):
        """ROE should return NaN when average equity is negative."""
        import pandas as pd
        import numpy as np
        from analisis_excel.ratio_calculator import RatioCalculator

        financial_data = {
            "balance": {
                "AC": pd.Series({"2023-12-31": 100}),
                "PC": pd.Series({"2023-12-31": 200}),
                "AT": pd.Series({"2023-12-31": 500}),
                "PT": pd.Series({"2023-12-31": 600}),
                "Patr": pd.Series({"2023-12-31": -100, "2022-12-31": -50}),
                "Efec": pd.Series({"2023-12-31": 10}),
                "Inv": pd.Series({"2023-12-31": 20}),
                "CxC": pd.Series({"2023-12-31": 30}),
                "CxP": pd.Series({"2023-12-31": 40}),
            },
            "income": {
                "Ventas": pd.Series({"2023-12-31": 1000}),
                "COGS": pd.Series({"2023-12-31": 700}),
                "Bruta": pd.Series({"2023-12-31": 300}),
                "EBIT": pd.Series({"2023-12-31": 100}),
                "Neta": pd.Series({"2023-12-31": 50}),
                "Interes": pd.Series({"2023-12-31": -20}),
                "Dep": pd.Series({"2023-12-31": 10}),
                "Amort": pd.Series({"2023-12-31": 5}),
                "DA": pd.Series({"2023-12-31": 15}),
            },
            "cash_flow": {
                "CFO": pd.Series({"2023-12-31": 80}),
                "CAPEX": pd.Series({"2023-12-31": 30}),
                "FCF": pd.Series({"2023-12-31": 50}),
                "CapexBuy": pd.Series({"2023-12-31": -30}),
            },
            "years": [2023],
        }

        calc = RatioCalculator(financial_data)
        ratios = calc.calculate_profitability_ratios()

        # ROE should be empty (NaN) since equity is negative
        roe = ratios.get("ROE", pd.Series(dtype=float))
        assert roe.empty or roe.isna().all(), "ROE should be NaN with negative equity"

    def test_ebitda_without_da(self):
        """EBITDA should equal EBIT when D&A data is missing."""
        import pandas as pd
        import numpy as np
        from analisis_excel.ratio_calculator import RatioCalculator

        financial_data = {
            "balance": {
                "AC": pd.Series({"2023-12-31": 100}),
                "PC": pd.Series({"2023-12-31": 50}),
                "AT": pd.Series({"2023-12-31": 500}),
                "PT": pd.Series({"2023-12-31": 200}),
                "Patr": pd.Series({"2023-12-31": 300}),
                "Efec": pd.Series({"2023-12-31": 10}),
                "Inv": pd.Series({"2023-12-31": 20}),
                "CxC": pd.Series({"2023-12-31": 30}),
                "CxP": pd.Series({"2023-12-31": 40}),
            },
            "income": {
                "Ventas": pd.Series({"2023-12-31": 1000}),
                "COGS": pd.Series({"2023-12-31": 700}),
                "Bruta": pd.Series({"2023-12-31": 300}),
                "EBIT": pd.Series({"2023-12-31": 100}),
                "Neta": pd.Series({"2023-12-31": 50}),
                "Interes": pd.Series({"2023-12-31": -20}),
                "Dep": pd.Series(dtype=float),
                "Amort": pd.Series(dtype=float),
                "DA": pd.Series(dtype=float),  # D&A missing
            },
            "cash_flow": {
                "CFO": pd.Series({"2023-12-31": 80}),
                "CAPEX": pd.Series({"2023-12-31": 30}),
                "FCF": pd.Series({"2023-12-31": 50}),
                "CapexBuy": pd.Series({"2023-12-31": -30}),
            },
            "years": [2023],
        }

        calc = RatioCalculator(financial_data)
        ratios = calc.calculate_profitability_ratios()

        # Margen EBITDA should equal Margen Operativo when D&A is missing
        ebitda_margin = ratios.get("Margen EBITDA", pd.Series(dtype=float))
        ebit_margin = ratios.get("Margen Operativo (EBIT)", pd.Series(dtype=float))

        if not ebitda_margin.empty and not ebit_margin.empty:
            col = "2023-12-31"
            if col in ebitda_margin.index and col in ebit_margin.index:
                assert abs(ebitda_margin[col] - ebit_margin[col]) < 1e-10, (
                    "EBITDA margin should equal EBIT margin when D&A is missing"
                )


class TestEquityEdgeCases:
    """Verify formula_builder handles equity <= 0 correctly."""

    def test_roe_formula_has_equity_guard(self):
        """ROE formulas should include IF(Patrimonio<=0,'N/A',...) guard."""
        fb_dir = Path(__file__).parent.parent / "analisis_excel" / "formula_builder"
        source = "\n".join(p.read_text(encoding="utf-8") for p in sorted(fb_dir.glob("*.py")))

        # Look for the ROE function section
        roe_section = source[source.index("def f_roe"):source.index("def f_roa")]

        assert "<=0" in roe_section, "ROE formula should have <=0 equity guard"
        assert "N/A" in roe_section, "ROE formula should return N/A for negative equity"

    def test_de_formula_has_equity_guard(self):
        """D/E formula should include IF(Patrimonio=0,'N/A',...) guard."""
        fb_dir = Path(__file__).parent.parent / "analisis_excel" / "formula_builder"
        source = "\n".join(p.read_text(encoding="utf-8") for p in sorted(fb_dir.glob("*.py")))

        # Look for the D/E function section
        de_section = source[source.index("def f_de"):source.index("def f_da")]

        assert "=0" in de_section, "D/E formula should have equity=0 guard"


class TestFlagsColumn:
    """Verify unusual value flag formulas exist in the codebase."""

    def test_flags_column_code_exists(self):
        """processor package should have flags column logic for key ratios."""
        proc_dir = Path(__file__).parent.parent / "analisis_excel" / "processor"
        source = "\n".join(p.read_text(encoding="utf-8") for p in sorted(proc_dir.glob("*.py")))

        # Check that flag formulas are present for key ratios
        assert "Margen inusualmente alto" in source or "Unusually high margin" in source
        assert "Alto endeudamiento" in source or "High leverage" in source
        assert "Riesgo de liquidez" in source or "Liquidity risk" in source
        assert "Capital trabajo negativo" in source or "Negative working capital" in source
        assert "Rentabilidad negativa" in source or "Negative profitability" in source
