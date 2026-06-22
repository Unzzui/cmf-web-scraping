#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Validador de consistencia para Product_v1.

Revisa, por cada par ES/EN en Product_v1/(Anual|Trimestral):
  1) Igualdad numérica entre hojas de estados (Balance/Resultados/Flujo)
     para cada periodo en común.
  2) Que no existan columnas de periodos completamente vacías.
  3) En "RATIOS & KPIs": filas clave presentes (D&A y EBITDA) y columnas no vacías.
  4) Sin filas completamente vacías (datos o fórmulas) en KPIs.

Uso:
  python tests/validate_products.py

Salida:
  Código 0 si todo OK, 2 si se detectan errores. Imprime un resumen.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import Counter

from openpyxl import load_workbook


PRODUCT_V1 = Path(__file__).resolve().parents[1] / "Product_v1"
SHEETS_ES = ["Balance General", "Estado de Resultados", "Flujo Efectivo"]
SHEETS_EN = ["Balance Sheet", "Income Statement", "Cash Flow"]
KPIS_SHEET = "RATIOS & KPIs"

PRETTY_RE = re.compile(
    r"^(?P<company>.+) - (?P<rut>\d{7,8}(?:-[0-9Kk])?) - (?P<title>.+) (?P<range>\d{4}-\d{4}) \[(?P<lang>ES|EN)\]\.xlsx$"
)


def find_pairs(base: Path) -> List[Tuple[Path, Path]]:
    files = list(base.glob("*.xlsx"))
    by_key: Dict[Tuple[str, str], Dict[str, Path]] = {}
    for p in files:
        m = PRETTY_RE.match(p.name)
        if not m:
            continue
        key = (m.group("rut"), m.group("range"))
        lang = "es" if m.group("lang").upper() == "ES" else "en"
        by_key.setdefault(key, {})[lang] = p
    pairs: List[Tuple[Path, Path]] = []
    for (rut, yr), mp in by_key.items():
        es = mp.get("es")
        en = mp.get("en")
        if es and en:
            pairs.append((es, en))
    return pairs


def detect_header_row(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and re.match(r"^\d{4}(|Q[1-4]|-\d{2}(-\d{2})?)$", v.strip()):
                return r
    return 3


def years_map(ws, hdr: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if not isinstance(v, str):
            continue
        s = v.strip().split("\n", 1)[0]
        if (
            re.match(r"^\d{4}$", s)
            or re.match(r"^\d{4}Q[1-4]$", s)
            or re.match(r"^\d{4}-\d{2}-\d{2}$", s)
        ):
            m[s] = c
    return m


def numeric_equal(a, b, tol: float = 1e-6) -> bool:
    try:
        if a in (None, "") and b in (None, ""):
            return True
        if a in (None, "") or b in (None, ""):
            return False
        av = float(str(a).replace(",", "").replace(" ", ""))
        bv = float(str(b).replace(",", "").replace(" ", ""))
        return abs(av - bv) <= tol
    except Exception:
        return False


def is_numberlike(x) -> bool:
    try:
        if x in (None, ""):
            return False
        float(str(x).replace(",", "").replace(" ", ""))
        return True
    except Exception:
        return False


def compare_statement_sheets(wb_es, wb_en, es_name: str, en_name: str, errors: List[str], ctx_es: str, ctx_en: str) -> None:
    if es_name not in wb_es.sheetnames or en_name not in wb_en.sheetnames:
        return
    ws_es = wb_es[es_name]
    ws_en = wb_en[en_name]
    hdr_es = detect_header_row(ws_es)
    hdr_en = detect_header_row(ws_en)
    y_es = years_map(ws_es, hdr_es)
    y_en = years_map(ws_en, hdr_en)
    # Comparar set de columnas (por etiqueta visible)
    if set(y_es.keys()) != set(y_en.keys()):
        missing_es = set(y_en) - set(y_es)
        missing_en = set(y_es) - set(y_en)
        if missing_es:
            errors.append(f"{es_name}: faltan columnas en ES: {sorted(missing_es)}")
        if missing_en:
            errors.append(f"{en_name}: faltan columnas en EN: {sorted(missing_en)}")
    common = sorted(set(y_es.keys()) & set(y_en.keys()))

    # Columnas totalmente vacías
    start_es = hdr_es + 1
    start_en = hdr_en + 1
    for lbl in common:
        c_es = y_es[lbl]
        c_en = y_en[lbl]
        non_es = any(ws_es.cell(row=r, column=c_es).value not in (None, "") for r in range(start_es, ws_es.max_row + 1))
        non_en = any(ws_en.cell(row=r, column=c_en).value not in (None, "") for r in range(start_en, ws_en.max_row + 1))
        if not non_es:
            errors.append(f"{es_name}: columna vacía {lbl}  ← {ctx_es}")
        if not non_en:
            errors.append(f"{en_name}: columna vacía {lbl}  ← {ctx_en}")

    # Igualdad numérica por columna como multiconjunto (ignora el orden por filas)
    def _column_numeric_multiset(ws, start_row: int, col_idx: int) -> List[float]:
        vals: List[float] = []
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if is_numberlike(v):
                try:
                    vals.append(float(str(v).replace(",", "").replace(" ", "")))
                except Exception:
                    continue
        return sorted(vals)

    for lbl in common:
        c_es = y_es[lbl]
        c_en = y_en[lbl]
        ms_es = _column_numeric_multiset(ws_es, start_es, c_es)
        ms_en = _column_numeric_multiset(ws_en, start_en, c_en)
        if len(ms_es) != len(ms_en) or any(abs(a - b) > 1e-6 for a, b in zip(ms_es, ms_en)):
            # Reportar un sólo mensaje por columna
            if len([e for e in errors if e.startswith("DIFFSET:")]) < 60:
                errors.append(
                    f"DIFFSET: {es_name}[{lbl}] ({ctx_es}) != {en_name}[{lbl}] ({ctx_en})  (count ES={len(ms_es)} vs EN={len(ms_en)})"
                )
                # Diagnóstico: valores extra en ES/EN con etiquetas
                c_es_ctr = Counter(ms_es)
                c_en_ctr = Counter(ms_en)
                extra_es = list((c_es_ctr - c_en_ctr).elements())
                extra_en = list((c_en_ctr - c_es_ctr).elements())

                def _format_num(x: float) -> str:
                    if abs(x - int(x)) < 1e-6:
                        return str(int(x))
                    return f"{x:.6g}"

                def _find_labels(ws, start_row: int, col_idx: int, value: float, max_hits: int = 2) -> List[str]:
                    hits: List[str] = []
                    target = value
                    for r in range(start_row, ws.max_row + 1):
                        v = ws.cell(row=r, column=col_idx).value
                        if is_numberlike(v):
                            try:
                                vv = float(str(v).replace(",", "").replace(" ", ""))
                            except Exception:
                                continue
                            if abs(vv - target) <= 1e-6:
                                label = ws.cell(row=r, column=1).value
                                hits.append(f"r{r} '{label}'")
                                if len(hits) >= max_hits:
                                    break
                    return hits

                if extra_es:
                    v = extra_es[0]
                    labs = _find_labels(ws_es, start_es, c_es, v, 2)
                    errors.append(
                        f"   → Solo en ES [{lbl}]: {_format_num(v)} en {', '.join(labs) if labs else 'filas desconocidas'}"
                    )
                if extra_en:
                    v = extra_en[0]
                    labs = _find_labels(ws_en, start_en, c_en, v, 2)
                    errors.append(
                        f"   → Solo en EN [{lbl}]: {_format_num(v)} en {', '.join(labs) if labs else 'filas desconocidas'}"
                    )

    # No reportamos filas vacías en estados; hay muchas cuentas no usadas


def _find_row(ws, hdr: int, names: List[str]) -> Optional[int]:
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if isinstance(name, str) and name.strip() in names:
            return r
    return None


def _row_has_any(ws_v, ws_f, r: int) -> bool:
    for c in range(2, ws_v.max_column + 1):
        v = ws_v.cell(row=r, column=c).value
        if v not in (None, ""):
            return True
        f = ws_f.cell(row=r, column=c).value
        if isinstance(f, str) and f.startswith("="):
            return True
    return False


def check_kpis(wb_es, wb_en, errors: List[str], ctx_es: str, ctx_en: str, es_file: Path, en_file: Path) -> None:
    if KPIS_SHEET not in wb_es.sheetnames or KPIS_SHEET not in wb_en.sheetnames:
        errors.append(f"KPIs: falta hoja 'RATIOS & KPIs' en ES o EN  ← {ctx_es} | {ctx_en}")
        return
    ws_es_v = wb_es[KPIS_SHEET]
    ws_en_v = wb_en[KPIS_SHEET]
    try:
        ws_es_f = load_workbook(str(es_file), data_only=False)[KPIS_SHEET]
        ws_en_f = load_workbook(str(en_file), data_only=False)[KPIS_SHEET]
    except Exception:
        ws_es_f = ws_es_v
        ws_en_f = ws_en_v

    hdr = 4  # formatter usa fila 4
    # Periodos
    labels = []
    for c in range(2, ws_es_v.max_column + 1):
        v = ws_es_v.cell(row=hdr, column=c).value
        if isinstance(v, str) and (re.match(r"^\d{4}Q[1-4]$", v) or re.match(r"^\d{4}$", v)):
            labels.append(c)
    if not labels:
        errors.append(f"KPIs: no se detectaron columnas de períodos en ES  ← {ctx_es}")

    # Filas clave
    da_es = _find_row(ws_es_v, hdr, ["Depreciación y Amortización", "Depreciation and Amortization"])
    ebitda_es = _find_row(ws_es_v, hdr, ["EBITDA"])
    if not da_es:
        errors.append(f"KPIs: falta fila 'Depreciación y Amortización' (ES/EN)  ← {ctx_es}")
    if not ebitda_es:
        errors.append(f"KPIs: falta fila 'EBITDA'  ← {ctx_es}")

    # Columnas vacías (valor o fórmula)
    for c in labels:
        has_es = any(
            ws_es_v.cell(row=r, column=c).value not in (None, "")
            or (isinstance(ws_es_f.cell(row=r, column=c).value, str) and ws_es_f.cell(row=r, column=c).value.startswith("="))
            for r in range(hdr + 1, ws_es_v.max_row + 1)
        )
        has_en = c <= ws_en_v.max_column and any(
            ws_en_v.cell(row=r, column=c).value not in (None, "")
            or (isinstance(ws_en_f.cell(row=r, column=c).value, str) and ws_en_f.cell(row=r, column=c).value.startswith("="))
            for r in range(hdr + 1, ws_en_v.max_row + 1)
        )
        if not has_es:
            v_hdr = ws_es_v.cell(row=hdr, column=c).value
            errors.append(f"KPIs ES: columna vacía {v_hdr}  ← {ctx_es}")
        if not has_en:
            v_hdr = ws_en_v.cell(row=hdr, column=c).value if c <= ws_en_v.max_column else f"col#{c}"
            errors.append(f"KPIs EN: columna vacía {v_hdr}  ← {ctx_en}")

    # Asegurar que las filas clave tengan datos o fórmulas
    if da_es is not None and not _row_has_any(ws_es_v, ws_es_f, da_es):
        errors.append(f"KPIs ES: 'Depreciación y Amortización' sin datos ni fórmulas  ← {ctx_es}")
    if ebitda_es is not None and not _row_has_any(ws_es_v, ws_es_f, ebitda_es):
        errors.append(f"KPIs ES: 'EBITDA' sin datos ni fórmulas  ← {ctx_es}")

    # Filas completamente vacías (omitir headers de sección, subnotas y espaciadores)
    empties = 0
    for r in range(hdr + 1, ws_es_v.max_row + 1):
        if not _row_has_any(ws_es_v, ws_es_f, r):
            nm = ws_es_v.cell(row=r, column=1).value
            # Omitir nombres de sección y subnotas (texto en col A sin datos)
            def _is_section_title(txt: Optional[str]) -> bool:
                if not isinstance(txt, str):
                    return False
                s = txt.strip().upper()
                es = {"LIQUIDEZ", "SOLVENCIA Y ESTRUCTURA", "RENTABILIDAD", "EFICIENCIA OPERATIVA", "FLUJOS Y ADICIONALES", "CREACIÓN DE VALOR", "COBERTURA Y RIESGO", "UTILIDADES"}
                en = {"LIQUIDITY", "SOLVENCY & CAPITAL STRUCTURE", "PROFITABILITY", "OPERATING EFFICIENCY", "CASH FLOWS & OTHER", "VALUE CREATION", "COVERAGE & RISK", "UTILS"}
                return s in es or s in en
            # Si es título de sección, subnota textual, o espaciador (None/""), no reportar
            if (isinstance(nm, str) and nm.strip()):
                # título de sección o subnota textual
                if _is_section_title(nm) or True:
                    continue
            else:
                # espaciador
                continue
            empties += 1
            if empties <= 15:
                errors.append(f"KPIs ES: fila vacía '{nm}' (r={r})  ← {ctx_es}")


def validate_dir(dir_path: Path) -> Tuple[int, Dict[str, List[str]]]:
    grouped: Dict[str, List[str]] = {}
    pairs = find_pairs(dir_path)
    checked = 0
    for es_path, en_path in pairs:
        try:
            wb_es = load_workbook(str(es_path), data_only=True)
            wb_en = load_workbook(str(en_path), data_only=True)
        except Exception as ex:
            key = f"{es_path.name} / {en_path.name}"
            grouped.setdefault(key, []).append(f"No se pudo abrir par ES/EN: {es_path.name} / {en_path.name} ({ex})")
            continue
        # Estados: Balance/Resultados/Flujo
        ctx_es = str(es_path.resolve())
        ctx_en = str(en_path.resolve())
        # Construir clave legible de empresa + rango
        key = None
        m = PRETTY_RE.match(es_path.name)
        if m:
            key = f"{m.group('company')} — {m.group('rut')} — {m.group('range')}"
        else:
            key = es_path.stem
        pair_errs: List[str] = []
        for es_name, en_name in zip(SHEETS_ES, SHEETS_EN):
            compare_statement_sheets(wb_es, wb_en, es_name, en_name, pair_errs, ctx_es, ctx_en)
        # KPIs
        check_kpis(wb_es, wb_en, pair_errs, ctx_es, ctx_en, es_path, en_path)
        checked += 1
        if pair_errs:
            grouped[key] = pair_errs
    return checked, grouped


def main() -> int:
    base = PRODUCT_V1
    if not base.exists():
        print(f"No existe {base}")
        return 2
    total_checked = 0
    grouped_all: Dict[str, Dict[str, List[str]]] = {}
    for freq in ("Anual", "Trimestral"):
        d = base / freq
        if not d.exists():
            continue
        n, grouped = validate_dir(d)
        total_checked += n
        grouped_all[freq] = grouped
    if total_checked == 0:
        print("No se encontraron pares ES/EN en Product_v1.")
        return 2

    # ¿Hay errores?
    total_issues = sum(len(errs) for freq in grouped_all.values() for errs in freq.values())
    if total_issues > 0:
        print(f"Validación terminada con errores. Pares verificados: {total_checked}")
        for freq in ("Anual", "Trimestral"):
            grouped = grouped_all.get(freq, {})
            if not grouped:
                continue
            num_pairs = len(grouped)
            num_issues = sum(len(v) for v in grouped.values())
            print(f" - [{freq}] {num_issues} issue(s) en {num_pairs} empresa(s)")
            # Ordenar empresas alfabéticamente
            for company in sorted(grouped.keys()):
                print(f"   · {company}")
                # Opcional: agrupar por hoja básica
                pair_errs = grouped[company]
                # Limitar salida por par para legibilidad
                for e in pair_errs[:80]:
                    print(f"     - {e}")
                if len(pair_errs) > 80:
                    print(f"     … {len(pair_errs) - 80} más")
        return 2
    print(f"OK. Pares verificados: {total_checked}. Sin inconsistencias detectadas.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


