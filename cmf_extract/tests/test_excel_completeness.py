"""Verificador de completitud de los Excel finales de análisis.

Para cada empresa en ``Product_v1/Total/`` revisa:

  * Que existan las hojas Balance General, Estado de Resultados, Flujo Efectivo.
  * Cobertura por período (% de celdas no vacías) por hoja.
  * Que items críticos del Balance estén poblados para cada período con data:
        Total de activos, Total de pasivos, Patrimonio total,
        Activos corrientes totales, Total de activos no corrientes,
        Pasivos corrientes totales, Total de pasivos no corrientes,
        Total de patrimonio y pasivos
  * Que la identidad contable cuadre:
        Activos corrientes totales + Total de activos no corrientes = Total de activos
        Total de pasivos + Patrimonio total = Total de patrimonio y pasivos
  * Que items críticos del Cash Flow no estén ausentes:
        Efectivo al principio del periodo, Efectivo al final del periodo,
        Incremento (disminución) neto de efectivo

Uso:
    .venv/bin/python -m cmf_extract.tests.test_excel_completeness
    .venv/bin/python -m cmf_extract.tests.test_excel_completeness --rut 93007000-9
    .venv/bin/python -m cmf_extract.tests.test_excel_completeness --product-v1-dir cmf_extract/Product_v1/Total

Reporta a stdout una tabla resumen y opcionalmente un CSV detallado.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


# --- Items críticos por hoja (deben estar siempre que haya data en el período) ---

CRITICAL_BALANCE = [
    "Total de activos",
    "Total de pasivos",
    "Patrimonio total",
    "Activos corrientes totales",
    "Total de activos no corrientes",
    "Pasivos corrientes totales",
    "Total de pasivos no corrientes",
    "Total de patrimonio y pasivos",
]

CRITICAL_INCOME = [
    "Ingresos de actividades ordinarias",
    "Ganancia bruta",
    "Ganancia (pérdida)",
]

CRITICAL_CASH_FLOW = [
    "Efectivo y equivalentes al efectivo al principio del periodo",
    "Efectivo y equivalentes al efectivo al final del periodo",
    "Incremento (disminución) neto de efectivo y equivalentes al efectivo",
]


# --- Tolerancia para identidades contables ---
# La tolerancia efectiva es max(IDENTITY_ABS_TOLERANCE, IDENTITY_REL_TOLERANCE * |total|).
# Esto absorbe redondeos en empresas pequeñas y permite restatements/ajustes
# históricos pequeños en empresas grandes sin generar ruido.
IDENTITY_ABS_TOLERANCE = 100        # 100 mil CLP
IDENTITY_REL_TOLERANCE = 0.005      # 0.5% del Total de activos


_PERIOD_RE = re.compile(r"^\d{4}(?:Q[1-4])?$")


# ---------------------------------------------------------------------------
# Dataclasses para reporte
# ---------------------------------------------------------------------------


@dataclass
class SheetReport:
    sheet_name: str
    found: bool = False
    total_rows: int = 0
    period_columns: list[str] = field(default_factory=list)
    cells_populated_per_period: dict[str, int] = field(default_factory=dict)
    missing_critical_per_period: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class IdentityCheck:
    period: str
    name: str  # e.g. "Activos = Corrientes + No Corrientes"
    ok: bool
    delta: Optional[float] = None  # diff si falla
    detail: str = ""


@dataclass
class CompanyReport:
    rut: str
    file: Path
    company_name: str = ""
    sheets: dict[str, SheetReport] = field(default_factory=dict)
    identity_failures: list[IdentityCheck] = field(default_factory=list)

    @property
    def has_problems(self) -> bool:
        if any(missing for sr in self.sheets.values()
               for missing in sr.missing_critical_per_period.values()):
            return True
        if self.identity_failures:
            return True
        return False


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def _detect_header_row(ws) -> int:
    """Devuelve la fila que contiene 'Cuenta' o las etiquetas YYYY/YYYYQn."""
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto"):
            return r
    return 3  # default usado por primary_csv_to_excel


def _row_map(ws, header_row: int) -> dict[str, int]:
    """Mapa Cuenta (exact, lowercase) → row number."""
    out: dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            key = v.strip().lower()
            if key and key not in out:
                out[key] = r
    return out


def _period_cols(ws, header_row: int) -> list[tuple[int, str]]:
    """Lista de (col_index, period_label) para columnas como '2024Q1'/'2024'."""
    out: list[tuple[int, str]] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and _PERIOD_RE.match(v.strip()):
            out.append((c, v.strip()))
    return out


def _to_num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------


def check_sheet(ws, sheet_name: str, critical_labels: list[str]) -> SheetReport:
    sr = SheetReport(sheet_name=sheet_name, found=True)
    sr.total_rows = max(0, ws.max_row - 3)
    hdr = _detect_header_row(ws)
    cols = _period_cols(ws, hdr)
    sr.period_columns = [lbl for _, lbl in cols]
    rmap = _row_map(ws, hdr)

    crit_keys = {lbl.lower(): lbl for lbl in critical_labels}

    for c_idx, lbl in cols:
        populated = 0
        for r in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=c_idx).value
            if v not in (None, "") and v != 0:
                populated += 1
        sr.cells_populated_per_period[lbl] = populated

        # Solo chequear críticos si el período tiene data sustancial.
        if populated >= 3:
            missing: list[str] = []
            for ck, label_orig in crit_keys.items():
                r = rmap.get(ck)
                if r is None:
                    missing.append(f"{label_orig} (fila no existe)")
                else:
                    v = ws.cell(row=r, column=c_idx).value
                    if v in (None, "", 0):
                        missing.append(label_orig)
            if missing:
                sr.missing_critical_per_period[lbl] = missing
    return sr


def check_balance_identities(ws) -> list[IdentityCheck]:
    """Verifica:
       Total activos ≈ Activos corrientes totales + Total de activos no corrientes
       Total patrimonio y pasivos ≈ Total de pasivos + Patrimonio total
    """
    hdr = _detect_header_row(ws)
    rmap = _row_map(ws, hdr)
    cols = _period_cols(ws, hdr)

    def _val(row_label: str, c_idx: int) -> Optional[float]:
        r = rmap.get(row_label.lower())
        if r is None:
            return None
        return _to_num(ws.cell(row=r, column=c_idx).value)

    failures: list[IdentityCheck] = []
    for c_idx, period in cols:
        total_act = _val("Total de activos", c_idx)
        ac = _val("Activos corrientes totales", c_idx)
        anc = _val("Total de activos no corrientes", c_idx)
        if all(v is not None for v in (total_act, ac, anc)):
            diff = (ac + anc) - total_act
            tol = max(IDENTITY_ABS_TOLERANCE, IDENTITY_REL_TOLERANCE * abs(total_act))
            if abs(diff) > tol:
                failures.append(IdentityCheck(
                    period=period,
                    name="Activos = Corrientes + No Corrientes",
                    ok=False, delta=diff,
                    detail=f"{ac:,.0f} + {anc:,.0f} = {ac+anc:,.0f} ≠ {total_act:,.0f}"
                           f"  (diff {diff:+,.0f}, tol ±{tol:,.0f})",
                ))

        tot_pas_pat = _val("Total de patrimonio y pasivos", c_idx)
        tot_pas = _val("Total de pasivos", c_idx)
        pat = _val("Patrimonio total", c_idx)
        if all(v is not None for v in (tot_pas_pat, tot_pas, pat)):
            diff = (tot_pas + pat) - tot_pas_pat
            tol = max(IDENTITY_ABS_TOLERANCE, IDENTITY_REL_TOLERANCE * abs(tot_pas_pat))
            if abs(diff) > tol:
                failures.append(IdentityCheck(
                    period=period,
                    name="Patrimonio+Pasivos = Total patrimonio y pasivos",
                    ok=False, delta=diff,
                    detail=f"{tot_pas:,.0f} + {pat:,.0f} = {tot_pas+pat:,.0f} ≠ {tot_pas_pat:,.0f}"
                           f"  (diff {diff:+,.0f}, tol ±{tol:,.0f})",
                ))
    return failures


# ---------------------------------------------------------------------------
# Per-company check
# ---------------------------------------------------------------------------


_RUT_RE = re.compile(r"(\d{7,8}-[\dkK])")
_SHEET_CHECKS = [
    ("Balance General", CRITICAL_BALANCE),
    ("Estado de Resultados", CRITICAL_INCOME),
    ("Flujo Efectivo", CRITICAL_CASH_FLOW),
]


def check_excel(path: Path) -> CompanyReport:
    rut_match = _RUT_RE.search(path.name)
    rut = rut_match.group(1) if rut_match else path.stem
    rep = CompanyReport(rut=rut, file=path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        rep.sheets["__error__"] = SheetReport("__error__", found=False)
        rep.identity_failures.append(IdentityCheck("-", "load_workbook", False,
                                                   detail=str(e)))
        return rep

    parts = path.stem.split(" - ", 2)
    rep.company_name = parts[0].strip() if parts else path.stem

    for sheet_name, critical in _SHEET_CHECKS:
        if sheet_name not in wb.sheetnames:
            rep.sheets[sheet_name] = SheetReport(sheet_name=sheet_name, found=False)
            continue
        ws = wb[sheet_name]
        rep.sheets[sheet_name] = check_sheet(ws, sheet_name, critical)

    if "Balance General" in wb.sheetnames:
        rep.identity_failures.extend(check_balance_identities(wb["Balance General"]))
    return rep


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------


def _print_company_summary(rep: CompanyReport) -> None:
    status = "❌" if rep.has_problems else "✅"
    print(f"\n{status} {rep.rut} — {rep.company_name}")
    for sheet_name, sr in rep.sheets.items():
        if not sr.found:
            print(f"  ✗ Hoja '{sheet_name}' NO encontrada")
            continue
        periods = len(sr.period_columns)
        avg_pop = (sum(sr.cells_populated_per_period.values())
                   / max(1, periods))
        missing_count = sum(len(v) for v in sr.missing_critical_per_period.values())
        marker = "✓" if missing_count == 0 else "⚠"
        print(f"  {marker} {sheet_name}: {periods} períodos, "
              f"avg {avg_pop:.0f} cells/período, "
              f"{missing_count} items críticos faltantes")
        for period, missing in sr.missing_critical_per_period.items():
            print(f"      {period} faltan: {', '.join(missing)[:120]}")
    if rep.identity_failures:
        print(f"  ✗ {len(rep.identity_failures)} identidades contables falladas:")
        for f in rep.identity_failures[:5]:
            print(f"      {f.period} {f.name}: {f.detail}")


def _build_missing_overrides_json(reports: list[CompanyReport],
                                  existing: dict) -> dict:
    """Genera/actualiza un JSON de overrides con las celdas críticas faltantes.

    Estructura:
        {RUT: {Sheet: {Label: {Period: null}}}}

    * Solo incluye los items críticos faltantes detectados.
    * Si ``existing`` tiene valores ya llenados, los preserva.
    * Si una celda que estaba en el template ya fue rellenada por el pipeline,
      se REMUEVE del JSON (ya no es un hueco).
    """
    # Deep clone para no mutar el existente.
    merged = json.loads(json.dumps(existing or {}))
    # Construir set de huecos actuales: (rut, sheet, label, period)
    current_gaps: set[tuple[str, str, str, str]] = set()
    for rep in reports:
        for sheet_name, sr in rep.sheets.items():
            for period, missing_labels in sr.missing_critical_per_period.items():
                for label in missing_labels:
                    # Algunos labels vienen con " (fila no existe)" como sufijo
                    label_clean = label.split(" (fila no existe)")[0].strip()
                    current_gaps.add((rep.rut, sheet_name, label_clean, period))

    # Añadir gaps nuevos al merged (como null si no estaban).
    for rut, sheet, label, period in current_gaps:
        merged.setdefault(rut, {}).setdefault(sheet, {}).setdefault(label, {})
        if period not in merged[rut][sheet][label]:
            merged[rut][sheet][label][period] = None

    # Limpiar: si una entrada estaba pero ya NO es un gap, sacarla.
    # (solo limpia las que están con valor null; si el usuario puso un número,
    # se mantiene aunque el gap ya no exista — útil para forzar overrides).
    for rut in list(merged.keys()):
        for sheet in list(merged[rut].keys()):
            for label in list(merged[rut][sheet].keys()):
                periods = merged[rut][sheet][label]
                if not isinstance(periods, dict):
                    continue
                for period in list(periods.keys()):
                    if period == "_force":
                        continue
                    if periods[period] is None:
                        if (rut, sheet, label, period) not in current_gaps:
                            del periods[period]
                if not periods:
                    del merged[rut][sheet][label]
            if not merged[rut][sheet]:
                del merged[rut][sheet]
        if not merged[rut]:
            del merged[rut]
    return merged


def _write_overrides_json(reports: list[CompanyReport], out_path: Path) -> None:
    """Escribe/actualiza el JSON de manual_overrides preservando valores existentes."""
    existing = {}
    if out_path.is_file():
        try:
            existing = json.loads(out_path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"⚠ No se pudo leer {out_path}: {e}; se sobrescribirá")
    merged = _build_missing_overrides_json(reports, existing)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(merged, indent=2, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )
    n_gaps = sum(
        1 for rut in merged.values() for sheet in rut.values()
        for label in sheet.values() for period, v in label.items()
        if period != "_force" and v is None
    )
    n_filled = sum(
        1 for rut in merged.values() for sheet in rut.values()
        for label in sheet.values() for period, v in label.items()
        if period != "_force" and v is not None
    )
    print(f"\n📝 manual_overrides actualizado: {out_path}")
    print(f"   {n_gaps} huecos pendientes de llenar (valor null)")
    print(f"   {n_filled} valores ya llenados (se aplicarán al re-correr)")


def _write_csv_report(reports: list[CompanyReport], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "rut", "company", "sheet", "period", "cells_populated",
            "missing_critical", "identity_failed",
        ])
        for rep in reports:
            for sn, sr in rep.sheets.items():
                if not sr.found:
                    w.writerow([rep.rut, rep.company_name, sn, "-", 0,
                                "SHEET_NOT_FOUND", ""])
                    continue
                for period, pop in sr.cells_populated_per_period.items():
                    missing = sr.missing_critical_per_period.get(period, [])
                    id_fail = "; ".join(
                        f.detail for f in rep.identity_failures
                        if f.period == period
                    )
                    w.writerow([
                        rep.rut, rep.company_name, sn, period, pop,
                        " | ".join(missing), id_fail,
                    ])
    print(f"\n📊 CSV detallado: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--product-v1-dir", default="cmf_extract/Product_v1/Total",
                    help="Directorio con los Excel finales de análisis.")
    ap.add_argument("--rut", default=None,
                    help="Verificar solo un RUT específico (ej. '93007000-9').")
    ap.add_argument("--csv", default=None,
                    help="Si se da, escribe un CSV detallado en esa ruta.")
    ap.add_argument("--export-missing", nargs="?", const="cmf_extract/manual_overrides.json",
                    default=None, metavar="PATH",
                    help="Genera/actualiza un JSON con los items críticos "
                         "faltantes. Si el archivo existe, preserva valores "
                         "ya llenados. Default path: cmf_extract/manual_overrides.json")
    ap.add_argument("--quiet", action="store_true",
                    help="Solo mostrar empresas con problemas.")
    args = ap.parse_args()

    d = Path(args.product_v1_dir)
    if not d.is_dir():
        print(f"Directorio no existe: {d}")
        sys.exit(2)

    pattern = "*.xlsx"
    files = sorted(p for p in d.glob(pattern)
                   if not p.name.startswith("~$"))
    if args.rut:
        files = [p for p in files if args.rut in p.name]
        if not files:
            print(f"No se encontró Excel para RUT {args.rut} en {d}")
            sys.exit(2)

    print(f"Verificando {len(files)} archivo(s) en {d}...\n")
    reports: list[CompanyReport] = []
    for f in files:
        rep = check_excel(f)
        reports.append(rep)
        if args.quiet and not rep.has_problems:
            continue
        _print_company_summary(rep)

    # Resumen global
    n_ok = sum(1 for r in reports if not r.has_problems)
    n_pb = len(reports) - n_ok
    print("\n" + "═" * 70)
    print(f"RESUMEN: {n_ok}/{len(reports)} sin problemas, {n_pb} con problemas")
    print("═" * 70)

    if args.csv:
        _write_csv_report(reports, Path(args.csv))

    if args.export_missing:
        _write_overrides_json(reports, Path(args.export_missing))

    sys.exit(1 if n_pb > 0 else 0)


if __name__ == "__main__":
    main()
