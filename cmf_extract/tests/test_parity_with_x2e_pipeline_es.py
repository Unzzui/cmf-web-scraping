#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import pytest


def _collect_final_excels() -> list[Path]:
    base = Path(os.getenv('CMF_PRODUCT_V1_DIR', 'Product_v1/Total'))
    if not base.exists():
        return []
    pattern = os.getenv('CMF_PYTEST_GLOB') or '*.xlsx'
    files = sorted(base.glob(pattern))
    # Prefer ES only
    es = [p for p in files if (' [ES]' in p.name) or p.name.endswith('_es.xlsx') or 'ES' in p.name]
    return es or files


def _parse_final_name(name: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    # Pretty pattern: "<Empresa> - <RUT> - ... <RANGE> [ES].xlsx"
    m = re.search(r" - (\d+-?[\dKk]) - .* (\d{4}(?:-\d{4}|-\d{4}Q[1-4])) \[(ES|EN)\]", name)
    if not m:
        return None, None, None
    rut, rng, lang = m.groups()
    return rut, rng, lang.lower()


def _period_sort_key(lbl: str) -> tuple[int, int]:
    s = str(lbl)
    mq = re.match(r"^(\d{4})Q([1-4])$", s)
    if mq:
        return (int(mq.group(1)), int(mq.group(2)))
    my = re.match(r"^(\d{4})$", s)
    if my:
        return (int(my.group(1)), 4)
    return (9999, 9)


def _excel_periods(ws) -> list[str]:
    # Header on row 3 per our writer
    hdr = 3
    labels = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if not isinstance(v, str):
            continue
        s = v.strip().split("\n", 1)[0]
        if re.match(r"^\d{4}Q[1-4]$", s):
            labels.append(s)
        elif re.match(r"^\d{4}$", s):
            # Normalize bare YYYY to YYYYQ4
            labels.append(f"{s}Q4")
    # unique and sorted desc
    uniq = []
    seen = set()
    for lb in labels:
        if lb not in seen:
            seen.add(lb)
            uniq.append(lb)
    return sorted(uniq, key=_period_sort_key, reverse=True)


def _find_out_consolidated_for_rut(rut: str, lang: str) -> Optional[Tuple[Path, str]]:
    base = Path('data/XBRL/Total')
    if not base.exists():
        return None
    rut_plain = rut.split('-')[0]
    candidates: list[Tuple[Path, str]] = []
    for company in base.iterdir():
        if not company.is_dir():
            continue
        if not company.name.startswith(rut_plain + '_') and not company.name.startswith(rut_plain):
            continue
        for p in company.glob('out_consolidated_*'):
            # detect stem via facts/pres presence
            # facts_{rut}_{ymrange}_{lang}.csv
            for facts_file in p.glob(f'facts_{rut_plain}_*_{lang}.csv'):
                stem = facts_file.name.removeprefix('facts_').removesuffix(f'_{lang}.csv')
                return (p, stem)
            # if no facts, try using estados filename
            for est in p.glob(f'estados_{rut_plain}_*_{lang}.xlsx'):
                stem = est.name.removeprefix('estados_').removesuffix(f'_{lang}.xlsx')
                return (p, stem)
    return None


EXCEL_FILES = _collect_final_excels()


@pytest.mark.skipif(not EXCEL_FILES, reason='No final Excel files found in Product_v1/Total')
@pytest.mark.parametrize('final_path', EXCEL_FILES[: int(os.getenv('CMF_TEST_MAX_FILES', '4'))])
def test_parity_with_x2e_pipeline(final_path: Path):
    # Parse final name
    from openpyxl import load_workbook
    rut, rng, lang = _parse_final_name(final_path.name)
    assert rut and lang, f'Nombre no parseable: {final_path.name}'

    # Locate the out_consolidated dir and stem used by x2e
    found = _find_out_consolidated_for_rut(rut, lang)
    assert found, f'No se halló out_consolidated para {rut} ({final_path.name})'
    out_dir, stem = found

    # Compose statements via xbrl_to_excel internals
    import importlib
    x2e = importlib.import_module('xbrl_to_excel')
    facts, pres = x2e.load_inputs(out_dir, stem, lang)
    tree = x2e.build_tree_and_order(pres)

    # Read final workbook
    wb = load_workbook(str(final_path), read_only=True, data_only=True)
    sheet_map = {
        'BALANCE': ['Balance General', 'Balance Sheet'],
        'RESULTADOS': ['Estado de Resultados', 'Income Statement'],
        'FLUJO': ['Flujo Efectivo', 'Cash Flow'],
    }

    for kind, names in sheet_map.items():
        name = next((n for n in names if n in wb.sheetnames), None)
        if not name:
            continue
        ws = wb[name]
        periods = _excel_periods(ws)
        # Build DF via compose_statement using same periods by restricting allowed range
        # Derive allowed_months from first/last label
        def _to_ym(lbl: str) -> Optional[str]:
            m = re.match(r'^(\d{4})Q([1-4])$', lbl)
            if m:
                y, q = int(m.group(1)), int(m.group(2))
                mo = {1:'03',2:'06',3:'09',4:'12'}[q]
                return f'{y}-{mo}'
            m2 = re.match(r'^(\d{4})$', lbl)
            if m2:
                return f"{m2.group(1)}-12"
            return None
        lo = _to_ym(periods[-1]) if periods else None
        hi = _to_ym(periods[0]) if periods else None
        allowed = (lo, hi) if (lo and hi) else None

        df = x2e.compose_statement(
            facts,
            x2e.select_role_tree(tree, kind),
            lang=lang,
            other_facts_raw=None,
            other_lang=None,
            max_dates=None,
            statement_kind=kind,
            allowed_months=allowed,
            presentation_data=pres,
            output_dir=out_dir,
        )
        assert not df.empty, f'DF vacío para {kind} ({final_path.name})'

        # Build map from Cuenta -> row values for compared periods
        headers = [str(c) for c in df.columns]
        period_idx = [headers.index(p) for p in periods if p in headers]
        assert period_idx, f'Sin columnas comunes para {kind} ({final_path.name})'
        values_by_name = {}
        for _, row in df.iterrows():
            nm = str(row['Cuenta'])
            vals = [row[headers[i]] for i in period_idx]
            values_by_name[nm] = vals

        # Compare with workbook values (must equal or both blank)
        hdr_row = 3
        # Build col index in ws for periods
        col_for = {}
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=hdr_row, column=c).value
            if isinstance(v, str):
                s = v.strip().split('\n', 1)[0]
                if s in periods:
                    col_for[s] = c
        mismatches = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            nm = ws.cell(row=r, column=1).value
            if not isinstance(nm, str) or not nm.strip():
                continue
            nm = nm.strip()
            if nm not in values_by_name:
                continue
            df_vals = values_by_name[nm]
            # Compare across periods
            for j, p in enumerate(periods):
                c = col_for.get(p)
                if not c:
                    continue
                v_ws = ws.cell(row=r, column=c).value
                v_df = df_vals[j] if j < len(df_vals) else None
                # Be strict: either equal numbers or both None/blank
                if pd.isna(v_df) or v_df in (None, ''):
                    if v_ws not in (None, ''):
                        mismatches.append((kind, nm, p, v_ws, v_df))
                else:
                    try:
                        v_df_f = float(str(v_df).replace(',', ''))
                        if not isinstance(v_ws, (int, float)) or abs(float(v_ws) - v_df_f) > 1e-6:
                            mismatches.append((kind, nm, p, v_ws, v_df))
                    except Exception:
                        mismatches.append((kind, nm, p, v_ws, v_df))

        assert not mismatches, (
            f"Diferencias con pipeline x2e en {final_path.name}:\n" +
            "\n".join(f"[{k}] {nm} @ {per}: excel={vws} vs x2e={vdf}" for (k,nm,per,vws,vdf) in mismatches[:80])
        )

