#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Validador ES-only de Excel final (Product_v1/*):
- Verifica encabezados de períodos (YYYY/YYYYQn), unicidad y orden natural
- Valida tipos numéricos en columnas de períodos (no strings tipo '#REF!')
- Asegura que filas de categoría ([sinopsis]/[abstract]/[resumen]) no tengan números
- Reglas contables básicas:
  * Balance: Total de activos == Total patrimonio y pasivos (tolerancia configurable)
  * Flujo: efectivo final ~= efectivo inicial + netos (operación+inversión+financiamiento) [+ FX si existe]

Uso:
  python tests/validate_final_excel_es.py [Product_v1_dir]  # default: Product_v1/Total

Retorna código 0 si todo OK; 2 si hay errores.
"""

from __future__ import annotations

import sys
import re
from pathlib import Path
from typing import List, Tuple, Optional
from openpyxl import load_workbook


PERIOD_RE_YEAR = re.compile(r"^\d{4}$")
PERIOD_RE_Q = re.compile(r"^(\d{4})Q([1-4])$")


def _detect_header_row(ws) -> int:
    for r in range(1, min(12, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 40) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and (PERIOD_RE_YEAR.match(v.strip().split("\n",1)[0]) or PERIOD_RE_Q.match(v.strip().split("\n",1)[0])):
                return r
    return 3


def _norm_hdr(v: str) -> Optional[str]:
    if not isinstance(v, str):
        return None
    s = v.strip().split("\n", 1)[0]
    if PERIOD_RE_Q.match(s):
        return s
    # Bare YYYY -> YYYYQ4 for backward compat
    if PERIOD_RE_YEAR.match(s):
        return f"{s}Q4"
    return None


def _period_sort_key(lbl: str) -> Tuple[int, int]:
    s = str(lbl)
    m = PERIOD_RE_Q.match(s)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    if PERIOD_RE_YEAR.match(s):
        return (int(s), 4)
    return (9999, 9)


def _sheet_periods(ws, hdr_row: int) -> List[str]:
    labels: List[str] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        lb = _norm_hdr(v)
        if lb:
            labels.append(lb)
    # Unicidad preservando primera ocurrencia
    seen = set()
    uniq = []
    for lb in labels:
        if lb not in seen:
            seen.add(lb)
            uniq.append(lb)
    return uniq


def _is_category(name: str) -> bool:
    if not isinstance(name, str):
        return False
    s = name.strip().lower()
    if s.startswith('[') and ']' in s:
        return True
    return any(tag in s for tag in ('[sinopsis]', '[abstract]', '[resumen]'))


def _numeric_or_blank(v) -> bool:
    if v in (None, ""):
        return True
    if isinstance(v, (int, float)):
        return True
    # Rechazar strings con errores Excel
    if isinstance(v, str) and v.strip().startswith('#'):
        return False
    # Strings que representan número (no deberían existir): rechazarlas para ser estrictos
    return False


def validate_workbook(path: Path, tol: float = 3000000.0) -> List[str]:
    errors: List[str] = []
    try:
        wb = load_workbook(str(path), read_only=True, data_only=False)
    except Exception as e:
        return [f"No se pudo abrir: {path.name} — {e}"]

    sheets_es = {
        'balance': 'Balance General',
        'resultados': 'Estado de Resultados',
        'flujo': 'Flujo Efectivo',
    }

    sheet_objs = {}
    hdr_rows = {}
    periods = {}

    for key, nm in sheets_es.items():
        if nm in wb.sheetnames:
            ws = wb[nm]
            sheet_objs[key] = ws
            hdr = _detect_header_row(ws)
            hdr_rows[key] = hdr
            periods[key] = _sheet_periods(ws, hdr)
            # Validar unicidad y orden
            p = periods[key]
            if p != sorted(p, key=_period_sort_key, reverse=True):
                errors.append(f"{path.name} [{nm}]: períodos no están en orden natural desc: {p}")
            if len(p) != len(set(p)):
                errors.append(f"{path.name} [{nm}]: períodos con duplicados: {p}")
        else:
            errors.append(f"{path.name}: falta hoja '{nm}'")

    # Validar consistencia entre hojas (conjunto de períodos intersecta y no vacío)
    try:
        keys_present = [k for k in ('balance','resultados','flujo') if k in periods]
        if len(keys_present) >= 2:
            inter = set(periods[keys_present[0]])
            for k in keys_present[1:]:
                inter &= set(periods[k])
            if not inter:
                errors.append(f"{path.name}: conjunto de períodos entre hojas no intersecta (posible desalineación)")
    except Exception:
        pass

    # Validar tipos numéricos por hoja y categorías
    for key, ws in sheet_objs.items():
        hdr = hdr_rows[key]
        # Columnas de período existentes en la hoja
        pcols = periods[key]
        col_index = {}
        for c in range(2, ws.max_column + 1):
            v = _norm_hdr(ws.cell(row=hdr, column=c).value)
            if v in pcols:
                col_index[v] = c
        # Recorrer filas de datos
        for r in range(hdr + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            is_cat = _is_category(name or '')
            for lb in pcols:
                c = col_index.get(lb)
                if not c:
                    continue
                v = ws.cell(row=r, column=c).value
                if is_cat:
                    # Categorías no deben tener números
                    if isinstance(v, (int, float)) and abs(float(v)) > 0:
                        errors.append(f"{path.name} [{ws.title}] fila {r}: categoría con valor en {lb}: {v}")
                else:
                    if not _numeric_or_blank(v):
                        errors.append(f"{path.name} [{ws.title}] fila {r}: valor no numérico en {lb}: {v}")

    # Reglas contables básicas
    # 1) Balance: Total de activos vs Total de patrimonio y pasivos
    if 'balance' in sheet_objs:
        ws = sheet_objs['balance']
        hdr = hdr_rows['balance']
        pcols = periods['balance']
        # Buscar filas por nombre
        row_tot_act = None
        row_tot_pyp = None
        for r in range(hdr + 1, ws.max_row + 1):
            nm = str(ws.cell(row=r, column=1).value or '').strip().lower()
            if nm == 'total de activos' and row_tot_act is None:
                row_tot_act = r
            if nm == 'total de patrimonio y pasivos' and row_tot_pyp is None:
                row_tot_pyp = r
        if row_tot_act and row_tot_pyp:
            for lb in pcols:
                c = None
                # localizar índice de columna lb
                for cc in range(2, ws.max_column + 1):
                    if _norm_hdr(ws.cell(row=hdr, column=cc).value) == lb:
                        c = cc
                        break
                if not c:
                    continue
                va = ws.cell(row=row_tot_act, column=c).value
                vp = ws.cell(row=row_tot_pyp, column=c).value
                if isinstance(va, (int, float)) and isinstance(vp, (int, float)):
                    if abs(float(va) - float(vp)) > tol:
                        errors.append(f"{path.name} [Balance] {lb}: Total de activos ({va}) != Total de patrimonio y pasivos ({vp})")

    # 2) Flujo: conciliación de efectivo
    if 'flujo' in sheet_objs:
        ws = sheet_objs['flujo']
        hdr = hdr_rows['flujo']
        pcols = periods['flujo']
        rows = {
            'ini': None,
            'fin': None,
            'op': None,
            'inv': None,
            'finan': None,
            'fx': None,
        }
        for r in range(hdr + 1, ws.max_row + 1):
            nm_raw = str(ws.cell(row=r, column=1).value or '').strip().lower()
            nm = nm_raw.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
            if 'efectivo y equivalentes al efectivo al principio' in nm and rows['ini'] is None:
                rows['ini'] = r
            if 'efectivo y equivalentes al efectivo al final' in nm and rows['fin'] is None:
                rows['fin'] = r
            if 'netos procedentes de (utilizados en) actividades de oper' in nm and rows['op'] is None:
                rows['op'] = r
            if 'netos procedentes de (utilizados en) actividades de inversion' in nm and rows['inv'] is None:
                rows['inv'] = r
            if 'netos procedentes de (utilizados en) actividades de financiacion' in nm and rows['finan'] is None:
                rows['finan'] = r
            if 'efectos de la variacion en la tasa de cambio' in nm and rows['fx'] is None:
                rows['fx'] = r
        if rows['ini'] and rows['fin'] and rows['op'] and rows['inv'] and rows['finan']:
            for lb in pcols:
                # localizar índice de columna lb
                c = None
                for cc in range(2, ws.max_column + 1):
                    if _norm_hdr(ws.cell(row=hdr, column=cc).value) == lb:
                        c = cc
                        break
                if not c:
                    continue
                v_ini = ws.cell(row=rows['ini'], column=c).value
                v_fin = ws.cell(row=rows['fin'], column=c).value
                v_op = ws.cell(row=rows['op'], column=c).value
                v_inv = ws.cell(row=rows['inv'], column=c).value
                v_fi = ws.cell(row=rows['finan'], column=c).value
                v_fx = ws.cell(row=rows['fx'], column=c).value if rows['fx'] else 0
                vals = (v_ini, v_fin, v_op, v_inv, v_fi, v_fx)
                if all(isinstance(x, (int, float)) for x in vals[:5]):
                    fx_value = float(v_fx) if isinstance(v_fx,(int,float)) else 0.0
                    expected_fin = float(v_ini) + float(v_op) + float(v_inv) + float(v_fi) + fx_value
                    if abs(expected_fin - float(v_fin)) > tol:
                        errors.append(f"{path.name} [Flujo] {lb}: final {v_fin} != inicio {v_ini} + flujos ({v_op}+{v_inv}+{v_fi}+{fx_value})")

    # 3) RATIOS & KPIs: verificar filas específicas
    if 'RATIOS & KPIs' in wb.sheetnames:
        ws = wb['RATIOS & KPIs']
        hdr = _detect_header_row(ws)
        pcols = _sheet_periods(ws, hdr)
        
        required_rows = {
            'dep_amort': 'Depreciación y Amortización',
            'acciones': 'Total número de acciones emitidas'
        }
        found_rows = {key: None for key in required_rows}

        for r in range(hdr + 1, ws.max_row + 1):
            nm = str(ws.cell(row=r, column=1).value or '').strip().lower()
            for key, text in required_rows.items():
                if text.lower() in nm:
                    found_rows[key] = r
        
        for key, row_num in found_rows.items():
            if row_num is None:
                errors.append(f"{path.name} [RATIOS & KPIs]: Falta la fila '{required_rows[key]}'")
            else:
                # Verificar que al menos un período tenga un valor numérico
                has_value = False
                for lb in pcols:
                    c = None
                    for cc in range(2, ws.max_column + 1):
                        if _norm_hdr(ws.cell(row=hdr, column=cc).value) == lb:
                            c = cc
                            break
                    if c:
                        val = ws.cell(row=row_num, column=c).value
                        if isinstance(val, (int, float)) and val > 0:
                            has_value = True
                            break
                if not has_value:
                    errors.append(f"{path.name} [RATIOS & KPIs]: La fila '{required_rows[key]}' no tiene valores numéricos positivos en ningún período.")

    try:
        wb.close()
    except Exception:
        pass
    return errors


def main() -> int:
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('Product_v1/Total')
    if not base.exists():
        print(f"No existe: {base}")
        return 2
    files: List[Path] = []
    if base.is_file() and str(base).endswith('.xlsx'):
        files = [base]
    else:
        files = sorted([p for p in base.glob('*.xlsx') if p.name.endswith('_[ES].xlsx') or p.name.endswith('_es.xlsx') or 'ES' in p.name])
        if not files:
            # Fallback general: todos los .xlsx
            files = sorted(base.glob('*.xlsx'))

    all_errors: List[Tuple[Path, List[str]]] = []
    for p in files:
        errs = validate_workbook(p)
        if errs:
            all_errors.append((p, errs))

    if not all_errors:
        print(f"OK: {len(files)} archivo(s) sin errores")
        return 0

    for p, errs in all_errors:
        print(f"\nErrores en {p.name}:")
        for e in errs:
            print(f"  - {e}")
    print(f"\nResumen: {sum(len(e) for _, e in all_errors)} error(es) en {len(all_errors)}/{len(files)} archivo(s)")
    return 2


if __name__ == '__main__':
    raise SystemExit(main())

