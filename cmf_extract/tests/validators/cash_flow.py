#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# Period labels like 2024 or 2024Q1
PERIOD_RE_YEAR = re.compile(r"^\d{4}$")
PERIOD_RE_Q = re.compile(r"^(\d{4})Q([1-4])$")


def _strip_accents_lower(s: str) -> str:
    if not isinstance(s, str):
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch)).lower()


def _detect_header_row(ws) -> int:
    # Heurística: buscar fila con encabezados de periodos o etiqueta Cuenta/Concepto
    for r in range(1, min(12, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 40) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip().split("\n", 1)[0]
                if PERIOD_RE_Q.match(s) or PERIOD_RE_YEAR.match(s):
                    return r
    return 3


def _norm_hdr(v: str) -> Optional[str]:
    if not isinstance(v, str):
        return None
    s = v.strip().split("\n", 1)[0]
    if PERIOD_RE_Q.match(s) or PERIOD_RE_YEAR.match(s):
        return s
    return None


def _periods(ws, hdr_row: int) -> List[str]:
    out: List[str] = []
    seen = set()
    for c in range(2, ws.max_column + 1):
        v = _norm_hdr(ws.cell(row=hdr_row, column=c).value)
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _find_rows(ws, hdr_row: int) -> Dict[str, int]:
    rows = {
        'ini': None,
        'fin': None,
        'fx': None,
        'net_change': None,
        'op_tot': None,
        'inv_tot': None,
        'fin_tot': None,
        # Subtotals by banking vs non-banking
        'op_nb': None, 'op_b': None,
        'inv_nb': None, 'inv_b': None,
        'fin_nb': None, 'fin_b': None,
    }
    for r in range(hdr_row + 1, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if not raw:
            continue
        nm = _strip_accents_lower(str(raw).strip())

        # Principio / Final de efectivo
        if rows['ini'] is None and 'efectivo y equivalentes al efectivo al principio' in nm:
            rows['ini'] = r
        if rows['fin'] is None and 'efectivo y equivalentes al efectivo al final' in nm:
            rows['fin'] = r

        # Efectos de variación en tasa de cambio
        if rows['fx'] is None and 'variacion en la tasa de cambio' in nm:
            rows['fx'] = r

        # Net change line (Incremento/Disminución neto de efectivo)
        if rows['net_change'] is None:
            if ('neto' in nm and 'efectivo' in nm) and (
                'incremento' in nm or 'disminucion' in nm or 'aumento' in nm
            ):
                rows['net_change'] = r

        # Totales por actividad
        if rows['op_tot'] is None:
            if (
                ('operacion' in nm and 'actividades' in nm and 'netos' in nm)
                or ('operacion' in nm and 'flujos' in nm and 'netos' in nm)
            ):
                rows['op_tot'] = r
        if rows['inv_tot'] is None:
            if (
                ('inversion' in nm and 'actividades' in nm and 'netos' in nm)
                or ('inversion' in nm and 'flujos' in nm and 'netos' in nm)
            ):
                rows['inv_tot'] = r
        if rows['fin_tot'] is None:
            if (
                (('financiacion' in nm or 'financiamiento' in nm) and 'actividades' in nm and 'netos' in nm)
                or ((
                    'financiacion' in nm or 'financiamiento' in nm
                ) and 'flujos' in nm and 'netos' in nm)
            ):
                rows['fin_tot'] = r

        # Subtotales (negocios no bancarios / servicios bancarios) por actividad
        is_nb = ('negocios no bancarios' in nm) or ('no bancario' in nm)
        is_b = ('servicios bancarios' in nm) or ('bancario' in nm)
        if is_nb or is_b:
            if 'operacion' in nm:
                if is_nb and rows['op_nb'] is None:
                    rows['op_nb'] = r
                if is_b and rows['op_b'] is None:
                    rows['op_b'] = r
            if 'inversion' in nm:
                if is_nb and rows['inv_nb'] is None:
                    rows['inv_nb'] = r
                if is_b and rows['inv_b'] is None:
                    rows['inv_b'] = r
            if 'financiacion' in nm or 'financiamiento' in nm:
                if is_nb and rows['fin_nb'] is None:
                    rows['fin_nb'] = r
                if is_b and rows['fin_b'] is None:
                    rows['fin_b'] = r
    return rows  # type: ignore[return-value]


def _col_map(ws, hdr_row: int, pcols: List[str]) -> Dict[str, int]:
    cmap: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        v = _norm_hdr(ws.cell(row=hdr_row, column=c).value)
        if v in pcols:
            cmap[v] = c
    return cmap


def _num(v) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def detect_cash_flow_structure(path: Path) -> Tuple[Optional[object], Optional[object], Optional[int], List[str]]:
    """
    Carga el workbook y detecta estructura básica de la hoja Flujo Efectivo.
    Retorna (wb, ws, header_row, period_labels). wb/ws pueden ser None si falla.
    """
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return None, None, None, []
    if 'Flujo Efectivo' not in wb.sheetnames:
        return wb, None, None, []
    ws = wb['Flujo Efectivo']
    hdr = _detect_header_row(ws)
    pcols = _periods(ws, hdr)
    return wb, ws, hdr, pcols


def validate_hierarchical_cash_flow(path: Path, tol: float = 3_000_000.0) -> List[str]:
    """
    Valida conciliación de efectivo considerando:
    - Línea directa de incremento neto
    - Totales por actividad (operación/inversión/financiación)
    - Subtotales por negocios no bancarios y bancarios
    Devuelve lista de errores formateados (vacía si todo OK).
    """
    errs: List[str] = []
    wb, ws, hdr, pcols = detect_cash_flow_structure(path)
    if ws is None or hdr is None:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        # No hay flujo → dejar que el validador base reporte esto
        return errs

    rows = _find_rows(ws, hdr)
    cmap = _col_map(ws, hdr, pcols)

    def get_row_val(row_idx: Optional[int], col_idx: int) -> Optional[float]:
        if not row_idx:
            return None
        return _num(ws.cell(row=row_idx, column=col_idx).value)

    # Primero: si existe net_change, validamos final = inicio + net_change (+ FX si corresponde, algunas empresas lo incluyen aparte)
    per_mismatches: List[str] = []
    for lb in pcols:
        c = cmap.get(lb)
        if not c:
            continue
        v_ini = get_row_val(rows['ini'], c)
        v_fin = get_row_val(rows['fin'], c)
        v_fx = get_row_val(rows['fx'], c) or 0.0
        v_net = get_row_val(rows['net_change'], c)

        # Estrategia A: usar net_change si existe
        if v_ini is not None and v_fin is not None and v_net is not None:
            expected = v_ini + v_net
            if abs(expected - v_fin) > tol and abs(expected + v_fx - v_fin) > tol:
                per_mismatches.append(
                    f"{lb}: final {v_fin} != inicio {v_ini} + neto {v_net} [+ FX {v_fx}]"
                )
            continue

        # Estrategia B: usar totales por actividad (op+inv+fin + FX)
        v_op = get_row_val(rows['op_tot'], c)
        v_inv = get_row_val(rows['inv_tot'], c)
        v_finac = get_row_val(rows['fin_tot'], c)

        # Fallback: sumar subtotales bancario + no bancario si no hay total por actividad
        if v_op is None:
            a = get_row_val(rows['op_nb'], c)
            b = get_row_val(rows['op_b'], c)
            if a is not None or b is not None:
                v_op = (a or 0.0) + (b or 0.0)
        if v_inv is None:
            a = get_row_val(rows['inv_nb'], c)
            b = get_row_val(rows['inv_b'], c)
            if a is not None or b is not None:
                v_inv = (a or 0.0) + (b or 0.0)
        if v_finac is None:
            a = get_row_val(rows['fin_nb'], c)
            b = get_row_val(rows['fin_b'], c)
            if a is not None or b is not None:
                v_finac = (a or 0.0) + (b or 0.0)

        if v_ini is not None and v_fin is not None and (
            v_op is not None or v_inv is not None or v_finac is not None
        ):
            total_flows = (v_op or 0.0) + (v_inv or 0.0) + (v_finac or 0.0) + (v_fx or 0.0)
            expected_fin = v_ini + total_flows
            if abs(expected_fin - v_fin) > tol:
                per_mismatches.append(
                    f"{lb}: final {v_fin} != inicio {v_ini} + flujos ({v_op or 0.0}+{v_inv or 0.0}+{v_finac or 0.0}+{v_fx or 0.0})"
                )

    if per_mismatches:
        errs.append(
            f"{path.name} [Flujo]: Inconsistencias en conciliación de efectivo → "
            + "; ".join(per_mismatches)
        )

    try:
        wb.close()
    except Exception:
        pass
    return errs


def flexible_cash_flow_validation(path: Path, tol: float = 3_000_000.0) -> List[str]:
    """
    Envoltura para la validación flexible del flujo de efectivo.
    """
    return validate_hierarchical_cash_flow(path, tol=tol)


def validate_cash_flow_consistency(path: Path, tol: float = 3_000_000.0) -> List[str]:
    """
    Interfaz principal usada por el test.
    Intenta distintas estrategias (net_change, totales, subtotales) para validar el flujo.
    """
    return flexible_cash_flow_validation(path, tol=tol)
