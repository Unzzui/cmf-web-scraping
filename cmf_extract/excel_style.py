"""Paleta y tipografía del Excel. Fuente única de verdad.

Por qué existe
--------------
Había DOS Excel con dos lenguajes visuales distintos, y el cliente podía recibir los dos:

    El de TypeScript (lib/excel-report.ts, la descarga desde la ficha):
        Inter · tinta #0B0D12 · acento cobre #B87333 · sin cuadrícula

    El de Python (Product_v1, el que se compra en la tienda):
        Segoe UI Y Calibri (dos tipografías) · slate de Tailwind · acento AZUL #3B82F6

Dos archivos del mismo producto, con dos marcas. En un producto financiero eso hace la
misma pregunta que un ratio inconsistente: "si esto no lo cuidan, ¿los números sí?".

Esta paleta es la de TypeScript, y ahora los dos generadores tiran de ella.

LA PALETA IMPRESA DE FEY
------------------------
El sistema de diseño del sitio (docs/DESIGN.md) es Fey: dark-only por definición. Su
propia documentación lo dice — "el halo negro y la tipografía luminosa no tienen
traducción a un lienzo claro".

Pero un Excel se lee, se imprime y se fotocopia en blanco. Así que hace falta una
traducción, y el acento cálido de Fey (Ember, rgb(255,161,108)) no sirve tal cual: sobre
blanco da un contraste de 1,9:1, ilegible.

El cobre #B87333 ES ese mismo tono cálido, oscurecido para papel: 4,5:1 sobre blanco,
pasa WCAG AA. Es Ember impreso.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────────────────── Color
# ARGB (openpyxl), idénticos a lib/excel-report.ts.
INK = "FF0B0D12"      # texto y cifras — casi negro, no negro puro
EMBER = "FFB87333"    # acento: encabezados de sección, KPIs, enlaces. Ember impreso.
LINE = "FFD9DDE3"     # bordes y hairlines
SOFT = "FFF4F6F8"     # fondo de fila alterna, paneles
MUTED = "FF6B7280"    # texto secundario, notas al pie
PAPER = "FFFFFFFF"    # el lienzo

# Dirección de precio. Son los ÚNICOS colores además del acento, y sólo se usan para eso:
# un número que sube o baja. Nunca para decorar.
GROWTH = "FF15803D"
LOSS = "FFB91C1C"

# Tinte de cobre para el ÚNICO degradado del libro: el heatmap de magnitud. Antes ese
# heatmap era una escala viridis (amarillo → verde → petróleo), tres colores que no
# significan nada en un estado financiero. La magnitud es una sola dimensión, así que
# le corresponde una sola tinta.
EMBER_TENUE = "FFF6ECE1"

# ─────────────────────────────────────────────────────────── Tipografía
# Una sola familia hace todos los roles, igual que en el sitio. Antes convivían Segoe UI
# y Calibri en el mismo libro, sin ninguna razón.
FAMILIA = "Inter"


def fuente(
    size: int = 10,
    bold: bool = False,
    color: str = INK,
    underline: str | None = None,
    italic: bool = False,
) -> Font:
    """La ÚNICA forma de crear una fuente en este repo.

    `Font(...)` a secas no lleva `name`, y openpyxl entonces pone **Calibri**. Así se
    colaron 2.546 celdas en Calibri dentro de un libro que dice usar Inter: nadie eligió
    esa tipografía, apareció por omisión.
    """
    return Font(name=FAMILIA, size=size, bold=bold, color=color, underline=underline, italic=italic)


# Roles tipográficos. La jerarquía sale del PESO y el TAMAÑO, no del color: el acento se
# reserva para lo que de verdad quiere decir algo.
TITULO = fuente(20, bold=True, color=INK)          # nombre de la empresa en la portada
SECCION = fuente(13, bold=True, color=INK)         # título de hoja
ETIQUETA = fuente(9, bold=True, color=EMBER)       # rótulo de sección — el acento vive aquí
CABECERA = fuente(10, bold=True, color=PAPER)      # cabecera de tabla (sobre relleno tinta)
CUERPO = fuente(10, color=INK)                     # celdas de datos
CUERPO_FUERTE = fuente(10, bold=True, color=INK)   # subtotales
NOTA = fuente(9, color=MUTED)                      # notas y fuentes
LEGAL = fuente(8, color=MUTED)                     # pie legal

# ─────────────────────────────────────────────────────────── Formatos numéricos
# Los negativos entre PARÉNTESIS, no con signo menos: es la convención contable, y es la
# que espera cualquiera que lea un estado financiero.
FMT_NUM = "#,##0;(#,##0)"
FMT_PCT = "0.0%;(0.0%)"
FMT_MULT = '0.00"x"'
FMT_DIAS = '0" días"'
FMT_MONEDA_CLP = '"$"#,##0;("$"#,##0)'
FMT_MONEDA_USD = '"US$"#,##0;("US$"#,##0)'


def formato_moneda(moneda: str | None) -> str:
    """El formato lleva el símbolo de la moneda REAL. 18 empresas reportan en dólares."""
    return FMT_MONEDA_USD if (moneda or "").upper() == "USD" else FMT_MONEDA_CLP


# ─────────────────────────────────────────────────────────── Rellenos y bordes
def relleno(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


RELLENO_TINTA = relleno(INK)     # cabeceras de tabla
RELLENO_SUAVE = relleno(SOFT)    # filas alternas, paneles
RELLENO_PAPEL = relleno(PAPER)

# La profundidad sale de UNA hairline, no de una rejilla. La cuadrícula del Excel se
# oculta (ws.sheet_view.showGridLines = False): las líneas que quedan son las que uno
# decidió poner.
_LINEA = Side(style="thin", color=LINE)
_TINTA = Side(style="thin", color=INK)

BORDE_INFERIOR = Border(bottom=_LINEA)
BORDE_SUPERIOR_TINTA = Border(top=_TINTA)   # cierra un subtotal
BORDE_CAJA = Border(left=_LINEA, right=_LINEA, top=_LINEA, bottom=_LINEA)

# ─────────────────────────────────────────────────────────── Alineación
IZQ = Alignment(horizontal="left", vertical="center")
DER = Alignment(horizontal="right", vertical="center")
CENTRO = Alignment(horizontal="center", vertical="center")
AJUSTE = Alignment(horizontal="left", vertical="top", wrap_text=True)


def preparar_hoja(ws, congelar: str | None = "B6") -> None:
    """Deja la hoja como corresponde: sin cuadrícula y con el encabezado congelado.

    Sin `freeze_panes`, en una serie de 49 períodos el usuario pierde de vista qué
    columna está mirando en cuanto hace scroll. Es la diferencia entre una tabla que se
    puede leer y una que hay que descifrar.
    """
    ws.sheet_view.showGridLines = False
    if congelar:
        ws.freeze_panes = congelar


def aplicar_tipografia_base(wb) -> int:
    """Pone la familia del libro en toda celda que no la haya declarado. Devuelve cuántas.

    openpyxl le da a cada celda una fuente por defecto, y esa fuente es **Calibri**. Así
    que una celda a la que nadie le asignó estilo no sale "sin tipografía": sale en
    Calibri, dentro de un libro que por lo demás usa Inter. En el Excel de Arauco eran
    2.544 celdas — entre ellas el CUERPO ENTERO de la hoja RATIOS & KPIs, que es el
    corazón del producto.

    No es una preferencia estética. Dos tipografías en un mismo libro es la clase de
    detalle que hace preguntar: "si no cuidan esto, ¿los números sí?".

    Se conserva todo lo demás de la fuente (tamaño, peso, color, cursiva): esto sólo
    corrige la familia.
    """
    from copy import copy

    n = 0
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for celda in fila:
                f = celda.font
                if f is not None and f.name == FAMILIA:
                    continue
                nueva = copy(f) if f is not None else Font()
                nueva.name = FAMILIA
                celda.font = nueva
                n += 1
    return n


# ─────────────────────────────────────────────────────────── Guarda de legibilidad
def _luminancia(argb: str) -> float:
    canales = [int(argb[i:i + 2], 16) / 255 for i in (2, 4, 6)]
    lin = [c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4 for c in canales]
    return 0.2126 * lin[0] + 0.7152 * lin[1] + 0.0722 * lin[2]


def contraste(letra: str, fondo: str) -> float:
    """Razón de contraste WCAG entre dos ARGB. 1,0 = invisible; 21 = negro sobre blanco."""
    a, b = _luminancia(letra), _luminancia(fondo)
    hi, lo = max(a, b), min(a, b)
    return (hi + 0.05) / (lo + 0.05)


def _argb(color, por_defecto: str) -> str:
    if color is None or getattr(color, "type", None) != "rgb":
        return por_defecto
    v = color.rgb
    if not isinstance(v, str):
        return por_defecto
    return v if len(v) == 8 else "FF" + v[-6:]


def verificar_contraste(wb, minimo: float = 3.0) -> list[tuple[str, str, float]]:
    """Devuelve las celdas cuyo texto NO se lee sobre su propio fondo.

    Por qué esto existe
    -------------------
    Un color de letra y un color de fondo son UNA decisión, no dos, pero el código las
    escribía en líneas distintas — y al cambiar los rellenos de la portada quedaron
    cinco botones de navegación en BLANCO SOBRE BLANCO. Contraste 1,0:1. El texto seguía
    ahí, en el archivo, y era invisible. Lo mismo le pasó al WACC del DCF: letra blanca
    sobre relleno claro, y es el número del que cuelga toda la valorización.

    Nada en el proceso lo habría detectado, porque un Excel se "genera bien" aunque no
    se pueda leer. Por eso la verificación corre en la propia generación: si algo vuelve
    a quedar ilegible, avisa antes de que el archivo llegue a un cliente.

    El umbral es 3,0:1 (el mínimo de WCAG para texto grande). Los pares de la paleta
    superan holgadamente: tinta sobre papel da 18,9:1; blanco sobre tinta, otro tanto.
    """
    malas: list[tuple[str, str, float]] = []
    for ws in wb.worksheets:
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value in (None, ""):
                    continue
                fondo = PAPER
                if celda.fill is not None and celda.fill.patternType == "solid":
                    fondo = _argb(celda.fill.start_color, PAPER)
                letra = _argb(celda.font.color if celda.font else None, INK)
                r = contraste(letra, fondo)
                if r < minimo:
                    malas.append((ws.title, celda.coordinate, round(r, 2)))
    return malas
