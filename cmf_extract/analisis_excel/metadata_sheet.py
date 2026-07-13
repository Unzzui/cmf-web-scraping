"""
Metadata Sheet Module
=====================

Generates a "Ficha Técnica" (or "Data Sheet" in English) worksheet that is
inserted at position 0 of an openpyxl Workbook, summarising key metadata
about the generated financial Excel report.

Usage
-----
    from analisis_excel.metadata_sheet import create_metadata_sheet

    create_metadata_sheet(
        wb=workbook,
        company_name="Banco de Chile",
        rut="97030000-7",
        periods=["2020", "2021", "2022", "2023Q1"],
        frequency="Total",
        lang="es",
        currency="Miles de CLP",
    )
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

try:
    from cmf_extract import excel_style as est
except ImportError:  # ejecutado desde dentro de cmf_extract/
    import excel_style as est

# ---------------------------------------------------------------------------
# Colour palette (shared with ExcelFormatter brand colours)
# ---------------------------------------------------------------------------
# Esta hoja tenía su propia paleta —azul marino, azul de enlace, tres grises de
# Tailwind— dentro de un libro cuyo sistema de diseño no contempla ninguno. Ahora los
# nombres siguen, pero apuntan a los tokens de excel_style: la Ficha Técnica se ve como
# el resto del producto.
#
# En ARGB de 8 dígitos, no de 6: con 6, openpyxl antepone "00" (alfa CERO) y el color
# queda declarado como transparente. Es la razón por la que la auditoría veía
# "00F4F6F8" como un color ajeno, siendo el mismo gris de la paleta.
_NAVY = est.INK
_DARK_GRAY = est.INK
_LIGHT_GRAY = est.SOFT
_TEXT_GRAY = est.MUTED
_BLUE_LINK = est.EMBER          # el acento cálido hace de enlace; no hay azul en Fey
_WHITE = est.PAPER
_BLACK = est.INK
_BORDER_COLOR = est.LINE

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    """Return a thin-sided Border using the project's standard colour."""
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _solid_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _apply_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> Any:
    """Write *value* to a cell and optionally style it.  Returns the cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    return cell


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_metadata_sheet(
    wb: Workbook,
    company_name: str,
    rut: str,
    periods: list[str],
    frequency: str,
    lang: str = "es",
    currency: str = "Miles de CLP",
) -> None:
    """Insert a professional metadata sheet at position 0 of *wb*.

    Parameters
    ----------
    wb:
        An openpyxl ``Workbook`` instance.  The sheet is inserted at index 0
        so it appears before any existing sheets (e.g. RATIOS & KPIs).
    company_name:
        Human-readable company name.
    rut:
        Chilean RUT identifier, e.g. ``"76129263-3"``.
    periods:
        Ordered list of period labels present in the workbook, e.g.
        ``["2014", "2015", ..., "2025Q1", "2025Q2"]``.
    frequency:
        One of ``"Total"``, ``"Anual"``, or ``"Trimestral"``.
    lang:
        ``"es"`` for Spanish (default) or ``"en"`` for English.
    currency:
        Currency description shown in the Moneda / Currency row.

    Returns
    -------
    None
        The sheet is added to *wb* as a side-effect.
    """
    # ------------------------------------------------------------------
    # Localisation strings
    # ------------------------------------------------------------------
    is_es = lang.lower() != "en"

    sheet_title = "Ficha Técnica" if is_es else "Data Sheet"
    heading = "FICHA TÉCNICA" if is_es else "DATA SHEET"
    subtitle = (
        "Resumen del archivo generado"
        if is_es
        else "Generated file summary"
    )
    disclaimer = (
        "Datos extraídos de reportes XBRL publicados por la Comisión para el "
        "Mercado Financiero (CMF) de Chile."
        if is_es
        else "Data extracted from XBRL reports published by Chile's Financial "
        "Market Commission (CMF)."
    )

    # Field labels
    labels: list[tuple[str, str]] = [
        ("Empresa" if is_es else "Company", company_name),
        ("RUT", rut),
        (
            "Generado" if is_es else "Generated",
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ),
        (
            "Fuente" if is_es else "Source",
            "CMF Chile - XBRL IFRS",
        ),
        (
            "Períodos" if is_es else "Periods",
            _period_range(periods),
        ),
        (
            "Moneda" if is_es else "Currency",
            currency,
        ),
        (
            "Frecuencia" if is_es else "Frequency",
            frequency,
        ),
        (
            "Versión" if is_es else "Version",
            "CMF Extract v3.0",
        ),
        ("Web", "findatachile.com"),
    ]

    # ------------------------------------------------------------------
    # Create worksheet and insert at position 0
    # ------------------------------------------------------------------
    ws = wb.create_sheet(title=sheet_title, index=0)

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions[get_column_letter(1)].width = 25
    ws.column_dimensions[get_column_letter(2)].width = 50

    # ------------------------------------------------------------------
    # Row 1 – Title
    # ------------------------------------------------------------------
    title_row = 1
    ws.row_dimensions[title_row].height = 30
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
    _apply_cell(
        ws,
        title_row,
        1,
        heading,
        font=est.fuente(bold=True, color=_WHITE, size=14),
        fill=_solid_fill(_NAVY),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=_thin_border(),
    )
    # The merged neighbour cell must also carry the border so the right
    # edge of the merge range is rendered.
    ws.cell(row=title_row, column=2).border = _thin_border()

    # ------------------------------------------------------------------
    # Row 2 – Subtitle
    # ------------------------------------------------------------------
    subtitle_row = 2
    ws.row_dimensions[subtitle_row].height = 22
    ws.merge_cells(
        start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=2
    )
    _apply_cell(
        ws,
        subtitle_row,
        1,
        subtitle,
        font=est.fuente(color=_WHITE, size=11),
        fill=_solid_fill(_DARK_GRAY),
        alignment=Alignment(horizontal="left", vertical="center", indent=1),
        border=_thin_border(),
    )
    ws.cell(row=subtitle_row, column=2).border = _thin_border()

    # ------------------------------------------------------------------
    # Data rows (label / value pairs)
    # ------------------------------------------------------------------
    label_font = est.fuente(bold=True, color=_BLACK, size=10)
    label_fill = _solid_fill(_LIGHT_GRAY)
    value_font = est.fuente(color=_BLACK, size=10)
    link_font = est.fuente(color=_BLUE_LINK, size=10, underline="single")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    border = _thin_border()

    for offset, (label, value) in enumerate(labels):
        data_row = 3 + offset
        ws.row_dimensions[data_row].height = 18

        # Column A – label
        _apply_cell(
            ws,
            data_row,
            1,
            label,
            font=label_font,
            fill=label_fill,
            alignment=left_align,
            border=border,
        )

        # Column B – value (hyperlink style for the website)
        is_web_row = label == "Web"
        _apply_cell(
            ws,
            data_row,
            2,
            value,
            font=link_font if is_web_row else value_font,
            alignment=left_align,
            border=border,
        )

    # ------------------------------------------------------------------
    # Empty spacer row
    # ------------------------------------------------------------------
    spacer_row = 3 + len(labels)
    ws.row_dimensions[spacer_row].height = 8

    # ------------------------------------------------------------------
    # Disclaimer row
    # ------------------------------------------------------------------
    disclaimer_row = spacer_row + 1
    ws.row_dimensions[disclaimer_row].height = 30
    ws.merge_cells(
        start_row=disclaimer_row,
        start_column=1,
        end_row=disclaimer_row,
        end_column=2,
    )
    _apply_cell(
        ws,
        disclaimer_row,
        1,
        disclaimer,
        font=est.fuente(italic=True, color=_TEXT_GRAY, size=9),
        alignment=Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
            indent=1,
        ),
        border=_thin_border(),
    )
    ws.cell(row=disclaimer_row, column=2).border = _thin_border()


# ---------------------------------------------------------------------------
# Internal utilities
# ---------------------------------------------------------------------------

def _period_range(periods: list[str]) -> str:
    """Return a concise string representing the span of *periods*.

    Examples
    --------
    >>> _period_range(["2014", "2015", "2025Q1"])
    '2014 - 2025Q1'
    >>> _period_range(["2023Q3"])
    '2023Q3'
    >>> _period_range([])
    'N/A'
    """
    if not periods:
        return "N/A"
    if len(periods) == 1:
        return periods[0]
    return f"{periods[0]} - {periods[-1]}"
