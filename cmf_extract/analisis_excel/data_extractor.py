"""
Data Extractor Module
====================

Módulo para extraer datos de estados financieros desde archivos Excel.
Maneja la identificación y extracción de conceptos contables específicos.
"""

import re
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from .utils.lang_map import load_mappings, guess_is_english

# ISO 4217. Lo que no esté aquí NO es una moneda, por mucho que una empresa lo haya
# escrito en la casilla de "moneda de presentación".
_MONEDAS_ISO = {"CLP", "USD", "EUR"}


class DataExtractor:
    """
    Extractor de datos de estados financieros desde archivos Excel.
    """
    
    def __init__(self, file_path: str):
        """
        Inicializa el extractor con un archivo Excel.
        
        Args:
            file_path: Ruta al archivo Excel de estados financieros
        """
        self.file_path = file_path
        self.df_bal = None
        self.df_pl = None
        self.df_cfs = None
        self.years = []
        self.currency_by_year: dict[int, str] = {}
        # Quality tracking
        self.missing_sheets: list[str] = []
        self.warnings: list[str] = []
        self.found_accounts: list[str] = []
        self.missing_accounts: list[str] = []
        self.estimated_accounts: list[str] = []
        # Income statement role: "310000" (function) or "320000" (nature)
        self.income_role: str = "310000"
        
    def _pick_sheet(self, wb, candidates: list[str]):
        for name in candidates:
            if name in wb.sheetnames:
                return wb[name]
        raise KeyError(f"No se encontró ninguna de las hojas: {candidates}")

    def _detect_header_row(self, sheet) -> int:
        """Detecta la fila de encabezado buscando una fila con fechas YYYY- en columnas 2..n o 'Cuenta/Concepto' en col A.

        Retorna número de fila 1-based. Por defecto, 3 (título, subtítulo, encabezado) si no detecta.
        """
        max_scan = min(10, sheet.max_row)
        for r in range(1, max_scan + 1):
            v0 = sheet.cell(row=r, column=1).value
            if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
                return r
            for c in range(2, min(sheet.max_column, 12) + 1):
                v = sheet.cell(row=r, column=c).value
                if isinstance(v, str) and re.match(r"^\d{4}-", v):
                    return r
        return 3

    def load_data(self) -> bool:
        """
        Carga los datos de las hojas de estados financieros.

        Returns:
            True si la carga fue exitosa, False en caso contrario
        """
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)

            # Validate required sheets with graceful fallback
            try:
                sh_bal = self._pick_sheet(wb, ["Balance General", "Balance Sheet"])
            except KeyError:
                self.missing_sheets.append("Balance")
                self.warnings.append("Balance sheet not found — cannot perform analysis")
                return False  # Balance is mandatory

            try:
                sh_pl = self._pick_sheet(wb, ["Estado de Resultados", "Estado Resultados (Función)", "Income Statement"])
            except KeyError:
                sh_pl = None
                self.missing_sheets.append("Estado de Resultados")
                self.warnings.append("Income Statement not found — only balance ratios available")

            try:
                sh_cfs = self._pick_sheet(wb, ["Flujo Efectivo", "Cash Flow"])
            except KeyError:
                sh_cfs = None
                self.missing_sheets.append("Flujo Efectivo")
                self.warnings.append("Cash Flow not found — ratios without cash flow data")

            hdr_bal = self._detect_header_row(sh_bal)
            self.df_bal = pd.read_excel(self.file_path, sheet_name=sh_bal.title, header=hdr_bal - 1)

            if sh_pl is not None:
                hdr_pl = self._detect_header_row(sh_pl)
                self.df_pl = pd.read_excel(self.file_path, sheet_name=sh_pl.title, header=hdr_pl - 1)
            else:
                self.df_pl = pd.DataFrame(columns=["Concepto"])

            if sh_cfs is not None:
                hdr_cfs = self._detect_header_row(sh_cfs)
                self.df_cfs = pd.read_excel(self.file_path, sheet_name=sh_cfs.title, header=hdr_cfs - 1)
            else:
                self.df_cfs = pd.DataFrame(columns=["Concepto"])

            # Estándar: primera columna como 'Concepto'
            self.df_bal.rename(columns={self.df_bal.columns[0]: "Concepto"}, inplace=True)
            self.df_pl.rename(columns={self.df_pl.columns[0]: "Concepto"}, inplace=True)
            self.df_cfs.rename(columns={self.df_cfs.columns[0]: "Concepto"}, inplace=True)

            # Si el archivo está en inglés, traducir cabeceras de conceptos a ES usando cuentas.json
            es_to_en, en_to_es = load_mappings()
            # Balance
            if guess_is_english(self.df_bal["Concepto"].astype(str).tolist()):
                rev = en_to_es.get("balance", {})
                self.df_bal["Concepto"] = self.df_bal["Concepto"].astype(str).map(lambda x: rev.get(x, rev.get(x.lower(), x)))
            # Estado resultados
            if guess_is_english(self.df_pl["Concepto"].astype(str).tolist()):
                rev = en_to_es.get("estado_resultados", {})
                self.df_pl["Concepto"] = self.df_pl["Concepto"].astype(str).map(lambda x: rev.get(x, rev.get(x.lower(), x)))
            # Flujo caja
            if guess_is_english(self.df_cfs["Concepto"].astype(str).tolist()):
                rev = en_to_es.get("flujo_caja", {})
                self.df_cfs["Concepto"] = self.df_cfs["Concepto"].astype(str).map(lambda x: rev.get(x, rev.get(x.lower(), x)))

            # Detect income statement role (310000=function, 320000=nature)
            self._detect_income_role()

            # Extraer años disponibles
            self._extract_years()

            return True
        except Exception as e:
            print(f"Error cargando datos: {e}")
            return False
    
    def _detect_income_role(self):
        """Auto-detect whether income statement uses function (310000) or nature (320000) layout."""
        if self.df_pl is None or self.df_pl.empty:
            return
        concepts_lc = self.df_pl["Concepto"].astype(str).str.strip().str.lower()
        has_cogs = concepts_lc.str.contains("costo de ventas|cost of sales", regex=True, na=False).any()
        has_gross = concepts_lc.str.contains("ganancia bruta|gross profit", regex=True, na=False).any()
        has_rawmat = concepts_lc.str.contains("materias primas|raw material", regex=True, na=False).any()
        has_emp_ben = concepts_lc.str.contains("beneficios a los empleados|employee benefit", regex=True, na=False).any()
        if (has_rawmat or has_emp_ben) and not has_cogs and not has_gross:
            self.income_role = "320000"
            self.warnings.append("Income Statement by Nature (320000) — no COGS or Gross Profit rows")
        else:
            self.income_role = "310000"

    def _extract_years(self):
        """Extrae los años disponibles soportando etiquetas YYYY, YYYYQn o YYYY-MM-DD."""
        years_set = set()

        def _year_of(label: str) -> int | None:
            s = str(label).strip().split("\n", 1)[0]
            # YYYYQn
            m = re.match(r"^(\d{4})Q[1-4]$", s)
            if m:
                return int(m.group(1))
            # YYYY-MM-DD
            m = re.match(r"^(\d{4})-\d{2}-\d{2}$", s)
            if m:
                return int(m.group(1))
            # YYYY
            m = re.match(r"^(\d{4})$", s)
            if m:
                return int(m.group(1))
            return None

        for df in [self.df_bal, self.df_pl, self.df_cfs]:
            for col in df.columns[1:]:
                y = _year_of(col)
                if y is not None:
                    years_set.add(y)

        self.years = sorted(years_set)

        # La moneda sale del XBRL, no de lo que la empresa haya escrito a mano.
        #
        # Antes se raspaba una fila de TEXTO LIBRE ("Moneda de presentación") de las
        # notas, y se tomaba tal cual lo que hubiera ahí. Arauco escribió "USA" en 2014,
        # y la Ficha Técnica del producto salió diciendo:
        #
        #     Moneda: 2014 USA; 2015-2026 USD
        #
        # "USA" no es una moneda: no existe en ISO 4217. Su XBRL de ese mismo período
        # declara `iso4217:USD`, explícito y sin ambigüedad. El dato correcto estaba en
        # el archivo fuente y el pipeline prefirió una casilla de texto.
        self.currency_by_year = self._monedas_desde_xbrl()
        if self.currency_by_year:
            return

        # Respaldo: sólo si no hay XBRL a mano. Y ahora se VALIDA contra ISO 4217, para
        # que un "USA" se descarte en vez de viajar hasta el cliente.
        try:
            # Buscar una fila que describa la moneda de presentación en df_pl/df_bal
            def _scan(df: pd.DataFrame) -> dict[int, str]:
                out: dict[int, str] = {}
                if df is None or 'Concepto' not in df.columns:
                    return out
                date_cols = [c for c in df.columns if c != 'Concepto']
                for _, row in df.iterrows():
                    concepto = str(row.get('Concepto', '')).lower()
                    if ('moneda' in concepto and 'present' in concepto) or ('currency' in concepto and 'presentation' in concepto):
                        for c in date_cols:
                            val = row.get(c)
                            if isinstance(val, str) and val.strip():
                                try:
                                    y = int(str(c)[:4])
                                    out[y] = val.strip()
                                except Exception:
                                    continue
                return out
            cur: dict[int, str] = {}
            cur.update(_scan(self.df_pl))
            cur.update(_scan(self.df_bal))
            self.currency_by_year = {
                anio: m.upper() for anio, m in cur.items()
                if m.strip().upper() in _MONEDAS_ISO
            }
        except Exception:
            self.currency_by_year = {}

    def _monedas_desde_xbrl(self) -> dict[int, str]:
        """Mapa año → moneda, leído de la unidad `<xbrli:measure>` del XBRL.

        El RUT sale del nombre del archivo (``estados_93458000-1_2026-2014_es.xlsx``),
        que es como el pipeline nombra sus salidas.
        """
        try:
            from cmf_extract.currency_detect import monedas_por_anio
        except ImportError:
            try:
                from currency_detect import monedas_por_anio
            except ImportError:
                return {}

        m = re.search(r"(\d{7,8}-[\dkK])", str(self.file_path))
        if not m:
            return {}
        rut = m.group(1).upper()

        raiz = Path(__file__).resolve().parents[2] / "data" / "XBRL" / "Total"
        if not raiz.is_dir():
            return {}
        carpeta = next((d for d in raiz.iterdir() if d.is_dir() and d.name.upper().startswith(rut)), None)
        if carpeta is None:
            return {}

        return monedas_por_anio(carpeta)
    
    def find_row_series(self, df: pd.DataFrame, concept_name: str) -> pd.Series:
        """
        Busca una serie de datos por concepto exacto o contiene.
        
        Args:
            df: DataFrame donde buscar
            concept_name: Nombre del concepto a buscar
            
        Returns:
            Serie con los valores numéricos del concepto
        """
        # Búsqueda exacta primero
        mask = df["Concepto"].astype(str).str.strip().str.lower() == concept_name.strip().lower()
        if mask.any():
            series = df[mask].iloc[0].drop(labels=["Concepto"])
            series.index = series.index.astype(str)
            return pd.to_numeric(series, errors="coerce")
        
        # Búsqueda por contiene
        mask = df["Concepto"].astype(str).str.contains(re.escape(concept_name), case=False, na=False)
        if mask.any():
            series = df[mask].iloc[0].drop(labels=["Concepto"])
            series.index = series.index.astype(str)
            return pd.to_numeric(series, errors="coerce")
        
        return pd.Series(dtype=float)
    
    def extract_balance_sheet_items(self) -> Dict[str, pd.Series]:
        """
        Extrae los conceptos principales del Balance General.
        
        Returns:
            Diccionario con los conceptos del balance general
        """
        concepts = {
            "AC": "Activos corrientes totales",
            "PC": "Pasivos corrientes totales", 
            "Efec": "Efectivo y equivalentes al efectivo",
            "Inv": "Inventarios corrientes",
            "AT": "Total de activos",
            "PT": "Total de pasivos",
            "Patr": "Patrimonio atribuible a los propietarios de la controladora",
            "CxC": "Deudores comerciales y otras cuentas por cobrar corrientes",
            "CxP": "Cuentas por pagar comerciales y otras cuentas por pagar"
        }
        
        result = {}
        for key, concept in concepts.items():
            series = self.find_row_series(self.df_bal, concept)
            if series.empty or series.dropna().empty:
                self.missing_accounts.append(concept)
            else:
                self.found_accounts.append(concept)
            result[key] = series

        # Cuentas opcionales (deuda financiera, utilidades retenidas, acciones)
        # para ratios avanzados; si no existen no se reportan como faltantes
        optional_concepts = {
            "OPFC": "Otros pasivos financieros corrientes",
            "OPFNC": "Otros pasivos financieros no corrientes",
            "UtilRet": "Ganancias (pérdidas) acumuladas",
            "Acciones": "Número de acciones suscritas",
        }
        for key, concept in optional_concepts.items():
            series = self.find_row_series(self.df_bal, concept)
            if not series.empty and not series.dropna().empty:
                result[key] = series
                self.found_accounts.append(concept)
        return result
    
    def extract_income_statement_items(self) -> Dict[str, pd.Series]:
        """
        Extrae los conceptos principales del Estado de Resultados.
        
        Returns:
            Diccionario con los conceptos del estado de resultados
        """
        concepts = {
            "Ventas": "Ingresos de actividades ordinarias",
            "COGS": "Costo de ventas",
            "Bruta": "Ganancia bruta",
            "EBIT": "Ganancias (pérdidas) de actividades operacionales",
            "Neta": "Ganancia (pérdida)",
            "Interes": "Costos financieros",
            "Dep": "Depreciación",
            "Amort": "Amortización"
        }
        
        items = {}
        for key, concept in concepts.items():
            series = self.find_row_series(self.df_pl, concept)
            if series.empty or series.dropna().empty:
                self.missing_accounts.append(concept)
            else:
                self.found_accounts.append(concept)
            items[key] = series

        # Complemento por naturaleza (320000): materias primas, cambio inventarios, trabajos capitalizados
        nat_concepts = {
            "RawMat": "Materias primas y consumibles utilizados",
            "InvChange": "Disminución (aumento) en inventarios de productos terminados y en proceso",
            "WorkCap": "Otros trabajos realizados por la entidad y capitalizados",
        }
        for key, concept in nat_concepts.items():
            ser = self.find_row_series(self.df_pl, concept)
            if not ser.empty:
                items[key] = ser
                self.found_accounts.append(nat_concepts[key])

        # Calcular D&A combinado
        dep = items.get("Dep", pd.Series(dtype=float))
        amort = items.get("Amort", pd.Series(dtype=float))
        if (dep.empty or dep.dropna().empty) and (amort.empty or amort.dropna().empty):
            self.missing_accounts.append("Depreciación y amortización (D&A)")
            self.warnings.append("D&A: not available — EBITDA Margin will show N/A")
        items["DA"] = dep.add(amort, fill_value=0)

        # Cuentas opcionales para ROIC (tasa impositiva efectiva implícita)
        tax_concepts = {
            "Impuesto": "Gasto por impuestos a las ganancias",
            "PreTax": "Ganancia (pérdida), antes de impuestos",
        }
        for key, concept in tax_concepts.items():
            ser = self.find_row_series(self.df_pl, concept)
            if not ser.empty and not ser.dropna().empty:
                items[key] = ser
                self.found_accounts.append(concept)

        return items
    
    def extract_cash_flow_items(self) -> Dict[str, pd.Series]:
        """
        Extrae los conceptos principales del Flujo de Efectivo.
        
        Returns:
            Diccionario con los conceptos del flujo de efectivo
        """
        concepts = {
            "CFO": "Flujos de efectivo netos procedentes de (utilizados en) operaciones",
            "CapexBuy": "Compras de propiedades, planta y equipo"
        }
        
        items = {}
        for key, concept in concepts.items():
            series = self.find_row_series(self.df_cfs, concept)
            if series.empty or series.dropna().empty:
                self.missing_accounts.append(concept)
            else:
                self.found_accounts.append(concept)
            items[key] = series

        # Calcular CAPEX y FCF
        capex_buy = items.get("CapexBuy", pd.Series(dtype=float))
        items["CAPEX"] = capex_buy.abs()
        
        cfo = items.get("CFO", pd.Series(dtype=float))
        items["FCF"] = cfo.subtract(items["CAPEX"], fill_value=0)
        
        return items
    
    def get_column_for_year(self, series: pd.Series, year: int) -> Optional[str]:
        """
        Encuentra la columna correspondiente a un año específico.
        
        Args:
            series: Serie de datos
            year: Año a buscar
            
        Returns:
            Nombre de la columna o None si no se encuentra
        """
        for col in series.index:
            if str(col).startswith(f"{year}-"):
                return col
        return None
    
    def get_all_financial_data(self) -> Dict[str, Dict[str, pd.Series]]:
        """
        Extrae todos los datos financieros de las tres hojas.
        
        Returns:
            Diccionario con todos los datos organizados por hoja
        """
        return {
            "balance": self.extract_balance_sheet_items(),
            "income": self.extract_income_statement_items(),
            "cash_flow": self.extract_cash_flow_items(),
            "years": self.years,
            "currency_by_year": self.currency_by_year,
            "_quality": {
                "missing_sheets": self.missing_sheets[:],
                "warnings": self.warnings[:],
                "found_accounts": self.found_accounts[:],
                "missing_accounts": self.missing_accounts[:],
                "estimated_accounts": self.estimated_accounts[:],
                "income_role": self.income_role,
            },
        }
