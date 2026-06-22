#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parsea la hoja "Taxonomía Ilustrada" de un Excel de la CMF y genera un JSON
con todas las secciones y sus tablas. Detecta:
- Fila de sección como: "[105000] Texto ..." (y toma la URL si viene en otra celda)
- Encabezado de tabla con columnas: Prefijo | Nombre | Etiqueta | Tipo | Referencias
- Fin de tabla por 4 filas vacías seguidas o al encontrar la siguiente sección

Uso:
  python parse_taxonomia_ilustrada.py -i /ruta/archivo.xlsx -o salida.json
  # Para partir cada sección en un archivo JSON independiente:
  python parse_taxonomia_ilustrada.py -i /ruta/archivo.xlsx --split out_dir

Requisitos:
  pip install openpyxl
"""

from __future__ import annotations
import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


EXPECTED_HEADERS = ["prefijo", "nombre", "etiqueta", "tipo", "referencias"]


def norm(val) -> str:
    """Normaliza celdas a str sin espacios extremos."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()


def is_url(text: str) -> bool:
    return isinstance(text, str) and text.lower().startswith("http")


def parse_section_header(row_vals: List[str]) -> Optional[Dict[str, str]]:
    """
    Detecta filas de sección tipo: "[105000] Nota - Comentarios..." (en cualquier celda no-URL).
    Si hay una URL en la fila, la captura en 'url'.
    """
    non_empty = [v for v in row_vals if v]
    if not non_empty:
        return None

    url = next((v for v in non_empty if is_url(v)), None)
    header_text = next((v for v in non_empty if not is_url(v)), None)
    if not header_text:
        return None

    m = re.match(r"^\[(\d+)\]\s*(.+)$", header_text)
    if not m:
        return None

    section_id = m.group(1)
    section_title = m.group(2).strip()
    return {"id": section_id, "title": section_title, "url": url}


def is_table_header(row_vals: List[str]) -> bool:
    """
    Verifica si la fila tiene los encabezados requeridos (insensible a mayúsculas/minúsculas).
    Acepta que haya otras celdas con contenido.
    """
    found = []
    for v in row_vals:
        vl = v.lower() if v else ""
        # pequeños alias tolerados para "referencias"
        if vl in EXPECTED_HEADERS or vl in ("referencia", "referencia(s)"):
            token = "referencias" if vl.startswith("referenc") else vl
            if token not in found:
                found.append(token)
    return all(h in found for h in EXPECTED_HEADERS)


def header_positions(row_vals: List[str]) -> Dict[str, int]:
    """
    Retorna el índice de cada columna esperada en la fila de encabezado.
    """
    pos = {}
    for idx, v in enumerate(row_vals):
        vl = (v or "").lower()
        if vl in EXPECTED_HEADERS:
            if vl not in pos:
                pos[vl] = idx
        elif vl in ("referencia", "referencia(s)"):
            if "referencias" not in pos:
                pos["referencias"] = idx
    return pos


def load_rows(xlsx_path: Path, sheet_name: Optional[str]) -> List[List[str]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else (
        wb["Taxonomía Ilustrada"] if "Taxonomía Ilustrada" in wb.sheetnames else wb.active
    )
    rows = []
    for r in sheet.iter_rows(values_only=True):
        rows.append([norm(c) for c in r])
    return rows


def parse_sheet(rows: List[List[str]], min_blank_streak: int = 4) -> List[Dict]:
    sections = []
    i, n = 0, len(rows)

    while i < n:
        header_info = parse_section_header(rows[i])
        if not header_info:
            i += 1
            continue

        # buscar encabezado de tabla
        j = i + 1
        while j < n and not is_table_header(rows[j]):
            # si encontramos otra sección antes del encabezado, guardamos sección vacía
            if parse_section_header(rows[j]):
                break
            j += 1

        # Si no hay encabezado o nos topamos con otra sección → items vacíos
        if j >= n or parse_section_header(rows[j]):
            sections.append({**header_info, "items": []})
            i = j
            continue

        # mapear posiciones de columnas
        pos = header_positions(rows[j])

        items = []
        k = j + 1
        blank_streak = 0
        while k < n:
            row = rows[k]

            # si aparece otra sección
            if parse_section_header(row):
                break

            # control de filas en blanco
            if all(not v for v in row):
                blank_streak += 1
                if blank_streak >= min_blank_streak:
                    k += 1
                    break
                k += 1
                continue
            else:
                blank_streak = 0

            def get(col_name: str) -> str:
                idx = pos.get(col_name)
                return row[idx] if (idx is not None and idx < len(row)) else ""

            item = {
                "prefijo": get("prefijo"),
                "nombre": get("nombre"),
                "etiqueta": get("etiqueta"),
                "tipo": get("tipo"),
                "referencias": get("referencias"),
            }
            # evitar filas 100% vacías
            if any(item.values()):
                items.append(item)

            k += 1

        sections.append({**header_info, "items": items})
        i = k

    return sections


def save_json(sections: List[Dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(sections, f, ensure_ascii=False, indent=2)


def save_split(sections: List[Dict], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    for s in sections:
        sid = s.get("id", "unknown")
        title = s.get("title", "").strip().replace("/", "-")
        name = f"{sid} - {title}.json" if title else f"{sid}.json"
        path = out_dir / name
        with path.open("w", encoding="utf-8") as f:
            json.dump(s, f, ensure_ascii=False, indent=2)


def main():
    ap = argparse.ArgumentParser(description="Parser de 'Taxonomía Ilustrada' (CMF) → JSON")
    ap.add_argument("-i", "--input", default="/home/unzzui/Documents/coding/CMF_extract/analisis_excel/utils/testeo_pdf/taxonomia/articles-28078_recurso_10.xlsx",  help="Ruta al Excel fuente (.xlsx)")
    ap.add_argument("-s", "--sheet", default=None, help="Nombre de hoja (default: 'Taxonomía Ilustrada' o la activa)")
    ap.add_argument("-o", "--output", default="/home/unzzui/Documents/coding/CMF_extract/analisis_excel/utils/testeo_pdf/taxonomia/taxonomia_ilustrada.json", help="Ruta del JSON de salida")
    ap.add_argument("--split", default=None, help="Directorio para guardar un JSON por sección")
    ap.add_argument("--min-blank-rows", type=int, default=4, help="Corte de tabla tras N filas vacías seguidas (default: 4)")
    args = ap.parse_args()

    xlsx_path = Path(args.input)
    rows = load_rows(xlsx_path, args.sheet)
    sections = parse_sheet(rows, min_blank_streak=args.min_blank_rows)

    if args.split:
        save_split(sections, Path(args.split))
        print(f"✔ Secciones guardadas individualmente en: {args.split} (total {len(sections)})")
    else:
        save_json(sections, Path(args.output))
        print(f"✔ JSON guardado en: {args.output} (secciones: {len(sections)})")


if __name__ == "__main__":
    main()
