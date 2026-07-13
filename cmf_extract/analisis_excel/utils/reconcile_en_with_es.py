"""
Reconcile EN statements with ES values
=====================================

Utilidad para tomar un workbook EN (Balance Sheet, Income Statement, Cash Flow)
y sobre-escribir sus valores con los del workbook ES pareado, verificando todas
las cuentas (filas) y todos los períodos (columnas). Si hay discrepancias, se
llenan con el dato en español. También se limpian filas EN que no tienen
contraparte en ES (opcional) para evitar diferencias de sets.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Optional, Tuple, Any
import os
import time
import tempfile
import zipfile

from openpyxl import load_workbook
from openpyxl.styles.fills import Fill as _Fill
from openpyxl.styles import PatternFill as _PatternFill
from openpyxl.styles import Font as _Font
from openpyxl.styles import Alignment as _Alignment
try:
    from cmf_extract import excel_style as est
except ImportError:
    import excel_style as est



SHEET_MAP = {
    "Balance General": "Balance Sheet",
    "Estado de Resultados": "Income Statement",
    "Flujo Efectivo": "Cash Flow",
}

def _find_sheet_by_name_or_prefix(wb, base_name: str):
    """Busca una hoja por nombre exacto o por prefijo (antes de ' —').
    Soporta variantes como 'Balance Sheet — EMPRESA'.
    """
    try:
        if base_name in wb.sheetnames:
            return wb[base_name]
        # normalizar nombres existentes y comparar por prefijo/strip
        target = base_name.strip().lower()
        for nm in wb.sheetnames:
            s = str(nm).strip()
            s_simple = s.split("—", 1)[0].strip().lower()
            if s_simple == target:
                return wb[nm]
            if s.lower().startswith(target + " ") or s.lower().startswith(target + "—"):
                return wb[nm]
        return None
    except Exception:
        return None


def _detect_header_row(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and re.match(r"^\d{4}(|Q[1-4]|-\d{2}(-\d{2})?)$", v.strip()):
                return r
    return 3


def _normalize_label(raw: Any) -> Optional[str]:
    if not isinstance(raw, str):
        return None
    s = raw.strip().split("\n", 1)[0]
    if re.match(r"^\d{4}Q[1-4]$", s):
        return s
    m = re.match(r"^(\d{4})-(\d{2})(?:-\d{2})?$", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        q = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(mo)
        return f"{y}{q}" if q else str(y)
    # Bare YYYY -> YYYYQ4 for backward compat
    if re.match(r"^\d{4}$", s):
        return f"{s}Q4"
    return s


def _labels_to_col(ws, hdr: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        lb = _normalize_label(v)
        if lb:
            m[lb] = c
    return m


def _copy_cell_style(src, dst) -> None:
    try:
        try:
            dst.font = src.font
        except Exception:
            pass
        try:
            # Solo asignar fill si es un objeto Fill válido
            f = getattr(src, 'fill', None)
            if f is not None:
                try:
                    # Clonar PatternFill para evitar proxies incompatibles
                    if isinstance(f, _PatternFill):
                        dst.fill = _PatternFill(patternType=f.patternType,
                                                fgColor=f.fgColor,
                                                bgColor=f.bgColor)
                    elif isinstance(f, _Fill):
                        dst.fill = f
                except Exception:
                    pass
        except Exception:
            pass
        try:
            dst.alignment = src.alignment
        except Exception:
            pass
        try:
            dst.border = src.border
        except Exception:
            pass
        try:
            dst.number_format = src.number_format
        except Exception:
            pass
    except Exception:
        pass


def _reorder_columns_to_match(ws_en, ws_es, hdr_en: int, hdr_es: int) -> bool:
    """Reordena columnas en ws_en para que el orden de períodos coincida exactamente con ws_es.
    Conserva valores y estilos por celda y copia dimensiones de columna.
    """
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        get_column_letter = None  # type: ignore

    # Lista objetivo desde ES (normalizada), manteniendo el orden exacto
    labels_es_ordered = []
    for c in range(2, ws_es.max_column + 1):
        lb = _normalize_label(ws_es.cell(row=hdr_es, column=c).value)
        if lb:
            labels_es_ordered.append(lb)

    # Función auxiliar para encontrar la columna actual de un label en EN (búsqueda dinámica)
    def _find_col_in_en(lbl: str) -> int | None:
        for c in range(2, ws_en.max_column + 1):
            v = _normalize_label(ws_en.cell(row=hdr_en, column=c).value)
            if v == lbl:
                return c
        return None

    # Reordenar secuencialmente: colocar cada etiqueta en su posición destino
    changed = False
    for pos, lbl in enumerate(labels_es_ordered, start=0):
        target_col = 2 + pos
        # Encontrar en EN la columna actual para lbl
        cur_col = _find_col_in_en(lbl)
        if cur_col is None or cur_col == target_col:
            continue
        changed = True
        # Extraer datos y estilos de la columna actual
        col_values = []
        col_styles = []
        max_r = ws_en.max_row
        for r in range(1, max_r + 1):
            cell = ws_en.cell(row=r, column=cur_col)
            col_values.append(cell.value)
            col_styles.append(cell)
        # Insertar columna en la posición objetivo
        ws_en.insert_cols(target_col, 1)
        # Copiar contenido + estilo a la nueva columna
        for r in range(1, max_r + 1):
            dst = ws_en.cell(row=r, column=target_col)
            dst.value = col_values[r - 1]
            _copy_cell_style(col_styles[r - 1], dst)
        # Copiar dimensiones (ancho/oculto/outline)
        try:
            if get_column_letter is not None:
                src_letter = get_column_letter(cur_col + (1 if target_col <= cur_col else 0))
                dst_letter = get_column_letter(target_col)
                src_dim = ws_en.column_dimensions.get(src_letter)
                if src_dim is not None:
                    _ = ws_en.column_dimensions[dst_letter]
                    dst_dim = ws_en.column_dimensions.get(dst_letter)
                    if dst_dim is not None:
                        if getattr(src_dim, 'width', None) is not None:
                            dst_dim.width = src_dim.width
                        try:
                            dst_dim.hidden = getattr(src_dim, 'hidden', False)
                        except Exception:
                            pass
                        try:
                            lvl = getattr(src_dim, 'outlineLevel', None)
                            if lvl is not None:
                                dst_dim.outlineLevel = lvl
                        except Exception:
                            pass
        except Exception:
            pass
        # Borrar la columna original (ajustando índice si se insertó antes)
        del_col = cur_col + 1 if target_col <= cur_col else cur_col
        ws_en.delete_cols(del_col, 1)
    return changed


def _parse_num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip()
        s = s.replace("\u00A0", " ").replace("−", "-").replace("–", "-")
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1]
        s = s.replace(",", "").replace(" ", "")
        if s in ("", "-", "+"):
            return None
        num = float(s)
        if neg:
            num = -num
        return num
    except Exception:
        return None


def _apply_generator_styles_base_sheet(ws, hdr_row: int, lang: str = "en") -> None:
    """Re-aplica estilos similares a los del generador (xbrl_to_excel.py) sobre una hoja base.

    Aplica:
      - Estilo de título (fila 1) y subtítulo (fila 2) si existen
      - Encabezado (fila hdr_row)
      - Anchos de columnas (A=36, resto=14)
      - Estilos por fila: categoría, total, subcategorías con alternancia
      - Formato numérico con miles y negativos en rojo
      - Congelar paneles bajo encabezado y después de la primera columna
    """
    if hdr_row is None or hdr_row <= 0:
        return

    # Paleta y tipografías (alineadas con xbrl_to_excel.py)
    brand_primary = "0F172A"   # Navy oscuro
    brand_secondary = "1F2937" # Gris azulado oscuro
    brand_gray_100 = "F7F7F7"
    brand_gray_150 = "F0F0F0"
    subcat_bg = "FAFAFA"
    subcat_alt_bg = "F5F5F5"
    total_bg = "E0E7FF"

    font_title = _Font(bold=True, color="FFFFFF", size=16, name=est.FAMILIA)
    font_header = _Font(bold=True, color="FFFFFF", size=11, name=est.FAMILIA)
    font_normal = _Font(size=10, name=est.FAMILIA)
    font_bold = _Font(bold=True, size=10, name=est.FAMILIA)
    font_white_small = _Font(color="111827", size=11, name=est.FAMILIA)

    align_center = _Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = _Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = _Alignment(horizontal="right", vertical="center")

    fill_title = _PatternFill("solid", fgColor=brand_primary)
    fill_header = _PatternFill("solid", fgColor=brand_secondary)
    fill_category = _PatternFill("solid", fgColor=brand_primary)
    fill_concept = _PatternFill("solid", fgColor=brand_gray_100)
    fill_concept_alt = _PatternFill("solid", fgColor=brand_gray_150)
    fill_subcat = _PatternFill("solid", fgColor=subcat_bg)
    fill_subcat_alt = _PatternFill("solid", fgColor=subcat_alt_bg)
    fill_total = _PatternFill("solid", fgColor=total_bg)

    # 1) Título y subtítulo (si existen filas 1 y 2)
    try:
        if hdr_row >= 3:
            c = ws.cell(row=1, column=1)
            c.font = font_title
            c.fill = fill_title
            c.alignment = align_center
            try:
                ws.row_dimensions[1].height = 26
            except Exception:
                pass
            c2 = ws.cell(row=2, column=1)
            c2.font = font_white_small
            c2.alignment = align_center
            try:
                ws.row_dimensions[2].height = 18
            except Exception:
                pass
    except Exception:
        pass

    # 2) Encabezado
    try:
        for col in range(1, ws.max_column + 1):
            hc = ws.cell(row=hdr_row, column=col)
            hc.font = font_header
            hc.fill = fill_header
            hc.alignment = align_center
        try:
            ws.row_dimensions[hdr_row].height = 22
        except Exception:
            pass
    except Exception:
        pass

    # 3) Anchos de columna
    try:
        from openpyxl.utils import get_column_letter
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            dim = ws.column_dimensions.get(letter)
            if dim is None:
                _ = ws.column_dimensions[letter]
                dim = ws.column_dimensions.get(letter)
            if dim is not None:
                try:
                    dim.width = 36 if col == 1 else 14
                except Exception:
                    pass
    except Exception:
        pass

    # 4) Congelar paneles bajo el encabezado y después de la primera columna
    try:
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)
    except Exception:
        pass

    # 5) Estilos por filas
    # Listas para detectar totales (idénticas a las del generador)
    cuentas_total_es = [
        'Ganancia bruta',
        'Ganancias (pérdidas) de actividades operacionales',
        'Ganancia (pérdida), antes de impuestos',
        'Ganancia (pérdida)',
        'Flujos de efectivo netos procedentes de (utilizados en) operaciones',
        'Flujos de efectivo netos procedentes de (utilizados en) actividades de inversión',
        'Flujos de efectivo netos procedentes de (utilizados en) actividades de financiación',
        'Efectivo y equivalentes al efectivo al final del periodo'
    ]
    cuentas_total_en = [
        'Gross profit',
        'Profit (loss) from operating activities',
        'Profit (loss)',
        'Net cash flows from (used in) operations',
        'Net cash flows from (used in) investing activities'
    ]
    cuentas_total_ifrs = [
        'ifrs-full:CashAndCashEquivalentsIfDifferentFromStatementOfFinancialPosition'
    ]

    data_start = hdr_row + 1
    for r in range(data_start, ws.max_row + 1):
        try:
            cuenta_raw = ws.cell(row=r, column=1).value
            cuenta = str(cuenta_raw) if cuenta_raw is not None else ""
            cuenta_stripped = cuenta.strip()
            is_category = cuenta_stripped.startswith('[') and (
                ']' in cuenta_stripped[:12]
            )
            cuenta_lower = cuenta_stripped.lower()
            is_total = (
                ('total' in cuenta_lower) or ('suma' in cuenta_lower) or ('subtotal' in cuenta_lower)
                or (cuenta_stripped in cuentas_total_es)
                or (cuenta_stripped in cuentas_total_en)
                or (cuenta_stripped in cuentas_total_ifrs)
            )
            r_index = r - data_start
            is_alternate = (r_index % 2 == 1)

            # Columna 1: etiqueta de cuenta
            c0 = ws.cell(row=r, column=1)
            if is_category:
                c0.font = _Font(bold=True, color="FFFFFF", size=11, name=est.FAMILIA)
                c0.fill = fill_category
                c0.alignment = align_left
            elif is_total:
                c0.font = font_bold
                c0.fill = fill_total
                c0.alignment = align_left
            else:
                c0.font = font_normal
                c0.fill = (fill_subcat_alt if is_alternate else fill_subcat)
                c0.alignment = _Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

            # Columnas de valores
            for c in range(2, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if is_category:
                    cell.fill = fill_category
                    cell.alignment = align_center
                    cell.number_format = "General"
                    continue
                # Determinar si es numérico
                v = cell.value
                num = None
                if isinstance(v, (int, float)):
                    num = float(v)
                else:
                    num = _parse_num(v)
                    if num is not None:
                        try:
                            cell.value = num
                        except Exception:
                            pass
                if num is not None:
                    # Numérico
                    if is_total:
                        cell.font = font_bold
                        cell.fill = fill_total
                        cell.alignment = align_right
                        cell.number_format = '#,##0'
                    else:
                        cell.font = font_normal
                        # Limpiar cualquier fill previo para que coincida con el generador (sin color de fondo)
                        try:
                            cell.fill = _PatternFill()
                        except Exception:
                            pass
                        cell.alignment = align_right
                        cell.number_format = '#,##0_);[Red](#,##0)' if num < 0 else '#,##0'
                else:
                    # Texto o vacío
                    cell.font = font_normal
                    cell.fill = (fill_concept_alt if is_alternate else fill_concept)
                    cell.alignment = align_left
                    cell.number_format = "General"
        except Exception:
            continue


def _apply_grouping_like_generator(ws, hdr_row: int) -> None:
    """Ajusta el outline de columnas (agrupación trimestral bajo cada año) como en xbrl_to_excel.

    - Marca columnas Q1..Q4 de cada año con outlineLevel=1
    - Oculta los bloques trimestrales de años que no son el último
    - Colapsa la columna 'año' (resumen) para años no últimos
    - Activa símbolos de outline visibles, summaryAbove y summaryLeft
    """
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    try:
        # Construir mapping de etiquetas normalizadas
        labels: dict[int, str] = {}
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=hdr_row, column=c).value
            lb = _normalize_label(v)
            if lb:
                labels[c] = lb
        if not labels:
            return
        # Índices por año
        years: set[int] = set()
        for lb in labels.values():
            import re as _re
            if _re.match(r"^\d{4}$", str(lb)):
                years.add(int(lb))
            else:
                m = _re.match(r"^(\d{4})Q[1-4]$", str(lb))
                if m:
                    years.add(int(m.group(1)))
        if not years:
            return
        latest_year = max(years)
        # Mapa año → columnas de quarters y columna resumen (año)
        year_col: dict[int, int] = {}
        year_quarters: dict[int, list[int]] = {}
        import re as _re2
        for c, lb in labels.items():
            if _re2.match(r"^\d{4}$", str(lb)):
                year_col[int(lb)] = c
            else:
                m = _re2.match(r"^(\d{4})Q([1-4])$", str(lb))
                if m:
                    y = int(m.group(1))
                    year_quarters.setdefault(y, []).append(c)
        # Aplicar outline
        for y, qs in year_quarters.items():
            if not qs:
                continue
            # Sort y aplicar nivel 1 y hidden según año
            qs_sorted = sorted(qs)
            hide_block = (y != latest_year)
            for c in qs_sorted:
                letter = get_column_letter(c)
                _ = ws.column_dimensions[letter]
                dim = ws.column_dimensions.get(letter)
                if dim is None:
                    continue
                try:
                    dim.outlineLevel = 1
                except Exception:
                    pass
                try:
                    dim.hidden = hide_block
                except Exception:
                    pass
            # Colapsar columna resumen (año)
            yc = year_col.get(y)
            if yc:
                letter_y = get_column_letter(yc)
                _ = ws.column_dimensions[letter_y]
                dimy = ws.column_dimensions.get(letter_y)
                if dimy is not None:
                    try:
                        dimy.collapsed = hide_block
                    except Exception:
                        pass
        # Configuración de outline del sheet: resumen arriba/izquierda y mostrar símbolos
        try:
            out = getattr(ws.sheet_properties, 'outlinePr', None)
            if out is None:
                from openpyxl.worksheet.properties import Outline
                ws.sheet_properties.outlinePr = Outline()
                out = ws.sheet_properties.outlinePr
            try:
                out.summaryBelow = False
            except Exception:
                pass
            try:
                out.summaryRight = False
            except Exception:
                pass
            try:
                out.applyStyles = True
            except Exception:
                pass
        except Exception:
            pass
        try:
            ws.sheet_view.showOutlineSymbols = True
        except Exception:
            pass
    except Exception:
        pass

def _find_es_counterpart(path_en: Path) -> Optional[Path]:
    name = path_en.name
    if name.endswith("_en.xlsx"):
        cand = path_en.with_name(name.replace("_en.xlsx", "_es.xlsx"))
        if cand.exists():
            return cand
    if "[EN].xlsx" in name:
        cand = path_en.with_name(name.replace("[EN].xlsx", "[ES].xlsx"))
        if cand.exists():
            return cand
    if " - Financial Analysis" in name:
        cand = path_en.with_name(name.replace(" - Financial Analysis", " - Análisis Financiero").replace("[EN]", "[ES]"))
        if cand.exists():
            return cand
    for p in path_en.parent.glob("*.xlsx"):
        if "[ES]" in p.name and path_en.stem.split("[")[0].strip() in p.stem:
            return p
    return None


def reconcile_en_with_es(en_file: Path, es_file: Optional[Path] = None, blank_unmatched_en_rows: bool = True) -> bool:
    """
    Copia valores de ES → EN para todas las hojas mapeadas y todos los periodos comunes.
    Si blank_unmatched_en_rows=True, limpia filas EN que no existen en ES (por índice) para evitar DIFFSET.
    Devuelve True si modificó EN, False si no pudo.
    """
    try:
        import time as _time
        _t_start = _time.perf_counter()
    except Exception:
        _time = None  # type: ignore
        _t_start = 0

    en_file = Path(en_file)
    if not en_file.exists():
        return False
    if es_file is None:
        es_file = _find_es_counterpart(en_file)
    if not es_file or not es_file.exists():
        return False
    wb_en = load_workbook(str(en_file))
    wb_es = load_workbook(str(es_file), read_only=False, data_only=True)

    def _atomic_save_excel(wb, target: Path) -> None:
        # Esperar si hay lock de LibreOffice en el directorio
        dir_path = target.parent
        lock_prefix = ".~lock."
        wait_deadline = time.time() + float(os.getenv("CMF_RECON_LOCK_TIMEOUT", "5"))
        while time.time() < wait_deadline:
            has_lock = any(p.name.startswith(lock_prefix) for p in dir_path.glob(".~lock.*"))
            if not has_lock:
                break
            time.sleep(0.25)

        # Guardar a un archivo temporal y reemplazar atómicamente
        attempts = 3
        last_err: Optional[Exception] = None
        for _ in range(attempts):
            tmp_fd = None
            tmp_path = None
            try:
                fd, tmp_path = tempfile.mkstemp(prefix=".tmp_recon_", suffix=".xlsx", dir=str(dir_path))
                tmp_fd = fd
                os.close(tmp_fd)
                tmp_fd = None
                wb.save(tmp_path)
                # Verificación básica de ZIP y contenido principal
                with zipfile.ZipFile(tmp_path, 'r') as zf:
                    if "[Content_Types].xml" not in zf.namelist():
                        raise KeyError("[Content_Types].xml missing in xlsx")
                os.replace(tmp_path, target)
                return
            except Exception as e:
                last_err = e
                time.sleep(0.4)
            finally:
                try:
                    if tmp_fd is not None:
                        os.close(tmp_fd)
                except Exception:
                    pass
                try:
                    if tmp_path and os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass
        if last_err:
            raise last_err

    try:
        modified = False
        modified_es = False
        # Caso 1: Workbooks de estados base (Balance/Resultados/Flujo)
        for es_name, en_name in SHEET_MAP.items():
            ws_es = _find_sheet_by_name_or_prefix(wb_es, es_name)
            ws_en = _find_sheet_by_name_or_prefix(wb_en, en_name)
            if ws_es is None or ws_en is None:
                continue
            hdr_es = _detect_header_row(ws_es)
            hdr_en = _detect_header_row(ws_en)
            cols_es = _labels_to_col(ws_es, hdr_es)
            cols_en = _labels_to_col(ws_en, hdr_en)
            # Mapear filas por concepto (ES) y por concepto traducido a ES (EN)
            try:
                from analisis_excel.utils.lang_map import load_mappings
                _, en_to_es = load_mappings()
                def _is_category(name: str) -> bool:
                    s = name.strip()
                    return s.startswith("[") and ("]" in s[:12])
                def _row_key_map(ws, hdr, group_key: str) -> dict[int, str]:
                    m: dict[int, str] = {}
                    rev = en_to_es.get(group_key, {})
                    for r in range(hdr + 1, ws.max_row + 1):
                        nm = ws.cell(row=r, column=1).value
                        if not isinstance(nm, str):
                            continue
                        s = nm.strip()
                        if _is_category(s):
                            continue
                        # traducir EN→ES si corresponde
                        s_es = rev.get(s, rev.get(s.lower(), s))
                        key = s_es.strip().lower()
                        if key:
                            m[r] = key
                    return m
                sheet_group = 'balance' if es_name == 'Balance General' else ('estado_resultados' if es_name.startswith('Estado de Resultados') else 'flujo_caja')
                rowkey_es = _row_key_map(ws_es, hdr_es, sheet_group)
                rowkey_en = _row_key_map(ws_en, hdr_en, sheet_group)
                en_row_by_key: dict[str, int] = {}
                for r, k in rowkey_en.items():
                    en_row_by_key.setdefault(k, r)
            except Exception:
                rowkey_es = {}
                en_row_by_key = {}
            # Asegurar que EN tenga todas las columnas de período presentes en ES (en el mismo orden)
            try:
                labels_es_ordered = []
                for c in range(2, ws_es.max_column + 1):
                    v = ws_es.cell(row=hdr_es, column=c).value
                    lb = _normalize_label(v)
                    if lb:
                        labels_es_ordered.append(lb)
                # Insertar en EN las etiquetas faltantes siguiendo el orden de ES
                from openpyxl.utils import get_column_letter
                for pos, lbl in enumerate(labels_es_ordered, start=0):
                    if lbl in cols_en:
                        continue
                    insert_at = 2 + pos  # columna 1 = Cuenta
                    ws_en.insert_cols(insert_at, 1)
                    # Header label
                    ws_en.cell(row=hdr_en, column=insert_at, value=lbl)
                    # Copiar estilo completo desde ES para esa etiqueta
                    try:
                        c_es = cols_es.get(lbl)
                        if c_es is not None:
                            max_r = max(ws_es.max_row, ws_en.max_row)
                            for r in range(1, max_r + 1):
                                s = ws_es.cell(row=r, column=c_es)
                                d = ws_en.cell(row=r, column=insert_at)
                                _copy_cell_style(s, d)
                            # Dimensiones de columna según ES
                            try:
                                src_letter = get_column_letter(c_es)
                                dst_letter = get_column_letter(insert_at)
                                src_dim = ws_es.column_dimensions.get(src_letter)
                                if src_dim is not None:
                                    _ = ws_en.column_dimensions[dst_letter]
                                    dst_dim = ws_en.column_dimensions.get(dst_letter)
                                    if dst_dim is not None:
                                        if getattr(src_dim, 'width', None) is not None:
                                            dst_dim.width = src_dim.width
                                        try:
                                            dst_dim.hidden = getattr(src_dim, 'hidden', False)
                                        except Exception:
                                            pass
                                        try:
                                            lvl = getattr(src_dim, 'outlineLevel', None)
                                            if lvl is not None:
                                                dst_dim.outlineLevel = lvl
                                        except Exception:
                                            pass
                                        try:
                                            dst_dim.collapsed = getattr(src_dim, 'collapsed', False)
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Actualizar mapping EN por shift
                    for y, cidx in list(cols_en.items()):
                        if cidx >= insert_at:
                            cols_en[y] = cidx + 1
                    cols_en[lbl] = insert_at
                    modified = True
            except Exception:
                pass
            common_labels = sorted(set(cols_es) & set(cols_en))
            # Recorrer filas por índice (conserva estructura de presentación)
            start_es = hdr_es + 1
            start_en = hdr_en + 1
            max_rows = min(ws_es.max_row - hdr_es, ws_en.max_row - hdr_en)
            for i in range(max_rows):
                r_es = start_es + i
                # Intentar mapear por concepto; si no existe, caer al índice
                r_en = start_en + i
                try:
                    key_es = rowkey_es.get(r_es)
                    if key_es and key_es in en_row_by_key:
                        r_en = en_row_by_key[key_es]
                except Exception:
                    pass
                # Saltar categorías [210000] etc.
                def _is_category(cell_val: Any) -> bool:
                    if not isinstance(cell_val, str):
                        return False
                    s = cell_val.strip()
                    return s.startswith("[") and "]" in s[:12]
                if _is_category(ws_es.cell(row=r_es, column=1).value) or _is_category(ws_en.cell(row=r_en, column=1).value):
                    continue
                for lbl in common_labels:
                    c_es = cols_es[lbl]
                    c_en = cols_en[lbl]
                    v_es = _parse_num(ws_es.cell(row=r_es, column=c_es).value)
                    v_en = _parse_num(ws_en.cell(row=r_en, column=c_en).value)
                    # Si difieren o EN tiene valor y ES none, forzar ES
                    equal = False
                    try:
                        if v_es is None and v_en is None:
                            equal = True
                        elif (v_es is not None) and (v_en is not None):
                            equal = abs(v_es - v_en) <= 1e-6
                    except Exception:
                        equal = False
                    if not equal:
                        # Escribir ES (incluye None → limpiar EN)
                        ws_en.cell(row=r_en, column=c_en).value = v_es
                        modified = True

            # Limpiar filas EN sin contraparte (cuando sea posible identificar por clave)
            if blank_unmatched_en_rows:
                try:
                    es_keys_set = set(k for k in rowkey_es.values())
                    for r in range(start_en, ws_en.max_row + 1):
                        k = rowkey_en.get(r)
                        if k and (k not in es_keys_set):
                            for lbl in common_labels:
                                c_en = cols_en[lbl]
                                if 2 <= c_en <= ws_en.max_column:
                                    ws_en.cell(row=r, column=c_en).value = None
                                    modified = True
                except Exception:
                    pass

            # Replicar diseño/agrupación básico desde ES → EN (encabezados y columnas)
            try:
                # 1) Merges en filas de encabezado (filas 1..hdr)
                merges_to_remove = [rng for rng in list(ws_en.merged_cells.ranges)
                                    if (rng.min_row <= hdr_en)]
                for rng in merges_to_remove:
                    ws_en.unmerge_cells(range_string=str(rng))
                for rng in ws_es.merged_cells.ranges:
                    if rng.min_row <= hdr_es:
                        ws_en.merge_cells(range_string=str(rng))
                        modified = True
            except Exception:
                pass
            # 1.b) Reordenar columnas EN para que coincidan exactamente con el orden ES
            try:
                if _reorder_columns_to_match(ws_en, ws_es, hdr_en, hdr_es):
                    modified = True
            except Exception:
                pass
            # 1.c) Eliminar columnas de períodos extra en EN que no existan en ES
            try:
                es_set = set(labels_es_ordered)
                to_delete = []
                for c in range(ws_en.max_column, 1, -1):
                    hv = ws_en.cell(row=hdr_en, column=c).value
                    if not isinstance(hv, str):
                        continue
                    lb = _normalize_label(hv)
                    if lb and re.match(r"^\d{4}(?:Q[1-4])?$", lb) and lb not in es_set:
                        to_delete.append(c)
                for c in to_delete:
                    ws_en.delete_cols(c, 1)
                    modified = True
            except Exception:
                pass

            # 1.d) Alinear textos de encabezado de períodos con ES, pero NO copiar estilos (estilos se re-aplican)
            try:
                cols_en = _labels_to_col(ws_en, hdr_en)
                # Actualizar textos de header fila hdr (c>=2)
                try:
                    for c in range(2, min(ws_en.max_column, ws_es.max_column) + 1):
                        se = ws_es.cell(row=hdr_es, column=c)
                        de = ws_en.cell(row=hdr_en, column=c)
                        if se.value is not None:
                            de.value = se.value
                    modified = True
                except Exception:
                    pass
            except Exception:
                pass
            # 1.c) Copiar configuración de outline del sheet (símbolos arriba/izquierda)
            try:
                es_out = getattr(ws_es.sheet_properties, 'outlinePr', None)
                if es_out is not None:
                    out = ws_en.sheet_properties.outlinePr
                    if out is None:
                        from openpyxl.worksheet.properties import Outline
                        ws_en.sheet_properties.outlinePr = Outline()
                        out = ws_en.sheet_properties.outlinePr
                    try:
                        out.summaryBelow = es_out.summaryBelow
                    except Exception:
                        pass
                    try:
                        out.summaryRight = es_out.summaryRight
                    except Exception:
                        pass
                    try:
                        out.applyStyles = getattr(es_out, 'applyStyles', None)
                    except Exception:
                        pass
                    modified = True
            except Exception:
                pass
            try:
                # 2) Anchos/oculto/outline por columna
                from openpyxl.utils import get_column_letter
                for col_idx in range(1, max(ws_es.max_column, ws_en.max_column) + 1):
                    col_letter = get_column_letter(col_idx)
                    src_dim = ws_es.column_dimensions.get(col_letter)
                    if src_dim is None:
                        continue
                    _ = ws_en.column_dimensions[col_letter]
                    dst_dim = ws_en.column_dimensions.get(col_letter)
                    if dst_dim is None:
                        continue
                    try:
                        if src_dim.width is not None:
                            dst_dim.width = src_dim.width
                        dst_dim.hidden = getattr(src_dim, 'hidden', False)
                        lvl = getattr(src_dim, 'outlineLevel', None)
                        if lvl is not None:
                            dst_dim.outlineLevel = lvl
                        try:
                            dst_dim.collapsed = getattr(src_dim, 'collapsed', False)
                        except Exception:
                            pass
                        modified = True
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                # 3) Freeze panes
                ws_en.freeze_panes = ws_es.freeze_panes
                modified = True
            except Exception:
                pass
            # 4) Re-aplicar estilos del generador sobre EN (en vez de copiar desde ES) y agrupar columnas
            try:
                _apply_generator_styles_base_sheet(ws_en, hdr_en, lang="en")
                _apply_grouping_like_generator(ws_en, hdr_en)
                modified = True
            except Exception:
                pass
            # 4.b) Asegurar misma agrupación visual en ES para que ambos libros queden consistentes
            try:
                _apply_grouping_like_generator(ws_es, hdr_es)
                modified_es = True
            except Exception:
                pass

        # Caso 2: Producto final (Product_v1) con hoja 'RATIOS & KPIs'
        try:
            if ("RATIOS & KPIs" in wb_es.sheetnames) and ("RATIOS & KPIs" in wb_en.sheetnames):
                ws_es = wb_es["RATIOS & KPIs"]
                ws_en = wb_en["RATIOS & KPIs"]
                # Orden en RATIOS: anual ascendente (2014..2024) seguido por trimestres recientes (2025Q1, 2025Q2)
                try:
                    hdr_es = 4
                    hdr_en = 4
                    # Construir orden objetivo: años asc + trimestres Q1..Q4 (si existen)
                    labels = []
                    for c in range(2, ws_es.max_column + 1):
                        lb = _normalize_label(ws_es.cell(row=hdr_es, column=c).value)
                        if lb:
                            labels.append(lb)
                    # Derivar años y trimestres de ES pero forzando orden años asc y luego Q1..Q4 del último año
                    years = sorted({int(lb) for lb in labels if isinstance(lb, str) and lb.isdigit()})
                    qs = [lb for lb in labels if isinstance(lb, str) and re.match(r"^\d{4}Q[1-4]$", lb)]
                    def _qkey(s: str) -> tuple[int, int]:
                        m = re.match(r"^(\d{4})Q([1-4])$", s)
                        return (int(m.group(1)), int(m.group(2))) if m else (9999, 9)
                    qs_sorted = sorted(qs, key=_qkey)
                    target = [str(y) for y in years] + qs_sorted
                    # Reordenar EN para que coincida con target
                    def _reorder_by_list(ws, hdr, target_order):
                        from openpyxl.utils import get_column_letter
                        def _find_col(lbl: str) -> int | None:
                            for c in range(2, ws.max_column + 1):
                                if _normalize_label(ws.cell(row=hdr, column=c).value) == lbl:
                                    return c
                            return None
                        for pos, lbl in enumerate(target_order, start=0):
                            tcol = 2 + pos
                            ccol = _find_col(lbl)
                            if ccol is None or ccol == tcol:
                                continue
                            # mover columna ccol → tcol conservando estilos
                            max_r = ws.max_row
                            col_vals = [ws.cell(row=r, column=ccol).value for r in range(1, max_r+1)]
                            col_cells = [ws.cell(row=r, column=ccol) for r in range(1, max_r+1)]
                            ws.insert_cols(tcol, 1)
                            for r in range(1, max_r+1):
                                dst = ws.cell(row=r, column=tcol)
                                dst.value = col_vals[r-1]
                                _copy_cell_style(col_cells[r-1], dst)
                            # copiar dimensiones
                            try:
                                src_letter = get_column_letter(ccol + (1 if tcol <= ccol else 0))
                                dst_letter = get_column_letter(tcol)
                                src_dim = ws.column_dimensions.get(src_letter)
                                if src_dim is not None:
                                    _ = ws.column_dimensions[dst_letter]
                                    dst_dim = ws.column_dimensions.get(dst_letter)
                                    if dst_dim is not None:
                                        if getattr(src_dim, 'width', None) is not None:
                                            dst_dim.width = src_dim.width
                                        try:
                                            dst_dim.hidden = getattr(src_dim, 'hidden', False)
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                            del_col = ccol + 1 if tcol <= ccol else ccol
                            ws.delete_cols(del_col, 1)
                    _reorder_by_list(ws_en, hdr_en, target)
                    modified = True
                except Exception:
                    pass
                # Copiar merges de filas 1-4
                try:
                    # Limpiar merges existentes en EN en filas 1-4
                    merges_to_remove = [rng for rng in list(ws_en.merged_cells.ranges)
                                        if (rng.min_row <= 4)]
                    for rng in merges_to_remove:
                        ws_en.unmerge_cells(range_string=str(rng))
                except Exception:
                    pass
                try:
                    for rng in ws_es.merged_cells.ranges:
                        if rng.min_row <= 4:
                            ws_en.merge_cells(range_string=str(rng))
                            modified = True
                except Exception:
                    pass
                # Copiar anchos/ocultos/outline por columna
                from openpyxl.utils import get_column_letter
                try:
                    for col_idx in range(1, ws_es.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        src_dim = ws_es.column_dimensions.get(col_letter)
                        dst_dim = ws_en.column_dimensions.get(col_letter)
                        if src_dim is not None:
                            if dst_dim is None:
                                # asegurar entrada
                                _ = ws_en.column_dimensions[col_letter]
                                dst_dim = ws_en.column_dimensions.get(col_letter)
                            try:
                                if src_dim.width is not None:
                                    dst_dim.width = src_dim.width
                                dst_dim.hidden = getattr(src_dim, 'hidden', False)
                                # Outline (si existiera)
                                lvl = getattr(src_dim, 'outlineLevel', None)
                                if lvl is not None:
                                    dst_dim.outlineLevel = lvl
                                try:
                                    dst_dim.collapsed = getattr(src_dim, 'collapsed', False)
                                except Exception:
                                    pass
                                modified = True
                            except Exception:
                                pass
                except Exception:
                    pass
                # Copiar configuración de outline del sheet
                try:
                    es_out = getattr(ws_es.sheet_properties, 'outlinePr', None)
                    if es_out is not None:
                        out = ws_en.sheet_properties.outlinePr
                        if out is None:
                            from openpyxl.worksheet.properties import Outline
                            ws_en.sheet_properties.outlinePr = Outline()
                            out = ws_en.sheet_properties.outlinePr
                        try:
                            out.summaryBelow = es_out.summaryBelow
                        except Exception:
                            pass
                        try:
                            out.summaryRight = es_out.summaryRight
                        except Exception:
                            pass
                        try:
                            out.applyStyles = getattr(es_out, 'applyStyles', None)
                        except Exception:
                            pass
                        modified = True
                except Exception:
                    pass
                # Copiar freeze panes
                try:
                    ws_en.freeze_panes = ws_es.freeze_panes
                    modified = True
                except Exception:
                    pass
                # Copiar estilos de filas 1-4 (manteniendo valores EN)
                try:
                    for r in range(1, min(5, max(ws_es.max_row, 4) + 1)):
                        for c in range(1, ws_en.max_column + 1):
                            src = ws_es.cell(row=r, column=c)
                            dst = ws_en.cell(row=r, column=c)
                            try:
                                dst.font = src.font
                                dst.fill = src.fill
                                dst.alignment = src.alignment
                                dst.border = src.border
                                dst.number_format = src.number_format
                                modified = True
                            except Exception:
                                pass
                except Exception:
                    pass
                # Copiar estilos de todas las celdas de datos de RATIOS (siempre)
                try:
                    _t_style2 = _time.perf_counter()
                    data_start = 5
                    for r in range(data_start, ws_en.max_row + 1):
                        for c in range(1, ws_en.max_column + 1):
                            s = ws_es.cell(row=r, column=c)
                            d = ws_en.cell(row=r, column=c)
                            try:
                                d.font = s.font
                                d.fill = s.fill
                                d.alignment = s.alignment
                                d.border = s.border
                                d.number_format = s.number_format
                            except Exception:
                                pass
                    modified = True
                    print(f"[recon] copia estilos datos RATIOS ⏱ {(_time.perf_counter()-_t_style2):.1f}s :: {en_file.name}")
                except Exception:
                    pass
        except Exception:
            pass

    except Exception:
        return False

    # Guardar si hubo cambios, asegurando cierre correcto y reemplazo atómico
    try:
        if modified:
            _atomic_save_excel(wb_en, en_file)
        # Si se modificó el ES (agrupación), guardarlo también
        if modified_es:
            _atomic_save_excel(wb_es, es_file)
        try:
            elapsed = (_time.perf_counter()-_t_start) if _time else 0.0  # type: ignore
        except Exception:
            elapsed = 0.0
        status = 'changed' if modified else 'nochange'
        if modified_es:
            status += '+es'
        print(f"[recon] total ⏱ {elapsed:.1f}s :: {en_file.name} :: {status}")
        return modified
    finally:
        try:
            wb_en.close()
        except Exception:
            pass
        try:
            wb_es.close()
        except Exception:
            pass


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Reconciliar EN con ES (copiar valores ES → EN)")
    ap.add_argument("en_path", help="Ruta al workbook EN")
    ap.add_argument("--es", dest="es_path", help="Ruta al workbook ES pareado (opcional)")
    ap.add_argument("--keep-extra", action="store_true", help="No limpiar filas EN sin contraparte ES")
    args = ap.parse_args()
    en = Path(args.en_path)
    es = Path(args.es_path) if args.es_path else None
    ok = reconcile_en_with_es(en, es, blank_unmatched_en_rows=not args.keep_extra)
    print("OK" if ok else "NOOP")


