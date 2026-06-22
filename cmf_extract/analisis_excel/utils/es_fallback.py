"""
ES Fallback Utilities
=====================

Construye una hoja oculta 'ES_DATA' con los valores de los estados financieros
en español y configura FormulaBuilder para que, cuando haya diferencias con el
libro analizado (típicamente EN), referencie esos valores ES.

Este módulo verifica todas las cuentas clave y todos los periodos visibles.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter


# Mapas de claves → nombres de conceptos (ES) por grupo
BALANCE_KEYS_TO_ES: Dict[str, str] = {
    # Activos básicos
    "AC": "Activos corrientes totales",
    "AT": "Total de activos",
    "ANC": "Total de activos no corrientes",
    "PPE": "Propiedades, planta y equipo",
    "Efec": "Efectivo y equivalentes al efectivo",
    "Inv": "Inventarios corrientes",
    "InvNC": "Inventarios, no corrientes",
    "CxC": "Deudores comerciales y otras cuentas por cobrar corrientes",
    "CxCNC": "Cuentas por cobrar no corrientes",
    "Intang": "Activos intangibles distintos de la plusvalía",
    "Plusvalia": "Plusvalía",
    "ActivDifImp": "Activos por impuestos diferidos",
    "PropInv": "Propiedad de inversión",
    "ActivDerUso": "Activos por derecho de uso",
    "InvAsoc": "Inversiones contabilizadas utilizando el método de la participación",
    # Pasivos básicos
    "PC": "Pasivos corrientes totales",
    "PT": "Total de pasivos",
    "PNC": "Total de pasivos no corrientes",
    "CxP": "Cuentas por pagar comerciales y otras cuentas por pagar",
    "CxPNC": "Cuentas por pagar no corrientes",
    "DeudaFinCorr": "Otros pasivos financieros corrientes",
    "DeudaFinNC": "Otros pasivos financieros no corrientes",
    "ArrCorr": "Pasivos por arrendamientos corrientes",
    "ArrNC": "Pasivos por arrendamientos no corrientes",
    "PasivDifImp": "Pasivo por impuestos diferidos",
    "ProvCorr": "Otras provisiones a corto plazo",
    "ProvNC": "Otras provisiones a largo plazo",
    "BenefEmpl": "Provisiones corrientes por beneficios a los empleados",
    "BenefEmplNC": "Provisiones no corrientes por beneficios a los empleados",
    # Patrimonio
    "Patr": "Patrimonio atribuible a los propietarios de la controladora",
    "PatrTotal": "Patrimonio total",
    "CapitalEmit": "Capital emitido y pagado",
    "GanAcum": "Ganancias (pérdidas) acumuladas",
    "PrimaEmis": "Prima de emisión",
    "AccPropias": "Acciones propias en cartera",
    "OtraPartPatr": "Otras participaciones en el patrimonio",
    "OtraReserv": "Otras reservas",
    "PartNoControl": "Participaciones no controladoras",
}

INCOME_KEYS_TO_ES: Dict[str, str] = {
    "Ventas": "Ingresos de actividades ordinarias",
    "COGS": "Costo de ventas",
    "Bruta": "Ganancia bruta",
    "EBIT": "Ganancias (pérdidas) de actividades operacionales",
    "Neta": "Ganancia (pérdida)",
    "NetaControl": "Ganancia (pérdida), atribuible a los propietarios de la controladora",
    "Interes": "Costos financieros",
    "IngFinanc": "Ingresos financieros",
    "Dep": "Depreciación",
    "Amort": "Amortización",
    "DepAmort": "Depreciación y amortización",
    "OtrosIng": "Otros ingresos",
    "CostDistrib": "Costos de distribución",
    "GastAdmin": "Gastos de administración",
    "OtrosGast": "Otros gastos, por función",
    "OtraGanPerd": "Otras ganancias (pérdidas)",
    "AntesImp": "Ganancia (pérdida), antes de impuestos",
    "ImpGanan": "Gasto por impuestos a las ganancias",
    "OperCont": "Ganancia (pérdida) procedente de operaciones continuadas",
    "OperDisc": "Ganancia (pérdida) procedente de operaciones discontinuadas",
    "DeterioBanco": "Deterioro de valor de ganancias y reversión de pérdidas por deterioro de valor (pérdidas por deterioro de valor) determinado de acuerdo con la NIIF 9",
    "PartAsoc": "Participación en las ganancias (pérdidas) de asociadas y negocios conjuntos que se contabilicen utilizando el método de la participación",
    "GanCambio": "Ganancias (pérdidas) de cambio en moneda extranjera",
    "ResUniReaj": "Resultados por unidades de reajuste",
    "GanBasAcc": "Ganancia (pérdida) por acción básica",
    "GanDilAcc": "Ganancias (pérdida) diluida por acción",
    # Naturaleza ([320000]) para proxy COGS
    "RawMat": "Materias primas y consumibles utilizados",
    "InvChange": "Disminución (aumento) en inventarios de productos terminados y en proceso",
    "WorkCap": "Otros trabajos realizados por la entidad y capitalizados",
}

CFS_KEYS_TO_ES: Dict[str, str] = {
    "CFO": "Flujos de efectivo netos procedentes de (utilizados en) actividades de operación",
    "CFI": "Flujos de efectivo netos procedentes de (utilizados en) actividades de inversión",
    "CFF": "Flujos de efectivo netos procedentes de (utilizados en) actividades de financiación",
    "CapexBuy": "Compras de propiedades, planta y equipo",
    "CapexSale": "Importes procedentes de la venta de propiedades, planta y equipo",
    "IntangBuy": "Compras de activos intangibles",
    "IntangSale": "Importes procedentes de ventas de activos intangibles",
    "DivPag": "Dividendos pagados",
    "DivRec": "Dividendos recibidos",
    "IntPag": "Intereses pagados",
    "IntRec": "Intereses recibidos",
    "ImpPag": "Impuestos a las ganancias pagados (reembolsados)",
    "EmisPrest": "Importes procedentes de préstamos",
    "ReembPrest": "Reembolsos de préstamos",
    "EmisAcc": "Importes procedentes de la emisión de acciones",
    "CompAcc": "Pagos por adquirir o rescatar las acciones de la entidad",
    "PagArrend": "Pagos de pasivos por arrendamientos",
    "CobrosVenta": "Cobros procedentes de las ventas de bienes y prestación de servicios",
    "PagosProveed": "Pagos a proveedores por el suministro de bienes y servicios",
    "PagosEmpl": "Pagos a y por cuenta de los empleados",
    "EfecInic": "Efectivo y equivalentes al efectivo al principio del periodo",
    "EfecFinal": "Efectivo y equivalentes al efectivo al final del periodo",
    "VarEfec": "Incremento (disminución) neto de efectivo y equivalentes al efectivo",
}


def _detect_header_row(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and re.match(r"^\d{4}(|Q[1-4]|-\d{2}(-\d{2})?)$", v.strip()):
                return r
    return 3


def _normalize_label(raw: Any) -> Optional[str]:
    if not isinstance(raw, str):
        return None
    s = raw.strip().split("\n", 1)[0]
    # YYYYQn
    if re.match(r"^\d{4}Q[1-4]$", s):
        return s
    # YYYY-MM or YYYY-MM-DD → YYYYQn o YYYY
    m = re.match(r"^(\d{4})-(\d{2})(?:-\d{2})?$", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        q = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(mo)
        return f"{y}{q}" if q else str(y)
    # Bare YYYY -> YYYYQ4 for backward compat
    if re.match(r"^\d{4}$", s):
        return f"{s}Q4"
    return s


def _period_sort_key(lbl: str) -> Tuple[int, int]:
    s = str(lbl)
    if re.match(r"^\d{4}Q[1-4]$", s):
        return (int(s[:4]), int(s[5]))
    if re.match(r"^\d{4}$", s):
        return (int(s), 4)
    try:
        return (int(s[:4]), 5)
    except Exception:
        return (9999, 9)


def _first_non_abstract_match(ws, hdr_row: int, target: str) -> Optional[int]:
    """Busca fila cuyo nombre coincida con target (ES). Evita [abstract]/[resumen]/[sinopsis]."""
    target_lc = target.strip().lower()
    for r in range(hdr_row + 1, ws.max_row + 1):
        nm = ws.cell(row=r, column=1).value
        if not isinstance(nm, str):
            continue
        s = nm.strip().lower()
        if "[abstract]" in s or "[resumen]" in s or "[sinopsis]" in s:
            continue
        # match exacta o contains (cubre variantes menores)
        if s == target_lc or target_lc in s:
            return r
    return None


def _labels_to_col_map(ws, hdr_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        lb = _normalize_label(v)
        if lb:
            m[lb] = c
    return m


def _parse_number(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip()
        s = s.replace("\u00A0", " ").replace("−", "-").replace("–", "-")
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1]
        s = s.replace(",", "").replace(" ", "").replace(".", "")
        if s in ("", "-", "+"):
            return None
        num = float(s)
        if neg:
            num = -num
        return num
    except Exception:
        return None


def _find_es_counterpart(path_en: Path) -> Optional[Path]:
    name = path_en.name
    if name.endswith("_en.xlsx"):
        cand = path_en.with_name(name.replace("_en.xlsx", "_es.xlsx"))
        if cand.exists():
            return cand
    if "[EN].xlsx" in name:
        cand = path_en.with_name(name.replace("[EN].xlsx", "[ES].xlsx"))
        if cand.exists():
            return cand
    # Heurística Products: " - Financial Analysis" → " - Análisis Financiero" y marcador [EN]→[ES]
    if " - Financial Analysis" in name:
        cand = path_en.with_name(name.replace(" - Financial Analysis", " - Análisis Financiero").replace("[EN]", "[ES]"))
        if cand.exists():
            return cand
    # Caso por defecto: buscar archivo con mismo prefijo y [ES]
    for p in path_en.parent.glob("*.xlsx"):
        if "[ES]" in p.name and path_en.stem.split("[")[0].strip() in p.stem:
            return p
    return None


def prepare_es_fallback(wb: Workbook, fb, current_file: Path) -> bool:
    """
    Prepara la hoja 'ES_DATA' y configura FormulaBuilder `fb` para fallback a ES.

    Crea mapas:
      - fb.es_data_sheet_name = 'ES_DATA'
      - fb.es_label_to_col: label → col letra en ES_DATA
      - fb.es_row_for_key: (grupo, clave) → fila en ES_DATA
      - fb.es_values: (grupo, clave, label) → valor numérico

    Devuelve True si se configuró fallback; False si no aplica o no se encontró ES.
    """
    try:
        # Solo tiene sentido cuando el input es inglés
        nm = current_file.name.lower()
        is_en = nm.endswith("_en.xlsx") or "[en]" in nm or " financial analysis" in nm
        if not is_en:
            return False
        es_path = _find_es_counterpart(current_file)
        if not es_path:
            return False
        wb_es = load_workbook(str(es_path), read_only=True, data_only=True)
        # Hojas ES esperadas
        ws_bal = wb_es["Balance General"] if "Balance General" in wb_es.sheetnames else None
        ws_pl = None
        for nm in ("Estado de Resultados", "Estado Resultados (Función)"):
            if nm in wb_es.sheetnames:
                ws_pl = wb_es[nm]; break
        ws_cfs = wb_es["Flujo Efectivo"] if "Flujo Efectivo" in wb_es.sheetnames else None
        if not (ws_bal and ws_pl and ws_cfs):
            return False

        hdr_bal = _detect_header_row(ws_bal)
        hdr_pl = _detect_header_row(ws_pl)
        hdr_cfs = _detect_header_row(ws_cfs)

        # Mapas etiqueta→columna índice por hoja ES
        cols_bal = _labels_to_col_map(ws_bal, hdr_bal)
        cols_pl = _labels_to_col_map(ws_pl, hdr_pl)
        cols_cfs = _labels_to_col_map(ws_cfs, hdr_cfs)

        labels_all = sorted(set(cols_bal) | set(cols_pl) | set(cols_cfs), key=_period_sort_key)
        if not labels_all:
            return False

        # Construir filas clave (grupo, clave) → fila en ES por búsqueda de concepto ES
        def build_row_map(ws, hdr, mapping: Dict[str, str]) -> Dict[str, Optional[int]]:
            out: Dict[str, Optional[int]] = {}
            for key, es_name in mapping.items():
                out[key] = _first_non_abstract_match(ws, hdr, es_name)
            return out

        rows_bal = build_row_map(ws_bal, hdr_bal, BALANCE_KEYS_TO_ES)
        rows_pl = build_row_map(ws_pl, hdr_pl, INCOME_KEYS_TO_ES)
        rows_cfs = build_row_map(ws_cfs, hdr_cfs, CFS_KEYS_TO_ES)

        # Crear hoja ES_DATA oculta en wb
        es_ws_name = "ES_DATA"
        if es_ws_name in wb.sheetnames:
            del wb[es_ws_name]
        ws_data = wb.create_sheet(es_ws_name)
        ws_data.sheet_state = "hidden"

        # Encabezados: GRUPO | KEY | Concepto | labels...
        ws_data.cell(row=1, column=1, value="GRUPO")
        ws_data.cell(row=1, column=2, value="KEY")
        ws_data.cell(row=1, column=3, value="Concepto")
        for i, lb in enumerate(labels_all, start=4):
            ws_data.cell(row=1, column=i, value=lb)
        es_label_to_col = {lb: get_column_letter(4 + idx) for idx, lb in enumerate(labels_all)}

        # Dump valores a ES_DATA y construir mapas
        es_row_for_key: Dict[Tuple[str, str], int] = {}
        es_values: Dict[Tuple[str, str, str], float] = {}

        def _write_group(group: str, ws_src, hdr_src: int, rows_map: Dict[str, Optional[int]], cols_map: Dict[str, int], names_map: Dict[str, str]):
            nonlocal es_row_for_key, es_values
            for key, rownum in rows_map.items():
                if not rownum:
                    continue
                out_row = ws_data.max_row + 1
                es_row_for_key[(group, key)] = out_row
                # Meta columnas
                ws_data.cell(row=out_row, column=1, value=group)
                ws_data.cell(row=out_row, column=2, value=key)
                ws_data.cell(row=out_row, column=3, value=names_map.get(key))
                # Periodos
                for lb in labels_all:
                    cidx = cols_map.get(lb)
                    # alias YYYY ↔ YYYYQ4
                    if not cidx:
                        m = re.match(r"^(\d{4})Q4$", lb)
                        if m:
                            cidx = cols_map.get(m.group(1))
                    if not cidx:
                        m2 = re.match(r"^(\d{4})$", lb)
                        if m2:
                            cidx = cols_map.get(f"{m2.group(1)}Q4")
                    val = ws_src.cell(row=rownum, column=cidx).value if cidx else None
                    num = _parse_number(val)
                    if num is not None:
                        ws_data.cell(row=out_row, column=3 + 1 + labels_all.index(lb), value=num)
                        es_values[(group, key, lb)] = float(num)
                        # duplicar alias
                        m = re.match(r"^(\d{4})Q4$", lb)
                        if m:
                            es_values[(group, key, m.group(1))] = float(num)
                        m2 = re.match(r"^(\d{4})$", lb)
                        if m2:
                            es_values[(group, key, f"{m2.group(1)}Q4")] = float(num)

        _write_group("BAL", ws_bal, hdr_bal, rows_bal, cols_bal, BALANCE_KEYS_TO_ES)
        _write_group("PL", ws_pl, hdr_pl, rows_pl, cols_pl, INCOME_KEYS_TO_ES)
        _write_group("CFS", ws_cfs, hdr_cfs, rows_cfs, cols_cfs, CFS_KEYS_TO_ES)

        # Configurar FormulaBuilder
        fb.es_fallback_active = True
        fb.es_data_sheet_name = es_ws_name
        fb.es_label_to_col = es_label_to_col
        fb.es_row_for_key = es_row_for_key
        fb.es_values = es_values
        return True
    except Exception:
        return False


