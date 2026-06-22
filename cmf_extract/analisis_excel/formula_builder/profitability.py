"""
Profitability Mixin
===================

Provides methods for building profitability ratio formulas,
including TTM (Trailing Twelve Months) helpers.
"""

import re
from typing import List, Optional, Tuple


class ProfitabilityMixin:
    """Mixin providing profitability ratio formula builders and TTM helpers."""

    def build_profitability_formulas(self) -> List[Tuple[str, str, callable, str]]:
        """Construye fórmulas para ratios de rentabilidad."""
        formulas = []

        # Helpers TTM: dados row en P&L y label Qn, TTM = YTD(Qn) - YTD(Qn-4)
        # Q4 is already annual data — no TTM adjustment needed.
        def ttm_ref(sheet, row: Optional[int], label_q: str) -> Optional[str]:
            try:
                # YTD actual
                cur = self.create_cell_reference_by_label(sheet, row, label_q)
                # Q4 is the full-year value: TTM = annual, no subtraction
                if label_q.endswith("Q4"):
                    return cur
                # YTD hace 4 trimestres
                prev4 = self._previous_year_same_quarter_label(label_q)
                if prev4:
                    prev = self.create_cell_reference_by_label(sheet, row, prev4)
                else:
                    prev = None
                if cur and prev:
                    return f"IFERROR({cur}-{prev},{cur})"
                return cur
            except Exception:
                return None

        # Exponer helper a otras secciones
        self.ttm_ref_pl = lambda key, lb: ttm_ref(self.sh_pl, self.rows_pl.get(key), lb) if self.rows_pl.get(key) else None
        self.ttm_ref_cfs = lambda key, lb: ttm_ref(self.sh_cfs, self.rows_cfs.get(key), lb) if self.rows_cfs.get(key) else None

        # Determinar etiquetas visibles (YYYY o YYYYQn) desde la hoja de P&L
        def visible_period_labels(sheet) -> List[str]:
            hdr = self.HDR_PL
            labels = []
            for c in range(2, sheet.max_column + 1):
                v = sheet.cell(row=hdr, column=c).value
                if isinstance(v, str):
                    labels.append(v.strip().split("\n", 1)[0])
            return labels

        labels_pl = visible_period_labels(self.sh_pl)
        has_quarters_in_pl = any(re.match(r"^\d{4}Q[1-4]$", lb) for lb in labels_pl)

        # Margen Bruto (TTM si hay quarters, si no anual)
        def f_margen_bruto():
            m = {}
            is_quarter = has_quarters_in_pl  # Siempre procesar trimestres cuando están disponibles


            # Si estamos usando vista por naturaleza ([320000]), calcular margen bruto como:
            # Usar exactamente la misma lógica que funciona para años, solo cambiar la columna
            if self._is_using_nature_based_income_statement():
                if is_quarter:
                    for lb in labels_pl:
                        if not re.match(r"^\d{4}Q[1-4]$", lb):
                            continue

                        # Encontrar la columna correspondiente al trimestre
                        col = None
                        for c in range(2, self.sh_pl.max_column + 1):
                            v = self.sh_pl.cell(row=self.HDR_PL, column=c).value
                            if isinstance(v, str) and v.strip().split("\n", 1)[0] == lb:
                                col = c
                                break

                        if not col:
                            continue

                        # Convertir número de columna a letra (B=2, C=3, etc.)
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col)

                        # Usar EXACTAMENTE la misma lógica que funciona para años
                        ven = self.create_cell_reference(self.sh_pl.title, col_letter, self.rows_pl["Ventas"])

                        # Construir misma fórmula exacta para trimestres
                        parts = []

                        # RawMat (Materias primas)
                        if self.rows_pl.get("RawMat"):
                            raw_mat = self.create_cell_reference(self.sh_pl.title, col_letter, self.rows_pl["RawMat"])
                            if raw_mat:
                                parts.append(f"IFERROR({raw_mat},0)")

                        # Gastos por beneficios empleados
                        emp_benefits_row = self._find_row_by_regex_in_pl(r"gastos.*beneficios.*empleados|beneficios.*empleados", exclude_abstract=True)
                        if emp_benefits_row:
                            emp_ben = self.create_cell_reference(self.sh_pl.title, col_letter, emp_benefits_row)
                            if emp_ben:
                                parts.append(f"IFERROR({emp_ben},0)")

                        # Depreciación y Amortización (usar función de año pero con columna específica)
                        da_row = self._find_row_by_regex_in_pl(r"gasto.*depreciaci.*amortizaci", exclude_abstract=True)
                        if da_row:
                            da_ref = self.create_cell_reference(self.sh_pl.title, col_letter, da_row)
                            if da_ref:
                                parts.append(f"IFERROR({da_ref},0)")

                        # Otros gastos por naturaleza
                        other_exp_row = self._find_row_by_regex_in_pl(r"otros gastos.*naturaleza", exclude_abstract=True)
                        if other_exp_row:
                            other_exp = self.create_cell_reference(self.sh_pl.title, col_letter, other_exp_row)
                            if other_exp:
                                parts.append(f"IFERROR({other_exp},0)")

                        # Cambio en inventarios
                        if self.rows_pl.get("InvChange"):
                            inv_change = self.create_cell_reference(self.sh_pl.title, col_letter, self.rows_pl["InvChange"])
                            if inv_change:
                                parts.append(f"IFERROR({inv_change},0)")

                        # Trabajos capitalizados (se resta)
                        if self.rows_pl.get("WorkCap"):
                            work_cap = self.create_cell_reference(self.sh_pl.title, col_letter, self.rows_pl["WorkCap"])
                            if work_cap:
                                parts.append(f"-IFERROR({work_cap},0)")

                        if ven and parts:
                            operating_costs = "+".join(parts)
                            m[lb] = f"IFERROR(({ven}-(({operating_costs})))/{ven},\"N/A\")"
                    # Completar claves anuales
                    for y in self.years:
                        if str(y) in m:
                            continue
                        ven_a = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Ventas"]
                        )

                        # Construir misma fórmula exacta para años
                        parts_a = []
                        col = self.find_year_column(self.sh_pl, y)

                        # RawMat (Materias primas)
                        if self.rows_pl.get("RawMat"):
                            raw_mat_a = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["RawMat"])
                            if raw_mat_a:
                                parts_a.append(f"IFERROR({raw_mat_a},0)")

                        # Gastos por beneficios empleados
                        emp_benefits_row = self._find_row_by_regex_in_pl(r"gastos.*beneficios.*empleados|beneficios.*empleados", exclude_abstract=True)
                        if emp_benefits_row:
                            emp_ben_a = self.create_cell_reference(self.sh_pl.title, col, emp_benefits_row)
                            if emp_ben_a:
                                parts_a.append(f"IFERROR({emp_ben_a},0)")

                        # Depreciación y Amortización
                        da_a = self._build_da_annual(y)
                        if da_a:
                            parts_a.append(f"IFERROR({da_a},0)")

                        # Otros gastos por naturaleza
                        other_exp_row = self._find_row_by_regex_in_pl(r"otros gastos.*naturaleza", exclude_abstract=True)
                        if other_exp_row:
                            other_exp_a = self.create_cell_reference(self.sh_pl.title, col, other_exp_row)
                            if other_exp_a:
                                parts_a.append(f"IFERROR({other_exp_a},0)")

                        # Cambio en inventarios
                        if self.rows_pl.get("InvChange"):
                            inv_change_a = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["InvChange"])
                            if inv_change_a:
                                parts_a.append(f"IFERROR({inv_change_a},0)")

                        # Trabajos capitalizados (se resta)
                        if self.rows_pl.get("WorkCap"):
                            work_cap_a = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["WorkCap"])
                            if work_cap_a:
                                parts_a.append(f"-IFERROR({work_cap_a},0)")

                        if ven_a:
                            if parts_a:
                                operating_costs_a = "+".join(parts_a)
                                m[str(y)] = f"IFERROR(({ven_a}-(({operating_costs_a})))/{ven_a},\"N/A\")"
                            else:
                                # Si no hay parts, usar solo ventas (margen bruto = 100%)
                                m[str(y)] = f"IFERROR({ven_a}/{ven_a},\"N/A\")"
                else:
                    for y in self.years:
                        ven = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Ventas"]
                        )

                        # Construir misma fórmula exacta para años (sin quarters)
                        parts = []
                        col = self.find_year_column(self.sh_pl, y)

                        # RawMat (Materias primas)
                        if self.rows_pl.get("RawMat"):
                            raw_mat = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["RawMat"])
                            if raw_mat:
                                parts.append(f"IFERROR({raw_mat},0)")

                        # Gastos por beneficios empleados
                        emp_benefits_row = self._find_row_by_regex_in_pl(r"gastos.*beneficios.*empleados|beneficios.*empleados", exclude_abstract=True)
                        if emp_benefits_row:
                            emp_ben = self.create_cell_reference(self.sh_pl.title, col, emp_benefits_row)
                            if emp_ben:
                                parts.append(f"IFERROR({emp_ben},0)")

                        # Depreciación y Amortización
                        da = self._build_da_annual(y)
                        if da:
                            parts.append(f"IFERROR({da},0)")

                        # Otros gastos por naturaleza
                        other_exp_row = self._find_row_by_regex_in_pl(r"otros gastos.*naturaleza", exclude_abstract=True)
                        if other_exp_row:
                            other_exp = self.create_cell_reference(self.sh_pl.title, col, other_exp_row)
                            if other_exp:
                                parts.append(f"IFERROR({other_exp},0)")

                        # Cambio en inventarios
                        if self.rows_pl.get("InvChange"):
                            inv_change = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["InvChange"])
                            if inv_change:
                                parts.append(f"IFERROR({inv_change},0)")

                        # Trabajos capitalizados (se resta)
                        if self.rows_pl.get("WorkCap"):
                            work_cap = self.create_cell_reference(self.sh_pl.title, col, self.rows_pl["WorkCap"])
                            if work_cap:
                                parts.append(f"-IFERROR({work_cap},0)")

                        if ven:
                            if parts:
                                operating_costs = "+".join(parts)
                                m[str(y)] = f"IFERROR(({ven}-(({operating_costs})))/{ven},\"N/A\")"
                            else:
                                # Si no hay parts, usar solo ventas (margen bruto = 100%)
                                m[str(y)] = f"IFERROR({ven}/{ven},\"N/A\")"
            else:
                # Método estándar usando Ganancia Bruta
                if is_quarter:
                    for lb in labels_pl:
                        if not re.match(r"^\d{4}Q[1-4]$", lb):
                            continue
                        gb = ttm_ref(self.sh_pl, self.rows_pl["Bruta"], lb)
                        ven = self.ttm_ref_pl("Ventas", lb)
                        if gb and ven:
                            m[lb] = f"IFERROR({gb}/{ven},\"N/A\")"
                    # Además, completar claves anuales (YYYY) usando referencias anuales
                    for y in self.years:
                        if str(y) in m:
                            continue
                        gb_a = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Bruta"]
                        )
                        ven_a = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Ventas"]
                        )
                        if gb_a and ven_a:
                            m[str(y)] = f"IFERROR({gb_a}/{ven_a},\"N/A\")"
                else:
                    for y in self.years:
                        gb = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Bruta"]
                        )
                        ven = self.create_cell_reference(
                            self.sh_pl.title,
                            self.find_year_column(self.sh_pl, y),
                            self.rows_pl["Ventas"]
                        )
                        if gb and ven:
                            m[str(y)] = f"IFERROR({gb}/{ven},\"N/A\")"
            return m

        formulas.append(("Margen Bruto", "pct", f_margen_bruto,
                        "Utilidad Bruta / Ventas"))

        # Margen Operativo (EBIT) (TTM en trimestral para vista funcional, directo para naturaleza)
        def f_margen_operativo():
            m = {}
            is_quarter = has_quarters_in_pl
            if is_quarter:
                for lb in labels_pl:
                    if not re.match(r"^\d{4}Q[1-4]$", lb):
                        continue

                    # Para vista por naturaleza, usar referencias directas al trimestre
                    if self._is_using_nature_based_income_statement():
                        ebit = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["EBIT"], lb)
                        ven = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], lb)
                    else:
                        # Para vista funcional, usar TTM para EBIT pero ventas del mismo período
                        ebit = ttm_ref(self.sh_pl, self.rows_pl["EBIT"], lb)
                        ven = self.create_cell_reference_by_label(self.sh_pl, self.rows_pl["Ventas"], lb)

                    if ebit and ven:
                        m[lb] = f"IFERROR({ebit}/{ven},\"N/A\")"
                for y in self.years:
                    if str(y) in m:
                        continue
                    ebit_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["EBIT"]
                    )
                    ven_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if ebit_a and ven_a:
                        m[str(y)] = f"IFERROR({ebit_a}/{ven_a},\"N/A\")"
            else:
                for y in self.years:
                    ebit = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["EBIT"]
                    )
                    ven = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if ebit and ven:
                        m[str(y)] = f"IFERROR({ebit}/{ven},\"N/A\")"
            return m

        formulas.append(("Margen Operativo (EBIT)", "pct", f_margen_operativo,
                        "EBIT / Ventas"))

        # Margen EBITDA
        def f_margen_ebitda():
            m = {}
            is_quarter = has_quarters_in_pl
            if is_quarter:
                for lb in labels_pl:
                    if not re.match(r"^\d{4}Q[1-4]$", lb):
                        continue

                    # SOLUCIÓN FINAL: Usar EBITDA ya calculado (M50) y ventas DIRECTAS del período

                    # Encontrar ventas del período SIN TTM (método manual más básico)
                    ven_col = None
                    for c in range(2, self.sh_pl.max_column + 1):
                        cell_val = self.sh_pl.cell(row=self.HDR_PL, column=c).value
                        if cell_val and str(cell_val).strip() == lb:
                            ven_col = c
                            break

                    if ven_col:
                        # Referencia DIRECTA a ventas del período (ej: C7 para 2025Q1)
                        ven = self.create_cell_reference(self.sh_pl.title, ven_col, self.rows_pl["Ventas"])

                        # Usar EBITDA ya calculado (fila anterior M50)
                        ebitda_ref = "INDIRECT(ADDRESS(ROW()-1,COLUMN()))"

                        m[lb] = f"IFERROR({ebitda_ref}/{ven},\"N/A\")"

                # Para años anuales (sin cambios)
                for y in self.years:
                    if str(y) in m:
                        continue
                    ebit_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["EBIT"]
                    )
                    da_a = self._build_da_annual(y)
                    ven_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if ebit_a and ven_a:
                        ebitda_a = f"({ebit_a}+({da_a}))" if da_a else f"({ebit_a})"
                        m[str(y)] = f"IFERROR({ebitda_a}/{ven_a},\"N/A\")"
            else:
                # Solo datos anuales
                for y in self.years:
                    ebit = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["EBIT"]
                    )
                    da = self._build_da_annual(y)
                    ven = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if ebit and ven:
                        if da:
                            ebitda = f"({ebit}+({da}))"
                        else:
                            ebitda = f"({ebit})"
                        m[str(y)] = f"IFERROR({ebitda}/{ven},\"N/A\")"
            return m

        formulas.append(("Margen EBITDA", "pct", f_margen_ebitda,
                        "EBITDA / Ventas"))

        # Margen Neto
        def f_margen_neto():
            m = {}
            is_quarter = has_quarters_in_pl
            if is_quarter:
                for lb in labels_pl:
                    if not re.match(r"^\d{4}Q[1-4]$", lb):
                        continue
                    net = ttm_ref(self.sh_pl, self.rows_pl["Neta"], lb)
                    ven = self.ttm_ref_pl("Ventas", lb)
                    if net and ven:
                        m[lb] = f"IFERROR({net}/{ven},\"N/A\")"
                for y in self.years:
                    if str(y) in m:
                        continue
                    net_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    ven_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if net_a and ven_a:
                        m[str(y)] = f"IFERROR({net_a}/{ven_a},\"N/A\")"
            else:
                for y in self.years:
                    net = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    ven = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Ventas"]
                    )
                    if net and ven:
                        m[str(y)] = f"IFERROR({net}/{ven},\"N/A\")"
            return m

        formulas.append(("Margen Neto", "pct", f_margen_neto,
                        "Utilidad Neta / Ventas"))

        # (Eliminado: EBITDA en Rentabilidad; se calculará en sección UTILIDADES)

        # ROE usando promedio del patrimonio (TTM en trimestral con promedio Patr [Qn, Qn-4])
        def f_roe():
            m = {}
            labels_pl = []
            hdr = self.HDR_PL
            for c in range(2, self.sh_pl.max_column + 1):
                v = self.sh_pl.cell(row=hdr, column=c).value
                if isinstance(v, str):
                    labels_pl.append(v.strip().split("\n", 1)[0])
            is_quarter = any(re.match(r"^\d{4}Q[1-4]$", lb) for lb in labels_pl)
            if is_quarter:
                for lb in labels_pl:
                    if not re.match(r"^\d{4}Q[1-4]$", lb):
                        continue
                    net = self.ttm_ref_pl("Neta", lb)
                    patr_now = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal["Patr"], lb)
                    lb_prev4 = self._previous_year_same_quarter_label(lb)
                    patr_prev4 = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal["Patr"], lb_prev4) if lb_prev4 else None
                    patr_avg = f"AVERAGE({patr_now},{patr_prev4})" if patr_now and patr_prev4 else (patr_now or patr_prev4)
                    if net and patr_avg:
                        m[lb] = f"IF({patr_avg}<=0,\"N/A\",IFERROR({net}/{patr_avg},\"N/A\"))"
                for y in self.years:
                    if str(y) in m:
                        continue
                    net_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    patr_avg_a = self.create_average_reference(self.sh_bal, self.rows_bal["Patr"], y)
                    if net_a and patr_avg_a:
                        m[str(y)] = f"IF({patr_avg_a}<=0,\"N/A\",IFERROR({net_a}/{patr_avg_a},\"N/A\"))"
            else:
                for y in self.years:
                    net = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    patr_avg = self.create_average_reference(self.sh_bal, self.rows_bal["Patr"], y)
                    if net and patr_avg:
                        m[str(y)] = f"IF({patr_avg}<=0,\"N/A\",IFERROR({net}/{patr_avg},\"N/A\"))"
            return m

        formulas.append(("ROE", "pct", f_roe,
                        "Utilidad Neta (TTM/anual) / Patrimonio Promedio"))

        # ROA usando promedio de activos totales (TTM en trimestral con promedio AT [Qn, Qn-4])
        def f_roa():
            m = {}
            labels_pl = []
            hdr = self.HDR_PL
            for c in range(2, self.sh_pl.max_column + 1):
                v = self.sh_pl.cell(row=hdr, column=c).value
                if isinstance(v, str):
                    labels_pl.append(v.strip().split("\n", 1)[0])
            is_quarter = any(re.match(r"^\d{4}Q[1-4]$", lb) for lb in labels_pl)
            if is_quarter:
                for lb in labels_pl:
                    if not re.match(r"^\d{4}Q[1-4]$", lb):
                        continue
                    net = self.ttm_ref_pl("Neta", lb)
                    at_now = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal["AT"], lb)
                    lb_prev4 = self._previous_year_same_quarter_label(lb)
                    at_prev4 = self.create_cell_reference_by_label(self.sh_bal, self.rows_bal["AT"], lb_prev4) if lb_prev4 else None
                    at_avg = f"AVERAGE({at_now},{at_prev4})" if at_now and at_prev4 else (at_now or at_prev4)
                    if net and at_avg:
                        m[lb] = f"IFERROR({net}/{at_avg},\"N/A\")"
                for y in self.years:
                    if str(y) in m:
                        continue
                    net_a = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    at_avg_a = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                    if net_a and at_avg_a:
                        m[str(y)] = f"IFERROR({net_a}/{at_avg_a},\"N/A\")"
            else:
                for y in self.years:
                    net = self.create_cell_reference(
                        self.sh_pl.title,
                        self.find_year_column(self.sh_pl, y),
                        self.rows_pl["Neta"]
                    )
                    at_avg = self.create_average_reference(self.sh_bal, self.rows_bal["AT"], y)
                    if net and at_avg:
                        m[str(y)] = f"IFERROR({net}/{at_avg},\"N/A\")"
            return m

        formulas.append(("ROA", "pct", f_roa,
                        "Utilidad Neta (TTM/anual) / Activos Totales Promedio"))

        return formulas
