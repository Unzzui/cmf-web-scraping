"""
Formula Builder Core
====================

Main FormulaBuilder class that combines all mixins via multiple inheritance.
"""

import re
import os
from typing import Dict, List, Optional, Tuple, Any
from ..utils.lang_map import load_mappings, guess_is_english
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from .cell_references import CellReferenceMixin
from .row_mappers import RowMapperMixin
from .liquidity import LiquidityMixin
from .solvency import SolvencyMixin
from .profitability import ProfitabilityMixin
from .efficiency import EfficiencyMixin
from .cash_flow import CashFlowMixin
from .value_creation import ValueCreationMixin
from .coverage_risk import CoverageRiskMixin
from .nature_based import NatureBasedMixin
from .derived_ratios import DerivedRatiosMixin


class FormulaBuilder(
    CellReferenceMixin,
    RowMapperMixin,
    LiquidityMixin,
    SolvencyMixin,
    ProfitabilityMixin,
    EfficiencyMixin,
    CashFlowMixin,
    ValueCreationMixin,
    CoverageRiskMixin,
    NatureBasedMixin,
    DerivedRatiosMixin,
):
    """
    Constructor de fórmulas Excel para análisis financiero.
    """

    def __init__(self, workbook: Workbook, financial_data: Dict[str, Any]):
        """
        Inicializa el constructor de fórmulas.

        Args:
            workbook: Libro de Excel abierto
            financial_data: Datos financieros extraídos
        """
        self.wb = workbook
        self.financial_data = financial_data
        self.years = financial_data.get("years", [])
        # Si se establece, fuerza a usar una etiqueta de período exacta (p. ej., '2024Q1')
        # cuando se busque la columna correspondiente en find_year_column
        self.period_override_label: Optional[str] = None

        # Obtener referencias a las hojas (compat es/en y nombres nuevos)
        def pick_sheet(wb: Workbook, names: list[str]):
            for n in names:
                if n in wb.sheetnames:
                    return wb[n]
            raise KeyError(f"No se encontró ninguna de las hojas: {names}")

        self.sh_bal = pick_sheet(workbook, ["Balance General", "Balance Sheet"])
        self.sh_pl  = pick_sheet(workbook, ["Estado de Resultados", "Estado Resultados (Función)", "Income Statement"])
        self.sh_cfs = pick_sheet(workbook, ["Flujo Efectivo", "Cash Flow"])

        # Detectar fila de encabezados por hoja (nuestro writer usa fila 3)
        def detect_header_row(sheet) -> int:
            import re
            max_scan = min(10, sheet.max_row)
            for r in range(1, max_scan + 1):
                v0 = sheet.cell(row=r, column=1).value
                if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
                    return r
                for c in range(2, min(sheet.max_column, 20) + 1):
                    v = sheet.cell(row=r, column=c).value
                    if isinstance(v, str):
                        s = v.strip().split("\n", 1)[0]
                        if re.match(r"^\d{4}-(\d{2}|\d{2}-\d{2})$", s):
                            return r
                        if re.match(r"^\d{4}(?:Q[1-4])?$", s):
                            return r
            return 3

        self.HDR_BAL = detect_header_row(self.sh_bal)
        self.HDR_PL  = detect_header_row(self.sh_pl)
        self.HDR_CFS = detect_header_row(self.sh_cfs)

        # Mapear conceptos a filas (después de detectar headers)
        self.rows_bal = self._map_balance_rows()
        self.rows_pl = self._map_income_rows()
        self.rows_cfs = self._map_cash_flow_rows()

        # Cargar mapeos ES<->EN para normalizar búsquedas en inglés
        self.es_to_en, self.en_to_es = load_mappings()

        # Inversos fila→clave para identificar el concepto a partir del row_num
        self._row_to_key_bal: Dict[int, str] = {row: key for key, row in self.rows_bal.items() if row}
        self._row_to_key_pl: Dict[int, str] = {row: key for key, row in self.rows_pl.items() if row}
        self._row_to_key_cfs: Dict[int, str] = {row: key for key, row in self.rows_cfs.items() if row}

        # Fallback a datos ES dentro del mismo libro (inyectados por el caller)
        # - Nombre de hoja oculta con valores ES
        # - Mapa etiqueta→columna (letra) dentro de esa hoja
        # - Mapa (grupo, clave)→fila dentro de esa hoja
        # - Mapa (grupo, clave, etiqueta)→valor numérico ES (para verificación dif)
        # Nota: grupo en {"BAL","PL","CFS"}
        self.es_data_sheet_name: Optional[str] = None
        self.es_label_to_col: Dict[str, str] = {}
        self.es_row_for_key: Dict[Tuple[str, str], int] = {}
        self.es_values: Dict[Tuple[str, str, str], float] = {}
        self.es_fallback_active: bool = False  # habilitado si el caller configura datos ES
        # Log de uso de ES_DATA por referencia
        self.es_usage_log_entries: List[str] = []

    def _find_row_in_sheet(self, sheet, df, concept_name: str) -> Optional[int]:
        """
        Encuentra la fila de un concepto en una hoja específica.

        Args:
            sheet: Hoja de Excel
            df: DataFrame correspondiente
            concept_name: Nombre del concepto a buscar

        Returns:
            Número de fila (1-based) o None si no se encuentra
        """
        # Normalizar concepto si la hoja está en inglés (conservar original por si se necesita fallback)
        series_concepts = df["Concepto"].astype(str)
        original_concept = concept_name
        if guess_is_english(series_concepts.tolist()):
            # Intentar convertir concepto ES -> EN antes de buscar
            for group_map in self.es_to_en.values():
                if concept_name in group_map:
                    concept_name = group_map[concept_name]
                    break
        # Búsqueda exacta
        mask = series_concepts.str.strip().str.lower() == concept_name.strip().lower()
        if mask.any():
            idx = df.index[mask][0]
            # Excel row = header_row + 1 (primera fila de datos) + idx (0-based)
            hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
            return int(idx) + hdr + 1

        # Búsqueda por contiene (prefiere filas no-abstract)
        mask = series_concepts.str.contains(re.escape(concept_name), case=False, na=False)
        if mask.any():
            candidate_indices = list(df.index[mask])
            hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
            # elegir la primera que no sea [abstract]/[resumen]
            for idx in candidate_indices:
                txt = str(df.loc[idx, "Concepto"]).strip().lower()
                if ("abstract" not in txt) and ("resumen" not in txt) and ("sinopsis" not in txt):
                    return int(idx) + hdr + 1
            # si todas son abstract/resumen, usar la primera
            idx = candidate_indices[0]
            return int(idx) + hdr + 1

        # Fallback por sinónimos comunes (ES), útil para totalizadores IFRS
        try:
            synonyms: dict[str, list[str]] = {
                "Activos corrientes totales": [r"^total .*activos corrientes", r"^total de activos corrientes"],
                "Pasivos corrientes totales": [r"^total .*pasivos corrientes", r"^total de pasivos corrientes"],
                "Total de activos": [r"^total .*activos$", r"^activos totales$"],
                "Total de pasivos": [r"^total .*pasivos$", r"^pasivos totales$"],
                "Total de activos no corrientes": [r"^total .*activos no corrientes"],
                "Total de pasivos no corrientes": [r"^total .*pasivos no corrientes"],
                "Patrimonio atribuible a los propietarios de la controladora": [r"^patrimonio.*propietarios de la controladora"],
            }
            patterns = synonyms.get(concept_name, [])
            if patterns:
                ser_lc = series_concepts.astype(str).str.strip().str.lower()
                hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
                for pat in patterns:
                    msk2 = ser_lc.str.contains(pat, regex=True, na=False)
                    if msk2.any():
                        # preferir no-abstract
                        cand = list(df.index[msk2])
                        for idx in cand:
                            txt = str(df.loc[idx, "Concepto"]).strip().lower()
                            if ("abstract" not in txt) and ("resumen" not in txt) and ("sinopsis" not in txt):
                                return int(idx) + hdr + 1
                        return int(cand[0]) + hdr + 1
        except Exception:
            pass

        # Fallback específico EN: buscar patrones comunes (sin depender de heurística de idioma)
        try:
            cn_lc = concept_name.strip().lower()
            english_series = series_concepts.astype(str).str.strip().str.lower()
            def _find_regex(pattern: str) -> Optional[int]:
                msk = english_series.str.contains(pattern, regex=True, na=False)
                if msk.any():
                    idx2 = df.index[msk][0]
                    hdr2 = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
                    return int(idx2) + hdr2 + 1
                return None
            if cn_lc in ("depreciación", "depreciacion"):

                row = _find_regex(r"\bdepreciation\b")
                if row:
                    return row
            if cn_lc in ("amortización", "amortizacion"):
                row = _find_regex(r"\bamorti[sz]ation\b")
                if row:
                    return row
            if cn_lc in ("depreciación y amortización", "depreciacion y amortizacion"):
                row = _find_regex(r"depreciation.*amorti[sz]ation|amorti[sz]ation.*depreciation")
                if row:
                    return row
        except Exception:
            pass

        # Fallback final: escanear directamente la hoja buscando coincidencia textual no-abstracta
        try:
            hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
            candidates = []
            try:
                candidates.append(concept_name.strip())
            except Exception:
                pass
            try:
                candidates.append(original_concept.strip())
            except Exception:
                pass
            # También intentar traducir EN->ES si corresponde
            try:
                for es_map in self.en_to_es.values():
                    if concept_name in es_map:
                        candidates.append(es_map[concept_name])
                        break
            except Exception:
                pass
            # Buscar por cualquiera de los candidatos
            cand_lc = [c.lower() for c in candidates if isinstance(c, str)]
            for r in range(hdr + 1, sheet.max_row + 1):
                nm = sheet.cell(row=r, column=1).value
                if not isinstance(nm, str):
                    continue
                s = nm.strip().lower()
                if ("abstract" in s) or ("resumen" in s) or ("sinopsis" in s):
                    continue
                for t in cand_lc:
                    if (s == t) or (t in s):
                        return r
        except Exception:
            pass
        return None

    def build_all_formulas(self) -> List[Tuple[str, List]]:
        """
        Construye todas las fórmulas organizadas por categoría.

        Returns:
            Lista de tuplas (categoría, lista_fórmulas)
        """
        return [
            ("LIQUIDEZ", self.build_liquidity_formulas()),
            ("SOLVENCIA Y ESTRUCTURA", self.build_solvency_formulas()),
            ("RENTABILIDAD", self.build_profitability_formulas()),
            ("EFICIENCIA OPERATIVA", self.build_efficiency_formulas()),
            ("FLUJOS Y ADICIONALES", self.build_cash_flow_formulas()),
            ("CREACIÓN DE VALOR", self.build_value_creation_formulas()),
            ("COBERTURA Y RIESGO", self.build_coverage_risk_formulas()),
            ("NATURALEZA", self.build_nature_based_formulas()),
        ]
