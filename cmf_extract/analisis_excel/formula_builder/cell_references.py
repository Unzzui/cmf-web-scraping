"""
Cell Reference Mixin
====================

Provides methods for finding columns and creating cell references
in Excel workbooks.
"""

import re
from typing import Dict, List, Optional, Tuple, Any
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


class CellReferenceMixin:
    """Mixin providing cell reference creation and column lookup methods."""

    def find_year_column(self, sheet, year: int) -> Optional[str]:
        """
        Encuentra la letra de columna para un año específico.

        Args:
            sheet: Hoja de Excel
            year: Año a buscar

        Returns:
            Letra de columna o None si no se encuentra
        """
        hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS

        # Si hay override de período exacto, intentar encontrarlo directamente
        if self.period_override_label:
            label_norm = self.period_override_label.strip().split("\n", 1)[0]
            for col in range(2, sheet.max_column + 1):
                val = sheet.cell(row=hdr, column=col).value
                if isinstance(val, str) and val.strip().split("\n", 1)[0] == label_norm:
                    return get_column_letter(col)

        def _match(v: str) -> bool:
            s = v.strip().split("\n", 1)[0]
            if s.startswith(f"{year}-"):
                return True
            if s == str(year):
                return True
            if re.match(rf"^{year}Q[1-4]$", s):
                return True
            return False

        # Preferencia: 'YYYYQ4' → 'YYYY' → 'YYYYQ3' → 'YYYYQ2' → 'YYYYQ1' → 'YYYY-MM'
        candidates = {"YYYY": None, "Q4": None, "Q3": None, "Q2": None, "Q1": None, "DATE": None}
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=hdr, column=col).value
            if val is None:
                continue

            # Manejar tanto strings como enteros (para años históricos en archivos especiales)
            if isinstance(val, int):
                if val == year:
                    candidates["YYYY"] = col
                continue
            elif isinstance(val, str):
                s = val.strip().split("\n", 1)[0]
                if s == str(year):
                    candidates["YYYY"] = col
                elif re.match(rf"^{year}Q4$", s):
                    candidates["Q4"] = col
                elif re.match(rf"^{year}Q3$", s):
                    candidates["Q3"] = col
                elif re.match(rf"^{year}Q2$", s):
                    candidates["Q2"] = col
                elif re.match(rf"^{year}Q1$", s):
                    candidates["Q1"] = col
                elif s.startswith(f"{year}-"):
                    candidates["DATE"] = col

        for key in ["Q4", "YYYY", "Q3", "Q2", "Q1", "DATE"]:
            if candidates[key] is not None:
                return get_column_letter(candidates[key])
        return None

    def create_cell_reference(self, sheet_name: str, col_letter: Optional[str],
                            row_num: Optional[int]) -> Optional[str]:
        """
        Crea una referencia de celda Excel.

        Args:
            sheet_name: Nombre de la hoja
            col_letter: Letra de columna
            row_num: Número de fila

        Returns:
            Referencia de celda o None si falta información
        """
        if col_letter is None or row_num is None:
            return None
        # Determinar hoja real por nombre
        sheet = self.sh_bal if sheet_name == self.sh_bal.title else self.sh_pl if sheet_name == self.sh_pl.title else self.sh_cfs if sheet_name == self.sh_cfs.title else None
        base_ref = f"'{sheet_name}'!{col_letter}{row_num}"

        # Intentar fallback ES cuando esté activo
        def _normalize_header_label(hv: Any) -> Optional[str]:
            if not isinstance(hv, str):
                return None
            s = hv.strip().split("\n", 1)[0]
            return s
        es_ref_current: Optional[str] = None
        es_ref_prev: Optional[str] = None
        label_current: Optional[str] = None
        try:
            if self.es_fallback_active and sheet is not None and self.es_data_sheet_name:
                hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
                col_idx = column_index_from_string(col_letter)
                label_current = _normalize_header_label(sheet.cell(row=hdr, column=col_idx).value)
                # Resolver grupo/clave
                if sheet is self.sh_bal:
                    group = "BAL"; key = self._row_to_key_bal.get(row_num)
                elif sheet is self.sh_pl:
                    group = "PL"; key = self._row_to_key_pl.get(row_num)
                else:
                    group = "CFS"; key = self._row_to_key_cfs.get(row_num)
                if key and label_current:
                    # Buscar columna en ES_DATA (considerar alias YYYY ↔ YYYYQ4)
                    def _col_for(lbl: str) -> Optional[str]:
                        if lbl in self.es_label_to_col:
                            return self.es_label_to_col[lbl]
                        # Aliasing simple: YYYYQ4 ↔ YYYY
                        m = re.match(r"^(\d{4})Q4$", lbl)
                        if m and m.group(1) in self.es_label_to_col:
                            return self.es_label_to_col[m.group(1)]
                        m2 = re.match(r"^(\d{4})$", lbl)
                        if m2 and f"{m2.group(1)}Q4" in self.es_label_to_col:
                            return self.es_label_to_col[f"{m2.group(1)}Q4"]
                        return None
                    col_es = _col_for(label_current)
                    row_es = self.es_row_for_key.get((group, key))
                    if col_es and row_es:
                        es_ref_current = f"'{self.es_data_sheet_name}'!{col_es}{row_es}"
        except Exception:
            es_ref_current = None

        # Determinar si hay TTM (solo P&L y CFS con etiqueta trimestral override)
        try:
            quarter_re = re.compile(r"^(\d{4})Q([1-4])$")
            is_ttm = bool(self.period_override_label and quarter_re.match(self.period_override_label) and sheet_name in (self.sh_pl.title, self.sh_cfs.title))
        except Exception:
            is_ttm = False

        # Si TTM, construir ref previa acorde a la fuente (base o ES)
        if is_ttm:
            prev_label = self._previous_year_same_quarter_label(self.period_override_label)  # type: ignore[arg-type]
            if es_ref_current is not None:
                try:
                    # Usar ES también para previo si existe
                    if prev_label:
                        def _col_for_prev(lbl: str) -> Optional[str]:
                            if lbl in self.es_label_to_col:
                                return self.es_label_to_col[lbl]
                            m = re.match(r"^(\d{4})Q4$", lbl)
                            if m and m.group(1) in self.es_label_to_col:
                                return self.es_label_to_col[m.group(1)]
                            m2 = re.match(r"^(\d{4})$", lbl)
                            if m2 and f"{m2.group(1)}Q4" in self.es_label_to_col:
                                return self.es_label_to_col[f"{m2.group(1)}Q4"]
                            return None
                        col_es_prev = _col_for_prev(prev_label)
                        if col_es_prev:
                            # group/key derivados arriba
                            if sheet_name == self.sh_bal.title:
                                group = "BAL"; key = self._row_to_key_bal.get(row_num)
                            elif sheet_name == self.sh_pl.title:
                                group = "PL"; key = self._row_to_key_pl.get(row_num)
                            else:
                                group = "CFS"; key = self._row_to_key_cfs.get(row_num)
                            row_es = self.es_row_for_key.get((group, key)) if key else None
                            if row_es:
                                es_ref_prev = f"'{self.es_data_sheet_name}'!{col_es_prev}{row_es}"
                except Exception:
                    es_ref_prev = None
            else:
                # base workbook refs
                try:
                    if prev_label:
                        prev_col = self._get_col_letter_by_label(self.sh_pl if sheet_name == self.sh_pl.title else self.sh_cfs, prev_label)
                        if prev_col:
                            prev_ref = f"'{sheet_name}'!{prev_col}{row_num}"
                            # Devolver TTM
                            return f"IFERROR({base_ref}-{prev_ref},{base_ref})"
                except Exception:
                    pass

        # Si está activo el fallback a ES_DATA, decidir si usarlo
        if self.es_fallback_active and es_ref_current is not None:
            try:
                # Si tenemos valor ES y valor actual, usar ES cuando difieran
                hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
                col_idx = column_index_from_string(col_letter)
                cur_val = sheet.cell(row=row_num, column=col_idx).value if sheet is not None else None
                # Determinar clave para buscar valor ES almacenado
                if sheet is self.sh_bal:
                    group = "BAL"; key = self._row_to_key_bal.get(row_num)
                elif sheet is self.sh_pl:
                    group = "PL"; key = self._row_to_key_pl.get(row_num)
                else:
                    group = "CFS"; key = self._row_to_key_cfs.get(row_num)
                lbl = label_current
                es_val = self.es_values.get((group, key, lbl)) if key and lbl else None
                # Fallback alias de etiqueta
                if es_val is None and key and lbl:
                    m = re.match(r"^(\d{4})Q4$", lbl)
                    if m:
                        es_val = self.es_values.get((group, key, m.group(1)))
                if es_val is None and key and lbl:
                    m2 = re.match(r"^(\d{4})$", lbl)
                    if m2:
                        es_val = self.es_values.get((group, key, f"{m2.group(1)}Q4"))
                def _is_close(a: Any, b: Any, tol: float = 1e-6) -> bool:
                    try:
                        if a in (None, "") and b in (None, ""):
                            return True
                        if a in (None, "") or b in (None, ""):
                            return False
                        av = float(str(a).replace(",", "").replace(" ", ""))
                        bv = float(str(b).replace(",", "").replace(" ", ""))
                        return abs(av - bv) <= tol
                    except Exception:
                        return False
                use_es = (es_val is not None) and (not _is_close(cur_val, es_val))
            except Exception:
                use_es = True
            if use_es:
                # Registrar en log
                try:
                    reason = "differs" if es_val is not None else "missing"
                    grp = "BAL" if sheet is self.sh_bal else ("PL" if sheet is self.sh_pl else "CFS")
                    ky = self._row_to_key_bal.get(row_num) if grp == "BAL" else self._row_to_key_pl.get(row_num) if grp == "PL" else self._row_to_key_cfs.get(row_num)
                    if is_ttm and es_ref_prev is not None:
                        self.es_usage_log_entries.append(
                            f"{grp}:{ky} | {sheet_name} | label={label_current} | TTM | base={base_ref} val={cur_val} vs es={es_val} → using {es_ref_current}-{es_ref_prev}"
                        )
                    else:
                        self.es_usage_log_entries.append(
                            f"{grp}:{ky} | {sheet_name} | label={label_current} | base={base_ref} val={cur_val} vs es={es_val} → using {es_ref_current}"
                        )
                except Exception:
                    pass
                if is_ttm and es_ref_prev is not None:
                    return f"IFERROR({es_ref_current}-{es_ref_prev},{es_ref_current})"
                return es_ref_current

        # Default: referencia base
        return base_ref

    def _get_col_letter_by_label(self, sheet, label: str) -> Optional[str]:
        hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
        label_norm = label.strip().split("\n", 1)[0]
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=hdr, column=col).value
            if isinstance(val, str) and val.strip().split("\n", 1)[0] == label_norm:
                return get_column_letter(col)
        return None

    def _previous_quarter_label(self, label: str) -> Optional[str]:
        m = re.match(r"^(\d{4})Q([1-4])$", label)
        if not m:
            return None
        y = int(m.group(1))
        q = int(m.group(2))
        if q == 1:
            return f"{y-1}Q4"
        return f"{y}Q{q-1}"

    def _previous_year_same_quarter_label(self, label: str) -> Optional[str]:
        m = re.match(r"^(\d{4})Q([1-4])$", label)
        if not m:
            return None
        y = int(m.group(1))
        q = int(m.group(2))
        return f"{y-1}Q{q}"

    def _year_start_quarter_label(self, label: str) -> Optional[str]:
        """Para un label trimestral YYYYQn, retorna el label del cierre del año previo (YYYY-1Q4)."""
        m = re.match(r"^(\d{4})Q([1-4])$", label)
        if not m:
            return None
        y = int(m.group(1))
        return f"{y-1}Q4"

    def _days_factor_for_label(self, label: str) -> str:
        """Días acumulados del período, como literal para incrustar en una fórmula.

        Delega en `_get_period_days` (NatureBasedMixin) para no mantener dos tablas
        de días que puedan divergir: ambos mixins conviven en FormulaBuilder.
        """
        return str(self._get_period_days(label))

    def _unwrap_iferror(self, excel_expr: str) -> str:
        """Quita el envoltorio IFERROR(X,"") si existe, tolerando espacios."""
        s = excel_expr.strip()
        try:
            import re as _re
            m = _re.match(r"^IFERROR\((.*?),(\s*)\"\"\)(\s*)$", s)
            if m:
                return m.group(1)
        except Exception:
            pass
        return s

    def create_cell_reference_by_label(self, sheet, row_num: Optional[int], label: str) -> Optional[str]:
        if row_num is None:
            return None
        col = self._get_col_letter_by_label(sheet, label)
        if col is None:
            return None
        # Preparar ref base y posible ref ES
        base_ref = f"'{sheet.title}'!{col}{row_num}"
        es_ref = None
        if self.es_fallback_active and self.es_data_sheet_name:
            try:
                if sheet is self.sh_bal:
                    group = "BAL"; key = self._row_to_key_bal.get(row_num)
                elif sheet is self.sh_pl:
                    group = "PL"; key = self._row_to_key_pl.get(row_num)
                else:
                    group = "CFS"; key = self._row_to_key_cfs.get(row_num)
                if key:
                    col_es = self.es_label_to_col.get(label)
                    if not col_es:
                        m = re.match(r"^(\d{4})Q4$", label)
                        if m and m.group(1) in self.es_label_to_col:
                            col_es = self.es_label_to_col[m.group(1)]
                        m2 = re.match(r"^(\d{4})$", label)
                        if not col_es and m2 and f"{m2.group(1)}Q4" in self.es_label_to_col:
                            col_es = self.es_label_to_col[f"{m2.group(1)}Q4"]
                    row_es = self.es_row_for_key.get((group, key))
                    if col_es and row_es:
                        es_ref = f"'{self.es_data_sheet_name}'!{col_es}{row_es}"
            except Exception:
                es_ref = None

        # Decidir usar ES si difiere
        if es_ref is not None:
            try:
                hdr = self.HDR_BAL if sheet is self.sh_bal else self.HDR_PL if sheet is self.sh_pl else self.HDR_CFS
                col_idx = column_index_from_string(col)
                cur_val = sheet.cell(row=row_num, column=col_idx).value
                if sheet is self.sh_bal:
                    group = "BAL"; key = self._row_to_key_bal.get(row_num)
                elif sheet is self.sh_pl:
                    group = "PL"; key = self._row_to_key_pl.get(row_num)
                else:
                    group = "CFS"; key = self._row_to_key_cfs.get(row_num)
                es_val = self.es_values.get((group, key, label)) if key else None
                if es_val is None and key:
                    m = re.match(r"^(\d{4})Q4$", label)
                    if m:
                        es_val = self.es_values.get((group, key, m.group(1)))
                if es_val is None and key:
                    m2 = re.match(r"^(\d{4})$", label)
                    if m2:
                        es_val = self.es_values.get((group, key, f"{m2.group(1)}Q4"))
                def _is_close(a: Any, b: Any, tol: float = 1e-6) -> bool:
                    try:
                        if a in (None, "") and b in (None, ""):
                            return True
                        if a in (None, "") or b in (None, ""):
                            return False
                        av = float(str(a).replace(",", "").replace(" ", ""))
                        bv = float(str(b).replace(",", "").replace(" ", ""))
                        return abs(av - bv) <= tol
                    except Exception:
                        return False
                if es_val is not None and not _is_close(cur_val, es_val):
                    # Registrar en log
                    try:
                        grp = "BAL" if sheet is self.sh_bal else ("PL" if sheet is self.sh_pl else "CFS")
                        ky = self._row_to_key_bal.get(row_num) if grp == "BAL" else self._row_to_key_pl.get(row_num) if grp == "PL" else self._row_to_key_cfs.get(row_num)
                        base_ref = f"'{sheet.title}'!{col}{row_num}"
                        self.es_usage_log_entries.append(
                            f"{grp}:{ky} | {sheet.title} | label={label} | base={base_ref} val={cur_val} vs es={es_val} → using {es_ref}"
                        )
                    except Exception:
                        pass
                    return es_ref
            except Exception:
                return es_ref
        return base_ref

    def create_average_reference(self, sheet, row: Optional[int], year: int) -> Optional[str]:
        """
        Crea referencia promedio entre año actual y anterior para items del balance.

        Args:
            sheet: Hoja de Excel
            row: Número de fila
            year: Año actual

        Returns:
            Fórmula de promedio o referencia simple
        """
        if row is None:
            return None

        col_now = self.find_year_column(sheet, year)
        col_prev = self.find_year_column(sheet, year - 1)

        now_ref = self.create_cell_reference(sheet.title, col_now, row) if col_now else None
        prev_ref = self.create_cell_reference(sheet.title, col_prev, row) if col_prev else None

        if now_ref and prev_ref:
            return f"AVERAGE({now_ref},{prev_ref})"
        elif now_ref:
            return now_ref
        elif prev_ref:
            return prev_ref

        return None
