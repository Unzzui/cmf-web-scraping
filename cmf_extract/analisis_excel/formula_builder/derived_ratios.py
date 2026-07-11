#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ratios derivados como FÓRMULAS Excel: CRECIMIENTO, DUPONT y CALIDAD Y SCORES.

Estas secciones antes se escribían como VALORES estáticos (calculados en Python
por ``RatioCalculator``). Aquí se generan como fórmulas Excel que referencian
directamente las celdas de los estados financieros, para transparencia y para
que se recalculen si el usuario edita los datos.

Semántica (replicando ``RatioCalculator`` para consistencia):
  - "Año" = la columna del trimestre MÁS ALTO presente del año (Q4 = YTD/año
    completo). Growth, DuPont y Piotroski son anuales → solo se llenan en esas
    columnas. Accruals y Deuda Neta/EBITDA son por trimestre (todas las columnas).
  - Promedios de balance = promedio(año, año-1), igual que ROE/ROA.
  - Piotroski F-Score = suma de 9 señales binarias entre dos años consecutivos;
    la señal de no-dilución referencia la fila "Total número de acciones
    emitidas" (en la propia hoja de ratios).

El mixin espera que el ``FormulaBuilder`` ya tenga inicializados:
``sh_bal/sh_pl/sh_cfs``, ``rows_bal/rows_pl/rows_cfs`` y
``create_cell_reference_by_label``.
"""
from __future__ import annotations

import re
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

# Estructura de una sección devuelta al procesador:
#   (nombre_seccion, [(nombre_ratio, tipo, {label_periodo: cuerpo_formula}), ...])
Section = Tuple[str, List[Tuple[str, str, Dict[str, str]]]]

_SHARES_LABELS = ("Total número de acciones emitidas", "Total number of shares issued")


class DerivedRatiosMixin:
    """Genera CRECIMIENTO / DUPONT / CALIDAD Y SCORES como fórmulas Excel."""

    # ── helpers de referencia de celda ────────────────────────────────────
    def _dr_bref(self, key: str, lb: str) -> Optional[str]:
        r = self.rows_bal.get(key)
        return self.create_cell_reference_by_label(self.sh_bal, r, lb) if r else None

    def _dr_pref(self, key: str, lb: str) -> Optional[str]:
        r = self.rows_pl.get(key)
        return self.create_cell_reference_by_label(self.sh_pl, r, lb) if r else None

    def _dr_cref(self, key: str, lb: str) -> Optional[str]:
        r = self.rows_cfs.get(key)
        return self.create_cell_reference_by_label(self.sh_cfs, r, lb) if r else None

    def _dr_ebitda(self, lb: str) -> Optional[str]:
        ebit = self._dr_pref("EBIT", lb)
        if not ebit:
            return None
        da = self._dr_pref("DepAmort", lb)
        return f"({ebit}+IFERROR({da},0))" if da else f"({ebit})"

    # ── helpers de período ────────────────────────────────────────────────
    @staticmethod
    def _dr_period_of_header(h: str) -> Optional[Tuple[int, int]]:
        """(año, trimestre) de un header 'YYYYQn' o 'YYYY' (→ Q4)."""
        m = re.match(r"^(\d{4})Q([1-4])$", h)
        if m:
            return int(m.group(1)), int(m.group(2))
        m = re.match(r"^(\d{4})$", h)
        if m:
            return int(m.group(1)), 4
        return None

    def _dr_context(self, headers: List[str]) -> Tuple[Dict[int, str], set]:
        """Mapa año → label anual (trimestre más alto) y el set de labels anuales."""
        annual: Dict[int, str] = {}
        for h in headers:
            if not isinstance(h, str):
                continue
            pq = self._dr_period_of_header(h.strip())
            if not pq:
                continue
            y, q = pq
            prev = annual.get(y)
            prev_q = self._dr_period_of_header(prev)[1] if prev else 0
            if q > prev_q:
                annual[y] = h.strip()
        return annual, set(annual.values())

    # ── generadores por ratio (devuelven {label: cuerpo_formula}) ──────────
    def _dr_yoy(self, annual: Dict[int, str], annual_labels: set,
                ref: Callable[[str], Optional[str]]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for lb in annual_labels:
            y = self._dr_period_of_header(lb)[0]
            pl = annual.get(y - 1)
            cur, prev = ref(lb), (ref(pl) if pl else None)
            if cur and prev:
                out[lb] = f'IF({prev}=0,"",IFERROR(({cur}-{prev})/ABS({prev}),""))'
        return out

    def _dr_cagr(self, annual: Dict[int, str], annual_labels: set, n: int) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for lb in annual_labels:
            y = self._dr_period_of_header(lb)[0]
            pl = annual.get(y - n)
            cur = self._dr_pref("Ventas", lb)
            base = self._dr_pref("Ventas", pl) if pl else None
            if cur and base:
                out[lb] = f'IF(OR({base}<=0,{cur}<=0),"",IFERROR(({cur}/{base})^(1/{n})-1,""))'
        return out

    def _dr_annual_ratio(self, annual_labels: set,
                         num: Callable[[str], Optional[str]],
                         den: Callable[[str], Optional[str]]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for lb in annual_labels:
            n, d = num(lb), den(lb)
            if n and d:
                out[lb] = f'IF({d}=0,"",IFERROR({n}/{d},""))'
        return out

    def _dr_avg_bal(self, annual: Dict[int, str], key: str, lb: str) -> Optional[str]:
        cur = self._dr_bref(key, lb)
        y = self._dr_period_of_header(lb)[0]
        pl = annual.get(y - 1)
        prev = self._dr_bref(key, pl) if pl else None
        if cur and prev:
            return f"(({cur}+{prev})/2)"
        return cur or prev

    def _dr_accruals(self, headers: List[str]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for h in headers:
            if not isinstance(h, str):
                continue
            lb = h.strip()
            neta, cfo, at = self._dr_pref("Neta", lb), self._dr_cref("CFO", lb), self._dr_bref("AT", lb)
            if neta and cfo and at:
                out[lb] = f'IF({at}=0,"",IFERROR(({neta}-{cfo})/{at},""))'
        return out

    def _dr_net_debt_ebitda(self, headers: List[str]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for h in headers:
            if not isinstance(h, str):
                continue
            lb = h.strip()
            ebitda = self._dr_ebitda(lb)
            dfc, dfnc, efec = self._dr_bref("DeudaFinCorr", lb), self._dr_bref("DeudaFinNC", lb), self._dr_bref("Efec", lb)
            if not ebitda or not (dfc or dfnc):
                continue
            debt = "+".join(f"IFERROR({x},0)" for x in (dfc, dfnc) if x)
            debt = f"({debt}-IFERROR({efec},0))" if efec else f"({debt})"
            out[lb] = f'IFERROR({debt}/{ebitda},"")'
        return out

    def _dr_piotroski(self, annual: Dict[int, str], annual_labels: set,
                      sh_ref: Callable[[str], Optional[str]]) -> Dict[str, str]:
        """F-Score 0-9: suma de 9 señales binarias entre año y año-1."""
        out: Dict[str, str] = {}
        for lb in annual_labels:
            y = self._dr_period_of_header(lb)[0]
            pl = annual.get(y - 1)
            if not pl:
                continue
            at_t, at_p = self._dr_bref("AT", lb), self._dr_bref("AT", pl)
            ne_t, ne_p = self._dr_pref("Neta", lb), self._dr_pref("Neta", pl)
            if not (at_t and at_p and ne_t and ne_p):
                continue
            cfo_t = self._dr_cref("CFO", lb)
            lt_t, lt_p = self._dr_bref("DeudaFinNC", lb), self._dr_bref("DeudaFinNC", pl)
            ac_t, ac_p = self._dr_bref("AC", lb), self._dr_bref("AC", pl)
            pc_t, pc_p = self._dr_bref("PC", lb), self._dr_bref("PC", pl)
            br_t, br_p = self._dr_pref("Bruta", lb), self._dr_pref("Bruta", pl)
            ve_t, ve_p = self._dr_pref("Ventas", lb), self._dr_pref("Ventas", pl)
            sh_t, sh_p = sh_ref(lb), sh_ref(pl)
            t = [
                f"IF({ne_t}/{at_t}>0,1,0)",                                                   # 1 ROA>0
                f"IF(AND(ISNUMBER({cfo_t}),{cfo_t}>0),1,0)" if cfo_t else "0",                # 2 CFO>0
                f"IF({ne_t}/{at_t}>{ne_p}/{at_p},1,0)",                                       # 3 ΔROA>0
                f"IF(AND(ISNUMBER({cfo_t}),{cfo_t}>{ne_t}),1,0)" if cfo_t else "0",           # 4 CFO>UN
                (f"IF(AND(ISNUMBER({lt_t}),ISNUMBER({lt_p}),{lt_t}/{at_t}<{lt_p}/{at_p}),1,0)"
                 if (lt_t and lt_p) else "0"),                                                # 5 ↓apalancamiento
                (f"IF(AND(ISNUMBER({ac_t}),ISNUMBER({ac_p}),ISNUMBER({pc_t}),ISNUMBER({pc_p}),"
                 f"{pc_t}<>0,{pc_p}<>0,{ac_t}/{pc_t}>{ac_p}/{pc_p}),1,0)"
                 if (ac_t and ac_p and pc_t and pc_p) else "0"),                              # 6 ↑liquidez
                (f"IF(AND(ISNUMBER({sh_t}),ISNUMBER({sh_p}),{sh_t}<={sh_p}),1,0)"
                 if (sh_t and sh_p) else "0"),                                                # 7 sin dilución
                (f"IF(AND(ISNUMBER({br_t}),ISNUMBER({br_p}),ISNUMBER({ve_t}),ISNUMBER({ve_p}),"
                 f"{ve_t}<>0,{ve_p}<>0,{br_t}/{ve_t}>{br_p}/{ve_p}),1,0)"
                 if (br_t and br_p and ve_t and ve_p) else "0"),                              # 8 ↑margen bruto
                (f"IF(AND(ISNUMBER({ve_t}),ISNUMBER({ve_p}),{ve_t}/{at_t}>{ve_p}/{at_p}),1,0)"
                 if (ve_t and ve_p) else "0"),                                                # 9 ↑rotación
            ]
            gate = (f"OR(NOT(ISNUMBER({at_t})),NOT(ISNUMBER({at_p})),"
                    f"NOT(ISNUMBER({ne_t})),NOT(ISNUMBER({ne_p})),{at_t}=0,{at_p}=0)")
            out[lb] = f'IF({gate},"",IFERROR({"+".join(t)},""))'
        return out

    # ── API pública ───────────────────────────────────────────────────────
    def build_derived_ratio_sections(self, headers: List[str], ws) -> List[Section]:
        """Devuelve las secciones CRECIMIENTO, DUPONT y CALIDAD Y SCORES como
        fórmulas. ``ws`` es la hoja de ratios (para ubicar la fila de acciones)."""
        annual, annual_labels = self._dr_context(headers)

        # Referencia a la fila de acciones ya escrita en la hoja de ratios.
        lb_to_col = {h.strip(): get_column_letter(i + 2)
                     for i, h in enumerate(headers) if isinstance(h, str)}
        shares_row = None
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(rr, 1).value
            if isinstance(v, str) and v.strip() in _SHARES_LABELS:
                shares_row = rr
                break

        def sh_ref(lb: str) -> Optional[str]:
            cl = lb_to_col.get(lb)
            return f"'{ws.title}'!{cl}{shares_row}" if (shares_row and cl) else None

        return [
            ("CRECIMIENTO", [
                ("Variación Ingresos (YoY)", "pct",
                 self._dr_yoy(annual, annual_labels, lambda lb: self._dr_pref("Ventas", lb))),
                ("Variación EBITDA (YoY)", "pct",
                 self._dr_yoy(annual, annual_labels, self._dr_ebitda)),
                ("Variación Utilidad Neta (YoY)", "pct",
                 self._dr_yoy(annual, annual_labels, lambda lb: self._dr_pref("Neta", lb))),
                ("CAGR Ingresos 3 Años", "pct", self._dr_cagr(annual, annual_labels, 3)),
                ("CAGR Ingresos 5 Años", "pct", self._dr_cagr(annual, annual_labels, 5)),
            ]),
            ("DUPONT", [
                ("Margen Neto (DuPont)", "pct",
                 self._dr_annual_ratio(annual_labels, lambda lb: self._dr_pref("Neta", lb),
                                       lambda lb: self._dr_pref("Ventas", lb))),
                ("Rotación de Activos (DuPont)", "ratio",
                 self._dr_annual_ratio(annual_labels, lambda lb: self._dr_pref("Ventas", lb),
                                       lambda lb: self._dr_avg_bal(annual, "AT", lb))),
                ("Multiplicador de Capital", "ratio",
                 self._dr_annual_ratio(annual_labels, lambda lb: self._dr_avg_bal(annual, "AT", lb),
                                       lambda lb: self._dr_avg_bal(annual, "Patr", lb))),
                ("ROE (DuPont)", "pct",
                 self._dr_annual_ratio(annual_labels, lambda lb: self._dr_pref("Neta", lb),
                                       lambda lb: self._dr_avg_bal(annual, "Patr", lb))),
            ]),
            ("CALIDAD Y SCORES", [
                ("Deuda Financiera Neta / EBITDA", "ratio", self._dr_net_debt_ebitda(headers)),
                ("Accruals (UN - CFO) / Activos", "pct", self._dr_accruals(headers)),
                ("Piotroski F-Score", "score", self._dr_piotroski(annual, annual_labels, sh_ref)),
            ]),
        ]
