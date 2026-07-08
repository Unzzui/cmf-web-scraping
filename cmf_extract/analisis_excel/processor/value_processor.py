"""
Value Processor Mixin
======================

Contains the ``_process_with_values`` method extracted from
``BulkProcessor``.
"""

import os
import pandas as pd
from pathlib import Path
from typing import Dict

from ..ratio_calculator import RatioCalculator


class ValueProcessorMixin:
    """Mixin that adds ``_process_with_values`` to BulkProcessor."""

    def _process_with_values(self, file_path: Path, financial_data: Dict,
                           company_name: str) -> str:
        """
        Procesa archivo con valores calculados (análisis estático).
        
        Args:
            file_path: Ruta del archivo original
            financial_data: Datos financieros extraídos
            company_name: Nombre de la empresa
            
        Returns:
            Nombre del archivo de salida
        """
        # Calcular ratios
        calculator = RatioCalculator(financial_data)
        all_ratios = calculator.calculate_all_ratios()
        
        # Crear nuevo workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Detectar idioma por heurística similar
        try:
            name_l = file_path.name.lower()
            is_en = name_l.endswith("_en.xlsx") or name_l.endswith("[en].xlsx") or "[en]" in name_l
        except Exception:
            is_en = False
        lang = "en" if is_en else "es"
        ws.title = "Financial Analysis (Values)" if lang == "en" else "Análisis Financiero (Valores)"
        
        years = financial_data.get("years", [])
        cols_total = 1 + len(years) + 3
        
        # Configurar estructura
        header_row = self.formatter.setup_worksheet_structure(ws, [str(y) for y in years], ws.title, lang=lang)
        
        # Escribir datos con valores
        current_row = header_row + 1
        
        section_map = {
            "LIQUIDEZ": "LIQUIDITY",
            "SOLVENCIA Y ESTRUCTURA": "SOLVENCY & CAPITAL STRUCTURE",
            "RENTABILIDAD": "PROFITABILITY",
            "EFICIENCIA OPERATIVA": "OPERATING EFFICIENCY",
            "FLUJOS Y ADICIONALES": "CASH FLOWS & OTHER",
            "CRECIMIENTO": "GROWTH",
            "DUPONT": "DUPONT",
            "CALIDAD Y SCORES": "QUALITY & SCORES",
        }

        ratio_map = {
            # Liquidez
            "Liquidez Corriente": "Current Ratio",
            "Prueba Ácida": "Quick Ratio",
            "Cash Ratio": "Cash Ratio",
            "Capital de Trabajo": "Working Capital",
            # Solvencia
            "Endeudamiento (D/E)": "Leverage (D/E)",
            "Apalancamiento (D/A)": "Leverage (D/A)",
            "Cobertura de Intereses": "Interest Coverage",
            "Deuda / EBITDA": "Debt / EBITDA",
            "Autonomía Financiera": "Equity Ratio",
            # Rentabilidad
            "Margen Bruto": "Gross Margin",
            "Margen Operativo (EBIT)": "Operating Margin (EBIT)",
            "Margen EBITDA": "EBITDA Margin",
            "Margen Neto": "Net Margin",
            # ROE / ROA y otros
            "ROE": "ROE",
            "ROA": "ROA",
            # Eficiencia
            "Rotación de Activos": "Asset Turnover",
            "Rotación de Inventarios": "Inventory Turnover",
            "Días de Inventario": "Days Inventory",
            "Rotación de Cuentas por Cobrar": "Receivables Turnover",
            "Período Promedio de Cobro": "Average Collection Period",
            "Rotación de Cuentas por Pagar": "Payables Turnover",
            "Período Promedio de Pago": "Average Payment Period",
            "Ciclo de Conversión de Efectivo": "Cash Conversion Cycle",
            # Flujos
            "Conversión de caja (CFO/Utilidad Neta)": "Cash Conversion (CFO/Net Income)",
            "Free Cash Flow (CFO - CAPEX)": "Free Cash Flow (CFO - CAPEX)",
            "AC / AT": "CA / TA",
            "PC / PT": "CL / TL",
            # Solvencia adicional
            "Deuda Financiera Neta / EBITDA": "Net Financial Debt / EBITDA",
            # Crecimiento
            "Variación Ingresos (YoY)": "Revenue Growth (YoY)",
            "Variación EBITDA (YoY)": "EBITDA Growth (YoY)",
            "Variación Utilidad Neta (YoY)": "Net Income Growth (YoY)",
            "CAGR Ingresos 3 Años": "Revenue CAGR 3Y",
            "CAGR Ingresos 5 Años": "Revenue CAGR 5Y",
            # DuPont
            "Margen Neto (DuPont)": "Net Margin (DuPont)",
            "Rotación de Activos (DuPont)": "Asset Turnover (DuPont)",
            "Multiplicador de Capital": "Equity Multiplier",
            "ROE (DuPont)": "ROE (DuPont)",
            # Calidad y scores
            "Accruals (UN - CFO) / Activos": "Accruals (NI - CFO) / Assets",
            "ROIC": "ROIC",
            "Altman Z''-Score (EM)": "Altman Z''-Score (EM)",
            "Piotroski F-Score": "Piotroski F-Score",
        }
        
        for section_name, ratios in all_ratios.items():
            sec_name = section_map.get(section_name, section_name) if lang == "en" else section_name
            # Encabezado de sección
            self.formatter.format_section_header(ws, current_row, cols_total, sec_name)
            current_row += 1
            
            # Ratios de la sección
            for ratio_name, ratio_series in ratios.items():
                r_name = ratio_map.get(ratio_name, ratio_name) if lang == "en" else ratio_name
                ws.cell(row=current_row, column=1, value=r_name)
                
                # Escribir valores por año
                # Soportar columnas como 'YYYY', 'YYYYQn' o 'YYYY-12'
                for j, year in enumerate(years, start=2):
                    # Prioridad: YYYY (anual), luego YYYYQ4, luego YYYY-12
                    candidates = [
                        str(year),
                        f"{year}Q4",
                        f"{year}-12",
                    ]
                    value = None
                    for cand in candidates:
                        if cand in ratio_series.index and pd.notna(ratio_series[cand]):
                            value = ratio_series[cand]
                            break
                    if value is not None:
                            ws.cell(row=current_row, column=j).value = value
                
                # Determinar tipo de ratio para formateo
                ratio_type = self._determine_ratio_type(r_name)
                self.formatter.format_ratio_row(ws, current_row, r_name, years, ratio_type)
                current_row += 1
        
        # Aplicar formateo condicional
        data_start_row = header_row + 1
        data_end_row = current_row - 1
        self.formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
        
        # Leyenda compacta al pie (solo trimestral)
        has_quarters_flag = any(isinstance(c, str) and ("Q" in c) for c in years)
        if has_quarters_flag:
            try:
                from openpyxl.styles import Font, Alignment
                ncols = 1 + len(years) + 3
                legend_es = [
                    "TTM: YTD(Qn) − YTD(Qn−4)",
                    "Promedios de balance: (Qn, Qn−4)",
                    "Días por trimestre: Q1=90, Q2=181, Q3=273, Q4=365",
                    "Compras TTM ≈ COGS TTM + ΔInventarios (Qn vs Qn−4)",
                ]
                legend_en = [
                    "TTM: YTD(Qn) − YTD(Qn−4)",
                    "Balance averages: (Qn, Qn−4)",
                    "Days per quarter: Q1=90, Q2=181, Q3=273, Q4=365",
                    "Purchases TTM ≈ COGS TTM + ΔInventory (Qn vs Qn−4)",
                ]
                lines = legend_en if lang == 'en' else legend_es
                row0 = data_end_row + 2
                for idx, line in enumerate(lines):
                    ws.merge_cells(start_row=row0 + idx, start_column=1, end_row=row0 + idx, end_column=ncols)
                    c = ws.cell(row=row0 + idx, column=1, value=f"• {line}")
                    c.font = Font(size=9, color="6B7280")
                    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            except Exception:
                pass

        # Congelar paneles
        self.formatter.add_freeze_panes(ws)
        # (Eliminado) Hoja README
        
        # Guardar archivo
        output_filename = (
            f"{company_name}_en - Financial Analysis (Values).xlsx" if lang == "en" else f"{company_name}_es - Análisis Financiero (Valores).xlsx"
        )
        output_path = self.output_dir / output_filename
        wb.save(str(output_path))
        
        return output_filename

