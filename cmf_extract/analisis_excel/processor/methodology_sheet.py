"""
Methodology Sheet Mixin
========================

Contains the ``_create_methodology_sheet`` method extracted from
``BulkProcessor``.
"""


class MethodologySheetMixin:
    """Mixin that adds ``_create_methodology_sheet`` to BulkProcessor."""

    def _create_methodology_sheet(self, wb, lang: str = "es") -> None:
        """
        Creates a Metodología / Methodology sheet in the workbook describing
        each financial ratio, its formula, interpretation, and data source.

        Args:
            wb: openpyxl Workbook instance (modified in-place)
            lang: "es" for Spanish, "en" for English
        """
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side
        )

        sheet_name = "METHODOLOGY" if lang == "en" else "METODOLOGÍA"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, index=len(wb.sheetnames))

        # ------------------------------------------------------------------
        # Palette
        # ------------------------------------------------------------------
        NAVY = "1F3864"
        WHITE = "FFFFFF"
        HEADER_BG = "2E4B8A"   # table header row
        ALT_ROW = "EEF2F8"     # alternating row tint
        CAT_BG = "D6E0F0"      # category sub-header
        BORDER_COLOR = "B0BEC5"

        thin = Side(style="thin", color=BORDER_COLOR)
        full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def _apply_border(cell) -> None:
            cell.border = full_border

        # ------------------------------------------------------------------
        # Column widths  (A=Ratio, B=Category, C=Formula, D=Interpretation, E=Source)
        # ------------------------------------------------------------------
        col_widths = {"A": 38, "B": 22, "C": 38, "D": 52, "E": 26}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ------------------------------------------------------------------
        # Row 1 — Title
        # ------------------------------------------------------------------
        title_text = "METHODOLOGY" if lang == "en" else "METODOLOGÍA"
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=14, color=WHITE)
        title_cell.fill = _fill(NAVY)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ------------------------------------------------------------------
        # Row 2 — Introduction
        # ------------------------------------------------------------------
        intro = (
            "This sheet describes each financial ratio, its formula, interpretation, and data source."
            if lang == "en" else
            "Esta hoja describe cada indicador financiero, su fórmula, interpretación y fuente de datos."
        )
        ws.merge_cells("A2:E2")
        intro_cell = ws["A2"]
        intro_cell.value = intro
        intro_cell.font = Font(size=10, italic=True, color="374151")
        intro_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 22

        # ------------------------------------------------------------------
        # Row 3 — blank spacer
        # ------------------------------------------------------------------
        ws.row_dimensions[3].height = 6

        # ------------------------------------------------------------------
        # Row 4 — Table header
        # ------------------------------------------------------------------
        if lang == "en":
            headers = ["Ratio / Indicator", "Category", "Formula", "Interpretation", "Data Source"]
        else:
            headers = ["Ratio / Indicador", "Categoría", "Fórmula", "Interpretación", "Fuente de Datos"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10, color=WHITE)
            cell.fill = _fill(HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            _apply_border(cell)
        ws.row_dimensions[4].height = 20

        # ------------------------------------------------------------------
        # Ratio data  (name_es, name_en, category_es, category_en,
        #              formula, interpretation_es, interpretation_en, source_es, source_en)
        # ------------------------------------------------------------------
        CAT_LIQ_ES = "Liquidez"
        CAT_LIQ_EN = "Liquidity"
        CAT_SOL_ES = "Solvencia"
        CAT_SOL_EN = "Solvency"
        CAT_PRF_ES = "Rentabilidad"
        CAT_PRF_EN = "Profitability"
        CAT_EFI_ES = "Eficiencia"
        CAT_EFI_EN = "Efficiency"
        CAT_FLU_ES = "Flujos de Caja"
        CAT_FLU_EN = "Cash Flows"

        BS_ES = "Balance General"
        BS_EN = "Balance Sheet"
        IS_ES = "Estado de Resultados"
        IS_EN = "Income Statement"
        CF_ES = "Flujo de Efectivo"
        CF_EN = "Cash Flow"
        BS_IS_ES = "Balance + Resultados"
        BS_IS_EN = "Balance + Income"
        CF_IS_ES = "Flujo + Resultados"
        CF_IS_EN = "Cash Flow + Income"

        # Each tuple: (name_es, name_en, cat_es, cat_en, formula,
        #              interp_es, interp_en, source_es, source_en)
        ratios: list[tuple[str, str, str, str, str, str, str, str, str]] = [
            # --- LIQUIDEZ / LIQUIDITY ---
            (
                "Liquidez Corriente", "Current Ratio",
                CAT_LIQ_ES, CAT_LIQ_EN,
                "AC / PC",
                ">1 indica solvencia de corto plazo",
                ">1 indicates short-term solvency",
                BS_ES, BS_EN,
            ),
            (
                "Prueba Ácida", "Quick Ratio",
                CAT_LIQ_ES, CAT_LIQ_EN,
                "(AC - Inventarios) / PC",
                ">1 prueba más conservadora de liquidez",
                ">1 more conservative liquidity test",
                BS_ES, BS_EN,
            ),
            (
                "Cash Ratio", "Cash Ratio",
                CAT_LIQ_ES, CAT_LIQ_EN,
                "Efectivo / PC",
                "Prueba más conservadora de liquidez",
                "Most conservative liquidity test",
                BS_ES, BS_EN,
            ),
            (
                "Capital de Trabajo", "Working Capital",
                CAT_LIQ_ES, CAT_LIQ_EN,
                "AC - PC",
                "Positivo = activos corrientes superan pasivos corrientes",
                "Positive = current assets exceed current liabilities",
                BS_ES, BS_EN,
            ),
            # --- SOLVENCIA / SOLVENCY ---
            (
                "Endeudamiento (D/E)", "Debt-to-Equity (D/E)",
                CAT_SOL_ES, CAT_SOL_EN,
                "PT / Patrimonio",
                "<2 generalmente saludable; >5 riesgo alto",
                "<2 generally healthy; >5 high risk",
                BS_ES, BS_EN,
            ),
            (
                "Apalancamiento (D/A)", "Leverage (D/A)",
                CAT_SOL_ES, CAT_SOL_EN,
                "PT / AT",
                "Proporción de activos financiada con deuda",
                "Portion of assets financed by debt",
                BS_ES, BS_EN,
            ),
            (
                "Cobertura de Intereses", "Interest Coverage",
                CAT_SOL_ES, CAT_SOL_EN,
                "EBIT / |Intereses|",
                ">3 cómodo; <1.5 riesgo de insolvencia",
                ">3 comfortable; <1.5 risky",
                IS_ES, IS_EN,
            ),
            (
                "Deuda / EBITDA", "Debt / EBITDA",
                CAT_SOL_ES, CAT_SOL_EN,
                "PT / EBITDA",
                "<3 saludable; >5 riesgo alto",
                "<3 healthy; >5 high risk",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Autonomía Financiera", "Equity Ratio",
                CAT_SOL_ES, CAT_SOL_EN,
                "Patrimonio / AT",
                "Mayor valor = menor apalancamiento",
                "Higher = less leverage",
                BS_ES, BS_EN,
            ),
            # --- RENTABILIDAD / PROFITABILITY ---
            (
                "Margen Bruto", "Gross Margin",
                CAT_PRF_ES, CAT_PRF_EN,
                "Ganancia Bruta / Ventas",
                "Referencia según industria",
                "Industry-specific benchmark",
                IS_ES, IS_EN,
            ),
            (
                "Margen Operativo", "Operating Margin",
                CAT_PRF_ES, CAT_PRF_EN,
                "EBIT / Ventas",
                "Eficiencia operativa de la empresa",
                "Operating efficiency",
                IS_ES, IS_EN,
            ),
            (
                "Margen EBITDA", "EBITDA Margin",
                CAT_PRF_ES, CAT_PRF_EN,
                "EBITDA / Ventas",
                "Rentabilidad proxy de caja",
                "Cash-proxy profitability",
                IS_ES, IS_EN,
            ),
            (
                "Margen Neto", "Net Margin",
                CAT_PRF_ES, CAT_PRF_EN,
                "Utilidad Neta / Ventas",
                "Rentabilidad final (bottom-line)",
                "Bottom-line profitability",
                IS_ES, IS_EN,
            ),
            (
                "ROE", "ROE",
                CAT_PRF_ES, CAT_PRF_EN,
                "Utilidad Neta / Patrimonio Promedio",
                "Retorno sobre el capital propio",
                "Return on equity capital",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "ROA", "ROA",
                CAT_PRF_ES, CAT_PRF_EN,
                "Utilidad Neta / Activos Promedio",
                "Retorno sobre activos totales",
                "Return on total assets",
                BS_IS_ES, BS_IS_EN,
            ),
            # --- EFICIENCIA / EFFICIENCY ---
            (
                "Rotación de Activos", "Asset Turnover",
                CAT_EFI_ES, CAT_EFI_EN,
                "Ventas / AT Promedio",
                "Ingresos por unidad de activo",
                "Revenue per unit of assets",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Rotación de Inventarios", "Inventory Turnover",
                CAT_EFI_ES, CAT_EFI_EN,
                "COGS / Inventario Promedio",
                "Rapidez con que se vende el inventario",
                "How fast inventory sells",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Días de Inventario", "Days Inventory Outstanding",
                CAT_EFI_ES, CAT_EFI_EN,
                "365 / Rotación Inventarios",
                "Días para vender el inventario",
                "Days to sell inventory",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Período Promedio de Cobro", "Avg Collection Period",
                CAT_EFI_ES, CAT_EFI_EN,
                "365 / (Ventas / CxC)",
                "Días para cobrar las cuentas por cobrar",
                "Days to collect receivables",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Período Promedio de Pago", "Avg Payment Period",
                CAT_EFI_ES, CAT_EFI_EN,
                "365 / (Compras / CxP)",
                "Días para pagar a proveedores",
                "Days to pay suppliers",
                BS_IS_ES, BS_IS_EN,
            ),
            (
                "Ciclo Conversión Efectivo", "Cash Conversion Cycle",
                CAT_EFI_ES, CAT_EFI_EN,
                "Días Inv + Días Cobro - Días Pago",
                "Duración del ciclo de conversión de caja",
                "Cash cycle length",
                BS_IS_ES, BS_IS_EN,
            ),
            # --- FLUJOS / CASH FLOWS ---
            (
                "Conversión de Caja", "Cash Conversion",
                CAT_FLU_ES, CAT_FLU_EN,
                "CFO / Utilidad Neta",
                ">1 = utilidades respaldadas por flujo de caja",
                ">1 = earnings backed by cash",
                CF_IS_ES, CF_IS_EN,
            ),
            (
                "Free Cash Flow", "Free Cash Flow",
                CAT_FLU_ES, CAT_FLU_EN,
                "CFO - CAPEX",
                "Caja disponible luego de reinversión",
                "Cash available after reinvestment",
                CF_ES, CF_EN,
            ),
        ]

        # ------------------------------------------------------------------
        # Write rows — group by category, inject a category header row
        # when the category changes.
        # ------------------------------------------------------------------
        current_row = 5
        current_cat = None
        category_order = [
            (CAT_LIQ_ES, CAT_LIQ_EN),
            (CAT_SOL_ES, CAT_SOL_EN),
            (CAT_PRF_ES, CAT_PRF_EN),
            (CAT_EFI_ES, CAT_EFI_EN),
            (CAT_FLU_ES, CAT_FLU_EN),
        ]
        cat_label_map = {cat_es: cat_en for cat_es, cat_en in category_order}

        data_row_index = 0  # counts only data rows (not category headers), for alternating color

        for (
            name_es, name_en, cat_es, cat_en, formula,
            interp_es, interp_en, source_es, source_en,
        ) in ratios:
            # --- Category header row when category changes ---
            if cat_es != current_cat:
                current_cat = cat_es
                cat_label = cat_en if lang == "en" else cat_es
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=5,
                )
                cat_cell = ws.cell(row=current_row, column=1, value=cat_label.upper())
                cat_cell.font = Font(bold=True, size=10, color="1F3864")
                cat_cell.fill = _fill(CAT_BG)
                cat_cell.alignment = Alignment(
                    horizontal="left", vertical="center", indent=1
                )
                cat_cell.border = full_border
                ws.row_dimensions[current_row].height = 18
                current_row += 1

            # --- Data row ---
            row_fill = _fill(ALT_ROW) if data_row_index % 2 == 0 else _fill(WHITE)
            data_row_index += 1

            ratio_name = name_en if lang == "en" else name_es
            category = cat_en if lang == "en" else cat_es
            interp = interp_en if lang == "en" else interp_es
            source = source_en if lang == "en" else source_es

            row_values = [ratio_name, category, formula, interp, source]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                cell.font = Font(size=9)
                _apply_border(cell)
            ws.row_dimensions[current_row].height = 30
            current_row += 1

        # ------------------------------------------------------------------
        # TTM note
        # ------------------------------------------------------------------
        current_row += 1  # blank spacer
        ttm_note = (
            "For quarterly data, P&L and Cash Flow ratios use Trailing Twelve Months (TTM) methodology: "
            "the four most recent quarters are summed to approximate annual figures, avoiding seasonal bias."
            if lang == "en" else
            "Para datos trimestrales, los ratios de Resultados y Flujo de Caja usan metodología TTM "
            "(Últimos Doce Meses): se suman los cuatro trimestres más recientes para aproximar cifras "
            "anuales y evitar sesgos estacionales."
        )
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=5,
        )
        ttm_cell = ws.cell(row=current_row, column=1, value=ttm_note)
        ttm_cell.font = Font(size=9, italic=True, color="374151")
        ttm_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 42

        # ------------------------------------------------------------------
        # Freeze the title + intro + spacer + header (first 4 rows)
        # ------------------------------------------------------------------
        ws.freeze_panes = "A5"

