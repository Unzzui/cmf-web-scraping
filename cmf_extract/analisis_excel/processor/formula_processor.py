"""
Formula Processor Mixin
========================

Contains the ``_process_with_formulas`` method extracted from
``BulkProcessor``.
"""

import os
import logging
import re
import time
from pathlib import Path
from typing import Dict, Optional, Callable, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment

from ..data_extractor import DataExtractor
from ..formula_builder import FormulaBuilder
from ..metadata_sheet import create_metadata_sheet


def _xbrl_search_root() -> Path:
    """Raíz donde buscar los facts XBRL (out_consolidated_*).

    Respeta CMF_XBRL_BASE_DIR (que apunta a .../data/XBRL/<frecuencia>), para no
    depender del CWD: el análisis puede correr con cwd distinto del repo donde
    viven los XBRL (p. ej. la GUF orquesta CMF_EXTRACT con cwd=cmf_extract pero
    los XBRL están en cmf-web-scraping/data/XBRL).
    """
    base = os.environ.get("CMF_XBRL_BASE_DIR")
    if base:
        p = Path(base)
        # Subir al directorio 'XBRL' para conservar el patrón **/<frecuencia>/...
        if p.name in ("Total", "Anual", "Trimestral"):
            return p.parent
        return p
    return Path("data/XBRL").resolve()

# Importar DCF functionality
try:
    import sys as _sys
    _sys.path.append(str(Path(__file__).parent.parent.parent))
    from dcf_patch import add_multi_period_dcf_functionality as add_dcf_functionality
except ImportError:
    add_dcf_functionality = None



class FormulaProcessorMixin:
    """Mixin that adds ``_process_with_formulas`` to BulkProcessor."""

    def _process_with_formulas(self, file_path: Path, financial_data: Dict, 
                             company_name: str,
                             progress_cb: Optional[Callable[[str, int, int], None]] = None) -> str:
        """
        Procesa archivo con fórmulas Excel (análisis dinámico).
        
        Args:
            file_path: Ruta del archivo original
            financial_data: Datos financieros extraídos
            company_name: Nombre de la empresa
            
        Returns:
            Nombre del archivo de salida
        """
        # Abrir workbook original
        wb = load_workbook(str(file_path))

        # Strip inherited outline/hidden column properties from source sheets
        # so all period columns are visible in the analysis output
        for _ws in wb.worksheets:
            for _dim in _ws.column_dimensions.values():
                if _dim.outlineLevel:
                    _dim.outlineLevel = 0
                if _dim.hidden:
                    _dim.hidden = False
                if _dim.collapsed:
                    _dim.collapsed = False
            if _ws.sheet_properties.outlinePr is not None:
                _ws.sheet_properties.outlinePr = None

        # Agregar DataFrames al financial_data para FormulaBuilder
        extractor = DataExtractor(str(file_path))
        extractor.load_data()
        financial_data["_df_bal"] = extractor.df_bal
        financial_data["_df_pl"] = extractor.df_pl
        financial_data["_df_cfs"] = extractor.df_cfs
        
        # Construir fórmulas
        formula_builder = FormulaBuilder(wb, financial_data)
        # Fallback a ES_DATA: desactivado por defecto. Activar con CMF_USE_ES_FALLBACK=1 si se requiere.
        try:
            import os as _os_fb
            if _os_fb.getenv('CMF_USE_ES_FALLBACK', '0') == '1':
                from ..utils.es_fallback import prepare_es_fallback
                prepare_es_fallback(wb, formula_builder, Path(str(file_path)))
        except Exception:
            pass
        # Detectar si headers incluyen trimestres (YYYYQn) y, si es así, correr una pasada por trimestre
        hdr_row = 4  # setup_worksheet_structure usa fila 4 para headers
        # Todavía no se ha creado la hoja de análisis, así que detectaremos usando el Balance
        # Extraer period labels de la hoja de balance
        def _period_labels_from_sheet(sheet):
            labels = []
            hdr = formula_builder.HDR_BAL
            for c in range(2, sheet.max_column + 1):
                v = sheet.cell(row=hdr, column=c).value
                if isinstance(v, str):
                    labels.append(v.strip().split("\n", 1)[0])
            return labels
        try:
            period_labels = _period_labels_from_sheet(formula_builder.sh_bal)
        except Exception:
            period_labels = []
        has_quarters = any(l.endswith(('Q1','Q2','Q3','Q4')) for l in period_labels)
        formula_blocks = formula_builder.build_all_formulas()
        if progress_cb:
            try:
                progress_cb('formulas_built', 2, 5)
            except Exception:
                pass
        
        # Crear hoja de análisis
        sheet_name = "RATIOS & KPIs"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, 0)
        
        years = financial_data.get("years", [])  # años reales para fórmulas
        period_vector = years[:]  # vector para formato (longitud de columnas visibles)
        # Modo combinado (Total): controla cómo se escriben fórmulas
        combined_mode = (os.getenv('CMF_ANALYSIS_COMBINED', '0') == '1')
        
        # No hay tratamientos especiales por empresa: comportamiento uniforme para todas

        # Detectar etiquetas de periodo visibles (YYYY o YYYYQn) desde Balance y normalizarlas
        def _quarter_from_month(m: int) -> str | None:
            try:
                mi = int(m)
            except Exception:
                return None
            return {3: 'Q1', 6: 'Q2', 9: 'Q3', 12: 'Q4'}.get(mi)

        def _normalize_label(raw: str) -> str | None:
            import re as _re
            s = str(raw).strip().split("\n", 1)[0]
            # YYYYQn
            if _re.match(r"^\d{4}Q[1-4]$", s):
                return s
            # YYYY-MM or YYYY-MM-DD
            m = _re.match(r"^(\d{4})-(\d{2})", s)
            if m:
                y = m.group(1)
                q = _quarter_from_month(m.group(2))
                return f"{y}{q}" if q else y
            # YYYY → normalize to YYYYQ4 for backward compat with old files
            if _re.match(r"^\d{4}$", s):
                return f"{s}Q4"
            return None

        try:
            # Detectar períodos desde encabezados del Balance para todas las empresas
            hdr_bal = formula_builder.HDR_BAL
            labels_raw = []
            for c in range(2, formula_builder.sh_bal.max_column + 1):
                v = formula_builder.sh_bal.cell(row=hdr_bal, column=c).value
                if isinstance(v, str):
                    labels_raw.append(v)
            labels_norm = [lb for lb in (_normalize_label(v) for v in labels_raw) if lb]

            # --- Filter out period labels whose source columns are entirely empty ---
            # Build raw_header → column_index mapping from Balance sheet
            _hdr_to_col: dict[str, int] = {}
            for c in range(2, formula_builder.sh_bal.max_column + 1):
                v = formula_builder.sh_bal.cell(row=hdr_bal, column=c).value
                if isinstance(v, str):
                    norm = _normalize_label(v)
                    if norm and norm not in _hdr_to_col:
                        _hdr_to_col[norm] = c

            def _col_has_data(col_idx: int, min_nonempty: int = 5) -> bool:
                """Check if a source column has meaningful data.

                Requires at least min_nonempty non-empty cells in the Balance sheet,
                since all financial ratios depend on balance sheet data.
                Scattered Cash Flow values alone are not sufficient.
                """
                count = 0
                sh = formula_builder.sh_bal
                for r in range(hdr_bal + 1, min(sh.max_row + 1, 200)):
                    v = sh.cell(row=r, column=col_idx).value
                    if v is not None and v != '' and v != 0:
                        count += 1
                        if count >= min_nonempty:
                            return True
                return False

            labels_filtered = []
            for lb in labels_norm:
                col_idx = _hdr_to_col.get(lb)
                if col_idx is None or _col_has_data(col_idx):
                    labels_filtered.append(lb)
            labels_norm = labels_filtered

            has_quarters = any('Q' in lb for lb in labels_norm)
            # De-duplicar preservando orden (bare YYYY ya normalizados a YYYYQ4 por _normalize_label)
            labels_norm = list(dict.fromkeys(labels_norm))
        except Exception:
            labels_norm = []
            has_quarters = False
        cols_total = 1 + len(years) + 3  # Indicador + años + Último + Promedio + Tendencia
        
        # Configurar estructura
        # Detectar idioma por heurística (sufijo en nombre de archivo o nombres de hojas en inglés)
        try:
            name_l = file_path.name.lower()
            # Si el nombre fuente incluye _en.xlsx o patrón [EN]
            is_en = name_l.endswith("_en.xlsx") or name_l.endswith("[en].xlsx") or "_en.xlsx" in name_l
            # Si no, mirar idioma por títulos de hojas del archivo fuente (balance, resultados, flujo)
            if not is_en:
                en_sheet_hits = any(s in wb.sheetnames for s in ["Balance Sheet", "Income Statement", "Cash Flow"])
                is_en = is_en or en_sheet_hits
            # Finalmente, si el input viene de Products con sufijo _en.xlsx (ruta completa)
            if not is_en and str(file_path).lower().endswith("_en.xlsx"):
                is_en = True
        except Exception:
            is_en = False

        lang = "en" if is_en else "es"

        # Incluir TODOS los períodos disponibles (años + trimestres) en orden cronológico
        import os as _os
        import re as _re

        # Sort all labels chronologically by (year, quarter)
        def _period_sort(lb):
            if _re.match(r"^\d{4}Q[1-4]$", lb):
                return (int(lb[:4]), int(lb[5]))
            if _re.match(r"^\d{4}$", lb):
                return (int(lb), 4)
            return (9999, 9)

        if labels_norm:
            # All labels are now YYYYQ[1-4] style (bare YYYY normalized to YYYYQ4)
            # Sort them chronologically
            display_labels = sorted(labels_norm, key=_period_sort)

            # Asegurar que FormulaBuilder tenga acceso a todos los años involucrados
            try:
                all_years_involved = list(set(int(lb[:4]) for lb in display_labels if lb[:4].isdigit()))
                formula_builder.years = sorted(all_years_involved)
            except Exception:
                pass

            # Para indexación: usar posición en display_labels
            years = list(range(len(display_labels)))
            period_vector = display_labels[:]
            # Recalcular total de columnas para encabezados/estilos
            cols_total = 1 + len(display_labels) + 3

            # Actualizar has_quarters
            has_quarters = any('Q' in lb for lb in display_labels)
        else:
            display_labels = labels_norm if has_quarters and labels_norm else [str(y) for y in years]
            if has_quarters and labels_norm:
                years = list(range(len(display_labels)))
                period_vector = years[:]
                # Si no hay years en financial_data, derivar de labels_norm
                try:
                    if not formula_builder.years:
                        import re as _re
                        derived_years = [int(s[:4]) for s in labels_norm if _re.match(r"^\d{4}(Q[1-4])?$", s)]
                        if derived_years:
                                formula_builder.years = sorted(list(set(derived_years)))
                except Exception:
                    pass

        # Detectar moneda de presentación por año desde facts (nota [110000] / ifrs:DescriptionOfPresentationCurrency)
        def _get_currency_map_from_facts() -> dict[int, str] | None:
            try:
                import pandas as _pd
                import re as _re
                from glob import glob
                # Resolver RUT y preferencia de idioma
                fname = file_path.name  # estados_<rut>_<YYYY-YYYY>_<lang>.xlsx
                m = _re.match(r"^estados_(?P<rut>[^_]+)_(?P<yr>\d{4}-\d{4})_(?P<lng>es|en)\.xlsx$", fname, flags=_re.I)
                rut_pat = m.group("rut") if m else ""
                yr_range = m.group("yr") if m else None
                lang_sfx = m.group("lng") if m else ("en" if lang=="en" else "es")
                rut_num = rut_pat.split('-')[0] if '-' in rut_pat else rut_pat
                # Ubicar facts consolidados del rut (preferir temp_consolidated_{rango})
                target_dir = None
                base_root = _xbrl_search_root()
                # 1) Dentro de data/XBRL/Total/<empresa> buscar temp_consolidated_{rango}
                comp_dirs = [p for p in base_root.glob(f"**/{rut_pat}_*") if p.is_dir()]
                if not comp_dirs:
                    comp_dirs = [p for p in base_root.glob("**/*") if p.is_dir() and rut_num in p.name]
                for comp in sorted(comp_dirs):
                    # buscar específicamente temp_consolidated_{rango}
                    if yr_range:
                        tdir = comp / f"temp_consolidated_{yr_range}"
                        if tdir.exists():
                            target_dir = tdir
                            break
                    # si no, buscar cualquier temp_consolidated_* o out_consolidated_*
                    subs = sorted(list(comp.glob("**/temp_consolidated_*")))
                    subs += sorted(list(comp.glob("**/out_consolidated_*")))
                    if subs:
                        target_dir = subs[-1]
                        break
                # 2) Fallback global: escanear temp_consolidated_* y out_consolidated_* y filtrar por ruta que contenga rut_num
                if not target_dir:
                    roots = glob(str(base_root / "**/temp_consolidated_*"), recursive=True)
                    roots += glob(str(base_root / "**/out_consolidated_*"), recursive=True)
                    for p in sorted(roots):
                        if (rut_pat and rut_pat in p) or (rut_num and rut_num in p):
                            target_dir = Path(p)
                # Si aún no hay directorio, abortar
                if not target_dir:
                    return None
                # Ubicar archivos facts (soportar sufijos variados como *_es.csv, *_es_v2.csv)
                path_es = next(iter(sorted(list(target_dir.glob("facts_*_es.csv")) + list(target_dir.glob("facts_*_es*.csv")))), None)
                path_en = next(iter(sorted(list(target_dir.glob("facts_*_en.csv")) + list(target_dir.glob("facts_*_en*.csv")))), None)
                prefer_es = (lang_sfx == 'es')
                ordered = [p for p in ([path_es, path_en] if prefer_es else [path_en, path_es]) if p]
                # Helper: normalizar header a etiqueta periodo
                def _to_lbl(col: str) -> str:
                    s = str(col).strip()
                    try:
                        d = _pd.to_datetime(s, errors='raise')
                        mth = int(d.strftime('%m'))
                        q = {3:'Q1',6:'Q2',9:'Q3',12:'Q4'}.get(mth)
                        return f"{d.year}{q}" if q else str(d.year)
                    except Exception:
                        # Handle bare YYYY -> YYYYQ4 for backward compat
                        import re as _re_lbl
                        if _re_lbl.match(r"^\d{4}$", s):
                            return f"{s}Q4"
                        return s
                # Buscar fila por Label o qname
                currency_by_year: dict[int, str] = {}
                for pth in ordered:
                    try:
                        df = _pd.read_csv(pth)
                    except Exception:
                        continue
                    # Intentar por qname/concept/name (EN estable). Acepta ifrs, ifrs-full, etc.
                    mask_q = _pd.Series([False]*len(df))
                    for colname in ["qname", "QName", "concept", "Concept", "name", "Name"]:
                        if colname in df.columns:
                            s = df[colname].astype(str)
                            mask_q = mask_q | s.str.contains(r"ifrs.*DescriptionOfPresentationCurrency", case=False, na=False, regex=True)
                    # Alternativa por Label ES/EN
                    ser_label = df.get('Label', _pd.Series(dtype=str)).astype(str)
                    mask_es = ser_label.str.contains(r"moneda\s+de\s+presentaci[oó]n", case=False, na=False, regex=True)
                    mask_en = ser_label.str.contains(r"description\s+of\s+presentation\s+currency", case=False, na=False, regex=True)
                    mask = mask_q | mask_es | mask_en
                    if not mask.any():
                        continue
                    row = df[mask].iloc[0]
                    meta_cols = {"Label","qname","QName","concept","Concept","name","Name","unitRef","decimals","precision","contextRef","id","uuid","Value","value"}
                    for col in df.columns:
                        if col in meta_cols:
                            continue
                        lb = _to_lbl(col)
                        try:
                            y = int(str(lb)[:4]) if lb and str(lb)[:4].isdigit() else None
                        except Exception:
                            y = None
                        if not y:
                            continue
                        val = row.get(col)
                        if isinstance(val, str) and val.strip():
                            # Extraer código de moneda si existe (p.ej., 'USD', 'CLP')
                            m3 = _re.search(r"\b([A-Z]{3})\b", val)
                            code = m3.group(1) if m3 else val.strip()
                            currency_by_year[y] = code
                    if currency_by_year:
                        break
                return currency_by_year or None
            except Exception:
                return None
        # Construir mapa de moneda por año (facts → fallback extractor)
        def _build_currency_spans_from_map(cmap: dict[int, str] | None) -> tuple[str | None, bool]:
            if not cmap:
                return (None, False)
            years_sorted = sorted(cmap.keys())
            spans = []
            cur = None
            start = None
            for y in years_sorted:
                c = cmap[y]
                if cur is None:
                    cur = c; start = y
                elif c != cur:
                    spans.append((start, years_sorted[years_sorted.index(y)-1], cur))
                    cur = c; start = y
            if cur is not None:
                spans.append((start, years_sorted[-1], cur))
            unique_codes = {code for (_,_,code) in spans}
            parts = [f"{a}-{b} {c}" if a != b else f"{a} {c}" for (a,b,c) in spans]
            return ("; ".join(parts), len(unique_codes) > 1)

        cmap_facts = _get_currency_map_from_facts()
        if not cmap_facts:
            try:
                cby = financial_data.get('currency_by_year') or {}
                cmap_facts = {int(y): str(cby[y]) for y in cby.keys()} if cby else None
            except Exception:
                cmap_facts = None
        spans_text, multi_currency = _build_currency_spans_from_map(cmap_facts)
        # Componer texto de unidad: mantener magnitud y solo agregar moneda si hay múltiples
        base_unit = ("Thousands CLP" if lang == 'en' else "Miles CLP")
        unit_text = base_unit if not (spans_text and multi_currency) else (base_unit + (" • Currency: " if lang=='en' else " • Moneda: ") + spans_text)
        header_row = self.formatter.setup_worksheet_structure(ws, period_vector, sheet_name, lang=lang, unit_text=unit_text)

        # Nota metodológica compacta (siempre en combinado)
        if has_quarters:
            try:
                from openpyxl.styles import Font, Alignment
                ncols = 1 + len(display_labels) + 3
                note_es = (
                    "Ratios anuales (YYYY) y trimestrales anualizados mediante factor de conversión temporal. "
                    "Q1×4, Q2×2, Q3×1.33, Q4×1 para estandarización anual de métricas YTD."
                )
                note_en = (
                    "Annual ratios (YYYY) and quarterly data annualized via temporal conversion factor. "
                    "Q1×4, Q2×2, Q3×1.33, Q4×1 for annual standardization of YTD metrics."
                )
                txt = note_en if lang == 'en' else note_es
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
                cell = ws.cell(row=3, column=1, value=txt)
                cell.font = Font(size=9, color="6B7280")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            except Exception:
                pass

        # Sobrescribir encabezados con TODAS las etiquetas de períodos (años + trimestres)
        import re as _re
        header_labels = list(dict.fromkeys(display_labels))  # mantener todos los períodos
        for i, lb in enumerate(header_labels, start=2):
            ws.cell(row=header_row, column=i, value=lb)
        
        # Lista de períodos para indexación (mantener como strings para incluir trimestres)
        try:
            year_labels: list[str] = [str(lb) for lb in header_labels]
        except Exception:
            year_labels = []
        
        # Add "Notas" (flags) column header
        flags_col = cols_total + 1
        ws.column_dimensions[get_column_letter(flags_col)].width = 28
        flags_header = ws.cell(row=header_row, column=flags_col,
                               value="Notes" if lang == "en" else "Notas")
        flags_header.font = self.formatter.bold_white_small
        flags_header.fill = self.formatter.subheader_fill
        flags_header.alignment = self.formatter.center
        flags_header.border = self.formatter.border

        # Escribir datos con fórmulas
        current_row = header_row + 1

        # Mapeo de secciones y nombres de ratios a EN cuando corresponda
        section_map = {
            "LIQUIDEZ": "LIQUIDITY",
            "SOLVENCIA Y ESTRUCTURA": "SOLVENCY & CAPITAL STRUCTURE",
            "RENTABILIDAD": "PROFITABILITY",
            "EFICIENCIA OPERATIVA": "OPERATING EFFICIENCY",
            "FLUJOS Y ADICIONALES": "CASH FLOWS & OTHER",
            "CREACIÓN DE VALOR": "VALUE CREATION",
            "COBERTURA Y RIESGO": "COVERAGE & RISK",
        }

        ratio_map = {
            # Liquidez
            "Liquidez Corriente": "Current Ratio",
            "Prueba Ácida": "Quick Ratio",
            "Cash Ratio": "Cash Ratio",
            "Capital de Trabajo": "Working Capital",
            # Solvencia
            "Endeudamiento (D/E)": "Leverage (D/E)",
            "Apalancamiento (D/A)": "Leverage (D/A)",
            "Cobertura de Intereses": "Interest Coverage",
            "Deuda / EBITDA": "Debt / EBITDA",
            "Autonomía Financiera": "Equity Ratio",
            # Rentabilidad
            "Margen Bruto": "Gross Margin",
            "Margen Operativo (EBIT)": "Operating Margin (EBIT)",
            "Margen EBITDA": "EBITDA Margin",
            "Margen Neto": "Net Margin",
            "EBITDA": "EBITDA",
            # ROE / ROA y otros
            "ROE": "ROE",
            "ROA": "ROA",
            # Eficiencia
            "Rotación de Activos": "Asset Turnover",
            "Rotación de Inventarios": "Inventory Turnover",
            "Días de Inventario": "Days Inventory",
            "Rotación de Cuentas por Cobrar": "Receivables Turnover",
            "Período Promedio de Cobro": "Average Collection Period",
            "Rotación de Cuentas por Pagar": "Payables Turnover",
            "Período Promedio de Pago": "Average Payment Period",
            "Ciclo de Conversión de Efectivo": "Cash Conversion Cycle",
            # Flujos
            "Conversión de caja (CFO/Utilidad Neta)": "Cash Conversion (CFO/Net Income)",
            "Free Cash Flow (CFO - CAPEX)": "Free Cash Flow (CFO - CAPEX)",
            "AC / AT": "CA / TA",
            "PC / PT": "CL / TL",
            # Creación de valor / Cobertura extra
            "ROIC": "ROIC",
            "EVA (WACC=10%)": "EVA (WACC=10%)",
            "Spread (ROIC - WACC)": "Spread (ROIC - WACC)",
            "Altman Z-Score": "Altman Z-Score",
            "Cobertura Servicio Deuda": "Debt Service Coverage",
            "Cobertura Gastos Fijos": "Fixed-Charge Coverage",
        }

        desc_map = {
            # Liquidez
            "Activo Corriente / Pasivo Corriente": "Current Assets / Current Liabilities",
            "(Activo Corriente - Inventarios) / Pasivo Corriente": "(Current Assets - Inventory) / Current Liabilities",
            "Efectivo y Equivalentes / Pasivo Corriente": "Cash & Equivalents / Current Liabilities",
            "Activo Corriente - Pasivo Corriente": "Current Assets - Current Liabilities",
            # Solvencia
            "Deuda Total / Patrimonio": "Total Liabilities / Equity",
            "Deuda Total / Activos Totales": "Total Liabilities / Total Assets",
            "EBIT / |Gastos por Intereses|": "EBIT / |Interest Expense|",
            "Deuda Total / (EBIT + Depreciación + Amortización)": "Total Debt / (EBIT + Depreciation + Amortization)",
            "Patrimonio / Activo Total": "Equity / Total Assets",
            "Patrimonio Total / Activo Total": "Total Equity / Total Assets",
            # Rentabilidad
            "Utilidad Bruta / Ventas": "Gross Profit / Revenue",
            "EBIT / Ventas": "EBIT / Revenue",
            "EBITDA / Ventas": "EBITDA / Revenue",
            "Utilidad Neta / Ventas": "Net Income / Revenue",
            "EBIT + Depreciación + Amortización (TTM si trimestral)": "EBIT + Depreciation + Amortization (TTM if quarterly)",
            "Utilidad Neta / Patrimonio Promedio": "Net Income / Average Equity",
            "Utilidad Neta / Activos Totales Promedio": "Net Income / Average Total Assets",
            # Eficiencia
            "Ventas / Activos Promedio": "Revenue / Average Assets",
            "Costo de Ventas / Inventario Promedio": "COGS / Average Inventory",
            "365 / Rotación de Inventarios": "365 / Inventory Turnover",
            "Días del período / Rotación de Inventarios": "Period days / Inventory Turnover",
            "Ventas / Cuentas por Cobrar Promedio": "Revenue / Average Receivables",
            "365 / Rotación de CxC": "365 / Receivables Turnover",
            "Días del período / Rotación de CxC": "Period days / Receivables Turnover",
            "Compras (≈ COGS + ΔInventario) / Cuentas por Pagar Promedio": "Purchases (≈ COGS + ΔInventory) / Average Payables",
            "365 / Rotación de CxP": "365 / Payables Turnover",
            "Días del período / Rotación de CxP": "Period days / Payables Turnover",
            "Días de Inventario + Días CxC - Días CxP": "Days Inventory + Days Receivables - Days Payables",
            # Flujos y adicionales
            "Flujo Operativo / Utilidad Neta": "Operating Cash Flow / Net Income",
            "CFO - CAPEX (Compras PPE)": "CFO - CAPEX (PPE Purchases)",
            "Activo Corriente / Activo Total": "Current Assets / Total Assets",
            "Pasivo Corriente / Pasivo Total": "Current Liabilities / Total Liabilities",
            # Creación de valor
            "NOPAT / Capital Invertido": "NOPAT / Invested Capital",
            "NOPAT - (Capital Invertido × WACC estimado)": "NOPAT - (Invested Capital × assumed WACC)",
            "ROIC - WACC estimado (10%)": "ROIC - assumed WACC (10%)",
            # Cobertura y riesgo
            "EBITDA / (Intereses + Amortización Deuda)": "EBITDA / (Interest + Debt Amortization)",
            "(EBIT + Gastos Fijos) / Gastos Fijos": "(EBIT + Fixed Expenses) / Fixed Expenses",
            "1.2×(CT/AT) + 1.4×(RE/AT) + 3.3×(EBIT/AT) + 0.6×(E/D) + 1.0×(S/AT)": "1.2×(CT/AT) + 1.4×(RE/AT) + 3.3×(EBIT/AT) + 0.6×(E/D) + 1.0×(S/AT)",
        }
        
        # Mapas para notas por sección y tags TTM/Instant
        section_note_es = {
            "LIQUIDEZ": "Base de cálculo: saldos instantáneos al cierre de Qn (promedios de balance cuando corresponda).",
            "SOLVENCIA Y ESTRUCTURA": "Base de cálculo: saldos instantáneos; 'Deuda/EBITDA' usa EBITDA TTM y Deuda Neta a cierre de Qn.",
            "RENTABILIDAD": "Base de cálculo: numeradores y ventas en TTM; promedios de balance (Qn, Qn−4).",
            "EFICIENCIA OPERATIVA": "Base de cálculo: COGS/Ventas en TTM; promedios (Qn, Qn−4); Días = DíasTrimestre(Qn)/Rotación.",
            "FLUJOS Y ADICIONALES": "Base de cálculo: CFO/FCF en TTM; ratios sobre Resultados en TTM.",
            "CREACIÓN DE VALOR": "Base de cálculo: componentes en TTM y promedios de balance según corresponda.",
            "COBERTURA Y RIESGO": "Base de cálculo: coberturas en TTM cuando aplica (EBIT, CFO, intereses).",
        }
        section_note_en = {
            "LIQUIDITY": "Computation basis: instant balances at Qn end (balance averages where applicable).",
            "SOLVENCY & CAPITAL STRUCTURE": "Computation basis: instant balances; 'Debt/EBITDA' uses TTM EBITDA and Net Debt at Qn end.",
            "PROFITABILITY": "Computation basis: numerators and revenue in TTM; balance averages (Qn, Qn−4).",
            "OPERATING EFFICIENCY": "Computation basis: COGS/Revenue in TTM; averages (Qn, Qn−4); Days = QuarterDays(Qn)/Turnover.",
            "CASH FLOWS & OTHER": "Computation basis: CFO/FCF in TTM; P&L-based ratios in TTM.",
            "VALUE CREATION": "Computation basis: TTM components and balance averages as applicable.",
            "COVERAGE & RISK": "Computation basis: coverage in TTM where applicable (EBIT, CFO, interest).",
        }

        ttm_es = " (TTM)"
        inst_es = " (Instantáneo)"
        ttm_en = " (TTM)"
        inst_en = " (Instant)"

        # Sets para tagging
        liq_instant = {"Liquidez Corriente", "Prueba Ácida", "Cash Ratio", "Capital de Trabajo"}
        solv_ttm = {"Deuda / EBITDA", "Cobertura de Intereses"}
        solv_instant = {"Endeudamiento (D/E)", "Apalancamiento (D/A)", "Autonomía Financiera"}
        rent_ttm = {"Margen Bruto", "Margen Operativo (EBIT)", "Margen EBITDA", "Margen Neto", "ROE", "ROA"}
        eff_ttm = {"Rotación de Activos", "Rotación de Inventarios", "Días de Inventario", "Rotación de Cuentas por Cobrar",
                   "Período Promedio de Cobro", "Rotación de Cuentas por Pagar", "Período Promedio de Pago", "Ciclo de Conversión de Efectivo"}
        flow_ttm = {"Conversión de caja (CFO/Utilidad Neta)", "Free Cash Flow (CFO - CAPEX)"}
        flow_instant = {"AC / AT", "PC / PT"}

        from openpyxl.comments import Comment
        from openpyxl.styles import Font, Alignment

        import time as _time
        t_sections_start = _time.perf_counter()
        for section_name, formulas in formula_blocks:
            sec_name = section_map.get(section_name, section_name) if lang == "en" else section_name
            # Encabezado de sección
            self.formatter.format_section_header(ws, current_row, cols_total, sec_name)
            current_row += 1

            # Subnota de sección (solo trimestral)
            # En modo Total, no mostrar subnotas de política trimestral
            if has_quarters and not combined_mode:
                try:
                    note = (section_note_en.get(sec_name) if lang == 'en' else section_note_es.get(sec_name)) or ""
                    if note:
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=cols_total)
                        cell = ws.cell(row=current_row, column=1, value=note)
                        cell.font = Font(size=9, color="6B7280")
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        current_row += 1
                except Exception:
                    pass
            
            # Fórmulas de la sección
            for name, ratio_type, func, description in formulas:
                # Base del nombre del ratio, traducido si corresponde
                ratio_name_base = ratio_map.get(name, name) if lang == "en" else name
                ratio_name = ratio_name_base
                # En modo Total, no agregar sufijos (TTM) o (Instantáneo) al nombre del ratio
                if has_quarters and not combined_mode:
                    if section_name == "LIQUIDEZ" and ratio_name_base in liq_instant:
                        ratio_name += (inst_en if lang == 'en' else inst_es)
                    elif section_name == "SOLVENCIA Y ESTRUCTURA" and ratio_name_base in solv_ttm:
                        ratio_name += (ttm_en if lang == 'en' else ttm_es)
                    elif section_name == "SOLVENCIA Y ESTRUCTURA" and ratio_name_base in solv_instant:
                        ratio_name += (inst_en if lang == 'en' else inst_es)
                    elif section_name == "RENTABILIDAD" and ratio_name_base in rent_ttm:
                        ratio_name += (ttm_en if lang == 'en' else ttm_es)
                    elif section_name == "EFICIENCIA OPERATIVA" and ratio_name_base in eff_ttm:
                        ratio_name += (ttm_en if lang == 'en' else ttm_es)
                    elif section_name == "FLUJOS Y ADICIONALES" and ratio_name_base in flow_ttm:
                        ratio_name += (ttm_en if lang == 'en' else ttm_es)
                    elif section_name == "FLUJOS Y ADICIONALES" and ratio_name_base in flow_instant:
                        ratio_name += (inst_en if lang == 'en' else inst_es)
                ws.cell(row=current_row, column=1, value=ratio_name)
                
                # Escribir fórmulas solo por año (YYYY) usando el vector fijo de años
                # Construir una vez el mapa base del ratio
                formula_builder.period_override_label = None
                try:
                    if not formula_builder.years:
                        formula_builder.years = year_labels[:]
                except Exception:
                    pass
                base_map = func()
                try:
                    base_map = {str(k): v for k, v in base_map.items()}
                except Exception:
                    pass
                import re as _re
                for idx, y_int in enumerate(year_labels):
                    j = 2 + idx
                    label_norm = str(y_int)
                    formula_str = None
                    
                    # 3) PRIORIDAD: Lógica trimestral personalizada para TODOS los trimestres
                    if _re.match(r"^\d{4}Q[1-4]$", label_norm):
                        try:
                            # Configurar período específico en FormulaBuilder
                            formula_builder.period_override_label = label_norm
                            year_for_quarter = int(label_norm[:4])
                            
                            # Generar fórmulas directamente para trimestres usando funciones específicas
                            # Helpers para referencias de balance y P&L por etiqueta
                            def _bal_ref(key: str) -> Optional[str]:
                                rownum = formula_builder.rows_bal.get(key)
                                return formula_builder.create_cell_reference_by_label(formula_builder.sh_bal, rownum, label_norm) if rownum else None
                            
                            def _pl_ref(key: str) -> Optional[str]:
                                rownum = formula_builder.rows_pl.get(key)
                                return formula_builder.create_cell_reference_by_label(formula_builder.sh_pl, rownum, label_norm) if rownum else None
                            
                            # Ratios de liquidez (datos instantáneos de balance)
                            if ratio_name_base in ("Liquidez Corriente", "Current Ratio"):
                                ca = _bal_ref("AC")
                                cl = _bal_ref("PC")
                                if ca and cl:
                                    formula_str = f"IFERROR({ca}/{cl},\"\")"
                                    
                            elif ratio_name_base in ("Prueba Ácida", "Quick Ratio", "Acid Test"):
                                ca = _bal_ref("AC")
                                inv = _bal_ref("Inv")
                                cl = _bal_ref("PC")
                                if ca and cl:
                                    inv_part = f"-IFERROR({inv},0)" if inv else ""
                                    formula_str = f"IFERROR(({ca}{inv_part})/{cl},\"\")"
                                    
                            elif ratio_name_base in ("Cash Ratio",):
                                cash = _bal_ref("Efec")
                                cl = _bal_ref("PC")
                                if cash and cl:
                                    formula_str = f"IFERROR({cash}/{cl},\"\")"
                                    
                            elif ratio_name_base in ("Capital de Trabajo", "Working Capital"):
                                ca = _bal_ref("AC")
                                cl = _bal_ref("PC")
                                if ca and cl:
                                    formula_str = f"IFERROR({ca}-{cl},\"\")"
                            
                            # Ratios de solvencia (balance instantáneo)
                            elif ratio_name_base in ("Endeudamiento (D/E)", "Deuda Total / Patrimonio", "Total Liabilities / Equity"):
                                tl = _bal_ref("PT")
                                equity = _bal_ref("Patr")
                                if tl and equity:
                                    formula_str = f"IFERROR({tl}/{equity},\"\")"
                                    
                            elif ratio_name_base in ("Apalancamiento (D/A)", "Deuda Total / Activos Totales", "Total Liabilities / Total Assets"):
                                tl = _bal_ref("PT")
                                ta = _bal_ref("AT")
                                if tl and ta:
                                    formula_str = f"IFERROR({tl}/{ta},\"\")"
                                    
                            elif ratio_name_base in ("Autonomía Financiera", "Patrimonio / Activo Total", "Equity / Total Assets"):
                                equity = _bal_ref("Patr")
                                ta = _bal_ref("AT")
                                if equity and ta:
                                    formula_str = f"IFERROR({equity}/{ta},\"\")"
                            
                            # Ratios de rentabilidad (usar YTD anualizado en lugar de TTM)
                            elif ratio_name_base in ("Margen Bruto", "Gross Profit / Revenue", "Utilidad Bruta / Ventas"):
                                # Para trimestres: usar datos YTD anualizados
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                gross_profit_ytd = _pl_ref("Bruta")
                                revenue_ytd = _pl_ref("Ventas")
                                if gross_profit_ytd and revenue_ytd:
                                    # Para márgenes, NO anualizar - usar YTD directamente
                                    formula_str = f"IFERROR({gross_profit_ytd}/{revenue_ytd},\"\")"
                                    
                            elif ratio_name_base in ("Margen Operativo (EBIT)", "EBIT / Revenue", "EBIT / Ventas"):
                                ebit_ytd = _pl_ref("EBIT")
                                revenue_ytd = _pl_ref("Ventas")
                                if ebit_ytd and revenue_ytd:
                                    formula_str = f"IFERROR({ebit_ytd}/{revenue_ytd},\"\")"
                                    
                            elif ratio_name_base in ("Margen Neto", "Net Income / Revenue", "Utilidad Neta / Ventas"):
                                net_income_ytd = _pl_ref("Neta")
                                revenue_ytd = _pl_ref("Ventas")
                                if net_income_ytd and revenue_ytd:
                                    formula_str = f"IFERROR({net_income_ytd}/{revenue_ytd},\"\")"
                            
                            # Ratios de eficiencia (usar datos anualizados)
                            elif ratio_name_base in ("Rotación de Activos Fijos", "PPE Turnover", "Ventas / PPE Promedio"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                revenue_ytd = _pl_ref("Ventas")
                                ppe = _bal_ref("PPE")
                                if revenue_ytd and ppe:
                                    # Anualizar ventas para comparabilidad
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({revenue_ytd}*{annualization_factor})/{ppe},\"\")"
                                    else:
                                        formula_str = f"IFERROR({revenue_ytd}/{ppe},\"\")"
                            
                            # ROE y ROA (anualizados para trimestres)
                            elif ratio_name_base in ("ROE", "Return on Equity", "Utilidad Neta / Patrimonio Promedio"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                net_income_ytd = _pl_ref("Neta")
                                equity = _bal_ref("Patr")
                                if net_income_ytd and equity:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({net_income_ytd}*{annualization_factor})/{equity},\"\")"
                                    else:
                                        formula_str = f"IFERROR({net_income_ytd}/{equity},\"\")"
                            
                            # Ratios de rotación y eficiencia (con días del trimestre)
                            elif ratio_name_base in ("Rotación de Activos", "Asset Turnover", "Ventas / Activos Promedio"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                revenue_ytd = _pl_ref("Ventas")
                                ta = _bal_ref("AT")
                                if revenue_ytd and ta:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({revenue_ytd}*{annualization_factor})/{ta},\"\")"
                                    else:
                                        formula_str = f"IFERROR({revenue_ytd}/{ta},\"\")"
                                        
                            elif ratio_name_base in ("Rotación de Inventarios", "Inventory Turnover", "Costo de Ventas / Inventario Promedio"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                cogs_ytd = _pl_ref("COGS")
                                inv = _bal_ref("Inv")
                                if cogs_ytd and inv:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({cogs_ytd}*{annualization_factor})/{inv},\"\")"
                                    else:
                                        formula_str = f"IFERROR({cogs_ytd}/{inv},\"\")"
                                        
                            elif ratio_name_base in ("Días de Inventario", "Days Inventory Outstanding"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                days_in_period = {1: 90, 2: 181, 3: 273, 4: 365}.get(quarter_num, 365)
                                
                                cogs_ytd = _pl_ref("COGS")
                                inv_current = _bal_ref("Inv")
                                if cogs_ytd and inv_current:
                                    if quarter_num != 4:
                                        # Para trimestres: usar YTD sin anualizar y días del período específico
                                        # Buscar inventario del año anterior mismo trimestre
                                        year_current = int(label_norm[:4])
                                        quarter_current = label_norm[5]
                                        label_prev_year = f"{year_current-1}Q{quarter_current}"
                                        inv_prev_year = formula_builder.create_cell_reference_by_label(formula_builder.sh_bal, formula_builder.rows_bal.get("Inv"), label_prev_year) if formula_builder.rows_bal.get("Inv") else None
                                        
                                        if inv_prev_year:
                                            inv_avg = f"AVERAGE({inv_current},{inv_prev_year})"
                                        else:
                                            inv_avg = inv_current
                                        
                                        # Días del período / (COGS YTD / Inventario promedio)
                                        formula_str = f"IFERROR({days_in_period}/({cogs_ytd}/{inv_avg}),\"\")"
                                    else:
                                        # Para año completo: usar promedio con año anterior
                                        year_current = int(label_norm)
                                        inv_prev_year = formula_builder.create_cell_reference(formula_builder.sh_bal.title, formula_builder.find_year_column(formula_builder.sh_bal, year_current-1), formula_builder.rows_bal.get("Inv")) if formula_builder.find_year_column(formula_builder.sh_bal, year_current-1) else None
                                        
                                        if inv_prev_year:
                                            inv_avg = f"AVERAGE({inv_current},{inv_prev_year})"
                                        else:
                                            inv_avg = inv_current
                                            
                                        formula_str = f"IFERROR(365/({cogs_ytd}/{inv_avg}),\"\")"
                                        
                            elif ratio_name_base in ("Rotación de Cuentas por Cobrar", "Receivables Turnover", "Ventas / Cuentas por Cobrar Promedio"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                revenue_ytd = _pl_ref("Ventas")
                                receivables = _bal_ref("CxC")
                                if revenue_ytd and receivables:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({revenue_ytd}*{annualization_factor})/{receivables},\"\")"
                                    else:
                                        formula_str = f"IFERROR({revenue_ytd}/{receivables},\"\")"
                                        
                            elif ratio_name_base in ("Período Promedio de Cobro", "Average Collection Period", "365 / Rotación de CxC"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                days_in_period = {1: 90, 2: 181, 3: 273, 4: 365}.get(quarter_num, 365)
                                
                                revenue_ytd = _pl_ref("Ventas")
                                receivables_current = _bal_ref("CxC")
                                if revenue_ytd and receivables_current:
                                    if quarter_num != 4:
                                        # Para trimestres: usar YTD sin anualizar y días del período específico
                                        year_current = int(label_norm[:4])
                                        quarter_current = label_norm[5]
                                        label_prev_year = f"{year_current-1}Q{quarter_current}"
                                        receivables_prev_year = formula_builder.create_cell_reference_by_label(formula_builder.sh_bal, formula_builder.rows_bal.get("CxC"), label_prev_year) if formula_builder.rows_bal.get("CxC") else None
                                        
                                        if receivables_prev_year:
                                            receivables_avg = f"AVERAGE({receivables_current},{receivables_prev_year})"
                                        else:
                                            receivables_avg = receivables_current
                                        
                                        # Días del período / (Ventas YTD / CxC promedio)
                                        formula_str = f"IFERROR({days_in_period}/({revenue_ytd}/{receivables_avg}),\"\")"
                                    else:
                                        # Para año completo: usar promedio con año anterior
                                        year_current = int(label_norm)
                                        receivables_prev_year = formula_builder.create_cell_reference(formula_builder.sh_bal.title, formula_builder.find_year_column(formula_builder.sh_bal, year_current-1), formula_builder.rows_bal.get("CxC")) if formula_builder.find_year_column(formula_builder.sh_bal, year_current-1) else None
                                        
                                        if receivables_prev_year:
                                            receivables_avg = f"AVERAGE({receivables_current},{receivables_prev_year})"
                                        else:
                                            receivables_avg = receivables_current
                                            
                                        formula_str = f"IFERROR(365/({revenue_ytd}/{receivables_avg}),\"\")"
                                        
                            elif ratio_name_base in ("Rotación de Cuentas por Pagar", "Payables Turnover"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                cogs_ytd = _pl_ref("COGS")
                                payables = _bal_ref("CxP")
                                if cogs_ytd and payables:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({cogs_ytd}*{annualization_factor})/{payables},\"\")"
                                    else:
                                        formula_str = f"IFERROR({cogs_ytd}/{payables},\"\")"
                                        
                            elif ratio_name_base in ("Período Promedio de Pago", "Average Payment Period", "365 / Rotación de CxP"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                days_in_period = {1: 90, 2: 181, 3: 273, 4: 365}.get(quarter_num, 365)
                                
                                cogs_ytd = _pl_ref("COGS")
                                payables_current = _bal_ref("CxP")
                                if cogs_ytd and payables_current:
                                    if quarter_num != 4:
                                        # Para trimestres: usar YTD sin anualizar y días del período específico
                                        year_current = int(label_norm[:4])
                                        quarter_current = label_norm[5]
                                        label_prev_year = f"{year_current-1}Q{quarter_current}"
                                        payables_prev_year = formula_builder.create_cell_reference_by_label(formula_builder.sh_bal, formula_builder.rows_bal.get("CxP"), label_prev_year) if formula_builder.rows_bal.get("CxP") else None
                                        
                                        if payables_prev_year:
                                            payables_avg = f"AVERAGE({payables_current},{payables_prev_year})"
                                        else:
                                            payables_avg = payables_current
                                        
                                        # Días del período / (COGS YTD / CxP promedio)
                                        formula_str = f"IFERROR({days_in_period}/({cogs_ytd}/{payables_avg}),\"\")"
                                    else:
                                        # Para año completo: usar promedio con año anterior
                                        year_current = int(label_norm)
                                        payables_prev_year = formula_builder.create_cell_reference(formula_builder.sh_bal.title, formula_builder.find_year_column(formula_builder.sh_bal, year_current-1), formula_builder.rows_bal.get("CxP")) if formula_builder.find_year_column(formula_builder.sh_bal, year_current-1) else None
                                        
                                        if payables_prev_year:
                                            payables_avg = f"AVERAGE({payables_current},{payables_prev_year})"
                                        else:
                                            payables_avg = payables_current
                                            
                                        formula_str = f"IFERROR(365/({cogs_ytd}/{payables_avg}),\"\")"
                                        
                            elif ratio_name_base in ("Ciclo de Conversión de Efectivo", "Cash Conversion Cycle"):
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                days_in_period = {1: 90, 2: 181, 3: 273, 4: 365}.get(quarter_num, 365)
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                cogs_ytd = _pl_ref("COGS")
                                revenue_ytd = _pl_ref("Ventas")
                                inv = _bal_ref("Inv")
                                receivables = _bal_ref("CxC")
                                payables = _bal_ref("CxP")
                                
                                if cogs_ytd and revenue_ytd and inv and receivables and payables:
                                    if annualization_factor != 1:
                                        # Días Inventario + Días CxC - Días CxP
                                        days_inv = f"{days_in_period}/(({cogs_ytd}*{annualization_factor})/{inv})"
                                        days_rec = f"{days_in_period}/(({revenue_ytd}*{annualization_factor})/{receivables})"
                                        days_pay = f"{days_in_period}/(({cogs_ytd}*{annualization_factor})/{payables})"
                                        formula_str = f"IFERROR(({days_inv})+({days_rec})-({days_pay}),\"\")"
                                    else:
                                        days_inv = f"365/({cogs_ytd}/{inv})"
                                        days_rec = f"365/({revenue_ytd}/{receivables})"
                                        days_pay = f"365/({cogs_ytd}/{payables})"
                                        formula_str = f"IFERROR(({days_inv})+({days_rec})-({days_pay}),\"\")"
                            
                            # Ratios adicionales
                            elif ratio_name_base in ("AC / AT", "CA / TA", "Current Assets / Total Assets"):
                                ca = _bal_ref("AC")
                                ta = _bal_ref("AT")
                                if ca and ta:
                                    formula_str = f"IFERROR({ca}/{ta},\"\")"
                                    
                            elif ratio_name_base in ("PC / PT", "CL / TL", "Current Liabilities / Total Liabilities"):
                                cl = _bal_ref("PC")
                                tl = _bal_ref("PT")
                                if cl and tl:
                                    formula_str = f"IFERROR({cl}/{tl},\"\")"
                                    
                            elif ratio_name_base in ("ROA", "Return on Assets", "Utilidad Neta / Activos Totales Promedio"):
                                # Para trimestres: usar utilidad neta YTD anualizada
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                net_income_ytd = _pl_ref("Neta")
                                ta = _bal_ref("AT")
                                if net_income_ytd and ta:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(({net_income_ytd}*{annualization_factor})/{ta},\"\")"
                                    else:
                                        formula_str = f"IFERROR({net_income_ytd}/{ta},\"\")"
                                    
                            elif ratio_name_base in ("ROIC", "Return on Invested Capital"):
                                # ROIC = EBIT(1-Tax) anualizado / (Patrimonio + Deuda)
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                ebit_ytd = _pl_ref("EBIT")
                                equity = _bal_ref("Patr")
                                debt = _bal_ref("DeudaFin")  # Deuda financiera
                                if ebit_ytd and equity:
                                    debt_part = f"+IFERROR({debt},0)" if debt else ""
                                    # Asumir tasa de impuestos del 27%
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR((({ebit_ytd}*{annualization_factor})*0.73)/({equity}{debt_part}),\"\")"
                                    else:
                                        formula_str = f"IFERROR(({ebit_ytd}*0.73)/({equity}{debt_part}),\"\")"
                                    
                            elif ratio_name_base in ("EVA (WACC=10%)", "Economic Value Added"):
                                # EVA = NOPAT anualizado - (Capital Invertido × WACC)
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                ebit_ytd = _pl_ref("EBIT")
                                equity = _bal_ref("Patr")
                                debt = _bal_ref("DeudaFin")
                                if ebit_ytd and equity:
                                    debt_part = f"+IFERROR({debt},0)" if debt else ""
                                    # NOPAT = EBIT × (1-0.27), Capital = Patrimonio + Deuda, WACC = 10%
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR((({ebit_ytd}*{annualization_factor})*0.73)-(({equity}{debt_part})*0.1),\"\")"
                                    else:
                                        formula_str = f"IFERROR(({ebit_ytd}*0.73)-(({equity}{debt_part})*0.1),\"\")"
                                    
                            elif ratio_name_base in ("Spread (ROIC - WACC)", "ROIC - WACC estimado (10%)"):
                                # Spread = ROIC anualizado - WACC
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                ebit_ytd = _pl_ref("EBIT")
                                equity = _bal_ref("Patr")
                                debt = _bal_ref("DeudaFin")
                                if ebit_ytd and equity:
                                    debt_part = f"+IFERROR({debt},0)" if debt else ""
                                    # ROIC - 10%
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR(((({ebit_ytd}*{annualization_factor})*0.73)/({equity}{debt_part}))-0.1,\"\")"
                                    else:
                                        formula_str = f"IFERROR((({ebit_ytd}*0.73)/({equity}{debt_part}))-0.1,\"\")"
                            
                            # Ratios de cobertura y riesgo
                            elif ratio_name_base in ("Cobertura Gastos Fijos", "Fixed Charge Coverage"):
                                # (EBIT + Gastos Fijos) / Gastos Fijos - usar YTD anualizado
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                ebit_ytd = _pl_ref("EBIT")
                                interest_ytd = _pl_ref("Interes")
                                if ebit_ytd and interest_ytd:
                                    if annualization_factor != 1:
                                        formula_str = f"IFERROR((({ebit_ytd}*{annualization_factor})+ABS(({interest_ytd}*{annualization_factor})))/ABS(({interest_ytd}*{annualization_factor})),\"\")"
                                    else:
                                        formula_str = f"IFERROR(({ebit_ytd}+ABS({interest_ytd}))/ABS({interest_ytd}),\"\")"
                                    
                            elif ratio_name_base in ("Altman Z-Score", "Z-Score"):
                                # Z = 1.2×(WC/TA) + 1.4×(RE/TA) + 3.3×(EBIT/TA) + 0.6×(E/D) + 1.0×(S/TA)
                                quarter_num = int(label_norm[5]) if len(label_norm) == 6 else 4
                                annualization_factor = {1: 4, 2: 2, 3: 1.333, 4: 1}.get(quarter_num, 1)
                                
                                ca = _bal_ref("AC")
                                cl = _bal_ref("PC")
                                retained_earnings = _bal_ref("ResAcum")  # Resultados Acumulados
                                ta = _bal_ref("AT")
                                equity = _bal_ref("Patr")
                                tl = _bal_ref("PT")
                                ebit_ytd = _pl_ref("EBIT")
                                sales_ytd = _pl_ref("Ventas")
                                
                                if ca and cl and ta and equity and tl and ebit_ytd and sales_ytd:
                                    wc_ta = f"({ca}-{cl})/{ta}"
                                    re_ta = f"IFERROR({retained_earnings}/{ta},0)" if retained_earnings else "0"
                                    
                                    if annualization_factor != 1:
                                        ebit_ta = f"({ebit_ytd}*{annualization_factor})/{ta}"
                                        s_ta = f"({sales_ytd}*{annualization_factor})/{ta}"
                                    else:
                                        ebit_ta = f"{ebit_ytd}/{ta}"
                                        s_ta = f"{sales_ytd}/{ta}"
                                    
                                    e_d = f"{equity}/{tl}"
                                    formula_str = f"IFERROR(1.2*({wc_ta})+1.4*({re_ta})+3.3*({ebit_ta})+0.6*({e_d})+1.0*({s_ta}),\"\")"
                            
                            # Limpiar override después del uso
                            formula_builder.period_override_label = None
                            
                        except Exception:
                            # Limpiar override en caso de error
                            formula_builder.period_override_label = None
                            pass
                    
                    # 2) Fallback: Usar fórmulas desde FormulaBuilder si no tenemos lógica personalizada
                    if not formula_str:
                        if label_norm in base_map:
                            formula_str = base_map[label_norm]
                        else:
                            for cand in [f"{label_norm}-12", f"{label_norm}-12-31", f"{label_norm}Q4"]:
                                if cand in base_map:
                                    formula_str = base_map[cand]
                                    break
                    
                    # 4) Mismo cálculo directo para columnas anuales (usando año YYYY)
                    if not formula_str:
                        try:
                            year_int = int(label_norm)
                            def _by_year(key: str) -> Optional[str]:
                                rownum = formula_builder.rows_bal.get(key)
                                col = formula_builder.find_year_column(formula_builder.sh_bal, year_int)
                                return formula_builder.create_cell_reference(formula_builder.sh_bal.title, col, rownum) if (rownum and col) else None
                            ca = _by_year("AC")
                            inv = _by_year("Inv")
                            cl = _by_year("PC")
                            cash = _by_year("Efec")
                            ta = _by_year("AT")
                            if ratio_name_base in ("Liquidez Corriente", "Current Ratio") and ca and cl:
                                formula_str = f"IFERROR({ca}/{cl},\"\")"
                            elif ratio_name_base in ("Prueba Ácida", "Quick Ratio") and ca and inv and cl:
                                formula_str = f"IFERROR(({ca}-IFERROR({inv},0))/{cl},\"\")"
                            elif ratio_name_base in ("Cash Ratio",) and cash and cl:
                                formula_str = f"IFERROR({cash}/{cl},\"\")"
                            elif ratio_name_base in ("Capital de Trabajo", "Working Capital") and ca and cl:
                                formula_str = f"IFERROR({ca}-{cl},\"\")"
                            elif ratio_name_base in ("AC / AT", "CA / TA") and ca and ta:
                                formula_str = f"IFERROR({ca}/{ta},\"\")"
                            elif ratio_name_base in ("PC / PT", "CL / TL") and cl:
                                tl_col = formula_builder.find_year_column(formula_builder.sh_bal, year_int)
                                tl = formula_builder.create_cell_reference(formula_builder.sh_bal.title, tl_col, formula_builder.rows_bal.get("PT")) if tl_col else None
                                if tl:
                                    formula_str = f"IFERROR({cl}/{tl},\"\")"
                        except Exception:
                            pass
                    # 5) Escribir fórmula si se obtuvo por mapa o cálculo directo (por cada año)
                    if formula_str:
                        ws.cell(row=current_row, column=j).value = f"={formula_str}"
                
                # Formatear fila con el nombre presentado
                self.formatter.format_ratio_row(ws, current_row, ratio_name, years, ratio_type)

                # --- Flags column: IF formulas for unusual values ---
                try:
                    # Use the "Latest" column cell as reference for flag
                    latest_col = 1 + len(year_labels) + 1
                    latest_ref = f"{get_column_letter(latest_col)}{current_row}"
                    flag_formula = None
                    # Map ratio base names to flag conditions
                    if name in ("Margen Bruto", "Gross Margin"):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}>0.9,'
                            f'"{"Unusually high margin" if lang=="en" else "Margen inusualmente alto"}",'
                            f'""))'
                        )
                    elif name in ("Margen Neto", "Net Margin"):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}<-0.5,'
                            f'"{"Significant loss" if lang=="en" else "Pérdida significativa"}",'
                            f'""))'
                        )
                    elif name in ("Endeudamiento (D/E)", "Leverage (D/E)"):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}>5,'
                            f'"{"High leverage" if lang=="en" else "Alto endeudamiento"}",'
                            f'""))'
                        )
                    elif name in ("Liquidez Corriente", "Current Ratio"):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}<0.5,'
                            f'"{"Liquidity risk" if lang=="en" else "Riesgo de liquidez"}",'
                            f'""))'
                        )
                    elif name in ("Capital de Trabajo", "Working Capital"):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}<0,'
                            f'"{"Negative working capital" if lang=="en" else "Capital trabajo negativo"}",'
                            f'""))'
                        )
                    elif name in ("ROE",):
                        flag_formula = (
                            f'=IF({latest_ref}="N/A","",IF({latest_ref}<0,'
                            f'"{"Negative profitability" if lang=="en" else "Rentabilidad negativa"}",'
                            f'""))'
                        )
                    if flag_formula:
                        fc = ws.cell(row=current_row, column=flags_col, value=flag_formula)
                        fc.font = Font(color="B45309", size=9, italic=True)
                        fc.alignment = Alignment(horizontal="left", vertical="center")
                        fc.border = self.formatter.border
                except Exception:
                    pass

                # Comentario hover con definición; en modo Total omitir política trimestral
                try:
                    # Texto de definición: en EN traducimos; en ES mantenemos el original
                    desc = desc_map.get(description, description) if lang == 'en' else description
                    if has_quarters and not combined_mode:
                        policy_es = "Política trimestral: TTM = YTD(Qn) − YTD(Qn−4). Promedios de balance: (Qn, Qn−4)."
                        policy_en = "Quarterly policy: TTM = YTD(Qn) − YTD(Qn−4). Balance averages: (Qn, Qn−4)."
                        pol = policy_en if lang == 'en' else policy_es
                        txt = f"{desc}\n{pol}"
                    else:
                        txt = desc
                    ws.cell(row=current_row, column=1).comment = Comment(txt, "Findata Chile")
                except Exception:
                    pass
                current_row += 1
        
        # Secciones escritas
        # Sección de utilidades al final: primero D&A (valores), luego EBITDA referenciando esa fila
        utils_title = "UTILIDADES" if lang == "es" else "UTILS"
        self.formatter.format_section_header(ws, current_row, cols_total, utils_title)
        current_row += 1
        from ..formula_builder import FormulaBuilder as _FB
        fb = _FB(wb, financial_data)
        headers_in_sheet = [ws.cell(row=header_row, column=c).value for c in range(2, 2 + len(year_labels))]

        # 1) D&A (Depreciación y Amortización)
        daname = "Depreciación y Amortización" if lang == "es" else "Depreciation and Amortization"
        ws.cell(row=current_row, column=1, value=daname)
        # Intentar localizar facts consolidados por RUT en data/XBRL/**
        da_values_map: dict[str, float] = {}
        # Inicializar fuera del try: si la extracción falla antes de poblarlo
        # (p. ej. empresas como SQM con facts sin la sección de acciones),
        # debe quedar ligado para los usos posteriores (evita UnboundLocalError).
        shares_values_map: dict[str, float] = {}
        try:
            # Opción para acelerar: saltar exploración de CSVs masivos de facts si así se configura
            if os.getenv('CMF_SKIP_DA_FACTS', '0') == '1':
                raise FileNotFoundError("CMF_SKIP_DA_FACTS=1")
            import re as _re
            from glob import glob
            import pandas as _pd
            fname = file_path.name  # estados_<rut>_<YYYY-YYYY>_<lang>.xlsx
            m = _re.match(r"^estados_(?P<rut>[^_]+)_(?P<yr>\d{4}-\d{4})_(?P<lng>es|en)\.xlsx$", fname, flags=_re.I)
            rut_pat = m.group("rut") if m else ""
            lang_sfx = m.group("lng") if m else ("en" if lang=="en" else "es")
            other_lang_sfx = "es" if lang_sfx == "en" else "en"

            # Ubicar out_consolidated del rut - priorizar Total sobre Anual para trimestres
            roots = glob(str(_xbrl_search_root() / "**/out_consolidated_*"), recursive=True)
            target_dir = None
            # Buscar primero en Total (que incluye trimestres), luego en Anual
            for path_type in ["Total", "Anual"]:
                for p in sorted(roots):
                    base = os.path.basename(p)
                    path_str = str(p)
                    # Búsqueda más flexible: buscar por RUT en el path o por patrón out_consolidated_
                    if f"/XBRL/{path_type}/" in path_str:
                        if (rut_pat in path_str and base.startswith("out_consolidated")) or base.startswith(f"out_consolidated_{rut_pat}"):
                            target_dir = Path(p)
                            # Using D&A facts from target_dir
                            break
                if target_dir:
                    break
            if not target_dir:
                raise FileNotFoundError("No out_consolidated dir found")

            def _to_lbl(col: str) -> str:
                s = str(col).strip()
                try:
                    d = _pd.to_datetime(s, errors='raise')
                    mth = int(d.strftime('%m'))
                    q = {3:'Q1',6:'Q2',9:'Q3',12:'Q4'}.get(mth)
                    return f"{d.year}{q}" if q else str(d.year)
                except Exception:
                    # Handle bare YYYY -> YYYYQ4 for backward compat
                    import re as _re_lbl
                    if _re_lbl.match(r"^\d{4}$", s):
                        return f"{s}Q4"
                    return s

            def _parse_shares(val):
                if val is None:
                    return None
                s = str(val).strip()
                if s == "":
                    return None

                # Negativos con paréntesis
                neg = s.startswith('(') and s.endswith(')')
                if neg:
                    s = s[1:-1].strip()

                # Quita separadores de miles (comas)
                s = s.replace(',', '')

                # Si tiene decimales, quita la parte decimal
                if '.' in s:
                    s = s.split('.')[0]

                # Solo dígitos
                if not s.isdigit():
                    return None

                v = int(s)
                return -v if neg else v


            def _parse_thousands(v: object) -> float | None:
                """Convierte strings numéricos variados a miles (float).

                Soporta:
                  - Separadores de miles: ",", ".", espacios (incluye NBSP)
                  - Negativos con paréntesis: (1234) → -1234
                  - Signos unicode: −, –
                Devuelve None si no es parseable.
                """
                if v is None or v == "":
                    return None
                try:
                    s = str(v).strip()
                    # Normalizar espacios y signos
                    s = s.replace('\u00A0', ' ').replace('−', '-').replace('–', '-')
                    neg = False
                    if s.startswith('(') and s.endswith(')'):
                        neg = True
                        s = s[1:-1]
                    # Eliminar separadores de miles comunes
                    s = s.replace(',', '').replace(' ', '').replace('.', '')
                    if s in ('', '-', '+'):
                        return None
                    num = float(s)
                    if neg:
                        num = -num
                    return num / 1000.0
                except Exception:
                    return None

            def _extract_da(df: _pd.DataFrame) -> dict[str, float]:
                out: dict[str, float] = {}
                ser = df['Label'].astype(str)
                # Filtrar abstracciones y totales que incluyen impairment
                not_abstract = ~ser.str.contains("[abstract]", case=False, na=False, regex=False) & \
                               ~ser.str.contains("[resumen]", case=False, na=False, regex=False)
                not_impair_total = ~ser.str.contains("impairment", case=False, na=False, regex=False) & \
                                   ~ser.str.contains("total", case=False, na=False, regex=False)

                # Combinada (ES/EN) sin abstract/total/impairment
                mask_comb = (
                    ser.str.contains(r"Gastos?\s+por\s+depreciaci[oó]n\s+y\s+amortizaci[oó]n", case=False, na=False) |
                    ser.str.contains(r"Depreciation\s+and\s+amorti[sz]ation\s+expense", case=False, na=False) |
                    ser.str.contains(r"Depreciation\s+and\s+amorti[sz]ation", case=False, na=False)
                ) & not_abstract & not_impair_total

                if mask_comb.any():
                    row = df.loc[mask_comb].iloc[0]
                    nonempty_found = False
                    for col in df.columns:
                        if col == 'Label':
                            continue
                        lbl = _to_lbl(col)
                        val = _parse_thousands(row.get(col))
                        if val is not None and not _pd.isna(val):
                            out[lbl] = val
                            nonempty_found = True
                    if nonempty_found:
                        return out
                # Separadas
                  
                m_dep = (
                    (
                        ser.str.contains(r"^\s*Depreciation\s+expense\s*$", case=False, na=False)
                        | ser.str.contains(r"^\s*Gastos?\s+por\s+depreciaci[oó]n\s*$", case=False, na=False)
                        | ser.str.contains(r"^\s*Depreciaci[oó]n\s*$", case=False, na=False)
                    )
                    & not_abstract
                )

                m_amo = (
                    (
                        ser.str.contains(r"^\s*Amorti[sz]ation\s+expense\s*$", case=False, na=False)
                        | ser.str.contains(r"^\s*Gastos?\s+por\s+amortizaci[oó]n\s*$", case=False, na=False)
                        | ser.str.contains(r"^\s*Amortizaci[oó]n\s*$", case=False, na=False)
                    )
                    & not_abstract
                )

                row_dep = df.loc[m_dep].iloc[0] if m_dep.any() else None
                row_amo = df.loc[m_amo].iloc[0] if m_amo.any() else None
                if row_dep is None and row_amo is None:
                    # Fallback 2: intentar Total D&A+Impairment - Impairment
                    try:
                        m_total = ser.str.contains(r"^\s*Total\s+depreciation.*impairment", case=False, na=False)
                        m_imp = ser.str.contains(r"impairment\s+loss\s*\(reversal.*\)\s*recognised\s*in\s*profit\s*or\s*loss", case=False, na=False)
                        row_total = df.loc[m_total].iloc[0] if m_total.any() else None
                        row_imp = df.loc[m_imp].iloc[0] if m_imp.any() else None
                        if row_total is not None:
                            for col in df.columns:
                                if col == 'Label':
                                    continue
                                lbl = _to_lbl(col)
                                vt = _parse_thousands(row_total.get(col))
                                vi = _parse_thousands(row_imp.get(col)) if row_imp is not None else 0
                                if vt is not None and not _pd.isna(vt):
                                    out[lbl] = vt - (vi or 0)
                            if out:
                                return out
                    except Exception:
                        return out
                for col in df.columns:
                    if col == 'Label':
                        continue
                    lbl = _to_lbl(col)
                    v_dep = _parse_thousands(row_dep.get(col)) if row_dep is not None else None
                    v_amo = _parse_thousands(row_amo.get(col)) if row_amo is not None else None
                    # considerar NaN como None
                    v_dep = None if (v_dep is not None and _pd.isna(v_dep)) else v_dep
                    v_amo = None if (v_amo is not None and _pd.isna(v_amo)) else v_amo
                    if v_dep is not None or v_amo is not None:
                        out[lbl] = (v_dep or 0) + (v_amo or 0)
                return out

            def _extract_shares(df: _pd.DataFrame) -> dict[str, float]:
                out: dict[str, float] = {}

                # Columnas tipo concepto/etiqueta
                qname_col = next((c for c in ['qname', 'QName', 'concept', 'Concept'] if c in df.columns), None)
                label_col = 'Label' if 'Label' in df.columns else None

                # Columnas meta a omitir (alineado con _extract_da)
                META_COLS = {'Label', 'qname', 'QName', 'concept', 'Concept', 'name', 'Name',
                            'unitRef', 'decimals', 'precision', 'contextRef', 'id', 'uuid',
                            'Value', 'value'}

                # ---------- 1) Preferir QNAME: "NumberOfSharesIssued" ----------
                mask_q = _pd.Series([False] * len(df))
                if qname_col:
                    mask_q = df[qname_col].astype(str).str.contains(
                        r"(?:^|[:/])NumberOfSharesIssued$", case=False, na=False, regex=True
                    )

                if mask_q.any():
                    row = df.loc[mask_q].iloc[0]
                    for col in df.columns:
                        if col in META_COLS:
                            continue
                        lbl = _to_lbl(col)
                        val = _parse_shares(row.get(col))
                        if val is not None and not _pd.isna(val):
                            out[lbl] = val
                    return out

                # ---------- 2) Fallback por LABEL: "Total número de acciones emitidas" ----------
                if label_col:
                    ser = df[label_col].astype(str)
                    # Excluir abstracts/resúmenes por si vienen en facts
                    not_abstract = ~ser.str.contains(r"\[(?:abstract|resumen)\]", case=False, na=False, regex=True)

                    mask_l = (
                        ser.str.contains(r"^\s*Total\s+n(?:ú|u)mero\s+de\s+acciones\s+emitidas\s*$", case=False, na=False, regex=True)
                        | ser.str.contains(r"^\s*Total\s+number\s+of\s+shares\s+issued\s*$", case=False, na=False, regex=True)  # opcional EN
                    ) & not_abstract

                    if mask_l.any():
                        row = df.loc[mask_l].iloc[0]
                        for col in df.columns:
                            if col in META_COLS:
                                continue
                            lbl = _to_lbl(col)
                            val = _parse_shares(row.get(col))
                            if val is not None and not _pd.isna(val):
                                out[lbl] = val
                        return out

                return out


            # Extraer del idioma actual y, si falta, del otro
            path_es = next(iter(target_dir.glob(f"facts_*_es.csv")), None)
            path_en = next(iter(target_dir.glob(f"facts_*_en.csv")), None)
            prefer_es = (os.getenv('CMF_DA_PREFER_ES', '0') == '1')
            ordered = [p for p in ([path_es, path_en] if prefer_es else ([path_es, path_en] if lang_sfx=='es' else [path_en, path_es])) if p]
            for idx, pth in enumerate(ordered):
                try:
                    if pth.exists():
                        df_any = _pd.read_csv(pth)
                        m = _extract_da(df_any)
                        # Primera fuente: agrega todo; siguientes solo rellenan faltantes
                        if idx == 0:
                            da_values_map.update(m)
                        else:
                            for k, v in m.items():
                                if k not in da_values_map or da_values_map[k] in (None, ""):
                                    da_values_map[k] = v
                except Exception:
                    pass

            # Extraer Shares
            shares_values_map: dict[str, float] = {}
            for idx, pth in enumerate(ordered):
                try:
                    if pth.exists():
                        df_any = _pd.read_csv(pth)
                        m = _extract_shares(df_any)
                        if idx == 0:
                            shares_values_map.update(m)
                        else:
                            for k, v in m.items():
                                if k not in shares_values_map or shares_values_map[k] is None:
                                    shares_values_map[k] = v
                except Exception:
                    pass
            if os.getenv('CMF_DEBUG_DA', '0') == '1': # Reusing debug flag
                # Shares map loaded
                pass
      
        except Exception:
            pass
        # Fallback adicional: leer de los libros fuente (ES/EN) si facts no aportan
        try:
            def _collect_da_from_wb(xlsx_path: Path) -> dict[str, float]:
                out: dict[str, float] = {}
                try:
                    wb_src = load_workbook(str(xlsx_path), read_only=True, data_only=True)
                except Exception:
                    return out
                # localizar hoja P&L
                pl_names = ["Estado de Resultados", "Estado Resultados (Función)", "Income Statement"]
                sh = None
                for nm in pl_names:
                    if nm in wb_src.sheetnames:
                        sh = wb_src[nm]
                        break
                if sh is None:
                    return out
                # detectar fila encabezado
                import re as _re
                hdr = 3
                for r in range(1, min(10, sh.max_row)+1):
                    v0 = sh.cell(row=r, column=1).value
                    if isinstance(v0, str) and v0.strip().lower() in ("cuenta","concepto","account"):
                        hdr = r
                        break
                    for c in range(2, min(sh.max_column, 25)+1):
                        v = sh.cell(row=r, column=c).value
                        if isinstance(v, str) and (_re.match(r"^\d{4}(Q[1-4])?$", v.strip()) or _re.match(r"^\d{4}-\d{2}-\d{2}$", v.strip())):
                            hdr = r
                            break
                # construir mapa de columnas por etiqueta visible
                def _to_lbl_from_header(h: str) -> str:
                    s = str(h).strip().split("\n",1)[0]
                    if _re.match(r"^\d{4}Q[1-4]$", s):
                        return s
                    m = _re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
                    if m:
                        y, mo = int(m.group(1)), int(m.group(2))
                        q = {3:'Q1',6:'Q2',9:'Q3',12:'Q4'}.get(mo)
                        return f"{y}{q}" if q else str(y)
                    # Bare YYYY -> YYYYQ4 for backward compat
                    if _re.match(r"^\d{4}$", s):
                        return f"{s}Q4"
                    return s
                col_map: dict[str, int] = {}
                for c in range(2, sh.max_column+1):
                    hv = sh.cell(row=hdr, column=c).value
                    if isinstance(hv, str):
                        col_map[_to_lbl_from_header(hv)] = c
                # localizar fila D&A (combinada o suma de separadas)
                da_row_idx = None
                dep_row_idx = None
                amo_row_idx = None
                for r in range(hdr+1, sh.max_row+1):
                    name = sh.cell(row=r, column=1).value
                    if not isinstance(name, str):
                        continue
                    name_s = name.strip().lower()
                    # combinar (evitar abstract/total/impairment)
                    if (("depreciation and amortisation expense" in name_s or "depreciation and amortization expense" in name_s or "gastos por depreciación y amortización" in name_s or "gasto por depreciación y amortización" in name_s)
                        and ("abstract" not in name_s) and ("total" not in name_s) and ("impairment" not in name_s)):
                        da_row_idx = r
                        break
                    if dep_row_idx is None and (name_s == "depreciation expense" or name_s == "depreciación" or name_s == "gastos por depreciación" or name_s == "gasto por depreciación"):
                        dep_row_idx = r
                    if amo_row_idx is None and (("amortisation" in name_s) or ("amortization" in name_s) or name_s == "amortización" or name_s == "gastos por amortización" or name_s == "gasto por amortización"):
                        amo_row_idx = r
                if da_row_idx:
                    for lbl, col in col_map.items():
                        val = sh.cell(row=da_row_idx, column=col).value
                        if isinstance(val, (int, float)):
                            out[lbl] = float(val)
                elif dep_row_idx or amo_row_idx:
                    for lbl, col in col_map.items():
                        v_dep = sh.cell(row=dep_row_idx, column=col).value if dep_row_idx else None
                        v_amo = sh.cell(row=amo_row_idx, column=col).value if amo_row_idx else None
                        s = 0.0
                        anyv = False
                        if isinstance(v_dep, (int, float)):
                            s += float(v_dep); anyv = True
                        if isinstance(v_amo, (int, float)):
                            s += float(v_amo); anyv = True
                        if anyv:
                            out[lbl] = s
                return out
            # si el map está vacío, intentar del libro actual y del del otro idioma dentro de Products/Trimestral o Anual
            if not da_values_map:
                da_values_map.update(_collect_da_from_wb(file_path))
            if not da_values_map:
                opp = None
                s = str(file_path)
                if s.endswith("_en.xlsx"):
                    opp = Path(s.replace("_en.xlsx", "_es.xlsx"))
                elif s.endswith("_es.xlsx"):
                    opp = Path(s.replace("_es.xlsx", "_en.xlsx"))
                if opp and opp.exists():
                    da_values_map.update(_collect_da_from_wb(opp))
        except Exception:
            pass

        # Escribir D&A: SIEMPRE preferir valores desde facts; si no hay, usar helpers (TTM para trimestral, anual para anual)
        # D&A map loaded
        for j, header_label in enumerate(headers_in_sheet, start=2):
            lbl = header_label if isinstance(header_label, str) else None
            val = da_values_map.get(lbl)
            # Fallbacks label ↔ map: 'YYYY' ↔ 'YYYYQ4'
            if (val is None) and isinstance(lbl, str):
                import re as _re
                m_year = _re.match(r"^(\d{4})$", lbl)
                if m_year:
                    alt = f"{m_year.group(1)}Q4"
                    val = da_values_map.get(alt)
                else:
                    m_q = _re.match(r"^(\d{4})Q4$", lbl)
                    if m_q:
                        alt2 = m_q.group(1)
                        val = da_values_map.get(alt2)
            if val is not None and val != "":
                ws.cell(row=current_row, column=j).value = val
            else:
                # Fallback a fórmulas si no hay fact
                try:
                    if isinstance(lbl, str) and lbl.endswith(('Q1','Q2','Q3','Q4')):
                        da = fb._build_da_ttm(lbl)
                        if da:
                            ws.cell(row=current_row, column=j).value = f"={da}"
                    else:
                        import re as _re
                        m = _re.match(r"^(\d{4})", lbl or "")
                        if m:
                            da = fb._build_da_annual(int(m.group(1)))
                            if da:
                                ws.cell(row=current_row, column=j).value = f"={da}"
                except Exception:
                    pass
        self.formatter.format_ratio_row(ws, current_row, daname, years, "number")
        da_row = current_row
        current_row += 1

        # Si D&A se recuperó desde los facts XBRL, corregir la nota de "CALIDAD DE
        # DATOS": el chequeo del extractor sólo mira el estado de resultados (donde
        # bajo IFRS la D&A no aparece como línea), pero el dato SÍ está disponible
        # desde la fuente XBRL. Evita la advertencia engañosa "EBITDA Margin N/A".
        try:
            if da_values_map and any(v not in (None, "") for v in da_values_map.values()):
                q = financial_data.get("_quality")
                if isinstance(q, dict):
                    q["warnings"] = [w for w in q.get("warnings", [])
                                     if "D&A" not in str(w) and "EBITDA Margin will show N/A" not in str(w)]
                    _drop = {"Depreciación y amortización (D&A)", "Depreciación", "Amortización",
                             "Depreciation and amortization (D&A)", "Depreciation", "Amortization"}
                    q["missing_accounts"] = [m for m in q.get("missing_accounts", [])
                                             if str(m).strip() not in _drop]
                    fa = q.setdefault("found_accounts", [])
                    _da_label = "Depreciación y amortización (D&A)" if lang == "es" else "Depreciation and amortization (D&A)"
                    if _da_label not in fa:
                        fa.append(_da_label)
        except Exception:
            pass

        # 1.5) Number of Shares Issued
        shares_name = "Total número de acciones emitidas" if lang == "es" else "Total number of shares issued"
        ws.cell(row=current_row, column=1, value=shares_name)
        
        for j, header_label in enumerate(headers_in_sheet, start=2):
            lbl = header_label if isinstance(header_label, str) else None
            val = shares_values_map.get(lbl)
            
            if (val is None) and isinstance(lbl, str):
                import re as _re
                m_year = _re.match(r"^(\d{4})$", lbl)
                if m_year:
                    alt = f"{m_year.group(1)}Q4"
                    val = shares_values_map.get(alt)
                else:
                    m_q = _re.match(r"^(\d{4})Q4$", lbl)
                    if m_q:
                        alt2 = m_q.group(1)
                        val = shares_values_map.get(alt2)

            if val is not None and val != "":
                ws.cell(row=current_row, column=j).value = val

        self.formatter.format_ratio_row(ws, current_row, shares_name, years, "number")
        shares_row = current_row # In case it's needed later
        current_row += 1

        # 2) EBITDA = EBIT (+ D&A de la fila anterior)
        kpi_name = "EBITDA"
        ws.cell(row=current_row, column=1, value=kpi_name)
        try:
            fb.build_profitability_formulas()  # ensure helpers set
            for j, header_label in enumerate(headers_in_sheet, start=2):
                if not isinstance(header_label, str):
                    continue
                lb = header_label.strip()
                # No se restringe por trimestres: solo años (YYYY)
                ebit_formula = None
                if lb.endswith(('Q1','Q2','Q3','Q4')):
                    # CAMBIO: Usar EBIT directo del período SIN TTM
                    ebit_col = None
                    for c in range(2, fb.sh_pl.max_column + 1):
                        cell_val = fb.sh_pl.cell(row=fb.HDR_PL, column=c).value
                        if cell_val and str(cell_val).strip() == lb:
                            ebit_col = c
                            break
                    if ebit_col:
                        ebit_col_letter = get_column_letter(ebit_col)
                        ebit_formula = fb.create_cell_reference(fb.sh_pl.title, ebit_col_letter, fb.rows_pl["EBIT"])
                else:
                    import re as _re
                    m = _re.match(r"^(\d{4})", lb)
                    if m:
                        y = int(m.group(1))
                        ebit_formula = fb.create_cell_reference(fb.sh_pl.title, fb.find_year_column(fb.sh_pl, y), fb.rows_pl.get("EBIT"))
                if ebit_formula:
                    da_cell = ws.cell(row=da_row, column=j).coordinate
                    ws.cell(row=current_row, column=j).value = f"=IFERROR({ebit_formula}+IFERROR({da_cell},0),\"\")"
        except Exception:
            pass
        self.formatter.format_ratio_row(ws, current_row, kpi_name, years, "number")
        ebitda_row = current_row
        current_row += 1

        if progress_cb:
            try:
                progress_cb('utilities', 3, 5)
            except Exception:
                pass

        # Reescribir ratios para que usen el EBITDA único de UTILIDADES
        # 2.1) Margen EBITDA
        try:
            # localizar fila de Margen EBITDA / EBITDA Margin
            margen_row = None
            for r in range(header_row+1, ws.max_row+1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip() in ("Margen EBITDA", "EBITDA Margin"):
                    margen_row = r
                    break
            if margen_row:
                for j, header_label in enumerate(headers_in_sheet, start=2):
                    if not isinstance(header_label, str):
                        continue
                    lb = header_label.strip()
                    # Solo años (YYYY)
                    # Ventas (TTM o anual)
                    ven_formula = None
                    if lb.endswith(('Q1','Q2','Q3','Q4')):
                        # CAMBIO: Usar ventas directas del período SIN TTM
                        ven_col = None
                        for c in range(2, fb.sh_pl.max_column + 1):
                            cell_val = fb.sh_pl.cell(row=fb.HDR_PL, column=c).value
                            if cell_val and str(cell_val).strip() == lb:
                                ven_col = c
                                break
                        if ven_col:
                            ven_col_letter = get_column_letter(ven_col)
                            ven_formula = fb.create_cell_reference(fb.sh_pl.title, ven_col_letter, fb.rows_pl["Ventas"])
                    else:
                        import re as _re
                        m = _re.match(r"^(\d{4})", lb)
                        if m:
                            y = int(m.group(1))
                            ven_formula = fb.create_cell_reference(fb.sh_pl.title, fb.find_year_column(fb.sh_pl, y), fb.rows_pl.get("Ventas"))
                    if ven_formula:
                        ebitda_cell = ws.cell(row=ebitda_row, column=j).coordinate
                        ws.cell(row=margen_row, column=j).value = f"=IFERROR({ebitda_cell}/{ven_formula},\"\")"
        except Exception:
            pass

        # 2.2) Deuda / EBITDA usando el EBITDA único
        try:
            deuda_row = None
            for r in range(header_row+1, ws.max_row+1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, str) and v.strip() in ("Deuda / EBITDA", "Debt / EBITDA"):
                    deuda_row = r
                    break
            if deuda_row:
                for j, header_label in enumerate(headers_in_sheet, start=2):
                    if not isinstance(header_label, str):
                        continue
                    lb = header_label.strip()
                    # Solo años (YYYY)
                    # Net Debt = (DeudaFinCorr + DeudaFinNC + ArrCorr + ArrNC) - Efectivo
                    def _by_label(sheet, rownum, label):
                        try:
                            col = fb._get_col_letter_by_label(sheet, label)
                            return f"'{sheet.title}'!{col}{rownum}" if col and rownum else None
                        except Exception:
                            return None
                    if lb.endswith(('Q1','Q2','Q3','Q4')):
                        dfc = _by_label(fb.sh_bal, fb.rows_bal.get("DeudaFinCorr"), lb)
                        dfnc= _by_label(fb.sh_bal, fb.rows_bal.get("DeudaFinNC"), lb)
                        arrc= _by_label(fb.sh_bal, fb.rows_bal.get("ArrCorr"), lb)
                        arrn= _by_label(fb.sh_bal, fb.rows_bal.get("ArrNC"), lb)
                        cash= _by_label(fb.sh_bal, fb.rows_bal.get("Efec"), lb)
                    else:
                        import re as _re
                        m = _re.match(r"^(\d{4})", lb)
                        y = int(m.group(1)) if m else None
                        def _by_year(sheet, rownum, year):
                            col = fb.find_year_column(sheet, year) if year else None
                            return fb.create_cell_reference(sheet.title, col, rownum) if col and rownum else None
                        dfc = _by_year(fb.sh_bal, fb.rows_bal.get("DeudaFinCorr"), y)
                        dfnc= _by_year(fb.sh_bal, fb.rows_bal.get("DeudaFinNC"), y)
                        arrc= _by_year(fb.sh_bal, fb.rows_bal.get("ArrCorr"), y)
                        arrn= _by_year(fb.sh_bal, fb.rows_bal.get("ArrNC"), y)
                        cash= _by_year(fb.sh_bal, fb.rows_bal.get("Efec"), y)
                    parts = [p for p in [dfc, dfnc, arrc, arrn] if p]
                    if parts:
                        net_debt = "+".join(parts)
                        if cash:
                            net_debt = f"({net_debt})-IFERROR({cash},0)"
                        ebitda_cell = ws.cell(row=ebitda_row, column=j).coordinate
                        # Parentizar el numerador completo
                        ws.cell(row=deuda_row, column=j).value = f"=IFERROR(({net_debt})/{ebitda_cell},\"\")"
        except Exception:
            pass

        # Aplicar formateo condicional
        data_start_row = header_row + 1
        data_end_row = current_row - 1
        self.formatter.apply_conditional_formatting(ws, data_start_row, data_end_row, years)
        
        # Agregar sección tooltip (localizando nombres si corresponde)
        tooltip_start = current_row + 2
        # En modo Total, no incluir bloque de nota trimestral en tooltips
        include_quarterly_note = (has_quarters and bool(labels_norm) and not combined_mode)
        if lang == "en":
            localized_blocks = []
            for sec, items in formula_blocks:
                sec_loc = section_map.get(sec, sec)
                new_items = []
                for (nm, kind, func, desc) in items:
                    nm_loc = ratio_map.get(nm, nm)
                    desc_loc = desc_map.get(desc, desc)
                    new_items.append((nm_loc, kind, func, desc_loc))
                localized_blocks.append((sec_loc, new_items))
            self.formatter.create_tooltip_section(ws, tooltip_start, localized_blocks, years, cols_total, lang=lang, include_quarterly_note=include_quarterly_note)
        else:
            self.formatter.create_tooltip_section(ws, tooltip_start, formula_blocks, years, cols_total, lang=lang, include_quarterly_note=include_quarterly_note)
        
        # Congelar paneles
        self.formatter.add_freeze_panes(ws)

        # (Eliminado) Hoja README
        
        # Crear hoja de NOTAS/NOTES (siempre, con calidad de datos + moneda si aplica)
        try:
            from openpyxl.styles import Font, Alignment
            notes_name = "NOTES" if lang == "en" else "NOTAS"
            # Construir cmap (facts → fallback)
            cmap = cmap_facts or {}
            # Fallback a extractor si no hay facts
            if not cmap:
                try:
                    cby = financial_data.get('currency_by_year') or {}
                    if cby:
                        cmap = {int(y): str(cby[y]) for y in cby.keys()}
                except Exception:
                    cmap = {}
            unique_codes = set(cmap.values())
            has_multi_currency = cmap and len(unique_codes) > 1
            if notes_name in wb.sheetnames:
                del wb[notes_name]
            ws_notes = wb.create_sheet(notes_name, index=len(wb.sheetnames))
            ws_notes.column_dimensions['A'].width = 30
            ws_notes.column_dimensions['B'].width = 20
            ws_notes.column_dimensions['C'].width = 40
            notes_row = 1

            # --- Data Quality section (always) ---
            quality = financial_data.get("_quality", {})
            if quality:
                notes_row = self.formatter.format_quality_section(
                    ws_notes, notes_row, quality, cols_total=3, lang=lang
                )
                notes_row += 1  # spacer

            # --- Disclaimer ---
            disclaimer = (
                "Data extracted from XBRL reports published by Chile's Financial Market Commission (CMF). findatachile.com"
                if lang == 'en' else
                "Datos extraídos de reportes XBRL publicados por la Comisión para el Mercado Financiero (CMF) de Chile. findatachile.com"
            )
            ws_notes.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)
            disc_cell = ws_notes.cell(row=notes_row, column=1, value=disclaimer)
            disc_cell.font = Font(italic=True, color="6B7280", size=9)
            disc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            notes_row += 2  # spacer

            # --- Currency information (if multiple currencies) ---
            if has_multi_currency:
                title = "Notes — Currency information" if lang == 'en' else "Notas — Información de moneda"
                ws_notes.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)
                ctitle = ws_notes.cell(row=notes_row, column=1, value=title)
                ctitle.font = Font(bold=True, size=13)
                ctitle.alignment = Alignment(horizontal="left", vertical="center")
                notes_row += 1
                src = "ifrs:DescriptionOfPresentationCurrency / Descripción de la moneda de presentación"
                ws_notes.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)
                source_label = "Facts" if (_get_currency_map_from_facts() or {}) else ("Workbook notes" if lang=='en' else "Notas del libro base")
                csrc = ws_notes.cell(row=notes_row, column=1, value=("Source: " if lang=='en' else "Fuente: ") + src + (f"  ({source_label})"))
                csrc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                notes_row += 1
                ws_notes.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)
                cspan = ws_notes.cell(row=notes_row, column=1, value=("Currency ranges: " if lang=='en' else "Rangos de moneda: ") + (spans_text or ""))
                cspan.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                notes_row += 1
                # Mensaje narrativo de transición
                def _transitions_text(spans: list[tuple[int,int,str]]) -> str:
                    parts = []
                    for (a,b,c) in spans:
                        rng = f"{a}-{b}" if a != b else f"{a}"
                        parts.append(f"{c} ({rng})")
                    def normalize_currency(code: str) -> str:
                        code_upper = code.upper().strip()
                        if code_upper in ['PESOS', 'CLP', 'PESO', 'PESOS CHILENOS', 'PESO CHILENO']:
                            return 'CLP'
                        return code_upper
                    normalized_codes = [normalize_currency(c) for _, _, c in spans]
                    unique_normalized = set(normalized_codes)
                    if len(unique_normalized) == 1:
                        if lang == 'en':
                            return f"Financial statements were reported in {spans[0][2]} (equivalent to {normalized_codes[0]}) throughout the period."
                        else:
                            return f"Los estados financieros se publicaron en {spans[0][2]} (equivalente a {normalized_codes[0]}) durante todo el período."
                    if lang == 'en':
                        if len(parts) >= 2:
                            return "Financial statements were reported in " + ", ".join(parts[:-1]) + ", then in " + parts[-1] + "."
                        else:
                            return "Financial statements were reported in " + parts[0] + "."
                    else:
                        if len(parts) >= 2:
                            return "Los estados financieros se publicaron en " + ", ".join(parts[:-1]) + ", y luego en " + parts[-1] + "."
                        else:
                            return "Los estados financieros se publicaron en " + parts[0] + "."
                ys = sorted(cmap.keys())
                spans = []
                cur_c = None; start_y = None
                for y in ys:
                    c = cmap[y]
                    if cur_c is None:
                        cur_c = c; start_y = y
                    elif c != cur_c:
                        spans.append((start_y, ys[ys.index(y)-1], cur_c))
                        cur_c = c; start_y = y
                if cur_c is not None:
                    spans.append((start_y, ys[-1], cur_c))
                ws_notes.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=3)
                ws_notes.cell(row=notes_row, column=1, value=_transitions_text(spans)).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                notes_row += 1
                if cmap:
                    ws_notes.cell(row=notes_row, column=1, value=("Year" if lang=='en' else "Año")).font = Font(bold=True)
                    ws_notes.cell(row=notes_row, column=2, value=("Currency" if lang=='en' else "Moneda")).font = Font(bold=True)
                    notes_row += 1
                    for y in sorted(cmap.keys()):
                        ws_notes.cell(row=notes_row, column=1, value=int(y))
                        ws_notes.cell(row=notes_row, column=2, value=str(cmap[y]))
                        notes_row += 1
            # Estilo profesional a tabla de notas
            try:
                self.formatter.format_notes_table(ws_notes, start_row=1, cols=3)
            except Exception:
                pass
        except Exception:
            pass

        # --- Methodology sheet (Metodología / Methodology) ---
        try:
            self._create_methodology_sheet(wb, lang=lang)
        except Exception:
            pass

        # --- Metadata sheet (Ficha Técnica) ---
        try:
            import re as _re_meta
            # Extract RUT from filename (format: estados_<RUT>_<range>_<lang>.xlsx)
            m_rut = _re_meta.match(r"^estados_(?P<rut>[^_]+)_", file_path.name)
            rut = m_rut.group("rut") if m_rut else ""
            # Detect frequency from environment or fallback
            freq = "Total"
            if os.getenv('CMF_ANALYSIS_COMBINED', '0') == '1':
                freq = "Total (Anual + Trimestral)" if lang == "es" else "Total (Annual + Quarterly)"
            elif has_quarters:
                freq = "Trimestral" if lang == "es" else "Quarterly"
            else:
                freq = "Anual" if lang == "es" else "Annual"
            create_metadata_sheet(
                wb=wb,
                company_name=company_name,
                rut=rut,
                periods=display_labels,
                frequency=freq,
                lang=lang,
                currency=unit_text or ("Miles de CLP" if lang == "es" else "Thousands CLP"),
            )
        except Exception:
            pass

        # Guardar archivo: incluir pista de idioma para evitar confusión en carpetas mixtas
        output_filename = (
            f"{company_name}_en - Financial Analysis (Formulas).xlsx" if lang == "en" else f"{company_name}_es - Análisis Financiero (Fórmulas).xlsx"
        )
        output_path = self.output_dir / output_filename

        # Agregar funcionalidad DCF antes de guardar
        try:
            if add_dcf_functionality is not None:
                if progress_cb:
                    try:
                        progress_cb('dcf_building', 4, 6)
                    except Exception:
                        pass
                add_dcf_functionality(wb, financial_data)
        except Exception:
            pass

        if progress_cb:
            try:
                progress_cb('saving', 5, 6)
            except Exception:
                pass
        wb.save(str(output_path))
        # Si hubo uso de ES_DATA, escribir log en Product_v1
        try:
            entries = getattr(formula_builder, 'es_usage_log_entries', None)
            if entries:
                log_path = self.output_dir / f"{company_name}_{'en' if lang=='en' else 'es'}_ES_DATA_usage.txt"
                with open(log_path, 'w', encoding='utf-8') as f:
                    f.write("Uso de ES_DATA (cuando valores difieren del libro base)\n")
                    f.write("Formato: GRUPO:KEY | SHEET | label= | base= val= vs es= → using ...\n\n")
                    for line in entries:
                        f.write(line + "\n")
        except Exception:
            pass
        
        return output_filename

