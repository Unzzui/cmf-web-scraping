"""
Excel Formatter Module
=====================

Módulo para formatear y estilizar hojas de análisis financiero en Excel.
Incluye estilos, colores, formateo condicional y estructura visual.
"""

from typing import Dict, List, Optional, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


class ExcelFormatter:
    """
    Formateador de hojas Excel para análisis financiero.
    """
    
    def __init__(self):
        """Inicializa el formateador con estilos predefinidos."""
        self._setup_styles()
    
    def _setup_styles(self):
        """Configura los estilos predefinidos."""
        # Colores de fondo
        # Alinear con paleta usada en xbrl_to_excel.py
        # brand_primary = '#0F172A' (navy oscuro), brand_secondary = '#1F2937' (gris azulado)
        self.header_fill = PatternFill("solid", fgColor="0F172A")
        self.subheader_fill = PatternFill("solid", fgColor="1F2937")
        
        self.section_fills = {
            "LIQUIDEZ": PatternFill("solid", fgColor="D1E7DD"),
            "SOLVENCIA Y ESTRUCTURA": PatternFill("solid", fgColor="FAD7A0"),
            "RENTABILIDAD": PatternFill("solid", fgColor="F8D7DA"),
            "EFICIENCIA OPERATIVA": PatternFill("solid", fgColor="D6EAF8"),
            "FLUJOS Y ADICIONALES": PatternFill("solid", fgColor="E8DAEF"),
            "CRECIMIENTO": PatternFill("solid", fgColor="FDEBD0"),
            "DUPONT": PatternFill("solid", fgColor="D4EFDF"),
            "CALIDAD Y SCORES": PatternFill("solid", fgColor="FCF3CF"),
        }
        # Estilo tabla notas
        self.notes_header_fill = PatternFill("solid", fgColor="EFEFEF")
        self.notes_header_font = Font(bold=True, color="000000", size=11)
        self.notes_cell_font = Font(color="000000", size=10)
        self.notes_alt_fill = PatternFill("solid", fgColor="F7FAFC")
        
        # Fuentes
        self.bold_white = Font(bold=True, color="FFFFFF", size=13)
        self.bold_white_small = Font(bold=True, color="FFFFFF", size=11)
        self.bold_dark = Font(bold=True, color="000000", size=11)
        self.normal = Font(color="000000", size=10)
        
        # Alineaciones
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Bordes
        thin = Side(style="thin", color="DDDDDD")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def setup_worksheet_structure(self, ws, periods: List[str], sheet_name: str = "RATIOS & KPIs", lang: str = "es", unit_text: str | None = None):
        """
        Configura la estructura básica de la hoja de análisis.
        
        Args:
            ws: Hoja de trabajo
            periods: Lista de períodos disponibles (años y/o trimestres)
            sheet_name: Nombre de la hoja
            lang: Idioma de la hoja ('es' o 'en')
        """
        num_periods = len(periods)
        cols_total = 1 + num_periods + 2 + 1  # Indicador + períodos + Último + Promedio + Tendencia
        
        # Configurar ancho de columnas
        for c in range(1, cols_total + 1):
            ws.column_dimensions[get_column_letter(c)].width = 30 if c == 1 else 14
        
        # Título principal (mejorado)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_total)
        title_text = (
            "Análisis Financiero — Ratios y Tendencias"
            if lang == "es" else
            "Financial Analysis — Ratios & Trends"
        )
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.fill = self.header_fill
        title_cell.font = self.bold_white
        title_cell.alignment = self.center
        
        # Subtítulo (alineado con estilo de xbrl_to_excel: información de unidad y período)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_total)
        # Construcción de texto de períodos como rango
        if periods:
            # Extraer primer y último período para mostrar rango
            first_period = periods[0]
            last_period = periods[-1]
            periods_text = f"{first_period} - {last_period}"
        else:
            periods_text = "-"
        # Unidad configurable; por defecto miles CLP / Thousands CLP
        unit_note = unit_text if unit_text else ("Miles CLP" if lang == "es" else "Thousands CLP")
        subtitle_text = (
            f"Unidad: {unit_note}    •    Períodos: {periods_text}" if lang == "es" else
            f"Unit: {unit_note}    •    Periods: {periods_text}"
        )
        subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
        # Mantener estética clara: texto oscuro sobre fondo claro
        subtitle_cell.font = self.bold_dark
        subtitle_cell.alignment = self.center
        
        # Fila de encabezados
        header_row = 4
        # Encabezados: mantener períodos en orden ascendente (antiguo → nuevo)
        if lang == "es":
            headers = ["Indicador"] + [str(p) for p in periods] + ["Último", "Promedio", "Tendencia"]
        else:
            headers = ["Indicator"] + [str(p) for p in periods] + ["Latest", "Average", "Trend"]
        
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=header)
            cell.font = self.bold_white_small
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border
        
        return header_row
    
    def format_section_header(self, ws, row: int, cols_total: int, section_name: str):
        """
        Formatea una fila de encabezado de sección.
        
        Args:
            ws: Hoja de trabajo
            row: Número de fila
            cols_total: Total de columnas
            section_name: Nombre de la sección
        """
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_total)
        section_cell = ws.cell(row=row, column=1, value=section_name)
        section_cell.font = self.bold_dark
        section_cell.alignment = self.left
        # Permitir claves en inglés para obtener el mismo color de sección
        fill = self.section_fills.get(section_name)
        if fill is None:
            en_to_es = {
                "LIQUIDITY": "LIQUIDEZ",
                "SOLVENCY & CAPITAL STRUCTURE": "SOLVENCIA Y ESTRUCTURA",
                "PROFITABILITY": "RENTABILIDAD",
                "OPERATING EFFICIENCY": "EFICIENCIA OPERATIVA",
                "CASH FLOWS & OTHER": "FLUJOS Y ADICIONALES",
                "VALUE CREATION": "CREACIÓN DE VALOR",
                "COVERAGE & RISK": "COBERTURA Y RIESGO",
                "GROWTH": "CRECIMIENTO",
                "DUPONT": "DUPONT",
                "QUALITY & SCORES": "CALIDAD Y SCORES",
            }
            fill = self.section_fills.get(en_to_es.get(section_name, ""), PatternFill("solid", fgColor="EFEFEF"))
        section_cell.fill = fill
        
        # Aplicar bordes a todas las celdas de la fila
        for c in range(1, cols_total + 1):
            ws.cell(row=row, column=c).border = self.border

    def format_notes_table(self, ws, start_row: int, cols: int = 3) -> None:
        """Aplica estilo de tabla profesional a la hoja de NOTAS."""
        # Encabezados en fila start_row y celdas siguientes según secciones creadas en bulk_processor
        # Estilizar título y subtítulos ya están configurados en bulk; aquí damos bordes a todo el bloque
        max_row = ws.max_row
        # Ajustar anchos típicos
        ws.column_dimensions[get_column_letter(1)].width = 30
        ws.column_dimensions[get_column_letter(2)].width = 18
        ws.column_dimensions[get_column_letter(3)].width = 50
        for r in range(start_row, max_row + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = self.border
                # Título y narrativa
                if r in (start_row, start_row + 2, start_row + 3):
                    cell.fill = self.notes_header_fill
                    cell.font = self.notes_header_font
                    cell.alignment = self.left
                    continue
                # Encabezado de tabla (Year, Currency) en fila start_row+4
                if r == start_row + 4:
                    cell.fill = self.subheader_fill
                    cell.font = self.bold_white_small
                    cell.alignment = self.center
                    continue
                # Filas de datos: alternar relleno suave y centrar año/moneda
                cell.font = self.notes_cell_font
                if (r - (start_row + 4)) % 2 == 1:
                    cell.fill = self.notes_alt_fill
                if c in (1, 2):
                    cell.alignment = self.center
                else:
                    cell.alignment = self.left
    
    def format_ratio_row(self, ws, row: int, indicator_name: str, years: List[int], 
                        ratio_type: str = "ratio"):
        """
        Formatea una fila de ratio financiero.
        
        Args:
            ws: Hoja de trabajo
            row: Número de fila
            indicator_name: Nombre del indicador
            years: Lista de años
            ratio_type: Tipo de ratio (ratio, pct, number, days)
        """
        # Celda del indicador
        indicator_cell = ws.cell(row=row, column=1, value=indicator_name)
        indicator_cell.alignment = self.left
        indicator_cell.border = self.border
        
        # Celdas de datos por año
        for j in range(2, 2 + len(years)):
            cell = ws.cell(row=row, column=j)
            cell.alignment = self.center
            cell.border = self.border
            self._apply_number_format(cell, ratio_type)
        
        # Columnas adicionales: Último, Promedio, Tendencia
        # Note: data cells may contain the text "N/A" for ratio/pct formulas when
        # the underlying financial data is unavailable or produces a division by zero.
        # Those cells will display as plain text and are excluded from conditional
        # formatting heatmaps automatically because they are non-numeric.
        # Call format_na_style() after writing values if grey styling is desired.
        last_col_idx = 1 + len(years) + 1
        avg_col_idx = 1 + len(years) + 2
        trend_col_idx = 1 + len(years) + 3
        
        year_start_col = 2
        year_end_col = 1 + len(years)
        rng = f"{get_column_letter(year_start_col)}{row}:{get_column_letter(year_end_col)}{row}"
        
        # Último valor
        last_cell = ws.cell(row=row, column=last_col_idx)
        last_cell.value = f"=LOOKUP(2,1/(--({rng}<>\"\")),{rng})"
        last_cell.alignment = self.center
        last_cell.border = self.border
        self._apply_number_format(last_cell, ratio_type)
        
        # Promedio
        avg_cell = ws.cell(row=row, column=avg_col_idx)
        avg_cell.value = f"=IFERROR(AVERAGE({rng}),\"\")"
        avg_cell.alignment = self.center
        avg_cell.border = self.border
        self._apply_number_format(avg_cell, ratio_type)
        
        # Tendencia
        trend_cell = ws.cell(row=row, column=trend_col_idx)
        trend_cell.value = f"=IFERROR(IF(" \
            f"(LOOKUP(2,1/(--({rng}<>\"\")),{rng}))>" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▲\"," \
            f"IF((LOOKUP(2,1/(--({rng}<>\"\")),{rng}))<" \
            f"(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>\"\")),{rng}))),{rng})),\"▼\",\"→\"))," \
            f"\"→\")"
        trend_cell.alignment = self.center
        trend_cell.border = self.border
    
    def _apply_number_format(self, cell, ratio_type: str):
        """
        Aplica formato numérico según el tipo de ratio.

        Args:
            cell: Celda a formatear
            ratio_type: Tipo de formato
        """
        if ratio_type == "pct":
            cell.number_format = "0.0%"
        elif ratio_type == "number":
            cell.number_format = "#,##0"
        elif ratio_type == "days":
            cell.number_format = "0"
        else:  # ratio
            cell.number_format = "0.00"

    def format_na_style(self, cell) -> None:
        """
        Applies a muted grey style to a cell whose evaluated value is the text
        "N/A".  Call this after the workbook has been calculated (or when writing
        static "N/A" strings) to visually distinguish unavailable ratios from
        zero values.

        Ratio and percentage formulas in formula_builder.py emit
        ``IFERROR(<division>, "N/A")`` so that a missing denominator is
        immediately identifiable.  This helper provides a consistent visual
        treatment for those cells.

        Args:
            cell: The openpyxl Cell object to style.

        Example::

            for col in ws.iter_cols(min_row=5, max_row=ws.max_row,
                                    min_col=2, max_col=ws.max_column):
                for cell in col:
                    if cell.value == "N/A":
                        formatter.format_na_style(cell)
        """
        cell.font = Font(color="999999", italic=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def format_quality_section(self, ws, start_row: int, quality_data: dict,
                              cols_total: int, lang: str = "es") -> int:
        """Write a 'Data Quality' summary block and return the next free row.

        Parameters
        ----------
        ws : Worksheet
        start_row : first row to write in
        quality_data : dict with keys 'found_accounts', 'missing_accounts',
                       'estimated_accounts', 'warnings', 'missing_sheets',
                       'total_expected' (int).
        cols_total : number of columns to span for the header
        lang : 'es' or 'en'

        Returns
        -------
        int  – the row immediately after the last written row.
        """
        is_en = lang == "en"
        border = self.border
        gray_fill = PatternFill("solid", fgColor="F3F4F6")
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        label_font = Font(bold=True, color="374151", size=10)
        value_font = Font(color="374151", size=10)

        # Section title
        r = start_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols_total)
        title = "DATA QUALITY" if is_en else "CALIDAD DE DATOS"
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, color="000000", size=11)
        c.fill = header_fill
        c.alignment = self.left
        for col in range(1, cols_total + 1):
            ws.cell(row=r, column=col).border = border
        r += 1

        found = quality_data.get("found_accounts", [])
        missing = quality_data.get("missing_accounts", [])
        estimated = quality_data.get("estimated_accounts", [])
        warnings = quality_data.get("warnings", [])
        total_expected = quality_data.get("total_expected", len(found) + len(missing))

        rows_data: list[tuple[str, str]] = [
            (
                "Accounts found" if is_en else "Cuentas encontradas",
                f"{len(found)}/{total_expected}",
            ),
            (
                "Missing accounts" if is_en else "Cuentas faltantes",
                ", ".join(missing[:8]) + ("..." if len(missing) > 8 else "") if missing else ("-" if is_en else "Ninguna"),
            ),
        ]
        if estimated:
            rows_data.append((
                "Estimates applied" if is_en else "Estimaciones aplicadas",
                ", ".join(estimated),
            ))
        if warnings:
            rows_data.append((
                "Warnings" if is_en else "Advertencias",
                "; ".join(warnings),
            ))

        for label, value in rows_data:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=cols_total)
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = label_font
            lc.fill = gray_fill
            lc.alignment = self.left
            lc.border = border
            vc = ws.cell(row=r, column=2, value=value)
            vc.font = value_font
            vc.fill = gray_fill
            vc.alignment = self.left
            vc.border = border
            for col in range(3, cols_total + 1):
                ws.cell(row=r, column=col).fill = gray_fill
                ws.cell(row=r, column=col).border = border
            r += 1

        return r

    def format_flags_column(self, ws, row: int, flags_col: int, formulas: dict,
                            lang: str = "es") -> None:
        """Write IF-based flag formulas in the Notes column for a ratio row.

        Parameters
        ----------
        ws : Worksheet
        row : the row of the ratio
        flags_col : 1-based column index for the notes/flags column
        formulas : dict mapping ratio_name -> (threshold_formula_str)
                   e.g. {"Margen Bruto": '=IF(B5>0.9,"Margen inusualmente alto","")'}
        lang : 'es' or 'en'
        """
        cell = ws.cell(row=row, column=flags_col)
        cell.alignment = self.left
        cell.border = self.border
        cell.font = Font(color="B45309", size=9, italic=True)

    def apply_conditional_formatting(self, ws, data_start_row: int, data_end_row: int,
                                   years: List[int]):
        """
        Aplica formateo condicional a la hoja.
        
        Args:
            ws: Hoja de trabajo
            data_start_row: Fila de inicio de datos
            data_end_row: Fila de fin de datos
            years: Lista de años
        """
        # Si no hay años, no aplicar CF para evitar rangos inválidos
        if not years:
            return

        # Heatmap en columnas de años
        year_start_col = 2
        year_end_col = 1 + len(years)
        year_range = f"{get_column_letter(year_start_col)}{data_start_row}:{get_column_letter(year_end_col)}{data_end_row}"
        
        ws.conditional_formatting.add(
            year_range,
            ColorScaleRule(
                start_type="percentile", start_value=5, start_color="FDE725",
                mid_type="percentile", mid_value=50, mid_color="5DC863",
                end_type="percentile", end_value=95, end_color="2A788E"
            )
        )
        
        # Barras de datos en columnas Último y Promedio
        last_col_letter = get_column_letter(1 + len(years) + 1)
        avg_col_letter = get_column_letter(1 + len(years) + 2)
        
        ws.conditional_formatting.add(
            f"{last_col_letter}{data_start_row}:{last_col_letter}{data_end_row}",
            DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
        )
        
        ws.conditional_formatting.add(
            f"{avg_col_letter}{data_start_row}:{avg_col_letter}{data_end_row}",
            DataBarRule(start_type="min", end_type="max", color="4F81BD", showValue=True)
        )
    
    def add_freeze_panes(self, ws):
        """Congela paneles en la posición apropiada."""
        ws.freeze_panes = "B5"
    
    def add_readme_sheet(self, wb: Workbook, lang: str = "es") -> None:
        # README eliminado por solicitud
        return
    
    def create_tooltip_section(self, ws, start_row: int, formula_blocks: List[Tuple], 
                             years: List[int], cols_total: int, lang: str = "es", 
                             include_quarterly_note: bool = True):
        """
        Crea la sección de tooltip con definiciones y fórmulas.
        
        Args:
            ws: Hoja de trabajo
            start_row: Fila de inicio para la sección
            formula_blocks: Bloques de fórmulas organizados por categoría
            years: Lista de años
            cols_total: Total de columnas
            lang: Idioma ('es' o 'en')
        """
        tip_title_row = start_row
        
        # Título de la sección tooltip
        ws.merge_cells(start_row=tip_title_row, start_column=1, 
                      end_row=tip_title_row, end_column=cols_total)
        tip_title_text = (
            "Definición y Fórmula (tooltip)" if lang == "es" else "Definition and Formula (tooltip)"
        )
        tip_title = ws.cell(row=tip_title_row, column=1, value=tip_title_text)
        tip_title.font = self.bold_dark
        tip_title.alignment = self.left
        tip_title.fill = PatternFill("solid", fgColor="EFEFEF")
        
        # Nota metodológica para trimestrales (TTM)
        if include_quarterly_note:
            note_row = tip_title_row + 1
            ttm_note = (
                "Metodología trimestral: Balance instantáneo (Qn). P&L y Flujos en TTM = YTD(Qn) − YTD(Qn−4). Promedios de saldos: (Qn, Qn−4)."
                if lang == "es" else
                "Quarterly methodology: Balance at quarter-end (Qn). P&L and Cash Flows in TTM = YTD(Qn) − YTD(Qn−4). Balance averages: (Qn, Qn−4)."
            )
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=cols_total)
            note_cell = ws.cell(row=note_row, column=1, value=ttm_note)
            note_cell.alignment = self.left
            tip_header_row = note_row + 1
        else:
            tip_header_row = tip_title_row + 1
        headers = (
            ["Indicador", "Fórmula (texto)", "Ejemplo de Fórmula Excel (último año)"]
            if lang == "es"
            else ["Indicator", "Formula (text)", "Example Excel Formula (latest year)"]
        )
        
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=tip_header_row, column=i, value=header)
            cell.font = self.bold_white_small
            cell.fill = self.subheader_fill
            cell.alignment = self.center
            cell.border = self.border
        
        # Contenido de la sección tooltip
        last_year = years[-1] if years else None
        tip_row = tip_header_row + 1
        
        for section_name, items in formula_blocks:
            # Encabezado de sección en tooltip
            section_cell = ws.cell(row=tip_row, column=1, value=section_name)
            section_cell.font = self.bold_dark
            section_cell.fill = self.section_fills.get(section_name, 
                                                     PatternFill("solid", fgColor="EFEFEF"))
            
            for c in range(1, 4):
                ws.cell(row=tip_row, column=c).border = self.border
            tip_row += 1
            
            # Items de la sección
            for name, kind, func, text_formula in items:
                # Nombre del indicador
                ws.cell(row=tip_row, column=1, value=name).alignment = self.left
                ws.cell(row=tip_row, column=1).border = self.border
                
                # Fórmula en texto
                ws.cell(row=tip_row, column=2, value=text_formula).alignment = self.left
                ws.cell(row=tip_row, column=2).border = self.border
                
                # Ejemplo de fórmula Excel
                example_formula = self._get_example_formula(func, last_year)
                ws.cell(row=tip_row, column=3, value=example_formula).alignment = self.left
                ws.cell(row=tip_row, column=3).border = self.border
                
                tip_row += 1
        
        return tip_row
    
    def _get_example_formula(self, func, year: Optional[int]) -> str:
        """
        Obtiene una fórmula de ejemplo para el último año.
        
        Args:
            func: Función que genera las fórmulas
            year: Año para el ejemplo
            
        Returns:
            Fórmula de ejemplo o string vacío
        """
        if year is None:
            return ""
        
        try:
            formula_map = func()
            formula = formula_map.get(str(year))
            if formula:
                return f"={formula}"
        except:
            pass
        
        return ""
