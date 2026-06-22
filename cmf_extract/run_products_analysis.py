#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Ejecutor de análisis sobre los Excel consolidados en Products/ generados por xbrl_to_excel.

Recorre Products/, ejecuta el análisis (con fórmulas) y guarda los resultados en Product_v1/
con nombres comerciales bonitos, incluyendo SIEMPRE el RUT y el nombre de la empresa.

Uso:
  python run_products_analysis.py [--input-dir Products] [--output-dir Product_v1] [--workers 4]

El nombre de salida seguirá la convención:
  <Empresa> - <RUT> - Análisis Financiero (Fórmulas) <AAAA-AAAA> [ES|EN].xlsx
"""

from __future__ import annotations

import argparse
import concurrent.futures
import os
import re
from pathlib import Path
from typing import Optional, Tuple, Any, List
import threading

from openpyxl import load_workbook

# Importar módulos de análisis
from analisis_excel.bulk_processor import BulkProcessor
try:
    from analisis_excel.utils.console_dashboard import ConsoleXBRLDashboard, get_global_dashboard
except Exception:
    ConsoleXBRLDashboard = None  # type: ignore
    def get_global_dashboard():  # type: ignore
        return None
def cross_fill_da(out_dir: Path) -> None:
    # Recorre pares ES/EN en out_dir y copia valores de D&A en UTILIDADES si faltan en uno de los dos
    from openpyxl import load_workbook
    debug = os.getenv('CMF_DEBUG_DA', '0') not in ('0', '', None)
    pairs = {}
    pretty_re = re.compile(r"^.+ - (?P<rut>\d{7,8}(?:-[0-9Kk])?) - .+ (?P<range>\d{4}-\d{4}) \[(?P<lang>ES|EN)\]\.xlsx$")
    for p in out_dir.glob("*.xlsx"):
        m = INPUT_PATTERN.match(p.name)
        if m:
            key = (m.group("rut"), m.group("range"))
            lst = pairs.setdefault(key, {})
            lst[m.group("lang")] = p
            if debug:
                print(f"[DA-XFILL] Found OUTPUT (input-style): {p.name}")
            continue
        m2 = pretty_re.match(p.name)
        if m2:
            key = (m2.group("rut"), m2.group("range"))
            lst = pairs.setdefault(key, {})
            lang = 'es' if m2.group('lang').upper() == 'ES' else 'en'
            lst[lang] = p
            if debug:
                print(f"[DA-XFILL] Found OUTPUT (pretty): {p.name}")
            continue
    if debug:
        print(f"[DA-XFILL] Pair count: {len(pairs)}")
    for (rut, yr), mp in pairs.items():
        if debug:
            print(f"[DA-XFILL] Pair: RUT={rut} RANGE={yr} files={{'es': mp.get('es'),'en': mp.get('en')}}")
        if len(mp) < 2:
            continue
        es_p = mp.get('es')
        en_p = mp.get('en')
        if not es_p or not en_p:
            continue
        wb_es = load_workbook(str(es_p))
        wb_en = load_workbook(str(en_p))
        try:
            for wb in (wb_es, wb_en):
                if "RATIOS & KPIs" not in wb.sheetnames:
                    continue
            sh_es = wb_es["RATIOS & KPIs"]
            sh_en = wb_en["RATIOS & KPIs"]
            def find_header(ws):
                for r in range(1, 8):
                    v = ws.cell(row=r, column=1).value
                    if isinstance(v, str) and v.strip().lower() in ("indicador", "indicator"):
                        return r
                return 4
            hdr_es = find_header(sh_es)
            hdr_en = find_header(sh_en)
            # construir mapa de columnas de períodos
            def period_cols(ws, hdr):
                cols = []
                for c in range(2, ws.max_column+1):
                    cols.append((c, str(ws.cell(row=hdr, column=c).value)))
                return cols
            cols_es = period_cols(sh_es, hdr_es)
            cols_en = period_cols(sh_en, hdr_en)
            # localizar fila D&A en cada archivo
            def find_da_row(ws, hdr):
                for r in range(hdr+1, ws.max_row+1):
                    name = ws.cell(row=r, column=1).value
                    if not isinstance(name, str):
                        continue
                    s = name.strip().lower()
                    if s in ("depreciación y amortización", "depreciation and amortization"):
                        return r
                return None
            row_es = find_da_row(sh_es, hdr_es)
            row_en = find_da_row(sh_en, hdr_en)
            if debug:
                print(f"[DA-XFILL] Headers: hdr_es={hdr_es} hdr_en={hdr_en}  rows: da_es={row_es} da_en={row_en}")
            if not row_es and not row_en:
                continue
            # copiado cruzado
            def copy_missing(src_ws, dst_ws, src_hdr, dst_hdr):
                src_cols = period_cols(src_ws, src_hdr)
                dst_cols = period_cols(dst_ws, dst_hdr)
                src_row = find_da_row(src_ws, src_hdr)
                dst_row = find_da_row(dst_ws, dst_hdr)
                if not src_row or not dst_row:
                    return False
                # mapear por etiqueta exacta
                lbl_to_src = { lbl: c for c, lbl in src_cols }
                changed = False
                copied = 0
                missing = 0
                for c_dst, lbl in dst_cols:
                    if not isinstance(lbl, str):
                        continue
                    v_dst = dst_ws.cell(row=dst_row, column=c_dst).value
                    if v_dst not in (None, ""):
                        continue
                    missing += 1
                    c_src = lbl_to_src.get(lbl)
                    if c_src:
                        v_src = src_ws.cell(row=src_row, column=c_src).value
                        if isinstance(v_src, (int, float)):
                            dst_ws.cell(row=dst_row, column=c_dst).value = v_src
                            changed = True
                            copied += 1
                if debug:
                    src_nm = 'ES' if src_ws is sh_es else 'EN'
                    dst_nm = 'ES' if dst_ws is sh_es else 'EN'
                    print(f"[DA-XFILL] {src_nm}→{dst_nm}: missing={missing} copied={copied}")
                return changed
            ch1 = copy_missing(sh_es, sh_en, hdr_es, hdr_en)
            ch2 = copy_missing(sh_en, sh_es, hdr_en, hdr_es)
            if ch1:
                if debug:
                    print(f"[DA-XFILL] Saved EN with cross-filled D&A: {en_p.name}")
                wb_en.save(str(en_p))
            if ch2:
                if debug:
                    print(f"[DA-XFILL] Saved ES with cross-filled D&A: {es_p.name}")
                wb_es.save(str(es_p))
        finally:
            try:
                wb_en.close()
            except Exception:
                pass
            try:
                wb_es.close()
            except Exception:
                pass


INPUT_PATTERN = re.compile(r"^estados_(?P<rut>\d{7,8}(?:-[0-9kK])?)_(?P<range>\d{4}-\d{4})_(?P<lang>es|en)\.xlsx$")


def find_input_excels(input_dir: Path, only_subdir: str | None = None, langs: tuple[str, ...] = ("es", "en")) -> list[Path]:
    """Encuentra entradas 'estados_*.xlsx'.

    - Si only_subdir es 'Anual' o 'Trimestral', solo busca en esa subcarpeta.
    - Si no se indica, busca en el propio input_dir (no en ambas subcarpetas) para evitar duplicados.
    """
    files: list[Path] = []

    def _collect_from(d: Path) -> None:
        if not d.exists() or not d.is_dir():
            return
        for p in d.glob("*.xlsx"):
            m = INPUT_PATTERN.match(p.name)
            if not m:
                continue
            lang = (m.group("lang") or "").lower()
            if langs and lang not in langs:
                continue
            files.append(p)

    # Si se especifica subcarpeta, limitar búsqueda a esa
    if only_subdir in ("Anual", "Trimestral"):
        _collect_from(input_dir / only_subdir)
    else:
        # Buscar solo en el directorio dado (no en ambas subcarpetas por defecto)
        _collect_from(input_dir)

    return sorted(files)


def ensure_products_inputs_from_xbrl(input_dir: Path, frequency: str, langs: tuple[str, ...]) -> None:
    """Garantiza que existan entradas 'estados_*.xlsx' en Products/<frequency> para todas las empresas en data/XBRL/<frequency>.

    Si faltan, genera los consolidados invocando el pipeline de batch_xbrl_to_excel solo para esa empresa/idioma y copia a Products.
    """
    try:
        from batch_xbrl_to_excel import find_datasets, generate_consolidated_company
    except Exception as e:
        print(f"[ensure] No se pudo importar batch_xbrl_to_excel: {e}")
        return

    # Directorio base de XBRL según frecuencia
    base_xbrl = Path("data/XBRL") / frequency
    if not base_xbrl.exists():
        print(f"[ensure] Base XBRL no existe: {base_xbrl}")
        return

    datasets = find_datasets(base_xbrl)
    if not datasets:
        print(f"[ensure] Sin datasets bajo {base_xbrl}")
        return

    # Agrupar por empresa
    company_to_datasets: dict[Path, list] = {}
    for ds in datasets:
        company_to_datasets.setdefault(ds.company_dir, []).append(ds)

    # Para cada empresa, validar que existan los archivos esperados en Products/<frequency>
    prod_subdir = input_dir / frequency
    prod_subdir.mkdir(parents=True, exist_ok=True)

    for company_dir, cds in company_to_datasets.items():
        if not cds:
            continue
        cds_sorted = sorted(cds, key=lambda d: d.yyyymm)
        min_yyyymm = min(d.yyyymm for d in cds_sorted)
        max_yyyymm = max(d.yyyymm for d in cds_sorted)

        # Detectar RUT con DV desde nombre del directorio, con fallback al numérico de dataset
        try:
            rut_prefix = company_dir.name.split('_', 1)[0]
            rut_with_dv = rut_prefix if '-' in rut_prefix else normalize_rut_with_dv(cds_sorted[0].rut)
        except Exception:
            rut_with_dv = normalize_rut_with_dv(cds_sorted[0].rut)

        # Nombre esperado en Products/<frequency>
        year_range = f"{max_yyyymm[:4]}-{min_yyyymm[:4]}"

        missing_langs: list[str] = []
        for lang in langs:
            expected = prod_subdir / f"estados_{rut_with_dv}_{year_range}_{lang}.xlsx"
            if not expected.exists():
                missing_langs.append(lang)

        if not missing_langs:
            continue

        # Generar solo los idiomas faltantes y copiar a Products
        try:
            print(f"[ensure] Faltan en Products/{frequency}: {company_dir.name}  idiomas={missing_langs}  rango={year_range}")
            cmf_dir = Path(__file__).resolve().parent
            generate_consolidated_company(company_dir, cds_sorted, cmf_dir, missing_langs, input_dir)
            # Verificar post-condición
            for lang in missing_langs:
                expected = prod_subdir / f"estados_{rut_with_dv}_{year_range}_{lang}.xlsx"
                if expected.exists():
                    print(f"[ensure] ✔ Generado: {expected.name}")
                else:
                    print(f"[ensure] ⚠ No se generó: {expected.name}")
        except Exception as ex:
            # No bloquear el resto por errores de una empresa
            print(f"[ensure] Error consolidando {company_dir.name}: {ex}")
            continue


def ensure_products_inputs_from_xbrl_combined(input_dir: Path, langs: tuple[str, ...]) -> None:
    """
    Garantiza que existan entradas combinadas 'estados_*.xlsx' en Products/Total
    uniendo datasets desde XBRL/Total (anuales + trimestrales).
    """
    try:
        from batch_xbrl_to_excel import find_datasets, generate_combined_company_from_total
    except Exception as e:
        print(f"[ensure-combined] No se pudo importar batch_xbrl_to_excel: {e}")
        return

    base_total = Path("data/XBRL/Total")
    if not base_total.exists():
        print(f"[ensure-combined] Directorio XBRL/Total no existe: {base_total}")
        return

    datasets = find_datasets(base_total)
    if not datasets:
        print(f"[ensure-combined] No se encontraron datasets en XBRL/Total")
        return

    # Agrupar por empresa
    company_to_datasets: dict[Path, List[DatasetInfo]] = {}
    for ds in datasets:
        company_to_datasets.setdefault(ds.company_dir, []).append(ds)

    print(f"[ensure-combined] Generando {len(company_to_datasets)} empresa(s) desde XBRL/Total...")
    import time as _time
    _t_ec_all = _time.perf_counter()

    # Generar archivos combinados
    for company_dir, company_datasets in company_to_datasets.items():
        _t_ec = _time.perf_counter()
        try:
            generate_combined_company_from_total(company_dir, company_datasets, Path("."), langs, input_dir)
            print(f"[ensure-combined] ✅ {company_dir.name}  ⏱ {(_time.perf_counter()-_t_ec):.1f}s")
        except Exception as e:
            print(f"[ensure-combined] ❌ {company_dir.name}: {e}")

    print(f"[ensure-combined] Completado en {(_time.perf_counter()-_t_ec_all):.1f}s. Archivos en: {input_dir}")


def ensure_product_v1_outputs(inputs: list[Path], output_dir: Path, frequency: str, workers: int) -> None:
    """Asegura que cada entrada de Products genere su correspondiente salida en Product_v1/<frequency>.

    Si falta el archivo pretty final, re-ejecuta el procesamiento solo para ese input.
    """
    effective_out = output_dir / frequency
    effective_out.mkdir(parents=True, exist_ok=True)

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _needs_processing(p: Path) -> tuple[Path, Path, bool]:
        m = INPUT_PATTERN.match(p.name)
        if not m:
            return p, effective_out / p.name, True
        rut = m.group("rut")
        rng = m.group("range")
        lang = m.group("lang")
        company = extract_company_name_from_excel(p)

        # Alinear la expectativa del nombre final con lo que realmente genera process_one
        range_effective = rng
        if frequency in ("Trimestral", "Total"):
            try:
                from openpyxl import load_workbook as _lw
                import re as _re
                def _detect_header_row(ws) -> int:
                    for r in range(1, min(10, ws.max_row) + 1):
                        v0 = ws.cell(row=r, column=1).value
                        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
                            return r
                        for c in range(2, min(ws.max_column, 30) + 1):
                            v = ws.cell(row=r, column=c).value
                            if isinstance(v, str) and _re.match(r"^\d{4}(|Q[1-4]|-\d{2}(-\d{2})?)$", v.strip()):
                                return r
                    return 3
                def _normalize_label(raw: str):
                    if not isinstance(raw, str):
                        return None
                    s = raw.strip().split("\n", 1)[0]
                    if _re.match(r"^\d{4}Q[1-4]$", s):
                        return s
                    m2 = _re.match(r"^(\d{4})-(\d{2})(?:-\d{2})?$", s)
                    if m2:
                        y, mo = int(m2.group(1)), int(m2.group(2))
                        q = {3:'Q1',6:'Q2',9:'Q3',12:'Q4'}.get(mo)
                        return f"{y}{q}" if q else str(y)
                    # Bare YYYY -> YYYYQ4 for backward compat
                    if _re.match(r"^\d{4}$", s):
                        return f"{s}Q4"
                    return None
                def _period_sort_key(lbl: str) -> tuple[int, int]:
                    s = str(lbl)
                    if _re.match(r"^\d{4}Q[1-4]$", s):
                        return (int(s[:4]), int(s[5]))
                    if _re.match(r"^\d{4}$", s):
                        return (int(s), 4)
                    try:
                        return (int(s[:4]), 5)
                    except Exception:
                        return (9999, 9)
                wb_probe = _lw(str(p), read_only=True, data_only=True)
                # Buscar una hoja conocida
                sheet_order = [
                    "Estado de Resultados", "Income Statement",
                    "Balance General", "Balance Sheet",
                    "Flujo Efectivo", "Cash Flow",
                ]
                sh_probe = None
                for nm in sheet_order:
                    if nm in wb_probe.sheetnames:
                        sh_probe = wb_probe[nm]
                        break
                if sh_probe is not None:
                    hdr = _detect_header_row(sh_probe)
                    labels = []
                    for c in range(2, sh_probe.max_column + 1):
                        v = sh_probe.cell(row=hdr, column=c).value
                        lb = _normalize_label(v) if isinstance(v, str) else None
                        if lb:
                            labels.append(lb)
                    if labels:
                        labels_sorted = sorted(set(labels), key=_period_sort_key, reverse=True)
                        if frequency == "Trimestral":
                            last_q = next((lb for lb in labels_sorted if _re.match(r"^\d{4}Q[1-4]$", lb)), None)
                            if last_q:
                                try:
                                    a, b = rng.split("-")
                                    yr_min = min(int(a), int(b))
                                    range_effective = f"{yr_min}-{last_q}"
                                except Exception:
                                    range_effective = f"{rng}-{last_q}"
                        else:  # Total
                            last_q = next((lb for lb in labels_sorted if _re.match(r"^\d{4}Q[1-4]$", lb)), None)
                            last_year = next((lb for lb in labels_sorted if _re.match(r"^\d{4}(Q4)?$", lb)), None)
                            last_label = last_q or last_year
                            if last_label:
                                try:
                                    a, b = rng.split("-")
                                    yr_min = min(int(a), int(b))
                                    range_effective = f"{yr_min}-{last_label}"
                                except Exception:
                                    range_effective = f"{rng}-{last_label}"
            except Exception:
                pass

        pretty = make_pretty_name(company, rut, range_effective, lang)
        dest = effective_out / pretty
        return p, dest, not dest.exists()

    pending: list[tuple[Path, Path]] = []
    for p in inputs:
        src, dest, missing = _needs_processing(p)
        if missing:
            print(f"[ensure_v1] Falta salida final para {src.name} → {dest.name}; re-ejecutando…")
            pending.append((src, dest))

    if not pending:
        return

    def _proc_one(p_src: Path) -> tuple[Path, str | None]:
        try:
            return process_one(p_src, effective_out, max(1, workers), frequency)
        except Exception as ex:
            return p_src, str(ex)

    with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
        futs = [ex.submit(_proc_one, p) for (p, _) in pending]
        for fut in as_completed(futs):
            out_path, err = fut.result()
            if err:
                print(f"[ensure_v1] ❌ Error con {out_path.name}: {err}")
            else:
                print(f"[ensure_v1] ✔ Generado {out_path.name}")

def extract_company_name_from_excel(xlsx_path: Path) -> Optional[str]:
    """Lee A1 de una hoja conocida para extraer el nombre de la empresa tras '—'."""
    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        sheet_order = [
            "Balance General", "Balance Sheet",
            "Estado de Resultados", "Income Statement",
            "Flujo Efectivo", "Cash Flow",
        ]
        sh = None
        for n in sheet_order:
            if n in wb.sheetnames:
                sh = wb[n]
                break
        if sh is None:
            sh = wb[wb.sheetnames[0]]
        a1 = sh.cell(row=1, column=1).value
        if isinstance(a1, str):
            # Título esperado: "<NombreHoja> — <Empresa>"
            if "—" in a1:
                return a1.split("—", 1)[1].strip()
            if "-" in a1:
                return a1.split("-", 1)[-1].strip()
        return None
    except Exception:
        return None


def _compute_rut_dv(rut_numeric: str) -> str:
    s = 0
    multiplier = 2
    for ch in reversed(rut_numeric):
        s += int(ch) * multiplier
        multiplier += 1
        if multiplier > 7:
            multiplier = 2
    dv_val = 11 - (s % 11)
    if dv_val == 11:
        return '0'
    if dv_val == 10:
        return 'K'
    return str(dv_val)


def normalize_rut_with_dv(rut: str) -> str:
    # Accept formats: 12345678 or 12345678-9
    parts = rut.split('-')
    num = parts[0]
    if len(parts) == 2 and parts[1]:
        dv = parts[1].upper()
        return f"{num}-{dv}"
    # Compute DV
    dv = _compute_rut_dv(num)
    return f"{num}-{dv}"


def make_pretty_name(company: Optional[str], rut: str, year_range: str, lang: str) -> str:
    # Rango en orden estético (menor - mayor)
    try:
        a, b = year_range.split("-")
        yr_min, yr_max = sorted([int(a), int(b)])
        range_pretty = f"{yr_min}-{yr_max}"
    except Exception:
        range_pretty = year_range
    company = (company or "Empresa").strip()
    rut = normalize_rut_with_dv(rut)
    lang_tag = "ES" if lang == "es" else "EN"
    # Título segmentado por idioma
    title = "Análisis Financiero" if lang == "es" else "Financial Analysis"
    return f"{company} - {rut} - {title} {range_pretty} [{lang_tag}].xlsx"


def process_one(input_file: Path, output_dir: Path, workers: int, frequency: str = "Anual",
                dash: Any | None = None) -> Tuple[Path, Optional[str]]:
    # Ejecutar análisis con fórmulas sobre un archivo
    # Permitir paralelismo interno en fórmulas (IO bound por openpyxl); configurable por env
    try:
        internal_workers = int(os.getenv('CMF_FORMULA_WORKERS', '2'))
    except Exception:
        internal_workers = 2
    import time as _time
    _t0_total = _time.perf_counter()
    
    # Mantener referencia del archivo original (para nombre bonito y dashboard)
    original_input_file = input_file
    # Antes de generar fórmulas, asegurar que EN y ES están alineados en columnas/estilos
    # Usar el archivo original para buscar contrapartes
    try:
        m_in = INPUT_PATTERN.match(original_input_file.name)
        if m_in:
            lang_cur = (m_in.group("lang") or "").lower()
            if lang_cur in ("es", "en"):
                other_lang = "es" if lang_cur == "en" else "en"
                other_name = original_input_file.name.replace(f"_{lang_cur}.xlsx", f"_{other_lang}.xlsx")
                counterpart = original_input_file.with_name(other_name)
                if counterpart.exists():
                    _t_align = _time.perf_counter()
                    from analisis_excel.utils.reconcile_en_with_es import reconcile_en_with_es as _recon
                    if lang_cur == "en":
                        _recon(original_input_file, counterpart, blank_unmatched_en_rows=True)
                    else:
                        _recon(counterpart, original_input_file, blank_unmatched_en_rows=True)
                    print(f"[timing] {original_input_file.name} pre-align ⏱ {(_time.perf_counter()-_t_align):.1f}s")
    except Exception:
        pass
    proc = BulkProcessor(input_directory=str(input_file.parent), output_directory=str(output_dir), max_workers=max(1, internal_workers))
    # Dashboard progreso por archivo
    total_steps = 5
    step_idx = 0
    def _progress(stage: str, cur: int, tot: int):
        nonlocal step_idx
        step_idx = max(step_idx, cur)
        if dash is not None:
            try:
                dash.update(str(original_input_file), estado="En progreso", progreso=stage, current=step_idx, total=total_steps, worker=threading.get_ident(), rut_display=original_input_file.name)
            except Exception:
                pass
    if dash is not None:
        try:
            empresa_name = original_input_file.parent.name.split('_', 1)[1].replace('_', ' ').strip() if '_' in original_input_file.parent.name else original_input_file.parent.name
            dash.update(str(original_input_file), estado="Preparando", progreso="start", current=0, total=total_steps, worker=threading.get_ident(), rut_display=original_input_file.name, empresa_name=empresa_name)
        except Exception:
            pass
    _t_form = _time.perf_counter()
    result = proc.process_single_file(input_file, progress_cb=_progress)
    print(f"[timing] {input_file.name} formulas ⏱ {(_time.perf_counter()-_t_form):.1f}s")
    # Salida inicial
    output_name = result.get("output_file") or f"{result.get('company_name','Analisis')}_Analisis_Formulas.xlsx"
    out_path = output_dir / output_name

    # Construir nombre bonito usando el archivo original
    m = INPUT_PATTERN.match(original_input_file.name)
    rut = m.group("rut") if m else "RUT"
    range_str = m.group("range") if m else "RANGO"
    lang = m.group("lang") if m else "es"
    company = extract_company_name_from_excel(input_file)
    if not company:
        # Fallback: usar nombre humano a partir del path (igual que xbrl_to_excel)
        try:
            # Buscar carpeta empresa tipo "<RUT-DV>_<NOMBRE_EMPRESA>"
            d = input_file.parent
            raw = d.name
            name_part = raw.split('_', 1)[1] if '_' in raw else raw
            company = name_part.replace('_', ' ').strip()
        except Exception:
            company = "Empresa"
    # Si es trimestral, intentar obtener el último label trimestral (YYYYQn) para el final del rango
    def _detect_header_row(ws) -> int:
        for r in range(1, min(10, ws.max_row) + 1):
            v0 = ws.cell(row=r, column=1).value
            if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
                return r
            for c in range(2, min(ws.max_column, 30) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and re.match(r"^\d{4}(|Q[1-4]|-\d{2}(-\d{2})?)$", v.strip()):
                    return r
        return 3
    def _normalize_label(raw: str) -> Optional[str]:
        """Normaliza etiquetas de períodos a formato estándar.

        Convierte diferentes formatos de fecha:
        - "2024Q3" → "2024Q3" (ya está normalizado)
        - "2024-09" → "2024Q3" (mes 9 = Q3)
        - "2024" → "2024Q4" (año completo = Q4)
        """
        if not isinstance(raw, str):
            return None
        s = raw.strip().split("\n", 1)[0]
        if re.match(r"^\d{4}Q[1-4]$", s):
            return s
        m = re.match(r"^(\d{4})-(\d{2})(?:-\d{2})?$", s)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            q = {3:'Q1',6:'Q2',9:'Q3',12:'Q4'}.get(mo)
            return f"{y}{q}" if q else str(y)
        # Bare YYYY -> YYYYQ4 for backward compat
        if re.match(r"^\d{4}$", s):
            return f"{s}Q4"
        return None
    def _period_sort_key(lbl: str) -> tuple[int, int]:
        s = str(lbl)
        if re.match(r"^\d{4}Q[1-4]$", s):
            return (int(s[:4]), int(s[5]))
        if re.match(r"^\d{4}$", s):
            return (int(s), 4)
        try:
            return (int(s[:4]), 5)
        except Exception:
            return (9999, 9)
    range_str_effective = range_str
    if frequency == "Trimestral":
        try:
            # Preferir el workbook de salida con hojas y etiquetas ya normalizadas
            wb_probe_path = out_path if out_path.exists() else input_file
            wb_probe = load_workbook(str(wb_probe_path), read_only=True, data_only=True)
            sheet_order = [
                "Estado de Resultados", "Income Statement",
                "Balance General", "Balance Sheet",
                "Flujo Efectivo", "Cash Flow",
            ]
            sh_probe = None
            for nm in sheet_order:
                if nm in wb_probe.sheetnames:
                    sh_probe = wb_probe[nm]
                    break
            if sh_probe is not None:
                hdr = _detect_header_row(sh_probe)
                labels = []
                for c in range(2, sh_probe.max_column + 1):
                    v = sh_probe.cell(row=hdr, column=c).value
                    lb = _normalize_label(v) if isinstance(v, str) else None
                    if lb:
                        labels.append(lb)
                if labels:
                    labels_sorted = sorted(set(labels), key=_period_sort_key, reverse=True)
                    # All labels are now YYYYQ[1-4] (bare YYYY normalized to YYYYQ4)
                    last_q = next((lb for lb in labels_sorted if re.match(r"^\d{4}Q[1-4]$", lb)), None)
                    last_year = next((lb for lb in labels_sorted if re.match(r"^\d{4}(Q4)?$", lb)), None)
                    last_label = last_q or last_year

                    if last_label:
                        try:
                            a, b = range_str.split("-")
                            yr_min = min(int(a), int(b))
                            range_str_effective = f"{yr_min}-{last_label}"
                        except Exception:
                            range_str_effective = f"{range_str}-{last_label}"
        except Exception:
            pass
    elif frequency == "Total":
        try:
            wb_probe_path = out_path if out_path.exists() else input_file
            wb_probe = load_workbook(str(wb_probe_path), read_only=True, data_only=True)
            sheet_order = [
                "Estado de Resultados", "Income Statement",
                "Balance General", "Balance Sheet",
                "Flujo Efectivo", "Cash Flow",
            ]
            sh_probe = None
            for nm in sheet_order:
                if nm in wb_probe.sheetnames:
                    sh_probe = wb_probe[nm]
                    break
            if sh_probe is not None:
                hdr = _detect_header_row(sh_probe)
                labels = []
                for c in range(2, sh_probe.max_column + 1):
                    v = sh_probe.cell(row=hdr, column=c).value
                    lb = _normalize_label(v) if isinstance(v, str) else None
                    if lb:
                        labels.append(lb)
                if labels:
                    labels_sorted = sorted(set(labels), key=_period_sort_key, reverse=True)
                    # All labels are now YYYYQ[1-4] (bare YYYY normalized to YYYYQ4)
                    last_q = next((lb for lb in labels_sorted if re.match(r"^\d{4}Q[1-4]$", lb)), None)
                    last_year = next((lb for lb in labels_sorted if re.match(r"^\d{4}(Q4)?$", lb)), None)
                    last_label = last_q or last_year

                    if last_label:
                        try:
                            a, b = range_str.split("-")
                            yr_min = min(int(a), int(b))
                            range_str_effective = f"{yr_min}-{last_label}"
                        except Exception:
                            range_str_effective = f"{range_str}-{last_label}"
        except Exception:
            pass
    pretty_name = make_pretty_name(company, rut, range_str_effective, lang)
    final_path = output_dir / pretty_name

    try:
        if out_path.exists() and out_path.is_file():
            if final_path.exists():
                final_path.unlink()
            _t_save = _time.perf_counter()
            os.replace(str(out_path), str(final_path))
            # Eliminar locks residuales si los hubiera
            try:
                for lockpat in (f"~${final_path.stem}.xlsx", f".~lock.{final_path.name}#"):
                    lp = final_path.parent / lockpat
                    if lp.exists():
                        lp.unlink()
            except Exception:
                pass
            print(f"[timing] {input_file.name} finalize ⏱ {(_time.perf_counter()-_t_save):.1f}s")
            # Add professional start sheet (merged from Phase 4 polish)
            try:
                from add_start_sheet_v4 import process_excel_file
                process_excel_file(str(final_path))
            except Exception as e_polish:
                print(f"[warning] Could not add start sheet to {final_path.name}: {e_polish}")
            # Limpieza: si existe archivo anterior sin sufijo Qn (para trimestral), eliminarlo
            if frequency in ("Trimestral", "Total") and range_str_effective != range_str:
                old_pretty = make_pretty_name(company, rut, range_str, lang)
                old_path = output_dir / old_pretty
                try:
                    if old_path.exists() and old_path.is_file() and old_path != final_path:
                        old_path.unlink()
                except Exception:
                    pass
        if dash is not None:
            try:
                dash.update(str(input_file), estado="Completado", progreso="listo", current=total_steps, total=total_steps, rut_display=input_file.name)
            except Exception:
                pass
        print(f"[timing] {input_file.name} total ⏱ {(_time.perf_counter()-_t0_total):.1f}s")
        return final_path, None
    except Exception as e:
        if dash is not None:
            try:
                dash.update(str(input_file), estado=f"Error: {str(e)[:40]}", rut_display=input_file.name)
            except Exception:
                pass
        return out_path, str(e)


def main() -> int:
    parser = argparse.ArgumentParser(description="Analiza Excels en Products/ y deja salidas en Product_v1/")
    parser.add_argument("--input-dir", type=Path, default=Path("Products"))
    parser.add_argument("--output-dir", type=Path, default=Path("Product_v1"))
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--frequency", choices=["Anual", "Trimestral", "Total"], default="Total",
                        help="Subcarpeta destino dentro de output-dir")
    parser.add_argument("--langs", type=str, default="es,en",
                        help="Idiomas a procesar (coma): es,en | es | en")
    args = parser.parse_args()

    effective_out = args.output_dir / args.frequency
    effective_out.mkdir(parents=True, exist_ok=True)

    # Asegurar que los directorios de entrada existan (base y subcarpeta de frecuencia)
    try:
        args.input_dir.mkdir(parents=True, exist_ok=True)
        if args.frequency == "Total":
            input_search_dir = args.input_dir / "Total"
        else:
            input_search_dir = args.input_dir
        input_search_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"Error creando directorios de entrada: {e}")
        return 1
    
    # Idiomas solicitados (ej. "es,en" o "es")
    try:
        langs_tuple = tuple(s.strip().lower() for s in (args.langs.split(',') if isinstance(args.langs, str) else ["es","en"]))
    except Exception:
        langs_tuple = ("es", "en")
    
    # Antes de listar entradas, asegurar que Products/<frequency> esté completo
    try:
        if args.frequency == "Total":
            ensure_products_inputs_from_xbrl_combined(args.input_dir, langs_tuple)
            # Configurar variables de entorno para modo combinado
            os.environ['CMF_ANALYSIS_COMBINED'] = '1'
            # Configurar número de trimestres para TTM (por defecto 2)
            if 'CMF_COMBINED_TTM_LAST_N' not in os.environ:
                os.environ['CMF_COMBINED_TTM_LAST_N'] = '2'
        else:
            ensure_products_inputs_from_xbrl(args.input_dir, args.frequency, langs_tuple)
    except Exception:
        pass

    # Filtrar entradas según la frecuencia elegida y los idiomas
    if args.frequency == "Total":
        # Buscar dentro de Products/Total
        inputs = find_input_excels(input_search_dir, only_subdir=None, langs=langs_tuple)
    else:
        # Buscar en Products/Anual o Products/Trimestral
        inputs = find_input_excels(input_search_dir, only_subdir=args.frequency, langs=langs_tuple)
    if not inputs:
        print(f"No se encontraron archivos tipo 'estados_...xlsx' en {input_search_dir}")
        return 1

    print(f"Procesando {len(inputs)} archivo(s) desde {args.input_dir} → {effective_out}")

    # Preparar dashboard (global si ya existe)
    dash = get_global_dashboard()
    created_local_dash = False
    if os.getenv('CMF_DASH_ENABLED', '0') == '1' and dash is None and ConsoleXBRLDashboard is not None and inputs:
        rows = []
        for p in inputs:
            rows.append({'key': str(p), 'rut': p.name, 'razon_social': p.parent.name})
        try:
            (args.output_dir / 'debug').mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        dash = ConsoleXBRLDashboard(rows, mute_stdout_logs=True, log_to_file=True,
                                    log_file_path=str((Path('data') / 'debug' / 'products_run.log').resolve()))
        dash.start()
        created_local_dash = True

    errors = []
    import time as _time
    _t_all = _time.perf_counter()
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futs = [ex.submit(process_one, p, effective_out, 1, args.frequency, dash) for p in inputs]
        for fut in concurrent.futures.as_completed(futs):
            out_path, err = fut.result()
            if err:
                print(f"❌ Error con {out_path.name}: {err}")
                errors.append((out_path, err))
            else:
                print(f"✔ {out_path.name}")

    print(f"Listo. Éxitos: {len(inputs) - len(errors)}  Errores: {len(errors)}  ⏱ Total: {(_time.perf_counter()-_t_all):.1f}s")

    # Asegurar que todos los inputs hayan producido el pretty final en Product_v1/<frequency>
    try:
        ensure_product_v1_outputs(inputs, args.output_dir, args.frequency, args.workers)
    except Exception as ex:
        print(f"[ensure_v1] Aviso: {ex}")

    # Post-verificación: completar D&A en Product_v1 si falta en un idioma usando el otro
    try:
        cross_fill_da(effective_out)
    except Exception as e:
        print(f"Aviso: no se pudo ejecutar verificación cruzada D&A: {e}")
    finally:
        if created_local_dash and dash is not None:
            try:
                dash.stop()
            except Exception:
                pass
    return 0 if not errors else 2


if __name__ == "__main__":
    raise SystemExit(main())

