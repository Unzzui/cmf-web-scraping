#!/usr/bin/env python3
"""
Script mejorado para poblar TODOS los datos históricos de WATTS SA
Usa un enfoque más agresivo de matching y completa toda la información disponible
"""

import pandas as pd
import openpyxl
import numpy as np
import re
from pathlib import Path
import json

def clean_numeric_value(value, from_facts=False):
    """Limpia y convierte valores numéricos"""
    if pd.isna(value) or value is None or value == 'None':
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ['none', 'nan', 'false', 'true']:
        return None
    
    # Remover formato de miles si es del facts CSV
    if from_facts:
        value_str = value_str.replace(',', '')
    
    try:
        return float(value_str)
    except:
        return None

def normalize_account_name(name):
    """Normaliza nombres de cuentas para matching"""
    if pd.isna(name) or not name:
        return ""
    
    name = str(name).strip()
    
    # Remover [sinopsis] y otros tags
    name = re.sub(r'\[.*?\]', '', name)
    
    # Normalizar espacios y puntuación
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'[,;:]', '', name)
    
    # Remover palabras comunes para mejor matching
    name = name.replace(' totales', '')
    name = name.replace(' total', '')
    name = name.replace(' corriente', '')
    name = name.replace(' corrientes', '')
    name = name.replace(' no corriente', '')
    name = name.replace(' no corrientes', '')
    
    return name.strip().lower()

def extract_all_excel_data(excel_path):
    """Extrae TODOS los datos del Excel, organizados por cuenta"""
    print(f"Extrayendo datos completos del Excel: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    account_data = {}  # {account_name: {date: value}}
    
    for sheet_name in ['Balance General', 'Estado de Resultados', 'Flujo Efectivo']:
        if sheet_name not in wb.sheetnames:
            continue
            
        print(f"\nProcesando hoja: {sheet_name}")
        ws = wb[sheet_name]
        
        # Encontrar fila de fechas
        date_row = 3
        date_columns = {}
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value:
                date_columns[col] = str(cell_value)
        
        print(f"  Fechas encontradas: {list(date_columns.values())}")
        
        # Extraer todas las cuentas y sus valores
        for row in range(4, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            
            if not account_name or str(account_name).startswith('='):
                continue
            
            account_key = f"{sheet_name}|{account_name}"
            
            if account_key not in account_data:
                account_data[account_key] = {
                    'sheet': sheet_name,
                    'original_name': account_name,
                    'normalized': normalize_account_name(account_name),
                    'values': {}
                }
            
            # Extraer todos los valores
            for col, date in date_columns.items():
                value = ws.cell(row=row, column=col).value
                clean_value = clean_numeric_value(value)
                
                if clean_value is not None:
                    account_data[account_key]['values'][date] = clean_value
    
    wb.close()
    
    print(f"\nTotal de cuentas extraídas: {len(account_data)}")
    
    # Contar valores por período
    period_counts = {}
    for account in account_data.values():
        for date in account['values']:
            period_counts[date] = period_counts.get(date, 0) + 1
    
    print("\nValores por período en Excel:")
    for date in sorted(period_counts.keys()):
        print(f"  {date}: {period_counts[date]} valores")
    
    return account_data

def load_and_analyze_facts(facts_path):
    """Carga y analiza el facts CSV"""
    print(f"\nAnalizando facts CSV: {facts_path}")
    facts_df = pd.read_csv(facts_path)
    
    date_columns = [col for col in facts_df.columns 
                   if col not in ['LabelKeyId', 'LabelKeyIdExt', 'SectionKey', 'Label', 'RoleCode']]
    
    print(f"  Columnas de fecha: {date_columns}")
    
    # Analizar contenido actual
    for date_col in date_columns:
        non_empty = facts_df[date_col].notna().sum()
        print(f"  {date_col}: {non_empty} valores no vacíos")
    
    return facts_df, date_columns

def aggressive_matching(excel_data, facts_df):
    """Matching agresivo basado en múltiples estrategias"""
    print("\n=== MATCHING AGRESIVO DE CUENTAS ===")
    
    matches = {}
    
    # Estrategia 1: Matching exacto por valor en períodos conocidos
    print("\nEstrategia 1: Matching por valores exactos...")
    
    date_mapping = {
        '2025Q2': '2025-06-30',
        '2025Q1': '2025-03-31',
        '2024': '2024-12-31',
        '2024Q3': '2024-09-30',
        '2024Q2': '2024-06-30',
        '2024Q1': '2024-03-31',
        '2023': '2023-12-31',
        '2023Q3': '2023-09-30',
        '2023Q2': '2023-06-30',
        '2023Q1': '2023-03-31',
        '2022': '2022-12-31',
        '2021': '2021-12-31'
    }
    
    for excel_key, excel_account in excel_data.items():
        excel_norm = excel_account['normalized']
        
        for idx, facts_row in facts_df.iterrows():
            facts_label = facts_row['Label']
            facts_norm = normalize_account_name(facts_label)
            
            # Verificar similitud de nombres
            score = 0
            if excel_norm and facts_norm:
                # Coincidencia exacta
                if excel_norm == facts_norm:
                    score = 100
                # Una contenida en otra
                elif excel_norm in facts_norm or facts_norm in excel_norm:
                    score = 80
                # Palabras clave comunes
                else:
                    excel_words = set(excel_norm.split())
                    facts_words = set(facts_norm.split())
                    common = excel_words & facts_words
                    if len(common) >= 2:
                        score = 60 * len(common) / max(len(excel_words), len(facts_words))
            
            if score >= 60:  # Umbral de similitud
                # Verificar coincidencia de valor
                for excel_date, excel_value in excel_account['values'].items():
                    if excel_date in date_mapping:
                        facts_date = date_mapping[excel_date]
                        facts_value = clean_numeric_value(facts_row.get(facts_date, None), from_facts=True)
                        
                        if facts_value is not None and excel_value is not None:
                            # Los valores del Excel están en miles
                            excel_value_units = excel_value * 1000
                            
                            # Verificar coincidencia con tolerancia
                            tolerance = max(abs(facts_value * 0.01), 1000)
                            if abs(facts_value - excel_value_units) < tolerance:
                                if idx not in matches or matches[idx]['score'] < score + 50:
                                    matches[idx] = {
                                        'excel_key': excel_key,
                                        'excel_account': excel_account,
                                        'facts_label': facts_label,
                                        'score': score + 50,  # Bonus por valor coincidente
                                        'matched_on': f"{excel_date}={excel_value}"
                                    }
                                break
    
    print(f"  Encontradas {len(matches)} coincidencias")
    
    # Estrategia 2: Para cuentas no mapeadas, usar solo nombre si es muy similar
    print("\nEstrategia 2: Matching por nombre para cuentas restantes...")
    
    unmapped_excel = set(excel_data.keys())
    for match in matches.values():
        if match['excel_key'] in unmapped_excel:
            unmapped_excel.remove(match['excel_key'])
    
    additional_matches = 0
    for excel_key in unmapped_excel:
        excel_account = excel_data[excel_key]
        excel_norm = excel_account['normalized']
        
        if not excel_norm:
            continue
        
        best_match = None
        best_score = 0
        
        for idx, facts_row in facts_df.iterrows():
            if idx in matches:  # Ya mapeado
                continue
                
            facts_label = facts_row['Label']
            facts_norm = normalize_account_name(facts_label)
            
            if excel_norm == facts_norm:
                best_match = idx
                best_score = 100
                break
            elif len(excel_norm) > 10 and len(facts_norm) > 10:
                # Para nombres largos, verificar similitud parcial
                if excel_norm[:15] == facts_norm[:15]:
                    if best_score < 70:
                        best_match = idx
                        best_score = 70
        
        if best_match is not None and best_score >= 70:
            matches[best_match] = {
                'excel_key': excel_key,
                'excel_account': excel_account,
                'facts_label': facts_df.iloc[best_match]['Label'],
                'score': best_score,
                'matched_on': 'name_only'
            }
            additional_matches += 1
    
    print(f"  Encontradas {additional_matches} coincidencias adicionales por nombre")
    print(f"\nTotal de coincidencias: {len(matches)}")
    
    return matches

def populate_all_data(facts_df, matches, date_columns):
    """Pobla TODOS los datos históricos disponibles"""
    print("\n=== POBLANDO TODOS LOS DATOS ===")
    
    date_mapping = {
        '2025Q2': '2025-06-30',
        '2025Q1': '2025-03-31',
        '2024': '2024-12-31',
        '2024Q3': '2024-09-30',
        '2024Q2': '2024-06-30',
        '2024Q1': '2024-03-31',
        '2023': '2023-12-31',
        '2023Q3': '2023-09-30',
        '2023Q2': '2023-06-30',
        '2023Q1': '2023-03-31',
        '2022': '2022-12-31',
        '2021': '2021-12-31'
    }
    
    populated_count = 0
    updated_count = 0
    
    for facts_idx, match_info in matches.items():
        excel_account = match_info['excel_account']
        
        for excel_date, excel_value in excel_account['values'].items():
            if excel_date not in date_mapping:
                continue
                
            facts_date = date_mapping[excel_date]
            if facts_date not in date_columns:
                continue
            
            current_value = facts_df.at[facts_idx, facts_date]
            
            # Convertir valor de miles a unidades
            value_units = int(excel_value * 1000)
            formatted_value = f"{value_units:,}"
            
            # Poblar si está vacío
            if pd.isna(current_value) or current_value == '' or current_value is None:
                facts_df.at[facts_idx, facts_date] = formatted_value
                populated_count += 1
            # Actualizar si es diferente (con tolerancia)
            else:
                current_clean = clean_numeric_value(current_value, from_facts=True)
                if current_clean is not None:
                    tolerance = max(abs(current_clean * 0.01), 1000)
                    if abs(current_clean - value_units) > tolerance:
                        facts_df.at[facts_idx, facts_date] = formatted_value
                        updated_count += 1
    
    print(f"  Valores nuevos poblados: {populated_count}")
    print(f"  Valores actualizados: {updated_count}")
    print(f"  Total modificaciones: {populated_count + updated_count}")
    
    return facts_df

def verify_final_state(facts_df, date_columns):
    """Verifica el estado final del facts"""
    print("\n=== ESTADO FINAL DEL FACTS ===")
    
    for date_col in date_columns:
        non_empty = facts_df[date_col].notna().sum()
        print(f"  {date_col}: {non_empty} valores no vacíos")
    
    # Verificar específicamente 2021 y 2022
    print("\nPeríodos históricos:")
    for year in ['2022-12-31', '2021-12-31']:
        if year in date_columns:
            count = facts_df[year].notna().sum()
            print(f"  {year}: {count} valores")

def main():
    # Rutas
    excel_path = '/home/unzzui/Documents/coding/CMF_extract/Products/Total/estados_76455830-8_2025-2021_es.xlsx'
    facts_path = '/home/unzzui/Documents/coding/CMF_extract/data/XBRL/Total/76455830-8_WATTS_SA/out_consolidated_76455830-8_202312-202506/facts_76455830-8_202312-202506_es.csv'
    
    # 1. Extraer TODOS los datos del Excel
    excel_data = extract_all_excel_data(excel_path)
    
    # 2. Cargar facts CSV
    facts_df, date_columns = load_and_analyze_facts(facts_path)
    
    # 3. Matching agresivo
    matches = aggressive_matching(excel_data, facts_df)
    
    # 4. Poblar TODOS los datos
    populated_df = populate_all_data(facts_df.copy(), matches, date_columns)
    
    # 5. Guardar resultados
    output_path = facts_path.replace('.csv', '_complete.csv')
    populated_df.to_csv(output_path, index=False)
    print(f"\n✓ Facts completo guardado en: {output_path}")
    
    # 6. Actualizar el original
    populated_df.to_csv(facts_path, index=False)
    print(f"✓ Facts original actualizado: {facts_path}")
    
    # 7. Verificar estado final
    verify_final_state(populated_df, date_columns)
    
    print("\n¡Proceso completado exitosamente!")
    print("Ahora puedes ejecutar xbrl_to_excel.py para generar el Excel final")

if __name__ == "__main__":
    main()