#!/usr/bin/env python3
"""
Versión v4: Combina el diseño superior de v2 con la robustez técnica de v3
"""

import os
import glob
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def extract_company_info(filename):
    """Extrae información de la empresa del nombre del archivo y formatea el RUT (76.129.263-3)."""
    try:
        name = os.path.basename(filename)
        # Regex: empresa - 76129263-3 - ... [ES].xlsx  (captura body y DV por separado)
        pattern = r'^(.+?)\s*-\s*([\d\.]+)-([0-9Kk])\s*-\s*.*?(\d{4}(?:-\d{4})?(?:Q\d)?)\s*\[ES\]'
        match = re.match(pattern, name)
        if match:
            company_name = match.group(1).strip()
            rut_body = match.group(2).replace('.', '')  # eliminar puntos si ya los tuviera
            dv = match.group(3).upper()
            # formatear con puntos de miles
            try:
                formatted_body = "{:,}".format(int(rut_body)).replace(",", ".")
            except ValueError:
                formatted_body = rut_body  # en caso de formatos inesperados
            rut = f"{formatted_body}-{dv}"
            period = match.group(4).strip()
            return company_name, rut, period

        # Fallback: dividir por ' - ' y limpiar
        base = name.replace('.xlsx', '').replace('[ES]', '').strip()
        parts = [p.strip() for p in base.split(' - ')]
        company_name = parts[0] if parts else "N/A"
        rut_raw = parts[1] if len(parts) > 1 else "N/A"

        # intentar separar cuerpo y DV si vienen juntos
        if rut_raw and '-' in rut_raw:
            body, dv = rut_raw.split('-', 1)
            dv = dv.upper()
            body = body.replace('.', '').strip()
            try:
                formatted_body = "{:,}".format(int(body)).replace(",", ".")
            except ValueError:
                formatted_body = body
            rut = f"{formatted_body}-{dv}"
        else:
            rut = rut_raw

        # intentar extraer periodo si no se obtuvo antes
        period_match = re.search(r'(\d{4}(?:-\d{4})?(?:Q\d)?)', name)
        period = period_match.group(1) if period_match else "N/A"

        return company_name, rut, period

    except Exception as e:
        print(f"Error extrayendo información: {e}")
        return filename.replace('.xlsx', ''), "N/A", "N/A"
def get_existing_sheets(workbook):
    """Mapea hojas existentes para navegación."""
    sheet_mapping = {
        'Balance General': ['Balance', 'Balance General', 'BS', 'Statement of Financial Position'],
        'Estado de Resultados': ['Estado de Resultados', 'PyG', 'P&L', 'Income Statement', 'IS'],
        'Flujo de Efectivo': ['Flujo de Efectivo', 'Cash Flow', 'CF', 'Flujo Efectivo'],
        'Ratios & KPIs': ['Ratios', 'KPIs', 'Ratios & KPIs', 'Indicators'],
        'Resumen Comparativo': ['Resumen', 'Summary', 'Comparativo', 'Resumen Comparativo']
    }
    
    available_sheets = []
    sheet_names = [sheet.title for sheet in workbook.worksheets]
    
    for display_name, possible_names in sheet_mapping.items():
        for sheet_name in sheet_names:
            if any(possible.lower() in sheet_name.lower() for possible in possible_names):
                available_sheets.append((display_name, sheet_name))
                break
    
    return available_sheets

def set_cell_with_merge(sheet, start_cell, end_cell, value, font=None, alignment=None, fill=None, border=None):
    """Función auxiliar para establecer valor y luego fusionar celdas de manera segura."""
    # Primero establecer el valor y formato en la celda inicial
    cell = sheet[start_cell]
    cell.value = value
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    
    # Luego fusionar si es necesario
    if start_cell != end_cell:
        sheet.merge_cells(f'{start_cell}:{end_cell}')

def create_professional_dashboard(workbook, company_name, rut, period, currency="", currency_changes=None):
    """
    Crea dashboard profesional con diseño superior y técnica robusta.
    """
    # Eliminar hojas existentes de manera robusta
    sheets_to_remove = []
    for sheet in workbook.worksheets:
        if ('Inicio' in sheet.title or 'Dashboard' in sheet.title or 
            sheet.title.startswith('📑') or sheet.title.startswith('🏠')):
            sheets_to_remove.append(sheet.title)
    
    for sheet_name in sheets_to_remove:
        try:
            del workbook[sheet_name]
        except KeyError:
            pass
    
    # Crear hoja nueva
    start_sheet = workbook.create_sheet("Inicio", 0)
    start_sheet.sheet_view.showGridLines = False
    
    # Configurar grid de 14 columnas con márgenes (A y N como márgenes)
    start_sheet.column_dimensions['A'].width = 2
    start_sheet.column_dimensions['N'].width = 2
    
    # Columnas B-M = contenido (8.5 unidades cada una)
    col_width = 8.5
    for col in range(2, 14):  # B hasta M
        start_sheet.column_dimensions[get_column_letter(col)].width = col_width
    
    # === ESTILOS PROFESIONALES ===
    header_primary = PatternFill("solid", fgColor="0F172A")
    header_secondary = PatternFill("solid", fgColor="1E293B") 
    accent_blue = PatternFill("solid", fgColor="3B82F6")
    accent_green = PatternFill("solid", fgColor="10B981")
    accent_purple = PatternFill("solid", fgColor="8B5CF6")
    accent_orange = PatternFill("solid", fgColor="F59E0B")
    
    card_bg = PatternFill("solid", fgColor="F8FAFC")
    card_header = PatternFill("solid", fgColor="E2E8F0")
    
    # Tipografías
    brand_font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
    title_font = Font(name='Segoe UI', size=14, bold=True, color='1E293B')
    subtitle_font = Font(name='Segoe UI', size=12, bold=True, color='475569')
    body_font = Font(name='Segoe UI', size=10, color='64748B')
    nav_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    
    # Alineaciones
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    # Bordes elegantes
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    row = 1
    
    # === HEADER CORPORATIVO ===
    set_cell_with_merge(start_sheet, 'B1', 'M2', 
                       "FinData Chile | Inteligencia Financiera Profesional",
                       brand_font, center, header_primary)
    
    row = 4
    
    # === HERO SECTION ===
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       f"ANÁLISIS FINANCIERO: {company_name.upper()}",
                       Font(name='Segoe UI', size=20, bold=True, color='1E293B'),
                       center, PatternFill("solid", fgColor="F1F5F9"))
    
    row += 2
    
    # === CARDS INFORMATIVOS ===
    # Card 1: Información Básica
    set_cell_with_merge(start_sheet, f'B{row}', f'E{row}',
                       "INFORMACIÓN CORPORATIVA",
                       subtitle_font, center, card_header, thin_border)
    
    # Card 2: Período y Fuente
    set_cell_with_merge(start_sheet, f'F{row}', f'I{row}',
                       "COBERTURA TEMPORAL",
                       subtitle_font, center, card_header, thin_border)
    
    # Card 3: Características
    set_cell_with_merge(start_sheet, f'J{row}', f'M{row}',
                       "ESPECIFICACIONES TÉCNICAS",
                       subtitle_font, center, card_header, thin_border)
    
    row += 1
    
    # Contenido de los cards
    # La moneda sale del XBRL, NO se asume.
    #
    # Antes esto decía "Moneda: CLP" fijo. Diecisiete empresas del catálogo reportan en
    # DÓLARES (SQM, COPEC, CMPC, LATAM, Colbún, Enel Américas, Engie, CAP, Vapores,
    # Blumar, Cintac, Iansa, Inversiones CMPC, Agrosuper, Enel Chile, Enel Generación y
    # Eléctrica Pehuenche). Para todas ellas, la portada del Excel que el cliente PAGA
    # afirmaba algo falso — y no es un detalle de formato: es un error de un factor de
    # 900 que el analista sólo descubre después de haber tomado una decisión.
    moneda_txt = f"Moneda: {currency}" if currency else "Moneda: (no determinada)"
    info_content = [
        f"Empresa: {company_name}",
        f"RUT: {rut}",
        "Mercado: Chile",
        moneda_txt,
    ]
    # Si la empresa CAMBIÓ de moneda a mitad de la serie, hay que decirlo: los años no
    # son comparables entre sí sin convertir. Enel Chile pasó de pesos a dólares en 2025.
    if currency_changes:
        info_content.append("Cambio de moneda: " + " · ".join(currency_changes))
    
    period_content = [
        f"Período: {period}",
        "Frecuencia: Trimestral",
        "Fuente: CMF Chile",
        "Estándar: IFRS"
    ]
    
    tech_content = [
        "Ratios: +30 indicadores",
        "Formato: Excel dinámico"
    ]
    
    for i, info in enumerate(info_content):
        cell_row = row + i
        set_cell_with_merge(start_sheet, f'B{cell_row}', f'E{cell_row}',
                           info, body_font, left, card_bg, thin_border)
    
    for i, info in enumerate(period_content):
        cell_row = row + i
        set_cell_with_merge(start_sheet, f'F{cell_row}', f'I{cell_row}',
                           info, body_font, left, card_bg, thin_border)
    
    for i, info in enumerate(tech_content):
        cell_row = row + i
        set_cell_with_merge(start_sheet, f'J{cell_row}', f'M{cell_row}',
                           info, body_font, left, card_bg, thin_border)
    
    row += 6
    
    # === NAVEGACIÓN PROFESIONAL ===
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "NAVEGACIÓN RÁPIDA - MÓDULOS DE ANÁLISIS",
                       Font(name='Segoe UI', size=14, bold=True, color='FFFFFF'),
                       center, header_secondary)
    
    row += 1
    
    # Obtener hojas disponibles
    available_sheets = get_existing_sheets(workbook)
    
    # Navegación en grid profesional 2x4
    nav_items = [
        ("Balance General", accent_blue),
        ("Estado de Resultados", accent_green),
        ("Flujo de Efectivo", accent_purple),
        ("Ratios & KPIs", accent_orange),
        ("Resumen Comparativo", accent_green)
    ]
    
    # Navegación
    for i, (sheet_label, color) in enumerate(nav_items):
        row_offset = i // 3
        col_idx = i % 3
        col_start = col_idx * 4 + 2  # Columnas B, F, J
        col_end = col_start + 3  # 4 columnas por botón

        nav_cell = start_sheet[f'{get_column_letter(col_start)}{row + row_offset}']
        nav_cell.value = sheet_label
        nav_cell.font = nav_font
        nav_cell.alignment = center
        nav_cell.fill = color
        nav_cell.border = thin_border

        for display_name, actual_name in available_sheets:
            if display_name == sheet_label:
                nav_cell.hyperlink = f"#'{actual_name}'!A1"
                break

        start_sheet.merge_cells(f'{get_column_letter(col_start)}{row + row_offset}:{get_column_letter(col_end)}{row + row_offset}')

    row += 2  # Two rows used for nav
    
    row += 3
    
    # === DESCRIPCIÓN EJECUTIVA ===
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "RESUMEN EJECUTIVO",
                       subtitle_font, center, card_header, thin_border)
    
    row += 1
    
    description = (
        "Este archivo contiene estados financieros oficiales bajo estándar IFRS reportados por la empresa "
        "a la CMF (Comisión para el Mercado Financiero de Chile). Los datos han sido procesados y "
        "complementados con análisis financiero avanzado que incluye más de 30 ratios financieros."
        "\n\nNOTAS TÉCNICAS IMPORTANTES:\n"
        "• Separadores decimales: Si las fórmulas no funcionan correctamente, configure Excel con separador "
        "decimal como punto (.) y separador de miles como coma (,). Ver instrucciones detalladas en contacto."
        "\n\nHerramienta profesional para inversiones institucionales, análisis de crédito e investigación académica."
    )
    
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row+5}',
                       description, body_font, wrap, card_bg, thin_border)
    
    row += 7
    
    # === FOOTER PROFESIONAL ===
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "Datos oficiales IFRS de CMF Chile. Solo fines educativos/profesionales. No constituye asesoría de inversión.",
                       Font(name='Segoe UI', size=9, color='6B7280'), center)
    
    row += 1
    
    # Sección de configuración técnica
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "CONFIGURACIÓN DE SEPARADORES DECIMALES",
                       Font(name='Segoe UI', size=11, bold=True, color='1E293B'),
                       center, PatternFill("solid", fgColor="FEF3C7"), thin_border)
    
    row += 1
    
    tech_instructions = (
        "Para que Excel reconozca correctamente los números con decimales separados por punto (ejemplo: 0.27):\n\n"
        "• En Windows: Archivo → Opciones → Avanzadas → Opciones de edición → Desmarcar 'Usar separadores del sistema' → "
        "Separador decimal: . (punto) → Separador de miles: , (coma)\n\n"
        "• En Mac: Excel → Preferencias → Avanzadas → Edición → Desmarcar 'Usar separadores del sistema' → "
        "Separador decimal: . (punto) → Separador de miles: , (coma)"
    )
    
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row+3}',
                       tech_instructions,
                       Font(name='Segoe UI', size=9, color='374151'),
                       wrap, PatternFill("solid", fgColor="FFFBEB"), thin_border)
    
    row += 5
    
    # Información de contacto y soporte
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "Web: www.findatachile.com | Soporte: contacto@findatachile.com | Resolveremos cualquier consulta o problema a la brevedad",
                       Font(name='Segoe UI', size=10, bold=True, color='3B82F6'), center)
    
    row += 2
    
    # Footer final corporativo
    set_cell_with_merge(start_sheet, f'B{row}', f'M{row}',
                       "FinData Chile - Transforming Financial Data into Strategic Intelligence",
                       Font(name='Segoe UI', size=12, bold=True, color='FFFFFF'),
                       center, header_primary)
    
    # Configurar alturas para mejor visualización
    start_sheet.row_dimensions[1].height = 40  # Header
    start_sheet.row_dimensions[4].height = 35  # Hero
    for r in range(5, row + 1):
        start_sheet.row_dimensions[r].height = 25
    
    return start_sheet

def detect_currency(rut_formateado):
    """Moneda de reporte de la empresa, leída de su XBRL.

    `rut_formateado` viene de la portada como "76.129.263-3". Las carpetas de XBRL usan
    "76129263-3", sin puntos.

    Si no se puede determinar, se devuelve vacío y la portada lo DICE. Preferimos "(no
    determinada)" a un "CLP" inventado: un dato ausente se nota; uno falso, no.
    """
    try:
        from currency_detect import resumen_moneda
    except ImportError:
        try:
            from cmf_extract.currency_detect import resumen_moneda
        except ImportError:
            return "", []

    rut = str(rut_formateado or "").replace(".", "").strip().upper()
    if not rut or "-" not in rut:
        return "", []

    raiz = Path(__file__).resolve().parent.parent / "data" / "XBRL" / "Total"
    if not raiz.is_dir():
        return "", []

    for carpeta in raiz.iterdir():
        if carpeta.is_dir() and carpeta.name.upper().startswith(rut + "_"):
            return resumen_moneda(carpeta)
    return "", []


def process_excel_file(file_path):
    """Procesa un archivo Excel individual."""
    try:
        print(f"Procesando: {os.path.basename(file_path)}")
        
        filename = os.path.basename(file_path)
        company_name, rut, period = extract_company_info(filename)

        currency, currency_changes = detect_currency(rut)
        if not currency:
            print(f"  ⚠️  No pude leer la moneda del XBRL de {rut}. La portada lo dirá en vez de asumir CLP.")

        workbook = load_workbook(file_path)

        # Crear nuevo dashboard profesional
        dashboard_sheet = create_professional_dashboard(
            workbook, company_name, rut, period, currency, currency_changes
        )
        
        workbook.save(file_path)
        print(f"  ✅ Hoja de inicio profesional creada exitosamente")
        return True
        
    except Exception as e:
        print(f"  ❌ Error procesando {file_path}: {e}")
        return False

def main():
    """CLI principal."""
    print("FinData Chile - Generador de Hoja de Inicio Profesional")
    print("=" * 60)
    print("1. Solo WATTS SA (prueba)")
    print("2. Todos los archivos [ES].xlsx")
    print("=" * 60)
    
    try:
        choice = input("Selecciona una opción (1 o 2): ").strip()
    except KeyboardInterrupt:
        print("\\n❌ Operación cancelada.")
        return
    
    if choice == "1":
        target_file = "/home/unzzui/Documents/coding/CMF_extract/Product_v1/Total/WATTS SA - 76455830-8 - Análisis Financiero 2023-2025Q1 [ES].xlsx"
        
        if not os.path.exists(target_file):
            print(f"❌ Error: El archivo WATTS SA no existe.")
            return
        
        print(f"\\n📁 Procesando archivo de prueba: WATTS SA")
        print("=" * 60)
        
        if process_excel_file(target_file):
            print("=" * 60)
            print(f"✅ Hoja de inicio creada exitosamente para WATTS SA")
        else:
            print("=" * 60)
            print(f"❌ Error procesando WATTS SA")
            
    elif choice == "2":
        target_dir = "/home/unzzui/Documents/coding/CMF_extract/Product_v1/Total"
        
        if not os.path.exists(target_dir):
            print(f"❌ Error: El directorio {target_dir} no existe.")
            return
        
        spanish_pattern = os.path.join(target_dir, "*[[]ES[]].xlsx")
        spanish_files = glob.glob(spanish_pattern)
        
        if not spanish_files:
            print(f"❌ No se encontraron archivos Excel en español.")
            return
        
        print(f"\\n📁 Directorio: {target_dir}")
        print(f"📊 Archivos encontrados: {len(spanish_files)}")
        
        try:
            confirm = input(f"\\n¿Procesar {len(spanish_files)} archivos? (s/n): ").strip().lower()
        except KeyboardInterrupt:
            print("\\n❌ Operación cancelada.")
            return
            
        if confirm in ['s', 'si', 'sí', 'y', 'yes']:
            print("=" * 60)
            
            success_count = 0
            for file_path in sorted(spanish_files):
                if process_excel_file(file_path):
                    success_count += 1
            
            print("=" * 60)
            print(f"✅ Procesamiento completado: {success_count}/{len(spanish_files)} archivos exitosos")
        else:
            print("❌ Operación cancelada.")
    else:
        print("❌ Opción inválida. Usa 1 o 2.")

if __name__ == "__main__":
    main()