#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesamiento masivo de XBRL → CSVs (Arelle) → Excel corporativo.

Recorre data/XBRL/(Anual|Trimestral)/**/<empresa_rut_y_nombre>/**/Estados_financieros_(XBRL)<RUT>_<yyyymm>_extracted
para encontrar instancias .xbrl, exporta factTable/presentation (ES/EN) con Arelle y
genera excels usando xbrl_to_excel.py.

Uso:
  python batch_xbrl_to_excel.py \
    --base-dir /ruta/a/data/XBRL \
    --arelle-dir /home/unzzui/Documents/Arelle \
    [--max] [--dry-run]

Procesa solo en español y escribe outputs en:
  <dataset_dir>/out_<RUT>_<yyyymm>/{facts_*.csv, presentation_*.csv, estados_*.xlsx}
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from dataclasses import dataclass
import time
from shutil import copy2
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple, Optional, Callable

# Import our perfect JSON-based ordering and data processing
from analisis_excel.utils.json_ordering import apply_perfect_json_ordering
from analisis_excel.utils.facts_processing import (
    apply_consolidated_processing, 
    apply_excel_processing_like_primary_csv
)


def _compute_rut_dv(rut_numeric: str) -> str:
    """Calcula el dígito verificador de un RUT chileno."""
    s = 0
    multiplier = 2
    for ch in reversed(rut_numeric):
        s += int(ch) * multiplier
        multiplier += 1
        if multiplier > 7:
            multiplier = 2
    dv_val = 11 - (s % 11)
    if dv_val == 11:
        return '0'
    if dv_val == 10:
        return 'K'
    return str(dv_val)


def normalize_rut_with_dv(rut: str) -> str:
    """Normaliza un RUT para incluir guión y dígito verificador."""
    # Accept formats: 12345678 or 12345678-9
    parts = rut.split('-')
    num = parts[0]
    if len(parts) == 2 and parts[1]:
        dv = parts[1].upper()
        return f"{num}-{dv}"
    # Compute DV
    dv = _compute_rut_dv(num)
    return f"{num}-{dv}"


DATASET_DIR_PATTERN = re.compile(r"Estados_financieros_\(XBRL\)(?P<rut>\d+)_(?P<yyyymm>\d{6})_extracted$")


@dataclass
class DatasetInfo:
    company_dir: Path
    dataset_dir: Path
    rut: str
    yyyyymm: str

    @property
    def stem(self) -> str:
        return f"{self.rut}_{self.yyyyymm}"
    
    @property
    def is_annual(self) -> bool:
        """Determina si este dataset es anual (Q4) basado en el mes."""
        try:
            month = int(self.yyyyymm[4:6])
            return month == 12  # Diciembre = anual
        except Exception:
            return False
    
    @property
    def year(self) -> int:
        """Extrae el año del yyyyymm."""
        try:
            return int(self.yyyyymm[:4])
        except Exception:
            return 0
    
    @property
    def month(self) -> int:
        """Extrae el mes del yyyyymm."""
        try:
            return int(self.yyyyymm[4:6])
        except Exception:
            return 0


def find_datasets(base_dir: Path) -> List[DatasetInfo]:
    """Busca datasets en toda la jerarquía bajo base_dir.

    Admite estructuras como:
      data/XBRL/Anual/<RUT-DV_EMPRESA>/Estados_financieros_(XBRL)<RUT>_<yyyymm>_extracted
      data/XBRL/Trimestral/<RUT-DV_EMPRESA>/Estados_financieros_(XBRL)<RUT>_<yyyymm>_extracted
      data/XBRL/Total/<RUT-DV_EMPRESA>/Estados_financieros_(XBRL)<RUT>_<yyyymm>_extracted
      data/XBRL/<RUT-DV_EMPRESA>/Estados_financieros_(XBRL)<RUT>_<yyyymm>_extracted
    """
    datasets: List[DatasetInfo] = []
    if not base_dir.exists():
        return datasets

    for dirpath, dirnames, filenames in os.walk(base_dir):
        current = Path(dirpath)
        m = DATASET_DIR_PATTERN.search(current.name)
        if not m:
            continue
        rut = m.group('rut')
        yyyymm = m.group('yyyymm')
        # El directorio de empresa es el padre inmediato del dataset
        company_dir = current.parent
        
        # Detectar si es dataset anual (Q4) o trimestral basado en el mes
        try:
            month = int(yyyymm[4:6])
            is_annual = (month == 12)  # Diciembre = anual
        except Exception:
            is_annual = False
            
        datasets.append(DatasetInfo(company_dir=company_dir, dataset_dir=current, rut=rut, yyyyymm=yyyymm))
    return datasets


def find_xbrl_file_typed(dataset_dir: Path, stem: str) -> Tuple[Path | None, str | None]:
    """Localiza el .xbrl del dataset y devuelve (ruta, tipo) con tipo en {'C','I'}.

    Prefiere el Consolidado (*_C.xbrl) y cae al Individual (*_I.xbrl) cuando el
    emisor no publica consolidado. Varios emisores (Bolsa de Comercio desde
    2023Q1, Metrogas desde 2024Q4, Chilquinta Distribución desde 2021Q3, …)
    dejaron de presentar estados consolidados; exigir *_C.xbrl hacía que esos
    períodos se descartaran en silencio y la serie quedara congelada.

    Nunca cae a un .xbrl arbitrario: si no hay ni C ni I, retorna (None, None).
    """
    for kind in ('C', 'I'):
        exact = dataset_dir / f"{stem}_{kind}.xbrl"
        if exact.exists():
            return exact, kind
    # El stem del directorio no siempre coincide con el del archivo: buscar por sufijo.
    for kind in ('C', 'I'):
        suffix = f"_{kind}.xbrl"
        for dirpath, dirnames, filenames in os.walk(dataset_dir):
            for name in sorted(filenames):
                if name.endswith(suffix):
                    return Path(dirpath) / name, kind
    return None, None


def find_xbrl_file(dataset_dir: Path, stem: str) -> Path | None:
    """Ruta del .xbrl a procesar (Consolidado, o Individual si no hay consolidado)."""
    path, _kind = find_xbrl_file_typed(dataset_dir, stem)
    return path


def dataset_statement_type(dataset_dir: Path, stem: str) -> str | None:
    """Tipo de estado financiero publicado en el dataset: 'C', 'I' o None."""
    _path, kind = find_xbrl_file_typed(dataset_dir, stem)
    return kind


def _arelle_timeout() -> int | None:
    # Timeout duro por export (Arelle puede colgarse bajando taxonomías de
    # la CMF si la red falla). Configurable vía CMF_ARELLE_TIMEOUT (segundos);
    # default 180 (3 min). 0 o negativo desactiva el timeout.
    try:
        timeout = int(os.getenv("CMF_ARELLE_TIMEOUT", "180"))
    except ValueError:
        timeout = 180
    return timeout if timeout > 0 else None


def run_cmd(cmd: Sequence[str], cwd: Path | None = None) -> None:
    subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        check=True,
        timeout=_arelle_timeout(),
    )


_worker_pool_broken = False


def run_arelle(arelle_python: Path, arelle_dir: Path, args: list[str]) -> None:
    """Ejecuta un export de Arelle.

    Con CMF_ARELLE_WORKER=1 usa el pool de workers persistentes (paga el
    arranque de intérprete + import de Arelle una vez por worker en lugar de
    una vez por dataset). Ante cualquier falla del pool cae al modo subprocess
    clásico y no vuelve a intentar el pool en esta corrida.
    """
    global _worker_pool_broken
    if os.getenv("CMF_ARELLE_WORKER", "0") == "1" and not _worker_pool_broken:
        try:
            from arelle_pool import ArelleWorkerPool
            pool = ArelleWorkerPool.get(arelle_python, arelle_dir)
            pool.run(args, timeout=float(_arelle_timeout() or 3600))
            return
        except Exception as exc:
            _worker_pool_broken = True
            print(f"[arelle] pool de workers falló ({exc}); "
                  f"usando subprocess clásico", file=sys.stderr)
    run_cmd([str(arelle_python), 'arelleCmdLine.py', *args], cwd=arelle_dir)


def run_arelle_exports(arelle_dir: Path, xbrl_file: Path, out_dir: Path, stem: str, langs: Sequence[str], facts_strategy: str = "es_only", force: bool = False) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # Absoluto: el subprocess corre con cwd=arelle_dir, donde una ruta
    # relativa al intérprete no resuelve.
    arelle_dir = Path(arelle_dir).resolve()
    # Elegir intérprete de Arelle (
    arelle_python = arelle_dir / '.venv' / 'bin' / 'python'
    if not arelle_python.exists():
        # Fallback a python del sistema en el dir de arelle
        arelle_python = Path(sys.executable)

    fact_cols = (
        "Label,localName,contextRef,unitRef,Dec,Prec,Lang,Value,"
        "entityIdentifier,periodStart,periodEnd,instant,endInstant,qname"
    )

    # Convertir rutas a absolutas ya que correremos con cwd=arelle_dir
    xbrl_abs = xbrl_file.resolve()
    out_abs = out_dir.resolve()

    def _up_to_date(inputs: list[Path], outputs: list[Path]) -> bool:
        try:
            latest_in = max(p.stat().st_mtime for p in inputs if p.exists())
            mt_outs = [p.stat().st_mtime for p in outputs if p.exists()]
            return len(mt_outs) == len(outputs) and all(m >= latest_in for m in mt_outs)
        except Exception:
            return False

    for lang in langs:
        label_lang = 'es-CL' if lang.startswith('es') else 'en'
        facts_csv = (out_abs / f"facts_{stem}_{lang}.csv")
        pres_csv = (out_abs / f"presentation_{stem}_{lang}.csv")
        facts_log = (out_abs / f"arelle_facts_{lang}.log")
        pre_log = (out_abs / f"arelle_pre_{lang}.log")

        # factTable (estrategia opcional para acelerar):
        #  - es_only: exportar facts solo para ES y reutilizar para EN
        #  - both: exportar facts para ambos idiomas
        want_facts = facts_strategy == "both" or lang.startswith('es')
        if not want_facts:
            # Create symlink for EN facts pointing to ES (saves ~22MB per dataset)
            try:
                es_facts = (out_abs / f"facts_{stem}_es.csv")
                if es_facts.exists() and not facts_csv.exists():
                    os.symlink(es_facts, facts_csv)
            except OSError:
                # Fallback to hardlink if symlinks not supported
                try:
                    if es_facts.exists() and not facts_csv.exists():
                        os.link(str(es_facts), str(facts_csv))
                except Exception:
                    pass

        need_facts = want_facts and (force or not _up_to_date([xbrl_abs], [facts_csv, facts_log]))
        need_pre = force or not _up_to_date([xbrl_abs], [pres_csv, pre_log])

        if need_facts and need_pre:
            # Una sola corrida de Arelle exporta ambos: la carga del DTS (lo
            # caro) se paga una vez en lugar de dos.
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--factTable', str(facts_csv.resolve()),
                '--factTableCols', fact_cols,
                '--pre', str(pres_csv.resolve()),
                '--logFile', str(facts_log.resolve()),
            ])
            pre_log.write_text("(exportado junto a facts; ver arelle_facts log)\n",
                               encoding="utf-8")
        elif need_facts:
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--factTable', str(facts_csv.resolve()),
                '--factTableCols', fact_cols,
                '--logFile', str(facts_log.resolve()),
            ])
        elif need_pre:
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--pre', str(pres_csv.resolve()),
                '--logFile', str(pre_log.resolve()),
            ])

        _verificar_facts_no_vacio(facts_csv, xbrl_file)


def _verificar_facts_no_vacio(facts_csv: Path, xbrl_file: Path) -> None:
    """Falla si Arelle exportó un CSV de facts con sólo la cabecera.

    Arelle escribe la cabecera y DESPUÉS los hechos, y termina con exit 0 aunque no
    haya podido resolver el DTS (taxonomía incompleta en el cache, .xsd del emisor
    ausente, etc.). El resultado es un CSV de 1 línea que el pipeline aceptaba como
    éxito: la empresa quedaba con huecos en el Balance de ese trimestre y nadie se
    enteraba. Medido el 2026-07-12: 231 de 10.174 exports (2,3%) estaban así,
    afectando a 58 empresas y dejando 74 Excel con columnas casi vacías.
    """
    if not facts_csv.exists():
        return
    try:
        with facts_csv.open('rb') as fh:
            lineas = sum(1 for _ in fh)
    except OSError:
        return
    if lineas <= 1:
        raise RuntimeError(
            f"Arelle exportó un facts CSV vacío (sólo cabecera) para "
            f"{xbrl_file.name}. Revisa el .log de Arelle: normalmente es la "
            f"taxonomía incompleta en ~/.config/arelle/cache o falta el "
            f"_shell.xsd del emisor en el dataset."
        )


def run_arelle_exports_progress(
    arelle_dir: Path,
    xbrl_file: Path,
    out_dir: Path,
    stem: str,
    langs: Sequence[str],
    facts_strategy: str = "es_only",
    force: bool = False,
    progress_cb: Optional[Callable[[str], None]] = None,
    offline: bool = False,
) -> None:
    """Igual a run_arelle_exports pero reporta progreso por etapa vía callback.

    Etapas reportadas: facts_<lang>, pre_<lang>

    Si ``offline=True`` se agrega ``--internetConnectivity=offline`` a Arelle,
    evitando llamadas a la red (taxonomía CMF debe estar ya en el HTTP cache).
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    # Absoluto: el subprocess corre con cwd=arelle_dir, donde una ruta
    # relativa al intérprete no resuelve.
    arelle_dir = Path(arelle_dir).resolve()
    arelle_python = arelle_dir / '.venv' / 'bin' / 'python'
    if not arelle_python.exists():
        arelle_python = Path(sys.executable)

    fact_cols = (
        "Label,localName,contextRef,unitRef,Dec,Prec,Lang,Value,"
        "entityIdentifier,periodStart,periodEnd,instant,endInstant,qname"
    )

    xbrl_abs = xbrl_file.resolve()
    out_abs = out_dir.resolve()
    offline_flags = ['--internetConnectivity=offline'] if offline else []

    def _up_to_date(inputs: list[Path], outputs: list[Path]) -> bool:
        try:
            latest_in = max(p.stat().st_mtime for p in inputs if p.exists())
            mt_outs = [p.stat().st_mtime for p in outputs if p.exists()]
            return len(mt_outs) == len(outputs) and all(m >= latest_in for m in mt_outs)
        except Exception:
            return False

    def _emit(stage: str):
        if progress_cb:
            try:
                progress_cb(stage)
            except Exception:
                pass

    for lang in langs:
        label_lang = 'es-CL' if lang.startswith('es') else 'en'
        facts_csv = (out_abs / f"facts_{stem}_{lang}.csv")
        pres_csv = (out_abs / f"presentation_{stem}_{lang}.csv")
        facts_log = (out_abs / f"arelle_facts_{lang}.log")
        pre_log = (out_abs / f"arelle_pre_{lang}.log")

        want_facts = facts_strategy == "both" or lang.startswith('es')
        if not want_facts:
            # Create symlink for EN facts pointing to ES (saves ~22MB per dataset)
            try:
                es_facts = (out_abs / f"facts_{stem}_es.csv")
                if es_facts.exists() and not facts_csv.exists():
                    os.symlink(es_facts, facts_csv)
            except OSError:
                # Fallback to hardlink if symlinks not supported
                try:
                    if es_facts.exists() and not facts_csv.exists():
                        os.link(str(es_facts), str(facts_csv))
                except Exception:
                    pass
            _emit(f"facts_{lang}")

        need_facts = want_facts and (force or not _up_to_date([xbrl_abs], [facts_csv, facts_log]))
        need_pre = force or not _up_to_date([xbrl_abs], [pres_csv, pre_log])

        if need_facts and need_pre:
            # Una sola corrida de Arelle exporta ambos: la carga del DTS (lo
            # caro) se paga una vez en lugar de dos.
            _emit(f"facts+pre_{lang}")
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--factTable', str(facts_csv.resolve()),
                '--factTableCols', fact_cols,
                '--pre', str(pres_csv.resolve()),
                '--logFile', str(facts_log.resolve()),
                *offline_flags,
            ])
            pre_log.write_text("(exportado junto a facts; ver arelle_facts log)\n",
                               encoding="utf-8")
        elif need_facts:
            _emit(f"facts_{lang}")
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--factTable', str(facts_csv.resolve()),
                '--factTableCols', fact_cols,
                '--logFile', str(facts_log.resolve()),
                *offline_flags,
            ])
        elif need_pre:
            _emit(f"pre_{lang}")
            run_arelle(arelle_python, arelle_dir, [
                '-f', str(xbrl_abs),
                f'--labelLang={label_lang}',
                '--pre', str(pres_csv.resolve()),
                '--logFile', str(pre_log.resolve()),
                *offline_flags,
            ])

        _verificar_facts_no_vacio(facts_csv, xbrl_file)


def generate_excels(cmf_dir: Path, out_dir: Path, stem: str, langs: Sequence[str]) -> None:
    script = cmf_dir / 'xbrl_to_excel.py'
    for lang in langs:
        run_cmd([sys.executable, str(script), str(out_dir), stem, lang], cwd=cmf_dir)


def _import_x2e_module(cmf_dir: Path):
    """Importa dinámicamente xbrl_to_excel para reutilizar normalización y árbol."""
    if str(cmf_dir) not in sys.path:
        sys.path.insert(0, str(cmf_dir))
    import importlib
    return importlib.import_module('xbrl_to_excel')


def _normalize_synopsis_name(name: str) -> str:
    """
    Normalize synopsis marker names to ensure consistent capitalization
    ESTA ES LA ÚNICA FUNCIÓN QUE DEBE HACER NORMALIZACIONES DE SYNOPSIS
    """
    if not name or '[sinopsis]' not in name.lower():
        return name
    
    # Common normalizations for synopsis markers - EXHAUSTIVE LIST
    normalizations = {
        # Principales categorías bancarias
        'negocios no bancarios': 'Negocios no bancarios',
        'servicios bancarios': 'Servicios bancarios',
        'activos bancarios': 'Activos bancarios', 
        'pasivos servicios bancarios': 'Pasivos servicios bancarios',
        
        # Estados financieros principales
        'estado de situación financiera': 'Estado de situación financiera',
        'estado de resultados': 'Estado de resultados',
        'estado de flujos de efectivo': 'Estado de flujos de efectivo',
        
        # Categorías principales
        'activos': 'Activos',
        'pasivos': 'Pasivos',
        'patrimonio': 'Patrimonio',
        'patrimonio y pasivos': 'Patrimonio y pasivos',
        'ganancia (pérdida)': 'Ganancia (pérdida)',
        
        # Subcategorías activos/pasivos
        'activos corrientes': 'Activos corrientes',
        'activos no corrientes': 'Activos no corrientes',
        'pasivos corrientes': 'Pasivos corrientes',
        'pasivos no corrientes': 'Pasivos no corrientes',
        
        # Flujos de efectivo - categorías principales
        'flujos de efectivo procedentes de (utilizados en) actividades de operación': 'Flujos de efectivo procedentes de (utilizados en) actividades de operación',
        'flujos de efectivo procedentes de (utilizados en) actividades de inversión': 'Flujos de efectivo procedentes de (utilizados en) actividades de inversión', 
        'flujos de efectivo procedentes de (utilizados en) actividades de financiación': 'Flujos de efectivo procedentes de (utilizados en) actividades de financiación',
        
        # Flujos de efectivo - subcategorías
        'cambios en activos y pasivos que afectan al flujo operacional': 'Cambios en activos y pasivos que afectan al flujo operacional',
        'cargos (abonos) a resultados que no significan movimientos de efectivo': 'Cargos (abonos) a resultados que no significan movimientos de efectivo',
        'clases de cobros por actividades de operación': 'Clases de cobros por actividades de operación',
        'clases de pagos': 'Clases de pagos',
    }
    
    # Extract the main part without [sinopsis]
    main_part = name.replace('[sinopsis]', '').strip()
    main_lower = main_part.lower()
    
    # Apply normalizations
    for key, normalized in normalizations.items():
        if main_lower == key:
            return f"{normalized} [sinopsis]"
    
    # For other cases, apply title case to first letter only
    if main_part:
        normalized_main = main_part[0].upper() + main_part[1:]
        return f"{normalized_main} [sinopsis]"
    
    return name

def _build_presentation_context_map(pres_file: Path) -> dict[str, list[str]]:
    """
    Construye mapa de contextos jerárquicos desde presentation.csv
    Returns: Dict mapping account names to list of their hierarchical contexts
    """
    if not pres_file.exists():
        return {}
    
    try:
        import pandas as pd
        pres = pd.read_csv(pres_file, encoding='utf-8')
    except Exception:
        return {}
    
    current_role_code = None
    hierarchy_stack = {}
    header_re = re.compile(r'^\s*\[(\d{6})\]')
    context_map = {}
    
    for idx, row in pres.iterrows():
        raw_label = row.get('Label') if 'Label' in pres.columns else row.iloc[0]
        label = str(raw_label).strip() if pd.notna(raw_label) else ''
        
        # Find account in indentation columns
        account_found = False
        account_level = -1
        account_name = ''
        
        # Search in unnamed columns (hierarchical indentation)
        for col_idx, col_name in enumerate(pres.columns):
            if col_idx == 0:  # Skip Label/first column
                continue
                
            cell_value = row.iloc[col_idx] if col_idx < len(row) else None
            if pd.notna(cell_value):
                cell_str = str(cell_value).strip()
                if (cell_str and not cell_str.startswith('[') and 
                    cell_str not in ['String', 'Monetary', 'TextBlock', 'Date', 'Boolean', 
                                   'Enumeration', 'TipoEEFF', 'terseLabel', 'negatedTerseLabel', 'netLabel']):
                    account_name = cell_str
                    account_level = col_idx - 1
                    account_found = True
                    break
        
        # If no account found in indentation columns, use Label
        if not account_found and label:
            m = header_re.match(label)
            if m:
                current_role_code = m.group(1)
                hierarchy_stack.clear()
                continue
            account_name = label
            account_level = 0
            account_found = True
        
        if not account_found:
            continue
        
        # Update hierarchy based on column level - normalize synopsis names
        normalized_account_name = _normalize_synopsis_name(account_name)
        hierarchy_stack = {k: v for k, v in hierarchy_stack.items() if k < account_level}
        hierarchy_stack[account_level] = normalized_account_name
        
        # Build full hierarchical context with normalization
        hierarchy_path = []
        for level in sorted(hierarchy_stack.keys()):
            element = hierarchy_stack[level]
            normalized_element = _normalize_synopsis_name(element)
            hierarchy_path.append(normalized_element)
        
        # Context is hierarchy MINUS current element (parents only)
        section_context = ' > '.join(hierarchy_path[:-1]) if len(hierarchy_path) > 1 else ''
        
        # Store the mapping: account_name -> [full_context_keys]
        if account_name and current_role_code:
            full_key = f"{current_role_code}||{section_context}||{normalized_account_name}"
            if account_name not in context_map:
                context_map[account_name] = []
            context_map[account_name].append(full_key)
    
    return context_map

def _aggregate_facts_for_company(company_datasets: List[DatasetInfo], lang: str, cmf_dir: Path):
    """
    Combina facts normalizados de múltiples períodos en un único DataFrame ancho (Label + fechas).
    
    MEJORA CRÍTICA: Evita contaminación entre contextos XBRL y roles diferentes:
    1. Solo cuenta fechas PURAS (YYYY-MM-DD) como períodos válidos
    2. Excluye qname y fechas con sufijos de miembro/dimensión del conteo
    3. Nunca sobrescribe datos existentes con cero datos
    4. Preserva la integridad de cuentas IFRS9 sensibles
    """
    import pandas as pd  # local import
    import re
    # Aumentar límite de tamaño de campo para manejar text blocks grandes
    try:
        csv.field_size_limit(2 ** 31 - 1)
    except Exception:
        pass
    x2e = _import_x2e_module(cmf_dir)

    # Procesar comenzando por los períodos más recientes para preservar mejor el orden
    # del facts original del último período disponible.
    ordered = sorted(company_datasets, key=lambda d: d.yyyyymm, reverse=True)

    # Usaremos una clave base por código de rol + etiqueta visible; pero
    # permitimos múltiples "buckets" por clave para conservar duplicados
    # con distintos valores (p. ej. mismas etiquetas dentro del mismo rol).
    key_to_values: dict[str, dict[str, object]] = {}
    # Nota: key_to_values quedó para compatibilidad con código previo;
    # ahora el almacenamiento real es por buckets en key_to_buckets.
    key_to_buckets: dict[str, list[dict[str, object]]] = {}
    key_to_label: dict[str, str] = {}
    key_to_role: dict[str, str] = {}
    all_dates: set[str] = set()
    min_ym: str | None = None
    max_ym: str | None = None
    # Tipo de estado financiero efectivamente usado por período ('C' consolidado /
    # 'I' individual). Se persiste como sidecar para que la Ficha Técnica y el gate
    # de subida sepan que parte de la serie no es consolidada.
    statement_types: dict[str, str] = {}
    
    # Helper para detectar fechas PURAS vs fechas con contexto
    def _is_pure_date_column(col_name: str) -> bool:
        """
        Acepta:
        - Fechas ISO exactas YYYY-MM-DD (ej: 2023-09-30)
        - Fechas en formato M/D/YYYY o MM/DD/YYYY (ej: 9/30/2023 o 09/30/2023)
        """
        col_name = str(col_name).strip()
        return bool(
            re.fullmatch(r'\d{4}-\d{2}-\d{2}', col_name) or    # 2023-09-30
            re.fullmatch(r'\d{1,2}/\d{1,2}/\d{4}', col_name)   # 9/30/2023 o 09/30/2023
        )

    

    print(f"      ║ Procesando {len(ordered)} períodos...")
    debug_contamination = []  # Para debuggear problemas de contaminación
    
    # Construir mapa de contextos jerárquicos desde presentation.csv del período más reciente
    presentation_context_map = {}
    if ordered:
        most_recent = ordered[0]  # Ya está ordenado por fecha descendente
        pres_path = most_recent.dataset_dir / f"out_{most_recent.stem}" / f"presentation_{most_recent.stem}_{lang}.csv"
        presentation_context_map = _build_presentation_context_map(pres_path)
        if presentation_context_map:
            print(f"      ║ Construido mapa de contextos jerárquicos desde {pres_path.name}: {len(presentation_context_map)} cuentas")
    
    for i, ds in enumerate(ordered, 1):
        if i % 10 == 0 or i == len(ordered):
            print(f"      ║ Progreso: {i}/{len(ordered)} períodos procesados...")
            
        facts_path = ds.dataset_dir / f"out_{ds.stem}" / f"facts_{ds.stem}_{lang}.csv"
        if not facts_path.exists():
            continue

        kind = dataset_statement_type(ds.dataset_dir, ds.stem)
        if kind:
            statement_types[ds.yyyyymm] = kind

        try:
            # Lectura rápida sin optimizaciones que causan lentitud
            facts_raw = pd.read_csv(facts_path, engine='python')
            facts_norm = x2e.normalize_facts(facts_raw)
        except Exception as e:
            print(f"      ⚠️ Error procesando {ds.stem}: {e}")
            continue

        # MEJORA CRÍTICA: Solo contar fechas PURAS, excluir qname y fechas con miembros
        all_cols = [c for c in facts_norm.columns if c != 'Label']
        pure_date_cols = [c for c in all_cols if _is_pure_date_column(c)]
        contaminated_cols = [c for c in all_cols if not _is_pure_date_column(c)]
        
        if len(contaminated_cols) > 0:
            debug_contamination.append(f"  {ds.stem}: {len(contaminated_cols)} cols contaminadas (ej: {contaminated_cols[0] if contaminated_cols else 'none'})")
            
        # Seguimiento del rol actual por encabezado [XXXXXX] y categoría/sección con tag [sinopsis], etc.
        current_role_code: str | None = None
        current_section: str | None = None
        current_main_category: str | None = None  # Categoría principal (operación/inversión/financiación)
        current_subcategory: str | None = None    # Subcategoría (Negocios no bancarios/Servicios bancarios)
        header_re = re.compile(r'^\s*\[(\d{6})\]')
        # Un "category" es una fila con tag entre corchetes que NO es header de rol
        category_tag_re = re.compile(r'^\s*(?!\[\d{6}\]).*\[[^\]]+\]')

        for _, row in facts_norm.iterrows():
            raw_label = row.get('Label')
            label = str(raw_label).strip() if pd.notna(raw_label) else ''
            if not label:
                continue

            # Actualizar rol si es encabezado; también registrar este encabezado como fila propia
            m = header_re.match(label)
            if m:
                current_role_code = m.group(1)
                # Al cambiar de rol, reiniciar secciones
                current_section = None
                current_main_category = None
                current_subcategory = None
                # Guardar encabezado como fila con su propia clave  
                normalized_header_label = _normalize_synopsis_name(label)
                key = f"{current_role_code}||||{normalized_header_label}"  # header sin sección específica
                if key not in key_to_buckets:
                    key_to_buckets[key] = [ {} ]  # header sin valores
                    key_to_label[key] = normalized_header_label  # Usar label normalizado
                    key_to_role[key] = current_role_code
                # Encabezados rara vez tienen valores, pero seguimos al siguiente row
                continue

            # Actualizar sección si es una categoría con tag (p.ej. "[sinopsis]")
            if category_tag_re.match(label):
                normalized_category = _normalize_synopsis_name(label)
                
                # Determinar si es categoría principal o subcategoría
                main_categories = [
                    'Flujos de efectivo procedentes de (utilizados en) actividades de operación [sinopsis]',
                    'Flujos de efectivo procedentes de (utilizados en) actividades de inversión [sinopsis]',
                    'Flujos de efectivo procedentes de (utilizados en) actividades de financiación [sinopsis]'
                ]
                
                subcategories = [
                    'Negocios no bancarios [sinopsis]',
                    'Servicios bancarios [sinopsis]',
                    'Servicios Bancarios [sinopsis]'  # Variación con B mayúscula
                ]
                
                if normalized_category in main_categories:
                    # Es una categoría principal
                    current_main_category = normalized_category
                    current_subcategory = None  # Reset subcategoría
                    current_section = normalized_category
                elif normalized_category in subcategories:
                    # Es una subcategoría
                    current_subcategory = normalized_category
                    # Construir sección completa con jerarquía
                    if current_main_category:
                        current_section = f"{current_main_category}||{normalized_category}"
                    else:
                        current_section = normalized_category
                else:
                    # Otra categoría no específica
                    current_section = normalized_category
                
                # Registrar también la categoría como fila propia (sin valores)
                role_code_for_cat = current_role_code or '000000'
                
                # 🔥 EVITAR registrar subcategorías bancarias sin categoría principal
                # "Servicios bancarios [sinopsis]" solo debe aparecer bajo categorías principales
                should_register_category = True
                if (normalized_category in subcategories and 
                    not current_main_category and 
                    role_code_for_cat == '510000'):
                    # Es una subcategoría bancaria sin categoría principal activa - NO registrar
                    should_register_category = False
                    # Saltando registro de subcategoría sin padre
                
                if should_register_category:
                    cat_key = f"{role_code_for_cat}||{current_section}||{normalized_category}"
                    if cat_key not in key_to_buckets:
                        key_to_buckets[cat_key] = [ {'__section__': current_section, '__qname__': str(row.get('qname')) if 'qname' in facts_norm.columns else ''} ]
                        key_to_label[cat_key] = current_section  # Use normalized section name as label too
                        key_to_role[cat_key] = role_code_for_cat

                # 🔥 NUEVO: Crear automáticamente subcategorías bancarias bajo cada categoría principal
                # Si encontramos una subcategoría bancaria, crearla bajo las 3 categorías principales
                # Normalizar variaciones de capitalización
                normalized_lower = normalized_category.lower()
                is_servicios_bancarios = (
                    'servicios bancarios [sinopsis]' in normalized_lower or
                    'servicios bancarios [sinopsis]' == normalized_lower
                )
                is_negocios_no_bancarios = (
                    'negocios no bancarios [sinopsis]' in normalized_lower and
                    'aseguradora' not in normalized_lower  # Excluir "Negocios no aseguradora y no bancarios"
                )
                
                # 🔥 FORZAR: Siempre agregar las subcategorías bancarias obligatorias
                # Independientemente de si se detectan en los datos, DEBEN existir
                if role_code_for_cat == '510000':
                    # Categorías principales que SIEMPRE deben tener subcategorías
                    main_categories_obligatorias = [
                        'Flujos de efectivo procedentes de (utilizados en) actividades de operación [sinopsis]',
                        'Flujos de efectivo procedentes de (utilizados en) actividades de inversión [sinopsis]',
                        'Flujos de efectivo procedentes de (utilizados en) actividades de financiación [sinopsis]'
                    ]
                    
                    # Subcategorías que DEBEN aparecer bajo cada categoría principal
                    subcategorias_obligatorias = [
                        'Negocios no bancarios [sinopsis]',
                        'Servicios bancarios [sinopsis]'
                    ]
                    
                    # Forzando creación de subcategorías obligatorias para Cash Flow
                    for main_category in main_categories_obligatorias:
                        # Asegurar que existe la categoría principal
                        main_key = f"{role_code_for_cat}||{main_category}||{main_category}"
                        if main_key not in key_to_buckets:
                            key_to_buckets[main_key] = [ {'__section__': main_category, '__qname__': ''} ]
                            key_to_label[main_key] = main_category
                            key_to_role[main_key] = role_code_for_cat
                            # Creada categoría principal
                        
                        # FORZAR creación de cada subcategoría
                        for subcategory in subcategorias_obligatorias:
                            subcategory_key = f"{role_code_for_cat}||{main_category}||{subcategory}"
                            if subcategory_key not in key_to_buckets:
                                key_to_buckets[subcategory_key] = [ {'__section__': f"{main_category}||{subcategory}", '__qname__': ''} ]
                                key_to_label[subcategory_key] = f"{main_category}||{subcategory}"
                                key_to_role[subcategory_key] = role_code_for_cat
                                # Subcategoría forzada
                
                # Mantener lógica original para casos detectados automáticamente
                if role_code_for_cat == '510000' and (is_servicios_bancarios or is_negocios_no_bancarios):
                    # Esta lógica ahora es redundante pero la mantenemos por compatibilidad
                    # Subcategoría detectada automáticamente
                    pass

            # Si no hay rol detectado aún, asignar rol desconocido '000000'
            role_code = current_role_code or '000000'
            
            # Usar el contexto jerárquico completo ya capturado
            section_context = current_section or ''
            
            # 🔥 NUEVO: Si la cuenta tiene jerarquía completa, crear TODAS las categorías intermedias
            # Por ejemplo, si tenemos: 510000||Flujos...||Servicios bancarios||Cuenta específica
            # Debemos asegurar que existe: 510000||Flujos...||Servicios bancarios [sinopsis]
            if role_code == '510000' and section_context and '||' in section_context:
                partes = section_context.split('||')
                if len(partes) >= 2:
                    categoria_principal = partes[0]
                    subcategoria = partes[1] if len(partes) > 1 else None
                    
                    # Si hay subcategoría bancaria, crear su entrada de categoría
                    if subcategoria and ('bancarios' in subcategoria.lower() or 'bancarios' in subcategoria.lower()):
                        # Normalizar nombre de subcategoría
                        if 'servicios' in subcategoria.lower():
                            subcategoria_normalizada = 'Servicios bancarios [sinopsis]'
                        elif 'negocios' in subcategoria.lower():
                            subcategoria_normalizada = 'Negocios no bancarios [sinopsis]'
                        else:
                            subcategoria_normalizada = subcategoria
                        
                        # Crear entrada de categoría para la subcategoría
                        subcat_key = f"{role_code}||{categoria_principal}||{subcategoria_normalizada}"
                        if subcat_key not in key_to_buckets:
                            key_to_buckets[subcat_key] = [ {'__section__': f"{categoria_principal}||{subcategoria_normalizada}", '__qname__': ''} ]
                            key_to_label[subcat_key] = f"{categoria_principal}||{subcategoria_normalizada}"
                            key_to_role[subcat_key] = role_code
                            # Auto-creada categoría intermedia
            

            # Usar directamente el contexto jerárquico completo capturado desde facts
            contexts_to_process = [section_context]
            
            # Normalizar el label también para consistencia
            normalized_label = _normalize_synopsis_name(label)
            
            # Procesar la fila actual con su contexto específico
            for context_section in contexts_to_process:
                # Si es jerarquía doble (contiene ||), usar directamente
                if '||' in context_section:
                    key = f"{role_code}||{context_section}||{normalized_label}"
                else:
                    # Jerarquía simple
                    key = f"{role_code}||{context_section}||{normalized_label}"

                # Inicializar buckets y metadatos
                if key not in key_to_buckets:
                    key_to_buckets[key] = []
                    key_to_label[key] = normalized_label  # Usar label normalizado
                    key_to_role[key] = role_code

                buckets = key_to_buckets[key]

                # Procesar todos los datos de esta fila en su contexto correspondiente
                current_data_cols = [dc for dc in pure_date_cols if pd.notna(row.get(dc)) and str(row.get(dc)).strip() != '']
                current_data_count = len(current_data_cols)

                # qname como discriminante si existe
                qname_val = str(row.get('qname')) if 'qname' in facts_norm.columns else ''

                def _norm_val(v):
                    import pandas as _pd
                    if v is None or _pd.isna(v):
                        return None
                    s = str(v).strip()
                    if s == '':
                        return None
                    # intentar numérico
                    try:
                        nv = _pd.to_numeric(s, errors='raise')
                        return float(nv)
                    except Exception:
                        return s

                # Representación de la fila actual: mapa fecha->valor normalizado
                current_map: dict[str, object] = {dc: _norm_val(row.get(dc)) for dc in current_data_cols}

                def _maps_equal(bmap: dict[str, object], cmap: dict[str, object]) -> bool:
                    # igualdad estricta en la unión de fechas: None vs valor se considera distinto
                    dates = set(bmap.keys()) | set(cmap.keys())
                    for d in dates:
                        vb = bmap.get(d, None)
                        vc = cmap.get(d, None)
                        if vb is None and vc is None:
                            continue
                        if vb == vc:
                            continue
                        return False
                    return True

                if current_data_count > 0:
                    placed = False
                    for b in buckets:
                        # respetar seccion/qname como partición de buckets
                        if b.get('__section__') != context_section or b.get('__qname__') != qname_val:
                            continue
                        
                        # MEJORA: Para cuentas de totales en todos los roles, ser más flexibles con conflictos
                        # Identificar cuentas de totales/subtotales que suelen tener variaciones mínimas
                        is_total_account = (
                            'Total' in normalized_label or 
                            'Patrimonio' in normalized_label or
                            'Subtotal' in normalized_label or
                            normalized_label.startswith('Total ') or
                            normalized_label.endswith(' total') or
                            normalized_label.endswith(' neto') or
                            'Resultado' in normalized_label
                        )
                        
                        # Debug para cuentas de totales
                        import os as _os
                        if is_total_account and _os.getenv('X2E_DEBUG') == '1':
                            print(f"        🔍 Procesando cuenta total: {normalized_label} (período {ds.stem})")
                        
                        # comparar solo intersección; manejar conflictos eligiendo el valor más alto
                        conflict = False
                        conflict_resolutions = {}  # Para almacenar resoluciones de conflictos
                        
                        for dc, vv in current_map.items():
                            if dc in b:
                                bv = b.get(dc)
                                if bv is not None and vv is not None and bv != vv:
                                    # Intentar resolver el conflicto eligiendo el valor más alto
                                    try:
                                        # Convertir strings con formato 69,069,688,917,000 a números
                                        bv_num = float(str(bv).replace(',', ''))
                                        vv_num = float(str(vv).replace(',', ''))
                                        
                                        # Para diferencias muy pequeñas, no considerar conflicto
                                        if abs(bv_num) > 0:
                                            rel_diff = abs((vv_num - bv_num) / bv_num)
                                            if rel_diff < 0.0001:  # Diferencia menor al 0.01%
                                                continue  # No es conflicto real
                                        
                                        # SIEMPRE preferir el valor más alto (en valor absoluto)
                                        if abs(vv_num) > abs(bv_num):
                                            conflict_resolutions[dc] = vv  # Marcar para actualizar con el valor más alto
                                        # Si el valor existente es más alto, mantenerlo (no hacer nada)
                                        
                                    except:
                                        # Si no se puede comparar numéricamente, considerar conflicto real
                                        if not is_total_account:
                                            conflict = True
                                            break
                        
                        if conflict:
                            continue  # Solo continuar si hay un conflicto no resuelto
                        
                        # Aplicar resoluciones de conflictos (valores más altos)
                        for dc, resolved_value in conflict_resolutions.items():
                            b[dc] = resolved_value
                            if _os.getenv('X2E_DEBUG') == '1':
                                print(f"          🔄 Resolviendo conflicto en {normalized_label} para {dc}: usando valor más alto {resolved_value}")
                        
                        # merge: agregar fechas nuevas o vacías
                        for dc, vv in current_map.items():
                            if dc not in b or b.get(dc) is None or str(b.get(dc)).strip() == '':
                                b[dc] = vv
                                all_dates.add(dc)
                                ym = str(dc)[:7]
                                min_ym = ym if (min_ym is None or ym < min_ym) else min_ym
                                max_ym = ym if (max_ym is None or ym > max_ym) else max_ym
                        placed = True
                        break

                    if not placed:
                        # Solo crear bucket nuevo si realmente tiene datos
                        if current_data_count > 0:
                            b_new: dict[str, object] = {'__section__': context_section, '__qname__': qname_val}
                            for dc, vv in current_map.items():
                                b_new[dc] = vv
                                all_dates.add(dc)
                                ym = str(dc)[:7]
                                min_ym = ym if (min_ym is None or ym < min_ym) else min_ym
                                max_ym = ym if (max_ym is None or ym > max_ym) else max_ym
                            buckets.append(b_new)

    # Debug de contaminación
    if debug_contamination:
        # DEBUG: Columnas contaminadas detectadas
        for dc in debug_contamination[:5]:  # Mostrar primeros 5
            print(f"      ║   {dc}")

    # OPTIMIZACIÓN: Construir DataFrame de forma más eficiente
    sorted_dates = sorted(all_dates, reverse=True)
    
    # 🔥 FORZAR CREACIÓN DE SUBCATEGORÍAS OBLIGATORIAS EN CASH FLOW
    # Asegurar que existen las 3 entradas de cada subcategoría bajo cada categoría principal
    # Verificando subcategorías obligatorias en Cash Flow...
    main_categories_cash_flow = [
        'Flujos de efectivo procedentes de (utilizados en) actividades de operación [sinopsis]',
        'Flujos de efectivo procedentes de (utilizados en) actividades de inversión [sinopsis]',
        'Flujos de efectivo procedentes de (utilizados en) actividades de financiación [sinopsis]'
    ]
    subcategorias_obligatorias = [
        'Negocios no bancarios [sinopsis]',
        'Servicios bancarios [sinopsis]'
    ]
    
    for main_cat in main_categories_cash_flow:
        for subcat in subcategorias_obligatorias:
            # Crear clave para la subcategoría bajo la categoría principal
            subcat_key = f"510000||{main_cat}||{subcat}"
            
            # Si no existe, crearla
            if subcat_key not in key_to_buckets:
                key_to_buckets[subcat_key] = [ {'__section__': f"{main_cat}||{subcat}", '__qname__': ''} ]
                key_to_label[subcat_key] = f"{main_cat}||{subcat}"
                key_to_role[subcat_key] = '510000'
                print(f"      ║   ✅ CREADA: {main_cat} → {subcat}")
    
    if not key_to_buckets:
        df = pd.DataFrame({'LabelKeyId': [], 'Label': [], 'RoleCode': []})
        for dc in sorted_dates:
            df[dc] = []
    else:
        # Usar construcción directa en lugar de records para mejor rendimiento
        # 🔧 NORMALIZACIÓN GLOBAL FINAL - Aplicar a todas las claves existentes
        print(f"      ║ 🔧 Aplicando normalización global a {len(key_to_label)} labels...")
        normalized_key_to_label = {}
        for key, label in key_to_label.items():
            normalized_label = _normalize_synopsis_name(label)
            normalized_key_to_label[key] = normalized_label
        key_to_label = normalized_key_to_label  # Reemplazar con versión normalizada
        
        final_labelkeyids: list[str] = []
        final_labelkeyext: list[str] = []
        final_sections: list[str] = []
        final_labels: list[str] = []
        final_roles: list[str] = []
        rows_by_date: dict[str, list[object]] = {dc: [] for dc in sorted_dates}

        for base_key, buckets in key_to_buckets.items():
            # Filtrar buckets vacíos (sin fechas con datos)
            non_empty_buckets = []
            for bucket in buckets:
                # Un bucket está vacío si no tiene datos en fechas puras
                has_data = False
                for dc in sorted_dates:
                    if bucket.get(dc) is not None:
                        has_data = True
                        break
                
                # Si es categoría/header, conservar aunque no tenga datos
                if not has_data:
                    section = bucket.get('__section__', '')
                    if section or '[' in key_to_label.get(base_key, ''):
                        has_data = True
                
                if has_data:
                    non_empty_buckets.append(bucket)
            
            # Procesar solo buckets no vacíos
            for idx_b, bucket in enumerate(non_empty_buckets):
                section_lbl = str(bucket.get('__section__') or '').strip()
                # Mantener LabelKeyId base para compatibilidad; extender en columna separada
                lkid = base_key
                lkid_ext = f"{base_key}||{section_lbl}" if section_lbl else base_key
                # Si hay duplicados de valores distintos en la misma sección, numerar con ##n
                if idx_b > 0:
                    lkid = f"{lkid}##{idx_b+1}"
                    lkid_ext = f"{lkid_ext}##{idx_b+1}"
                final_labelkeyids.append(lkid)
                final_labelkeyext.append(lkid_ext)
                final_sections.append(section_lbl)
                final_labels.append(key_to_label.get(base_key, ''))
                final_roles.append(key_to_role.get(base_key, ''))
                for dc in sorted_dates:
                    rows_by_date[dc].append(bucket.get(dc))

        data_dict = {
            'LabelKeyId': final_labelkeyids,
            'LabelKeyIdExt': final_labelkeyext,
            'SectionKey': final_sections,
            'Label': final_labels,
            'RoleCode': final_roles,
        }
        for dc in sorted_dates:
            data_dict[dc] = rows_by_date[dc]

        df = pd.DataFrame(data_dict)
        pure_dates_count = len([d for d in sorted_dates if _is_pure_date_column(d)])
        print(f"      ║ DataFrame consolidado creado: {len(df)} cuentas, {pure_dates_count} fechas puras de {len(sorted_dates)} totales")
    
    # LIMPIEZA: eliminar filas sin rol/código o con datos perdidos (párrafos sueltos)
    try:
        import pandas as _pd
        # Rol válido: seis dígitos y distinto de '000000'
        role_series = df.get('RoleCode', _pd.Series(dtype=object)).astype(str).str.strip()
        labelkey_series = df.get('LabelKeyId', _pd.Series(dtype=object)).astype(str).str.strip()
        role_mask = role_series.str.fullmatch(r'\d{6}', na=False) & (role_series != '000000')
        lkid_mask = labelkey_series.str.contains(r'^\d{6}\|\|', regex=True, na=False)
        keep_mask = role_mask & lkid_mask
        before_rows = len(df)
        df = df[keep_mask].copy()
        removed = before_rows - len(df)
        if removed > 0:
            print(f"      ║ Limpieza facts: removidas {removed} filas sin RoleCode/LabelKeyId válidos")

        # Limpieza adicional: descartar filas cuyo Label parece un párrafo (texto narrativo),
        # pero mantener headers [XXXXXX] y conceptos con tags (p.ej. "[bloque de texto]").
        lbl = df['Label'].astype(str)
        header_mask = lbl.str.match(r'^\s*\[(\d{6})\]')
        has_tag_mask = lbl.str.contains(r'\[[^\]]+\]', regex=True)
        multiline_mask = lbl.str.contains(r'[\r\n]', regex=True)
        bullet_mask = lbl.str.match(r'^\s*[-•–—]\s+')
        very_long_mask = lbl.str.len() > 180
        # contar palabras de forma robusta
        try:
            word_count = lbl.str.split().map(len)
        except Exception:
            word_count = _pd.Series([0] * len(lbl), index=lbl.index)
        many_words_mask = word_count > 30
        punct_count = lbl.str.count(r'[\.!?;:,]')
        para_punct_mask = (punct_count > 3) & (lbl.str.len() > 100)

        # Solo eliminar "párrafo" si además NO tiene datos numéricos en ninguna fecha pura
        date_cols_clean = [c for c in df.columns if re.fullmatch(r'\d{4}-\d{2}-\d{2}', str(c))]
        def _row_has_data(idx: int) -> bool:
            try:
                row = df.loc[idx, date_cols_clean]
                # considerar no vacío y no string vacía
                return any((x is not None) and (str(x).strip() != '') for x in row.values)
            except Exception:
                return False
        has_data_mask = df.index.to_series().map(_row_has_data)

        paragraph_like = (~header_mask) & (~has_tag_mask) & (
            multiline_mask | bullet_mask | very_long_mask | many_words_mask | para_punct_mask
        ) & (~has_data_mask)

        before_rows2 = len(df)
        df = df[~paragraph_like].copy()
        removed2 = before_rows2 - len(df)
        if removed2 > 0:
            print(f"      ║ Limpieza facts: removidas {removed2} filas con Label tipo párrafo")
    except Exception:
        # En caso de cualquier problema, continuar sin bloquear el flujo
        pass

    # Reordenar filas para respetar la presentación del período más reciente
    try:
        latest_ds = max(company_datasets, key=lambda d: d.yyyyymm)
        pres_path = latest_ds.dataset_dir / f"out_{latest_ds.stem}" / f"presentation_{latest_ds.stem}_{lang}.csv"
        if pres_path.exists():
            import pandas as _pd
            pres_df = _pd.read_csv(pres_path, engine='python')
            # Extraer secuencia de labels por RoleCode recorriendo la columna 'Label'
            role_orders: dict[str, list[str]] = {}
            current_role: str | None = None
            for _, prow in pres_df.iterrows():
                lbl = str(prow.get('Label') or '').strip()
                if not lbl:
                    continue
                m = re.match(r'^\s*\[(\d{6})\]', lbl)
                if m:
                    current_role = m.group(1)
                    role_orders.setdefault(current_role, []).append(lbl)
                else:
                    if current_role:
                        role_orders.setdefault(current_role, []).append(lbl)

            # Orden de roles según primera aparición en presentation
            role_rank: dict[str, int] = {}
            rank = 0
            for rc in role_orders.keys():
                role_rank[rc] = rank
                rank += 1

            # ✨ APLICAR ORDENAMIENTO PERFECTO BASADO EN JSON ✨
            # Usar el mismo algoritmo perfecto que genera primary_roles_csv.py
            try:
                import os as _os
                debug = _os.getenv('X2E_DEBUG') == '1'
                company_dir_path = company_datasets[0].company_dir
                df = apply_perfect_json_ordering(df, company_dir_path, lang=lang, enable_log=debug)
                if debug:
                    print(f"      ║ ✨ Ordenamiento JSON perfecto aplicado: {len(df)} filas")
            except Exception as json_err:
                debug = _os.getenv('X2E_DEBUG') == '1'
                if debug:
                    print(f"      ║ ⚠ Error en ordenamiento JSON perfecto: {json_err}")
                # Fallback: usar ordenamiento básico anterior basado en presentation
                def _label_idx(rc: str, lbl: str) -> int:
                    seq = role_orders.get(rc)
                    if not seq:
                        return 10**9
                    try:
                        return seq.index(lbl)
                    except ValueError:
                        return 10**9
                def _dup_idx(lkid: str) -> int:
                    m = re.search(r'##(\d+)$', str(lkid))
                    return int(m.group(1)) if m else 1
                df = df.copy()
                df['__role_rank'] = df['RoleCode'].astype(str).map(lambda rc: role_rank.get(rc, 10**6))
                df['__label_rank'] = df.apply(lambda r: _label_idx(str(r['RoleCode']), str(r['Label'])), axis=1)
                df['__dup_rank'] = df['LabelKeyId'].astype(str).map(_dup_idx)
                df.sort_values(['__role_rank', '__label_rank', '__dup_rank'], inplace=True, kind='stable')
                df.drop(columns=['__role_rank', '__label_rank', '__dup_rank'], inplace=True, errors='ignore')
    except Exception:
        pass
    # Anotar rango permitido (YYYY-MM) para este consolidado en un archivo lateral para que xbrl_to_excel lo use
    try:
        rng_path = cmf_dir / '.tmp_allowed_range.txt'
        if min_ym and max_ym:
            rng_path.write_text(f"{min_ym},{max_ym}", encoding='utf-8')
            print(f"      ║ Rango de datos: {min_ym} a {max_ym}")
    except Exception:
        pass

    # Sidecar con el tipo de estado por período. Los períodos 'I' (Individual) no
    # son comparables con los 'C' (Consolidado): el individual excluye filiales.
    try:
        if statement_types and company_datasets:
            individual = sorted(ym for ym, k in statement_types.items() if k == 'I')
            payload = {
                'by_period': dict(sorted(statement_types.items())),
                'individual_periods': individual,
                'mixed': bool(individual) and len(individual) < len(statement_types),
            }
            sidecar = company_datasets[0].company_dir / 'statement_types.json'
            sidecar.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
            if individual:
                print(f"      ║ ⚠ {len(individual)} período(s) sin consolidado, se usó Individual "
                      f"({individual[0]}–{individual[-1]}): serie NO homogénea")
    except Exception:
        pass


    # Liberar memoria
    key_to_values.clear()
    key_to_buckets.clear()
    key_to_label.clear()
    key_to_role.clear()
    all_dates.clear()

    return df


def generate_consolidated_company(company_dir: Path, company_datasets: List[DatasetInfo], cmf_dir: Path, langs: Sequence[str], products_dir: Path) -> None:
    """Genera un Excel consolidado por empresa combinando todos los períodos disponibles."""
    import pandas as pd  # local import

    if not company_datasets:
        return

    rut = company_datasets[0].rut
    # Intentar extraer RUT con dígito verificador desde el nombre del directorio de la empresa
    # Formato esperado: "<rut>-<dv>_<NOMBRE_EMPRESA>"
    try:
        company_dir_name = company_dir.name
        rut_prefix = company_dir_name.split('_', 1)[0]
        rut_with_dv = rut_prefix if '-' in rut_prefix else rut
    except Exception:
        rut_with_dv = rut
    min_yyyymm = min(ds.yyyyymm for ds in company_datasets)
    max_yyyymm = max(ds.yyyyymm for ds in company_datasets)
    latest_ds = max(company_datasets, key=lambda d: d.yyyyymm)

    stem_range = f"{rut_with_dv}_{min_yyyymm}-{max_yyyymm}"
    out_dir = company_dir / f"out_consolidated_{stem_range}"
    out_dir.mkdir(parents=True, exist_ok=True)

    x2e = _import_x2e_module(cmf_dir)

    # Determinar frecuencia para etiquetado/filtrado de periodos
    try:
        freq = "Anual" if any(p.name == 'Anual' for p in company_dir.parents) or 'Anual' in str(company_dir) else (
            "Trimestral" if any(p.name == 'Trimestral' for p in company_dir.parents) or 'Trimestral' in str(company_dir) else "Anual"
        )
    except Exception:
        freq = "Anual"

    # Configurar variables de entorno para el writer según frecuencia
    import os as _os
    _os.environ['X2E_DECEMBER_AS_YEAR'] = '1' if freq == 'Anual' else '0'
    _os.environ['X2E_KEEP_ONLY_QUARTERS'] = '1' if freq == 'Trimestral' else '0'
    # En anual, mantener todas las fechas; en trimestral, conservar heurística por defecto
    if freq == 'Anual':
        _os.environ['X2E_KEEP_ALL_DATES'] = '1'

    for lang in langs:
        # 1) Agregar facts multi-período
        df_facts = _aggregate_facts_for_company(company_datasets, lang, cmf_dir)
        if df_facts.empty:
            continue
        
        # ✨ APLICAR PROCESAMIENTO CONSOLIDADO (TODOS LOS ROLES) ✨
        # Procesar los facts consolidados manteniendo TODOS los roles (incluye notas)
        try:
            import os as _os
            debug = _os.getenv('X2E_DEBUG') == '1'
            company_dir_path = Path(cmf_dir) / company_datasets[0].company_dir.name
            df_facts_processed = apply_consolidated_processing(
                df_facts, company_dir_path, lang=lang, enable_log=debug
            )
            if df_facts_processed is not None and not df_facts_processed.empty:
                df_facts = df_facts_processed
                if debug:
                    print(f"      ║ ✨ Procesamiento PRIMARY_CSV aplicado: {len(df_facts)} filas")
        except Exception as e:
            debug = _os.getenv('X2E_DEBUG') == '1'
            if debug:
                print(f"      ║ ⚠ Error en procesamiento PRIMARY_CSV: {e}")
        
        facts_csv = out_dir / f"facts_{stem_range}_{lang}.csv"
        # Ordenar columnas de fechas de más nuevo a más antiguo y mantener columnas no-fecha al inicio
        try:
            cols = list(df_facts.columns)
            # Columnas de fecha reales (YYYY-MM-DD)
            date_cols = [c for c in cols if re.fullmatch(r'\d{4}-\d{2}-\d{2}', str(c))]
            # Mantener todas las demás columnas (Label, LabelKeyId, RoleCode, etc.) al inicio en el mismo orden
            non_date_cols = [c for c in cols if c not in date_cols]

            def _sort_key(lbl: str) -> tuple[int, int]:
                return _import_x2e_module(cmf_dir)._period_sort_key(str(lbl))  # reuse helper
            date_cols_sorted = sorted(date_cols, key=_sort_key, reverse=True)
            df_facts = df_facts[non_date_cols + date_cols_sorted]
        except Exception:
            pass
        df_facts.to_csv(facts_csv, index=False)

        # 2) Tomar la presentación del período más reciente como árbol base
        latest_pres_csv = latest_ds.dataset_dir / f"out_{latest_ds.stem}" / f"presentation_{latest_ds.stem}_{lang}.csv"
        if not latest_pres_csv.exists():
            # Fallback: si no existe, intentar con otro dataset
            for ds in sorted(company_datasets, key=lambda d: d.yyyyymm, reverse=True):
                alt = ds.dataset_dir / f"out_{ds.stem}" / f"presentation_{ds.stem}_{lang}.csv"
                if alt.exists():
                    latest_pres_csv = alt
                    break

        pres_csv = out_dir / f"presentation_{stem_range}_{lang}.csv"
        if latest_pres_csv.exists():
            # copiar contenido
            pd.read_csv(latest_pres_csv, engine='python').to_csv(pres_csv, index=False)
        else:
            # Si no hay presentación, crear mínima dummy para no fallar
            pd.DataFrame({'roleUri': [None], 'order': [1], 'depth': [1], 'Label': ['dummy'], 'presLabel': ['dummy']}).to_csv(pres_csv, index=False)

        # 3) Generar Excel con el writer corporativo existente
        # SKIP si CMF_SKIP_OLD_EXCEL está configurado (para usar primary_csv_to_excel en su lugar)
        import os as _os
        if _os.getenv('CMF_SKIP_OLD_EXCEL') != '1':
            generate_excels(cmf_dir, out_dir, stem_range, [lang])
        else:
            if _os.getenv('X2E_DEBUG') == '1':
                print(f"      ║ ⏩ Saltando generación de Excel antiguo (CMF_SKIP_OLD_EXCEL=1)")

    # Post-fix: alinear y rellenar columnas entre ES/EN ya generados (consolidado)
    # También saltar si CMF_SKIP_OLD_EXCEL está configurado
    if _os.getenv('CMF_SKIP_OLD_EXCEL') != '1':
        try:
            post_align_language_excels(out_dir, stem_range)
        except Exception:
            pass

    # 4) Copiar a Products con nombre claro, garantizando único por empresa/idioma
    # Determinar frecuencia para destinar en subcarpeta (Anual/Trimestral)
    # SKIP si CMF_SKIP_OLD_EXCEL está configurado (para usar primary_csv_to_excel en su lugar)
    
    if _os.getenv('CMF_SKIP_OLD_EXCEL') != '1':
        for lang in langs:
            src = out_dir / f"estados_{stem_range}_{lang}.xlsx"
            freq_dir = products_dir / freq
            freq_dir.mkdir(parents=True, exist_ok=True)
            dst = freq_dir / f"estados_{rut_with_dv}_{max_yyyymm[:4]}-{min_yyyymm[:4]}_{lang}.xlsx"
            if src.exists():
                # Eliminar versiones previas del mismo rut/idioma para evitar duplicados
                for f in freq_dir.glob(f"estados_{rut_with_dv}_*-*_*.xlsx"):
                    if f.name.endswith(f"_{lang}.xlsx"):
                        try:
                            f.unlink()
                        except Exception:
                            pass
                copy2(src, dst)
    else:
        if _os.getenv('X2E_DEBUG') == '1':
            print(f"      ║ ⏩ Saltando copia a Products (CMF_SKIP_OLD_EXCEL=1) - se usará primary_csv_to_excel")


def generate_combined_company_from_total(company_dir: Path, company_datasets: List[DatasetInfo], cmf_dir: Path, langs: Sequence[str], products_dir: Path) -> dict[str, Path]:
    """Genera un Excel desde XBRL/Total con columnas anuales (Q4) + trimestres recientes agrupados."""
    import pandas as pd  # local import
    import os as _os

    # Define debug based on environment variable
    debug = _os.getenv('X2E_DEBUG') == '1'

    if not company_datasets:
        return {}

    rut = company_datasets[0].rut
    try:
        company_dir_name = company_dir.name
        rut_prefix = company_dir_name.split('_', 1)[0]
        rut_with_dv = rut_prefix if '-' in rut_prefix else rut
    except Exception:
        rut_with_dv = rut
    
    min_yyyymm = min(ds.yyyyymm for ds in company_datasets)
    max_yyyymm = max(ds.yyyyymm for ds in company_datasets)
    year_range = f"{max_yyyymm[:4]}-{min_yyyymm[:4]}"
    ym_range = f"{min_yyyymm}-{max_yyyymm}"
    
    # Crear directorio Products/Total (evitar anidar si ya es 'Total')
    if products_dir.name.lower() == "total":
        products_total_dir = products_dir
    else:
        products_total_dir = products_dir / "Total"
    products_total_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"   📊 Generando desde XBRL/Total: {company_dir.name}")
    print(f"      Rango: {year_range}")
    print(f"      Períodos: {len(company_datasets)}")
    

    stem_tmp = f"{rut_with_dv}_{ym_range}"
    # Crear directorio out_consolidated_* en el directorio de la empresa
    out_consolidated_dir = company_dir / f"out_consolidated_{year_range}"
    out_consolidated_dir.mkdir(parents=True, exist_ok=True)
    
    # Generar para cada idioma
    created_outputs: dict[str, Path] = {}
    import time as _time
    _t_company = _time.perf_counter()
    
    for lang in langs:
        try:
            # Configurar variables de entorno para modo combinado
            _os.environ['X2E_COMBINED'] = '1'
            _os.environ['X2E_KEEP_ALL_DATES'] = '1'
            # Mapear diciembre a 'YYYY' para que exista columna anual por año
            _os.environ['X2E_DECEMBER_AS_YEAR'] = '1'
            # Evitar que se recorten años antiguos por umbral de no vacíos
            _os.environ['X2E_AUTO_TRIM_EMPTY_TAIL'] = '0'
            
            # Generar Excel consolidado
            # Usar la función main de xbrl_to_excel.py con argumentos sintéticos
            import sys as _sys
            from xbrl_to_excel import main as x2e_main
            
            # Generar facts CSV en out_consolidated
            facts_csv = out_consolidated_dir / f"facts_{rut_with_dv}_{ym_range}_{lang}.csv"
            facts_consolidated = _aggregate_facts_for_company(company_datasets, lang, cmf_dir)
            if facts_consolidated is not None and not facts_consolidated.empty:
                # ✨ APLICAR PROCESAMIENTO CONSOLIDADO (TODOS LOS ROLES) ✨
                try:
                    company_dir_path = company_datasets[0].company_dir
                    facts_consolidated_processed = apply_consolidated_processing(
                        facts_consolidated, company_dir_path, lang=lang, enable_log=debug
                    )
                    if facts_consolidated_processed is not None and not facts_consolidated_processed.empty:
                        facts_consolidated = facts_consolidated_processed
                        if debug:
                            print(f"      ║ ✨ Procesamiento PRIMARY_CSV aplicado en out_consolidated: {len(facts_consolidated)} filas")
                except Exception as e:
                    if debug:
                        print(f"      ║ ⚠ Error en procesamiento PRIMARY_CSV consolidado: {e}")
                
                facts_consolidated.to_csv(facts_csv, index=False)
            
            # ✨ EJECUTAR GENERATE_PRIMARY_ROLES_CSV.PY COMPLETO ✨
            print(f"      ║ 🔄 Ejecutando generate_primary_roles_csv.py...")
            try:
                company_dir_for_primary = company_datasets[0].company_dir
                print(f"      ║ 📁 Company dir: {company_dir_for_primary}")
                
                # Ejecutar el script completo de generate_primary_roles_csv.py
                import subprocess
                import sys
                
                cmd = [
                    sys.executable, 
                    "generate_primary_roles_csv.py", 
                    "--company-dir", str(company_dir_for_primary),
                    "--lang", lang
                ]
                
                print(f"      ║ 🚀 Ejecutando: {' '.join(cmd)}")
                result = subprocess.run(cmd, cwd=cmf_dir, capture_output=True, text=True)
                
                if result.returncode == 0:
                    print(f"      ║ ✅ generate_primary_roles_csv.py ejecutado exitosamente")
                    if result.stdout.strip():
                        for line in result.stdout.strip().split('\n'):
                            if line.strip():
                                print(f"      ║ 📝 {line}")
                    
                    # Buscar el archivo generado
                    import glob
                    pattern = str(company_dir_for_primary / "out_consolidated*" / f"primary_roles_*_{lang}.csv")
                    generated_files = glob.glob(pattern)
                    if generated_files:
                        for file_path in generated_files:
                            file_size = Path(file_path).stat().st_size
                            print(f"      ║ 📄 Archivo generado: {file_path} ({file_size} bytes)")
                    else:
                        print(f"      ║ ⚠ No se encontraron archivos primary_roles generados")
                        
                else:
                    print(f"      ║ ❌ Error ejecutando generate_primary_roles_csv.py (código: {result.returncode})")
                    if result.stderr:
                        for line in result.stderr.strip().split('\n'):
                            if line.strip():
                                print(f"      ║ ⚠ {line}")
                    if result.stdout:
                        for line in result.stdout.strip().split('\n'):
                            if line.strip():
                                print(f"      ║ 📝 {line}")
                                
            except Exception as e:
                print(f"      ║ ❌ Error ejecutando generate_primary_roles_csv.py: {e}")
                import traceback
                if debug:
                    traceback.print_exc()
                    
            # Generar presentation CSV (usar el del dataset más reciente)
            latest_ds = max(company_datasets, key=lambda d: d.yyyyymm)
            latest_pres_csv = latest_ds.dataset_dir / f"out_{latest_ds.stem}" / f"presentation_{latest_ds.stem}_{lang}.csv"
            if latest_pres_csv.exists():
                pres_csv = out_consolidated_dir / f"presentation_{rut_with_dv}_{ym_range}_{lang}.csv"
                import pandas as pd
                pd.read_csv(latest_pres_csv, engine='python').to_csv(pres_csv, index=False)
             
            # Llamar xbrl_to_excel.py con argumentos sintéticos
            old_argv = _sys.argv.copy()
            try:
                _sys.argv = [
                    "xbrl_to_excel.py",
                    str(out_consolidated_dir),
                    f"{rut_with_dv}_{ym_range}",
                    lang
                ]
                x2e_main()
                print(f"      ✅ {lang}: Excel generado en {out_consolidated_dir}")
                 
            finally:
                _sys.argv = old_argv
                 
        except Exception as e:
            print(f"      ❌ {lang}: Error - {e}")
            continue

    # Alinear columnas ES/EN en out_consolidated
    try:
        if out_consolidated_dir.exists():
            post_align_language_excels(out_consolidated_dir, stem_tmp)
            print(f"      🔄 Columnas alineadas entre idiomas")
    except Exception as e:
        print(f"      ❌ Error al alinear columnas: {e}")
    
    # Mantener el directorio out_consolidated para uso posterior
    print(f"      📁 Directorio consolidado creado: {out_consolidated_dir}")
    
    # COPIAR A PRODUCTS/TOTAL (esto faltaba!)
    # SKIP si CMF_SKIP_OLD_EXCEL está configurado (para usar primary_csv_to_excel en su lugar)
    
    if _os.getenv('CMF_SKIP_OLD_EXCEL') != '1':
        # Normalizar RUT para archivos de destino (con guión y dígito verificador)
        rut_normalized = normalize_rut_with_dv(rut_with_dv.split('-')[0] if '-' in rut_with_dv else rut_with_dv)
        
        for lang in langs:
            # El archivo fuente usa el formato: estados_{rut}_{periodoInicio-periodoFin}_{lang}.xlsx
            src = out_consolidated_dir / f"estados_{rut_with_dv}_{ym_range}_{lang}.xlsx"
            if src.exists():
                # El archivo destino debe usar el formato: estados_{rut_normalizado}_{año_inicio}-{año_fin}_{lang}.xlsx
                dst = products_total_dir / f"estados_{rut_normalized}_{year_range}_{lang}.xlsx"
                try:
                    copy2(src, dst)
                    print(f"      📁 Copiado a Products/Total: {dst.name}")
                    created_outputs[lang] = dst
                except Exception as e:
                    print(f"      ❌ Error copiando a Products/Total: {e}")
            else:
                print(f"      ⚠️ Archivo fuente no encontrado: {src.name}")
                # Intentar listar los archivos que sí existen para debug
                existing_files = list(out_consolidated_dir.glob(f"estados_{rut_with_dv}_*_{lang}.xlsx"))
                if existing_files:
                    print(f"      📋 Archivos disponibles: {[f.name for f in existing_files]}")
                    # Usar el primer archivo disponible si existe
                    actual_src = existing_files[0]
                    dst = products_total_dir / f"estados_{rut_normalized}_{year_range}_{lang}.xlsx"
                    try:
                        copy2(actual_src, dst)
                        print(f"      📁 Copiado archivo alternativo a Products/Total: {dst.name}")
                        created_outputs[lang] = dst
                    except Exception as e:
                        print(f"      ❌ Error copiando archivo alternativo a Products/Total: {e}")
    else:
        if debug:
            print(f"      ║ ⏩ Saltando copia a Products/Total (CMF_SKIP_OLD_EXCEL=1) - se usará primary_csv_to_excel")
    
    return created_outputs


def _detect_header_row(ws) -> int:
    import re
    for r in range(1, min(10, ws.max_row) + 1):
        v0 = ws.cell(row=r, column=1).value
        if isinstance(v0, str) and v0.strip().lower() in ("cuenta", "concepto", "account"):
            return r
        for c in range(2, min(ws.max_column, 20) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and re.match(r"^\d{4}-", v):
                return r
    return 3



def post_align_language_excels(out_dir: Path, stem: str) -> None:
    """Solo aplica formato/agrupación al Excel en español (ya no hay inglés)."""
    es_path = out_dir / f"estados_{stem}_es.xlsx"
    if not es_path.exists():
        return

    try:
        wb_es = load_workbook(str(es_path))
        
        # Solo aplicar agrupación al español
        def _apply_grouping_like_writer(ws, hdr_row: int) -> None:
            """Agrupa columnas trimestrales (YYYYQ1..Q4) bajo su año, ocultando bloques salvo el último año."""
            try:
                from openpyxl.utils import get_column_letter
                import re as _re
                
                labels: dict[int, str] = {}
                for c in range(2, ws.max_column + 1):
                    v = ws.cell(row=hdr_row, column=c).value
                    if isinstance(v, str):
                        s = v.strip().split("\n", 1)[0]
                        if _re.match(r"^\d{4}(?:Q[1-4])?$", s):
                            labels[c] = s
                
                if not labels:
                    return {}
                    
                years: set[int] = set()
                for s in labels.values():
                    if _re.match(r"^\d{4}$", s):
                        years.add(int(s))
                    else:
                        m = _re.match(r"^(\d{4})Q[1-4]$", s)
                        if m:
                            years.add(int(m.group(1)))
                
                if not years:
                    return {}
                    
                latest_year = max(years)
                year_col: dict[int, int] = {}
                year_quarters: dict[int, list[int]] = {}
                
                # Map columns: Q4 is the summary column, Q1-Q3 are grouped under it
                year_quarter_map: dict[int, dict[int, int]] = {}  # year → {quarter: col_idx}
                for c, s in labels.items():
                    m = _re.match(r"^(\d{4})Q([1-4])$", s)
                    if m:
                        y = int(m.group(1))
                        q = int(m.group(2))
                        year_quarter_map.setdefault(y, {})[q] = c
                    elif _re.match(r"^\d{4}$", s):
                        # Bare year → treat as Q4 summary
                        year_quarter_map.setdefault(int(s), {})[4] = c

                for y, qmap in year_quarter_map.items():
                    hide_block = (y != latest_year)
                    # Only group Q1-Q3 (not Q4 — Q4 is the summary, always visible)
                    inner_cols = sorted([ci for q, ci in qmap.items() if q != 4])
                    q4_col = qmap.get(4)

                    for c in inner_cols:
                        letter = get_column_letter(c)
                        dim = ws.column_dimensions.get(letter)
                        if dim is not None:
                            try:
                                dim.outlineLevel = 1
                                dim.hidden = hide_block
                            except Exception:
                                pass

                    # Q4 column is summary (collapsed marker, never hidden)
                    if q4_col and inner_cols:
                        ly = get_column_letter(q4_col)
                        dimy = ws.column_dimensions.get(ly)
                        if dimy is not None:
                            try:
                                dimy.collapsed = hide_block
                            except Exception:
                                pass

                # Outline properties: summary to the RIGHT (Q4 is after Q1-Q3)
                try:
                    from openpyxl.worksheet.properties import Outline
                    if not hasattr(ws.sheet_properties, 'outlinePr') or ws.sheet_properties.outlinePr is None:
                        ws.sheet_properties.outlinePr = Outline()
                    ws.sheet_properties.outlinePr.summaryBelow = False
                    ws.sheet_properties.outlinePr.summaryRight = True
                    ws.sheet_properties.outlinePr.applyStyles = True
                    ws.sheet_view.showOutlineSymbols = True
                except Exception:
                    pass
            except Exception:
                pass

        # Aplicar agrupación a hojas en español
        for sheet_name in ["Balance General", "Estado de Resultados", "Flujo Efectivo"]:
            if sheet_name in wb_es.sheetnames:
                ws = wb_es[sheet_name]
                hdr_row = _detect_header_row(ws)
                _apply_grouping_like_writer(ws, hdr_row)

        wb_es.save(str(es_path))
    except Exception:
        pass


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Procesamiento masivo de XBRL → Excel')
    # Por defecto procesar Total (se puede pasar Trimestral mediante --base-dir)
    parser.add_argument('--base-dir', type=Path, default=Path('data/XBRL/Total'), help='Raíz con carpetas por empresa o frecuencia (Anual/Trimestral)')
    parser.add_argument('--arelle-dir', type=Path, default=Path('/home/unzzui/Documents/Arelle'), help='Directorio de Arelle')
    parser.add_argument('--cmf-dir', type=Path, default=Path(__file__).resolve().parent, help='Raíz del repo con xbrl_to_excel.py')
    parser.add_argument('--langs', nargs='+', default=['es'], choices=['es'], help='Solo español disponible')
    parser.add_argument('--products-dir', type=Path, default=Path(__file__).resolve().parent / 'Products', help='Directorio destino para Excel finales (solo consolidados por empresa)')
    parser.add_argument('--max', type=int, default=0, help='Limitar número de datasets (0 = sin límite)')
    parser.add_argument('--dry-run', action='store_true', help='Solo listar sin ejecutar Arelle/Excel')
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)

    datasets = find_datasets(args.base_dir)
    # Si no hay datasets en Anual y el usuario no forzó base-dir, intentar Trimestral como fallback
    if not datasets and str(args.base_dir).endswith('data/XBRL/Anual'):
        alt_base = Path('data/XBRL/Trimestral')
        if alt_base.exists():
            print("No se hallaron datasets en Anual. Probando Trimestral…")
            datasets = find_datasets(alt_base)
    if args.max and len(datasets) > args.max:
        datasets = datasets[: args.max]

    if not datasets:
        print(f"No se encontraron datasets en: {args.base_dir}")
        return 1

    print(f"Se encontraron {len(datasets)} dataset(s) para procesar en {args.base_dir}")

    processed = 0
    errors: List[Tuple[DatasetInfo, str]] = []

    # Asegurar carpeta Products
    args.products_dir.mkdir(parents=True, exist_ok=True)

    # Agrupar por empresa (directorio raíz) para consolidado
    company_to_datasets: dict[Path, List[DatasetInfo]] = {}
    for ds in datasets:
        company_to_datasets.setdefault(ds.company_dir, []).append(ds)

    # Ejecutar por empresa en paralelo (limitar workers para no saturar)
    env_workers = os.getenv('CMF_WORKERS', 'auto')
    if env_workers.lower() != 'auto':
        try:
            max_workers = max(1, int(env_workers))
        except Exception:
            max_workers = min(8, os.cpu_count() or 4)
    else:
        max_workers = min(8, os.cpu_count() or 4)

    # Preparar Dashboard de consola (por dataset)
    try:
        from analisis_excel.utils.console_dashboard import ConsoleXBRLDashboard, get_global_dashboard
    except Exception:
        ConsoleXBRLDashboard = None  # type: ignore
        def get_global_dashboard():  # type: ignore
            return None

    def _company_name_from_dir(p: Path) -> str:
        nm = p.name
        if '_' in nm:
            try:
                return nm.split('_', 1)[1].replace('_', ' ').strip()
            except Exception:
                return nm
        return nm

    def _row_display(ds: DatasetInfo) -> tuple[str, str]:
        """Devuelve (empresa_name, rut_display) para el dashboard."""
        empresa_name = _company_name_from_dir(ds.company_dir)
        rut_display = ds.stem
        return empresa_name, rut_display

    # Reusar dashboard global si existe; si no, crear local
    dashboard = get_global_dashboard()
    created_local_dash = False
    if dashboard is None and ConsoleXBRLDashboard is not None and datasets:
        items = []
        for ds in datasets:
            # Mostrar y usar siempre el identificador único del dataset (ds.stem)
            rut_disp = ds.stem
            items.append({
                'key': ds.stem,           # clave única para actualizar esta fila
                'rut': rut_disp,          # se muestra en la columna RUT (dataset id)
                'razon_social': _company_name_from_dir(ds.company_dir),
            })
        # Crear directorio de logs si no existe
        try:
            (Path('data') / 'debug').mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        dashboard = ConsoleXBRLDashboard(items, mute_stdout_logs=True, log_to_file=True,
                                          log_file_path=str((Path('data') / 'debug' / 'xbrl_run.log').resolve()))
        dashboard.start()
        created_local_dash = True

    def _update_dash(ds: DatasetInfo, **kwargs):
        if dashboard is None:
            return
        # Usar ds.stem como clave estable y también mostrarlo como rut_display
        key = ds.stem
        try:
            empresa_name, rut_display = _row_display(ds)
            dashboard.update(key, empresa_name=empresa_name, rut_display=key, periodo=ds.yyyyymm, **kwargs)
        except Exception:
            pass

    def process_dataset(ds: DatasetInfo) -> Tuple[DatasetInfo, str | None]:
        stem = ds.stem
        try:
            t_total_start = time.perf_counter()
            xbrl_file = find_xbrl_file(ds.dataset_dir, stem)
            if not xbrl_file or not xbrl_file.exists():
                _update_dash(ds, estado="Error: sin .xbrl")
                return ds, f"No se encontró .xbrl en {ds.dataset_dir}"

            out_dir = ds.dataset_dir / f"out_{stem}"

            if args.dry_run:
                print(f"[DRY-RUN] Arelle → {out_dir}; Excel por {stem} idiomas={args.langs}")
                _update_dash(ds, estado="Completado (dry-run)", progreso="arelle+skip", current=1, total=1)
                return ds, None

            # Solo exportar facts en español
            facts_strategy = 'es_only'
            t_arelle_start = time.perf_counter()
            # Solo español: facts + presentation = 2 pasos
            langs = list(args.langs)
            total_steps = 2 * len(langs)  # facts_es + pre_es = 2

            step_idx = 0
            _update_dash(ds, estado="Preparando", worker=threading.get_ident(), archivos='-', progreso='—', current=0, total=total_steps)

            # Intentar usar versión con callback de progreso; si no existe, fallback
            try:
                run_arelle_exports_progress  # type: ignore[name-defined]
                def _progress(stage: str):
                    nonlocal step_idx
                    stage_clean = stage.replace('_', ' ')
                    if stage_clean.lower().startswith(('facts', 'pre')):
                        step_idx += 1
                    _update_dash(ds, estado="En progreso", progreso=stage_clean, current=step_idx, total=total_steps)
                run_arelle_exports_progress(args.arelle_dir, xbrl_file, out_dir, stem, langs, facts_strategy, progress_cb=_progress)  # type: ignore[name-defined]
            except NameError:
                _update_dash(ds, estado="Descargando (arelle)", progreso="arelle")
                run_arelle_exports(args.arelle_dir, xbrl_file, out_dir, stem, langs, facts_strategy=facts_strategy)
                step_idx = total_steps
                _update_dash(ds, estado="En progreso", progreso="pre", current=step_idx, total=total_steps)
            t_arelle = time.perf_counter() - t_arelle_start
            # OMITIR generación de Excel por PERÍODO (solo se generará el consolidado más adelante)
            # Si en algún momento deseas reactivarlo, habilita la siguiente línea:
            # generate_excels(args.cmf_dir, out_dir, stem, args.langs)
            # y opcionalmente:
            # post_align_language_excels(out_dir, stem)
            t_total = time.perf_counter() - t_total_start
            print(f"✔ OK: {stem} → {out_dir}  ⏱ total {t_total:.1f}s (arelle {t_arelle:.1f}s)")
            _update_dash(ds, estado="Completado", progreso=f"{len(langs)} idioma(s)", current=total_steps, total=total_steps)
            return ds, None
        except subprocess.CalledProcessError as cpe:
            _update_dash(ds, estado=f"Error: exit {cpe.returncode}")
            return ds, f"Error ejecutando comando (exit {cpe.returncode})"
        except Exception as ex:
            _update_dash(ds, estado=f"Error: {str(ex)[:40]}")
            return ds, str(ex)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_dataset, ds) for ds in datasets]
        for fut in as_completed(futures):
            ds, err = fut.result()
            if err:
                errors.append((ds, err))
                print(f"❌ {ds.stem}: {err}")
            else:
                processed += 1

    # Consolidado por empresa - DESHABILITADO para evitar duplicados
    # La función generate_combined_company_from_total es más completa y hace todo lo necesario
    created_outputs: dict[str, Path] = {}
    # for company_dir, cds in company_to_datasets.items():
    #     cds_sorted = sorted(cds, key=lambda d: d.yyyyymm)
    #     try:
    #         generate_consolidated_company(company_dir, cds_sorted, args.cmf_dir, args.langs, args.products_dir)
    #     except Exception as ex:
    #         print(f"⚠ Consolidado falló para {company_dir.name}: {ex}")
    
    # Solo trabajamos con español - no hay reconciliación necesaria

    # Consolidado COMBINADO por empresa (combina XBRL/Total)
    for company_dir, cds in company_to_datasets.items():
        cds_sorted = sorted(cds, key=lambda d: d.yyyyymm)
        try:
            generated_excels_outputs = generate_combined_company_from_total(company_dir, cds_sorted, args.cmf_dir, args.langs, args.products_dir)
            # Actualizar created_outputs con los archivos generados
            created_outputs.update(generated_excels_outputs)
        except Exception as ex:
            print(f"⚠ Consolidado COMBINADO falló para {company_dir.name}: {ex}")

    print(f"\nResumen: {processed} procesado(s), {len(errors)} con error(es)")
    if errors:
        for ds, msg in errors[:10]:
            print(f"  - {ds.stem}: {msg}")
    if created_local_dash and dashboard is not None:
        try:
            dashboard.stop()
        except Exception:
            pass
    return 0 if not errors else 2


if __name__ == '__main__':
    raise SystemExit(main())