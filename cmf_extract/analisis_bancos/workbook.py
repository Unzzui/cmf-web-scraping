"""Arma el workbook de analisis bancario a partir de un BankData.

Reutiliza cmf_extract/excel_style.py para todo el lenguaje visual y replica el formato del
Product_v1 de empresas: cabecera con titulo/subtitulo sobre tinta, bandas de seccion, columnas
Ultimo/Promedio/Tendencia, portada tipo dashboard y ficha en dos columnas. Los ratios se
escriben como formulas Excel vivas que referencian las celdas de los estados.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cmf_extract import excel_style as est
from cmf_extract.analisis_bancos import concept_map as cm

SHEET_BALANCE = "Estado de Situacion"
SHEET_RESULTADO = "Estado de Resultados"
SHEET_CAPITAL = "Adecuacion de Capital"
SHEET_RATIOS = "RATIOS & KPIs"
SHEET_VALUACION = "Valuacion"
SHEET_FICHA = "Ficha Tecnica"
SHEET_METODO = "Metodologia"
SHEET_INICIO = "Inicio"

# Parametros de valuacion (editables en la hoja)
RF_DEFAULT = 0.055
ERP_DEFAULT = 0.065
BETA_DEFAULT = 0.9
G_DEFAULT = 0.03

# Formatos numericos de la hoja de ratios (planos, como en el producto de empresas)
RFMT_PCT = "0.0%"
RFMT_MULT = "0.00"
RFMT_NUM = "#,##0"

_BORDER = est.BORDE_CAJA  # hairline en las cuatro caras (color est.LINE)


def _period_label(y: int, m: int) -> str:
    return f"{y}-{m:02d}"


def _assign_columns(data) -> None:
    for i, (y, m) in enumerate(data.periods):
        data.col_of[(y, m)] = get_column_letter(2 + i)


def _sheet_header(ws, title: str, subtitle: str, ncols: int, note: str | None = None) -> None:
    """Bloque de cabecera comun: titulo (fila 1, sobre tinta), subtitulo (fila 2), nota (fila 3).
    La fila 4 queda para la cabecera de tabla."""
    ws.sheet_view.showGridLines = False
    ncols = max(ncols, 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font = est.fuente(13, bold=True, color=est.PAPER)
    c.fill = est.RELLENO_TINTA
    c.alignment = est.CENTRO
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitle)
    c.font = est.fuente(11, bold=True, color=est.INK)
    c.alignment = est.CENTRO
    ws.row_dimensions[2].height = 18
    if note:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        c = ws.cell(3, 1, note)
        c.font = est.NOTA
        c.alignment = est.IZQ


def _table_header(ws, headers: list[str], row: int = 4) -> None:
    """Fila de cabecera de tabla: blanco-bold sobre tinta, centrada, con borde."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.font = est.fuente(11, bold=True, color=est.PAPER)
        c.fill = est.RELLENO_TINTA
        c.alignment = est.CENTRO if i > 1 else est.IZQ
        c.border = _BORDER
    ws.row_dimensions[row].height = 20


def _write_statement(ws, data, accounts, values, statement, title) -> None:
    ncols = 1 + len(data.periods)
    _sheet_header(ws, title, f"{data.nombre}  -  RUT {data.rut or '-'}  -  cifras en $ (moneda total)",
                  ncols, note="Saldos de fin de mes. Fuente: API CMF, plan de cuentas Compendio 2022.")
    _table_header(ws, ["Cuenta"] + [_period_label(y, m) for (y, m) in data.periods])
    ws.column_dimensions["A"].width = 64
    for (y, m), col in data.col_of.items():
        ws.column_dimensions[col].width = 15
    for r, acc in enumerate(accounts, start=5):
        es_total = acc.codigo.endswith("000000")
        label = ws.cell(r, 1, f"{acc.codigo}  {acc.descripcion}")
        label.font = est.CUERPO_FUERTE if es_total else est.CUERPO
        label.alignment = est.IZQ
        label.border = _BORDER
        if es_total:
            label.fill = est.RELLENO_SUAVE
        data.row_of[(statement, acc.codigo)] = r
        for (y, m), col in data.col_of.items():
            cell = ws.cell(r, ws[f"{col}1"].column, values.get((acc.codigo, (y, m))))
            cell.number_format = est.FMT_NUM
            cell.font = est.CUERPO_FUERTE if es_total else est.CUERPO
            cell.alignment = est.DER
            cell.border = _BORDER
            if es_total:
                cell.fill = est.RELLENO_SUAVE
    ws.freeze_panes = "B5"


def _write_capital(ws, data) -> dict:
    ncols = 1 + len(data.periods)
    _sheet_header(ws, "Adecuacion de Capital (Basilea)", f"{data.nombre}  -  indicadores de solvencia",
                  ncols, note="De la CMF; cobertura irregular, los meses sin dato quedan en blanco.")
    _table_header(ws, ["Indicador"] + [_period_label(y, m) for (y, m) in data.periods])
    ws.column_dimensions["A"].width = 42
    for col in data.col_of.values():
        ws.column_dimensions[col].width = 15
    filas = [
        ("irs", "IRS - Indice de solvencia"),
        ("ire", "IRE - Indice de endeudamiento"),
        ("capital_basico", "Capital basico"),
        ("patrimonio_efectivo", "Patrimonio efectivo"),
        ("apr", "Activos ponderados por riesgo (APR)"),
    ]
    cap_row_of: dict[str, int] = {}
    for r, (key, label) in enumerate(filas, start=5):
        c = ws.cell(r, 1, label)
        c.font = est.CUERPO
        c.alignment = est.IZQ
        c.border = _BORDER
        cap_row_of[key] = r
        for (y, m), col in data.col_of.items():
            cell = ws.cell(r, ws[f"{col}1"].column, data.capital.get((y, m), {}).get(key))
            cell.number_format = est.FMT_NUM
            cell.font = est.CUERPO
            cell.alignment = est.DER
            cell.border = _BORDER
    ws.freeze_panes = "B5"
    return cap_row_of


# ---- referencias de celdas para las formulas de ratios ----

def _ref(data, concept: str, col: str):
    codes = cm.codes_for(concept)
    stmt = cm.statement_for(concept)
    sheet = SHEET_BALANCE if stmt == "balance" else SHEET_RESULTADO
    parts = []
    for code in codes:
        row = data.row_of.get((stmt, code))
        if row is None:
            return None
        parts.append(f"'{sheet}'!{col}{row}")
    return "(" + "+".join(parts) + ")"


def _ann(ref: str, m: int) -> str:
    return f"({ref}*12/{m})"


def _cap(cap_row_of, field: str, col: str) -> str:
    return f"'{SHEET_CAPITAL}'!{col}{cap_row_of[field]}"


def _ratio_defs():
    def roe(d, col, y, m, cr):
        ni, pat = _ref(d, "resultado_ejercicio", col), _ref(d, "patrimonio", col)
        return f"{_ann(ni, m)}/{pat}" if ni and pat else None

    def roa(d, col, y, m, cr):
        ni, act = _ref(d, "resultado_ejercicio", col), _ref(d, "activos_total", col)
        return f"{_ann(ni, m)}/{act}" if ni and act else None

    def nim(d, col, y, m, cr):
        nii, nir = _ref(d, "ingreso_neto_intereses", col), _ref(d, "ingreso_neto_reajustes", col)
        act = _ref(d, "activos_total", col)
        return f"(({nii}+{nir})*12/{m})/{act}" if nii and nir and act else None

    def margen_op(d, col, y, m, cr):
        ro, ing = _ref(d, "resultado_operacional", col), _ref(d, "total_ingresos_operacionales", col)
        return f"{ro}/{ing}" if ro and ing else None

    def eficiencia(d, col, y, m, cr):
        g, ing = _ref(d, "total_gastos_operacionales", col), _ref(d, "total_ingresos_operacionales", col)
        return f"ABS({g})/{ing}" if g and ing else None

    def costo_riesgo(d, col, y, m, cr):
        gp, colo = _ref(d, "gasto_perdidas_crediticias", col), _ref(d, "colocaciones", col)
        return f"ABS({_ann(gp, m)})/{colo}" if gp and colo else None

    def prov_col(d, col, y, m, cr):
        pr, colo = _ref(d, "provisiones_colocaciones", col), _ref(d, "colocaciones", col)
        return f"ABS({pr})/{colo}" if pr and colo else None

    def ldr(d, col, y, m, cr):
        colo, dep = _ref(d, "colocaciones", col), _ref(d, "depositos_clientes", col)
        return f"{colo}/{dep}" if colo and dep else None

    def col_act(d, col, y, m, cr):
        colo, act = _ref(d, "colocaciones", col), _ref(d, "activos_total", col)
        return f"{colo}/{act}" if colo and act else None

    def dep_pas(d, col, y, m, cr):
        dep, pas = _ref(d, "depositos_clientes", col), _ref(d, "pasivos_total", col)
        return f"{dep}/{pas}" if dep and pas else None

    def apalancamiento(d, col, y, m, cr):
        act, pat = _ref(d, "activos_total", col), _ref(d, "patrimonio", col)
        return f"{act}/{pat}" if act and pat else None

    def irs(d, col, y, m, cr):
        return _cap(cr, "irs", col)

    def ire(d, col, y, m, cr):
        return _cap(cr, "ire", col)

    def cap_apr(d, col, y, m, cr):
        return f"{_cap(cr, 'capital_basico', col)}/{_cap(cr, 'apr', col)}"

    def _yoy(concept):
        def f(d, col, y, m, cr):
            prev = d.col_of.get((y - 1, m))
            if not prev:
                return None
            a, b = _ref(d, concept, col), _ref(d, concept, prev)
            return f"{a}/{b}-1" if a and b else None
        return f

    return [
        ("RENTABILIDAD", "ROE (anualizado)", RFMT_PCT, roe),
        ("RENTABILIDAD", "ROA (anualizado)", RFMT_PCT, roa),
        ("RENTABILIDAD", "Margen de interes neto (NIM)", RFMT_PCT, nim),
        ("RENTABILIDAD", "Margen operacional", RFMT_PCT, margen_op),
        ("EFICIENCIA", "Indice de eficiencia (gastos/ingresos)", RFMT_PCT, eficiencia),
        ("RIESGO DE CREDITO", "Costo de riesgo (anualizado)", RFMT_PCT, costo_riesgo),
        ("RIESGO DE CREDITO", "Provisiones / colocaciones", RFMT_PCT, prov_col),
        ("ESTRUCTURA", "Colocaciones / depositos (LDR)", RFMT_PCT, ldr),
        ("ESTRUCTURA", "Colocaciones / activos", RFMT_PCT, col_act),
        ("ESTRUCTURA", "Depositos / pasivos", RFMT_PCT, dep_pas),
        ("ESTRUCTURA", "Apalancamiento (activos / patrimonio)", RFMT_MULT, apalancamiento),
        ("CAPITAL (BASILEA)", "IRS - Indice de solvencia", RFMT_MULT, irs),
        ("CAPITAL (BASILEA)", "IRE - Indice de endeudamiento", RFMT_MULT, ire),
        ("CAPITAL (BASILEA)", "Capital basico / APR", RFMT_PCT, cap_apr),
        ("CRECIMIENTO", "Colocaciones YoY", RFMT_PCT, _yoy("colocaciones")),
        ("CRECIMIENTO", "Depositos YoY", RFMT_PCT, _yoy("depositos_clientes")),
        ("CRECIMIENTO", "Resultado del ejercicio YoY", RFMT_PCT, _yoy("resultado_ejercicio")),
    ]


def _section_band(ws, row: int, ncols: int, label: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, label)
    c.font = est.fuente(11, bold=True, color=est.INK)
    c.fill = est.RELLENO_SUAVE
    c.alignment = est.IZQ
    for cc in range(1, ncols + 1):
        ws.cell(row, cc).border = _BORDER


def _write_ratios(ws, data, cap_row_of) -> None:
    nper = len(data.periods)
    ncols = 1 + nper + 3  # Indicador + periodos + Ultimo/Promedio/Tendencia
    _sheet_header(
        ws, "Analisis Financiero - Ratios Bancarios",
        f"{data.nombre}  -  ratios trazables sobre los estados",
        ncols,
        note="Resultado anualizado = acumulado YTD / mes * 12. Formulas vivas: se recalculan si se editan los estados.",
    )
    headers = (["Indicador"] + [_period_label(y, m) for (y, m) in data.periods]
               + ["Ultimo", "Promedio", "Tendencia"])
    _table_header(ws, headers)
    ws.column_dimensions["A"].width = 34
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    last_col = get_column_letter(1 + nper) if nper else "B"
    ci_last, ci_avg, ci_trend = 2 + nper, 3 + nper, 4 + nper

    r = 5
    cat_actual = None
    for cat, label, fmt, build in _ratio_defs():
        if cat != cat_actual:
            _section_band(ws, r, ncols, cat)
            r += 1
            cat_actual = cat
        lc = ws.cell(r, 1, label)
        lc.font = est.CUERPO
        lc.alignment = est.IZQ
        lc.border = _BORDER
        for (y, m), col in data.col_of.items():
            body = build(data, col, y, m, cap_row_of)
            cell = ws.cell(r, ws[f"{col}1"].column)
            if body:
                cell.value = f'=IFERROR({body},"")'
                cell.number_format = fmt
            cell.font = est.CUERPO
            cell.alignment = est.DER
            cell.border = _BORDER
        if nper:
            rng = f"B{r}:{last_col}{r}"
            last = ws.cell(r, ci_last, f'=IFERROR(LOOKUP(2,1/(--({rng}<>"")),{rng}),"")')
            avg = ws.cell(r, ci_avg, f'=IFERROR(AVERAGE({rng}),"")')
            trend = ws.cell(
                r, ci_trend,
                f'=IFERROR(IF((LOOKUP(2,1/(--({rng}<>"")),{rng}))>'
                f'(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>"")),{rng}))),{rng})),"▲",'
                f'IF((LOOKUP(2,1/(--({rng}<>"")),{rng}))<'
                f'(LOOKUP(2,1/(--({rng}<LOOKUP(2,1/(--({rng}<>"")),{rng}))),{rng})),"▼","→")),"→")',
            )
            for cell in (last, avg):
                cell.number_format = fmt
                cell.font = est.CUERPO
                cell.alignment = est.DER
                cell.border = _BORDER
            trend.font = est.CUERPO
            trend.alignment = est.CENTRO
            trend.border = _BORDER
        r += 1
    ws.freeze_panes = "B5"


def _write_valuacion(ws, data) -> None:
    ws.sheet_view.showGridLines = False
    _sheet_header(
        ws, "Valuacion - Exceso de retorno / P-B garantizado", data.nombre, 2,
        note="P/B garantizado = (ROE - g)/(Ke - g); valor patrimonial = P/B * patrimonio contable. Ke por CAPM.",
    )
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 24
    if not data.periods:
        return
    y, m = data.periods[-1]
    col = data.col_of[(y, m)]
    pat_row = data.row_of.get(("balance", "380000000")) or data.row_of.get(("balance", "300000000"))
    ni_row = data.row_of.get(("resultado", "590000000"))
    pat_ref = f"'{SHEET_BALANCE}'!{col}{pat_row}" if pat_row else None
    ni_ref = f"'{SHEET_RESULTADO}'!{col}{ni_row}" if ni_row else None
    beta = data.market.get("beta") or BETA_DEFAULT

    rows: dict[str, int] = {}
    state = {"row": 5}

    def section(label: str) -> None:
        c = ws.cell(state["row"], 1, label)
        c.font = est.fuente(11, bold=True, color=est.INK)
        c.fill = est.RELLENO_SUAVE
        c.alignment = est.IZQ
        ws.cell(state["row"], 2).fill = est.RELLENO_SUAVE
        ws.cell(state["row"], 1).border = _BORDER
        ws.cell(state["row"], 2).border = _BORDER
        state["row"] += 1

    def put(label, value, fmt=None, font=est.CUERPO, key=None) -> None:
        r = state["row"]
        a = ws.cell(r, 1, label)
        a.font = font
        a.alignment = est.IZQ
        a.border = _BORDER
        b = ws.cell(r, 2)
        b.border = _BORDER
        b.alignment = est.DER
        if value is not None:
            b.value = value
            if fmt:
                b.number_format = fmt
            b.font = font
        if key:
            rows[key] = r
        state["row"] += 1

    section("PARAMETROS")
    put("Tasa libre de riesgo (Rf)", RF_DEFAULT, RFMT_PCT, key="rf")
    put("Premio por riesgo (ERP)", ERP_DEFAULT, RFMT_PCT, key="erp")
    put("Beta", beta, "0.00", key="beta")
    put("Crecimiento largo plazo (g)", G_DEFAULT, RFMT_PCT, key="g")
    put("Costo de capital (Ke = Rf + Beta*ERP)",
        f"=B{rows['rf']}+B{rows['beta']}*B{rows['erp']}", RFMT_PCT, est.CUERPO_FUERTE, key="ke")
    section("BASE CONTABLE")
    put("Patrimonio contable", f"={pat_ref}" if pat_ref else None, est.FMT_NUM, key="pat")
    put("Resultado anualizado", f"={_ann(ni_ref, m)}" if ni_ref else None, est.FMT_NUM, key="ni")
    put("ROE sostenible", f"=IFERROR(B{rows['ni']}/B{rows['pat']},\"\")", RFMT_PCT,
        est.CUERPO_FUERTE, key="roe")
    section("VALUACION")
    put("P/B garantizado = (ROE-g)/(Ke-g)",
        f"=IFERROR((B{rows['roe']}-B{rows['g']})/(B{rows['ke']}-B{rows['g']}),\"\")",
        RFMT_MULT, est.CUERPO_FUERTE, key="pb")
    put("Valor patrimonial intrinseco", f"=IFERROR(B{rows['pb']}*B{rows['pat']},\"\")",
        est.FMT_NUM, est.CUERPO_FUERTE, key="vi")

    mc = data.market.get("market_cap")
    if mc:
        section("MERCADO (banco listado)")
        put("Capitalizacion de mercado", mc, est.FMT_NUM, key="mc")
        put("P/B de mercado", f"=IFERROR(B{rows['mc']}/B{rows['pat']},\"\")", RFMT_MULT, key="pbm")
        put("Prima (+) / Descuento (-) vs intrinseco",
            f"=IFERROR(B{rows['vi']}/B{rows['mc']}-1,\"\")", RFMT_PCT, est.CUERPO_FUERTE)


def _write_ficha(ws, data) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 52
    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, "FICHA TECNICA")
    c.font = est.fuente(14, bold=True, color=est.PAPER)
    c.fill = est.RELLENO_TINTA
    c.alignment = est.CENTRO
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:B2")
    c = ws.cell(2, 1, "Resumen de la institucion")
    c.font = est.fuente(11, color=est.PAPER)
    c.fill = est.RELLENO_TINTA
    c.alignment = est.IZQ
    ws.row_dimensions[2].height = 20
    p = data.profile
    filas = [
        ("Institucion", data.nombre),
        ("Codigo CMF", data.codigo_institucion),
        ("RUT", data.rut or "-"),
        ("SWIFT", p.get("swift") or data.swift or "-"),
        ("Sitio web", p.get("sitio_web") or "-"),
        ("Direccion", p.get("direccion") or "-"),
        ("Empleados", p.get("empleados")),
        ("Sucursales", p.get("sucursales")),
        ("Oficinas", p.get("oficinas")),
        ("Cajeros", p.get("cajeros")),
        ("Ficha al", str(p.get("fecha_publicacion") or "-")),
    ]
    for i, (label, val) in enumerate(filas, start=3):
        a = ws.cell(i, 1, label)
        a.font = est.fuente(10, bold=True, color=est.INK)
        a.fill = est.RELLENO_SUAVE
        a.alignment = est.IZQ
        a.border = _BORDER
        b = ws.cell(i, 2, val)
        b.font = est.CUERPO
        b.alignment = est.IZQ
        b.border = _BORDER
        ws.row_dimensions[i].height = 18


def _write_metodologia(ws, data) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 118
    c = ws.cell(1, 1, "METODOLOGIA")
    c.font = est.fuente(13, bold=True, color=est.PAPER)
    c.fill = est.RELLENO_TINTA
    c.alignment = est.IZQ
    ws.row_dimensions[1].height = 24
    notas = [
        "Fuente: API oficial de Bancos de la CMF (api.cmfchile.cl), tablas bank_*.",
        "Cobertura: plan de cuentas del Compendio de Normas Contables (desde enero 2022).",
        "Periodicidad: mensual. Los saldos de balance son de fin de mes (instantaneos).",
        "El estado de resultados es ACUMULADO del ejercicio (YTD): el valor de un mes es el",
        "   acumulado enero-mes. Los ratios anualizan el resultado: valor YTD / mes * 12.",
        "Los ratios se escriben como formulas Excel que referencian las celdas de los estados",
        "   (trazables y recalculables si se editan los estados).",
        "Adecuacion de capital (IRS/IRE): de la CMF; cobertura irregular, meses sin dato en blanco.",
        "Valuacion: modelo de exceso de retorno. P/B garantizado = (ROE - g)/(Ke - g); Ke por",
        "   CAPM. La comparacion con mercado solo aplica a bancos listados.",
    ]
    for i, n in enumerate(notas, start=3):
        cell = ws.cell(i, 1, n)
        cell.font = est.NOTA
        cell.alignment = est.IZQ


def _write_inicio(ws, data, nav_sheets: list[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["N"].width = 2
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 8.5

    def band(row, text, font, fill, height, align=est.CENTRO):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        c = ws.cell(row, 2, text)
        c.font = font
        if fill:
            c.fill = fill
        c.alignment = align
        ws.row_dimensions[row].height = height

    band(1, "FinData Chile  |  Inteligencia Financiera Profesional",
         est.fuente(18, bold=True, color=est.PAPER), est.RELLENO_TINTA, 40)
    band(4, f"ANALISIS BANCARIO: {data.nombre}", est.TITULO, est.RELLENO_SUAVE, 35)

    rng = ""
    if data.periods:
        y0, m0 = data.periods[0]
        y1, m1 = data.periods[-1]
        rng = f"{_period_label(y0, m0)} a {_period_label(y1, m1)}"

    # Tres tarjetas (cabecera + cuerpo) en columnas B-E, F-I, J-M
    cards = [
        ("INSTITUCION", [f"Nombre: {data.nombre}", f"RUT: {data.rut or '-'}",
                         f"Codigo CMF: {data.codigo_institucion}",
                         f"SWIFT: {data.profile.get('swift') or data.swift or '-'}"]),
        ("COBERTURA TEMPORAL", [f"Periodo: {rng}", f"Meses: {len(data.periods)}",
                                "Frecuencia: mensual", "Fuente: CMF Bancos"]),
        ("ESPECIFICACIONES", ["Plan: Compendio 2022", "Moneda: $ (total)",
                              "Ratios: bancarios", "Formato: formulas vivas"]),
    ]
    for idx, (titulo, lineas) in enumerate(cards):
        c0 = 2 + idx * 4
        ws.merge_cells(start_row=6, start_column=c0, end_row=6, end_column=c0 + 3)
        h = ws.cell(6, c0, titulo)
        h.font = est.ETIQUETA
        h.fill = est.RELLENO_SUAVE
        h.alignment = est.CENTRO
        for c in range(c0, c0 + 4):
            ws.cell(6, c).border = _BORDER
        for j, linea in enumerate(lineas):
            rr = 7 + j
            ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0 + 3)
            cell = ws.cell(rr, c0, linea)
            cell.font = est.fuente(10, color=est.MUTED)
            cell.fill = est.RELLENO_SUAVE
            cell.alignment = est.IZQ
            for c in range(c0, c0 + 4):
                ws.cell(rr, c).border = _BORDER
    for r in range(6, 11):
        ws.row_dimensions[r].height = 22

    band(13, "NAVEGACION - MODULOS DE ANALISIS",
         est.fuente(14, bold=True, color=est.PAPER), est.RELLENO_TINTA, 26)
    row = 14
    for i, name in enumerate(nav_sheets):
        c0 = 2 + (i % 3) * 4
        if i and i % 3 == 0:
            row += 1
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 3)
        cell = ws.cell(row, c0, name)
        cell.font = est.fuente(11, bold=True, color=est.PAPER)
        cell.fill = est.RELLENO_TINTA
        cell.alignment = est.CENTRO
        cell.hyperlink = f"#'{name}'!A1"
        for c in range(c0, c0 + 4):
            ws.cell(row, c).border = _BORDER
        ws.row_dimensions[row].height = 24
    band(row + 2, "FinData Chile  -  findatachile.com",
         est.fuente(12, bold=True, color=est.PAPER), est.RELLENO_TINTA, 26)


def build_workbook(data) -> Workbook:
    _assign_columns(data)
    wb = Workbook()
    wb.remove(wb.active)

    ws_bal = wb.create_sheet(SHEET_BALANCE)
    _write_statement(ws_bal, data, data.balance_accounts, data.balance_values, "balance",
                     "Estado de Situacion")
    ws_res = wb.create_sheet(SHEET_RESULTADO)
    _write_statement(ws_res, data, data.resultado_accounts, data.resultado_values, "resultado",
                     "Estado de Resultados")
    cap_row_of = _write_capital(wb.create_sheet(SHEET_CAPITAL), data)
    _write_ratios(wb.create_sheet(SHEET_RATIOS), data, cap_row_of)
    _write_valuacion(wb.create_sheet(SHEET_VALUACION), data)
    _write_ficha(wb.create_sheet(SHEET_FICHA), data)
    _write_metodologia(wb.create_sheet(SHEET_METODO), data)

    nav = list(wb.sheetnames)
    ws_ini = wb.create_sheet(SHEET_INICIO)
    _write_inicio(ws_ini, data, nav)
    wb.move_sheet(SHEET_INICIO, -(len(wb.sheetnames) - 1))

    est.aplicar_tipografia_base(wb)
    return wb
